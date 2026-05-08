from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Any

_BOX_HEADER_RE = re.compile(r"包装箱\s*(\d+)\s*数量")
_BOX_TOTAL_LABEL = "包装箱总数"
_SKU_TOTAL_LABEL = "SKU 总数"
_SKU_HEADER_LABEL = "SKU"
_SOURCE_BOX_SEQUENCE_COLUMN = "箱序号"
_BOX_SPEC_LABELS = {
    "length": "包装箱长度",
    "width": "包装箱宽度",
    "height": "包装箱高度",
    "weight": "包装箱重量",
}
_MAX_WEIGHT_LB = 40


def _resolve_path(raw: str | Path) -> Path:
    text = str(raw or "").strip()
    if not text:
        raise RuntimeError("路径不能为空")
    path = Path(text).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    else:
        path = path.resolve()
    return path


def _require_pandas():
    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError("缺少 pandas 依赖，无法填充 step2 Excel") from exc
    return pd


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法填充 step2 Excel") from exc
    return load_workbook


def _normalize_box_id(value: Any) -> Any:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return ""
    try:
        numeric = float(text)
    except Exception:
        return text
    rounded = round(numeric)
    if abs(numeric - rounded) > 1e-6:
        return text
    return int(rounded)


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def _parse_positive_int(value: Any) -> int | None:
    text = _normalize_text(value)
    if not text:
        return None
    match = re.search(r"(-?\d+(?:\.\d+)?)", text)
    if not match:
        return None
    try:
        numeric = float(match.group(1))
    except Exception:
        return None
    rounded = round(numeric)
    if rounded <= 0 or abs(numeric - rounded) > 1e-6:
        return None
    return int(rounded)


def _read_source_data(source_file: Path):
    pd = _require_pandas()
    try:
        df = pd.read_excel(source_file)
    except FileNotFoundError as exc:
        raise RuntimeError(f"找不到文件: {source_file}") from exc
    except Exception as exc:
        raise RuntimeError(f"读取文件失败: {exc}") from exc

    expected_columns = [
        "箱子编号",
        _SOURCE_BOX_SEQUENCE_COLUMN,
        "MSKU",
        "FBA产品名称",
        "本地sku",
        "本地sku所属店铺",
        "FNSKU",
        "装箱数量",
        "长",
        "宽",
        "高",
        "毛重",
    ]
    if len(df.columns) >= len(expected_columns):
        df.columns = expected_columns[: len(df.columns)]

    required_cols = (_SOURCE_BOX_SEQUENCE_COLUMN, "MSKU", "装箱数量")
    for column in required_cols:
        if column not in df.columns:
            raise RuntimeError(f"找不到必要列: {column}")

    df[_SOURCE_BOX_SEQUENCE_COLUMN] = df[_SOURCE_BOX_SEQUENCE_COLUMN].map(_normalize_box_id)
    df["MSKU"] = df["MSKU"].map(lambda value: str(value or "").strip())
    df = df[(df[_SOURCE_BOX_SEQUENCE_COLUMN] != "") & (df["MSKU"] != "")]
    if df.empty:
        raise RuntimeError(f"源数据没有有效装箱信息: {source_file.name}")

    pivot = df.groupby([_SOURCE_BOX_SEQUENCE_COLUMN, "MSKU"])["装箱数量"].sum().reset_index()
    pivot_table = pivot.pivot(index="MSKU", columns=_SOURCE_BOX_SEQUENCE_COLUMN, values="装箱数量").fillna(0)
    box_info = df.groupby(_SOURCE_BOX_SEQUENCE_COLUMN)[["长", "宽", "高", "毛重"]].first()
    return pivot_table, box_info


def _sheet_has_step2_anchors(ws: Any) -> bool:
    has_sku_total = False
    has_box_one = False
    for row in ws.iter_rows():
        for cell in row:
            text = _normalize_text(cell.value)
            if not text:
                continue
            if _SKU_TOTAL_LABEL in text:
                has_sku_total = True
            match = _BOX_HEADER_RE.search(text)
            if match and int(match.group(1)) == 1:
                has_box_one = True
            if has_sku_total and has_box_one:
                return True
    return False


def _select_target_sheet(workbook: Any) -> Any:
    for worksheet in workbook.worksheets:
        if _sheet_has_step2_anchors(worksheet):
            return worksheet
    available = [str(name) for name in workbook.sheetnames]
    raise RuntimeError(
        "找不到包含 SKU 总数 和 包装箱 1 数量 的 step2 模板 sheet: "
        f"{available}"
    )


def _find_box_total(ws: Any) -> int:
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = _normalize_text(ws.cell(row=row, column=col).value)
            if _BOX_TOTAL_LABEL not in text:
                continue
            inline_total = _parse_positive_int(text)
            if inline_total is not None:
                return inline_total
            for right_col in range(col + 1, max_col + 1):
                total = _parse_positive_int(ws.cell(row=row, column=right_col).value)
                if total is not None:
                    return total
            raise RuntimeError("已找到 包装箱总数 标签，但同一行右侧未找到有效数字")
    raise RuntimeError("找不到 包装箱总数")


def _find_box_header_row(ws: Any) -> tuple[int, dict[int, int]]:
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    best_row = 0
    best_columns: dict[int, int] = {}
    best_has_sku = False

    for row in range(1, max_row + 1):
        row_columns: dict[int, int] = {}
        row_has_sku = False
        for col in range(1, max_col + 1):
            text = _normalize_text(ws.cell(row=row, column=col).value)
            if not text:
                continue
            if text == _SKU_HEADER_LABEL:
                row_has_sku = True
            match = _BOX_HEADER_RE.search(text)
            if match:
                row_columns[int(match.group(1))] = col
        if not row_columns:
            continue
        if len(row_columns) > len(best_columns) or (
            len(row_columns) == len(best_columns) and row_has_sku and not best_has_sku
        ):
            best_row = row
            best_columns = row_columns
            best_has_sku = row_has_sku

    if not best_columns:
        raise RuntimeError("找不到 step2 模板中的包装箱数量表头行")
    if 1 not in best_columns:
        raise RuntimeError("包装箱数量表头行中缺少 包装箱 1 数量")
    return best_row, best_columns


def _validate_box_columns(box_columns: dict[int, int], total_boxes: int) -> list[int]:
    found = sorted(int(box_num) for box_num in box_columns.keys())
    expected = list(range(1, int(total_boxes) + 1))
    missing = [box_num for box_num in expected if box_num not in box_columns]
    if missing:
        raise RuntimeError(
            "模板缺少必要箱号列: "
            f"total={total_boxes}, missing={missing}, found={found}"
        )
    return expected


def _find_sku_header_cell(ws: Any, header_row: int) -> tuple[int, int]:
    max_col = int(ws.max_column or 0)
    for col in range(1, max_col + 1):
        if _normalize_text(ws.cell(row=header_row, column=col).value) == _SKU_HEADER_LABEL:
            return header_row, col

    max_row = int(ws.max_row or 0)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if _normalize_text(ws.cell(row=row, column=col).value) == _SKU_HEADER_LABEL:
                return row, col
    raise RuntimeError("找不到模板中的 SKU 列标题")


def _find_box_spec_rows(ws: Any) -> dict[str, int]:
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    found: dict[str, int] = {}
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = _normalize_text(ws.cell(row=row, column=col).value)
            if not text:
                continue
            for key, label in _BOX_SPEC_LABELS.items():
                if key not in found and label in text:
                    found[key] = row
        if len(found) == len(_BOX_SPEC_LABELS):
            break

    missing = [label for key, label in _BOX_SPEC_LABELS.items() if key not in found]
    if missing:
        raise RuntimeError(f"找不到模板标签行: {', '.join(missing)}")
    return found


def _build_sku_row_map(
    ws: Any,
    *,
    sku_col: int,
    start_row: int,
    end_row: int,
) -> dict[str, int]:
    if end_row < start_row:
        raise RuntimeError("模板中的 SKU 明细区域为空")

    sku_to_row: dict[str, int] = {}
    for row in range(start_row, end_row + 1):
        text = _normalize_text(ws.cell(row=row, column=sku_col).value)
        if not text:
            continue
        if text == _SKU_HEADER_LABEL or _BOX_TOTAL_LABEL in text or text.startswith("包装箱"):
            continue
        sku_to_row.setdefault(text, row)

    if not sku_to_row:
        raise RuntimeError("模板中未识别到任何 SKU 行")
    return sku_to_row


def _normalize_box_sequence(values: Any, *, field_name: str) -> list[int]:
    normalized: list[int] = []
    for raw in list(values):
        box_num = _normalize_box_id(raw)
        if not isinstance(box_num, int):
            raise RuntimeError(f"{field_name} 包含非整数箱号: {raw}")
        normalized.append(box_num)
    deduped = sorted(set(normalized))
    return deduped


def _validate_source_boxes(
    pivot_table: Any,
    box_info: Any,
    *,
    total_boxes: int,
) -> list[int]:
    source_boxes = _normalize_box_sequence(pivot_table.columns, field_name="托运单数量表")
    box_info_boxes = _normalize_box_sequence(box_info.index, field_name="托运单箱规表")

    if source_boxes != box_info_boxes:
        raise RuntimeError(
            "托运单数量表与箱规表箱号不一致: "
            f"qty={source_boxes}, box_info={box_info_boxes}"
        )

    if not source_boxes:
        raise RuntimeError("托运单未解析到有效箱号")

    if int(source_boxes[-1]) != int(total_boxes):
        raise RuntimeError(
            "托运单最大箱号与包装箱总数不一致: "
            f"total={total_boxes}, found={source_boxes}"
        )
    return source_boxes


def _validate_template_skus(pivot_table: Any, sku_to_row: dict[str, int]) -> list[str]:
    source_skus = [str(item or "").strip() for item in list(pivot_table.index)]
    missing = [sku for sku in source_skus if sku and sku not in sku_to_row]
    if missing:
        preview = ", ".join(missing[:10])
        suffix = " ..." if len(missing) > 10 else ""
        raise RuntimeError(
            f"模板缺少托运单中的 MSKU: count={len(missing)}, missing={preview}{suffix}"
        )
    return source_skus


def _filled_template_path(template_file: Path) -> Path:
    return template_file.with_name(f"{template_file.stem}.filled{template_file.suffix}")


def _to_int(value: Any, factor: float = 1.0) -> int:
    try:
        numeric = float(value or 0)
    except Exception:
        return 0
    return int(round(numeric * factor)) if numeric else 0


def fill_multi_box_step2_template(consignment_excel_path: str | Path, template_path: str | Path) -> dict[str, Any]:
    consignment_file = _resolve_path(consignment_excel_path)
    template_file = _resolve_path(template_path)
    if not consignment_file.is_file():
        raise RuntimeError(f"找不到托运单 Excel: {consignment_file}")
    if not template_file.is_file():
        raise RuntimeError(f"找不到 step2 模板文件: {template_file}")

    pivot_table, box_info = _read_source_data(consignment_file)

    filled_file = _filled_template_path(template_file)
    shutil.copy2(template_file, filled_file)

    load_workbook = _require_openpyxl()
    try:
        workbook = load_workbook(filled_file)
    except Exception as exc:
        raise RuntimeError(f"读取 step2 模板失败: {exc}") from exc

    worksheet = _select_target_sheet(workbook)
    total_boxes = _find_box_total(worksheet)
    header_row, box_columns = _find_box_header_row(worksheet)
    expected_box_numbers = _validate_box_columns(box_columns, total_boxes)
    spec_rows = _find_box_spec_rows(worksheet)
    sku_header_row, sku_col = _find_sku_header_cell(worksheet, header_row)
    data_end_row = min(spec_rows.values()) - 1
    sku_to_row = _build_sku_row_map(
        worksheet,
        sku_col=sku_col,
        start_row=max(header_row, sku_header_row) + 1,
        end_row=data_end_row,
    )
    source_skus = _validate_template_skus(pivot_table, sku_to_row)
    source_box_numbers = _validate_source_boxes(pivot_table, box_info, total_boxes=total_boxes)

    for sku in source_skus:
        row = sku_to_row[sku]
        for box_num in source_box_numbers:
            qty = int(round(float(pivot_table.loc[sku, box_num] or 0)))
            if qty > 0:
                worksheet.cell(row=row, column=box_columns[box_num], value=qty)

    cm_to_inch = 0.393701
    kg_to_lb = 2.20462
    weight_capped_boxes: list[dict[str, int]] = []
    for box_num in source_box_numbers:
        col = box_columns[box_num]
        length_inch = _to_int(box_info.loc[box_num, "长"], cm_to_inch)
        width_inch = _to_int(box_info.loc[box_num, "宽"], cm_to_inch)
        height_inch = _to_int(box_info.loc[box_num, "高"], cm_to_inch)
        weight_lb = _to_int(box_info.loc[box_num, "毛重"], kg_to_lb)
        capped_weight_lb = min(weight_lb, _MAX_WEIGHT_LB)
        if capped_weight_lb != weight_lb:
            weight_capped_boxes.append(
                {
                    "box_no": int(box_num),
                    "original_weight_lb": int(weight_lb),
                    "capped_weight_lb": int(capped_weight_lb),
                }
            )

        worksheet.cell(row=spec_rows["length"], column=col, value=length_inch)
        worksheet.cell(row=spec_rows["width"], column=col, value=width_inch)
        worksheet.cell(row=spec_rows["height"], column=col, value=height_inch)
        worksheet.cell(row=spec_rows["weight"], column=col, value=capped_weight_lb)

    try:
        workbook.save(filled_file)
    except Exception as exc:
        raise RuntimeError(f"保存 step2 填充文件失败: {exc}") from exc

    return {
        "filled_template_path": str(filled_file),
        "weight_capped_boxes": weight_capped_boxes,
    }


__all__ = ["fill_multi_box_step2_template"]
