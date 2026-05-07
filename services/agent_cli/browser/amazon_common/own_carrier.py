from __future__ import annotations

import re
import time
from datetime import date, timedelta
from pathlib import Path
from typing import Any

from services.browser.browser.shadow_dom import SHADOW_DOM_HELPERS_JS


_OWN_CARRIER_SELECTORS = (
    '[data-testid="own-carrier-box"]',
    'kat-box[data-testid="own-carrier-box"]',
    '[data-testid="own-carrier-box-body"]',
)
_LAYOUT_PHASE_3_1 = "phase_3_1"
_LAYOUT_PHASE_3_2 = "phase_3_2"
_LAYOUT_LEGACY = "legacy"
_CROSS_BORDER_NON_PCP_BOX_SELECTOR = '[data-testid="cross-border-non-pcp-box-test-id"]'
_PCP_OPTION_SELECTOR = 'kat-option[value="PCP"]'
_NPCP_OPTION_SELECTOR = 'kat-option[value="nPCP"]'
_PLAN_DELIVERY_WINDOW_INPUT_SELECTOR = '[data-testid="plan-delivery-window-input"]'
_DELIVERY_WINDOW_MODAL_CONFIRM_BUTTON_SELECTOR = '[data-testid="delivery-window-modal-confirm-button"]'
_SHIPPING_CARRIER_NON_PCP_DROPDOWN_SELECTOR = '[data-testid="shipping-carrier-non-pcp-dropdown"]'
_NON_PCP_CARRIER_CHOICES_SELECTOR = '[data-testid="non-pcp-carrier-choices"]'
_TRANSPORTATION_MODE_DROPDOWN_SELECTOR = '[data-testid="transportation-mode-dropdown"]'
_DELIVERY_WINDOW_SECTION_SELECTOR = 'kat-accordion-item[data-testid="delivery-window-section-accordion-header"]'
_DELIVERY_WINDOW_DATE_PICKER_SELECTOR = 'kat-date-picker[data-testid="delivery-window-date-picker"]'
_PHASE_3_2_SHIP_DATE_PICKER_SELECTOR = 'kat-date-picker[data-testid="kat-ship-date-picker"]'
_PHASE_3_2_SHIP_DATE_PICKER_FALLBACK_SELECTORS = (
    "kat-date-picker#sendByDatePicker",
    'kat-date-picker[kat-aria-label="发货日期"]',
)
_DIALOG_SELECTOR = ".dialog"
_DELIVERY_WINDOW_LINK_SELECTOR = '[data-testid="arrival-edit-delivery-window-link"]'
_PLACEMENT_GROUP_SELECTED_ICON_SELECTOR = '.placement-group-tile-selected-icon'
_AMAZON_OPERATIONS_CENTER_TITLE = "亚马逊运营中心"
_AMAZON_OPERATIONS_CENTER_READY_NOTICE = "已进入亚马逊运营中心步骤"
_DATE_INPUT_SELECTOR = 'input[type="text"]'
_CALENDAR_SELECTOR = ".calendar"
_CALENDAR_MONTH_SELECTOR = ".cal-month"
_CALENDAR_PREV_SELECTOR = '[part="calendar-prev-month"]'
_CALENDAR_NEXT_SELECTOR = '[part="calendar-next-month"]'
_CONTINUE_INPUT_TRACKING_TEXT = "继续输入追踪详情"
_TRACKING_INPUT_TEXT = "输入追踪编码"
_FINAL_SUCCESS_NOTICE = "恭喜第三步完成，现在需要输入追踪编码，请运行第四阶段脚本"
_RETURN_TO_STEP2_BUTTON_TEXT = "创建货件并稍后包装"
_RETURN_TO_STEP2_NOTICE = "亚马逊店铺页面店铺出现bug，已返回第二步开头，请执行第二阶段CLI"
_SHIPMENT_SUMMARY_SELECTOR = '[data-testid="shipment-summary"]'
_SHIPMENT_NAME_SELECTOR = '[data-testid="shipment-name"]'
_SHIPMENT_ID_SELECTOR = '[data-testid="shipment-id"]'
_SHIPMENT_TRACKING_ID_SELECTOR = '[data-testid="purchase-order-id"]'
_SHIPMENT_SEND_TO_ADDRESS_SELECTOR = '[data-testid="send-to-address"]'

_MONTH_NAME_TO_NUMBER = {
    "一月": 1,
    "二月": 2,
    "三月": 3,
    "四月": 4,
    "五月": 5,
    "六月": 6,
    "七月": 7,
    "八月": 8,
    "九月": 9,
    "十月": 10,
    "十一月": 11,
    "十二月": 12,
}


class _ReturnedToStep2Start(RuntimeError):
    pass


class _OwnCarrierWorkflowError(RuntimeError):
    def __init__(self, message: str, *, shipment_summary_excel_path: str = "") -> None:
        super().__init__(str(message or "").strip())
        self.shipment_summary_excel_path = str(shipment_summary_excel_path or "").strip()


def _execute_page_script(driver: Any, script_body: str, *args):
    return driver.execute_script(SHADOW_DOM_HELPERS_JS + script_body, *args)


def _read_selector_text(driver: Any, selector: str) -> str:
    raw_text = _execute_page_script(
        driver,
        """
function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

const root = deepQuerySelector(arguments[0]);
if (!root) return '';
return cleanText(root.innerText || root.textContent || '');
""",
        selector,
    )
    return str(raw_text or "").strip()


def _require_openpyxl_workbook():
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法生成货件摘要 Excel") from exc
    return Workbook


def _has_selector(driver: Any, selector: str) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
return Boolean(deepQuerySelector(arguments[0]));
""",
            selector,
        )
    )


def _click_first_matching(driver: Any, selectors: tuple[str, ...]) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const selectors = Array.isArray(arguments[0]) ? arguments[0] : [];

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  try {
    el.click();
    return true;
  } catch (error) {}
  try {
    el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
    return true;
  } catch (error) {}
  return false;
}

for (const selector of selectors) {
  const el = deepQuerySelector(String(selector || ''));
  if (el && clickElement(el)) return true;
}
return false;
""",
            list(selectors),
        )
    )


def _click_dropdown_by_selector(driver: Any, selector: str) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const selector = String(arguments[0] || '');

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  try {
    el.click();
    return true;
  } catch (error) {}
  try {
    el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
    return true;
  } catch (error) {}
  return false;
}

const root = deepQuerySelector(selector);
if (!root) return false;
const header = deepQuerySelector('.select-header[part="dropdown-header"]', root);
if (header && clickElement(header)) return true;
return clickElement(root);
""",
            selector,
        )
    )


def _read_dropdown_value(driver: Any, selector: str) -> str:
    raw_value = _execute_page_script(
        driver,
        """
const root = deepQuerySelector(String(arguments[0] || ''));
if (!root) return '';
for (const value of [root.value, root.getAttribute('value')]) {
  if (value === undefined || value === null) continue;
  const text = String(value).trim();
  if (text) return text;
}
return '';
""",
        selector,
    )
    return str(raw_value or "").strip()


def _dropdown_value_equals(driver: Any, selector: str, expected_value: str) -> bool:
    return _read_dropdown_value(driver, selector).upper() == str(expected_value or "").strip().upper()


def _select_dropdown_option_by_value(
    driver: Any,
    dropdown_selector: str,
    option_value: str,
    *,
    option_text: str = "",
) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const dropdownSelector = String(arguments[0] || '');
const targetValue = String(arguments[1] || '').trim();
const targetText = String(arguments[2] || '').replace(/\\s+/g, ' ').trim();

function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  try {
    el.click();
    return true;
  } catch (error) {}
  try {
    el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
    return true;
  } catch (error) {}
  return false;
}

function readOptionText(el) {
  const lightText = cleanText(el.innerText || el.textContent || '');
  if (lightText) return lightText;
  if (el.shadowRoot) {
    const shadowText = cleanText(el.shadowRoot.innerText || el.shadowRoot.textContent || '');
    if (shadowText) return shadowText;
  }
  return cleanText(el.getAttribute('value') || el.getAttribute('label') || '');
}

function collectOptions(root) {
  const options = [];
  const seen = new Set();
  for (const selector of ['kat-option', '[role="option"]']) {
    for (const el of deepQuerySelectorAll(selector, root)) {
      if (seen.has(el)) continue;
      seen.add(el);
      options.push(el);
    }
  }
  return options;
}

const roots = [];
const dropdown = deepQuerySelector(dropdownSelector);
if (dropdown) roots.push(dropdown);
roots.push(document);

const seen = new Set();
for (const root of roots) {
  for (const el of collectOptions(root)) {
    if (seen.has(el)) continue;
    seen.add(el);
    const value = String(el.value || el.getAttribute('value') || '').trim();
    const text = readOptionText(el);
    if (value !== targetValue && (!targetText || !text.includes(targetText))) continue;
    if (clickElement(el)) return true;
  }
}
return false;
""",
            dropdown_selector,
            option_value,
            option_text,
        )
    )


def _extract_labeled_value(text: str, label: str) -> str:
    safe_text = str(text or "").strip()
    safe_label = str(label or "").strip()
    if not safe_text or not safe_label:
        return safe_text
    for separator in ("：", ":"):
        prefix = f"{safe_label}{separator}"
        if safe_text.startswith(prefix):
            return safe_text[len(prefix):].strip()
    return safe_text


def _collect_shipment_summaries(driver: Any) -> list[dict[str, str]]:
    payload = _execute_page_script(
        driver,
        """
function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

function readWithin(root, selector) {
  if (!root) return '';
  const element = deepQuerySelector(selector, root);
  if (!element) return '';
  return cleanText(element.innerText || element.textContent || '');
}

const roots = Array.from(deepQuerySelectorAll(arguments[0]) || []);
return roots.map((root) => ({
  shipment_name: readWithin(root, arguments[1]),
  shipment_id: readWithin(root, arguments[2]),
  shipment_tracking_id: readWithin(root, arguments[3]),
  send_to_address: readWithin(root, arguments[4]),
}));
""",
        _SHIPMENT_SUMMARY_SELECTOR,
        _SHIPMENT_NAME_SELECTOR,
        _SHIPMENT_ID_SELECTOR,
        _SHIPMENT_TRACKING_ID_SELECTOR,
        _SHIPMENT_SEND_TO_ADDRESS_SELECTOR,
    )
    if not isinstance(payload, list) or not payload:
        raise RuntimeError("未找到货件摘要区域 shipment-summary")

    summaries: list[dict[str, str]] = []
    for index, item in enumerate(payload, start=1):
        row = dict(item or {})
        summary = {
            "shipment_name": _extract_labeled_value(str(row.get("shipment_name") or "").strip(), "货件名称"),
            "shipment_id": _extract_labeled_value(str(row.get("shipment_id") or "").strip(), "货件编号"),
            "shipment_tracking_id": _extract_labeled_value(
                str(row.get("shipment_tracking_id") or "").strip(),
                "货件追踪编号",
            ),
            "send_to_address": _extract_labeled_value(
                str(row.get("send_to_address") or "").strip(),
                "收货地址",
            ),
        }
        missing = [key for key, value in summary.items() if not str(value or "").strip()]
        if missing:
            raise RuntimeError(f"第{index}个货件摘要字段缺失: {', '.join(missing)}")
        summaries.append(summary)
    return summaries


def _sanitize_excel_filename(raw_name: str) -> str:
    safe_name = re.sub(r'[<>:"/\\\\|?*\\x00-\\x1F]+', "_", str(raw_name or "").strip())
    safe_name = safe_name.strip(" .")
    return safe_name or f"shipment-summary-{int(time.time())}"


def _reserve_output_file(directory: Path, filename: str) -> Path:
    target = directory / filename
    if not target.exists():
        return target
    stem = target.stem
    suffix = target.suffix
    for index in range(1, 1000):
        candidate = directory / f"{stem}-{index}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法为货件摘要 Excel 预留文件名: {target}")


def _write_shipment_summary_excel(
    output_dir: Path,
    summaries: list[dict[str, str]],
    *,
    consignment_no: str = "",
) -> str:
    Workbook = _require_openpyxl_workbook()
    safe_output_dir = Path(output_dir).expanduser().resolve()
    safe_output_dir.mkdir(parents=True, exist_ok=True)
    safe_consignment_no = str(consignment_no or "").strip()
    file_stub = (
        f"shipment-summary-{safe_consignment_no}"
        if safe_consignment_no
        else f"shipment-summary-{int(time.time())}"
    )
    excel_path = _reserve_output_file(
        safe_output_dir,
        f"{_sanitize_excel_filename(file_stub)}.xlsx",
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "货件摘要"
    headers = ("序号", "货件名称", "货件编号", "货件追踪编号", "收货地址")
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)
    for row_index, summary in enumerate(list(summaries or []), start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)
        worksheet.cell(row=row_index, column=2, value=str(summary.get("shipment_name") or "").strip())
        worksheet.cell(row=row_index, column=3, value=str(summary.get("shipment_id") or "").strip())
        worksheet.cell(row=row_index, column=4, value=str(summary.get("shipment_tracking_id") or "").strip())
        worksheet.cell(row=row_index, column=5, value=str(summary.get("send_to_address") or "").strip())
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 36
    worksheet.column_dimensions["C"].width = 24
    worksheet.column_dimensions["D"].width = 24
    worksheet.column_dimensions["E"].width = 96

    try:
        workbook.save(excel_path)
    except Exception as exc:
        raise RuntimeError(f"保存货件摘要 Excel 失败: {exc}") from exc
    return str(excel_path)


def _click_button_by_text(driver: Any, text: str) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const targetText = String(arguments[0] || '').replace(/\\s+/g, ' ').trim();

function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

function isVisible(el) {
  if (!el) return false;
  if (typeof el.getClientRects === 'function' && el.getClientRects().length > 0) return true;
  return Boolean(el.offsetWidth || el.offsetHeight);
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  try {
    el.click();
    return true;
  } catch (error) {}
  try {
    el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
    return true;
  } catch (error) {}
  return false;
}

for (const selector of ['button', 'kat-button', '[role="button"]']) {
  for (const el of deepQuerySelectorAll(selector)) {
    const text = cleanText(el.innerText || el.textContent || el.getAttribute('label') || '');
    if (!text || !text.includes(targetText) || !isVisible(el)) continue;
    if (clickElement(el)) return true;
  }
}
return false;
""",
            text,
        )
    )


def _click_select_header(
    driver: Any,
    *,
    index: int | None = None,
) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const targetIndex = Number.isInteger(arguments[0]) ? arguments[0] : null;

function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

function isVisible(el) {
  if (!el) return false;
  if (typeof el.getClientRects === 'function' && el.getClientRects().length > 0) return true;
  return Boolean(el.offsetWidth || el.offsetHeight);
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  try {
    el.click();
    return true;
  } catch (error) {}
  try {
    el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
    return true;
  } catch (error) {}
  return false;
}

const headers = Array.from(
  deepQuerySelectorAll('.select-header[part="dropdown-header"]')
).filter(isVisible);

if (targetIndex !== null && targetIndex >= 0 && targetIndex < headers.length) {
  return clickElement(headers[targetIndex]);
}
return false;
""",
            index,
        )
    )


def _select_option_by_text(driver: Any, text: str) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const targetText = String(arguments[0] || '').replace(/\\s+/g, ' ').trim();

function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

function isVisible(el) {
  if (!el) return false;
  if (typeof el.getClientRects === 'function' && el.getClientRects().length > 0) return true;
  return Boolean(el.offsetWidth || el.offsetHeight);
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  try {
    el.click();
    return true;
  } catch (error) {}
  try {
    el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
    return true;
  } catch (error) {}
  return false;
}

// Read text from element, including shadow DOM content
function readOptionText(el) {
  // 1. Try light DOM text first
  const lightText = cleanText(el.innerText || el.textContent || '');
  if (lightText) return lightText;
  // 2. Penetrate into shadowRoot to read rendered text
  if (el.shadowRoot) {
    const shadowText = cleanText(el.shadowRoot.innerText || el.shadowRoot.textContent || '');
    if (shadowText) return shadowText;
  }
  // 3. Fall back to value / label attributes
  return cleanText(el.getAttribute('value') || el.getAttribute('label') || '');
}

for (const selector of ['kat-option', '[role="option"]']) {
  for (const el of deepQuerySelectorAll(selector)) {
    const text = readOptionText(el);
    if (!text || !text.includes(targetText) || !isVisible(el)) continue;
    if (clickElement(el)) return true;
  }
}
return false;
""",
            text,
        )
    )


def _select_non_amazon_partner_carrier_type(driver: Any) -> bool:
    return _select_option_by_text(driver, "非亚马逊合作承运人")


def _click_own_carrier_entry(driver: Any) -> bool:
    return _click_first_matching(driver, _OWN_CARRIER_SELECTORS)


def _dialog_visible(driver: Any) -> bool:
    return _has_selector(driver, _DIALOG_SELECTOR)


def _click_confirm_modal(driver: Any) -> bool:
    return _click_button_by_text(driver, "Confirm") or _click_button_by_text(driver, "确认")


def _confirm_modal_if_visible(driver: Any) -> bool:
    if not _dialog_visible(driver):
        return True
    return _click_confirm_modal(driver)


def _delivery_window_link_visible(driver: Any) -> bool:
    return _has_selector(driver, _DELIVERY_WINDOW_LINK_SELECTOR)


def _has_pcp_npcp_options(driver: Any) -> bool:
    return _has_selector(driver, _PCP_OPTION_SELECTOR) and _has_selector(driver, _NPCP_OPTION_SELECTOR)


def _detect_own_carrier_layout(driver: Any) -> str:
    if _has_selector(driver, _CROSS_BORDER_NON_PCP_BOX_SELECTOR):
        return _LAYOUT_PHASE_3_2
    if _has_pcp_npcp_options(driver):
        return _LAYOUT_PHASE_3_1
    return _LAYOUT_LEGACY


def _amazon_operations_center_selected(driver: Any) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

const selectedIconSelector = String(arguments[0] || '');
const expectedTitle = cleanText(arguments[1]);
for (const icon of deepQuerySelectorAll(selectedIconSelector)) {
  let current = icon;
  for (let depth = 0; depth < 8 && current; depth += 1) {
    const title = deepQuerySelector('h5', current);
    const text = cleanText(title?.innerText || title?.textContent || '');
    if (text === expectedTitle) {
      return true;
    }
    const rootNode = typeof current.getRootNode === 'function' ? current.getRootNode() : null;
    current = current.parentElement || (rootNode && rootNode.host ? rootNode.host : null);
  }
}
return false;
""",
            _PLACEMENT_GROUP_SELECTED_ICON_SELECTOR,
            _AMAZON_OPERATIONS_CENTER_TITLE,
        )
    )


def _date_picker_visible(driver: Any) -> bool:
    return _delivery_window_link_visible(driver) or _has_selector(driver, _CALENDAR_SELECTOR)


def _calendar_visible(driver: Any) -> bool:
    return _has_selector(driver, _CALENDAR_SELECTOR)


def _click_date_input(driver: Any) -> bool:
    return _click_first_matching(driver, (_DELIVERY_WINDOW_LINK_SELECTOR, _DATE_INPUT_SELECTOR))


def _click_phase_3_1_date_input(driver: Any) -> bool:
    return _click_first_matching(driver, (_PLAN_DELIVERY_WINDOW_INPUT_SELECTOR,))


_PHASE_3_2_DATE_PICKER_SCOPE_JS = r"""
const sectionSelector = String(arguments[0] || '');
const pickerSelector = String(arguments[1] || '');

function scopedQuerySelectorAll(selector, root) {
  const results = [];
  const seen = new Set();

  function add(el) {
    if (!el || seen.has(el)) return;
    seen.add(el);
    results.push(el);
  }

  function walk(node) {
    if (!node) return;
    if (node.matches && node.matches(selector)) add(node);
    if (node.querySelectorAll) {
      for (const match of node.querySelectorAll(selector)) add(match);
      for (const child of node.querySelectorAll('*')) {
        if (child.shadowRoot) walk(child.shadowRoot);
      }
    }
    if (node.shadowRoot) walk(node.shadowRoot);
  }

  walk(root);
  return results;
}

function scopedQuerySelector(selector, root) {
  return scopedQuerySelectorAll(selector, root)[0] || null;
}

function cleanText(value) {
  return String(value || '').replace(/\s+/g, ' ').trim();
}

function getDeliveryWindowDatePicker() {
  const section = deepQuerySelector(sectionSelector);
  if (!section) return null;
  return deepQuerySelector(pickerSelector, section);
}

function getScopedCalendar() {
  const picker = getDeliveryWindowDatePicker();
  if (!picker) return null;
  const calendars = scopedQuerySelectorAll('.calendar', picker);
  return calendars.length ? calendars[calendars.length - 1] : null;
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  const eventTypes = ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click'];
  for (const type of eventTypes) {
    try {
      const EventClass = type.startsWith('pointer') && window.PointerEvent ? PointerEvent : MouseEvent;
      el.dispatchEvent(new EventClass(type, { bubbles: true, cancelable: true, view: window }));
    } catch (error) {
      try {
        el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
      } catch (innerError) {}
    }
  }
  try {
    el.click();
    return true;
  } catch (error) {}
  return true;
}
"""


_PHASE_3_2_SHIP_DATE_SCOPE_JS = r"""
const shipDatePickerSelector = String(arguments[0] || '');
const shipDatePickerFallbackSelectors = Array.isArray(arguments[1]) ? arguments[1] : [];

function scopedQuerySelectorAll(selector, root) {
  const results = [];
  const seen = new Set();

  function add(el) {
    if (!el || seen.has(el)) return;
    seen.add(el);
    results.push(el);
  }

  function walk(node) {
    if (!node) return;
    if (node.matches && node.matches(selector)) add(node);
    if (node.querySelectorAll) {
      for (const match of node.querySelectorAll(selector)) add(match);
      for (const child of node.querySelectorAll('*')) {
        if (child.shadowRoot) walk(child.shadowRoot);
      }
    }
    if (node.shadowRoot) walk(node.shadowRoot);
  }

  walk(root);
  return results;
}

function scopedQuerySelector(selector, root) {
  return scopedQuerySelectorAll(selector, root)[0] || null;
}

function cleanText(value) {
  return String(value || '').replace(/\s+/g, ' ').trim();
}

function getShipDatePicker() {
  for (const selector of [shipDatePickerSelector, ...shipDatePickerFallbackSelectors]) {
    const safeSelector = String(selector || '').trim();
    if (!safeSelector) continue;
    const picker = deepQuerySelector(safeSelector);
    if (picker) return picker;
  }
  return null;
}

function getScopedCalendar() {
  const shipDatePicker = getShipDatePicker();
  if (!shipDatePicker) return null;
  const calendars = scopedQuerySelectorAll('.calendar', shipDatePicker);
  return calendars.length ? calendars[calendars.length - 1] : null;
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  const eventTypes = ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click'];
  for (const type of eventTypes) {
    try {
      const EventClass = type.startsWith('pointer') && window.PointerEvent ? PointerEvent : MouseEvent;
      el.dispatchEvent(new EventClass(type, { bubbles: true, cancelable: true, view: window }));
    } catch (error) {
      try {
        el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
      } catch (innerError) {}
    }
  }
  try {
    el.click();
    return true;
  } catch (error) {}
  return true;
}
"""


def _execute_phase_3_2_date_picker_script(driver: Any, script_body: str, *args):
    return _execute_page_script(
        driver,
        _PHASE_3_2_DATE_PICKER_SCOPE_JS + script_body,
        _DELIVERY_WINDOW_SECTION_SELECTOR,
        _DELIVERY_WINDOW_DATE_PICKER_SELECTOR,
        *args,
    )


def _execute_phase_3_2_ship_date_script(driver: Any, script_body: str, *args):
    return _execute_page_script(
        driver,
        _PHASE_3_2_SHIP_DATE_SCOPE_JS + script_body,
        _PHASE_3_2_SHIP_DATE_PICKER_SELECTOR,
        list(_PHASE_3_2_SHIP_DATE_PICKER_FALLBACK_SELECTORS),
        *args,
    )


def _click_phase_3_2_date_input(driver: Any) -> bool:
    return bool(
        _execute_phase_3_2_date_picker_script(
            driver,
            """
const picker = getDeliveryWindowDatePicker();
if (!picker) return false;

for (const selector of ['input', '[part="input"]', '.container']) {
  const target = scopedQuerySelector(selector, picker);
  if (target && clickElement(target)) return true;
}
return clickElement(picker);
""",
        )
    )


def _click_phase_3_2_ship_date_input(driver: Any) -> bool:
    return bool(
        _execute_phase_3_2_ship_date_script(
            driver,
            """
const shipDatePicker = getShipDatePicker();
if (!shipDatePicker) return false;

for (const selector of ['input', '[part="input"]', '.container']) {
  const target = scopedQuerySelector(selector, shipDatePicker);
  if (target && clickElement(target)) return true;
}
return clickElement(shipDatePicker);
""",
        )
    )


def _click_update_button(driver: Any) -> bool:
    return _click_button_by_text(driver, "更新")


def _click_delivery_window_modal_confirm_button(driver: Any) -> bool:
    return _click_first_matching(driver, (_DELIVERY_WINDOW_MODAL_CONFIRM_BUTTON_SELECTOR,)) or _click_update_button(driver)


def _open_carrier_dropdown(driver: Any) -> bool:
    return _click_select_header(driver, index=1)


def _open_phase_3_1_carrier_dropdown(driver: Any) -> bool:
    return _click_dropdown_by_selector(driver, _SHIPPING_CARRIER_NON_PCP_DROPDOWN_SELECTOR)


def _select_other_carrier(driver: Any) -> bool:
    return _select_option_by_text(driver, "其他")


def _open_transport_mode_dropdown(driver: Any) -> bool:
    return _click_select_header(driver, index=2)


def _select_transport_mode_option(driver: Any, option_text: str) -> bool:
    return _select_option_by_text(driver, option_text)


def _phase_3_2_transport_target_value(mode: dict[str, Any]) -> str:
    canonical = str(mode.get("canonical") or "").strip().upper()
    if canonical in {"AIR", "OCEAN", "GROUND"}:
        return canonical
    ui_text = str(mode.get("ui_text") or "").strip()
    if ui_text == "空运":
        return "AIR"
    if ui_text == "海运":
        return "OCEAN"
    if ui_text == "陆运":
        return "GROUND"
    raise RuntimeError(f"phase_3_2 不支持的运输方式: {ui_text or canonical}")


def _click_accept_fees_button(driver: Any) -> bool:
    return _click_button_by_text(driver, "接受费用并确认发货")


def _click_continue_input_tracking_button(driver: Any) -> bool:
    return _click_button_by_text(driver, _CONTINUE_INPUT_TRACKING_TEXT)


def _button_text_visible(driver: Any, text: str) -> bool:
    return bool(
        _execute_page_script(
            driver,
            """
const targetText = String(arguments[0] || '').replace(/\\s+/g, ' ').trim();

function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

for (const selector of ['button', 'kat-button', '[role="button"]']) {
  for (const el of deepQuerySelectorAll(selector)) {
    const text = cleanText(el.innerText || el.textContent || el.getAttribute('label') || '');
    if (text.includes(targetText)) return true;
  }
}
return false;
""",
            text,
        )
    )


def _continue_input_tracking_visible(driver: Any) -> bool:
    return _button_text_visible(driver, _CONTINUE_INPUT_TRACKING_TEXT)


def _tracking_input_visible(driver: Any) -> bool:
    return _button_text_visible(driver, _TRACKING_INPUT_TEXT) or bool(
        _execute_page_script(
            driver,
            """
const targetText = String(arguments[0] || '').replace(/\\s+/g, ' ').trim();

function cleanText(value) {
  return String(value || '').replace(/\\s+/g, ' ').trim();
}

for (const el of deepQuerySelectorAll('div, span, label, kat-text, p')) {
  const text = cleanText(el.innerText || el.textContent || '');
  if (text.includes(targetText)) return true;
}
return false;
""",
            _TRACKING_INPUT_TEXT,
        )
    )


def _return_to_step2_start_visible(driver: Any) -> bool:
    return _button_text_visible(driver, _RETURN_TO_STEP2_BUTTON_TEXT)


def _raise_if_returned_to_step2_start(driver: Any | None) -> None:
    if driver is None or not hasattr(driver, "execute_script"):
        return
    if _return_to_step2_start_visible(driver):
        raise _ReturnedToStep2Start(_RETURN_TO_STEP2_NOTICE)


def _read_calendar_month_label(driver: Any) -> str:
    return _read_selector_text(driver, _CALENDAR_MONTH_SELECTOR)


def _click_calendar_prev_month(driver: Any) -> bool:
    return _click_first_matching(driver, (_CALENDAR_PREV_SELECTOR,))


def _click_calendar_next_month(driver: Any) -> bool:
    return _click_first_matching(driver, (_CALENDAR_NEXT_SELECTOR,))


def _click_calendar_day(driver: Any, target_date: date) -> bool:
    label_prefix = f"{target_date.year}年{target_date.month}月{target_date.day}日"
    return bool(
        _execute_page_script(
            driver,
            """
const labelPrefix = String(arguments[0] || '').trim();

function isDisabledDay(button) {
  if (!button) return true;
  if (button.disabled || button.getAttribute('aria-disabled') === 'true') return true;
  const cell = deepClosest(button, 'td') || button.parentElement;
  return Boolean(cell && cell.classList && cell.classList.contains('disabled'));
}

function clickElement(el) {
  if (!el) return false;
  try {
    el.scrollIntoView({ block: 'center', inline: 'center' });
  } catch (error) {}
  const eventTypes = ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click'];
  for (const type of eventTypes) {
    try {
      const EventClass = type.startsWith('pointer') && window.PointerEvent ? PointerEvent : MouseEvent;
      el.dispatchEvent(new EventClass(type, { bubbles: true, cancelable: true, view: window }));
    } catch (error) {
      try {
        el.dispatchEvent(new MouseEvent(type, { bubbles: true, cancelable: true, view: window }));
      } catch (innerError) {}
    }
  }
  try {
    el.click();
    return true;
  } catch (error) {}
  return true;
}

for (const button of deepQuerySelectorAll('button[aria-label]')) {
  const label = String(button.getAttribute('aria-label') || '').trim();
  if (!labelPrefix || !label.startsWith(labelPrefix)) continue;
  if (isDisabledDay(button)) continue;
  if (clickElement(button)) return true;
}
return false;
""",
            label_prefix,
        )
    )


def _calendar_day_selected(driver: Any, target_date: date) -> bool:
    label_prefix = f"{target_date.year}年{target_date.month}月{target_date.day}日"
    return bool(
        _execute_page_script(
            driver,
            """
const labelPrefix = String(arguments[0] || '').trim();
for (const button of deepQuerySelectorAll('button[aria-label]')) {
  const label = String(button.getAttribute('aria-label') || '').trim();
  if (!labelPrefix || !label.startsWith(labelPrefix)) continue;
  const cell = deepClosest(button, 'td') || button.parentElement;
  if (button.getAttribute('aria-pressed') === 'true') return true;
  if (button.classList && button.classList.contains('selected')) return true;
  if (cell && cell.classList && cell.classList.contains('selected')) return true;
}
return false;
""",
            label_prefix,
        )
    )


def _phase_3_2_calendar_visible(driver: Any) -> bool:
    return bool(_execute_phase_3_2_date_picker_script(driver, "return Boolean(getScopedCalendar());"))


def _read_phase_3_2_calendar_month_label(driver: Any) -> str:
    raw_label = _execute_phase_3_2_date_picker_script(
        driver,
        """
const calendar = getScopedCalendar();
if (!calendar) return '';
const month = scopedQuerySelector('.cal-month', calendar);
return cleanText(month ? (month.innerText || month.textContent || '') : '');
""",
    )
    return str(raw_label or "").strip()


def _click_phase_3_2_calendar_prev_month(driver: Any) -> bool:
    return bool(
        _execute_phase_3_2_date_picker_script(
            driver,
            """
const calendar = getScopedCalendar();
if (!calendar) return false;
return clickElement(scopedQuerySelector('[part="calendar-prev-month"]', calendar));
""",
        )
    )


def _click_phase_3_2_calendar_next_month(driver: Any) -> bool:
    return bool(
        _execute_phase_3_2_date_picker_script(
            driver,
            """
const calendar = getScopedCalendar();
if (!calendar) return false;
return clickElement(scopedQuerySelector('[part="calendar-next-month"]', calendar));
""",
        )
    )


def _click_phase_3_2_calendar_day(driver: Any, target_date: date) -> bool:
    label_prefix = f"{target_date.year}年{target_date.month}月{target_date.day}日"
    return bool(
        _execute_phase_3_2_date_picker_script(
            driver,
            """
const labelPrefix = String(arguments[2] || '').trim();
const calendar = getScopedCalendar();
if (!calendar) return false;

function isDisabledDay(button) {
  if (!button) return true;
  if (button.disabled || button.getAttribute('aria-disabled') === 'true') return true;
  const cell = deepClosest(button, 'td') || button.parentElement;
  return Boolean(cell && cell.classList && cell.classList.contains('disabled'));
}

for (const button of scopedQuerySelectorAll('button[aria-label]', calendar)) {
  const label = String(button.getAttribute('aria-label') || '').trim();
  if (!labelPrefix || !label.startsWith(labelPrefix)) continue;
  if (isDisabledDay(button)) continue;
  if (clickElement(button)) return true;
}
return false;
""",
            label_prefix,
        )
    )


def _phase_3_2_calendar_day_selected(driver: Any, target_date: date) -> bool:
    label_prefix = f"{target_date.year}年{target_date.month}月{target_date.day}日"
    return bool(
        _execute_phase_3_2_date_picker_script(
            driver,
            """
const labelPrefix = String(arguments[2] || '').trim();
const calendar = getScopedCalendar();
if (!calendar) return false;
for (const button of scopedQuerySelectorAll('button[aria-label]', calendar)) {
  const label = String(button.getAttribute('aria-label') || '').trim();
  if (!labelPrefix || !label.startsWith(labelPrefix)) continue;
  const cell = deepClosest(button, 'td') || button.parentElement;
  if (button.getAttribute('aria-pressed') === 'true') return true;
  if (button.classList && button.classList.contains('selected')) return true;
  if (cell && cell.classList && cell.classList.contains('selected')) return true;
}
return false;
""",
            label_prefix,
        )
    )


def _date_picker_value_matches(value: str, target_date: date) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    match = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", text)
    if not match:
        return False
    year, month, day = (int(part) for part in match.groups())
    return (year, month, day) == (target_date.year, target_date.month, target_date.day)


def _read_phase_3_2_ship_date_value(driver: Any) -> str:
    raw_value = _execute_phase_3_2_ship_date_script(
        driver,
        """
const picker = getShipDatePicker();
if (!picker) return '';

for (const value of [picker.value, picker.getAttribute('value')]) {
  if (value === undefined || value === null) continue;
  const text = String(value).trim();
  if (text) return text;
}

const input = scopedQuerySelector('input', picker);
if (!input) return '';
for (const value of [input.value, input.getAttribute('value')]) {
  if (value === undefined || value === null) continue;
  const text = String(value).trim();
  if (text) return text;
}
return '';
""",
    )
    return str(raw_value or "").strip()


def _phase_3_2_ship_date_value_matches(driver: Any, ship_date: date) -> bool:
    return _date_picker_value_matches(_read_phase_3_2_ship_date_value(driver), ship_date)


def _phase_3_2_ship_date_calendar_visible(driver: Any) -> bool:
    return bool(_execute_phase_3_2_ship_date_script(driver, "return Boolean(getScopedCalendar());"))


def _read_phase_3_2_ship_date_calendar_month_label(driver: Any) -> str:
    raw_label = _execute_phase_3_2_ship_date_script(
        driver,
        """
const calendar = getScopedCalendar();
if (!calendar) return '';
const month = scopedQuerySelector('.cal-month', calendar);
return cleanText(month ? (month.innerText || month.textContent || '') : '');
""",
    )
    return str(raw_label or "").strip()


def _click_phase_3_2_ship_date_calendar_prev_month(driver: Any) -> bool:
    return bool(
        _execute_phase_3_2_ship_date_script(
            driver,
            """
const calendar = getScopedCalendar();
if (!calendar) return false;
return clickElement(scopedQuerySelector('[part="calendar-prev-month"]', calendar));
""",
        )
    )


def _click_phase_3_2_ship_date_calendar_next_month(driver: Any) -> bool:
    return bool(
        _execute_phase_3_2_ship_date_script(
            driver,
            """
const calendar = getScopedCalendar();
if (!calendar) return false;
return clickElement(scopedQuerySelector('[part="calendar-next-month"]', calendar));
""",
        )
    )


def _click_phase_3_2_ship_date_calendar_day(driver: Any, target_date: date) -> bool:
    label_prefix = f"{target_date.year}年{target_date.month}月{target_date.day}日"
    return bool(
        _execute_phase_3_2_ship_date_script(
            driver,
            """
const labelPrefix = String(arguments[2] || '').trim();
const calendar = getScopedCalendar();
if (!calendar) return false;

function isDisabledDay(button) {
  if (!button) return true;
  if (button.disabled || button.getAttribute('aria-disabled') === 'true') return true;
  const cell = deepClosest(button, 'td') || button.parentElement;
  return Boolean(cell && cell.classList && cell.classList.contains('disabled'));
}

for (const button of scopedQuerySelectorAll('button[aria-label]', calendar)) {
  const label = String(button.getAttribute('aria-label') || '').trim();
  if (!labelPrefix || !label.startsWith(labelPrefix)) continue;
  if (isDisabledDay(button)) continue;
  if (clickElement(button)) return true;
}
return false;
""",
            label_prefix,
        )
    )


def _phase_3_2_ship_date_calendar_day_selected(driver: Any, target_date: date) -> bool:
    label_prefix = f"{target_date.year}年{target_date.month}月{target_date.day}日"
    return bool(
        _execute_phase_3_2_ship_date_script(
            driver,
            """
const labelPrefix = String(arguments[2] || '').trim();
const calendar = getScopedCalendar();
if (!calendar) return false;
for (const button of scopedQuerySelectorAll('button[aria-label]', calendar)) {
  const label = String(button.getAttribute('aria-label') || '').trim();
  if (!labelPrefix || !label.startsWith(labelPrefix)) continue;
  const cell = deepClosest(button, 'td') || button.parentElement;
  if (button.getAttribute('aria-pressed') === 'true') return true;
  if (button.classList && button.classList.contains('selected')) return true;
  if (cell && cell.classList && cell.classList.contains('selected')) return true;
}
return false;
""",
            label_prefix,
        )
    )


def _parse_calendar_month_label(label: str) -> tuple[int, int]:
    text = str(label or "").replace(" ", "").strip()
    if not text:
        raise RuntimeError("日历月份标题为空")

    year_text = ""
    for char in text:
        if char.isdigit():
            year_text += char
    if len(year_text) < 4:
        raise RuntimeError(f"无法解析日历年份: {label}")
    year = int(year_text[:4])

    month_text = text.replace(str(year), "")
    month_text = month_text.strip()
    if month_text in _MONTH_NAME_TO_NUMBER:
        return year, _MONTH_NAME_TO_NUMBER[month_text]
    if month_text.endswith("月"):
        numeric_text = month_text[:-1]
        if numeric_text.isdigit():
            month = int(numeric_text)
            if 1 <= month <= 12:
                return year, month
    raise RuntimeError(f"无法解析日历月份: {label}")


def normalize_transport_mode(raw_mode: str) -> dict[str, Any]:
    raw_text = str(raw_mode or "").strip()
    normalized = raw_text.upper().replace(" ", "")
    if raw_text in ("空运",) or normalized == "AIR":
        return {"canonical": "AIR", "ui_text": "空运", "offset_days": 14}
    if raw_text in ("海运",) or normalized == "OCEAN":
        return {"canonical": "OCEAN", "ui_text": "海运", "offset_days": 30}
    if raw_text in ("低价商城", "陆运") or normalized == "GROUND":
        return {"canonical": "GROUND", "ui_text": "陆运", "offset_days": 1}
    if raw_text in ("DHL/UPS快递（空运速派）", "DHL/UPS快递(空运速派)") or normalized in (
        "EXPRESS_AIR",
        "DHL_UPS",
    ):
        return {"canonical": "EXPRESS_AIR", "ui_text": "空运", "offset_days": 7}
    raise RuntimeError(f"不支持的 transport_mode: {raw_text}")


def calculate_pickup_date(raw_mode: str, *, today: date | None = None) -> date:
    mode = normalize_transport_mode(raw_mode)
    base_date = today or date.today()
    return base_date + timedelta(days=int(mode["offset_days"]))


def _wait_for_condition(
    step_name: str,
    checker,
    *,
    timeout_seconds: int,
    interval_seconds: float = 0.5,
    driver: Any | None = None,
) -> None:
    deadline = time.time() + max(1, int(timeout_seconds or 0))
    while time.time() < deadline:
        _raise_if_returned_to_step2_start(driver)
        if checker():
            return
        time.sleep(max(0.2, float(interval_seconds or 0.5)))
    _raise_if_returned_to_step2_start(driver)
    raise RuntimeError(f"等待{step_name}超时")


def _wait_for_click(
    step_name: str,
    clicker,
    *,
    timeout_seconds: int,
    interval_seconds: float = 0.5,
    driver: Any | None = None,
) -> None:
    deadline = time.time() + max(1, int(timeout_seconds or 0))
    while time.time() < deadline:
        _raise_if_returned_to_step2_start(driver)
        if clicker():
            return
        time.sleep(max(0.2, float(interval_seconds or 0.5)))
    _raise_if_returned_to_step2_start(driver)
    raise RuntimeError(f"等待{step_name}超时")


def _wait_phase_3_2_refresh_after_selection() -> None:
    time.sleep(1.0)


def _wait_phase_3_2_after_carrier_before_dates() -> None:
    time.sleep(5.0)


def _select_pickup_date(driver: Any, target_date: date, *, timeout_seconds: int = 60) -> None:
    _raise_if_returned_to_step2_start(driver)
    if not _calendar_visible(driver):
        _wait_for_click("日期输入框", lambda: _click_date_input(driver), timeout_seconds=timeout_seconds, driver=driver)
        _wait_for_condition("日历出现", lambda: _calendar_visible(driver), timeout_seconds=timeout_seconds, driver=driver)

    deadline = time.time() + max(1, int(timeout_seconds or 0))
    clicked_target = False
    while time.time() < deadline:
        _raise_if_returned_to_step2_start(driver)
        if _calendar_day_selected(driver, target_date):
            return
        if _click_calendar_day(driver, target_date):
            clicked_target = True
            time.sleep(0.2)
            if _calendar_day_selected(driver, target_date):
                return
            continue
        current_year, current_month = _parse_calendar_month_label(_read_calendar_month_label(driver))
        current_tuple = (current_year, current_month)
        target_tuple = (target_date.year, target_date.month)
        if current_tuple < target_tuple:
            if not _click_calendar_next_month(driver):
                raise RuntimeError("日历切换到下一月失败")
        elif current_tuple > target_tuple:
            if not _click_calendar_prev_month(driver):
                raise RuntimeError("日历切换到上一月失败")
        else:
            time.sleep(0.5)
            continue
        time.sleep(0.5)
    _raise_if_returned_to_step2_start(driver)
    if clicked_target:
        raise RuntimeError(f"等待目标日期选中超时: {target_date.isoformat()}")
    raise RuntimeError(f"等待目标日期出现或可点击超时: {target_date.isoformat()}")


def _select_phase_3_2_pickup_date(driver: Any, target_date: date, *, timeout_seconds: int = 60) -> None:
    _raise_if_returned_to_step2_start(driver)
    if not _phase_3_2_calendar_visible(driver):
        _wait_for_click("计划送达日期输入框", lambda: _click_phase_3_2_date_input(driver), timeout_seconds=timeout_seconds, driver=driver)
        _wait_for_condition("送达时段日历出现", lambda: _phase_3_2_calendar_visible(driver), timeout_seconds=timeout_seconds, driver=driver)

    deadline = time.time() + max(1, int(timeout_seconds or 0))
    clicked_target = False
    while time.time() < deadline:
        _raise_if_returned_to_step2_start(driver)
        if _phase_3_2_calendar_day_selected(driver, target_date):
            return
        if _click_phase_3_2_calendar_day(driver, target_date):
            clicked_target = True
            time.sleep(0.2)
            if _phase_3_2_calendar_day_selected(driver, target_date):
                return
            continue
        current_year, current_month = _parse_calendar_month_label(_read_phase_3_2_calendar_month_label(driver))
        current_tuple = (current_year, current_month)
        target_tuple = (target_date.year, target_date.month)
        if current_tuple < target_tuple:
            if not _click_phase_3_2_calendar_next_month(driver):
                raise RuntimeError("送达时段日历切换到下一月失败")
        elif current_tuple > target_tuple:
            if not _click_phase_3_2_calendar_prev_month(driver):
                raise RuntimeError("送达时段日历切换到上一月失败")
        else:
            time.sleep(0.5)
            continue
        time.sleep(0.5)
    _raise_if_returned_to_step2_start(driver)
    if clicked_target:
        raise RuntimeError(f"等待送达时段目标日期选中超时: {target_date.isoformat()}")
    raise RuntimeError(f"等待送达时段目标日期出现或可点击超时: {target_date.isoformat()}")


def _select_phase_3_2_ship_date(driver: Any, ship_date: date, *, timeout_seconds: int = 60) -> None:
    _raise_if_returned_to_step2_start(driver)
    if _phase_3_2_ship_date_value_matches(driver, ship_date):
        return
    if not _phase_3_2_ship_date_calendar_visible(driver):
        _wait_for_click("发货日期输入框", lambda: _click_phase_3_2_ship_date_input(driver), timeout_seconds=timeout_seconds, driver=driver)
        _wait_for_condition("发货日期日历出现", lambda: _phase_3_2_ship_date_calendar_visible(driver), timeout_seconds=timeout_seconds, driver=driver)

    deadline = time.time() + max(1, int(timeout_seconds or 0))
    clicked_target = False
    while time.time() < deadline:
        _raise_if_returned_to_step2_start(driver)
        if _phase_3_2_ship_date_value_matches(driver, ship_date):
            return
        if _phase_3_2_ship_date_calendar_day_selected(driver, ship_date):
            return
        if _click_phase_3_2_ship_date_calendar_day(driver, ship_date):
            clicked_target = True
            time.sleep(0.2)
            if _phase_3_2_ship_date_calendar_day_selected(driver, ship_date):
                return
            continue
        current_year, current_month = _parse_calendar_month_label(_read_phase_3_2_ship_date_calendar_month_label(driver))
        current_tuple = (current_year, current_month)
        target_tuple = (ship_date.year, ship_date.month)
        if current_tuple < target_tuple:
            if not _click_phase_3_2_ship_date_calendar_next_month(driver):
                raise RuntimeError("发货日期日历切换到下一月失败")
        elif current_tuple > target_tuple:
            if not _click_phase_3_2_ship_date_calendar_prev_month(driver):
                raise RuntimeError("发货日期日历切换到上一月失败")
        else:
            time.sleep(0.5)
            continue
        time.sleep(0.5)
    _raise_if_returned_to_step2_start(driver)
    if clicked_target:
        raise RuntimeError(f"等待发货日期目标日期选中超时: {ship_date.isoformat()}")
    raise RuntimeError(f"等待发货日期目标日期出现或可点击超时: {ship_date.isoformat()}")


def _select_pickup_date_for_layout(
    driver: Any,
    target_date: date,
    *,
    layout: str,
    timeout_seconds: int,
) -> None:
    safe_timeout = min(timeout_seconds, 60)
    if layout == _LAYOUT_PHASE_3_1:
        _wait_for_click(
            "非亚马逊合作承运人选项",
            lambda: _select_non_amazon_partner_carrier_type(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        _wait_for_click(
            "计划送达日期输入框",
            lambda: _click_phase_3_1_date_input(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        _wait_for_condition("日历出现", lambda: _calendar_visible(driver), timeout_seconds=safe_timeout, driver=driver)
        _select_pickup_date(driver, target_date, timeout_seconds=safe_timeout)
        _wait_for_click(
            "日期更新按钮",
            lambda: _click_delivery_window_modal_confirm_button(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        return

    if layout == _LAYOUT_PHASE_3_2:
        _wait_for_click(
            "计划送达日期输入框",
            lambda: _click_phase_3_2_date_input(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        _wait_for_condition("送达时段日历出现", lambda: _phase_3_2_calendar_visible(driver), timeout_seconds=safe_timeout, driver=driver)
        _select_phase_3_2_pickup_date(driver, target_date, timeout_seconds=safe_timeout)
        _wait_phase_3_2_refresh_after_selection()
        return

    _wait_for_condition(
        "日期选择器出现",
        lambda: _date_picker_visible(driver),
        timeout_seconds=safe_timeout,
        driver=driver,
    )
    _select_pickup_date(driver, target_date, timeout_seconds=safe_timeout)
    _wait_for_click("更新按钮", lambda: _click_update_button(driver), timeout_seconds=min(timeout_seconds, 10), driver=driver)


def _prepare_phase_3_2_own_carrier_entry(driver: Any, *, timeout_seconds: int) -> None:
    _wait_for_click(
        "非合作承运人入口",
        lambda: _click_first_matching(driver, (_CROSS_BORDER_NON_PCP_BOX_SELECTOR,)),
        timeout_seconds=min(timeout_seconds, 10),
        driver=driver,
    )
    _wait_for_click(
        "非合作承运人确认弹窗",
        lambda: _confirm_modal_if_visible(driver),
        timeout_seconds=min(timeout_seconds, 10),
        driver=driver,
    )


def _wait_for_dropdown_value(
    driver: Any,
    *,
    dropdown_selector: str,
    expected_value: str,
    label: str,
    timeout_seconds: int,
) -> None:
    deadline = time.time() + max(1, int(timeout_seconds or 0))
    while time.time() < deadline:
        _raise_if_returned_to_step2_start(driver)
        if _dropdown_value_equals(driver, dropdown_selector, expected_value):
            return
        time.sleep(0.2)
    _raise_if_returned_to_step2_start(driver)
    current_value = _read_dropdown_value(driver, dropdown_selector)
    raise RuntimeError(f"{label}未选择为 {expected_value}，当前值: {current_value or '空'}")


def _select_phase_3_2_dropdown_value(
    driver: Any,
    *,
    dropdown_selector: str,
    option_value: str,
    option_text: str,
    step_name: str,
    value_label: str,
    timeout_seconds: int,
) -> None:
    safe_timeout = min(timeout_seconds, 10)
    _wait_for_click(
        f"{step_name}选项",
        lambda: _select_dropdown_option_by_value(
            driver,
            dropdown_selector,
            option_value,
            option_text=option_text,
        ),
        timeout_seconds=safe_timeout,
        driver=driver,
    )
    _wait_for_dropdown_value(
        driver,
        dropdown_selector=dropdown_selector,
        expected_value=option_value,
        label=value_label,
        timeout_seconds=safe_timeout,
    )
    _wait_phase_3_2_refresh_after_selection()


def _select_carrier_mode_for_layout(
    driver: Any,
    mode: dict[str, Any],
    *,
    layout: str,
    timeout_seconds: int,
) -> None:
    if layout == _LAYOUT_PHASE_3_1:
        _wait_for_click(
            "非合作承运人下拉框",
            lambda: _open_phase_3_1_carrier_dropdown(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        _wait_for_click("其他承运人选项", lambda: _select_other_carrier(driver), timeout_seconds=min(timeout_seconds, 10), driver=driver)
        return

    if layout == _LAYOUT_PHASE_3_2:
        transport_value = _phase_3_2_transport_target_value(mode)
        for _index in range(2):
            _select_phase_3_2_dropdown_value(
                driver,
                dropdown_selector=_TRANSPORTATION_MODE_DROPDOWN_SELECTOR,
                option_value=transport_value,
                option_text=str(mode["ui_text"]),
                step_name=f"{mode['ui_text']}运输方式",
                value_label="运输方式",
                timeout_seconds=timeout_seconds,
            )

            _select_phase_3_2_dropdown_value(
                driver,
                dropdown_selector=_NON_PCP_CARRIER_CHOICES_SELECTOR,
                option_value="OTHER",
                option_text="其他",
                step_name="非合作承运人",
                value_label="非合作承运人",
                timeout_seconds=timeout_seconds,
            )
        return

    _wait_for_click("承运人下拉框", lambda: _open_carrier_dropdown(driver), timeout_seconds=min(timeout_seconds, 10), driver=driver)
    _wait_for_click("其他承运人选项", lambda: _select_other_carrier(driver), timeout_seconds=min(timeout_seconds, 10), driver=driver)
    _wait_for_click(
        "运输方式下拉框",
        lambda: _open_transport_mode_dropdown(driver),
        timeout_seconds=min(timeout_seconds, 10),
        driver=driver,
    )
    _wait_for_click(
        f"{mode['ui_text']}运输方式选项",
        lambda: _select_transport_mode_option(driver, str(mode["ui_text"])),
        timeout_seconds=min(timeout_seconds, 10),
        driver=driver,
    )


def probe_own_carrier_ready(session: Any, *, timeout_seconds: int = 10) -> dict[str, Any]:
    deadline = time.time() + max(1, min(int(timeout_seconds or 0), 10))
    while time.time() < deadline:
        if _delivery_window_link_visible(session.driver):
            return {"ready": True, "notice": "已进入日期选择器步骤"}
        layout = _detect_own_carrier_layout(session.driver)
        if layout in {_LAYOUT_PHASE_3_1, _LAYOUT_PHASE_3_2}:
            return {"ready": True, "notice": f"已识别自己的承运人页面布局: {layout}"}
        if _amazon_operations_center_selected(session.driver):
            return {"ready": True, "notice": _AMAZON_OPERATIONS_CENTER_READY_NOTICE}
        time.sleep(0.5)
    return {"ready": False, "notice": ""}


def confirm_own_carrier_shipment(
    session: Any,
    transport_mode: str,
    *,
    consignment_no: str = "",
    timeout_seconds: int = 60,
    today: date | None = None,
) -> dict[str, str]:
    mode = normalize_transport_mode(transport_mode)
    ship_date = today or date.today()
    target_date = calculate_pickup_date(transport_mode, today=ship_date)
    driver = session.driver
    shipment_summary_excel_path = ""

    try:
        layout = _detect_own_carrier_layout(driver)
        if layout == _LAYOUT_LEGACY and not _delivery_window_link_visible(driver):
            _wait_for_click(
                "自己的承运人入口",
                lambda: _click_own_carrier_entry(driver),
                timeout_seconds=min(timeout_seconds, 10),
                driver=driver,
            )
            _wait_for_condition("弹窗出现", lambda: _dialog_visible(driver), timeout_seconds=min(timeout_seconds, 10), driver=driver)
            _wait_for_click(
                "弹窗 Confirm 按钮",
                lambda: _click_confirm_modal(driver),
                timeout_seconds=min(timeout_seconds, 10),
                driver=driver,
            )
            layout = _detect_own_carrier_layout(driver)
        if layout == _LAYOUT_PHASE_3_2:
            _prepare_phase_3_2_own_carrier_entry(driver, timeout_seconds=timeout_seconds)
            _select_carrier_mode_for_layout(driver, mode, layout=layout, timeout_seconds=timeout_seconds)
            _wait_phase_3_2_after_carrier_before_dates()
            _select_phase_3_2_ship_date(driver, ship_date, timeout_seconds=timeout_seconds)
            _wait_phase_3_2_refresh_after_selection()
            _select_pickup_date_for_layout(driver, target_date, layout=layout, timeout_seconds=timeout_seconds)
        else:
            _select_pickup_date_for_layout(driver, target_date, layout=layout, timeout_seconds=timeout_seconds)
            _select_carrier_mode_for_layout(driver, mode, layout=layout, timeout_seconds=timeout_seconds)
        _wait_for_click(
            "接受费用并确认发货按钮",
            lambda: _click_accept_fees_button(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        _wait_for_condition(
            _CONTINUE_INPUT_TRACKING_TEXT,
            lambda: _continue_input_tracking_visible(driver),
            timeout_seconds=max(30, int(timeout_seconds or 0)),
            interval_seconds=1.0,
            driver=driver,
        )
        shipment_summaries = _collect_shipment_summaries(driver)
        shipment_summary_excel_path = _write_shipment_summary_excel(
            Path(getattr(session, "output_dir", Path.cwd())),
            shipment_summaries,
            consignment_no=consignment_no,
        )
        _wait_for_click(
            _CONTINUE_INPUT_TRACKING_TEXT,
            lambda: _click_continue_input_tracking_button(driver),
            timeout_seconds=min(timeout_seconds, 10),
            driver=driver,
        )
        _wait_for_condition(
            _TRACKING_INPUT_TEXT,
            lambda: _tracking_input_visible(driver),
            timeout_seconds=60,
            interval_seconds=1.0,
            driver=driver,
        )
        return {
            "notice": _FINAL_SUCCESS_NOTICE,
            "shipment_summary_excel_path": shipment_summary_excel_path,
        }
    except _ReturnedToStep2Start:
        payload = {"notice": _RETURN_TO_STEP2_NOTICE}
        if shipment_summary_excel_path:
            payload["shipment_summary_excel_path"] = shipment_summary_excel_path
        return payload
    except Exception as exc:
        raise _OwnCarrierWorkflowError(
            str(exc).strip() or exc.__class__.__name__,
            shipment_summary_excel_path=shipment_summary_excel_path,
        ) from exc


__all__ = [
    "calculate_pickup_date",
    "confirm_own_carrier_shipment",
    "normalize_transport_mode",
    "probe_own_carrier_ready",
]
