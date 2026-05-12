from __future__ import annotations

import argparse
import asyncio
import json
import re
import shutil
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_output import configure_utf8_stdio
from services.agent_cli.mabang.summarize_fba_delivery_tax_sku import (
    EXPORT_TAX_PRODUCTS_PATH,
    EXPORT_TAX_PRODUCTS_SHEET,
    TAX_PRODUCT_NAME_COLUMN,
    TAX_PRODUCT_SKU_COLUMN,
)
from services.agent_cli.mabang.validate_export_tax_products import validate_export_tax_products
from services.mabang.stock_sku_export import export_stock_sku_names
from shared.infra.net import close_all_network_clients

SOURCE = "export_tax_products_import"
DEFAULT_BACKUP_DIR = Path("artifacts") / "export_tax_products_backup"
SKU_TOKEN_SPLIT_PATTERN = re.compile(r"[\s,，;；]+")
WHITESPACE_PATTERN = re.compile(r"\s+")


class JsonArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise ValueError(str(message or "").strip() or "参数解析失败")


def _write_json(payload: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(dict(payload or {}), ensure_ascii=False) + "\n")
    sys.stdout.flush()


def _exception_text(exc: Exception) -> str:
    message = str(exc or "").strip()
    return message or exc.__class__.__name__


def _clean_cell(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() == "nan":
        return ""
    return text


def _sku_match_key(value: Any) -> str:
    return WHITESPACE_PATTERN.sub("", _clean_cell(value))


def normalize_input_skus(values: list[str] | tuple[str, ...] | None) -> list[str]:
    unique: OrderedDict[str, str] = OrderedDict()
    for raw_value in values or []:
        for token in SKU_TOKEN_SPLIT_PATTERN.split(str(raw_value or "")):
            sku = _clean_cell(token)
            key = _sku_match_key(sku)
            if not key or key in unique:
                continue
            unique[key] = key
    return list(unique.values())


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _backup_products_file(products_path: Path, backup_dir: str | Path | None = None) -> Path:
    directory = Path(DEFAULT_BACKUP_DIR if backup_dir is None else backup_dir)
    directory.mkdir(parents=True, exist_ok=True)
    stem = products_path.stem or "export_tax_products"
    suffix = products_path.suffix or ".xlsx"
    target = directory / f"{stem}_{_timestamp()}{suffix}"
    index = 2
    while target.exists():
        target = directory / f"{stem}_{_timestamp()}_{index}{suffix}"
        index += 1
    shutil.copy2(products_path, target)
    return target


def _load_products_workbook(products_path: Path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入出口退税产品表") from exc

    try:
        workbook = load_workbook(products_path)
    except Exception as exc:
        raise RuntimeError(f"读取出口退税产品表失败: {products_path}, error={exc}") from exc

    if EXPORT_TAX_PRODUCTS_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"出口退税产品表缺少 sheet: {EXPORT_TAX_PRODUCTS_SHEET}")

    worksheet = workbook[EXPORT_TAX_PRODUCTS_SHEET]
    headers = [_clean_cell(cell.value) for cell in worksheet[1]]
    missing_columns = [
        column
        for column in (TAX_PRODUCT_SKU_COLUMN, TAX_PRODUCT_NAME_COLUMN)
        if column not in headers
    ]
    if missing_columns:
        raise RuntimeError(f"出口退税产品表缺少列: {', '.join(missing_columns)}")

    sku_col = headers.index(TAX_PRODUCT_SKU_COLUMN) + 1
    name_col = headers.index(TAX_PRODUCT_NAME_COLUMN) + 1
    return workbook, worksheet, sku_col, name_col


def _existing_sku_keys(worksheet: Any, *, sku_col: int) -> set[str]:
    keys: set[str] = set()
    for row_index in range(2, worksheet.max_row + 1):
        key = _sku_match_key(worksheet.cell(row=row_index, column=sku_col).value)
        if key:
            keys.add(key)
    return keys


def _append_product_row(worksheet: Any, *, sku_col: int, name_col: int, sku: str, product_name: str) -> None:
    row_index = worksheet.max_row + 1
    worksheet.cell(row=row_index, column=sku_col, value=sku)
    worksheet.cell(row=row_index, column=name_col, value=product_name)


async def import_export_tax_products(
    skus: list[str] | tuple[str, ...],
    *,
    products_path: str | Path | None = None,
    backup_dir: str | Path | None = None,
) -> dict[str, Any]:
    requested_skus = normalize_input_skus(skus)
    if not requested_skus:
        raise ValueError("sku 不能为空")

    source_path = Path(EXPORT_TAX_PRODUCTS_PATH if products_path is None else products_path)
    validate_export_tax_products(source_path)
    workbook, worksheet, sku_col, name_col = _load_products_workbook(source_path)

    existing_keys = _existing_sku_keys(worksheet, sku_col=sku_col)
    skipped_duplicate_skus: list[str] = []
    lookup_skus: list[str] = []
    for sku in requested_skus:
        key = _sku_match_key(sku)
        if key in existing_keys:
            skipped_duplicate_skus.append(sku)
            continue
        lookup_skus.append(sku)

    stock_names_by_key: dict[str, str] = {}
    if lookup_skus:
        stock_result = await export_stock_sku_names(lookup_skus, delivery_no="export_tax_products_import")
        stock_names_by_key = dict(getattr(stock_result, "names_by_key", {}) or {})

    imported_skus: list[str] = []
    skipped_not_found_skus: list[str] = []
    imported_rows: list[tuple[str, str]] = []
    for sku in lookup_skus:
        product_name = _clean_cell(stock_names_by_key.get(_sku_match_key(sku)))
        if not product_name:
            skipped_not_found_skus.append(sku)
            continue
        imported_rows.append((sku, product_name))

    backup_path = ""
    if imported_rows:
        backup = _backup_products_file(source_path, backup_dir=backup_dir)
        backup_path = str(backup)
        for sku, product_name in imported_rows:
            _append_product_row(
                worksheet,
                sku_col=sku_col,
                name_col=name_col,
                sku=sku,
                product_name=product_name,
            )
            imported_skus.append(sku)
        workbook.save(source_path)
        try:
            validate_export_tax_products(source_path)
        except Exception as exc:
            shutil.copy2(backup, source_path)
            raise RuntimeError(f"导入后校验失败，已恢复备份: {backup}, error={exc}") from exc

    return {
        "success": True,
        "requested_sku_count": len(requested_skus),
        "imported_count": len(imported_skus),
        "skipped_duplicate_count": len(skipped_duplicate_skus),
        "skipped_not_found_count": len(skipped_not_found_skus),
        "imported_skus": imported_skus,
        "skipped_duplicate_skus": skipped_duplicate_skus,
        "skipped_not_found_skus": skipped_not_found_skus,
        "products_path": str(source_path),
        "backup_path": backup_path,
        "source": SOURCE,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = JsonArgumentParser(
        prog="python -m services.agent_cli.mabang.import_export_tax_products"
    )
    parser.add_argument("--sku", action="append", default=[])
    parser.add_argument("--products-path", default=str(EXPORT_TAX_PRODUCTS_PATH))
    parser.add_argument("--backup-dir", default=str(DEFAULT_BACKUP_DIR))
    return parser


async def _run_async(args: argparse.Namespace) -> dict[str, Any]:
    return await import_export_tax_products(
        list(getattr(args, "sku", []) or []),
        products_path=getattr(args, "products_path", ""),
        backup_dir=getattr(args, "backup_dir", ""),
    )


def main(argv: list[str] | None = None) -> int:
    configure_utf8_stdio()
    try:
        args = build_parser().parse_args(argv)
        payload = asyncio.run(_run_async(args))
    except Exception as exc:
        payload = {
            "success": False,
            "exception": _exception_text(exc),
        }
    finally:
        try:
            asyncio.run(close_all_network_clients())
        except Exception:
            pass

    _write_json(payload)
    return 0 if bool(payload.get("success")) else 1


if __name__ == "__main__":
    raise SystemExit(main())
