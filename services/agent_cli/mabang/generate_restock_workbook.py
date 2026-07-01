from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from collections import OrderedDict
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_output import configure_utf8_stdio
from services.agent_cli.mabang.summarize_fba_delivery_tax_sku import (
    DELIVERY_CSV_DIR,
    ITEM_SPLIT_PATTERN,
    SKU_SHIP_QTY_COLUMN,
    _clean_cell,
    _decimal_to_cell_value,
    _parse_sku_quantity_item,
    _require_delivery_no,
    _read_delivery_rows,
    _sku_match_key,
    find_latest_delivery_csv,
)
from shared.infra.net import close_all_network_clients

OUTPUT_DIR = Path("artifacts") / "mabang_purchase_summary"
SOURCE = "fba_purchase_summary"

MASTER_REQUIRED_HEADERS = ("库存sku", "产品名称", "型号", "原价", "厂家", "备用厂家")
MASTER_HEADER_ALIASES = {
    "库存SKU": "库存sku",
}
MASTER_SKU_SHEET_NAME = "SKU表"
CONTRACT_SHEET_NAME = "供应商合同信息"
CONTRACT_REQUIRED_HEADERS = ("供货方", "单位", "合同产品名称")
CONTRACT_PREFIX_HEADER = "合同编号前缀"
CONTRACT_TAX_RATE_HEADER = "税率"
MANUFACTURER_COLUMNS = (
    "库存sku",
    "产品名称",
    "来源SP单号",
    "库存sku（第一行）",
    "产品名称（第一行）",
    "型号",
    "原价",
    "厂家",
    "单位",
    "合同产品名称",
    "合同编号前缀",
    "税率",
    "数量",
    "总价",
)
UNMATCHED_COLUMNS = ("库存sku", "来源SP单号", "数量", "问题说明")
SUMMARY_SHEET_NAME = "采购汇总"
UNMATCHED_SHEET_NAME = "未匹配"
EMPTY_MANUFACTURER_SHEET_NAME = "未填写厂家"
INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
FLOAT_NOISE_TOLERANCE = Decimal("0.000000001")


class MasterProducts(OrderedDict[str, dict[str, Any]]):
    def __init__(self) -> None:
        super().__init__()
        self.deduped_duplicate_sku_count = 0
        self.deduped_duplicate_row_count = 0
        self.deduped_duplicate_sku_examples: list[str] = []
        self.skipped_empty_sku_row_count = 0
        self.skipped_empty_sku_rows: list[int] = []
        self.unmerged_empty_model_sku_count = 0
        self.unmerged_empty_model_skus: list[str] = []
        self.contracts_by_manufacturer: dict[str, dict[str, str]] = {}
        self.contract_lookup_enabled = False
        self.contract_mapping_count = 0
        self.contract_unmapped_manufacturer_count = 0
        self.contract_unmapped_manufacturer_examples: list[str] = []
        self.contract_conflict_manufacturer_count = 0
        self.contract_conflict_manufacturer_examples: list[str] = []
        self.contract_prefix_conflict_manufacturer_count = 0
        self.contract_prefix_conflict_manufacturer_examples: list[str] = []
        self.contract_tax_rate_conflict_manufacturer_count = 0
        self.contract_tax_rate_conflict_manufacturer_examples: list[str] = []
        self._contract_sheet_warning = ""
        self._contract_prefix_header_warning = ""
        self._contract_tax_rate_header_warning = ""
        self._contract_unmapped_manufacturers: OrderedDict[str, None] = OrderedDict()
        self._contract_conflict_manufacturers: set[str] = set()
        self._contract_prefix_conflict_manufacturers: OrderedDict[str, None] = OrderedDict()
        self._contract_tax_rate_conflict_manufacturers: OrderedDict[str, None] = OrderedDict()
        self.warnings: list[str] = []


class JsonArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise ValueError(str(message or "").strip() or "参数解析失败")


def _write_json(payload: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(dict(payload or {}), ensure_ascii=False) + "\n")
    sys.stdout.flush()


def _exception_text(exc: Exception) -> str:
    message = str(exc or "").strip()
    return message or exc.__class__.__name__


def _decimal_from_cell(value: Any, *, field_name: str, row_number: int, source_path: Path) -> Decimal:
    text = _clean_cell(value)
    if not text:
        raise RuntimeError(f"出口退税总表 {source_path.name} 第{row_number}行 {field_name} 不能为空")
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise RuntimeError(
            f"出口退税总表 {source_path.name} 第{row_number}行 {field_name}非数字: {text}"
        ) from exc
    if not result.is_finite():
        raise RuntimeError(f"出口退税总表 {source_path.name} 第{row_number}行 {field_name}非数字: {text}")
    return result


def _header_indexes(
    header_values: tuple[Any, ...],
    *,
    source_path: Path,
    required_headers: tuple[str, ...] = MASTER_REQUIRED_HEADERS,
    header_aliases: dict[str, str] | None = MASTER_HEADER_ALIASES,
    sheet_label: str = "",
) -> dict[str, int]:
    indexes: dict[str, int] = {}
    duplicates: list[str] = []
    aliases = dict(header_aliases or {})
    for column_index, raw_header in enumerate(header_values, start=1):
        raw_name = _clean_cell(raw_header)
        if not raw_name:
            continue
        header = aliases.get(raw_name, raw_name)
        if header in indexes:
            duplicates.append(header)
            continue
        indexes[header] = column_index

    location = f" {sheet_label}" if sheet_label else ""
    missing = [header for header in required_headers if header not in indexes]
    if missing:
        raise RuntimeError(f"出口退税总表 {source_path.name}{location} 缺少必需列: {', '.join(missing)}")
    if duplicates:
        duplicate_text = ", ".join(dict.fromkeys(duplicates))
        raise RuntimeError(f"出口退税总表 {source_path.name}{location} 第1行表头重复: {duplicate_text}")
    return indexes


def _master_row_record(
    row_values: dict[str, Any],
    *,
    row_number: int,
    source_path: Path,
) -> tuple[dict[str, Any], tuple[Any, ...]]:
    sku = _clean_cell(row_values["库存sku"])
    if not sku:
        raise RuntimeError(f"出口退税总表 {source_path.name} 第{row_number}行 库存sku 不能为空")

    original_price = _decimal_from_cell(
        row_values["原价"],
        field_name="原价",
        row_number=row_number,
        source_path=source_path,
    )
    product_name = _clean_cell(row_values["产品名称"])
    model = _clean_cell(row_values["型号"])
    manufacturer = _clean_cell(row_values["厂家"])
    backup_manufacturer = _clean_cell(row_values["备用厂家"])

    return (
        {
            "stock_sku": sku,
            "product_name": product_name,
            "model": model,
            "original_price": original_price,
            "manufacturer": manufacturer,
        },
        (sku, product_name, model, original_price, manufacturer, backup_manufacturer),
    )


def _append_duplicate_warning(products: MasterProducts, duplicate_keys: set[str]) -> None:
    products.deduped_duplicate_sku_count = len(duplicate_keys)
    if products.deduped_duplicate_row_count <= 0:
        return

    products.warnings.append(
        "出口退税总表存在完全相同的重复库存sku，已自动去重: "
        f"sku_count={products.deduped_duplicate_sku_count}, "
        f"row_count={products.deduped_duplicate_row_count}, "
        f"examples={', '.join(products.deduped_duplicate_sku_examples)}"
    )


def _append_empty_sku_warning(products: MasterProducts) -> None:
    if products.skipped_empty_sku_row_count <= 0:
        return
    rows = ", ".join(str(row_number) for row_number in products.skipped_empty_sku_rows)
    products.warnings.append(
        "出口退税总表存在库存sku为空的行，已忽略: "
        f"count={products.skipped_empty_sku_row_count}, rows={rows}"
    )


def _append_empty_model_warning(products: MasterProducts) -> None:
    if products.unmerged_empty_model_sku_count <= 0:
        return
    examples = ", ".join(products.unmerged_empty_model_skus)
    products.warnings.append(
        "出口退税总表存在型号为空的库存sku，已按 SKU 粒度保留不合并: "
        f"count={products.unmerged_empty_model_sku_count}, examples={examples}"
    )


def _append_contract_warnings(products: MasterProducts) -> None:
    if products._contract_sheet_warning:
        products.warnings.append(products._contract_sheet_warning)
    if products.contract_conflict_manufacturer_count > 0:
        products.warnings.append(
            f"出口退税总表 {CONTRACT_SHEET_NAME} sheet 存在同一供货方对应不同单位或合同产品名称，"
            "相关厂家字段已留空: "
            f"count={products.contract_conflict_manufacturer_count}, "
            f"examples={', '.join(products.contract_conflict_manufacturer_examples)}"
        )
    if products._contract_prefix_header_warning:
        products.warnings.append(products._contract_prefix_header_warning)
    if products.contract_prefix_conflict_manufacturer_count > 0:
        products.warnings.append(
            f"出口退税总表 {CONTRACT_SHEET_NAME} sheet 存在同一供货方对应不同合同编号前缀，"
            "合同编号前缀已留空: "
            f"count={products.contract_prefix_conflict_manufacturer_count}, "
            f"examples={', '.join(products.contract_prefix_conflict_manufacturer_examples)}"
        )
    if products._contract_tax_rate_header_warning:
        products.warnings.append(products._contract_tax_rate_header_warning)
    if products.contract_tax_rate_conflict_manufacturer_count > 0:
        products.warnings.append(
            f"出口退税总表 {CONTRACT_SHEET_NAME} sheet 存在同一供货方对应不同税率，"
            "税率已留空: "
            f"count={products.contract_tax_rate_conflict_manufacturer_count}, "
            f"examples={', '.join(products.contract_tax_rate_conflict_manufacturer_examples)}"
        )


def _append_contract_unmapped_warning(products: MasterProducts) -> None:
    products.contract_unmapped_manufacturer_count = len(products._contract_unmapped_manufacturers)
    products.contract_unmapped_manufacturer_examples = list(products._contract_unmapped_manufacturers)[:20]
    if products.contract_unmapped_manufacturer_count <= 0:
        return
    products.warnings.append(
        f"出口退税总表 {CONTRACT_SHEET_NAME} sheet 未找到部分厂家对应的供货方映射，"
        "单位和合同产品名称已留空: "
        f"count={products.contract_unmapped_manufacturer_count}, "
        f"examples={', '.join(products.contract_unmapped_manufacturer_examples)}"
    )


def _contract_fields_for_manufacturer(products: MasterProducts, manufacturer: str) -> tuple[str, str, str, str]:
    if manufacturer in products._contract_conflict_manufacturers:
        return "", "", "", ""
    contract = products.contracts_by_manufacturer.get(manufacturer)
    if contract is not None:
        return (
            contract["unit"],
            contract["contract_product_name"],
            contract.get("contract_no_prefix", ""),
            contract.get("tax_rate", ""),
        )
    if products.contract_lookup_enabled:
        products._contract_unmapped_manufacturers.setdefault(manufacturer, None)
    return "", "", "", ""


def _load_contract_mappings(workbook: Any, *, source_path: Path, products: MasterProducts) -> None:
    if CONTRACT_SHEET_NAME not in workbook.sheetnames:
        products._contract_sheet_warning = (
            f"出口退税总表缺少 sheet: {CONTRACT_SHEET_NAME}，单位和合同产品名称将留空"
        )
        return

    worksheet = workbook[CONTRACT_SHEET_NAME]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration:
        header_values = ()
    try:
        indexes = _header_indexes(
            header_values,
            source_path=source_path,
            required_headers=CONTRACT_REQUIRED_HEADERS,
            header_aliases={},
            sheet_label=f"{CONTRACT_SHEET_NAME} sheet",
        )
    except RuntimeError as exc:
        products._contract_sheet_warning = f"{_exception_text(exc)}，单位和合同产品名称将留空"
        return

    products.contract_lookup_enabled = True
    prefix_column_index = indexes.get(CONTRACT_PREFIX_HEADER)
    if prefix_column_index is None:
        products._contract_prefix_header_warning = (
            f"出口退税总表 {source_path.name} {CONTRACT_SHEET_NAME} sheet 缺少列: "
            f"{CONTRACT_PREFIX_HEADER}，{CONTRACT_PREFIX_HEADER}将留空"
        )
    tax_rate_column_index = indexes.get(CONTRACT_TAX_RATE_HEADER)
    if tax_rate_column_index is None:
        products._contract_tax_rate_header_warning = (
            f"出口退税总表 {source_path.name} {CONTRACT_SHEET_NAME} sheet 缺少列: "
            f"{CONTRACT_TAX_RATE_HEADER}，{CONTRACT_TAX_RATE_HEADER}将留空"
        )
    conflicts: OrderedDict[str, None] = OrderedDict()
    prefix_conflicts: OrderedDict[str, None] = OrderedDict()
    tax_rate_conflicts: OrderedDict[str, None] = OrderedDict()
    contracts: dict[str, dict[str, str]] = {}
    for row in rows:
        row_values = {
            header: row[indexes[header] - 1] if indexes[header] - 1 < len(row) else None
            for header in CONTRACT_REQUIRED_HEADERS
        }
        if prefix_column_index is not None:
            row_values[CONTRACT_PREFIX_HEADER] = (
                row[prefix_column_index - 1] if prefix_column_index - 1 < len(row) else None
            )
        if tax_rate_column_index is not None:
            row_values[CONTRACT_TAX_RATE_HEADER] = (
                row[tax_rate_column_index - 1] if tax_rate_column_index - 1 < len(row) else None
            )
        if not any(_clean_cell(value) for value in row_values.values()):
            continue
        supplier = _clean_cell(row_values["供货方"])
        if not supplier or supplier in conflicts:
            continue
        contract = {
            "unit": _clean_cell(row_values["单位"]),
            "contract_product_name": _clean_cell(row_values["合同产品名称"]),
            "contract_no_prefix": _clean_cell(row_values.get(CONTRACT_PREFIX_HEADER)),
            "tax_rate": _clean_cell(row_values.get(CONTRACT_TAX_RATE_HEADER)),
        }
        existing = contracts.get(supplier)
        if existing is None:
            contracts[supplier] = contract
            continue
        if (
            existing["unit"] == contract["unit"]
            and existing["contract_product_name"] == contract["contract_product_name"]
        ):
            if existing["contract_no_prefix"] != contract["contract_no_prefix"]:
                existing["contract_no_prefix"] = ""
                prefix_conflicts[supplier] = None
            if existing["tax_rate"] != contract["tax_rate"]:
                existing["tax_rate"] = ""
                tax_rate_conflicts[supplier] = None
            continue
        contracts.pop(supplier, None)
        conflicts[supplier] = None
        prefix_conflicts.pop(supplier, None)
        tax_rate_conflicts.pop(supplier, None)

    products.contracts_by_manufacturer = contracts
    products.contract_mapping_count = len(contracts)
    products._contract_conflict_manufacturers = set(conflicts)
    products.contract_conflict_manufacturer_count = len(conflicts)
    products.contract_conflict_manufacturer_examples = list(conflicts)[:20]
    products._contract_prefix_conflict_manufacturers = prefix_conflicts
    products.contract_prefix_conflict_manufacturer_count = len(prefix_conflicts)
    products.contract_prefix_conflict_manufacturer_examples = list(prefix_conflicts)[:20]
    products._contract_tax_rate_conflict_manufacturers = tax_rate_conflicts
    products.contract_tax_rate_conflict_manufacturer_count = len(tax_rate_conflicts)
    products.contract_tax_rate_conflict_manufacturer_examples = list(tax_rate_conflicts)[:20]


def load_master_products(master_xlsx: str | Path) -> MasterProducts:
    source_path = Path(master_xlsx).expanduser()
    if not source_path.is_file():
        raise RuntimeError(f"找不到出口退税总表: {source_path}")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取出口退税总表") from exc

    try:
        workbook = load_workbook(source_path, data_only=True, read_only=True)
    except Exception as exc:
        raise RuntimeError(f"读取出口退税总表失败: {source_path}, error={exc}") from exc
    try:
        if MASTER_SKU_SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError(f"出口退税总表 {source_path.name} 缺少 sheet: {MASTER_SKU_SHEET_NAME}")
        worksheet = workbook[MASTER_SKU_SHEET_NAME]
        # Read-only workbooks must be consumed sequentially. Random single-cell
        # access reparses the stream and becomes extremely slow on large master files.
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            header_values = ()
        indexes = _header_indexes(header_values, source_path=source_path)
        products = MasterProducts()
        signatures: dict[str, tuple[Any, ...]] = {}
        first_row_numbers: dict[str, int] = {}
        duplicate_keys: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            row_values = {
                header: row[indexes[header] - 1] if indexes[header] - 1 < len(row) else None
                for header in MASTER_REQUIRED_HEADERS
            }
            if not any(_clean_cell(value) for value in row_values.values()):
                continue
            if not _clean_cell(row_values["库存sku"]):
                products.skipped_empty_sku_row_count += 1
                if len(products.skipped_empty_sku_rows) < 20:
                    products.skipped_empty_sku_rows.append(row_number)
                continue

            record, signature = _master_row_record(row_values, row_number=row_number, source_path=source_path)
            sku = str(record["stock_sku"])
            key = _sku_match_key(sku)
            if key in products:
                if signatures[key] == signature:
                    products.deduped_duplicate_row_count += 1
                    if key not in duplicate_keys and len(products.deduped_duplicate_sku_examples) < 20:
                        products.deduped_duplicate_sku_examples.append(sku)
                    duplicate_keys.add(key)
                    continue
                raise RuntimeError(
                    f"出口退税总表 {source_path.name} 库存sku重复且字段不一致: "
                    f"{sku}, 首次行={first_row_numbers[key]}, 冲突行={row_number}"
                )

            products[key] = record
            signatures[key] = signature
            first_row_numbers[key] = row_number
        _load_contract_mappings(workbook, source_path=source_path, products=products)
    finally:
        workbook.close()

    if not products:
        raise RuntimeError(f"出口退税总表 {source_path.name} 没有有效库存sku")
    _append_empty_sku_warning(products)
    _append_duplicate_warning(products, duplicate_keys)
    _append_contract_warnings(products)
    return products


def _find_required_delivery_csv(delivery_no: str, *, csv_dir: str | Path | None = None) -> Path:
    target = _require_delivery_no(delivery_no)
    directory = Path(DELIVERY_CSV_DIR if csv_dir is None else csv_dir)
    csv_path = find_latest_delivery_csv(target, csv_dir=directory)
    if csv_path is None:
        raise RuntimeError(f"本地未找到发货单 CSV: {directory / f'{target}_*.csv'}")
    return csv_path


def summarize_tax_sku_quantities_in_delivery_order(csv_path: str | Path) -> OrderedDict[str, Decimal]:
    headers, rows = _read_delivery_rows(csv_path)
    if SKU_SHIP_QTY_COLUMN not in headers:
        raise RuntimeError(f"发货单 CSV 缺少列: {SKU_SHIP_QTY_COLUMN}")

    summary: OrderedDict[str, Decimal] = OrderedDict()
    for index, row in enumerate(rows, start=2):
        cell_value = str(row.get(SKU_SHIP_QTY_COLUMN) or "").strip()
        if not cell_value:
            continue
        for raw_item in ITEM_SPLIT_PATTERN.split(cell_value):
            item = str(raw_item or "").strip()
            if not item:
                continue
            sku, quantity = _parse_sku_quantity_item(item, row_number=index)
            if sku not in summary:
                summary[sku] = Decimal("0")
            summary[sku] += quantity

    if not summary:
        raise RuntimeError(f"发货单 CSV 未解析到有效 {SKU_SHIP_QTY_COLUMN}")
    return summary


def summarize_delivery_quantities(
    delivery_nos: list[str],
    *,
    csv_dir: str | Path | None = None,
) -> tuple[OrderedDict[str, Decimal], OrderedDict[str, list[str]], list[str], list[str]]:
    if not delivery_nos:
        raise ValueError("至少需要一个 --delivery-no")

    normalized_delivery_nos: list[str] = []
    csv_paths: list[str] = []
    summary: OrderedDict[str, Decimal] = OrderedDict()
    sku_sources: OrderedDict[str, list[str]] = OrderedDict()
    for delivery_no in delivery_nos:
        target = _require_delivery_no(delivery_no)
        normalized_delivery_nos.append(target)
        csv_path = _find_required_delivery_csv(target, csv_dir=csv_dir)
        csv_paths.append(str(csv_path))

        for sku, quantity in summarize_tax_sku_quantities_in_delivery_order(csv_path).items():
            if sku not in summary:
                summary[sku] = Decimal("0")
                sku_sources[sku] = []
            summary[sku] += quantity
            if quantity > 0 and target not in sku_sources[sku]:
                sku_sources[sku].append(target)

    positive_summary = OrderedDict((sku, quantity) for sku, quantity in summary.items() if quantity > 0)
    if not positive_summary:
        raise RuntimeError("发货单 CSV 汇总后没有正数 SKU 发货量")
    positive_sources = OrderedDict((sku, sku_sources.get(sku, [])) for sku in positive_summary)
    return positive_summary, positive_sources, normalized_delivery_nos, csv_paths


def _safe_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    cleaned = _clean_cell(raw_title) or EMPTY_MANUFACTURER_SHEET_NAME
    base = INVALID_SHEET_TITLE_CHARS.sub("_", cleaned).strip("'").strip() or EMPTY_MANUFACTURER_SHEET_NAME
    base = base[:31]
    title = base
    index = 2
    while title in used_titles:
        suffix = f"_{index}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(title)
    return title


def build_restock_rows(
    summary: OrderedDict[str, Decimal],
    sku_sources: OrderedDict[str, list[str]],
    products: MasterProducts,
) -> tuple[list[list[Any]], OrderedDict[str, list[list[Any]]], list[list[Any]], int, int]:
    summary_entries: list[dict[str, Any]] = []
    manufacturer_entries: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    merge_entries: dict[tuple[str, str], dict[str, Any]] = {}
    unmatched_rows: list[list[Any]] = []
    matched_sku_count = 0
    unmatched_sku_count = 0

    for sku, quantity in summary.items():
        product = products.get(_sku_match_key(sku))
        quantity_value = _decimal_to_cell_value(quantity)
        source_values = list(sku_sources.get(sku, []))
        source_text = "\n".join(source_values)
        if product is None:
            unmatched_sku_count += 1
            unmatched_rows.append([sku, source_text, quantity_value, "出口退税总表未找到库存sku"])
            continue

        matched_sku_count += 1
        manufacturer = _clean_cell(product.get("manufacturer")) or EMPTY_MANUFACTURER_SHEET_NAME
        entries = manufacturer_entries.setdefault(manufacturer, [])
        stock_sku = _clean_cell(product.get("stock_sku")) or sku
        product_name = _clean_cell(product.get("product_name"))
        model = _clean_cell(product.get("model"))
        original_price = product["original_price"]
        unit, contract_product_name, contract_no_prefix, tax_rate = _contract_fields_for_manufacturer(
            products,
            manufacturer,
        )
        if not model:
            products.unmerged_empty_model_sku_count += 1
            if len(products.unmerged_empty_model_skus) < 20:
                products.unmerged_empty_model_skus.append(stock_sku)
            entries.append(
                {
                    "stock_skus": [stock_sku],
                    "stock_sku_keys": {_sku_match_key(stock_sku)},
                    "product_names": [product_name],
                    "source_delivery_nos": source_values,
                    "model": model,
                    "original_price": original_price,
                    "manufacturer": manufacturer,
                    "unit": unit,
                    "contract_product_name": contract_product_name,
                    "contract_no_prefix": contract_no_prefix,
                    "tax_rate": tax_rate,
                    "quantity": quantity,
                    "first_sku": stock_sku,
                }
            )
            summary_entries.append(entries[-1])
            continue

        merge_key = (manufacturer, model)
        entry = merge_entries.get(merge_key)
        if entry is None:
            entry = {
                "stock_skus": [stock_sku],
                "stock_sku_keys": {_sku_match_key(stock_sku)},
                "product_names": [product_name],
                "source_delivery_nos": source_values,
                "model": model,
                "original_price": original_price,
                "manufacturer": manufacturer,
                "unit": unit,
                "contract_product_name": contract_product_name,
                "contract_no_prefix": contract_no_prefix,
                "tax_rate": tax_rate,
                "quantity": quantity,
                "first_sku": stock_sku,
            }
            merge_entries[merge_key] = entry
            entries.append(entry)
            summary_entries.append(entry)
            continue

        if entry["original_price"] != original_price:
            raise RuntimeError(
                "出口退税总表同一厂家同一型号的原价不一致: "
                f"厂家={manufacturer}, 型号={model}, 首个SKU={entry['first_sku']}, 冲突SKU={stock_sku}"
            )
        stock_key = _sku_match_key(stock_sku)
        if stock_key not in entry["stock_sku_keys"]:
            entry["stock_skus"].append(stock_sku)
            entry["product_names"].append(product_name)
            entry["stock_sku_keys"].add(stock_key)
        for source_value in source_values:
            if source_value not in entry["source_delivery_nos"]:
                entry["source_delivery_nos"].append(source_value)
        entry["quantity"] += quantity

    def entry_to_row(entry: dict[str, Any]) -> list[Any]:
        total_price = entry["original_price"] * entry["quantity"]
        return [
            "\n".join(entry["stock_skus"]),
            "\n".join(entry["product_names"]),
            "\n".join(entry["source_delivery_nos"]),
            entry["stock_skus"][0] if entry["stock_skus"] else "",
            entry["product_names"][0] if entry["product_names"] else "",
            entry["model"],
            _decimal_to_cell_value(entry["original_price"]),
            entry["manufacturer"],
            entry["unit"],
            entry["contract_product_name"],
            entry["contract_no_prefix"],
            entry["tax_rate"],
            _decimal_to_cell_value(entry["quantity"]),
            _decimal_to_cell_value(total_price),
        ]

    summary_rows = [entry_to_row(entry) for entry in summary_entries]
    manufacturer_rows: OrderedDict[str, list[list[Any]]] = OrderedDict()
    for manufacturer, entries in manufacturer_entries.items():
        manufacturer_rows[manufacturer] = [entry_to_row(entry) for entry in entries]

    _append_empty_model_warning(products)
    _append_contract_unmapped_warning(products)
    return summary_rows, manufacturer_rows, unmatched_rows, matched_sku_count, unmatched_sku_count


def _decimal_places(value: Decimal) -> int:
    if value == value.to_integral_value():
        return 0
    normalized = value.normalize()
    return max(0, -normalized.as_tuple().exponent)


def _total_price_number_format(value: Any) -> str | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    if not decimal_value.is_finite():
        return None
    rounded_to_cents = decimal_value.quantize(Decimal("0.01"))
    if abs(decimal_value - rounded_to_cents) <= FLOAT_NOISE_TOLERANCE:
        return "0.00"
    decimal_places = max(2, _decimal_places(decimal_value))
    return "0." + ("0" * decimal_places)


def _write_rows(worksheet: Any, columns: tuple[str, ...], rows: list[list[Any]]) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    worksheet.append(list(columns))
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    worksheet.freeze_panes = "A2"
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = wrap_alignment
    for price_column_name in ("售价", "总价", "总价（售价）"):
        if price_column_name not in columns:
            continue
        total_price_column = columns.index(price_column_name) + 1
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=total_price_column)
            number_format = _total_price_number_format(cell.value)
            if number_format:
                cell.number_format = number_format
    for column_index in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 15
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 15


def _output_file_name(delivery_nos: list[str]) -> str:
    joined = "_".join(delivery_nos)
    if len(joined) > 120:
        joined = "_".join(delivery_nos[:3]) + f"_and_{len(delivery_nos) - 3}_more"
    return f"{joined}_purchase_summary.xlsx"


def write_restock_workbook(
    summary_rows: list[list[Any]],
    manufacturer_rows: OrderedDict[str, list[list[Any]]],
    unmatched_rows: list[list[Any]],
    *,
    delivery_nos: list[str],
    output_dir: str | Path | None = None,
) -> Path:
    directory = Path(OUTPUT_DIR if output_dir is None else output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    output_path = directory / _output_file_name(delivery_nos)

    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入 xlsx") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)

    used_titles = {SUMMARY_SHEET_NAME, UNMATCHED_SHEET_NAME}
    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    _write_rows(summary_sheet, MANUFACTURER_COLUMNS, summary_rows)

    unmatched_sheet = workbook.create_sheet(UNMATCHED_SHEET_NAME)
    _write_rows(unmatched_sheet, UNMATCHED_COLUMNS, unmatched_rows)

    for manufacturer, rows in manufacturer_rows.items():
        sheet_title = _safe_sheet_title(manufacturer, used_titles)
        worksheet = workbook.create_sheet(sheet_title)
        _write_rows(worksheet, MANUFACTURER_COLUMNS, rows)

    workbook.save(output_path)
    return output_path


def generate_restock_workbook(
    delivery_nos: list[str],
    *,
    master_xlsx: str | Path,
    csv_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    summary, sku_sources, normalized_delivery_nos, csv_paths = summarize_delivery_quantities(
        delivery_nos,
        csv_dir=csv_dir,
    )
    products = load_master_products(master_xlsx)
    summary_rows, manufacturer_rows, unmatched_rows, matched_sku_count, unmatched_sku_count = build_restock_rows(
        summary,
        sku_sources,
        products,
    )
    output_xlsx = write_restock_workbook(
        summary_rows,
        manufacturer_rows,
        unmatched_rows,
        delivery_nos=normalized_delivery_nos,
        output_dir=output_dir,
    )
    return {
        "success": True,
        "delivery_nos": normalized_delivery_nos,
        "csv_paths": csv_paths,
        "master_xlsx": str(Path(master_xlsx).expanduser()),
        "output_xlsx": str(output_xlsx),
        "sku_count": len(summary),
        "sku_source_count": sum(1 for sources in sku_sources.values() if sources),
        "matched_sku_count": matched_sku_count,
        "unmatched_sku_count": unmatched_sku_count,
        "manufacturer_count": len(manufacturer_rows),
        "deduped_duplicate_sku_count": products.deduped_duplicate_sku_count,
        "deduped_duplicate_row_count": products.deduped_duplicate_row_count,
        "deduped_duplicate_sku_examples": products.deduped_duplicate_sku_examples,
        "skipped_empty_sku_row_count": products.skipped_empty_sku_row_count,
        "skipped_empty_sku_rows": products.skipped_empty_sku_rows,
        "unmerged_empty_model_sku_count": products.unmerged_empty_model_sku_count,
        "unmerged_empty_model_skus": products.unmerged_empty_model_skus,
        "contract_mapping_count": products.contract_mapping_count,
        "contract_unmapped_manufacturer_count": products.contract_unmapped_manufacturer_count,
        "contract_unmapped_manufacturer_examples": products.contract_unmapped_manufacturer_examples,
        "contract_conflict_manufacturer_count": products.contract_conflict_manufacturer_count,
        "contract_conflict_manufacturer_examples": products.contract_conflict_manufacturer_examples,
        "contract_prefix_conflict_manufacturer_count": products.contract_prefix_conflict_manufacturer_count,
        "contract_prefix_conflict_manufacturer_examples": products.contract_prefix_conflict_manufacturer_examples,
        "contract_tax_rate_conflict_manufacturer_count": products.contract_tax_rate_conflict_manufacturer_count,
        "contract_tax_rate_conflict_manufacturer_examples": products.contract_tax_rate_conflict_manufacturer_examples,
        "warnings": products.warnings,
        "source": SOURCE,
    }


generate_purchase_summary_workbook = generate_restock_workbook


def build_parser(
    *,
    prog: str = "python -m services.agent_cli.mabang.generate_restock_workbook",
) -> argparse.ArgumentParser:
    parser = JsonArgumentParser(prog=prog)
    parser.add_argument("--delivery-no", action="append", default=[])
    parser.add_argument("--master-xlsx", required=True)
    return parser


def main(
    argv: list[str] | None = None,
    *,
    prog: str = "python -m services.agent_cli.mabang.generate_restock_workbook",
) -> int:
    configure_utf8_stdio()
    delivery_nos: list[str] = []
    master_xlsx = ""
    try:
        args = build_parser(prog=prog).parse_args(argv)
        delivery_nos = list(getattr(args, "delivery_no", []) or [])
        master_xlsx = str(getattr(args, "master_xlsx", "") or "")
        payload = generate_restock_workbook(delivery_nos, master_xlsx=master_xlsx)
    except Exception as exc:
        payload = {
            "success": False,
            "delivery_nos": delivery_nos,
            "master_xlsx": master_xlsx,
            "exception": _exception_text(exc),
            "source": SOURCE,
        }
    finally:
        try:
            asyncio.run(close_all_network_clients())
        except Exception:
            pass

    _write_json(payload)
    return 0 if bool(payload.get("success")) else 1


if __name__ == "__main__":
    raise SystemExit(main())
