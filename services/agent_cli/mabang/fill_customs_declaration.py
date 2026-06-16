from __future__ import annotations

import argparse
import copy
import json
import re
import shutil
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_output import configure_utf8_stdio
from services.agent_cli.mabang.restock_workbook import (
    SUMMARY_HEADERS,
    SUMMARY_WORKSHEET_NAME,
    find_summary_sheet,
    summary_column_indexes,
)
from services.amazon.amazon_logistic.sources.consignment_excel import (
    find_consignment_excel,
    resolve_column,
)

SOURCE = "customs_declaration_fill"
DEFAULT_TEMPLATE_PATH = Path("data") / "customs_declaration" / "custom_declaration_documents.xlsx"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "customs_declaration"
SOURCE_WORKSHEET_NAME = SUMMARY_WORKSHEET_NAME
INPUT_HEADERS = SUMMARY_HEADERS
TARGET_HEADERS = (
    "序号",
    "品名",
    "海关HS编码",
    "检疫附加码",
    "品牌",
    "型号",
    "品牌类型",
    "出口享惠情况",
    "其他申报要素",
)
CUSTOMS_DECLARATION_SHEET = "报关单"
DECLARATION_ELEMENTS_SHEET = "申报要素"
CUSTOMS_DETAIL_HEADER_LABELS = (
    "项号",
    "商品编号",
    "商品名称及规格型号",
    "数量及单位",
    "单价/总价/币制",
    "最终目的国（地区）",
    "净重",
    "毛重",
    "件数",
)
CUSTOMS_DETAIL_BLOCK_ROWS = 3
MAX_CUSTOMS_PRODUCT_ROWS = 50
CUSTOMS_DETAIL_BLANK_BLOCKS_TO_KEEP = 2
FORMULA_BLANK_ROWS_TO_KEEP = 5
SUPPORTED_DESTINATION_COUNTRIES = ("日本", "澳大利亚", "德国", "英国", "美国", "加拿大")
CONSIGNMENT_BOX_SEQUENCE_ALIASES = ("箱序号", "箱子编号", "箱号", "Box No", "Box Number")
CONSIGNMENT_GROSS_WEIGHT_ALIASES = ("毛重", "Gross Weight", "gross_weight", "weight")
ONE_DECIMAL = Decimal("0.1")
TWO_DECIMALS = Decimal("0.01")
THREE_DECIMALS = Decimal("0.001")
RMB_UPPER_DIGITS = ("零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖")
RMB_INTEGER_UNITS = ("", "拾", "佰", "仟")
RMB_SECTION_UNITS = ("", "万", "亿", "兆")
RMB_FRACTION_UNITS = ("角", "分", "厘")
SP_NO_PATTERN = re.compile(r"(SP\d+)", re.IGNORECASE)
CELL_REF_PATTERN = re.compile(
    r"(?<![A-Za-z0-9_])"
    r"(?P<sheet>(?:'[^']+'|[^=+\-*/^&(),:;\s]+)!)?"
    r"(?P<column>\$?[A-Z]{1,3})"
    r"(?P<row_absolute>\$?)"
    r"(?P<row>\d+)"
)


class JsonArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise ValueError(str(message or "").strip() or "参数解析失败")


@dataclass(frozen=True)
class SourceDeclarationRow:
    row_number: int
    source_name: str
    model: str
    quantity: Any
    commodity_name: str
    sale_price: Any
    total_price: Any
    unit: str


@dataclass(frozen=True)
class ClassificationResult:
    hs_code: str
    declaration_element: str


@dataclass(frozen=True)
class CustomsDetailLayout:
    header_row: int
    item_no_col: int
    quantity_col: int
    unit_col: int
    price_col: int
    destination_country_col: int
    net_weight_col: int
    gross_weight_col: int
    package_count_col: int
    block_count: int


@dataclass(frozen=True)
class ConsignmentWeightInfo:
    excel_path: str
    box_count: int
    total_gross_weight: Decimal


@dataclass(frozen=True)
class WeightAllocation:
    gross_weights: list[Decimal]
    net_weights: list[Decimal]


@dataclass(frozen=True)
class InputDeclarationBundle:
    input_path: Path
    sp_no: str
    destination_country: str
    source_rows: list[SourceDeclarationRow]
    consignment_weight_info: ConsignmentWeightInfo
    weight_allocation: WeightAllocation


@dataclass(frozen=True)
class FormulaSheetConfig:
    sheet_name: str
    required_headers: tuple[str, ...]
    owned_header_offsets: tuple[tuple[str, int], ...]
    summary_markers: tuple[str, ...]


FORMULA_SHEET_CONFIGS = (
    FormulaSheetConfig(
        sheet_name="发票",
        required_headers=(
            "標記號碼\nMark & No",
            "貨物名稱\nDescription",
            "型号\nModel",
            "數量\nQuantity",
            "單價\nUnit price",
            "總金額\nAmount",
        ),
        owned_header_offsets=(
            ("貨物名稱\nDescription", 0),
            ("型号\nModel", 0),
            ("數量\nQuantity", 0),
            ("數量\nQuantity", 1),
            ("單價\nUnit price", 0),
            ("總金額\nAmount", 0),
            ("總金額\nAmount", 1),
        ),
        summary_markers=("TOTAL:",),
    ),
    FormulaSheetConfig(
        sheet_name="箱单",
        required_headers=(
            "箱号\nCtn.No.",
            "货物名称及规格\nDescription",
            "型号\nModel",
            "箱数：\nPkg：",
            "数量：\nGe.Quantity",
            "毛重(KG)：\nG.W.(KG):",
            "净重(KG)：\nN.W.(KG):",
        ),
        owned_header_offsets=(
            ("货物名称及规格\nDescription", 0),
            ("型号\nModel", 0),
            ("数量：\nGe.Quantity", 0),
            ("数量：\nGe.Quantity", 1),
            ("毛重(KG)：\nG.W.(KG):", 0),
            ("净重(KG)：\nN.W.(KG):", 0),
        ),
        summary_markers=("合计\nTotal",),
    ),
    FormulaSheetConfig(
        sheet_name="合同",
        required_headers=(
            "Name of commodity",
            "Model",
            "Quantity",
            "Unit",
            "Unit Price",
            "Amount",
        ),
        owned_header_offsets=(
            ("Name of commodity", 0),
            ("Model", 0),
            ("Quantity", 0),
            ("Unit", 0),
            ("Unit Price", 0),
            ("Amount", 0),
            ("Amount", 1),
        ),
        summary_markers=("总      值", "Total Amount:"),
    ),
)


def _write_json(payload: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(dict(payload or {}), ensure_ascii=False) + "\n")
    sys.stdout.flush()


def _exception_text(exc: Exception) -> str:
    message = str(exc or "").strip()
    return message or exc.__class__.__name__


def _clean_cell(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() == "nan":
        return ""
    return text


def _decimal_to_json_number(value: Decimal) -> int | float:
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _decimal_to_excel_number(value: Decimal) -> int | float:
    return _decimal_to_json_number(value)


def _parse_decimal(value: Any, *, field_name: str) -> Decimal:
    text = _clean_cell(value)
    if not text:
        raise ValueError(f"{field_name} 不能为空")
    try:
        numeric = Decimal(text)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} 无法解析为数字: {value}") from exc
    return numeric


def _parse_positive_decimal(value: Any, *, field_name: str) -> Decimal:
    numeric = _parse_decimal(value, field_name=field_name)
    if numeric <= 0:
        raise ValueError(f"{field_name} 必须大于 0: {value}")
    return numeric


def calculate_total_amount(rows: list[SourceDeclarationRow]) -> Decimal:
    total = Decimal("0")
    for row in rows:
        amount = _parse_decimal(row.total_price, field_name=f"第{row.row_number}行总价")
        if amount < 0:
            raise ValueError(f"第{row.row_number}行总价 不能小于 0: {row.total_price}")
        total += amount
    return total.quantize(THREE_DECIMALS, rounding=ROUND_HALF_UP)


def _convert_four_digit_rmb_section(value: int) -> str:
    if not 0 <= value <= 9999:
        raise ValueError(f"人民币金额分组超出范围: {value}")

    parts: list[str] = []
    zero_pending = False
    remaining = value
    for divisor, unit in (
        (1000, RMB_INTEGER_UNITS[3]),
        (100, RMB_INTEGER_UNITS[2]),
        (10, RMB_INTEGER_UNITS[1]),
        (1, RMB_INTEGER_UNITS[0]),
    ):
        digit = remaining // divisor
        remaining %= divisor
        if digit:
            if zero_pending and parts:
                parts.append("零")
            parts.append(f"{RMB_UPPER_DIGITS[digit]}{unit}")
            zero_pending = False
        elif parts:
            zero_pending = True
    return "".join(parts)


def _convert_integer_rmb(value: int) -> str:
    if value == 0:
        return "零"
    if value < 0:
        raise ValueError("人民币金额不能小于 0")

    sections: list[int] = []
    remaining = value
    while remaining:
        sections.append(remaining % 10000)
        remaining //= 10000
    if len(sections) > len(RMB_SECTION_UNITS):
        raise ValueError(f"人民币金额整数部分过大: {value}")

    parts: list[str] = []
    zero_pending = False
    for section_index in range(len(sections) - 1, -1, -1):
        section = sections[section_index]
        if section == 0:
            if parts:
                zero_pending = True
            continue
        if parts and (zero_pending or section < 1000):
            parts.append("零")
        parts.append(f"{_convert_four_digit_rmb_section(section)}{RMB_SECTION_UNITS[section_index]}")
        zero_pending = False
    return "".join(parts).rstrip("零")


def _convert_fraction_rmb(fraction_units: int, *, integer_part: int) -> str:
    if fraction_units == 0:
        return "整"

    digits = [
        fraction_units // 100,
        (fraction_units // 10) % 10,
        fraction_units % 10,
    ]
    parts: list[str] = []
    zero_pending = False
    for index, digit in enumerate(digits):
        if digit:
            if parts and zero_pending:
                parts.append("零")
            elif not parts and integer_part and index > 0:
                parts.append("零" * index)
            parts.append(f"{RMB_UPPER_DIGITS[digit]}{RMB_FRACTION_UNITS[index]}")
            zero_pending = False
        elif parts:
            zero_pending = True
    return "".join(parts)


def amount_to_chinese_upper_rmb(amount: Decimal) -> str:
    if amount < 0:
        raise ValueError("人民币金额不能小于 0")

    rounded = amount.quantize(THREE_DECIMALS, rounding=ROUND_HALF_UP)
    amount_in_li = int((rounded * Decimal("1000")).to_integral_value(rounding=ROUND_HALF_UP))
    integer_part = amount_in_li // 1000
    fraction_units = amount_in_li % 1000
    integer_text = _convert_integer_rmb(integer_part)
    fraction_text = _convert_fraction_rmb(fraction_units, integer_part=integer_part)
    return f"人民币{integer_text}圆{fraction_text}"


def _parse_box_sequence(value: Any) -> int:
    numeric = _parse_positive_decimal(value, field_name="箱序号")
    integral = numeric.to_integral_value()
    if numeric != integral:
        raise ValueError(f"箱序号必须为正整数: {value}")
    return int(integral)


def _round_one_decimal(value: Decimal) -> Decimal:
    return value.quantize(ONE_DECIMAL, rounding=ROUND_HALF_UP)


def _round_two_decimals(value: Decimal) -> Decimal:
    return value.quantize(TWO_DECIMALS, rounding=ROUND_HALF_UP)


def extract_sp_no_from_filename(path: str | Path) -> str:
    name = Path(path).name
    match = SP_NO_PATTERN.search(name)
    if not match:
        raise ValueError(f"文件名中缺少 SP 单号: {name}")
    return match.group(1).upper()


def extract_destination_country_from_filename(path: str | Path) -> str:
    name = Path(path).name
    for country in SUPPORTED_DESTINATION_COUNTRIES:
        if country in name:
            return country
    supported = "、".join(SUPPORTED_DESTINATION_COUNTRIES)
    raise ValueError(f"文件名中缺少目的国，支持: {supported}")


def _load_workbook(path: Path, *, data_only: bool = False):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法处理 xlsx 文件") from exc

    try:
        return load_workbook(path, data_only=data_only)
    except Exception as exc:
        raise RuntimeError(f"读取 xlsx 文件失败: {path}, error={exc}") from exc


def read_source_rows(input_xlsx: str | Path) -> list[SourceDeclarationRow]:
    path = Path(input_xlsx)
    if not path.is_file():
        raise FileNotFoundError(f"输入 xlsx 不存在: {path}")

    workbook = _load_workbook(path, data_only=True)
    try:
        worksheet = find_summary_sheet(workbook, path)
        column_indexes = summary_column_indexes(worksheet, input_path=path, sheet_name=worksheet.title)

        rows: list[SourceDeclarationRow] = []
        for row_number in range(2, worksheet.max_row + 1):
            if not _clean_cell(worksheet.cell(row=row_number, column=column_indexes["日期"]).value):
                break
            rows.append(
                SourceDeclarationRow(
                    row_number=row_number,
                    source_name=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["品名"]).value),
                    model=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["规格型号"]).value),
                    quantity=worksheet.cell(row=row_number, column=column_indexes["发货量"]).value,
                    commodity_name=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["商品名称"]).value),
                    sale_price=worksheet.cell(row=row_number, column=column_indexes["售价"]).value,
                    total_price=worksheet.cell(row=row_number, column=column_indexes["总价"]).value,
                    unit=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["单位"]).value),
                )
            )

        if not rows:
            raise ValueError(f"输入 xlsx 的 sheet {worksheet.title} 未解析到有效数据: {path.name}")
        return rows
    finally:
        workbook.close()


def classify_declaration(row: SourceDeclarationRow) -> ClassificationResult:
    commodity_name = row.commodity_name
    source_name = row.source_name
    source_name_upper = source_name.upper()

    if "编织表带" in commodity_name or "尼龙表带" in commodity_name:
        return ClassificationResult("9113900090", "尼龙")

    if "表带" in commodity_name:
        if "米兰尼斯" in source_name or "米兰" in source_name:
            return ClassificationResult("9113200000", "贱金属")
        if "编织" in source_name or "尼龙" in source_name:
            return ClassificationResult("9113900090", "尼龙")
        if "真皮" in source_name or "尖尾皮带" in source_name:
            return ClassificationResult("9113900090", "皮革")
        if "硅胶" in source_name:
            return ClassificationResult("9113900090", "硅胶")
        return ClassificationResult("", "")

    if "表壳" in commodity_name:
        if "金属" in source_name:
            return ClassificationResult("9111800000", "贱金属")
        if "PC" in source_name_upper:
            return ClassificationResult("9111800000", "PC")
        if "TPU" in source_name_upper:
            return ClassificationResult("9111800000", "TPU")
        return ClassificationResult("9111800000", "PC")

    if "包装盒" in commodity_name:
        return ClassificationResult("4819200000", "纸质+塑料")

    if "手表保护套" in commodity_name:
        if "PC" in source_name_upper:
            return ClassificationResult("3926909090", "PC")
        if "TPU" in source_name_upper:
            return ClassificationResult("3926909090", "TPU")
        return ClassificationResult("3926909090", "PC")

    return ClassificationResult("", "")


def _find_target_header_row(worksheet: Any) -> int:
    expected = list(TARGET_HEADERS)
    for row_index in range(1, worksheet.max_row + 1):
        headers = [
            _clean_cell(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, len(expected) + 1)
        ]
        if headers == expected:
            return row_index
    raise ValueError(f"模板申报要素 sheet 缺少表头: {expected}")


def _clear_old_declaration_rows(worksheet: Any, *, header_row: int) -> None:
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        for column_index in range(1, len(TARGET_HEADERS) + 1):
            worksheet.cell(row=row_index, column=column_index).value = None


def _write_declaration_rows(worksheet: Any, *, header_row: int, rows: list[SourceDeclarationRow]) -> list[str]:
    header_to_col = {header: index + 1 for index, header in enumerate(TARGET_HEADERS)}
    notice: list[str] = []

    for offset, row in enumerate(rows, start=1):
        target_row = header_row + offset
        classification = classify_declaration(row)
        if not classification.hs_code or not classification.declaration_element:
            notice.append(
                f"第{row.row_number}行未匹配申报规则: "
                f"商品名称={row.commodity_name}, 品名={row.source_name}"
            )

        worksheet.cell(row=target_row, column=header_to_col["序号"], value=offset)
        worksheet.cell(row=target_row, column=header_to_col["品名"], value=row.commodity_name)
        worksheet.cell(row=target_row, column=header_to_col["海关HS编码"]).value = classification.hs_code or None
        worksheet.cell(row=target_row, column=header_to_col["检疫附加码"]).value = None
        worksheet.cell(row=target_row, column=header_to_col["品牌"], value="0无品牌")
        worksheet.cell(row=target_row, column=header_to_col["型号"], value=row.model)
        worksheet.cell(row=target_row, column=header_to_col["品牌类型"], value="0无品牌")
        worksheet.cell(row=target_row, column=header_to_col["出口享惠情况"], value="0不享惠")
        worksheet.cell(row=target_row, column=header_to_col["其他申报要素"]).value = (
            classification.declaration_element or None
        )

    return notice


def _merged_top_left(worksheet: Any, *, row: int, column: int) -> tuple[int, int]:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return merged_range.min_row, merged_range.min_col
    return row, column


def _write_merged_safe(worksheet: Any, *, row: int, column: int, value: Any) -> None:
    target_row, target_col = _merged_top_left(worksheet, row=row, column=column)
    worksheet.cell(row=target_row, column=target_col).value = value


def _find_cell_exact_text(worksheet: Any, text: str) -> tuple[int, int]:
    expected = _clean_cell(text)
    for row_index in range(1, worksheet.max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            if _clean_cell(worksheet.cell(row=row_index, column=column_index).value) == expected:
                return row_index, column_index
    raise ValueError(f"{worksheet.title} sheet 找不到单元格: {text}")


def _write_uppercase_amount(workbook: Any, amount_upper: str) -> None:
    if "发票" not in workbook.sheetnames:
        raise ValueError("报关资料模板缺少 sheet: 发票")
    if "合同" not in workbook.sheetnames:
        raise ValueError("报关资料模板缺少 sheet: 合同")

    invoice_sheet = workbook["发票"]
    invoice_total_row, invoice_total_col = _find_cell_exact_text(invoice_sheet, "TOTAL:")
    if invoice_total_col <= 1:
        raise ValueError("发票 sheet TOTAL: 左侧没有可填写中文金额的单元格")
    _write_merged_safe(
        invoice_sheet,
        row=invoice_total_row,
        column=invoice_total_col - 1,
        value=amount_upper,
    )

    contract_sheet = workbook["合同"]
    contract_total_row, contract_total_col = _find_cell_exact_text(contract_sheet, "Total Amount:")
    _write_merged_safe(
        contract_sheet,
        row=contract_total_row + 1,
        column=contract_total_col,
        value=amount_upper,
    )


def _detail_block_has_template_content(worksheet: Any, *, start_row: int, max_col: int) -> bool:
    end_row = start_row + CUSTOMS_DETAIL_BLOCK_ROWS - 1
    for row_index in range(start_row, end_row + 1):
        for column_index in range(1, max_col + 1):
            if worksheet.cell(row=row_index, column=column_index).value is not None:
                return True

    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row >= start_row and merged_range.min_row <= end_row:
            return True
    return False


def _count_customs_detail_blocks(worksheet: Any, *, header_row: int) -> int:
    block_count = 0
    start_row = header_row + 1
    max_col = max(worksheet.max_column, 1)
    while start_row + CUSTOMS_DETAIL_BLOCK_ROWS - 1 <= worksheet.max_row:
        if not _detail_block_has_template_content(worksheet, start_row=start_row, max_col=max_col):
            break
        block_count += 1
        start_row += CUSTOMS_DETAIL_BLOCK_ROWS
    return block_count


def _find_customs_detail_layout(worksheet: Any) -> CustomsDetailLayout:
    for row_index in range(1, worksheet.max_row + 1):
        header_to_col: dict[str, int] = {}
        for column_index in range(1, worksheet.max_column + 1):
            value = _clean_cell(worksheet.cell(row=row_index, column=column_index).value)
            if value and value not in header_to_col:
                header_to_col[value] = column_index
        if all(label in header_to_col for label in CUSTOMS_DETAIL_HEADER_LABELS):
            block_count = _count_customs_detail_blocks(worksheet, header_row=row_index)
            if block_count <= 0:
                raise ValueError("报关单明细区没有可写入的商品明细块")
            quantity_col = header_to_col["数量及单位"]
            return CustomsDetailLayout(
                header_row=row_index,
                item_no_col=header_to_col["项号"],
                quantity_col=quantity_col,
                unit_col=quantity_col + 1,
                price_col=header_to_col["单价/总价/币制"],
                destination_country_col=header_to_col["最终目的国（地区）"],
                net_weight_col=header_to_col["净重"],
                gross_weight_col=header_to_col["毛重"],
                package_count_col=header_to_col["件数"],
                block_count=block_count,
            )
    raise ValueError(f"报关单 sheet 缺少明细表头: {list(CUSTOMS_DETAIL_HEADER_LABELS)}")


def _customs_detail_start_row(layout: CustomsDetailLayout, *, index: int) -> int:
    return layout.header_row + 1 + (index - 1) * CUSTOMS_DETAIL_BLOCK_ROWS


def _clear_customs_detail_owned_fields(
    worksheet: Any,
    *,
    layout: CustomsDetailLayout,
    start_index: int,
) -> None:
    for index in range(start_index, layout.block_count + 1):
        start_row = _customs_detail_start_row(layout, index=index)
        for row_index in range(start_row, start_row + CUSTOMS_DETAIL_BLOCK_ROWS):
            for column_index in range(1, layout.package_count_col + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if cell.__class__.__name__ == "MergedCell":
                    continue
                cell.value = None


def _resize_customs_package_count_merge(
    worksheet: Any,
    *,
    layout: CustomsDetailLayout,
    row_count: int,
) -> None:
    first_row = _customs_detail_start_row(layout, index=1)
    last_row = _customs_detail_start_row(layout, index=row_count) + CUSTOMS_DETAIL_BLOCK_ROWS - 1
    for merged_range in list(worksheet.merged_cells.ranges):
        if (
            merged_range.min_col <= layout.package_count_col <= merged_range.max_col
            and merged_range.min_row <= first_row <= merged_range.max_row
        ):
            worksheet.unmerge_cells(str(merged_range))
            break
    worksheet.merge_cells(
        start_row=first_row,
        start_column=layout.package_count_col,
        end_row=last_row,
        end_column=layout.package_count_col,
    )


def _delete_unused_customs_detail_blocks(
    worksheet: Any,
    *,
    layout: CustomsDetailLayout,
    row_count: int,
) -> None:
    _clear_customs_detail_owned_fields(
        worksheet,
        layout=layout,
        start_index=row_count + 1,
    )
    first_deleted_block = row_count + CUSTOMS_DETAIL_BLANK_BLOCKS_TO_KEEP + 1
    if first_deleted_block <= layout.block_count:
        _delete_rows_preserving_template(
            worksheet,
            start_row=_customs_detail_start_row(layout, index=first_deleted_block),
            amount=(layout.block_count - first_deleted_block + 1) * CUSTOMS_DETAIL_BLOCK_ROWS,
        )


def _calculate_net_weight(gross_weight: Decimal) -> Decimal:
    net_weight = gross_weight - ONE_DECIMAL
    if net_weight > 0:
        return _round_one_decimal(net_weight)
    net_weight = gross_weight - TWO_DECIMALS
    if net_weight > 0:
        return _round_two_decimals(net_weight)
    return gross_weight


def allocate_weights_by_quantity(
    rows: list[SourceDeclarationRow],
    *,
    total_gross_weight: Decimal,
) -> WeightAllocation:
    quantities = [
        _parse_positive_decimal(row.quantity, field_name=f"第{row.row_number}行发货量")
        for row in rows
    ]
    total_quantity = sum(quantities, Decimal("0"))
    if total_quantity <= 0:
        raise ValueError("发货量合计必须大于 0")

    rounded_total_gross = _round_one_decimal(total_gross_weight)
    gross_weights = [
        _round_one_decimal(rounded_total_gross * quantity / total_quantity)
        for quantity in quantities
    ]
    diff = rounded_total_gross - sum(gross_weights, Decimal("0"))
    if diff:
        max_index = max(range(len(gross_weights)), key=lambda index: gross_weights[index])
        gross_weights[max_index] = _round_one_decimal(gross_weights[max_index] + diff)
        if gross_weights[max_index] <= 0:
            raise ValueError("毛重分配结果无效，最大项调整后小于等于 0")

    net_weights = [_calculate_net_weight(weight) for weight in gross_weights]
    return WeightAllocation(gross_weights=gross_weights, net_weights=net_weights)


def _write_customs_detail_rows(
    worksheet: Any,
    *,
    rows: list[SourceDeclarationRow],
    destination_country: str,
    weight_allocation: WeightAllocation,
    box_count: int,
) -> int:
    layout = _find_customs_detail_layout(worksheet)
    if len(rows) > layout.block_count:
        raise ValueError(
            "报关单明细区容量不足: "
            f"需要 {len(rows)} 行商品，模板可容纳 {layout.block_count} 行商品"
        )
    if len(weight_allocation.gross_weights) != len(rows) or len(weight_allocation.net_weights) != len(rows):
        raise ValueError("重量分配结果数量与商品行数不一致")

    for index, row in enumerate(rows, start=1):
        start_row = _customs_detail_start_row(layout, index=index)
        quantity_row = start_row + 2
        _write_merged_safe(worksheet, row=start_row, column=layout.item_no_col, value=index)
        _write_merged_safe(worksheet, row=start_row, column=layout.price_col, value=row.sale_price)
        _write_merged_safe(worksheet, row=start_row, column=layout.destination_country_col, value=destination_country)
        _write_merged_safe(
            worksheet,
            row=start_row,
            column=layout.net_weight_col,
            value=_decimal_to_excel_number(weight_allocation.net_weights[index - 1]),
        )
        _write_merged_safe(
            worksheet,
            row=start_row,
            column=layout.gross_weight_col,
            value=_decimal_to_excel_number(weight_allocation.gross_weights[index - 1]),
        )
        _write_merged_safe(worksheet, row=quantity_row, column=layout.quantity_col, value=row.quantity)
        _write_merged_safe(worksheet, row=quantity_row, column=layout.unit_col, value=row.unit)

    _write_merged_safe(
        worksheet,
        row=_customs_detail_start_row(layout, index=1),
        column=layout.package_count_col,
        value=box_count,
    )
    _resize_customs_package_count_merge(
        worksheet,
        layout=layout,
        row_count=len(rows),
    )
    _write_merged_safe(
        worksheet,
        row=_customs_detail_start_row(layout, index=1),
        column=layout.package_count_col,
        value=box_count,
    )
    _delete_unused_customs_detail_blocks(
        worksheet,
        layout=layout,
        row_count=len(rows),
    )
    return len(rows)


def _normalize_header_text(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_cell(value)).lower()


def _normalize_formula_label(label: str) -> str:
    return _normalize_header_text(label)


def _row_header_map(worksheet: Any, *, row_index: int) -> dict[str, int]:
    header_to_col: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        value = _normalize_header_text(worksheet.cell(row=row_index, column=column_index).value)
        if value and value not in header_to_col:
            header_to_col[value] = column_index
    return header_to_col


def _find_formula_header_row(worksheet: Any, config: FormulaSheetConfig) -> tuple[int, dict[str, int]]:
    required = [_normalize_formula_label(label) for label in config.required_headers]
    for row_index in range(1, worksheet.max_row + 1):
        header_to_col = _row_header_map(worksheet, row_index=row_index)
        if all(label in header_to_col for label in required):
            return row_index, header_to_col
    raise ValueError(f"{config.sheet_name} sheet 缺少公式区表头: {list(config.required_headers)}")


def _formula_owned_columns(header_to_col: dict[str, int], config: FormulaSheetConfig) -> list[int]:
    columns: list[int] = []
    for raw_label, offset in config.owned_header_offsets:
        label = _normalize_formula_label(raw_label)
        if label not in header_to_col:
            raise ValueError(f"{config.sheet_name} sheet 缺少公式列: {raw_label}")
        column_index = header_to_col[label] + offset
        if column_index not in columns:
            columns.append(column_index)
    return columns


def _find_formula_summary_row(worksheet: Any, config: FormulaSheetConfig, *, start_row: int) -> int:
    markers = [_normalize_formula_label(marker) for marker in config.summary_markers]
    for row_index in range(start_row, worksheet.max_row + 1):
        values = [
            _normalize_header_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, worksheet.max_column + 1)
        ]
        if any(marker and any(marker in value for value in values) for marker in markers):
            return row_index
    raise ValueError(f"{config.sheet_name} sheet 缺少公式区合计行: {list(config.summary_markers)}")


def _translate_formula_row_references(formula: Any, *, row_delta: int) -> Any:
    if not isinstance(formula, str) or not formula.startswith("=") or row_delta == 0:
        return formula

    def replace(match: re.Match[str]) -> str:
        row_absolute = match.group("row_absolute")
        row_number = int(match.group("row"))
        if row_absolute:
            new_row = row_number
        else:
            new_row = row_number + row_delta
            if new_row <= 0:
                new_row = row_number
        return (
            f"{match.group('sheet') or ''}"
            f"{match.group('column')}"
            f"{row_absolute}"
            f"{new_row}"
        )

    return CELL_REF_PATTERN.sub(replace, formula)


def _formula_sheet_name(sheet_reference: str | None) -> str:
    if not sheet_reference:
        return ""
    text = sheet_reference[:-1] if sheet_reference.endswith("!") else sheet_reference
    if text.startswith("'") and text.endswith("'"):
        text = text[1:-1].replace("''", "'")
    return text


def _translate_formula_after_row_insert(
    formula: Any,
    *,
    current_sheet_name: str,
    insert_at: int,
    amount: int,
) -> Any:
    if not isinstance(formula, str) or not formula.startswith("=") or amount <= 0:
        return formula

    def replace(match: re.Match[str]) -> str:
        sheet_reference = match.group("sheet")
        referenced_sheet = _formula_sheet_name(sheet_reference)
        row_absolute = match.group("row_absolute")
        row_number = int(match.group("row"))
        if referenced_sheet and referenced_sheet != current_sheet_name:
            new_row = row_number
        elif row_absolute or row_number < insert_at:
            new_row = row_number
        else:
            new_row = row_number + amount
        return (
            f"{sheet_reference or ''}"
            f"{match.group('column')}"
            f"{row_absolute}"
            f"{new_row}"
        )

    return CELL_REF_PATTERN.sub(replace, formula)


def _translate_formula_after_row_delete(
    formula: Any,
    *,
    current_sheet_name: str,
    delete_start: int,
    amount: int,
) -> Any:
    if not isinstance(formula, str) or not formula.startswith("=") or amount <= 0:
        return formula

    delete_end = delete_start + amount - 1

    def replace(match: re.Match[str]) -> str:
        sheet_reference = match.group("sheet")
        referenced_sheet = _formula_sheet_name(sheet_reference)
        row_absolute = match.group("row_absolute")
        row_number = int(match.group("row"))
        if referenced_sheet and referenced_sheet != current_sheet_name:
            new_row = row_number
        elif row_absolute or row_number <= delete_end:
            new_row = row_number
        else:
            new_row = row_number - amount
        return (
            f"{sheet_reference or ''}"
            f"{match.group('column')}"
            f"{row_absolute}"
            f"{new_row}"
        )

    return CELL_REF_PATTERN.sub(replace, formula)


def _copy_cell_style(source_cell: Any, target_cell: Any) -> None:
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy.copy(source_cell.protection)
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)


def _range_intersects_columns(min_col: int, max_col: int, columns: set[int]) -> bool:
    return any(min_col <= column <= max_col for column in columns)


def _merged_range_exists(worksheet: Any, bounds: tuple[int, int, int, int]) -> bool:
    min_col, min_row, max_col, max_row = bounds
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_col == min_col
            and merged_range.min_row == min_row
            and merged_range.max_col == max_col
            and merged_range.max_row == max_row
        ):
            return True
    return False


def _print_area_bounds(worksheet: Any) -> tuple[int, int, int, int] | None:
    print_area = _clean_cell(worksheet.print_area)
    if not print_area:
        return None
    if "," in print_area:
        raise ValueError(f"{worksheet.title} sheet 打印区域不是单段区域: {print_area}")
    range_text = print_area.rsplit("!", 1)[-1]
    try:
        from openpyxl.utils.cell import range_boundaries
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法解析打印区域") from exc
    return range_boundaries(range_text)


def _adjust_print_area_after_row_delete(worksheet: Any, *, start_row: int, amount: int) -> int | None:
    bounds = _print_area_bounds(worksheet)
    if bounds is None:
        return None
    min_col, min_row, max_col, max_row = bounds
    delete_end = start_row + amount - 1
    if delete_end < min_row:
        min_row = max(1, min_row - amount)
        max_row = max(min_row, max_row - amount)
    elif start_row <= max_row:
        removed_count = min(max_row, delete_end) - max(min_row, start_row) + 1
        if removed_count > 0:
            max_row = max(min_row, max_row - removed_count)

    try:
        from openpyxl.utils.cell import get_column_letter
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法更新打印区域") from exc
    worksheet.print_area = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    return max_row


def _adjust_row_breaks_after_row_delete(
    worksheet: Any,
    *,
    start_row: int,
    amount: int,
    print_area_end_row: int | None,
) -> None:
    if amount <= 0 or not worksheet.row_breaks.brk:
        return

    try:
        from openpyxl.worksheet.pagebreak import Break
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法更新分页符") from exc

    delete_end = start_row + amount - 1
    adjusted_breaks = []
    for row_break in worksheet.row_breaks.brk:
        break_id = int(row_break.id)
        if start_row <= break_id <= delete_end:
            continue
        if break_id > delete_end:
            break_id -= amount
        if print_area_end_row is not None and break_id > print_area_end_row:
            continue
        adjusted_breaks.append(
            Break(
                id=break_id,
                min=row_break.min,
                max=row_break.max,
                man=row_break.man,
                pt=row_break.pt,
            )
        )
    worksheet.row_breaks.brk = adjusted_breaks


def _adjust_print_settings_after_row_delete(worksheet: Any, *, start_row: int, amount: int) -> None:
    print_area_end_row = _adjust_print_area_after_row_delete(
        worksheet,
        start_row=start_row,
        amount=amount,
    )
    _adjust_row_breaks_after_row_delete(
        worksheet,
        start_row=start_row,
        amount=amount,
        print_area_end_row=print_area_end_row,
    )


def _insert_rows_preserving_template(worksheet: Any, *, insert_at: int, amount: int) -> None:
    if amount <= 0:
        return

    old_max_row = worksheet.max_row
    shifted_merged_ranges: list[tuple[int, int, int, int]] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row < insert_at:
            continue
        shifted_merged_ranges.append(
            (
                merged_range.min_col,
                merged_range.min_row + amount,
                merged_range.max_col,
                merged_range.max_row + amount,
            )
        )
        worksheet.unmerge_cells(str(merged_range))

    worksheet.insert_rows(insert_at, amount=amount)

    for bounds in shifted_merged_ranges:
        if not _merged_range_exists(worksheet, bounds):
            worksheet.merge_cells(
                start_column=bounds[0],
                start_row=bounds[1],
                end_column=bounds[2],
                end_row=bounds[3],
            )

    for row_index in range(insert_at + amount, old_max_row + amount + 1):
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.value = _translate_formula_after_row_insert(
                cell.value,
                current_sheet_name=worksheet.title,
                insert_at=insert_at,
                amount=amount,
            )


def _delete_rows_preserving_template(worksheet: Any, *, start_row: int, amount: int) -> None:
    if amount <= 0:
        return

    end_row = start_row + amount - 1
    old_max_row = worksheet.max_row
    retained_merged_ranges: list[tuple[int, int, int, int]] = []
    ranges_to_unmerge: list[str] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row < start_row:
            continue
        if merged_range.min_row > end_row:
            retained_merged_ranges.append(
                (
                    merged_range.min_col,
                    merged_range.min_row - amount,
                    merged_range.max_col,
                    merged_range.max_row - amount,
                )
            )
            ranges_to_unmerge.append(str(merged_range))
            continue
        if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
            ranges_to_unmerge.append(str(merged_range))
            continue
        if merged_range.min_row < start_row and merged_range.max_row <= end_row:
            retained_merged_ranges.append(
                (
                    merged_range.min_col,
                    merged_range.min_row,
                    merged_range.max_col,
                    start_row - 1,
                )
            )
            ranges_to_unmerge.append(str(merged_range))
            continue
        raise ValueError(f"{worksheet.title} sheet 合并单元格跨越删除区域: {merged_range}")

    for merged_range_ref in ranges_to_unmerge:
        worksheet.unmerge_cells(merged_range_ref)

    worksheet.delete_rows(start_row, amount=amount)

    for bounds in retained_merged_ranges:
        if not _merged_range_exists(worksheet, bounds):
            worksheet.merge_cells(
                start_column=bounds[0],
                start_row=bounds[1],
                end_column=bounds[2],
                end_row=bounds[3],
            )

    for row_index in range(start_row, max(start_row, old_max_row - amount + 1)):
        for column_index in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.value = _translate_formula_after_row_delete(
                cell.value,
                current_sheet_name=worksheet.title,
                delete_start=start_row,
                amount=amount,
            )
    _adjust_print_settings_after_row_delete(
        worksheet,
        start_row=start_row,
        amount=amount,
    )


def _copy_formula_row(
    worksheet: Any,
    *,
    source_row: int,
    target_row: int,
    owned_columns: list[int],
) -> None:
    row_delta = target_row - source_row
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    owned_column_set = set(owned_columns)
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row != source_row or merged_range.max_row != source_row:
            continue
        if not _range_intersects_columns(merged_range.min_col, merged_range.max_col, owned_column_set):
            continue
        target_bounds = (
            merged_range.min_col,
            target_row,
            merged_range.max_col,
            target_row,
        )
        if not _merged_range_exists(worksheet, target_bounds):
            worksheet.merge_cells(
                start_row=target_row,
                start_column=merged_range.min_col,
                end_row=target_row,
                end_column=merged_range.max_col,
            )

    for column_index in owned_columns:
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_row_index, target_col_index = _merged_top_left(
            worksheet,
            row=target_row,
            column=column_index,
        )
        target_cell = worksheet.cell(row=target_row_index, column=target_col_index)
        _copy_cell_style(source_cell, target_cell)
        target_cell.value = _translate_formula_row_references(source_cell.value, row_delta=row_delta)


def _clear_formula_rows(
    worksheet: Any,
    *,
    owned_columns: list[int],
    start_row: int,
    end_row: int,
) -> None:
    if start_row > end_row:
        return
    for row_index in range(start_row, end_row + 1):
        for column_index in owned_columns:
            _write_merged_safe(worksheet, row=row_index, column=column_index, value=None)


def _update_formula_summary(
    worksheet: Any,
    *,
    config: FormulaSheetConfig,
    summary_row: int,
    detail_start_row: int,
    detail_end_row: int,
) -> None:
    if config.sheet_name == "发票":
        _write_merged_safe(worksheet, row=summary_row, column=8, value=f"=H{detail_start_row}")
        _write_merged_safe(worksheet, row=summary_row, column=9, value=f"=SUM(I{detail_start_row}:I{detail_end_row})")
        return
    if config.sheet_name == "箱单":
        _write_merged_safe(worksheet, row=summary_row, column=7, value=f"=SUM(G{detail_start_row}:G{detail_end_row})")
        _write_merged_safe(worksheet, row=summary_row, column=8, value=f"=SUM(H{detail_start_row}:H{detail_end_row})")
        return
    if config.sheet_name == "合同":
        _write_merged_safe(worksheet, row=summary_row, column=8, value=f"=H{detail_start_row}")
        _write_merged_safe(worksheet, row=summary_row, column=9, value=f"=SUM(I{detail_start_row}:K{detail_end_row})")


def _fill_formula_sheet(
    worksheet: Any,
    *,
    config: FormulaSheetConfig,
    row_count: int,
) -> int:
    header_row, header_to_col = _find_formula_header_row(worksheet, config)
    formula_start_row = header_row + 1
    summary_row = _find_formula_summary_row(worksheet, config, start_row=formula_start_row + 1)
    available_rows = summary_row - formula_start_row
    if available_rows <= 0:
        raise ValueError(f"{config.sheet_name} sheet 公式区没有可复制的明细行")

    owned_columns = _formula_owned_columns(header_to_col, config)
    if row_count > available_rows:
        missing_rows = row_count - available_rows
        _insert_rows_preserving_template(worksheet, insert_at=summary_row, amount=missing_rows)
        summary_row += missing_rows
        available_rows = row_count

    for offset in range(row_count):
        _copy_formula_row(
            worksheet,
            source_row=formula_start_row,
            target_row=formula_start_row + offset,
            owned_columns=owned_columns,
        )

    _clear_formula_rows(
        worksheet,
        owned_columns=owned_columns,
        start_row=formula_start_row + row_count,
        end_row=formula_start_row + available_rows - 1,
    )
    unused_row_count = available_rows - row_count
    deleted_row_count = max(0, unused_row_count - FORMULA_BLANK_ROWS_TO_KEEP)
    if deleted_row_count > 0:
        _delete_rows_preserving_template(
            worksheet,
            start_row=formula_start_row + row_count + FORMULA_BLANK_ROWS_TO_KEEP,
            amount=deleted_row_count,
        )
        summary_row -= deleted_row_count
    _update_formula_summary(
        worksheet,
        config=config,
        summary_row=summary_row,
        detail_start_row=formula_start_row,
        detail_end_row=formula_start_row + row_count - 1,
    )
    return row_count


def _fill_formula_sheets(workbook: Any, *, row_count: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for config in FORMULA_SHEET_CONFIGS:
        if config.sheet_name not in workbook.sheetnames:
            raise ValueError(f"报关资料模板缺少 sheet: {config.sheet_name}")
        result[config.sheet_name] = _fill_formula_sheet(
            workbook[config.sheet_name],
            config=config,
            row_count=row_count,
        )
    return result


def read_consignment_weight_info(excel_path: str | Path) -> ConsignmentWeightInfo:
    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError("缺少 pandas 依赖，无法读取装箱数据 Excel") from exc

    source_path = Path(excel_path).expanduser()
    if not source_path.is_absolute():
        source_path = (Path.cwd() / source_path).resolve()
    else:
        source_path = source_path.resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"装箱数据 Excel 不存在: {source_path}")

    try:
        with pd.ExcelFile(source_path) as workbook:
            if not workbook.sheet_names:
                raise RuntimeError("装箱数据 Excel 没有可用 sheet")
            df = pd.read_excel(workbook, sheet_name=workbook.sheet_names[0])
    except Exception as exc:
        raise RuntimeError(f"读取装箱数据 Excel 失败: {source_path.name}, error={exc}") from exc

    if df.empty:
        raise RuntimeError(f"装箱数据 Excel 没有数据: {source_path.name}")

    columns = [str(column).strip() for column in list(df.columns)]
    df.columns = columns
    box_sequence_col = resolve_column(columns, CONSIGNMENT_BOX_SEQUENCE_ALIASES)
    gross_weight_col = resolve_column(columns, CONSIGNMENT_GROSS_WEIGHT_ALIASES)
    missing = [
        name
        for name, column in (("箱序号", box_sequence_col), ("毛重", gross_weight_col))
        if not column
    ]
    if missing:
        raise RuntimeError(f"装箱数据 Excel 缺少必需列: {', '.join(missing)}")

    weights_by_box: dict[int, Decimal] = {}
    for row_number, row in enumerate(df.to_dict(orient="records"), start=2):
        try:
            box_sequence = _parse_box_sequence(row.get(box_sequence_col))
            gross_weight = _parse_positive_decimal(row.get(gross_weight_col), field_name="毛重")
        except Exception as exc:
            raise RuntimeError(f"装箱数据第{row_number}行解析失败: {exc}") from exc
        existing = weights_by_box.get(box_sequence)
        if existing is not None and existing != gross_weight:
            raise RuntimeError(
                f"装箱数据同一箱序号存在不同毛重: 箱序号={box_sequence}, "
                f"毛重={existing} / {gross_weight}"
            )
        weights_by_box[box_sequence] = gross_weight

    if not weights_by_box:
        raise RuntimeError(f"装箱数据未解析到有效箱序号: {source_path.name}")

    box_count = max(weights_by_box)
    total_gross_weight = _round_one_decimal(sum(weights_by_box.values(), Decimal("0")))
    if total_gross_weight <= 0:
        raise RuntimeError(f"装箱数据毛重合计必须大于 0: {source_path.name}")
    return ConsignmentWeightInfo(
        excel_path=str(source_path),
        box_count=box_count,
        total_gross_weight=total_gross_weight,
    )


def _resolve_consignment_excel_path(sp_no: str, consignment_excel: str | Path | None = None) -> Path:
    if consignment_excel:
        path = Path(consignment_excel).expanduser()
        if not path.is_absolute():
            path = (Path.cwd() / path).resolve()
        else:
            path = path.resolve()
        if not path.is_file():
            raise FileNotFoundError(f"装箱数据 Excel 不存在: {path}")
        return path
    return Path(find_consignment_excel(sp_no)).resolve()


def _coerce_input_paths(input_xlsx: str | Path | list[str | Path] | tuple[str | Path, ...]) -> list[Path]:
    if isinstance(input_xlsx, (str, Path)):
        raw_paths: list[str | Path] = [input_xlsx]
    else:
        raw_paths = list(input_xlsx or [])
    paths = [Path(path) for path in raw_paths if _clean_cell(path)]
    if not paths:
        raise ValueError("至少需要提供一个 input-xlsx")
    return paths


def read_input_bundle(
    input_xlsx: str | Path,
    *,
    consignment_excel: str | Path | None = None,
) -> InputDeclarationBundle:
    input_path = Path(input_xlsx)
    sp_no = extract_sp_no_from_filename(input_path)
    destination_country = extract_destination_country_from_filename(input_path)
    source_rows = read_source_rows(input_path)
    consignment_excel_path = _resolve_consignment_excel_path(sp_no, consignment_excel)
    consignment_weight_info = read_consignment_weight_info(consignment_excel_path)
    weight_allocation = allocate_weights_by_quantity(
        source_rows,
        total_gross_weight=consignment_weight_info.total_gross_weight,
    )
    return InputDeclarationBundle(
        input_path=input_path,
        sp_no=sp_no,
        destination_country=destination_country,
        source_rows=source_rows,
        consignment_weight_info=consignment_weight_info,
        weight_allocation=weight_allocation,
    )


def _validate_bundle_destinations(bundles: list[InputDeclarationBundle]) -> str:
    destinations = {bundle.destination_country for bundle in bundles}
    if len(destinations) != 1:
        detail = ", ".join(f"{bundle.sp_no}={bundle.destination_country}" for bundle in bundles)
        raise ValueError(f"多个备货单目的国不一致: {detail}")
    return bundles[0].destination_country


def _combine_weight_allocations(bundles: list[InputDeclarationBundle]) -> WeightAllocation:
    gross_weights: list[Decimal] = []
    net_weights: list[Decimal] = []
    for bundle in bundles:
        gross_weights.extend(bundle.weight_allocation.gross_weights)
        net_weights.extend(bundle.weight_allocation.net_weights)
    return WeightAllocation(gross_weights=gross_weights, net_weights=net_weights)


def _build_output_filename(sp_nos: list[str]) -> str:
    if len(sp_nos) == 1:
        return f"{sp_nos[0]}_custom_declaration_documents.xlsx"
    joined = "_".join(sp_nos)
    if len(sp_nos) > 3 or len(joined) > 120:
        joined = "_".join(sp_nos[:3])
        return f"{joined}_multi_custom_declaration_documents.xlsx"
    return f"{joined}_custom_declaration_documents.xlsx"


def fill_customs_declaration(
    input_xlsx: str | Path | list[str | Path] | tuple[str | Path, ...],
    *,
    template_xlsx: str | Path | None = None,
    output_dir: str | Path | None = None,
    consignment_excel: str | Path | None = None,
) -> dict[str, Any]:
    input_paths = _coerce_input_paths(input_xlsx)
    if len(input_paths) > 1 and consignment_excel:
        raise ValueError("多备货单模式不支持 --consignment-excel；请按 SP 单号准备本地装箱数据")
    bundles = [
        read_input_bundle(input_path, consignment_excel=consignment_excel if len(input_paths) == 1 else None)
        for input_path in input_paths
    ]
    sp_nos = [bundle.sp_no for bundle in bundles]
    destination_country = _validate_bundle_destinations(bundles)
    source_rows = [row for bundle in bundles for row in bundle.source_rows]
    if len(source_rows) > MAX_CUSTOMS_PRODUCT_ROWS:
        raise ValueError(
            "商品数超过报关资料模板容量: "
            f"需要 {len(source_rows)} 行商品，最多支持 {MAX_CUSTOMS_PRODUCT_ROWS} 行商品"
        )
    total_amount = calculate_total_amount(source_rows)
    total_amount_upper = amount_to_chinese_upper_rmb(total_amount)
    weight_allocation = _combine_weight_allocations(bundles)
    box_count = sum(bundle.consignment_weight_info.box_count for bundle in bundles)
    total_gross_weight = _round_one_decimal(
        sum(
            (bundle.consignment_weight_info.total_gross_weight for bundle in bundles),
            Decimal("0"),
        )
    )

    template_path = Path(DEFAULT_TEMPLATE_PATH if template_xlsx is None else template_xlsx)
    if not template_path.is_file():
        raise FileNotFoundError(f"报关资料模板不存在: {template_path}")

    directory = Path(DEFAULT_OUTPUT_DIR if output_dir is None else output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    output_path = directory / _build_output_filename(sp_nos)
    shutil.copy2(template_path, output_path)

    workbook = _load_workbook(output_path)
    if DECLARATION_ELEMENTS_SHEET not in workbook.sheetnames:
        raise ValueError(f"报关资料模板缺少 sheet: {DECLARATION_ELEMENTS_SHEET}")
    if CUSTOMS_DECLARATION_SHEET not in workbook.sheetnames:
        raise ValueError(f"报关资料模板缺少 sheet: {CUSTOMS_DECLARATION_SHEET}")

    worksheet = workbook[DECLARATION_ELEMENTS_SHEET]
    header_row = _find_target_header_row(worksheet)
    _clear_old_declaration_rows(worksheet, header_row=header_row)
    notice = _write_declaration_rows(worksheet, header_row=header_row, rows=source_rows)
    customs_detail_row_count = _write_customs_detail_rows(
        workbook[CUSTOMS_DECLARATION_SHEET],
        rows=source_rows,
        destination_country=destination_country,
        weight_allocation=weight_allocation,
        box_count=box_count,
    )
    formula_sheets = _fill_formula_sheets(workbook, row_count=len(source_rows))
    _write_uppercase_amount(workbook, total_amount_upper)
    workbook.save(output_path)

    payload: dict[str, Any] = {
        "success": True,
        "sp_no": sp_nos[0] if len(sp_nos) == 1 else "_".join(sp_nos),
        "sp_nos": sp_nos,
        "destination_country": destination_country,
        "input_xlsx": str(bundles[0].input_path),
        "input_xlsx_paths": [str(bundle.input_path) for bundle in bundles],
        "output_xlsx": str(output_path),
        "consignment_excel_paths": {
            bundle.sp_no: bundle.consignment_weight_info.excel_path for bundle in bundles
        },
        "box_count": box_count,
        "total_gross_weight": _decimal_to_json_number(total_gross_weight),
        "total_amount": _decimal_to_json_number(total_amount),
        "total_amount_upper": total_amount_upper,
        "row_count": len(source_rows),
        "customs_detail_row_count": customs_detail_row_count,
        "formula_sheet_row_count": len(source_rows),
        "formula_sheets": formula_sheets,
        "unmatched_count": len(notice),
        "notice": notice,
        "source": SOURCE,
    }
    if len(bundles) == 1:
        payload["consignment_excel_path"] = bundles[0].consignment_weight_info.excel_path
    return payload


def build_parser() -> argparse.ArgumentParser:
    parser = JsonArgumentParser(
        prog="python -m services.agent_cli.mabang.fill_customs_declaration"
    )
    parser.add_argument("--input-xlsx", required=True, action="append")
    parser.add_argument("--template-xlsx", default=str(DEFAULT_TEMPLATE_PATH))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--consignment-excel", default="")
    return parser


def main(argv: list[str] | None = None) -> int:
    configure_utf8_stdio()
    try:
        args = build_parser().parse_args(argv)
        payload = fill_customs_declaration(
            getattr(args, "input_xlsx", ""),
            template_xlsx=getattr(args, "template_xlsx", ""),
            output_dir=getattr(args, "output_dir", ""),
            consignment_excel=getattr(args, "consignment_excel", ""),
        )
    except Exception as exc:
        payload = {
            "success": False,
            "exception": _exception_text(exc),
        }

    _write_json(payload)
    return 0 if bool(payload.get("success")) else 1


if __name__ == "__main__":
    raise SystemExit(main())
