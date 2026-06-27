from __future__ import annotations

import csv
import math
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from services.agent_cli.mabang.restock_workbook import clean_cell
from services.amazon.amazon_logistic.sources.consignment_excel import resolve_column

DELIVERY_CSV_DIR = Path("artifacts") / "mabang_fba_delivery"
DELIVERY_MSKU_COLUMN = "MSKU"
SKU_SHIP_QTY_COLUMN = "SKU发货量"
CONSIGNMENT_MSKU_COLUMN = "MSKU"
CONSIGNMENT_QUANTITY_COLUMN = "装箱数量"
CONSIGNMENT_BOX_ALIASES = ("箱序号", "箱子编号", "箱号", "Box No", "Box Number")
CONSIGNMENT_LENGTH_ALIASES = ("长", "长度", "Length", "length")
CONSIGNMENT_WIDTH_ALIASES = ("宽", "Width", "width")
CONSIGNMENT_HEIGHT_ALIASES = ("高", "Height", "height")
CONSIGNMENT_GROSS_WEIGHT_ALIASES = ("毛重", "Gross Weight", "gross_weight", "weight")
ITEM_SPLIT_PATTERN = re.compile(r"[，,\r\n;；]+")
SKU_QTY_PATTERN = re.compile(r"^\s*(?P<sku>.+?)\s*(?:×|x|X|\*)\s*(?P<qty>\d+(?:\.\d+)?)\s*$")
EXPECTED_STOCK_HEADERS = ("SKU", "产品名称", "发货量", "规则型号", "单价")


@dataclass(frozen=True)
class ConsignmentBoxInfo:
    box_no: str
    gross_weight: str
    length: str
    width: str
    height: str


@dataclass(frozen=True)
class ConsignmentMskuRow:
    row_number: int
    box_info: ConsignmentBoxInfo
    msku: str
    quantity: Decimal


@dataclass(frozen=True)
class ConsignmentMskuQuantityRow:
    row_number: int
    msku: str
    quantity: Decimal


def _parse_decimal(value: Any, *, field_name: str, row_context: str) -> Decimal:
    text = clean_cell(value)
    if not text:
        raise ValueError(f"{row_context} 缺少 {field_name}")
    try:
        return Decimal(text)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{row_context} 的 {field_name} 无法解析为数字: {value}") from exc


def _decimal_to_number(value: Decimal | None) -> int | float | str:
    if value is None:
        return ""
    if value == value.to_integral_value():
        return int(value)
    text = format(value.normalize(), "f").rstrip("0").rstrip(".")
    try:
        return float(text)
    except ValueError:
        return text


def _decimal_text(value: Decimal | None) -> str:
    if value is None:
        return ""
    if value == value.to_integral_value():
        return str(int(value))
    return format(value.normalize(), "f").rstrip("0").rstrip(".")


def _normalize_sku_key(value: Any) -> str:
    return clean_cell(value).upper()


def find_latest_delivery_csv(sp_no: str, *, csv_dir: str | Path | None = None) -> Path | None:
    target = clean_cell(sp_no).upper()
    directory = Path(DELIVERY_CSV_DIR if csv_dir is None else csv_dir)
    if not directory.is_dir():
        return None
    candidates = [path for path in directory.glob(f"{target}_*.csv") if path.is_file()]
    if not candidates:
        return None
    return max(candidates, key=lambda path: (path.stat().st_mtime, path.name))


def resolve_delivery_csv_path(sp_no: str, delivery_csv: str | Path | None = None) -> Path:
    if delivery_csv:
        path = Path(delivery_csv).expanduser()
        if not path.is_file():
            raise FileNotFoundError(f"找不到发货单 CSV: {path}")
        return path.resolve()
    path = find_latest_delivery_csv(sp_no)
    if path is None:
        raise FileNotFoundError(f"本地未找到发货单 CSV: {DELIVERY_CSV_DIR / f'{clean_cell(sp_no).upper()}_*.csv'}")
    return path.resolve()


def _parse_sku_quantity_item(raw_item: str, *, row_number: int) -> tuple[str, Decimal]:
    item = str(raw_item or "").strip()
    if not item:
        raise ValueError(f"第{row_number}行 {SKU_SHIP_QTY_COLUMN} 存在空项目")
    match = SKU_QTY_PATTERN.match(item)
    if not match:
        raise ValueError(f"第{row_number}行 {SKU_SHIP_QTY_COLUMN} 格式无法解析: {item}")
    sku = clean_cell(match.group("sku"))
    if not sku:
        raise ValueError(f"第{row_number}行 {SKU_SHIP_QTY_COLUMN} 缺少 SKU: {item}")
    quantity = _parse_decimal(match.group("qty"), field_name="数量", row_context=f"第{row_number}行 {SKU_SHIP_QTY_COLUMN}")
    if quantity < 0:
        raise ValueError(f"第{row_number}行 {SKU_SHIP_QTY_COLUMN} 数量不能小于 0: {item}")
    return sku, quantity


def _read_delivery_rows(csv_path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    source_path = Path(csv_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"找不到发货单 CSV: {source_path}")
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with source_path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                headers = [clean_cell(name) for name in list(reader.fieldnames or [])]
                rows = [{clean_cell(key): clean_cell(value) for key, value in row.items()} for row in reader]
                return headers, rows
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"读取发货单 CSV 失败: {source_path}, error={last_error}") from last_error


def read_delivery_msku_components(csv_path: str | Path) -> OrderedDict[str, OrderedDict[str, Decimal]]:
    headers, rows = _read_delivery_rows(csv_path)
    missing = [column for column in (DELIVERY_MSKU_COLUMN, SKU_SHIP_QTY_COLUMN) if column not in headers]
    if missing:
        raise ValueError(f"发货单 CSV 缺少列: {', '.join(missing)}")

    components: OrderedDict[str, OrderedDict[str, Decimal]] = OrderedDict()
    for index, row in enumerate(rows, start=2):
        msku = clean_cell(row.get(DELIVERY_MSKU_COLUMN))
        cell_value = clean_cell(row.get(SKU_SHIP_QTY_COLUMN))
        if not cell_value:
            continue
        if not msku:
            raise ValueError(f"第{index}行 {SKU_SHIP_QTY_COLUMN} 有值但 MSKU 为空")
        msku_components = components.setdefault(msku, OrderedDict())
        for raw_item in ITEM_SPLIT_PATTERN.split(cell_value):
            item = str(raw_item or "").strip()
            if not item:
                continue
            sku, quantity = _parse_sku_quantity_item(item, row_number=index)
            sku_key = _normalize_sku_key(sku)
            existing_sku = next((key for key in msku_components if _normalize_sku_key(key) == sku_key), sku)
            msku_components[existing_sku] = msku_components.get(existing_sku, Decimal("0")) + quantity
    if not components:
        raise ValueError(f"发货单 CSV 未解析到有效 {DELIVERY_MSKU_COLUMN} + {SKU_SHIP_QTY_COLUMN}")
    return components


def _resolve_required_column(columns: list[str], aliases: tuple[str, ...] | str, *, label: str) -> str:
    alias_tuple = (aliases,) if isinstance(aliases, str) else aliases
    column = resolve_column(columns, alias_tuple)
    if not column:
        raise ValueError(f"装箱数据缺少列: {label}")
    return column


def _consistent_box_info(existing: ConsignmentBoxInfo, current: ConsignmentBoxInfo, *, row_number: int) -> None:
    if existing == current:
        return
    raise ValueError(
        f"装箱数据同一箱序号存在不同箱规: 箱序号={existing.box_no}, "
        f"row={row_number}, existing={existing}, current={current}"
    )


def _load_consignment_dataframe(excel_path: str | Path):
    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError("缺少 pandas 依赖，无法读取装箱数据 Excel") from exc

    source_path = Path(excel_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"找不到装箱数据 Excel: {source_path}")
    try:
        with pd.ExcelFile(source_path) as workbook:
            sheet_name = "FBA装箱任务" if "FBA装箱任务" in workbook.sheet_names else workbook.sheet_names[0]
            df = pd.read_excel(workbook, sheet_name=sheet_name, dtype=str)
    except Exception as exc:
        raise RuntimeError(f"读取装箱数据 Excel 失败: {source_path.name}, error={exc}") from exc
    if df.empty:
        raise ValueError(f"装箱数据没有有效数据: {source_path.name}")
    columns = [clean_cell(column) for column in list(df.columns)]
    df.columns = columns
    return source_path, df, columns


def read_consignment_msku_rows(excel_path: str | Path) -> list[ConsignmentMskuRow]:
    source_path, df, columns = _load_consignment_dataframe(excel_path)
    box_col = _resolve_required_column(columns, CONSIGNMENT_BOX_ALIASES, label="箱序号")
    msku_col = _resolve_required_column(columns, CONSIGNMENT_MSKU_COLUMN, label=CONSIGNMENT_MSKU_COLUMN)
    quantity_col = _resolve_required_column(columns, CONSIGNMENT_QUANTITY_COLUMN, label=CONSIGNMENT_QUANTITY_COLUMN)
    length_col = _resolve_required_column(columns, CONSIGNMENT_LENGTH_ALIASES, label="长")
    width_col = _resolve_required_column(columns, CONSIGNMENT_WIDTH_ALIASES, label="宽")
    height_col = _resolve_required_column(columns, CONSIGNMENT_HEIGHT_ALIASES, label="高")
    weight_col = _resolve_required_column(columns, CONSIGNMENT_GROSS_WEIGHT_ALIASES, label="毛重")

    rows: list[ConsignmentMskuRow] = []
    box_info_by_no: dict[str, ConsignmentBoxInfo] = {}
    for index, row in df.iterrows():
        row_number = int(index) + 2
        box_no = clean_cell(row.get(box_col))
        msku = clean_cell(row.get(msku_col))
        quantity_text = clean_cell(row.get(quantity_col))
        if not any((box_no, msku, quantity_text)):
            continue
        row_context = f"装箱数据第{row_number}行"
        if not box_no:
            raise ValueError(f"{row_context} 缺少箱序号")
        if not msku:
            raise ValueError(f"{row_context} 缺少 MSKU")
        quantity = _parse_decimal(quantity_text, field_name="装箱数量", row_context=row_context)
        if quantity <= 0:
            raise ValueError(f"{row_context} 装箱数量必须大于 0: {quantity_text}")
        box_info = ConsignmentBoxInfo(
            box_no=box_no,
            gross_weight=clean_cell(row.get(weight_col)),
            length=clean_cell(row.get(length_col)),
            width=clean_cell(row.get(width_col)),
            height=clean_cell(row.get(height_col)),
        )
        for field_name, value in (
            ("毛重", box_info.gross_weight),
            ("长", box_info.length),
            ("宽", box_info.width),
            ("高", box_info.height),
        ):
            if not value:
                raise ValueError(f"{row_context} 缺少 {field_name}")
        existing = box_info_by_no.get(box_no)
        if existing is not None:
            _consistent_box_info(existing, box_info, row_number=row_number)
        else:
            box_info_by_no[box_no] = box_info
        rows.append(ConsignmentMskuRow(row_number=row_number, box_info=box_info, msku=msku, quantity=quantity))
    if not rows:
        raise ValueError(f"装箱数据未解析到有效 MSKU 和装箱数量: {source_path.name}")
    return rows


def read_consignment_msku_quantities(excel_path: str | Path) -> OrderedDict[str, Decimal]:
    source_path, df, columns = _load_consignment_dataframe(excel_path)
    msku_col = _resolve_required_column(columns, CONSIGNMENT_MSKU_COLUMN, label=CONSIGNMENT_MSKU_COLUMN)
    quantity_col = _resolve_required_column(columns, CONSIGNMENT_QUANTITY_COLUMN, label=CONSIGNMENT_QUANTITY_COLUMN)

    quantities: OrderedDict[str, Decimal] = OrderedDict()
    for index, row in df.iterrows():
        row_number = int(index) + 2
        msku = clean_cell(row.get(msku_col))
        quantity_text = clean_cell(row.get(quantity_col))
        if not any((msku, quantity_text)):
            continue
        row_context = f"装箱数据第{row_number}行"
        if not msku:
            raise ValueError(f"{row_context} 缺少 MSKU")
        quantity = _parse_decimal(quantity_text, field_name="装箱数量", row_context=row_context)
        if quantity <= 0:
            raise ValueError(f"{row_context} 装箱数量必须大于 0: {quantity_text}")
        quantities[msku] = quantities.get(msku, Decimal("0")) + quantity
    if not quantities:
        raise ValueError(f"装箱数据未解析到有效 MSKU 和装箱数量: {source_path.name}")
    return quantities


def _is_expected_stock_total_or_trailing_row(row_values_by_header: dict[str, str]) -> bool:
    if clean_cell(row_values_by_header.get("SKU")):
        return False
    meaningful_headers = {
        header
        for header, value in row_values_by_header.items()
        if clean_cell(value)
    }
    if not meaningful_headers:
        return True
    return meaningful_headers <= {"发货量"}


def read_expected_stock_sku_quantities(input_xlsx: str | Path) -> OrderedDict[str, Decimal]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取备货单预期数量") from exc

    source_path = Path(input_xlsx).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"输入 xlsx 不存在: {source_path}")
    workbook = load_workbook(source_path, data_only=True, read_only=True)
    try:
        worksheet = workbook.worksheets[0]
        headers = [
            clean_cell(worksheet.cell(row=1, column=column_index).value)
            for column_index in range(1, worksheet.max_column + 1)
        ]
        if "SKU" not in headers:
            raise ValueError(f"备货单第一个 sheet 缺少预期数量表头: SKU")
        sku_col = headers.index("SKU") + 1
        header_indexes: dict[str, int] = {"SKU": sku_col}
        for header in EXPECTED_STOCK_HEADERS:
            if header == "SKU":
                continue
            try:
                offset = headers[sku_col - 1 :].index(header)
            except ValueError as exc:
                raise ValueError(f"备货单第一个 sheet 缺少预期数量表头: {header}") from exc
            header_indexes[header] = sku_col + offset

        quantities: OrderedDict[str, Decimal] = OrderedDict()
        display_skus: dict[str, str] = {}
        for row_number in range(2, worksheet.max_row + 1):
            sku = clean_cell(worksheet.cell(row=row_number, column=header_indexes["SKU"]).value)
            quantity_cell = worksheet.cell(row=row_number, column=header_indexes["发货量"]).value
            row_values_by_header = {
                header: clean_cell(worksheet.cell(row=row_number, column=column_index).value)
                for header, column_index in header_indexes.items()
            }
            if _is_expected_stock_total_or_trailing_row(row_values_by_header):
                continue
            if not sku:
                raise ValueError(f"备货单第一个 sheet 第{row_number}行 SKU 不能为空")
            quantity = _parse_decimal(quantity_cell, field_name="发货量", row_context=f"备货单第一个 sheet 第{row_number}行 SKU={sku}")
            if quantity < 0:
                raise ValueError(f"备货单第一个 sheet 第{row_number}行 SKU={sku} 发货量不能小于 0")
            sku_key = _normalize_sku_key(sku)
            display_skus.setdefault(sku_key, sku)
            quantities[sku_key] = quantities.get(sku_key, Decimal("0")) + quantity
        if not quantities:
            raise ValueError(f"备货单第一个 sheet 未解析到预期库存 SKU 发货量: {source_path.name}")
        return OrderedDict((display_skus[key], value) for key, value in quantities.items())
    finally:
        workbook.close()


def _unit_components(components: OrderedDict[str, Decimal], *, msku: str) -> OrderedDict[str, Decimal]:
    if not components:
        raise ValueError(f"发货单 MSKU 缺少 SKU 组成: {msku}")
    integers: list[int] = []
    for sku, quantity in components.items():
        if quantity <= 0:
            raise ValueError(f"发货单 MSKU 组成数量必须大于 0: MSKU={msku}, SKU={sku}, quantity={_decimal_text(quantity)}")
        if quantity != quantity.to_integral_value():
            raise ValueError(f"发货单 MSKU 组成数量不是整数: MSKU={msku}, SKU={sku}, quantity={_decimal_text(quantity)}")
        integers.append(abs(int(quantity)))
    divisor = 0
    for value in integers:
        divisor = value if divisor == 0 else math.gcd(divisor, value)
    if divisor <= 0:
        raise ValueError(f"发货单 MSKU 无法推导单位组成: {msku}")
    return OrderedDict((sku, quantity / Decimal(divisor)) for sku, quantity in components.items())


def build_actual_stock_sku_quantities(
    delivery_components: OrderedDict[str, OrderedDict[str, Decimal]],
    consignment_quantities: OrderedDict[str, Decimal],
) -> tuple[OrderedDict[str, Decimal], list[str]]:
    actual: OrderedDict[str, Decimal] = OrderedDict()
    display_skus: dict[str, str] = {}
    issues: list[str] = []

    for msku in consignment_quantities:
        components = delivery_components.get(msku)
        if components is None:
            issues.append(f"WMS 装箱数据 MSKU 在发货单 CSV 中不存在: {msku}")
            continue
        try:
            unit_components = _unit_components(components, msku=msku)
        except Exception as exc:
            issues.append(str(exc))
            continue
        for sku, unit_quantity in unit_components.items():
            actual_quantity = consignment_quantities[msku] * unit_quantity
            if actual_quantity != actual_quantity.to_integral_value():
                issues.append(
                    f"MSKU 拆分后库存 SKU 数量不是整数: MSKU={msku}, SKU={sku}, quantity={_decimal_text(actual_quantity)}"
                )
                continue
            sku_key = _normalize_sku_key(sku)
            display_skus.setdefault(sku_key, sku)
            actual[sku_key] = actual.get(sku_key, Decimal("0")) + actual_quantity

    for msku in delivery_components:
        if msku not in consignment_quantities:
            issues.append(f"发货单 CSV MSKU 在 WMS 装箱数据中不存在: {msku}")

    return OrderedDict((display_skus[key], value) for key, value in actual.items()), issues


def build_quantity_validation_rows(
    *,
    sp_no: str,
    expected_quantities: OrderedDict[str, Decimal],
    actual_quantities: OrderedDict[str, Decimal],
    issues: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    expected_by_key = {_normalize_sku_key(sku): (sku, quantity) for sku, quantity in expected_quantities.items()}
    actual_by_key = {_normalize_sku_key(sku): (sku, quantity) for sku, quantity in actual_quantities.items()}
    for sku_key in sorted(set(expected_by_key) | set(actual_by_key)):
        expected_sku, expected = expected_by_key.get(sku_key, (actual_by_key.get(sku_key, (sku_key, Decimal("0")))[0], Decimal("0")))
        actual_sku, actual = actual_by_key.get(sku_key, (expected_sku, Decimal("0")))
        sku = expected_sku or actual_sku
        diff = actual - expected
        if diff == 0:
            status = "一致"
            issue = ""
        elif expected and not actual:
            status = "缺少实际"
            issue = "WMS 实际发货量少于备货单预期"
        elif actual and not expected:
            status = "多出实际"
            issue = "WMS 实际发货量未在备货单预期中找到"
        else:
            status = "差异"
            issue = "预期发货量与实际发货量不一致"
        rows.append(
            {
                "SP单号": sp_no,
                "SKU": sku,
                "预期发货量": expected,
                "实际发货量": actual,
                "差异": diff,
                "状态": status,
                "问题说明": issue,
            }
        )
    for issue in issues:
        rows.append(
            {
                "SP单号": sp_no,
                "SKU": "",
                "预期发货量": None,
                "实际发货量": None,
                "差异": None,
                "状态": "无法校验",
                "问题说明": issue,
            }
        )
    return rows


def summarize_quantity_validation(rows: list[dict[str, Any]]) -> dict[str, int]:
    sku_rows = [row for row in rows if clean_cell(row.get("SKU"))]
    matched_count = sum(1 for row in sku_rows if row.get("状态") == "一致")
    mismatch_count = sum(1 for row in sku_rows if row.get("状态") in {"差异", "缺少实际", "多出实际"})
    unresolved_count = sum(1 for row in rows if row.get("状态") == "无法校验")
    return {
        "total_sku_count": len(sku_rows),
        "matched_count": matched_count,
        "mismatch_count": mismatch_count,
        "unresolved_count": unresolved_count,
    }


def status_from_summary(summary: dict[str, int]) -> str:
    if int(summary.get("unresolved_count") or 0) > 0:
        return "incomplete"
    if int(summary.get("mismatch_count") or 0) > 0:
        return "mismatch"
    return "passed"


def write_quantity_validation_report(
    *,
    output_path: str | Path,
    rows: list[dict[str, Any]],
    source_rows: list[dict[str, Any]],
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法生成数量校验报告") from exc

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "数量校验"
    headers = ["SP单号", "SKU", "预期发货量", "实际发货量", "差异", "状态", "问题说明"]
    worksheet.append(headers)
    for row in rows:
        worksheet.append(
            [
                row.get("SP单号", ""),
                row.get("SKU", ""),
                _decimal_to_number(row.get("预期发货量")),
                _decimal_to_number(row.get("实际发货量")),
                _decimal_to_number(row.get("差异")),
                row.get("状态", ""),
                row.get("问题说明", ""),
            ]
        )
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_letter, width in {
        "A": 18,
        "B": 28,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 52,
    }.items():
        worksheet.column_dimensions[column_letter].width = width
    worksheet.freeze_panes = "A2"

    source_sheet = workbook.create_sheet("数据来源")
    source_sheet.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    source_sheet.append([])
    source_sheet.append(["SP单号", "备货单", "发货单CSV", "WMS装箱数据", "状态", "问题"])
    for source in source_rows:
        source_sheet.append(
            [
                source.get("SP单号", ""),
                source.get("备货单", ""),
                source.get("发货单CSV", ""),
                source.get("WMS装箱数据", ""),
                source.get("状态", ""),
                source.get("问题", ""),
            ]
        )
    for cell in source_sheet[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_letter, width in {
        "A": 18,
        "B": 48,
        "C": 48,
        "D": 48,
        "E": 14,
        "F": 64,
    }.items():
        source_sheet.column_dimensions[column_letter].width = width
    workbook.save(target_path)
    return target_path
