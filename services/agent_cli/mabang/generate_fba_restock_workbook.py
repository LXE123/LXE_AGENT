from __future__ import annotations

import argparse
import asyncio
import re
from collections import OrderedDict
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_cli import (
    JsonArgumentParser,
    configure_utf8_stdio,
    exception_text as _exception_text,
    write_json as _write_json,
)
from services.agent_cli.mabang import generate_restock_workbook as _purchase

DELIVERY_CSV_DIR = _purchase.DELIVERY_CSV_DIR
OUTPUT_DIR = Path("artifacts") / "mabang_restock_workbook"
SOURCE = "fba_restock_workbook"
RESTOCK_SHEET_NAME = "备货单"
COUNTRY_COLUMN = "国家"
DEFAULT_COUNTRY_NAME = "未知国家"
RESTOCK_FILE_LABEL = "新棱镜备货"
INVALID_FILE_NAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
RESTOCK_COLUMNS = (
    "库存sku",
    "产品名称",
    "库存sku（第一行）",
    "产品名称（第一行）",
    "型号",
    "原价",
    "均价",
    "售价",
    "售价(均价)",
    "毛利率",
    "厂家",
    "单位",
    "合同产品名称",
    "数量",
    "总价",
    "总价（均价）",
    "总价（售价）",
    "总价（售价(均价)）",
)
RESTOCK_UNMATCHED_COLUMNS = ("库存sku", "数量", "问题说明")
MIN_GROSS_MARGIN = Decimal("0.2")
MAX_GROSS_MARGIN = Decimal("0.5")
PRICING_BASIS = "tax_exclusive_cost"

close_all_network_clients = _purchase.close_all_network_clients


def _normalize_single_delivery_no(delivery_nos: str | list[str]) -> str:
    values = [delivery_nos] if isinstance(delivery_nos, str) else list(delivery_nos or [])
    if len(values) != 1:
        raise ValueError("备货单一次只能处理一个 --delivery-no；多个 SP 请拆分运行")
    return _purchase._require_delivery_no(values[0])


def _find_required_delivery_csv(delivery_no: str, *, csv_dir: str | Path | None = None) -> Path:
    target = _purchase._require_delivery_no(delivery_no)
    directory = Path(DELIVERY_CSV_DIR if csv_dir is None else csv_dir)
    csv_path = _purchase.find_latest_delivery_csv(target, csv_dir=directory)
    if csv_path is None:
        raise RuntimeError(f"本地未找到发货单 CSV: {directory / f'{target}_*.csv'}")
    return csv_path


def summarize_single_delivery_quantities(
    delivery_no: str,
    *,
    csv_dir: str | Path | None = None,
) -> tuple[OrderedDict[str, Decimal], OrderedDict[str, list[str]], str, str]:
    target = _purchase._require_delivery_no(delivery_no)
    csv_path = _find_required_delivery_csv(target, csv_dir=csv_dir)
    summary = OrderedDict(
        (sku, quantity)
        for sku, quantity in _purchase.summarize_tax_sku_quantities_in_delivery_order(csv_path).items()
        if quantity > 0
    )
    if not summary:
        raise RuntimeError("发货单 CSV 汇总后没有正数 SKU 发货量")
    sku_sources = OrderedDict((sku, [target]) for sku in summary)
    return summary, sku_sources, target, str(csv_path)


def _append_cross_manufacturer_model_warning(
    warnings: list[str],
    summary_rows: list[list[Any]],
) -> int:
    model_manufacturers: OrderedDict[str, OrderedDict[str, None]] = OrderedDict()
    for row in summary_rows:
        model = _purchase._clean_cell(_purchase_row_value(row, "型号"))
        manufacturer = _purchase._clean_cell(_purchase_row_value(row, "厂家"))
        if not model:
            continue
        manufacturers = model_manufacturers.setdefault(model, OrderedDict())
        manufacturers[manufacturer] = None

    conflicts = [
        (model, list(manufacturers))
        for model, manufacturers in model_manufacturers.items()
        if len(manufacturers) > 1
    ]
    if not conflicts:
        return 0

    examples = "; ".join(
        f"{model}: {', '.join(manufacturers)}"
        for model, manufacturers in conflicts[:20]
    )
    warnings.append(
        "不同厂家有相同型号，已保留为不同行，请业务人员核查: "
        f"count={len(conflicts)}, examples={examples}"
    )
    return len(conflicts)


def _safe_file_name_part(value: Any, *, fallback: str) -> str:
    cleaned = INVALID_FILE_NAME_CHARS.sub("_", _purchase._clean_cell(value)).strip(". ")
    return cleaned or fallback


def _date_prefix(value: date) -> str:
    return f"{value.month}.{value.day}"


def _delivery_country_metadata(csv_path: str | Path) -> tuple[str, list[str]]:
    headers, rows = _purchase._read_delivery_rows(csv_path)
    warnings: list[str] = []
    if COUNTRY_COLUMN not in headers:
        warnings.append(f"发货单 CSV 缺少 `{COUNTRY_COLUMN}` 字段，备货单文件名国家已使用 `{DEFAULT_COUNTRY_NAME}`")
        return DEFAULT_COUNTRY_NAME, warnings

    countries: list[str] = []
    for row in rows:
        country = _purchase._clean_cell(row.get(COUNTRY_COLUMN))
        if country and country not in countries:
            countries.append(country)

    if not countries:
        warnings.append(f"发货单 CSV `{COUNTRY_COLUMN}` 字段为空，备货单文件名国家已使用 `{DEFAULT_COUNTRY_NAME}`")
        return DEFAULT_COUNTRY_NAME, warnings

    if len(countries) > 1:
        warnings.append(
            f"发货单 CSV 存在多个不同国家，备货单文件名已使用第一条非空国家 `{countries[0]}`，"
            f"请业务人员核查: {', '.join(countries[:20])}"
        )
    return countries[0], warnings


def _output_file_name(delivery_no: str, *, country: str, today: date) -> str:
    delivery_no_part = _safe_file_name_part(delivery_no, fallback="SP")
    label_part = _safe_file_name_part(RESTOCK_FILE_LABEL, fallback="备货")
    country_part = _safe_file_name_part(country, fallback=DEFAULT_COUNTRY_NAME)
    return f"{_date_prefix(today)}-{delivery_no_part}-{label_part}-{country_part}.xlsx"


def _purchase_row_value(row: list[Any], column: str) -> Any:
    column_index = _purchase.MANUFACTURER_COLUMNS.index(column)
    return row[column_index] if column_index < len(row) else ""


def _parse_gross_margin(value: Any) -> Decimal:
    text = _purchase._clean_cell(value)
    if not text:
        raise ValueError("毛利率必须是 0.2～0.5 之间的数字")
    try:
        gross_margin = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("毛利率必须是 0.2～0.5 之间的数字") from exc
    if not gross_margin.is_finite() or gross_margin < MIN_GROSS_MARGIN or gross_margin > MAX_GROSS_MARGIN:
        raise ValueError("毛利率必须是 0.2～0.5 之间的数字")
    return gross_margin


def _decimal_from_restock_row(value: Any, *, field_name: str, manufacturer: str, model: str) -> Decimal:
    text = _purchase._clean_cell(value)
    if not text:
        raise RuntimeError(f"备货单售价计算缺少{field_name}: 厂家={manufacturer}, 型号={model}")
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise RuntimeError(
            f"备货单售价计算{field_name}非数字: 厂家={manufacturer}, 型号={model}, value={text}"
        ) from exc
    if not result.is_finite():
        raise RuntimeError(f"备货单售价计算{field_name}非数字: 厂家={manufacturer}, 型号={model}, value={text}")
    return result


def _tax_multiplier_from_rate(value: Any, *, manufacturer: str, model: str) -> Decimal:
    text = _purchase._clean_cell(value)
    if not text:
        raise RuntimeError(f"备货单售价计算缺少税率: 厂家={manufacturer}, 型号={model}")
    try:
        if text.endswith("%"):
            tax_rate = Decimal(text[:-1].strip())
            multiplier = Decimal("1") + (tax_rate / Decimal("100"))
        else:
            numeric = Decimal(text)
            if numeric < Decimal("1"):
                multiplier = Decimal("1") + numeric
            elif numeric < Decimal("2"):
                multiplier = numeric
            else:
                multiplier = Decimal("1") + (numeric / Decimal("100"))
    except (InvalidOperation, ValueError) as exc:
        raise RuntimeError(
            f"备货单售价计算税率无法解析: 厂家={manufacturer}, 型号={model}, value={text}"
        ) from exc
    if not multiplier.is_finite() or multiplier <= 0:
        raise RuntimeError(f"备货单售价计算税率无法解析: 厂家={manufacturer}, 型号={model}, value={text}")
    return multiplier


def _sale_price_for_row(row: list[Any], *, gross_margin: Decimal) -> Decimal:
    manufacturer = _purchase._clean_cell(_purchase_row_value(row, "厂家"))
    model = _purchase._clean_cell(_purchase_row_value(row, "型号"))
    original_price = _decimal_from_restock_row(
        _purchase_row_value(row, "原价"),
        field_name="原价",
        manufacturer=manufacturer,
        model=model,
    )
    tax_multiplier = _tax_multiplier_from_rate(
        _purchase_row_value(row, "税率"),
        manufacturer=manufacturer,
        model=model,
    )
    sale_price = original_price / tax_multiplier / (Decimal("1") - gross_margin)
    return sale_price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _sale_price_for_average_row(row: list[Any], *, gross_margin: Decimal) -> Decimal | str:
    manufacturer = _purchase._clean_cell(_purchase_row_value(row, "厂家"))
    if not _purchase._is_zhengfei_manufacturer(manufacturer):
        return ""
    model = _purchase._clean_cell(_purchase_row_value(row, "型号"))
    average_price = _decimal_from_restock_row(
        _purchase_row_value(row, "均价"),
        field_name="均价",
        manufacturer=manufacturer,
        model=model,
    )
    tax_multiplier = _tax_multiplier_from_rate(
        _purchase_row_value(row, "税率"),
        manufacturer=manufacturer,
        model=model,
    )
    sale_price = average_price / tax_multiplier / (Decimal("1") - gross_margin)
    return sale_price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _project_restock_rows(rows: list[list[Any]], *, gross_margin: Decimal) -> list[list[Any]]:
    projected_rows: list[list[Any]] = []
    for row in rows:
        sale_price = _sale_price_for_row(row, gross_margin=gross_margin)
        average_sale_price = _sale_price_for_average_row(row, gross_margin=gross_margin)
        quantity = _decimal_from_restock_row(
            _purchase_row_value(row, "数量"),
            field_name="数量",
            manufacturer=_purchase._clean_cell(_purchase_row_value(row, "厂家")),
            model=_purchase._clean_cell(_purchase_row_value(row, "型号")),
        )
        sale_total_price = (sale_price * quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        average_sale_total_price = (
            (average_sale_price * quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if isinstance(average_sale_price, Decimal)
            else ""
        )
        values = {
            "库存sku": _purchase_row_value(row, "库存sku"),
            "产品名称": _purchase_row_value(row, "产品名称"),
            "库存sku（第一行）": _purchase_row_value(row, "库存sku（第一行）"),
            "产品名称（第一行）": _purchase_row_value(row, "产品名称（第一行）"),
            "型号": _purchase_row_value(row, "型号"),
            "原价": _purchase_row_value(row, "原价"),
            "均价": _purchase_row_value(row, "均价"),
            "售价": _purchase._decimal_to_cell_value(sale_price),
            "售价(均价)": (
                _purchase._decimal_to_cell_value(average_sale_price)
                if isinstance(average_sale_price, Decimal)
                else ""
            ),
            "毛利率": _purchase._decimal_to_cell_value(gross_margin),
            "厂家": _purchase_row_value(row, "厂家"),
            "单位": _purchase_row_value(row, "单位"),
            "合同产品名称": _purchase_row_value(row, "合同产品名称"),
            "数量": _purchase_row_value(row, "数量"),
            "总价": _purchase_row_value(row, "总价"),
            "总价（均价）": _purchase_row_value(row, "总价（均价）"),
            "总价（售价）": _purchase._decimal_to_cell_value(sale_total_price),
            "总价（售价(均价)）": (
                _purchase._decimal_to_cell_value(average_sale_total_price)
                if isinstance(average_sale_total_price, Decimal)
                else ""
            ),
        }
        projected_rows.append([values[column] for column in RESTOCK_COLUMNS])
    return projected_rows


def _drop_unmatched_source_column(rows: list[list[Any]]) -> list[list[Any]]:
    return [
        [value for index, value in enumerate(row) if index != 1]
        for row in rows
    ]


def write_fba_restock_workbook(
    restock_rows: list[list[Any]],
    unmatched_rows: list[list[Any]],
    *,
    delivery_no: str,
    country: str,
    gross_margin: Decimal,
    today: date,
    output_dir: str | Path | None = None,
) -> Path:
    directory = Path(OUTPUT_DIR if output_dir is None else output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    output_path = directory / _output_file_name(delivery_no, country=country, today=today)

    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入 xlsx") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)

    restock_sheet = workbook.create_sheet(RESTOCK_SHEET_NAME)
    _purchase._write_rows(
        restock_sheet,
        RESTOCK_COLUMNS,
        _project_restock_rows(restock_rows, gross_margin=gross_margin),
        append_total=True,
    )

    unmatched_sheet = workbook.create_sheet(_purchase.UNMATCHED_SHEET_NAME)
    _purchase._write_rows(unmatched_sheet, RESTOCK_UNMATCHED_COLUMNS, _drop_unmatched_source_column(unmatched_rows))

    workbook.save(output_path)
    return output_path


def _manufacturer_count(rows: list[list[Any]]) -> int:
    manufacturers: OrderedDict[str, None] = OrderedDict()
    for row in rows:
        manufacturer = _purchase._clean_cell(_purchase_row_value(row, "厂家"))
        manufacturers[manufacturer] = None
    return len(manufacturers)


def generate_fba_restock_workbook(
    delivery_nos: str | list[str],
    *,
    master_xlsx: str | Path,
    gross_margin: Any,
    csv_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    today: date | None = None,
) -> dict[str, Any]:
    parsed_gross_margin = _parse_gross_margin(gross_margin)
    delivery_no = _normalize_single_delivery_no(delivery_nos)
    summary, sku_sources, normalized_delivery_no, csv_path = summarize_single_delivery_quantities(
        delivery_no,
        csv_dir=csv_dir,
    )
    country, country_warnings = _delivery_country_metadata(csv_path)
    products = _purchase.load_master_products(master_xlsx)
    products.warnings.extend(country_warnings)
    restock_rows, _manufacturer_rows, unmatched_rows, matched_sku_count, unmatched_sku_count = (
        _purchase.build_restock_rows(summary, sku_sources, products)
    )
    cross_manufacturer_model_count = _append_cross_manufacturer_model_warning(
        products.warnings,
        restock_rows,
    )
    output_xlsx = write_fba_restock_workbook(
        restock_rows,
        unmatched_rows,
        delivery_no=normalized_delivery_no,
        country=country,
        gross_margin=parsed_gross_margin,
        today=today or date.today(),
        output_dir=output_dir,
    )
    return {
        "success": True,
        "delivery_no": normalized_delivery_no,
        "delivery_nos": [normalized_delivery_no],
        "csv_path": csv_path,
        "csv_paths": [csv_path],
        "country": country,
        "master_xlsx": str(Path(master_xlsx).expanduser()),
        "output_xlsx": str(output_xlsx),
        "sku_count": len(summary),
        "sku_source_count": sum(1 for sources in sku_sources.values() if sources),
        "matched_sku_count": matched_sku_count,
        "unmatched_sku_count": unmatched_sku_count,
        "manufacturer_count": _manufacturer_count(restock_rows),
        "cross_manufacturer_model_count": cross_manufacturer_model_count,
        "gross_margin": str(parsed_gross_margin),
        "pricing_basis": PRICING_BASIS,
        "deduped_duplicate_sku_count": products.deduped_duplicate_sku_count,
        "deduped_duplicate_row_count": products.deduped_duplicate_row_count,
        "deduped_duplicate_sku_examples": products.deduped_duplicate_sku_examples,
        "skipped_empty_sku_row_count": products.skipped_empty_sku_row_count,
        "skipped_empty_sku_rows": products.skipped_empty_sku_rows,
        "unmerged_empty_model_sku_count": products.unmerged_empty_model_sku_count,
        "unmerged_empty_model_skus": products.unmerged_empty_model_skus,
        "contract_mapping_count": products.contract_mapping_count,
        "contract_unmapped_manufacturer_count": products.contract_unmapped_manufacturer_count,
        "contract_unmapped_manufacturer_examples": products.contract_unmapped_manufacturer_examples,
        "contract_conflict_manufacturer_count": products.contract_conflict_manufacturer_count,
        "contract_conflict_manufacturer_examples": products.contract_conflict_manufacturer_examples,
        "warnings": products.warnings,
        "source": SOURCE,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = JsonArgumentParser(
        prog="python -m services.agent_cli.mabang.generate_fba_restock_workbook"
    )
    parser.add_argument("--delivery-no", action="append", default=[])
    parser.add_argument("--master-xlsx", required=True)
    parser.add_argument("--gross-margin", required=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    configure_utf8_stdio()
    delivery_nos: list[str] = []
    master_xlsx = ""
    gross_margin = ""
    try:
        args = build_parser().parse_args(argv)
        delivery_nos = list(getattr(args, "delivery_no", []) or [])
        master_xlsx = str(getattr(args, "master_xlsx", "") or "")
        gross_margin = str(getattr(args, "gross_margin", "") or "")
        payload = generate_fba_restock_workbook(
            delivery_nos,
            master_xlsx=master_xlsx,
            gross_margin=gross_margin,
        )
    except Exception as exc:
        payload = {
            "success": False,
            "delivery_nos": delivery_nos,
            "master_xlsx": master_xlsx,
            "gross_margin": gross_margin,
            "exception": _exception_text(exc),
            "source": SOURCE,
        }
    finally:
        try:
            asyncio.run(close_all_network_clients())
        except Exception:
            pass

    _write_json(payload)
    return 0 if bool(payload.get("success")) else 1


if __name__ == "__main__":
    raise SystemExit(main())
