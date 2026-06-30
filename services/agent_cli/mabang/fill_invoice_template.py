from __future__ import annotations

import argparse
import asyncio
import copy
import json
import sys
from collections import OrderedDict
from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_output import configure_utf8_stdio
from services.agent_cli.mabang import shipment_quantity_validation as quantity_validation
from services.agent_cli.mabang.fill_customs_declaration import (
    SourceDeclarationRow,
    _clean_cell,
    classify_declaration,
    extract_destination_country_from_filename,
    extract_sp_no_from_filename,
)
from services.agent_cli.mabang.restock_workbook import (
    MERGE_DETAIL_HEADERS,
    SUMMARY_HEADERS,
    find_merge_detail_sheet,
    find_summary_sheet,
    summary_column_indexes,
)
from services.agent_cli.mabang.shipment_quantity_validation import (
    CONSIGNMENT_BOX_ALIASES,
    CONSIGNMENT_GROSS_WEIGHT_ALIASES,
    CONSIGNMENT_HEIGHT_ALIASES,
    CONSIGNMENT_LENGTH_ALIASES,
    CONSIGNMENT_MSKU_COLUMN,
    CONSIGNMENT_QUANTITY_COLUMN,
    CONSIGNMENT_WIDTH_ALIASES,
    DELIVERY_CSV_DIR,
    DELIVERY_MSKU_COLUMN,
    ITEM_SPLIT_PATTERN,
    SKU_SHIP_QTY_COLUMN,
    SKU_QTY_PATTERN,
    ConsignmentBoxInfo,
    ConsignmentMskuRow,
    find_latest_delivery_csv,
    read_consignment_msku_rows,
    resolve_delivery_csv_path,
)
from services.amazon.amazon_logistic.sources.consignment_excel import (
    find_consignment_excel,
)
from services.mabang.stock_sku_export import (
    STOCK_SKU_COLUMN,
    export_stock_sku_names,
    normalize_sku_key,
)
from shared.infra.net import close_all_network_clients

SOURCE = "invoice_template_fill"
DEFAULT_TEMPLATE_PATH = Path("data") / "invoice_Template" / "invoice_Template.xlsx"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "invoice_template"
STOCK_SKU_OUTPUT_DIR = Path("artifacts") / "mabang_stock_sku"
INVOICE_TEMPLATE_SHEET = "WS-通用发票导入模版"
STOCK_SKU_IMAGE_COLUMN = "库存sku图片"
IMAGE_MAX_WIDTH = 80
IMAGE_MAX_HEIGHT = 80
IMAGE_ROW_HEIGHT = 65
IMAGE_COLUMN_WIDTH = 14
UNKNOWN_DECLARED_PRICE_TEXT = "没有该材质的计算价格方式"
INPUT_HEADERS = SUMMARY_HEADERS
MERGE_PRICE_QUANTIZE = Decimal("0.01")
REPORT_TOTAL_QUANTIZE = Decimal("0.001")
MATERIAL_TRANSLATIONS = {
    "硅胶": "silicone",
    "尼龙": "nylon",
    "贱金属": "base metal",
    "金属": "metal",
    "纸质+塑料": "paper+plastic",
    "皮革": "leather",
    "PC": "PC",
    "TPU": "TPU",
}

INVOICE_TEMPLATE_HEADERS = (
    "货箱编号*",
    "PO Number*",
    "单件货箱重量(KG)*",
    "货箱长度(CM)*",
    "货箱宽度(CM)*",
    "货箱高度(CM)*",
    "产品英文品名*",
    "产品中文品名*",
    "产品申报单价*",
    "单箱产品申报数量*",
    "单位*",
    "产品材质*",
    "产品海关编码*",
    "产品用途*",
    "产品品牌*",
    "品牌类型*",
    "产品型号*",
    "产品销售链接*",
    "产品图片*",
    "预计配送时间段*",
)


@dataclass(frozen=True)
class InvoiceSourceRow:
    row_number: int
    sku: str
    source_name: str
    model: str
    quantity: Any
    purchase_price: Any
    commodity_name: str
    sale_price: Any
    total_price: Any
    unit: str

    def to_declaration_row(self) -> SourceDeclarationRow:
        return SourceDeclarationRow(
            row_number=self.row_number,
            source_name=self.source_name,
            model=self.model,
            quantity=self.quantity,
            commodity_name=self.commodity_name,
            sale_price=self.sale_price,
            total_price=self.total_price,
            unit=self.unit,
            sku=self.sku,
            purchase_price=self.purchase_price,
        )


@dataclass(frozen=True)
class InvoiceBoxRow:
    source: InvoiceSourceRow
    box_info: ConsignmentBoxInfo
    quantity: Decimal


@dataclass(frozen=True)
class InvoiceActualRowsResult:
    invoice_rows: list[InvoiceBoxRow]
    summary_comparison_rows: list[dict[str, Any]]
    actual_quantities: OrderedDict[str, Decimal]
    quantity_issues: list[quantity_validation.QuantityValidationIssue]


StockSkuMergeInfo = quantity_validation.StockSkuMergeInfo


class JsonArgumentParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:
        raise ValueError(str(message or "").strip() or "参数解析失败")


def _write_json(payload: dict[str, Any]) -> None:
    sys.stdout.write(json.dumps(dict(payload or {}), ensure_ascii=False) + "\n")
    sys.stdout.flush()


def _exception_text(exc: Exception) -> str:
    message = str(exc or "").strip()
    return message or exc.__class__.__name__


def _load_workbook(path: str | Path, *, data_only: bool = False):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法处理 xlsx 文件") from exc

    source_path = Path(path).expanduser()
    try:
        return load_workbook(source_path, data_only=data_only)
    except Exception as exc:
        raise RuntimeError(f"读取 xlsx 文件失败: {source_path}, error={exc}") from exc


def read_invoice_source_rows(input_xlsx: str | Path) -> list[InvoiceSourceRow]:
    path = Path(input_xlsx).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"输入 xlsx 不存在: {path}")

    workbook = _load_workbook(path, data_only=True)
    try:
        worksheet = find_summary_sheet(workbook, path)
        column_indexes = summary_column_indexes(worksheet, input_path=path, sheet_name=worksheet.title)

        rows: list[InvoiceSourceRow] = []
        for row_number in range(2, worksheet.max_row + 1):
            if not _clean_cell(worksheet.cell(row=row_number, column=column_indexes["日期"]).value):
                break
            rows.append(
                InvoiceSourceRow(
                    row_number=row_number,
                    sku=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["SKU"]).value),
                    source_name=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["品名"]).value),
                    model=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["规格型号"]).value),
                    quantity=worksheet.cell(row=row_number, column=column_indexes["发货量"]).value,
                    purchase_price=worksheet.cell(row=row_number, column=column_indexes["单价"]).value,
                    commodity_name=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["商品名称"]).value),
                    sale_price=worksheet.cell(row=row_number, column=column_indexes["售价"]).value,
                    total_price=worksheet.cell(row=row_number, column=column_indexes["总价"]).value,
                    unit=_clean_cell(worksheet.cell(row=row_number, column=column_indexes["单位"]).value),
                )
            )
        if not rows:
            raise ValueError(f"输入 xlsx 的 sheet {worksheet.title} 未解析到有效数据: {path.name}")
        return rows
    finally:
        workbook.close()


def read_stock_sku_merge_infos(input_xlsx: str | Path) -> OrderedDict[str, StockSkuMergeInfo]:
    return quantity_validation.read_stock_sku_merge_infos(input_xlsx)


def infer_missing_summary_models(
    source_rows: list[InvoiceSourceRow],
    stock_sku_merge_infos: OrderedDict[str, StockSkuMergeInfo],
) -> tuple[list[InvoiceSourceRow], list[str]]:
    inferred_rows: list[InvoiceSourceRow] = []
    notices: list[str] = []
    for row in source_rows:
        if _clean_cell(row.model):
            inferred_rows.append(row)
            continue
        sku_key = normalize_sku_key(row.sku)
        if not sku_key:
            raise ValueError(f"汇总表第{row.row_number}行 SKU 不能为空，无法补全规格型号")
        merge_info = stock_sku_merge_infos.get(sku_key)
        if merge_info is None:
            raise ValueError(f"汇总表第{row.row_number}行 SKU={row.sku} 规格型号为空，且财务合并明细表找不到同 SKU")
        merge_model = merge_info.merge_key[0]
        inferred_rows.append(replace(row, model=merge_model))
        notices.append(f"汇总表第{row.row_number}行 SKU={row.sku} 规格型号为空，已按财务合并明细表补为 {merge_model}")
    return inferred_rows, notices


def unique_source_skus(rows: list[InvoiceSourceRow]) -> list[str]:
    unique: OrderedDict[str, str] = OrderedDict()
    for row in rows:
        key = normalize_sku_key(row.sku)
        if not key or key in unique:
            continue
        unique[key] = row.sku
    return list(unique.values())


def resolve_consignment_excel_path(sp_no: str, consignment_excel: str | Path | None = None) -> Path:
    if consignment_excel:
        path = Path(consignment_excel).expanduser()
        if not path.is_file():
            raise FileNotFoundError(f"找不到装箱数据 Excel: {path}")
        return path.resolve()
    return Path(find_consignment_excel(sp_no)).resolve()


def _parse_decimal(value: Any, *, field_name: str, row_context: str) -> Decimal:
    text = _clean_cell(value)
    if not text:
        raise ValueError(f"{row_context} 缺少 {field_name}")
    try:
        return Decimal(text)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"{row_context} 的 {field_name} 无法解析为数字: {value}") from exc


def _box_sort_key(box_no: str) -> tuple[int, str]:
    try:
        return 0, f"{int(Decimal(str(box_no))):08d}"
    except Exception:
        return 1, str(box_no)


def _decimal_to_quantity(value: Decimal, *, context: str) -> Decimal:
    if value != value.to_integral_value():
        raise ValueError(f"{context} 计算得到非整数库存 SKU 数量: {value}")
    return value


def _model_from_merge_info(merge_info: StockSkuMergeInfo) -> str:
    model = _clean_cell(merge_info.merge_key[0])
    if not model:
        raise ValueError(f"财务合并明细表第{merge_info.row_number}行 SKU={merge_info.sku} 缺少规则型号")
    return model


def _source_rows_by_model(
    rows: list[InvoiceSourceRow],
    stock_sku_merge_infos: OrderedDict[str, StockSkuMergeInfo],
) -> tuple[OrderedDict[str, InvoiceSourceRow], dict[int, str]]:
    by_model: OrderedDict[str, InvoiceSourceRow] = OrderedDict()
    model_by_summary_row: dict[int, str] = {}
    duplicates: list[str] = []
    for row in rows:
        sku_key = normalize_sku_key(row.sku)
        if not sku_key:
            raise ValueError(f"汇总表第{row.row_number}行 SKU 不能为空")
        merge_info = stock_sku_merge_infos.get(sku_key)
        if merge_info is None:
            raise ValueError(f"汇总表第{row.row_number}行 SKU={row.sku} 不在备货单第一个表格中，无法确定规则型号")
        model = _model_from_merge_info(merge_info)
        if model in by_model:
            existing_row = by_model[model]
            duplicates.append(
                f"规则型号={model}: 第{existing_row.row_number}行 SKU={existing_row.sku}, "
                f"第{row.row_number}行 SKU={row.sku}"
            )
            continue
        by_model[model] = row
        model_by_summary_row[row.row_number] = model
    if duplicates:
        preview = "; ".join(duplicates[:10])
        suffix = " ..." if len(duplicates) > 10 else ""
        raise ValueError(f"汇总表同一个规则型号存在多个代表 SKU，无法归并: {preview}{suffix}")
    return by_model, model_by_summary_row


def _parse_nonnegative_decimal(value: Any, *, field_name: str, row_context: str) -> Decimal:
    numeric = _parse_decimal(value, field_name=field_name, row_context=row_context)
    if numeric < 0:
        raise ValueError(f"{row_context} 的 {field_name} 不能小于 0: {value}")
    return numeric


def _actual_total(quantity: Decimal, sale_price: Decimal) -> Decimal:
    return (quantity * sale_price).quantize(REPORT_TOTAL_QUANTIZE, rounding=ROUND_HALF_UP)


def _build_summary_comparison_rows(
    *,
    sp_no: str,
    source_rows: list[InvoiceSourceRow],
    model_by_summary_row: dict[int, str],
    actual_by_model: OrderedDict[str, Decimal],
) -> list[dict[str, Any]]:
    comparison_rows: list[dict[str, Any]] = []
    for summary_row in source_rows:
        model = model_by_summary_row[summary_row.row_number]
        row_context = f"汇总表第{summary_row.row_number}行 SKU={summary_row.sku}"
        expected_quantity = _parse_nonnegative_decimal(
            summary_row.quantity,
            field_name="发货量",
            row_context=row_context,
        )
        purchase_price = _parse_nonnegative_decimal(
            summary_row.purchase_price,
            field_name="单价",
            row_context=row_context,
        ).quantize(MERGE_PRICE_QUANTIZE, rounding=ROUND_HALF_UP)
        sale_price = _parse_nonnegative_decimal(
            summary_row.sale_price,
            field_name="售价",
            row_context=row_context,
        )
        original_total = _parse_nonnegative_decimal(
            summary_row.total_price,
            field_name="总价",
            row_context=row_context,
        )
        actual_quantity = actual_by_model.get(model, Decimal("0"))
        recalculated_total = _actual_total(actual_quantity, sale_price)
        diff_quantity = actual_quantity - expected_quantity
        diff_total = recalculated_total - original_total
        if actual_quantity == 0:
            status = "未发货不写入"
        elif actual_quantity == expected_quantity and diff_total == 0:
            status = "一致"
        else:
            status = "数量变化"
        issue = "映射来源：汇总表 SKU 命中第一个表格型号组"
        if status == "未发货不写入":
            issue += "；实际发货量为 0，发票模板不写入该行"
        comparison_rows.append(
            {
                "SP单号": sp_no,
                "汇总表行号": summary_row.row_number,
                "汇总表SKU": summary_row.sku,
                "品名": summary_row.source_name,
                "规格型号": model,
                "单价": purchase_price,
                "售价": sale_price,
                "原发货量": expected_quantity,
                "实际发货量": actual_quantity,
                "数量差异": diff_quantity,
                "原总价": original_total,
                "重算总价": recalculated_total,
                "金额差异": diff_total,
                "状态": status,
                "问题说明": issue,
            }
        )
    return comparison_rows


def build_actual_invoice_box_rows(
    source_rows: list[InvoiceSourceRow],
    stock_sku_merge_infos: OrderedDict[str, StockSkuMergeInfo],
    delivery_components: OrderedDict[str, OrderedDict[str, Decimal]],
    consignment_rows: list[ConsignmentMskuRow],
    *,
    delivery_msku_ship_quantities: dict[str, Decimal | None] | None = None,
    sp_no: str = "",
) -> InvoiceActualRowsResult:
    source_by_model, model_by_summary_row = _source_rows_by_model(source_rows, stock_sku_merge_infos)
    consignment_totals: OrderedDict[str, Decimal] = OrderedDict()
    for row in consignment_rows:
        consignment_totals[row.msku] = consignment_totals.get(row.msku, Decimal("0")) + row.quantity

    actual_quantities, issues = quantity_validation.build_actual_stock_sku_quantities(
        delivery_components,
        consignment_totals,
        delivery_msku_ship_quantities=delivery_msku_ship_quantities,
    )
    blocking_issues = [issue.message for issue in issues if issue.status == "无法校验"]
    if blocking_issues:
        preview = "; ".join(blocking_issues[:10])
        suffix = " ..." if len(blocking_issues) > 10 else ""
        raise ValueError(f"无法按实际发货量生成发票模板: {preview}{suffix}")

    actual_by_model: OrderedDict[str, Decimal] = OrderedDict()
    for sku, actual_quantity in actual_quantities.items():
        if actual_quantity == 0:
            continue
        sku_key = normalize_sku_key(sku)
        merge_info = stock_sku_merge_infos.get(sku_key)
        if merge_info is None:
            raise ValueError(f"实际装箱库存 SKU 不在备货单第一个表格中: SKU={sku}")
        model = _model_from_merge_info(merge_info)
        if model not in source_by_model:
            raise ValueError(
                "实际装箱库存 SKU 型号组在汇总表中没有代表 SKU: "
                f"SKU={sku}, 规则型号={model}"
            )
        actual_by_model[model] = actual_by_model.get(model, Decimal("0")) + actual_quantity

    unit_components_by_msku: dict[str, OrderedDict[str, Decimal]] = {}
    aggregate: OrderedDict[tuple[str, str], dict[str, Any]] = OrderedDict()
    for consignment_row in consignment_rows:
        if consignment_row.msku not in unit_components_by_msku:
            components = delivery_components.get(consignment_row.msku)
            if components is None:
                raise ValueError(f"装箱数据 MSKU 在发货单中不存在: {consignment_row.msku}")
            unit_components_by_msku[consignment_row.msku] = quantity_validation.unit_components_for_msku(
                components,
                msku=consignment_row.msku,
            )
        for sku, unit_quantity in unit_components_by_msku[consignment_row.msku].items():
            quantity = _decimal_to_quantity(
                consignment_row.quantity * unit_quantity,
                context=f"箱序号={consignment_row.box_info.box_no}, MSKU={consignment_row.msku}, SKU={sku}",
            )
            if quantity == 0:
                continue
            sku_key = normalize_sku_key(sku)
            merge_info = stock_sku_merge_infos.get(sku_key)
            if merge_info is None:
                raise ValueError(f"拆分得到的库存 SKU 不在备货单第一个表格中: SKU={sku}, MSKU={consignment_row.msku}")
            model = _model_from_merge_info(merge_info)
            source_row = source_by_model.get(model)
            if source_row is None:
                raise ValueError(
                    "拆分得到的库存 SKU 型号组在汇总表中没有代表 SKU: "
                    f"SKU={sku}, MSKU={consignment_row.msku}, 规则型号={model}"
                )
            aggregate_key = (consignment_row.box_info.box_no, model)
            if aggregate_key not in aggregate:
                aggregate[aggregate_key] = {
                    "source": source_row,
                    "box_info": consignment_row.box_info,
                    "quantity": Decimal("0"),
                }
            aggregate[aggregate_key]["quantity"] += quantity

    invoice_rows = [
        InvoiceBoxRow(
            source=value["source"],
            box_info=value["box_info"],
            quantity=value["quantity"],
        )
        for value in aggregate.values()
    ]
    if not invoice_rows:
        raise ValueError("按装箱数据拆分后未生成发票明细行")

    summary_comparison_rows = _build_summary_comparison_rows(
        sp_no=sp_no,
        source_rows=source_rows,
        model_by_summary_row=model_by_summary_row,
        actual_by_model=actual_by_model,
    )
    sorted_rows = sorted(invoice_rows, key=lambda row: (_box_sort_key(row.box_info.box_no), row.source.row_number))
    return InvoiceActualRowsResult(
        invoice_rows=sorted_rows,
        summary_comparison_rows=summary_comparison_rows,
        actual_quantities=actual_quantities,
        quantity_issues=issues,
    )


def build_invoice_box_rows(
    source_rows: list[InvoiceSourceRow],
    stock_sku_merge_infos: OrderedDict[str, StockSkuMergeInfo],
    delivery_components: OrderedDict[str, OrderedDict[str, Decimal]],
    consignment_rows: list[ConsignmentMskuRow],
) -> list[InvoiceBoxRow]:
    return build_actual_invoice_box_rows(
        source_rows,
        stock_sku_merge_infos,
        delivery_components,
        consignment_rows,
    ).invoice_rows


def translate_commodity_name(row: InvoiceSourceRow) -> str:
    name = row.commodity_name
    if "编织表带" in name:
        return "Braided Watch Band"
    if "表带" in name:
        return "Watch Band"
    if "表壳" in name:
        return "Watch Case"
    if "包装盒" in name:
        return "Packaging Box"
    if "手表保护套" in name:
        return "Watch Protective Case"
    return ""


def translated_material(material: str) -> str:
    material_text = _clean_cell(material)
    if not material_text:
        return ""
    english = MATERIAL_TRANSLATIONS.get(material_text)
    if not english:
        return ""
    return f"{material_text}/{english}"


def usage_for(row: InvoiceSourceRow) -> str:
    name = row.commodity_name
    if "表带" in name:
        return "装饰/decorate"
    if "表壳" in name:
        return "装饰/decorate"
    if "手表保护套" in name:
        return "保护手表用/Used for protecting watches"
    if "包装盒" in name:
        return "装表带表壳/Watch strap and case"
    return ""


def _decimal_to_cell_value(value: Decimal) -> int | float | str:
    if value == value.to_integral_value():
        return int(value)
    text = format(value.normalize(), "f").rstrip("0").rstrip(".")
    try:
        return float(text)
    except ValueError:
        return text


def declared_price_for(row: InvoiceSourceRow, material: str) -> int | float | str:
    material_text = _clean_cell(material)
    unit_price: Decimal | None = None
    if "表带" in row.commodity_name:
        if material_text == "硅胶":
            unit_price = Decimal("0.35")
        elif material_text == "尼龙":
            unit_price = Decimal("0.5")
        elif material_text == "皮革":
            unit_price = Decimal("0.5")
        elif material_text in {"金属", "贱金属"}:
            unit_price = Decimal("1")
    elif "表壳" in row.commodity_name:
        unit_price = Decimal("0.35")
    elif "包装盒" in row.commodity_name:
        unit_price = Decimal("0.32")
    elif "手表保护套" in row.commodity_name:
        unit_price = Decimal("0.35")
    if unit_price is not None:
        return _decimal_to_cell_value(unit_price)
    return UNKNOWN_DECLARED_PRICE_TEXT


def _find_invoice_template_header_row(worksheet: Any) -> int:
    expected = list(INVOICE_TEMPLATE_HEADERS)
    for row_index in range(1, worksheet.max_row + 1):
        actual = [
            _clean_cell(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, len(expected) + 1)
        ]
        if actual == expected:
            return row_index
    raise ValueError(f"{INVOICE_TEMPLATE_SHEET} sheet 缺少表头: {expected}")


def _copy_row_style(worksheet: Any, *, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column_index in range(1, len(INVOICE_TEMPLATE_HEADERS) + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)


def _clear_data_rows(worksheet: Any, *, header_row: int) -> None:
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        for column_index in range(1, len(INVOICE_TEMPLATE_HEADERS) + 1):
            worksheet.cell(row=row_index, column=column_index).value = None


def _image_bytes(image: Any) -> bytes:
    try:
        return image._data()
    except Exception as exc:
        raise RuntimeError(f"读取库存SKU图片失败: {exc}") from exc


def load_stock_sku_images(xlsx_paths: list[str] | tuple[str, ...]) -> dict[str, bytes]:
    images_by_key: dict[str, bytes] = {}
    for raw_path in xlsx_paths:
        source_path = Path(raw_path).expanduser()
        if not source_path.is_file():
            raise FileNotFoundError(f"找不到库存SKU导出xlsx: {source_path}")
        workbook = _load_workbook(source_path, data_only=True)
        try:
            worksheet = workbook.worksheets[0]
            headers = [_clean_cell(worksheet.cell(row=1, column=column_index).value) for column_index in range(1, worksheet.max_column + 1)]
            if STOCK_SKU_COLUMN not in headers:
                raise RuntimeError(f"库存SKU导出xlsx缺少列: {STOCK_SKU_COLUMN}")
            if STOCK_SKU_IMAGE_COLUMN not in headers:
                raise RuntimeError(f"库存SKU导出xlsx缺少列: {STOCK_SKU_IMAGE_COLUMN}")
            sku_col = headers.index(STOCK_SKU_COLUMN) + 1
            image_col = headers.index(STOCK_SKU_IMAGE_COLUMN) + 1
            for image in list(getattr(worksheet, "_images", []) or []):
                marker = getattr(getattr(image, "anchor", None), "_from", None)
                if marker is None:
                    continue
                image_row = int(marker.row) + 1
                image_column = int(marker.col) + 1
                if image_column != image_col:
                    continue
                sku = _clean_cell(worksheet.cell(row=image_row, column=sku_col).value)
                key = normalize_sku_key(sku)
                if key and key not in images_by_key:
                    images_by_key[key] = _image_bytes(image)
        finally:
            workbook.close()
    return images_by_key


def _add_image_to_cell(worksheet: Any, *, image_bytes: bytes, row: int, column: int) -> None:
    try:
        from openpyxl.drawing.image import Image
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 图片依赖，无法写入产品图片") from exc

    image = Image(BytesIO(image_bytes))
    if image.width and image.height:
        scale = min(IMAGE_MAX_WIDTH / image.width, IMAGE_MAX_HEIGHT / image.height, 1)
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
    worksheet.row_dimensions[row].height = max(float(worksheet.row_dimensions[row].height or 0), IMAGE_ROW_HEIGHT)
    column_letter = get_column_letter(column)
    worksheet.column_dimensions[column_letter].width = max(
        float(worksheet.column_dimensions[column_letter].width or 0),
        IMAGE_COLUMN_WIDTH,
    )
    worksheet.add_image(image, f"{column_letter}{row}")


def write_invoice_template(
    rows: list[InvoiceBoxRow],
    *,
    sp_no: str,
    destination_country: str,
    stock_sku_xlsx_paths: list[str],
    template_path: str | Path | None = None,
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    source_template = Path(DEFAULT_TEMPLATE_PATH if template_path is None else template_path).expanduser()
    if not source_template.is_file():
        raise FileNotFoundError(f"找不到发票模板: {source_template}")
    target_dir = Path(DEFAULT_OUTPUT_DIR if output_dir is None else output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    output_path = target_dir / f"{sp_no}_invoice_Template.xlsx"

    workbook = _load_workbook(source_template)
    try:
        if INVOICE_TEMPLATE_SHEET not in workbook.sheetnames:
            raise ValueError(f"发票模板缺少 sheet: {INVOICE_TEMPLATE_SHEET}")
        worksheet = workbook[INVOICE_TEMPLATE_SHEET]
        header_row = _find_invoice_template_header_row(worksheet)
        first_data_row = header_row + 1
        header_to_col = {
            _clean_cell(worksheet.cell(row=header_row, column=column_index).value): column_index
            for column_index in range(1, len(INVOICE_TEMPLATE_HEADERS) + 1)
        }
        _clear_data_rows(worksheet, header_row=header_row)
        images_by_key = load_stock_sku_images(stock_sku_xlsx_paths)

        notice: list[str] = []
        image_matched_count = 0
        image_missing_count = 0
        for offset, invoice_row in enumerate(rows):
            row = invoice_row.source
            target_row = first_data_row + offset
            _copy_row_style(worksheet, source_row=first_data_row, target_row=target_row)
            classification = classify_declaration(row.to_declaration_row())
            material = classification.declaration_element
            hs_code = classification.hs_code
            if not material or not hs_code:
                notice.append(f"第{row.row_number}行未匹配申报规则: SKU={row.sku}, 商品名称={row.commodity_name}, 品名={row.source_name}")
            english_name = translate_commodity_name(row)
            if not english_name:
                notice.append(f"第{row.row_number}行缺少英文品名规则: SKU={row.sku}, 商品名称={row.commodity_name}")
            declared_price = declared_price_for(row, material)
            if declared_price == UNKNOWN_DECLARED_PRICE_TEXT:
                notice.append(f"第{row.row_number}行没有该材质的计算价格方式: SKU={row.sku}, 产品材质={material or ''}")
            material_value = translated_material(material)
            if material and not material_value:
                notice.append(f"第{row.row_number}行缺少产品材质英文映射: SKU={row.sku}, 产品材质={material}")
            usage = usage_for(row)
            if not usage:
                notice.append(f"第{row.row_number}行缺少产品用途规则: SKU={row.sku}, 商品名称={row.commodity_name}")

            values = {
                "货箱编号*": invoice_row.box_info.box_no,
                "单件货箱重量(KG)*": invoice_row.box_info.gross_weight,
                "货箱长度(CM)*": invoice_row.box_info.length,
                "货箱宽度(CM)*": invoice_row.box_info.width,
                "货箱高度(CM)*": invoice_row.box_info.height,
                "产品英文品名*": english_name,
                "产品中文品名*": row.commodity_name,
                "产品申报单价*": declared_price,
                "单箱产品申报数量*": _decimal_to_cell_value(invoice_row.quantity),
                "单位*": row.unit,
                "产品材质*": material_value,
                "产品海关编码*": hs_code,
                "产品用途*": usage,
                "产品品牌*": "无",
                "品牌类型*": "无",
                "产品型号*": row.sku,
            }
            for header, value in values.items():
                worksheet.cell(row=target_row, column=header_to_col[header], value=value)

            image_key = normalize_sku_key(row.sku)
            image_data = images_by_key.get(image_key)
            if image_data:
                _add_image_to_cell(
                    worksheet,
                    image_bytes=image_data,
                    row=target_row,
                    column=header_to_col["产品图片*"],
                )
                image_matched_count += 1
            else:
                image_missing_count += 1
                notice.append(f"第{row.row_number}行缺少产品图片: SKU={row.sku}")

        workbook.save(output_path)
    finally:
        workbook.close()

    return {
        "success": True,
        "sp_no": sp_no,
        "destination_country": destination_country,
        "output_xlsx": str(output_path),
        "row_count": len(rows),
        "stock_sku_xlsx_paths": list(stock_sku_xlsx_paths),
        "image_matched_count": image_matched_count,
        "image_missing_count": image_missing_count,
        "notice": notice,
        "source": SOURCE,
    }


def _empty_quantity_validation_summary() -> dict[str, int]:
    return {
        "total_sku_count": 0,
        "matched_count": 0,
        "mismatch_count": 0,
        "unresolved_count": 0,
        "not_shipped_msku_count": 0,
    }


def _build_validation_report_filename(sp_no: str) -> str:
    return f"{sp_no}_invoice_quantity_validation_report.xlsx"


def _build_invoice_quantity_validation_report(
    *,
    input_path: Path,
    sp_no: str,
    delivery_csv_path: Path,
    consignment_excel_path: Path,
    delivery_components: OrderedDict[str, OrderedDict[str, Decimal]],
    actual_result: InvoiceActualRowsResult,
    output_dir: str | Path,
) -> dict[str, Any]:
    expected_quantities = quantity_validation.read_expected_stock_sku_quantities(input_path)
    rows = quantity_validation.build_quantity_validation_rows(
        sp_no=sp_no,
        expected_quantities=expected_quantities,
        actual_quantities=actual_result.actual_quantities,
        issues=actual_result.quantity_issues,
        stock_sku_msku_sources=quantity_validation.build_stock_sku_msku_sources(delivery_components),
    )
    summary = quantity_validation.summarize_quantity_validation(rows)
    status = quantity_validation.status_from_summary(summary)
    source_info = {
        "SP单号": sp_no,
        "备货单": str(input_path),
        "发货单CSV": str(delivery_csv_path),
        "WMS装箱数据": str(consignment_excel_path),
        "状态": status,
        "问题": "; ".join(
            str(row.get("问题说明") or "")
            for row in rows
            if row.get("状态") in {"差异", "缺少实际", "多出实际", "无法校验"}
        ),
    }
    report_path = quantity_validation.write_quantity_validation_report(
        output_path=Path(output_dir) / _build_validation_report_filename(sp_no),
        rows=rows,
        source_rows=[source_info],
        summary_comparison_rows=actual_result.summary_comparison_rows,
    )
    return {
        "validation_report_xlsx": str(report_path),
        "quantity_validation_status": status,
        "quantity_validation_summary": summary,
    }


async def fill_invoice_template(
    input_xlsx: str | Path,
    *,
    template_path: str | Path | None = None,
    output_dir: str | Path | None = None,
    stock_sku_output_dir: str | Path | None = None,
    delivery_csv: str | Path | None = None,
    consignment_excel: str | Path | None = None,
) -> dict[str, Any]:
    input_path = Path(input_xlsx).expanduser()
    sp_no = extract_sp_no_from_filename(input_path)
    destination_country = extract_destination_country_from_filename(input_path)
    source_rows = read_invoice_source_rows(input_path)
    stock_sku_merge_infos = read_stock_sku_merge_infos(input_path)
    source_rows, inferred_model_notices = infer_missing_summary_models(source_rows, stock_sku_merge_infos)
    delivery_csv_path = resolve_delivery_csv_path(sp_no, delivery_csv)
    consignment_excel_path = resolve_consignment_excel_path(sp_no, consignment_excel)
    delivery_infos = quantity_validation.read_delivery_msku_infos(delivery_csv_path)
    delivery_components = OrderedDict((msku, info.components) for msku, info in delivery_infos.items())
    delivery_msku_ship_quantities = {
        msku: info.msku_ship_quantity
        for msku, info in delivery_infos.items()
    }
    consignment_rows = read_consignment_msku_rows(consignment_excel_path)
    actual_result = build_actual_invoice_box_rows(
        source_rows,
        stock_sku_merge_infos,
        delivery_components,
        consignment_rows,
        delivery_msku_ship_quantities=delivery_msku_ship_quantities,
        sp_no=sp_no,
    )
    invoice_rows = actual_result.invoice_rows
    skus = unique_source_skus([row.source for row in invoice_rows])
    stock_result = await export_stock_sku_names(
        skus,
        delivery_no=f"{sp_no}_invoice",
        output_dir=STOCK_SKU_OUTPUT_DIR if stock_sku_output_dir is None else stock_sku_output_dir,
    )
    stock_sku_xlsx_paths = list(getattr(stock_result, "xlsx_paths", []) or [])
    payload = write_invoice_template(
        invoice_rows,
        sp_no=sp_no,
        destination_country=destination_country,
        stock_sku_xlsx_paths=stock_sku_xlsx_paths,
        template_path=template_path,
        output_dir=output_dir,
    )
    try:
        validation_payload = _build_invoice_quantity_validation_report(
            input_path=input_path,
            sp_no=sp_no,
            delivery_csv_path=delivery_csv_path,
            consignment_excel_path=consignment_excel_path,
            delivery_components=delivery_components,
            actual_result=actual_result,
            output_dir=DEFAULT_OUTPUT_DIR if output_dir is None else output_dir,
        )
    except Exception as exc:
        validation_payload = {
            "quantity_validation_status": "error",
            "quantity_validation_summary": _empty_quantity_validation_summary(),
            "validation_report_error": _exception_text(exc),
        }
    payload["notice"] = inferred_model_notices + list(payload.get("notice", []))
    payload["input_xlsx"] = str(input_path)
    payload["delivery_csv_path"] = str(delivery_csv_path)
    payload["consignment_excel_path"] = str(consignment_excel_path)
    payload["box_count"] = len({row.box_info.box_no for row in invoice_rows})
    payload["source_row_count"] = len(source_rows)
    payload["invoice_row_count"] = len(invoice_rows)
    payload["quantity_basis"] = "actual"
    payload.update(validation_payload)
    payload["merge_group_count"] = len({
        _model_from_merge_info(stock_sku_merge_infos[normalize_sku_key(row.sku)])
        for row in source_rows
    })
    payload["merged_sku_count"] = max(0, len(stock_sku_merge_infos) - payload["merge_group_count"])
    return payload


def build_parser() -> argparse.ArgumentParser:
    parser = JsonArgumentParser(
        prog="python -m services.agent_cli.mabang.fill_invoice_template"
    )
    parser.add_argument("--input-xlsx", default="")
    parser.add_argument("--delivery-csv", default="")
    parser.add_argument("--consignment-excel", default="")
    return parser


async def _run_async(args: argparse.Namespace) -> dict[str, Any]:
    input_xlsx = str(getattr(args, "input_xlsx", "") or "").strip()
    if not input_xlsx:
        raise ValueError("input_xlsx 不能为空")
    return await fill_invoice_template(
        input_xlsx,
        delivery_csv=str(getattr(args, "delivery_csv", "") or "").strip() or None,
        consignment_excel=str(getattr(args, "consignment_excel", "") or "").strip() or None,
    )


def main(argv: list[str] | None = None) -> int:
    configure_utf8_stdio()
    try:
        args = build_parser().parse_args(argv)
        payload = asyncio.run(_run_async(args))
    except Exception as exc:
        payload = {
            "success": False,
            "exception": _exception_text(exc),
        }
    finally:
        try:
            asyncio.run(close_all_network_clients())
        except Exception:
            pass

    _write_json(payload)
    return 0 if bool(payload.get("success")) else 1


if __name__ == "__main__":
    raise SystemExit(main())
