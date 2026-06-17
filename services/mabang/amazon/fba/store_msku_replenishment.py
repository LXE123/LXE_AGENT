from __future__ import annotations

import math
import re
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any

from services.mabang import config as mabang_settings
from services.mabang.amazon.fba.amazon_inventory import (
    AMAZON_FBA_TOTAL_COLUMN,
    load_amazon_inventory_snapshot,
)
from services.mabang.amazon.fba.unlinked_shipments import (
    DEFAULT_SNAPSHOT_DIR as DEFAULT_UNLINKED_SHIPMENTS_SNAPSHOT_DIR,
    load_unlinked_shipment_quantities,
)
from services.mabang.amazon.fba.replenishment_template import (
    DEFAULT_TEMPLATE_NAME,
    ReplenishmentTemplate,
    calculate_weighted_daily_sales,
    effective_params_for_msku,
    get_template,
    replenishment_days_from_template,
    sea_companion_air_day_candidates_from_template,
    sea_companion_air_enabled_from_template,
    sea_day_candidates_from_template,
    sea_daily_sales_meets_min_from_template,
    sea_enabled_from_template,
    sea_min_daily_sales_inclusive_from_template,
    sea_min_net_quantity_from_template,
    trend_group_from_template,
    validate_template,
)

DEFAULT_SALES_ANALYSIS_DIR = Path("artifacts") / "mabang_store_msku_analysis"
DEFAULT_ACTUAL_INVENTORY_DIR = Path("artifacts") / "mabang_store_msku_inventory"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "mabang_store_msku_replenishment"
SOURCE = "mabang_store_msku_replenishment"
EXCEL_ROW_HEIGHT = 15
EXCEL_COLUMN_WIDTH = 15
SALES_REPORT_RE = re.compile(r"^(?P<source_time>\d{12})-(?P<store>.+)_sales_analysis\.xlsx$", re.IGNORECASE)
INVENTORY_REPORT_RE = re.compile(r"^(?P<source_time>\d{12})-(?P<store>.+)_actual_inventory\.xlsx$", re.IGNORECASE)
UNLINKED_SNAPSHOT_RE = re.compile(r"^(?P<snapshot_time>\d{12})-(?P<store>.+)_unlinked_shipments_snapshot\.xlsx$", re.IGNORECASE)
UNLINKED_SNAPSHOT_MISSING_WARNING = "未找到与备货数据同日的未关联货件快照，本次未扣减未关联货件"
UNLINKED_SNAPSHOT_IGNORED_NON_SAME_DAY_WARNING = "未找到与备货数据同日的未关联货件快照，已忽略非同日未关联货件快照，本次未扣减未关联货件"
AMAZON_DEDUCTED_REPLENISH_COLUMN = "补货量（减去 FBA 总库存[亚马逊后台数据]和未关联货件）"
DETAIL_SHEET = "MSKU明细"
SUMMARY_SHEET = "链接备货汇总"
INVENTORY_SHORTAGE_SHEET = "真实库存不足"
CLEARANCE_SHEET = "清货"
AIR_URGENT_SHEET = "空运（急发）"
AIR_SHEET = "空运"
SEA_SHEET = "海运"
NO_SHIP_SHEET = "暂不建议发货"
SAMPLE_INSUFFICIENT_SHEET = "样本不足"
CLEARANCE_KEYWORD = "清货"
INVENTORY_SHEETS = ("真实库存-组合sku", "真实库存-库存sku")
REPORT_SHEETS = (
    AIR_URGENT_SHEET,
    AIR_SHEET,
    SEA_SHEET,
    INVENTORY_SHORTAGE_SHEET,
    CLEARANCE_SHEET,
    NO_SHIP_SHEET,
    SUMMARY_SHEET,
    SAMPLE_INSUFFICIENT_SHEET,
)
DETAIL_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    "本地SKU名称",
    "产品名称",
    "备注",
    "子SKU",
    "商品链接",
    "SKU类型",
    "模板名称",
    "命中规则",
    "销量趋势",
    "趋势分组",
    "加权日销",
    "可销售天数",
    "FBA总库存",
    "未关联数量",
    "真实库存数量",
    "单品重量(g)",
    "补货天数",
    "补货量",
    "补货量（减去 FBA 总库存）",
    "补货量（减去 FBA 总库存和未关联货件）",
    AMAZON_FBA_TOTAL_COLUMN,
    AMAZON_DEDUCTED_REPLENISH_COLUMN,
    "海运天数",
    "海运建议量",
    "同时空运天数",
    "同时空运建议量",
    "预计总重量kg",
    "决策原因",
)
AIR_DETAIL_COLUMNS = tuple(
    column
    for column in DETAIL_COLUMNS
    if column not in {"海运天数", "海运建议量", "同时空运天数", "同时空运建议量"}
)
INVENTORY_SHORTAGE_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    "本地SKU名称",
    "产品名称",
    "备注",
    "子SKU",
    "商品链接",
    "SKU类型",
    "运输渠道",
    "模板名称",
    "命中规则",
    "销量趋势",
    "趋势分组",
    "加权日销",
    "可销售天数",
    "FBA总库存",
    "未关联数量",
    "真实库存数量",
    "单品重量(g)",
    "补货天数",
    "补货量",
    "补货量（减去 FBA 总库存）",
    "补货量（减去 FBA 总库存和未关联货件）",
    AMAZON_FBA_TOTAL_COLUMN,
    AMAZON_DEDUCTED_REPLENISH_COLUMN,
    "库存缺口",
    "海运天数",
    "海运建议量",
    "同时空运天数",
    "同时空运建议量",
    "预计总重量kg",
    "决策原因",
)
CLEARANCE_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    "本地SKU名称",
    "产品名称",
    "备注",
    "子SKU",
    "商品链接",
    "SKU类型",
    "运输渠道",
    "模板名称",
    "命中规则",
    "销量趋势",
    "趋势分组",
    "加权日销",
    "可销售天数",
    "FBA总库存",
    "未关联数量",
    "真实库存数量",
    "单品重量(g)",
    "补货天数",
    "补货量",
    "补货量（减去 FBA 总库存）",
    "补货量（减去 FBA 总库存和未关联货件）",
    AMAZON_FBA_TOTAL_COLUMN,
    AMAZON_DEDUCTED_REPLENISH_COLUMN,
    "海运天数",
    "海运建议量",
    "同时空运天数",
    "同时空运建议量",
    "预计总重量kg",
    "决策原因",
)
SUMMARY_COLUMNS = (
    "父ASIN",
    "商品链接",
    "MSKU数",
    "最大加权日销",
    "合计加权日销",
    "最小可销售天数",
    "链接未关联数量汇总",
    "链接真实本地库存汇总",
    "总补货量",
    "空运（急发）补货量",
    "空运补货量",
    "海运建议量",
    "涉及运输方式",
    "决策备注",
)
INVENTORY_REQUIRED_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    "商品链接",
    "FBA总库存",
    "加权日销",
    "可销售天数",
    "真实库存数量",
    "子SKU",
)
SALES_REQUIRED_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    "7天销量",
    "14天销量",
    "30天销量",
    "销量趋势",
    "单品重量(g)(cm)",
)
TRANSPORT_ORDER = (AIR_URGENT_SHEET, AIR_SHEET, SEA_SHEET, NO_SHIP_SHEET, SAMPLE_INSUFFICIENT_SHEET)
FLOAT_TOLERANCE = 1e-9
TWO_DECIMAL_COLUMNS = {
    "加权日销",
    "可销售天数",
    "单品重量(g)",
    "预计总重量kg",
    "最大加权日销",
    "合计加权日销",
    "最小可销售天数",
}
INTEGER_COLUMNS = {
    "MSKU数",
    "FBA总库存",
    "真实库存数量",
    "总补货量",
    "空运（急发）补货量",
    "空运补货量",
    "海运建议量",
    "补货天数",
    "补货量",
    "补货量（减去 FBA 总库存）",
    "补货量（减去 FBA 总库存和未关联货件）",
    AMAZON_FBA_TOTAL_COLUMN,
    AMAZON_DEDUCTED_REPLENISH_COLUMN,
    "海运天数",
    "同时空运天数",
    "同时空运建议量",
    "未关联数量",
    "链接未关联数量汇总",
    "链接真实本地库存汇总",
    "库存缺口",
}
WEIGHT_RE = re.compile(r"[-+]?\d+(?:,\d{3})*(?:\.\d+)?|[-+]?\d+(?:\.\d+)?")
URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)


class StoreMskuReplenishmentError(ValueError):
    pass


@dataclass(frozen=True)
class ReportFile:
    path: Path
    source_data_time: str
    source_datetime: datetime


@dataclass(frozen=True)
class MatchedReports:
    source_data_time: str
    sales_analysis_path: Path
    actual_inventory_path: Path


@dataclass(frozen=True)
class UnlinkedSnapshotFile:
    path: Path
    snapshot_time: str
    snapshot_datetime: datetime
    store_part: str


@dataclass(frozen=True)
class SalesDetail:
    trend: str
    weight_grams: float | None
    sales_7d: float
    sales_14d: float
    sales_30d: float


@dataclass(frozen=True)
class InventoryInputRow:
    msku: str
    parent_asin: str
    asin: str
    local_sku: str
    local_sku_name: str
    product_name: str
    remark: str
    product_link: str
    sku_type: str
    weighted_daily_sales: float
    sales_days: float | None
    fba_total_inventory: float
    actual_inventory: float | None
    child_skus: str


@dataclass(frozen=True)
class ReplenishmentRow:
    msku: str
    parent_asin: str
    asin: str
    local_sku: str
    local_sku_name: str
    product_name: str
    remark: str
    product_link: str
    sku_type: str
    template_name: str
    matched_rule: str
    sales_trend: str
    trend_group: str
    weighted_daily_sales: float
    sales_days: float | None
    fba_total_inventory: float
    unlinked_quantity: float
    actual_inventory: float | None
    weight_grams: float | None
    replenish_days: int | None
    replenish_quantity: int | None
    original_replenish_quantity: int | None
    sea_days: int | None
    sea_quantity: int | None
    estimated_weight_kg: float | None
    decision_reason: str
    child_skus: str
    sheet_name: str
    companion_air_days: int | None = None
    companion_air_quantity: int | None = None
    sea_net_quantity: int | None = None
    fba_deducted_replenish_quantity: int | None = None
    amazon_fba_total_inventory: float | None = None
    amazon_deducted_replenish_quantity: int | None = None
    transport_channel: str = ""

    def to_detail_payload(self) -> dict[str, Any]:
        display_weight_kg = _report_estimated_weight_kg(self)
        fba_deducted_quantity = self.fba_deducted_replenish_quantity
        if fba_deducted_quantity is None:
            fba_deducted_quantity = self.replenish_quantity
        return {
            "MSKU": self.msku,
            "父ASIN": self.parent_asin,
            "ASIN": self.asin,
            "本地SKU": self.local_sku,
            "本地SKU名称": self.local_sku_name,
            "产品名称": self.product_name,
            "备注": self.remark,
            "子SKU": self.child_skus,
            "商品链接": self.product_link,
            "SKU类型": self.sku_type,
            "运输渠道": self.transport_channel,
            "模板名称": self.template_name,
            "命中规则": self.matched_rule,
            "销量趋势": self.sales_trend,
            "趋势分组": self.trend_group,
            "加权日销": _display_float(self.weighted_daily_sales),
            "可销售天数": _display_optional_float(self.sales_days),
            "FBA总库存": _display_quantity(self.fba_total_inventory),
            "未关联数量": _display_quantity(self.unlinked_quantity),
            "真实库存数量": _display_optional_quantity(self.actual_inventory),
            "单品重量(g)": _display_optional_float(self.weight_grams),
            "补货天数": _display_optional_int(self.replenish_days),
            "补货量": _display_optional_int(self.original_replenish_quantity),
            "补货量（减去 FBA 总库存）": _display_optional_int(fba_deducted_quantity),
            "补货量（减去 FBA 总库存和未关联货件）": _display_optional_int(self.replenish_quantity),
            AMAZON_FBA_TOTAL_COLUMN: _display_optional_quantity(self.amazon_fba_total_inventory),
            AMAZON_DEDUCTED_REPLENISH_COLUMN: _display_optional_int(self.amazon_deducted_replenish_quantity),
            "海运天数": _display_optional_int(self.sea_days),
            "海运建议量": _display_optional_int(self.sea_quantity),
            "同时空运天数": _display_optional_int(self.companion_air_days),
            "同时空运建议量": _display_optional_int(self.companion_air_quantity),
            "预计总重量kg": _display_optional_float(display_weight_kg),
            "决策原因": self.decision_reason,
        }


@dataclass(frozen=True)
class StoreMskuReplenishmentResult:
    store_name: str
    source_data_time: str
    sales_analysis_xlsx_path: str
    actual_inventory_xlsx_path: str
    template_name: str
    template_version: int
    row_count: int
    link_count: int
    air_urgent_count: int
    air_count: int
    sea_count: int
    clearance_count: int
    no_ship_count: int
    sample_insufficient_count: int
    report_xlsx_path: str
    unlinked_shipments_snapshot_path: str = ""
    unlinked_shipments_snapshot_warning: str = ""
    amazon_inventory_snapshot_path: str = ""
    amazon_inventory_validation: dict[str, Any] | None = None
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        payload = {
            "success": True,
            "store_name": self.store_name,
            "source_data_time": self.source_data_time,
            "sales_analysis_xlsx_path": self.sales_analysis_xlsx_path,
            "actual_inventory_xlsx_path": self.actual_inventory_xlsx_path,
            "template_name": self.template_name,
            "template_version": self.template_version,
            "row_count": self.row_count,
            "link_count": self.link_count,
            "air_urgent_count": self.air_urgent_count,
            "air_count": self.air_count,
            "sea_count": self.sea_count,
            "clearance_count": self.clearance_count,
            "no_ship_count": self.no_ship_count,
            "sample_insufficient_count": self.sample_insufficient_count,
            "report_xlsx_path": self.report_xlsx_path,
            "source": self.source,
        }
        if self.unlinked_shipments_snapshot_path:
            payload["unlinked_shipments_snapshot_path"] = self.unlinked_shipments_snapshot_path
        if self.unlinked_shipments_snapshot_warning:
            payload["unlinked_shipments_snapshot_warning"] = self.unlinked_shipments_snapshot_warning
        if self.amazon_inventory_snapshot_path:
            payload["amazon_inventory_snapshot_path"] = self.amazon_inventory_snapshot_path
        if self.amazon_inventory_validation:
            payload["amazon_inventory_validation"] = self.amazon_inventory_validation
        return payload


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _safe_file_part(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("._-") or "store_msku"


def _safe_unlinked_snapshot_store_part(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip(" ._-") or "store"


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _configured_path(name: str, default: Path) -> Path:
    return mabang_settings.configured_path(name, default)


def _sales_analysis_dir(input_dir: str | Path | None = None) -> Path:
    return Path(input_dir) if input_dir is not None else _configured_path(
        "MABANG_STORE_MSKU_ANALYSIS_OUTPUT_DIR",
        DEFAULT_SALES_ANALYSIS_DIR,
    )


def _actual_inventory_dir(input_dir: str | Path | None = None) -> Path:
    return Path(input_dir) if input_dir is not None else _configured_path(
        "MABANG_STORE_MSKU_INVENTORY_OUTPUT_DIR",
        DEFAULT_ACTUAL_INVENTORY_DIR,
    )


def _output_dir(output_dir: str | Path | None = None) -> Path:
    path = Path(output_dir) if output_dir is not None else _configured_path(
        "MABANG_STORE_MSKU_REPLENISHMENT_OUTPUT_DIR",
        DEFAULT_OUTPUT_DIR,
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _unlinked_shipments_snapshot_dir(input_dir: str | Path | None = None) -> Path:
    return Path(input_dir) if input_dir is not None else _configured_path(
        "MABANG_FBA_UNLINKED_SHIPMENTS_SNAPSHOT_DIR",
        DEFAULT_UNLINKED_SHIPMENTS_SNAPSHOT_DIR,
    )


def _number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        number = float(value)
        return 0.0 if math.isnan(number) else number
    text = _clean_text(value).replace(",", "")
    if not text or text.lower() == "nan":
        return 0.0
    try:
        number = float(text)
    except (TypeError, ValueError):
        return 0.0
    return 0.0 if math.isnan(number) else number


def _optional_number(value: Any) -> float | None:
    text = _clean_text(value)
    if not text:
        return None
    return _number(value)


def _display_float(value: float) -> float:
    return round(float(value), 2)


def _display_quantity(value: float) -> int | float:
    number = float(value or 0)
    if math.isnan(number):
        return 0
    rounded = round(number, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _display_optional_quantity(value: float | None) -> int | float | str:
    if value is None:
        return ""
    return _display_quantity(value)


def _display_optional_float(value: float | None) -> float | str:
    if value is None:
        return ""
    return _display_float(value)


def _display_optional_int(value: int | None) -> int | str:
    return "" if value is None else int(value)


def _row_key(row: dict[str, Any] | InventoryInputRow) -> tuple[str, str, str, str]:
    if isinstance(row, InventoryInputRow):
        return (row.msku, row.parent_asin, row.asin, row.local_sku)
    return (
        _clean_text(row.get("MSKU")),
        _clean_text(row.get("父ASIN")),
        _clean_text(row.get("ASIN")),
        _clean_text(row.get("本地SKU")),
    )


def _extract_product_url(value: Any) -> str:
    match = URL_RE.search(_clean_text(value))
    if not match:
        return ""
    return match.group(0).rstrip("),，;；。")


def _product_url_prefix(value: Any, asin: Any) -> str:
    url = _extract_product_url(value)
    if not url:
        return ""

    clean_asin = _clean_text(asin)
    if clean_asin and clean_asin in url:
        prefix = url.split(clean_asin, 1)[0]
        if prefix:
            return prefix

    slash_index = url.rfind("/")
    if slash_index < 0:
        return ""
    return url[: slash_index + 1]


def _parse_report_file(path: Path, pattern: re.Pattern[str]) -> ReportFile | None:
    match = pattern.match(path.name)
    if not match:
        return None
    source_data_time = match.group("source_time")
    try:
        source_datetime = datetime.strptime(source_data_time, "%Y%m%d%H%M")
    except ValueError:
        return None
    return ReportFile(path=path, source_data_time=source_data_time, source_datetime=source_datetime)


def _parse_unlinked_snapshot_file(path: Path) -> UnlinkedSnapshotFile | None:
    match = UNLINKED_SNAPSHOT_RE.match(path.name)
    if not match:
        return None
    snapshot_time = match.group("snapshot_time")
    try:
        snapshot_datetime = datetime.strptime(snapshot_time, "%Y%m%d%H%M")
    except ValueError:
        return None
    return UnlinkedSnapshotFile(
        path=path,
        snapshot_time=snapshot_time,
        snapshot_datetime=snapshot_datetime,
        store_part=match.group("store"),
    )


def _find_report_files(directory: Path, pattern: re.Pattern[str], safe_store_name: str) -> list[ReportFile]:
    if not directory.is_dir():
        return []
    reports: list[ReportFile] = []
    for path in directory.glob(f"*-{safe_store_name}_*.xlsx"):
        parsed = _parse_report_file(path, pattern)
        if parsed is not None:
            reports.append(parsed)
    return reports


def find_same_day_unlinked_shipments_snapshot(
    store_name: str,
    source_data_time: str,
    *,
    snapshot_dir: str | Path | None = None,
) -> tuple[Path | None, str]:
    source_date = _clean_text(source_data_time)[:8]
    clean_store_name = normalize_store_name(store_name)
    expected_store_part = _safe_unlinked_snapshot_store_part(clean_store_name)
    directory = _unlinked_shipments_snapshot_dir(snapshot_dir)
    if not directory.is_dir():
        return None, UNLINKED_SNAPSHOT_MISSING_WARNING

    candidates: list[UnlinkedSnapshotFile] = []
    ignored_dates: list[UnlinkedSnapshotFile] = []
    for path in directory.glob("*_unlinked_shipments_snapshot.xlsx"):
        parsed = _parse_unlinked_snapshot_file(path)
        if parsed is None or parsed.store_part != expected_store_part:
            continue
        if parsed.snapshot_time[:8] == source_date:
            candidates.append(parsed)
        else:
            ignored_dates.append(parsed)

    if candidates:
        selected = sorted(candidates, key=lambda item: item.snapshot_time, reverse=True)[0]
        return selected.path, ""
    if ignored_dates:
        return None, UNLINKED_SNAPSHOT_IGNORED_NON_SAME_DAY_WARNING
    return None, UNLINKED_SNAPSHOT_MISSING_WARNING


def validate_unlinked_shipments_snapshot_same_day(path: str | Path, source_data_time: str) -> Path:
    source_date = _clean_text(source_data_time)[:8]
    snapshot_path = Path(path)
    parsed = _parse_unlinked_snapshot_file(snapshot_path)
    if parsed is None:
        raise StoreMskuReplenishmentError(f"未关联货件快照文件名无法识别日期: {snapshot_path}")
    snapshot_date = parsed.snapshot_time[:8]
    if snapshot_date != source_date:
        raise StoreMskuReplenishmentError(
            f"未关联货件快照日期与备货数据日期不一致: source_date={source_date}, "
            f"snapshot_date={snapshot_date}, path={snapshot_path}"
        )
    return snapshot_path


def find_matching_report_files(
    store_name: str,
    *,
    sales_analysis_dir: str | Path | None = None,
    actual_inventory_dir: str | Path | None = None,
) -> MatchedReports:
    clean_store_name = normalize_store_name(store_name)
    safe_store_name = _safe_file_part(clean_store_name)
    sales_reports = _find_report_files(_sales_analysis_dir(sales_analysis_dir), SALES_REPORT_RE, safe_store_name)
    inventory_reports = _find_report_files(_actual_inventory_dir(actual_inventory_dir), INVENTORY_REPORT_RE, safe_store_name)
    sales_by_time = {item.source_data_time: item for item in sales_reports}
    inventory_by_time = {item.source_data_time: item for item in inventory_reports}
    common_times = sorted(set(sales_by_time) & set(inventory_by_time), reverse=True)
    if not common_times:
        sales_times = ", ".join(sorted(sales_by_time, reverse=True)[:5]) or "无"
        inventory_times = ", ".join(sorted(inventory_by_time, reverse=True)[:5]) or "无"
        raise StoreMskuReplenishmentError(
            f"未找到同源时间的销量分析和真实库存报表: store={clean_store_name}, "
            f"sales_times={sales_times}, inventory_times={inventory_times}"
        )
    source_data_time = common_times[0]
    return MatchedReports(
        source_data_time=source_data_time,
        sales_analysis_path=sales_by_time[source_data_time].path,
        actual_inventory_path=inventory_by_time[source_data_time].path,
    )


def _headers_and_rows(xlsx_path: str | Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取xlsx") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"xlsx不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise StoreMskuReplenishmentError(f"xlsx缺少sheet: {sheet_name}, path={source_path}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        if not any(headers):
            raise StoreMskuReplenishmentError(f"xlsx表头为空: sheet={sheet_name}, path={source_path}")
        records: list[dict[str, Any]] = []
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            if any(_clean_text(value) for value in row.values()):
                records.append(row)
    except StoreMskuReplenishmentError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取xlsx失败: {source_path}, sheet={sheet_name}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass
    return headers, records


def _require_columns(headers: list[str], required_columns: tuple[str, ...], *, path: Path, sheet_name: str) -> None:
    missing = [column for column in required_columns if column not in headers]
    if missing:
        raise StoreMskuReplenishmentError(
            f"xlsx缺少列: {', '.join(missing)}, sheet={sheet_name}, path={path}"
        )


def load_inventory_rows(xlsx_path: str | Path) -> list[InventoryInputRow]:
    source_path = Path(xlsx_path)
    rows: list[InventoryInputRow] = []
    for sheet_name in INVENTORY_SHEETS:
        headers, records = _headers_and_rows(source_path, sheet_name)
        _require_columns(headers, INVENTORY_REQUIRED_COLUMNS, path=source_path, sheet_name=sheet_name)
        sku_type = "组合sku" if sheet_name == "真实库存-组合sku" else "库存sku"
        for record in records:
            rows.append(
                InventoryInputRow(
                    msku=_clean_text(record.get("MSKU")),
                    parent_asin=_clean_text(record.get("父ASIN")) or "未填写父ASIN",
                    asin=_clean_text(record.get("ASIN")),
                    local_sku=_clean_text(record.get("本地SKU")),
                    local_sku_name=_clean_text(record.get("本地SKU名称")),
                    product_name=_clean_text(record.get("产品名称")),
                    remark=_clean_text(record.get("备注")),
                    product_link=_clean_text(record.get("商品链接")),
                    sku_type=sku_type,
                    weighted_daily_sales=_number(record.get("加权日销")),
                    sales_days=_optional_number(record.get("可销售天数")),
                    fba_total_inventory=_number(record.get("FBA总库存")),
                    actual_inventory=_optional_number(record.get("真实库存数量")),
                    child_skus=_clean_text(record.get("子SKU")),
                )
            )
    return rows


def parse_weight_grams(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if not math.isnan(number) and number > 0:
            return number
    match = WEIGHT_RE.search(_clean_text(value))
    if not match:
        return None
    try:
        number = float(match.group(0).replace(",", ""))
    except ValueError:
        return None
    return number if number > 0 else None


def load_sales_details(xlsx_path: str | Path) -> dict[tuple[str, str, str, str], SalesDetail]:
    source_path = Path(xlsx_path)
    headers, records = _headers_and_rows(source_path, DETAIL_SHEET)
    _require_columns(headers, SALES_REQUIRED_COLUMNS, path=source_path, sheet_name=DETAIL_SHEET)
    details: dict[tuple[str, str, str, str], SalesDetail] = {}
    for record in records:
        key = _row_key(record)
        if key in details:
            raise StoreMskuReplenishmentError(
                f"销量分析MSKU明细存在重复行: MSKU={key[0]}, 父ASIN={key[1]}, ASIN={key[2]}, 本地SKU={key[3]}"
            )
        details[key] = SalesDetail(
            trend=_clean_text(record.get("销量趋势")),
            weight_grams=parse_weight_grams(record.get("单品重量(g)(cm)")),
            sales_7d=_number(record.get("7天销量")),
            sales_14d=_number(record.get("14天销量")),
            sales_30d=_number(record.get("30天销量")),
        )
    return details


def trend_group(trend: str) -> str | None:
    return trend_group_from_template(trend, get_template(DEFAULT_TEMPLATE_NAME))


def replenishment_days(weighted_daily_sales: float, mapped_trend: str) -> int:
    return replenishment_days_from_template(
        weighted_daily_sales,
        mapped_trend,
        get_template(DEFAULT_TEMPLATE_NAME).params,
    )


def _sea_day_candidates(weighted_daily_sales: float) -> list[int]:
    return sea_day_candidates_from_template(weighted_daily_sales, get_template(DEFAULT_TEMPLATE_NAME).params)


def _safe_unlinked_quantity(value: float | int | None) -> float:
    try:
        quantity = float(value or 0)
    except (TypeError, ValueError):
        return 0.0
    if math.isnan(quantity) or quantity < 0:
        return 0.0
    return quantity


def _remaining_quantity(quantity: int | float, deduction: int | float) -> int:
    remaining = float(quantity or 0) - float(deduction or 0)
    return 0 if remaining <= 0 else math.ceil(remaining)


def _deduct_companion_then_sea(
    *,
    companion_air_quantity: int,
    sea_quantity: int,
    deduction: float,
) -> tuple[int, int]:
    remaining_deduction = max(0.0, float(deduction or 0))
    companion_remaining = companion_air_quantity
    sea_remaining = sea_quantity
    if remaining_deduction <= 0:
        return companion_remaining, sea_remaining

    companion_after = companion_remaining - remaining_deduction
    if companion_after > 0:
        return math.ceil(companion_after), sea_remaining

    remaining_deduction = abs(companion_after)
    companion_remaining = 0
    sea_after = sea_remaining - remaining_deduction
    sea_remaining = 0 if sea_after <= 0 else math.ceil(sea_after)
    return companion_remaining, sea_remaining


def _report_estimated_weight_kg(row: ReplenishmentRow) -> float | None:
    if row.sheet_name not in {AIR_URGENT_SHEET, AIR_SHEET, SEA_SHEET, CLEARANCE_SHEET}:
        return None
    if row.replenish_quantity is None or row.weight_grams is None:
        return None
    return row.replenish_quantity * row.weight_grams / 1000


def _with_clearance_split(row: ReplenishmentRow) -> ReplenishmentRow:
    if row.sheet_name not in {AIR_URGENT_SHEET, AIR_SHEET, SEA_SHEET}:
        return row
    if row.replenish_quantity is None or row.replenish_quantity <= 0:
        return row
    if CLEARANCE_KEYWORD not in row.remark:
        return row
    return replace(
        row,
        sheet_name=CLEARANCE_SHEET,
        transport_channel=row.sheet_name,
        decision_reason=f"{row.decision_reason}；备注包含{CLEARANCE_KEYWORD}，移入清货sheet",
    )


def _with_inventory_deductions(
    row: ReplenishmentRow,
    *,
    min_weight_kg: float | None = None,
    min_sea_quantity: float | None = None,
) -> ReplenishmentRow:
    if row.sheet_name not in {AIR_URGENT_SHEET, AIR_SHEET, SEA_SHEET}:
        return row
    if row.replenish_quantity is None:
        return row

    original_quantity = int(row.original_replenish_quantity or row.replenish_quantity)
    if row.sheet_name == SEA_SHEET:
        original_sea_quantity = int(row.sea_quantity or original_quantity)
        original_companion_air_quantity = int(row.companion_air_quantity or 0)
        fba_companion_air_quantity, fba_sea_quantity = _deduct_companion_then_sea(
            companion_air_quantity=original_companion_air_quantity,
            sea_quantity=original_sea_quantity,
            deduction=row.fba_total_inventory,
        )
        final_companion_air_quantity, final_sea_quantity = _deduct_companion_then_sea(
            companion_air_quantity=fba_companion_air_quantity,
            sea_quantity=fba_sea_quantity,
            deduction=row.unlinked_quantity,
        )
        fba_deducted_quantity = fba_companion_air_quantity + fba_sea_quantity
        final_quantity = final_companion_air_quantity + final_sea_quantity
    else:
        fba_deducted_quantity = _remaining_quantity(original_quantity, row.fba_total_inventory)
        final_quantity = _remaining_quantity(fba_deducted_quantity, row.unlinked_quantity)
        fba_sea_quantity = 0
        final_sea_quantity = 0
        final_companion_air_quantity = 0

    reason_suffix = f"；补货量={original_quantity}，FBA总库存={_display_quantity(row.fba_total_inventory)}，补货量（减去 FBA 总库存）={fba_deducted_quantity}"
    if row.unlinked_quantity > 0:
        reason_suffix += (
            f"，未关联数量={_display_quantity(row.unlinked_quantity)}，"
            f"补货量（减去 FBA 总库存和未关联货件）={final_quantity}"
        )
    else:
        reason_suffix += f"，补货量（减去 FBA 总库存和未关联货件）={final_quantity}"

    if fba_deducted_quantity <= 0:
        return replace(
            row,
            replenish_quantity=0,
            fba_deducted_replenish_quantity=0,
            sea_quantity=0 if row.sheet_name == SEA_SHEET else row.sea_quantity,
            companion_air_quantity=0 if row.companion_air_quantity is not None else row.companion_air_quantity,
            sea_net_quantity=0 if row.sea_net_quantity is not None else row.sea_net_quantity,
            estimated_weight_kg=None,
            decision_reason=f"{row.decision_reason}{reason_suffix}，FBA总库存已覆盖本次建议量",
            sheet_name=NO_SHIP_SHEET,
        )
    if final_quantity <= 0:
        return replace(
            row,
            replenish_quantity=0,
            fba_deducted_replenish_quantity=fba_deducted_quantity,
            sea_quantity=0 if row.sheet_name == SEA_SHEET else row.sea_quantity,
            companion_air_quantity=0 if row.companion_air_quantity is not None else row.companion_air_quantity,
            sea_net_quantity=0 if row.sea_net_quantity is not None else row.sea_net_quantity,
            estimated_weight_kg=None,
            decision_reason=f"{row.decision_reason}{reason_suffix}，FBA总库存和未关联数量已覆盖本次建议量",
            sheet_name=NO_SHIP_SHEET,
        )

    estimated_weight_kg = row.estimated_weight_kg
    if row.sheet_name == SEA_SHEET and row.weight_grams is not None:
        estimated_weight_kg = final_sea_quantity * row.weight_grams / 1000
        if min_sea_quantity is not None and final_sea_quantity < min_sea_quantity:
            return replace(
                row,
                replenish_quantity=final_quantity,
                fba_deducted_replenish_quantity=fba_deducted_quantity,
                sea_quantity=final_sea_quantity,
                companion_air_quantity=(
                    final_companion_air_quantity
                    if row.companion_air_quantity is not None
                    else row.companion_air_quantity
                ),
                sea_net_quantity=final_sea_quantity if row.sea_net_quantity is not None else row.sea_net_quantity,
                estimated_weight_kg=None,
                decision_reason=(
                    f"{row.decision_reason}{reason_suffix}，"
                    f"扣减FBA总库存和未关联货件后，海运数量不足{min_sea_quantity:g}件"
                ),
                sheet_name=NO_SHIP_SHEET,
            )
        if min_weight_kg is not None and estimated_weight_kg <= min_weight_kg:
            return replace(
                row,
                replenish_quantity=final_quantity,
                fba_deducted_replenish_quantity=fba_deducted_quantity,
                sea_quantity=final_sea_quantity,
                companion_air_quantity=(
                    final_companion_air_quantity
                    if row.companion_air_quantity is not None
                    else row.companion_air_quantity
                ),
                sea_net_quantity=final_sea_quantity if row.sea_net_quantity is not None else row.sea_net_quantity,
                estimated_weight_kg=None,
                decision_reason=(
                    f"{row.decision_reason}{reason_suffix}，"
                    f"扣减FBA总库存和未关联货件后，海运重量不足{min_weight_kg:g}kg"
                ),
                sheet_name=NO_SHIP_SHEET,
            )

    return _with_clearance_split(
        replace(
            row,
            replenish_quantity=final_quantity,
            fba_deducted_replenish_quantity=fba_deducted_quantity,
            sea_quantity=final_sea_quantity if row.sheet_name == SEA_SHEET else row.sea_quantity,
            companion_air_quantity=(
                final_companion_air_quantity if row.companion_air_quantity is not None else row.companion_air_quantity
            ),
            sea_net_quantity=final_sea_quantity if row.sea_net_quantity is not None else row.sea_net_quantity,
            estimated_weight_kg=estimated_weight_kg,
            decision_reason=f"{row.decision_reason}{reason_suffix}",
        )
    )


def _with_amazon_inventory_snapshot(row: ReplenishmentRow, amazon_fba_inventory: float | None) -> ReplenishmentRow:
    if amazon_fba_inventory is None:
        return replace(row, decision_reason=f"{row.decision_reason}；未匹配 Amazon 后台库存")
    amazon_deducted_quantity = None
    if row.original_replenish_quantity is not None:
        amazon_deducted_quantity = _remaining_quantity(
            _remaining_quantity(row.original_replenish_quantity, amazon_fba_inventory),
            row.unlinked_quantity,
        )
    return replace(
        row,
        amazon_fba_total_inventory=amazon_fba_inventory,
        amazon_deducted_replenish_quantity=amazon_deducted_quantity,
    )


def calculate_replenishment_row(
    row: InventoryInputRow,
    sales_detail: SalesDetail,
    template: ReplenishmentTemplate | None = None,
    unlinked_quantity: float = 0.0,
) -> ReplenishmentRow:
    active_template = template or get_template(DEFAULT_TEMPLATE_NAME)
    unlinked_qty = _safe_unlinked_quantity(unlinked_quantity)
    params, matched_rule = effective_params_for_msku(active_template, row.msku)
    weighted_daily_sales = calculate_weighted_daily_sales(
        sales_7d=sales_detail.sales_7d,
        sales_14d=sales_detail.sales_14d,
        sales_30d=sales_detail.sales_30d,
        params=params,
    )
    sales_days = None if weighted_daily_sales <= 0 else row.fba_total_inventory / weighted_daily_sales
    rule_name = matched_rule or "默认规则"
    mapped_trend = trend_group_from_template(sales_detail.trend, active_template)
    if mapped_trend is None:
        return ReplenishmentRow(
            msku=row.msku,
            parent_asin=row.parent_asin,
            asin=row.asin,
            local_sku=row.local_sku,
            local_sku_name=row.local_sku_name,
            product_name=row.product_name,
            remark=row.remark,
            product_link=row.product_link,
            sku_type=row.sku_type,
            template_name=active_template.name,
            matched_rule=rule_name,
            sales_trend=sales_detail.trend,
            trend_group="样本不足",
            weighted_daily_sales=weighted_daily_sales,
            sales_days=sales_days,
            fba_total_inventory=row.fba_total_inventory,
            unlinked_quantity=unlinked_qty,
            actual_inventory=row.actual_inventory,
            weight_grams=sales_detail.weight_grams,
            replenish_days=None,
            replenish_quantity=None,
            original_replenish_quantity=None,
            sea_days=None,
            sea_quantity=None,
            estimated_weight_kg=None,
            decision_reason="销量趋势为样本不足，不计算备货量",
            child_skus=row.child_skus,
            sheet_name=SAMPLE_INSUFFICIENT_SHEET,
        )

    replenish_day_count = replenishment_days_from_template(weighted_daily_sales, mapped_trend, params)
    replenish_quantity = math.ceil(weighted_daily_sales * replenish_day_count)
    base_kwargs = {
        "msku": row.msku,
        "parent_asin": row.parent_asin,
        "asin": row.asin,
        "local_sku": row.local_sku,
        "local_sku_name": row.local_sku_name,
        "product_name": row.product_name,
        "remark": row.remark,
        "product_link": row.product_link,
        "sku_type": row.sku_type,
        "template_name": active_template.name,
        "matched_rule": rule_name,
        "sales_trend": sales_detail.trend,
        "trend_group": mapped_trend,
        "weighted_daily_sales": weighted_daily_sales,
        "sales_days": sales_days,
        "fba_total_inventory": row.fba_total_inventory,
        "unlinked_quantity": unlinked_qty,
        "actual_inventory": row.actual_inventory,
        "weight_grams": sales_detail.weight_grams,
        "replenish_days": replenish_day_count,
        "replenish_quantity": replenish_quantity,
        "original_replenish_quantity": replenish_quantity,
        "child_skus": row.child_skus,
    }

    if sales_days is None:
        return ReplenishmentRow(
            **base_kwargs,
            sea_days=None,
            sea_quantity=None,
            estimated_weight_kg=None,
            decision_reason="可销售天数为空，暂不建议发货",
            sheet_name=NO_SHIP_SHEET,
        )
    air_urgent_days = float(params["shipping"]["air_urgent_sales_days_lte"])
    air_days = float(params["shipping"]["air_sales_days_lte"])
    if sales_days <= air_urgent_days + FLOAT_TOLERANCE:
        return _with_inventory_deductions(
            ReplenishmentRow(
                **base_kwargs,
                sea_days=None,
                sea_quantity=None,
                estimated_weight_kg=None,
                decision_reason=f"可销售天数={sales_days:.2f} <= {air_urgent_days:g}，建议空运（急发）",
                sheet_name=AIR_URGENT_SHEET,
            )
        )
    if sales_days <= air_days + FLOAT_TOLERANCE:
        return _with_inventory_deductions(
            ReplenishmentRow(
                **base_kwargs,
                sea_days=None,
                sea_quantity=None,
                estimated_weight_kg=None,
                decision_reason=f"{air_urgent_days:g} < 可销售天数={sales_days:.2f} <= {air_days:g}，建议空运",
                sheet_name=AIR_SHEET,
            )
        )

    if not sea_enabled_from_template(params):
        return ReplenishmentRow(
            **base_kwargs,
            sea_days=None,
            sea_quantity=None,
            estimated_weight_kg=None,
            decision_reason=f"可销售天数={sales_days:.2f} > {air_days:g}，但模板已关闭海运，暂不建议发货",
            sheet_name=NO_SHIP_SHEET,
        )

    min_daily_sales = float(params["sea"]["min_daily_sales"])
    if not sea_daily_sales_meets_min_from_template(weighted_daily_sales, params):
        operator_text = "<" if sea_min_daily_sales_inclusive_from_template(params) else "<="
        return ReplenishmentRow(
            **base_kwargs,
            sea_days=None,
            sea_quantity=None,
            estimated_weight_kg=None,
            decision_reason=f"可销售天数={sales_days:.2f} > {air_days:g}，但加权日销={weighted_daily_sales:.2f} {operator_text} {min_daily_sales:g}，不建议海运",
            sheet_name=NO_SHIP_SHEET,
        )
    if sales_detail.weight_grams is None:
        return ReplenishmentRow(
            **base_kwargs,
            sea_days=None,
            sea_quantity=None,
            estimated_weight_kg=None,
            decision_reason=f"可销售天数 > {air_days:g} 且加权日销 > {min_daily_sales:g}，但缺少单品重量，暂不建议海运",
            sheet_name=NO_SHIP_SHEET,
        )

    tried: list[str] = []
    min_weight_kg = float(params["sea"]["min_weight_kg"])
    companion_air_enabled = sea_companion_air_enabled_from_template(params)
    min_net_quantity = sea_min_net_quantity_from_template(params)
    for sea_days in sea_day_candidates_from_template(weighted_daily_sales, params):
        sea_target_quantity = math.ceil(weighted_daily_sales * sea_days)
        sea_quantity = sea_target_quantity
        companion_air_days: int | None = None
        companion_air_quantity: int | None = None
        sea_net_quantity: int | None = None
        actual_sea_quantity = sea_quantity
        if companion_air_enabled:
            companion_air_candidates = sea_companion_air_day_candidates_from_template(weighted_daily_sales, params)
            if not companion_air_candidates:
                tried.append(f"{sea_days}天: 未匹配海运同时空运分档")
                continue
            companion_air_days = companion_air_candidates[0]
            companion_air_quantity = math.ceil(weighted_daily_sales * companion_air_days)
            sea_quantity = max(0, sea_target_quantity - companion_air_quantity)
            sea_net_quantity = sea_quantity
            actual_sea_quantity = sea_net_quantity
        estimated_weight_kg = actual_sea_quantity * sales_detail.weight_grams / 1000
        if companion_air_enabled:
            total_replenish_quantity = sea_quantity + int(companion_air_quantity or 0)
            tried.append(
                f"{sea_days}天: 海运目标ceil({weighted_daily_sales:.2f}*{sea_days})={sea_target_quantity}, "
                f"同时空运ceil({weighted_daily_sales:.2f}*{companion_air_days})={companion_air_quantity}, "
                f"海运建议量={sea_quantity}, 补货量={total_replenish_quantity}, 海运重量={estimated_weight_kg:.2f}kg"
            )
            sea_kwargs = {
                **base_kwargs,
                "replenish_days": sea_days,
                "replenish_quantity": total_replenish_quantity,
                "original_replenish_quantity": total_replenish_quantity,
            }
            return _with_inventory_deductions(
                ReplenishmentRow(
                    **sea_kwargs,
                    sea_days=sea_days,
                    sea_quantity=sea_quantity,
                    companion_air_days=companion_air_days,
                    companion_air_quantity=companion_air_quantity,
                    sea_net_quantity=sea_net_quantity,
                    estimated_weight_kg=estimated_weight_kg,
                    decision_reason=(
                        f"可销售天数 > {air_days:g}，满足海运条件；"
                        f"按海运备货天数{sea_days}天计算海运目标量，"
                        f"拆分为海运建议量和同时空运建议量；计算方法："
                        + "；".join(tried)
                    ),
                    sheet_name=SEA_SHEET,
                ),
                min_weight_kg=min_weight_kg,
                min_sea_quantity=min_net_quantity,
            )

        tried.append(f"{sea_days}天: ceil({weighted_daily_sales:.2f}*{sea_days})={sea_quantity}, 重量={estimated_weight_kg:.2f}kg")
        if estimated_weight_kg > min_weight_kg:
            sea_kwargs = {
                **base_kwargs,
                "replenish_days": sea_days,
                "replenish_quantity": sea_quantity,
                "original_replenish_quantity": sea_quantity,
            }
            return _with_inventory_deductions(
                ReplenishmentRow(
                    **sea_kwargs,
                    sea_days=sea_days,
                    sea_quantity=sea_quantity,
                    estimated_weight_kg=estimated_weight_kg,
                    decision_reason=f"可销售天数 > {air_days:g}，满足海运条件；按海运备货天数{sea_days}天计算补货量；计算方法：" + "；".join(tried),
                    sheet_name=SEA_SHEET,
                ),
                min_weight_kg=min_weight_kg,
            )

    return ReplenishmentRow(
        **base_kwargs,
        sea_days=None,
        sea_quantity=None,
        estimated_weight_kg=None,
        decision_reason=f"可销售天数 > {air_days:g}，但海运重量未超过{min_weight_kg:g}kg；计算方法：" + "；".join(tried),
        sheet_name=NO_SHIP_SHEET,
    )


def calculate_replenishment_rows(
    inventory_rows: list[InventoryInputRow],
    sales_details: dict[tuple[str, str, str, str], SalesDetail],
    template: ReplenishmentTemplate | None = None,
    unlinked_quantities: dict[str, float] | None = None,
    amazon_inventory_quantities: dict[str, float] | None = None,
) -> list[ReplenishmentRow]:
    result: list[ReplenishmentRow] = []
    active_template = template or get_template(DEFAULT_TEMPLATE_NAME)
    quantity_by_msku = dict(unlinked_quantities or {})
    amazon_quantity_by_msku = dict(amazon_inventory_quantities or {})
    use_amazon_inventory = amazon_inventory_quantities is not None
    for row in inventory_rows:
        key = _row_key(row)
        sales_detail = sales_details.get(key)
        if sales_detail is None:
            raise StoreMskuReplenishmentError(
                f"销量分析MSKU明细缺少匹配行: MSKU={key[0]}, 父ASIN={key[1]}, ASIN={key[2]}, 本地SKU={key[3]}"
            )
        replenishment_row = calculate_replenishment_row(
            row,
            sales_detail,
            active_template,
            quantity_by_msku.get(row.msku, 0.0),
        )
        if use_amazon_inventory:
            replenishment_row = _with_amazon_inventory_snapshot(
                replenishment_row,
                amazon_quantity_by_msku.get(row.msku),
            )
        result.append(replenishment_row)
    return result


def summarize_links(rows: list[ReplenishmentRow]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.sheet_name == CLEARANCE_SHEET:
            continue
        group = groups.setdefault(
            row.parent_asin,
            {
                "父ASIN": row.parent_asin,
                "商品链接": "",
                "MSKU数": 0,
                "最大加权日销": 0.0,
                "合计加权日销": 0.0,
                "最小可销售天数": None,
                "链接未关联数量汇总": 0.0,
                "链接真实本地库存汇总": 0.0,
                "总补货量": 0,
                "空运（急发）补货量": 0,
                "空运补货量": 0,
                "海运建议量": 0,
                "涉及运输方式": set(),
                "决策备注": [],
                "商品链接前缀": "",
                "_order": len(groups),
            },
        )
        group["MSKU数"] += 1
        link_prefix = _product_url_prefix(row.product_link, row.asin)
        if link_prefix and not group["商品链接前缀"]:
            group["商品链接前缀"] = link_prefix
            if row.parent_asin != "未填写父ASIN":
                group["商品链接"] = f"{link_prefix}{row.parent_asin}"
        if row.sheet_name == SAMPLE_INSUFFICIENT_SHEET:
            group["涉及运输方式"].add(SAMPLE_INSUFFICIENT_SHEET)
        else:
            group["总补货量"] += row.replenish_quantity or 0
            group["涉及运输方式"].add(row.sheet_name)
        if row.sheet_name == AIR_URGENT_SHEET:
            group["空运（急发）补货量"] += row.replenish_quantity or 0
        elif row.sheet_name == AIR_SHEET:
            group["空运补货量"] += row.replenish_quantity or 0
        elif row.sheet_name == SEA_SHEET:
            group["海运建议量"] += row.sea_quantity or 0
        group["最大加权日销"] = max(group["最大加权日销"], row.weighted_daily_sales)
        group["合计加权日销"] += row.weighted_daily_sales
        if row.sales_days is not None:
            current_min = group["最小可销售天数"]
            group["最小可销售天数"] = row.sales_days if current_min is None else min(current_min, row.sales_days)
        if row.actual_inventory is not None:
            group["链接真实本地库存汇总"] += row.actual_inventory
        group["链接未关联数量汇总"] += row.unlinked_quantity
        if row.sheet_name in {SEA_SHEET, NO_SHIP_SHEET, SAMPLE_INSUFFICIENT_SHEET}:
            group["决策备注"].append(f"{row.msku}: {row.decision_reason}")

    summaries: list[dict[str, Any]] = []
    for group in groups.values():
        transport_values = [name for name in TRANSPORT_ORDER if name in group["涉及运输方式"]]
        summaries.append(
            {
                "父ASIN": group["父ASIN"],
                "商品链接": group["商品链接"],
                "MSKU数": group["MSKU数"],
                "最大加权日销": _display_float(group["最大加权日销"]),
                "合计加权日销": _display_float(group["合计加权日销"]),
                "最小可销售天数": _display_optional_float(group["最小可销售天数"]),
                "链接未关联数量汇总": _display_quantity(group["链接未关联数量汇总"]),
                "链接真实本地库存汇总": _display_quantity(group["链接真实本地库存汇总"]),
                "总补货量": group["总补货量"],
                "空运（急发）补货量": group["空运（急发）补货量"],
                "空运补货量": group["空运补货量"],
                "海运建议量": group["海运建议量"],
                "涉及运输方式": "、".join(transport_values),
                "决策备注": "；".join(group["决策备注"]),
            }
        )
    return sorted(summaries, key=lambda item: (-int(item["总补货量"]), groups[item["父ASIN"]]["_order"]))


def _inventory_shortage_payload(row: ReplenishmentRow) -> dict[str, Any] | None:
    if row.sheet_name not in {AIR_URGENT_SHEET, AIR_SHEET, SEA_SHEET}:
        return None
    if row.actual_inventory is None or row.replenish_quantity is None:
        return None
    shortage_quantity = row.replenish_quantity - row.actual_inventory
    if shortage_quantity <= 0:
        return None
    payload = {
        **row.to_detail_payload(),
        "运输渠道": row.sheet_name,
        "库存缺口": _display_quantity(shortage_quantity),
    }
    return {column: payload.get(column, "") for column in INVENTORY_SHORTAGE_COLUMNS}


def inventory_shortage_rows(rows: list[ReplenishmentRow]) -> list[dict[str, Any]]:
    shortages = [_inventory_shortage_payload(row) for row in rows]
    return [row for row in shortages if row is not None]


def _write_table(worksheet: Any, headers: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_format.defaultRowHeight = EXCEL_ROW_HEIGHT
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = EXCEL_ROW_HEIGHT
    for index, header in enumerate(headers, start=1):
        if header in TWO_DECIMAL_COLUMNS:
            for cells in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
                cells[0].number_format = "0.00"
        elif header in INTEGER_COLUMNS:
            for cells in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
                value = cells[0].value
                if isinstance(value, float) and not value.is_integer():
                    cells[0].number_format = "0.00"
                else:
                    cells[0].number_format = "0"
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = EXCEL_COLUMN_WIDTH


def write_replenishment_report(rows: list[ReplenishmentRow], report_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入备货计算报告") from exc

    target_path = Path(report_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_rows = {
        AIR_URGENT_SHEET: [row for row in rows if row.sheet_name == AIR_URGENT_SHEET],
        AIR_SHEET: [row for row in rows if row.sheet_name == AIR_SHEET],
        SEA_SHEET: [row for row in rows if row.sheet_name == SEA_SHEET],
        CLEARANCE_SHEET: [row for row in rows if row.sheet_name == CLEARANCE_SHEET],
        NO_SHIP_SHEET: [row for row in rows if row.sheet_name == NO_SHIP_SHEET],
        SAMPLE_INSUFFICIENT_SHEET: [row for row in rows if row.sheet_name == SAMPLE_INSUFFICIENT_SHEET],
    }
    workbook = Workbook()
    try:
        specs: list[tuple[str, tuple[str, ...], list[dict[str, Any]]]] = [
            *[
                (
                    sheet_name,
                    AIR_DETAIL_COLUMNS if sheet_name in {AIR_URGENT_SHEET, AIR_SHEET} else DETAIL_COLUMNS,
                    [row.to_detail_payload() for row in sorted(sheet_rows[sheet_name], key=_detail_sort_key)],
                )
                for sheet_name in (
                    AIR_URGENT_SHEET,
                    AIR_SHEET,
                    SEA_SHEET,
                )
            ],
            (
                INVENTORY_SHORTAGE_SHEET,
                INVENTORY_SHORTAGE_COLUMNS,
                sorted(inventory_shortage_rows(rows), key=_shortage_sort_key),
            ),
            (
                CLEARANCE_SHEET,
                CLEARANCE_COLUMNS,
                [
                    row.to_detail_payload()
                    for row in sorted(sheet_rows[CLEARANCE_SHEET], key=_detail_sort_key)
                ],
            ),
            (
                NO_SHIP_SHEET,
                DETAIL_COLUMNS,
                [row.to_detail_payload() for row in sorted(sheet_rows[NO_SHIP_SHEET], key=_detail_sort_key)],
            ),
            (SUMMARY_SHEET, SUMMARY_COLUMNS, summarize_links(rows)),
            (
                SAMPLE_INSUFFICIENT_SHEET,
                DETAIL_COLUMNS,
                [
                    row.to_detail_payload()
                    for row in sorted(sheet_rows[SAMPLE_INSUFFICIENT_SHEET], key=_detail_sort_key)
                ],
            ),
        ]
        for index, (sheet_name, headers, payload_rows) in enumerate(specs):
            worksheet = workbook.active if index == 0 else workbook.create_sheet()
            worksheet.title = sheet_name
            _write_table(worksheet, headers, payload_rows)
        workbook.save(target_path)
    finally:
        workbook.close()
    return target_path


def _detail_sort_key(row: ReplenishmentRow) -> tuple[int, str, str]:
    quantity = row.replenish_quantity or row.sea_quantity or 0
    return (-quantity, row.parent_asin, row.msku)


def _shortage_sort_key(row: dict[str, Any]) -> tuple[float, str, str]:
    return (-float(row.get("库存缺口") or 0), _clean_text(row.get("父ASIN")), _clean_text(row.get("MSKU")))


def calculate_store_msku_replenishment(
    store_name: str,
    *,
    template_name: str | None = None,
    sales_analysis_dir: str | Path | None = None,
    actual_inventory_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    unlinked_shipments_snapshot_path: str | Path | None = None,
    unlinked_shipments_snapshot_dir: str | Path | None = None,
    amazon_inventory_snapshot_path: str | Path | None = None,
) -> StoreMskuReplenishmentResult:
    clean_store_name = normalize_store_name(store_name)
    template = validate_template(get_template(template_name or DEFAULT_TEMPLATE_NAME)).template
    reports = find_matching_report_files(
        clean_store_name,
        sales_analysis_dir=sales_analysis_dir,
        actual_inventory_dir=actual_inventory_dir,
    )
    inventory_rows = load_inventory_rows(reports.actual_inventory_path)
    sales_details = load_sales_details(reports.sales_analysis_path)
    requested_snapshot_path = _clean_text(unlinked_shipments_snapshot_path)
    unlinked_snapshot_warning = ""
    if requested_snapshot_path:
        selected_snapshot_path: Path | None = validate_unlinked_shipments_snapshot_same_day(
            requested_snapshot_path,
            reports.source_data_time,
        )
    else:
        selected_snapshot_path, unlinked_snapshot_warning = find_same_day_unlinked_shipments_snapshot(
            clean_store_name,
            reports.source_data_time,
            snapshot_dir=unlinked_shipments_snapshot_dir,
        )
    unlinked_quantities = (
        load_unlinked_shipment_quantities(selected_snapshot_path, store_name=clean_store_name)
        if selected_snapshot_path
        else {}
    )
    requested_amazon_snapshot_path = _clean_text(amazon_inventory_snapshot_path)
    amazon_snapshot_data = (
        load_amazon_inventory_snapshot(requested_amazon_snapshot_path, store_name=clean_store_name)
        if requested_amazon_snapshot_path
        else None
    )
    if amazon_snapshot_data is not None and amazon_snapshot_data.snapshot_date != reports.source_data_time[:8]:
        raise StoreMskuReplenishmentError(
            "Amazon 后台库存快照日期与备货数据日期不一致: "
            f"snapshot_date={amazon_snapshot_data.snapshot_date}, source_data_date={reports.source_data_time[:8]}"
        )
    replenishment_rows = calculate_replenishment_rows(
        inventory_rows,
        sales_details,
        template,
        unlinked_quantities,
        amazon_snapshot_data.quantities_by_msku if amazon_snapshot_data is not None else None,
    )
    report_path = _output_dir(output_dir) / f"{reports.source_data_time}-{_safe_file_part(clean_store_name)}_replenishment.xlsx"
    write_replenishment_report(replenishment_rows, report_path)
    summary_rows = summarize_links(replenishment_rows)

    return StoreMskuReplenishmentResult(
        store_name=clean_store_name,
        source_data_time=reports.source_data_time,
        sales_analysis_xlsx_path=str(reports.sales_analysis_path),
        actual_inventory_xlsx_path=str(reports.actual_inventory_path),
        template_name=template.name,
        template_version=template.version,
        row_count=len(replenishment_rows),
        link_count=len(summary_rows),
        air_urgent_count=sum(1 for row in replenishment_rows if row.sheet_name == AIR_URGENT_SHEET),
        air_count=sum(1 for row in replenishment_rows if row.sheet_name == AIR_SHEET),
        sea_count=sum(1 for row in replenishment_rows if row.sheet_name == SEA_SHEET),
        clearance_count=sum(1 for row in replenishment_rows if row.sheet_name == CLEARANCE_SHEET),
        no_ship_count=sum(1 for row in replenishment_rows if row.sheet_name == NO_SHIP_SHEET),
        sample_insufficient_count=sum(1 for row in replenishment_rows if row.sheet_name == SAMPLE_INSUFFICIENT_SHEET),
        report_xlsx_path=str(report_path),
        unlinked_shipments_snapshot_path=str(selected_snapshot_path or ""),
        unlinked_shipments_snapshot_warning=unlinked_snapshot_warning,
        amazon_inventory_snapshot_path=requested_amazon_snapshot_path,
        amazon_inventory_validation=(
            amazon_snapshot_data.validation.to_payload()
            if amazon_snapshot_data is not None and amazon_snapshot_data.validation is not None
            else None
        ),
    )


__all__ = [
    "AIR_DETAIL_COLUMNS",
    "AIR_SHEET",
    "AIR_URGENT_SHEET",
    "CLEARANCE_COLUMNS",
    "CLEARANCE_SHEET",
    "DETAIL_COLUMNS",
    "INVENTORY_SHORTAGE_COLUMNS",
    "INVENTORY_SHORTAGE_SHEET",
    "NO_SHIP_SHEET",
    "REPORT_SHEETS",
    "SAMPLE_INSUFFICIENT_SHEET",
    "SEA_SHEET",
    "SOURCE",
    "SUMMARY_COLUMNS",
    "SUMMARY_SHEET",
    "UNLINKED_SNAPSHOT_IGNORED_NON_SAME_DAY_WARNING",
    "UNLINKED_SNAPSHOT_MISSING_WARNING",
    "InventoryInputRow",
    "ReplenishmentRow",
    "SalesDetail",
    "StoreMskuReplenishmentError",
    "StoreMskuReplenishmentResult",
    "calculate_replenishment_row",
    "calculate_replenishment_rows",
    "calculate_store_msku_replenishment",
    "find_matching_report_files",
    "find_same_day_unlinked_shipments_snapshot",
    "inventory_shortage_rows",
    "load_inventory_rows",
    "load_sales_details",
    "normalize_store_name",
    "parse_weight_grams",
    "replenishment_days",
    "summarize_links",
    "trend_group",
    "validate_unlinked_shipments_snapshot_same_day",
    "write_replenishment_report",
]
