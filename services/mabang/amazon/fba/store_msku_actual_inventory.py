from __future__ import annotations

import asyncio
import json
import math
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_FLOOR, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from shared.infra.net import erp_http_session, external_http_session
from services.mabang import config as mabang_settings
from services.mabang.auth_constants import (
    MABANG_MEMCACHE_COOKIE_NAME as MEMCACHE_COOKIE_NAME,
    PRIVATE_AMZ_HOST,
    PRIVATE_HOST,
)

from ...auth import get_auth_context
from ...cookies import build_cookie_header, extract_named_cookies
from ...errors import MabangAuthError, MabangBusinessError, MabangRequestError

DEFAULT_STORE_MSKU_DIR = Path("artifacts") / "mabang_store_msku"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "mabang_store_msku_inventory"
DEFAULT_COMBO_EXPORT_URL = "https://private.mabangerp.com/index.php?mod=combosku.doExportFileNew"
DEFAULT_COMBO_LIST_URL = "https://private-amz.mabangerp.com/index.php?mod=combosku.getCombosSkuList"
DEFAULT_COMBO_EXPORT_TEMPLATE_URL = "https://private.mabangerp.com/index.php"
DEFAULT_WAREHOUSE_SEARCH_URL = "https://private-amz.mabangerp.com/index.php?mod=warehouse.searchwarehousestock"
DEFAULT_WAREHOUSE_EXPORT_URL = (
    "https://private-amz.mabangerp.com/index.php?mod=warehouse.doexportwarehousestock&flag=1&showRmbColumn=0"
)
DEFAULT_PRIVATE_ORIGIN = "https://private.mabangerp.com"
DEFAULT_PRIVATE_REFERER = "https://private.mabangerp.com/"
DEFAULT_PRIVATE_AMZ_ORIGIN = "https://private-amz.mabangerp.com"
DEFAULT_PRIVATE_AMZ_REFERER = "https://private-amz.mabangerp.com/"
SOURCE = "mabang_store_msku_actual_inventory"
EXCEL_ROW_HEIGHT = 15
EXCEL_COLUMN_WIDTH = 15
SENTINEL_COMBO_SKU = "HSP022"
WAREHOUSE_ID = "1014318"
SOURCE_FILE_RE = re.compile(r"^(?P<source_time>\d{12})-(?P<store>.+)_msku_data\.xlsx$", re.IGNORECASE)
WHITESPACE_PATTERN = re.compile(r"\s+")
AUTH_FAIL_STATUS = {401, 403}
SALES_COLUMNS = ("7天销量", "14天销量", "30天销量")
FBA_STOCK_COLUMNS = ("可售", "待入库", "预留", "在途", "待调仓", "调仓中")
SOURCE_COLUMNS = ("MSKU", "父ASIN", "ASIN", "本地SKU", "商品链接", *SALES_COLUMNS, *FBA_STOCK_COLUMNS)
COMBO_SKU_COLUMN = "组合sku编码"
COMBO_COMPONENT_COUNT_COLUMN = "关联sku个数"
STOCK_SKU_COLUMN = "库存SKU编号"
AVAILABLE_STOCK_COLUMN = "可用库存量"
BASE_OUTPUT_COLUMNS = ("MSKU", "父ASIN", "ASIN", "本地SKU", "商品链接", "真实库存数量", "子SKU")
INVENTORY_OUTPUT_COLUMNS = (
    "MSKU",
    "父ASIN",
    "ASIN",
    "本地SKU",
    "商品链接",
    "FBA总库存",
    "加权日销",
    "可销售天数",
    "真实库存数量",
    "子SKU",
)
OUTPUT_COLUMNS = BASE_OUTPUT_COLUMNS
TWO_DECIMAL_COLUMNS = {"加权日销", "可销售天数"}
ACTUAL_INVENTORY_HIGHLIGHT_COLOR = "FFF2CC"
INVENTORY_HIGHLIGHT_COLUMNS = {"MSKU", "真实库存数量"}
COMBO_EXPORT_FIELDS: tuple[tuple[str, str], ...] = (
    (COMBO_SKU_COLUMN, "uq101"),
    (COMBO_COMPONENT_COUNT_COLUMN, "uq136"),
    ("关联sku信息", "uq138"),
)


class StoreMskuActualInventoryError(MabangBusinessError):
    pass


class StoreMskuActualInventoryAuthError(StoreMskuActualInventoryError, MabangAuthError):
    pass


class StoreMskuActualInventoryTimeoutError(StoreMskuActualInventoryError):
    pass


@dataclass(frozen=True)
class SourceMskuFile:
    path: Path
    source_data_time: str
    source_datetime: datetime


@dataclass(frozen=True)
class StoreMskuRow:
    msku: str
    parent_asin: str
    asin: str
    local_sku: str
    product_link: str
    sales_7d: Decimal = Decimal("0")
    sales_14d: Decimal = Decimal("0")
    sales_30d: Decimal = Decimal("0")
    fba_sellable: Decimal = Decimal("0")
    fba_inbound: Decimal = Decimal("0")
    fba_reserved: Decimal = Decimal("0")
    fba_in_transit: Decimal = Decimal("0")
    fba_pending_transfer: Decimal = Decimal("0")
    fba_transferring: Decimal = Decimal("0")


@dataclass(frozen=True)
class ComboComponent:
    stock_sku: str
    quantity: Decimal


@dataclass(frozen=True)
class ComboSku:
    combo_sku: str
    components: tuple[ComboComponent, ...]


@dataclass(frozen=True)
class ActualInventoryRow:
    msku: str
    parent_asin: str
    asin: str
    local_sku: str
    product_link: str
    actual_inventory: Decimal | None
    child_skus: str
    is_combo_sku: bool = False
    fba_total_inventory: Decimal = Decimal("0")
    weighted_daily_sales: Decimal = Decimal("0")
    sales_days: Decimal | None = None


@dataclass(frozen=True)
class ActualInventoryRowGroups:
    combo_inventory_rows: list[ActualInventoryRow]
    stock_inventory_rows: list[ActualInventoryRow]
    no_local_sku_rows: list[ActualInventoryRow]
    no_inventory_rows: list[ActualInventoryRow]

    @property
    def inventory_rows(self) -> list[ActualInventoryRow]:
        return [*self.combo_inventory_rows, *self.stock_inventory_rows]


@dataclass(frozen=True)
class ActualInventoryResult:
    store_name: str
    source_xlsx_path: str
    source_data_time: str
    local_sku_count: int
    combo_sku_count: int
    stock_sku_count: int
    missing_stock_skus: list[str]
    xlsx_path: str
    inventory_row_count: int = 0
    no_local_sku_count: int = 0
    no_inventory_row_count: int = 0
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "source_xlsx_path": self.source_xlsx_path,
            "source_data_time": self.source_data_time,
            "local_sku_count": self.local_sku_count,
            "combo_sku_count": self.combo_sku_count,
            "stock_sku_count": self.stock_sku_count,
            "inventory_row_count": self.inventory_row_count,
            "no_local_sku_count": self.no_local_sku_count,
            "no_inventory_row_count": self.no_inventory_row_count,
            "missing_stock_sku_count": len(self.missing_stock_skus),
            "missing_stock_skus": list(self.missing_stock_skus),
            "xlsx_path": self.xlsx_path,
            "source": self.source,
        }


def _configured_text(name: str, default: str) -> str:
    return mabang_settings.configured_text(name, default)


def _clean_text(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() == "nan":
        return ""
    return text


def normalize_sku_key(value: Any) -> str:
    return WHITESPACE_PATTERN.sub("", _clean_text(value))


def _safe_file_part(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("._-") or "store_msku_inventory"


def _timestamp_text(value: datetime | None = None) -> str:
    return (value or datetime.now()).strftime("%Y%m%d%H%M")


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _unique_text(values: list[str] | tuple[str, ...] | Any) -> list[str]:
    unique: OrderedDict[str, str] = OrderedDict()
    for value in values or []:
        text = _clean_text(value)
        key = normalize_sku_key(text)
        if not key or key in unique:
            continue
        unique[key] = text
    return list(unique.values())


def _decimal_value(value: Any, *, default: Decimal = Decimal("0")) -> Decimal:
    text = _clean_text(value).replace(",", "")
    if not text:
        return default
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def _display_decimal(value: Decimal | None) -> float | int | str:
    if value is None:
        return ""
    normalized = value.normalize()
    if normalized == normalized.to_integral_value():
        return int(normalized)
    return float(normalized)


def _display_two_decimal(value: Decimal | None) -> float | str:
    if value is None:
        return ""
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _display_quantity(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral_value():
        return str(int(normalized))
    return format(normalized, "f")


def _resolve_store_msku_dir(input_dir: str | Path | None = None) -> Path:
    if input_dir is not None:
        return Path(input_dir)
    configured = str(mabang_settings.MABANG_STORE_MSKU_OUTPUT_DIR or "").strip()
    return Path(configured) if configured else DEFAULT_STORE_MSKU_DIR


def _resolve_output_dir(output_dir: str | Path | None = None) -> Path:
    if output_dir is not None:
        path = Path(output_dir)
    else:
        configured = str(mabang_settings.MABANG_STORE_MSKU_INVENTORY_OUTPUT_DIR or "").strip()
        path = Path(configured) if configured else DEFAULT_OUTPUT_DIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def find_latest_store_msku_file(store_name: str, *, input_dir: str | Path | None = None) -> SourceMskuFile:
    clean_store_name = normalize_store_name(store_name)
    directory = _resolve_store_msku_dir(input_dir)
    safe_store_name = _safe_file_part(clean_store_name)
    if not directory.is_dir():
        raise StoreMskuActualInventoryError(f"未找到本地店铺MSKU数据文件: {clean_store_name}")

    candidates: list[SourceMskuFile] = []
    for path in directory.glob(f"*-{safe_store_name}_msku_data.xlsx"):
        match = SOURCE_FILE_RE.match(path.name)
        if not match:
            continue
        source_data_time = match.group("source_time")
        try:
            source_datetime = datetime.strptime(source_data_time, "%Y%m%d%H%M")
        except ValueError:
            continue
        candidates.append(SourceMskuFile(path=path, source_data_time=source_data_time, source_datetime=source_datetime))
    if not candidates:
        raise StoreMskuActualInventoryError(f"未找到本地店铺MSKU数据文件: {clean_store_name}")
    return max(candidates, key=lambda item: (item.source_datetime, item.path.name))


def load_store_msku_rows(xlsx_path: str | Path) -> list[StoreMskuRow]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取店铺MSKU数据Excel") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        header_values = next(values, None)
        headers = [_clean_text(cell) for cell in list(header_values or [])]
        missing = [column for column in SOURCE_COLUMNS if column not in headers]
        if missing:
            raise StoreMskuActualInventoryError(f"店铺MSKU数据缺少列: {', '.join(missing)}")

        rows: list[StoreMskuRow] = []
        for row_values in values:
            row = dict(zip(headers, list(row_values or []), strict=False))
            if not any(_clean_text(value) for value in row.values()):
                continue
            rows.append(
                StoreMskuRow(
                    msku=_clean_text(row.get("MSKU")),
                    parent_asin=_clean_text(row.get("父ASIN")),
                    asin=_clean_text(row.get("ASIN")),
                    local_sku=_clean_text(row.get("本地SKU")),
                    product_link=_clean_text(row.get("商品链接")),
                    sales_7d=_decimal_value(row.get("7天销量")),
                    sales_14d=_decimal_value(row.get("14天销量")),
                    sales_30d=_decimal_value(row.get("30天销量")),
                    fba_sellable=_decimal_value(row.get("可售")),
                    fba_inbound=_decimal_value(row.get("待入库")),
                    fba_reserved=_decimal_value(row.get("预留")),
                    fba_in_transit=_decimal_value(row.get("在途")),
                    fba_pending_transfer=_decimal_value(row.get("待调仓")),
                    fba_transferring=_decimal_value(row.get("调仓中")),
                )
            )
    except StoreMskuActualInventoryError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取店铺MSKU数据Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    return rows


def _private_request_headers(cookie_header: str) -> dict[str, str]:
    return {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": _configured_text("MABANG_COMBO_SKU_EXPORT_ORIGIN", DEFAULT_PRIVATE_ORIGIN),
        "Referer": _configured_text("MABANG_COMBO_SKU_EXPORT_REFERER", DEFAULT_PRIVATE_REFERER),
        "Cookie": cookie_header,
    }


def _private_amz_post_headers(cookie_header: str) -> dict[str, str]:
    return {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": _configured_text("MABANG_WAREHOUSE_STOCK_ORIGIN", DEFAULT_PRIVATE_AMZ_ORIGIN),
        "Referer": _configured_text("MABANG_WAREHOUSE_STOCK_REFERER", DEFAULT_PRIVATE_AMZ_REFERER),
        "Cookie": cookie_header,
    }


def _private_html_post_headers(cookie_header: str) -> dict[str, str]:
    return {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Content-Type": "application/x-www-form-urlencoded",
        "Origin": "null",
        "Cookie": cookie_header,
    }


def _private_amz_get_headers(cookie_header: str) -> dict[str, str]:
    return {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "Cookie": cookie_header,
    }


async def _read_json_response(resp: Any, *, action: str) -> dict[str, Any]:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise StoreMskuActualInventoryAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code >= 400:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")

    data: Any = None
    if text:
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            data = None
    if data is None:
        try:
            data = await resp.json(content_type=None)
        except Exception:
            data = None
    if not isinstance(data, dict):
        raise StoreMskuActualInventoryError(f"{action}返回非JSON对象")
    if data.get("success") is False:
        message = _clean_text(data.get("msg") or data.get("message") or data.get("error") or "unknown")
        raise StoreMskuActualInventoryError(f"{action}业务异常: {message}")
    return data


async def _read_optional_json_response(resp: Any, *, action: str) -> dict[str, Any]:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise StoreMskuActualInventoryAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code >= 400:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")
    if not text:
        return {}
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return {}
    if isinstance(data, dict) and data.get("success") is False:
        message = _clean_text(data.get("msg") or data.get("message") or data.get("error") or "unknown")
        raise StoreMskuActualInventoryError(f"{action}业务异常: {message}")
    return data if isinstance(data, dict) else {}


async def _read_http_ok_response(resp: Any, *, action: str) -> None:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise StoreMskuActualInventoryAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code < 200 or status_code >= 300:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")


async def _resolve_private_auth() -> tuple[str, str]:
    context = await get_auth_context(scope="erp")
    cookie_header = build_cookie_header(
        context.cookies_by_domain,
        request_host=PRIVATE_HOST,
        extra_cookies={"exportv2": "1"},
    )
    if not cookie_header:
        raise StoreMskuActualInventoryAuthError("未获取到 private.mabangerp.com Cookie")

    values = extract_named_cookies(context.cookies_by_domain, (MEMCACHE_COOKIE_NAME,))
    memcache_key = _clean_text(values.get(MEMCACHE_COOKIE_NAME))
    if not memcache_key:
        raise StoreMskuActualInventoryAuthError(f"缺少关键 Cookie: {MEMCACHE_COOKIE_NAME}")
    return cookie_header, memcache_key


async def _resolve_private_amz_cookie() -> str:
    context = await get_auth_context(scope="private_amz")
    cookie_header = build_cookie_header(
        context.cookies_by_domain,
        request_host=PRIVATE_AMZ_HOST,
        extra_cookies={"mabang_lite_rowsPerPage": "100"},
    )
    if not cookie_header:
        raise StoreMskuActualInventoryAuthError("未获取到 private-amz.mabangerp.com Cookie")
    return cookie_header


def _combo_export_url() -> str:
    return _configured_text("MABANG_COMBO_SKU_EXPORT_URL", DEFAULT_COMBO_EXPORT_URL)


def _combo_list_url() -> str:
    return _configured_text("MABANG_COMBO_SKU_LIST_URL", DEFAULT_COMBO_LIST_URL)


def _combo_export_template_url() -> str:
    return _configured_text("MABANG_COMBO_SKU_EXPORT_TEMPLATE_URL", DEFAULT_COMBO_EXPORT_TEMPLATE_URL)


def _warehouse_search_url() -> str:
    return _configured_text("MABANG_WAREHOUSE_STOCK_SEARCH_URL", DEFAULT_WAREHOUSE_SEARCH_URL)


def _warehouse_export_url() -> str:
    return _configured_text("MABANG_WAREHOUSE_STOCK_EXPORT_URL", DEFAULT_WAREHOUSE_EXPORT_URL)


def combo_query_skus(local_skus: list[str] | tuple[str, ...]) -> list[str]:
    return _unique_text([*list(local_skus or []), SENTINEL_COMBO_SKU])


def _ids_lines(values: list[str]) -> str:
    return "\r\n".join(values) + "\r\n"


def _ids_text(values: list[str]) -> str:
    return "\r\n".join(values)


def _combo_list_prewarm_form_data(local_skus: list[str]) -> list[tuple[str, str]]:
    return [
        ("searchLike", "comboSku"),
        ("operate", "Like"),
        ("searchKeywords", ""),
        ("labelId", ""),
        ("timeStart", ""),
        ("timeEnd", ""),
        ("searchStatus", ""),
        ("isBatchSearch", "1"),
        ("selecttype", "comboSku"),
        ("stockData", _ids_text(combo_query_skus(local_skus))),
        ("page", ""),
        ("rowsPerPage", ""),
    ]


def _combo_export_template_prewarm_form_data(local_skus: list[str]) -> list[tuple[str, str]]:
    return [
        ("mod", "export.exportTemplate"),
        ("data", _ids_text(combo_query_skus(local_skus))),
        ("type", "1"),
        ("menu", "combosku"),
        ("exportUrl", _combo_export_url()),
        ("sessid", ""),
        ("showRmbColumn", "2"),
    ]


def _combo_step1_form_data(local_skus: list[str], *, memcache_key: str) -> list[tuple[str, str]]:
    query_skus = combo_query_skus(local_skus)
    form: list[tuple[str, str]] = [
        ("backUrl", ""),
        ("orderIds", _ids_lines(query_skus)),
    ]
    form.extend(("fieldlabel", uq) for _, uq in COMBO_EXPORT_FIELDS)
    for name, uq in COMBO_EXPORT_FIELDS:
        form.extend(
            [
                ("map-name[]", name),
                ("map-uq[]", uq),
                ("map-text[]", ""),
            ]
        )
    form.extend(
        [
            ("templateName", ""),
            ("templateId", "0"),
            ("datasOpen", "1"),
            ("memcacheKey", memcache_key),
            ("showRmbColumn", "2"),
            ("pageSave", "1"),
            ("operateType", "19"),
            ("params", ""),
            ("InterfaceUrl", ""),
            ("mainMenu", ""),
            ("hiddenPage", "1"),
            ("hiddenPageSize", ""),
            ("tableBase", ""),
            ("isMerage", "1"),
            ("version", "v2"),
            ("step", "1"),
        ]
    )
    return form


def _step2_form_data(*, sn: str, sub_no: int) -> list[tuple[str, str]]:
    return [
        ("tableBase", ""),
        ("isMerage", "1"),
        ("version", "v2"),
        ("sn", sn),
        ("sub_no", str(sub_no)),
        ("step", "2"),
        ("1", "1"),
    ]


def _step3_form_data(*, sn: str) -> list[tuple[str, str]]:
    return [
        ("tableBase", ""),
        ("isMerage", "1"),
        ("sn", sn),
        ("version", "v2"),
        ("step", "3"),
        ("1", "1"),
    ]


def _step4_form_data(*, sn: str, task_id: str) -> list[tuple[str, str]]:
    return [
        ("tableBase", ""),
        ("isMerage", "1"),
        ("sn", sn),
        ("version", "v2"),
        ("step", "4"),
        ("taskId", task_id),
        ("1", "1"),
    ]


async def _post_combo_export(
    form_data: list[tuple[str, str]],
    *,
    cookie_header: str,
    action: str,
) -> dict[str, Any]:
    async with erp_http_session.post(
        _combo_export_url(),
        data=form_data,
        headers=_private_request_headers(cookie_header),
    ) as resp:
        return await _read_json_response(resp, action=action)


async def prewarm_combo_sku_export(
    local_skus: list[str],
    *,
    private_cookie_header: str,
    private_amz_cookie_header: str,
    delay_sec: float = 1.0,
) -> None:
    async with erp_http_session.post(
        _combo_list_url(),
        data=_combo_list_prewarm_form_data(local_skus),
        headers=_private_amz_post_headers(private_amz_cookie_header),
    ) as resp:
        await _read_http_ok_response(resp, action="组合SKU预热 1")

    await asyncio.sleep(max(0.0, float(delay_sec)))

    async with erp_http_session.post(
        _combo_export_template_url(),
        data=_combo_export_template_prewarm_form_data(local_skus),
        headers=_private_html_post_headers(private_cookie_header),
    ) as resp:
        await _read_http_ok_response(resp, action="组合SKU预热 2")


def _int_value(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _normalize_step1_response(payload: dict[str, Any]) -> tuple[str, int, int | None]:
    sn = _clean_text(payload.get("sn"))
    if not sn:
        raise StoreMskuActualInventoryError("组合SKU导出 Step 1 返回缺少 sn")
    subtask_num = _int_value(payload.get("subtask_num"))
    if not subtask_num or subtask_num <= 0:
        raise StoreMskuActualInventoryError(f"组合SKU导出 Step 1 返回 subtask_num 无效: {payload.get('subtask_num')}")
    chunk_num = _int_value(payload.get("chunkNum"))
    return sn, subtask_num, chunk_num


def _validate_step2_response(payload: dict[str, Any], *, sub_no: int) -> None:
    sub_items = payload.get("subO")
    if not isinstance(sub_items, list) or not sub_items:
        raise StoreMskuActualInventoryError(f"组合SKU导出 Step 2 返回缺少 subO: sub_no={sub_no}")
    for item in sub_items:
        if not isinstance(item, dict):
            raise StoreMskuActualInventoryError(f"组合SKU导出 Step 2 返回 subO 格式异常: sub_no={sub_no}")
        if _clean_text(item.get("success")) != "1":
            raise StoreMskuActualInventoryError(f"组合SKU导出 Step 2 失败: sub_no={sub_no}, subO={item}")


def _normalize_step3_response(payload: dict[str, Any]) -> str:
    task_id = _clean_text(payload.get("taskId"))
    if not task_id:
        raise StoreMskuActualInventoryError("组合SKU导出 Step 3 返回缺少 taskId")
    return task_id


async def _wait_for_combo_file_url(
    *,
    sn: str,
    task_id: str,
    cookie_header: str,
    timeout_sec: float,
    poll_interval_sec: float,
) -> str:
    safe_timeout = max(0.0, float(timeout_sec))
    safe_interval = max(0.1, float(poll_interval_sec))
    deadline = asyncio.get_running_loop().time() + safe_timeout
    last_state = "unknown"

    while True:
        payload = await _post_combo_export(
            _step4_form_data(sn=sn, task_id=task_id),
            cookie_header=cookie_header,
            action="组合SKU导出 Step 4",
        )
        state = _clean_text(payload.get("state"))
        last_state = state or "unknown"
        file_url = _clean_text(payload.get("file_url"))
        if state == "1":
            if not file_url:
                raise StoreMskuActualInventoryError(f"组合SKU导出完成但缺少 file_url: taskId={task_id}")
            return file_url

        now = asyncio.get_running_loop().time()
        if now >= deadline:
            break
        await asyncio.sleep(min(safe_interval, max(0.0, deadline - now)))

    raise StoreMskuActualInventoryTimeoutError(
        f"组合SKU导出超时(taskId={task_id}, timeout={safe_timeout:g}s, last_state={last_state})"
    )


async def _download_xlsx_from_url(file_url: str, target_path: Path, *, action: str) -> Path:
    url = _clean_text(file_url)
    if not url:
        raise ValueError("file_url 不能为空")
    target_path.parent.mkdir(parents=True, exist_ok=True)
    headers = {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream,*/*"
    }
    async with external_http_session.get(url, headers=headers) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"{action}失败(status={status_code}): {msg}")
        if not body:
            raise StoreMskuActualInventoryError(f"{action}返回空文件")
    target_path.write_bytes(body)
    return target_path


async def export_combo_sku_xlsx(
    local_skus: list[str],
    *,
    store_name: str,
    cookie_header: str,
    memcache_key: str,
    output_dir: str | Path | None = None,
    timeout_sec: float = 180,
    poll_interval_sec: float = 3,
) -> Path:
    step1 = await _post_combo_export(
        _combo_step1_form_data(local_skus, memcache_key=memcache_key),
        cookie_header=cookie_header,
        action="组合SKU导出 Step 1",
    )
    sn, subtask_num, _chunk_num = _normalize_step1_response(step1)
    for sub_no in range(1, subtask_num + 1):
        step2 = await _post_combo_export(
            _step2_form_data(sn=sn, sub_no=sub_no),
            cookie_header=cookie_header,
            action="组合SKU导出 Step 2",
        )
        _validate_step2_response(step2, sub_no=sub_no)
    step3 = await _post_combo_export(
        _step3_form_data(sn=sn),
        cookie_header=cookie_header,
        action="组合SKU导出 Step 3",
    )
    task_id = _normalize_step3_response(step3)
    file_url = await _wait_for_combo_file_url(
        sn=sn,
        task_id=task_id,
        cookie_header=cookie_header,
        timeout_sec=timeout_sec,
        poll_interval_sec=poll_interval_sec,
    )
    directory = _resolve_output_dir(output_dir)
    target_path = directory / f"{_timestamp_text()}-{_safe_file_part(store_name)}_combo_sku.xlsx"
    return await _download_xlsx_from_url(file_url, target_path, action="下载组合SKU导出xlsx")


def parse_combo_sku_xlsx(xlsx_path: str | Path) -> dict[str, ComboSku]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取组合SKU导出xlsx") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"组合SKU导出xlsx不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        missing = [column for column in (COMBO_SKU_COLUMN, COMBO_COMPONENT_COUNT_COLUMN) if column not in headers]
        if missing:
            raise StoreMskuActualInventoryError(f"组合SKU导出xlsx缺少列: {', '.join(missing)}")

        combos: dict[str, ComboSku] = {}
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            combo_sku = _clean_text(row.get(COMBO_SKU_COLUMN))
            if not combo_sku:
                continue
            component_count = _int_value(row.get(COMBO_COMPONENT_COUNT_COLUMN))
            if component_count is None or component_count <= 0:
                raise StoreMskuActualInventoryError(f"组合SKU关联sku个数无效: {combo_sku}")
            components: list[ComboComponent] = []
            for index in range(1, component_count + 1):
                stock_sku_column = f"关联sku编号{index}"
                quantity_column = f"关联sku捆绑数量{index}"
                if stock_sku_column not in headers or quantity_column not in headers:
                    raise StoreMskuActualInventoryError(
                        f"组合SKU导出xlsx缺少列: {stock_sku_column}, {quantity_column}"
                    )
                stock_sku = _clean_text(row.get(stock_sku_column))
                quantity = _decimal_value(row.get(quantity_column), default=Decimal("0"))
                if not stock_sku:
                    raise StoreMskuActualInventoryError(f"组合SKU缺少关联sku编号: {combo_sku}, index={index}")
                if quantity <= 0:
                    raise StoreMskuActualInventoryError(f"组合SKU关联sku捆绑数量无效: {combo_sku}, {stock_sku}")
                components.append(ComboComponent(stock_sku=stock_sku, quantity=quantity))
            combos[normalize_sku_key(combo_sku)] = ComboSku(combo_sku=combo_sku, components=tuple(components))
    except StoreMskuActualInventoryError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取组合SKU导出xlsx失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass
    return combos


def filter_combo_map_for_source(
    combo_map: dict[str, ComboSku],
    *,
    source_local_skus: list[str],
) -> dict[str, ComboSku]:
    sentinel_key = normalize_sku_key(SENTINEL_COMBO_SKU)
    if sentinel_key not in combo_map:
        raise StoreMskuActualInventoryError(f"组合SKU导出结果缺少哨兵SKU: {SENTINEL_COMBO_SKU}")
    source_keys = {normalize_sku_key(sku) for sku in source_local_skus if normalize_sku_key(sku)}
    return {key: combo for key, combo in combo_map.items() if key in source_keys}


def _warehouse_search_form_data(stock_skus: list[str]) -> list[tuple[str, str]]:
    return [
        ("stockOrderby", ""),
        ("parentCategoryId", ""),
        ("categoryId", ""),
        ("third_category_id", ""),
        ("warehouseIds[]", WAREHOUSE_ID),
        ("stockName", "nameCN"),
        ("stockNameValue", ""),
        ("statusIN[]", "3"),
        ("inventoryAlertId", "0"),
        ("livenessType", ""),
        ("isNewType", ""),
        ("gridcodeStr", ""),
        ("stockSkuStr", _ids_lines(stock_skus)),
        ("page", "1"),
        ("rowsPerPage", "50"),
        ("warehouseId", "undefined"),
        ("startTime", ""),
        ("endTime", ""),
        ("isIdn", "1"),
        ("warehouseIdArr", ""),
        ("stockQuantitylt", ""),
        ("stockQuantitygt", ""),
        ("stockWarningQuantitylt", ""),
        ("stockWarningQuantitygt", ""),
        ("saleAvailableDayslt", ""),
        ("saleAvailableDaysgt", ""),
    ]


async def search_warehouse_stock(stock_skus: list[str], *, cookie_header: str) -> None:
    if not stock_skus:
        return
    async with erp_http_session.post(
        _warehouse_search_url(),
        data=_warehouse_search_form_data(stock_skus),
        headers=_private_amz_post_headers(cookie_header),
    ) as resp:
        await _read_optional_json_response(resp, action="库存SKU搜索")


async def download_warehouse_stock_xlsx(
    *,
    cookie_header: str,
    store_name: str,
    output_dir: str | Path | None = None,
) -> Path:
    directory = _resolve_output_dir(output_dir)
    target_path = directory / f"{_timestamp_text()}-{_safe_file_part(store_name)}_warehouse_stock.xlsx"
    async with erp_http_session.get(
        _warehouse_export_url(),
        headers=_private_amz_get_headers(cookie_header),
    ) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code in AUTH_FAIL_STATUS:
            raise StoreMskuActualInventoryAuthError(f"库存SKU导出鉴权失败(status={status_code})")
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"库存SKU导出请求失败(status={status_code}): {msg}")
        if not body:
            raise StoreMskuActualInventoryError("库存SKU导出返回空文件")
    target_path.write_bytes(body)
    return target_path


def parse_stock_inventory_xlsx(xlsx_path: str | Path) -> dict[str, Decimal]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取库存SKU导出xlsx") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"库存SKU导出xlsx不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        missing = [column for column in (STOCK_SKU_COLUMN, AVAILABLE_STOCK_COLUMN) if column not in headers]
        if missing:
            raise StoreMskuActualInventoryError(f"库存SKU导出xlsx缺少列: {', '.join(missing)}")

        quantities: dict[str, Decimal] = {}
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            stock_sku = _clean_text(row.get(STOCK_SKU_COLUMN))
            key = normalize_sku_key(stock_sku)
            if not key:
                continue
            quantities[key] = quantities.get(key, Decimal("0")) + _decimal_value(row.get(AVAILABLE_STOCK_COLUMN))
    except StoreMskuActualInventoryError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取库存SKU导出xlsx失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass
    return quantities


def stock_skus_for_inventory(local_skus: list[str], combo_map: dict[str, ComboSku]) -> list[str]:
    unique: OrderedDict[str, str] = OrderedDict()
    for local_sku in local_skus:
        key = normalize_sku_key(local_sku)
        if not key:
            continue
        combo = combo_map.get(key)
        if combo is None:
            unique.setdefault(key, local_sku)
            continue
        for component in combo.components:
            component_key = normalize_sku_key(component.stock_sku)
            if component_key:
                unique.setdefault(component_key, component.stock_sku)
    return list(unique.values())


def _combo_child_text(combo: ComboSku) -> str:
    return ", ".join(f"{component.stock_sku} * {_display_quantity(component.quantity)}" for component in combo.components)


def _fba_total_inventory(row: StoreMskuRow) -> Decimal:
    return (
        row.fba_sellable
        + row.fba_inbound
        + row.fba_reserved
        + row.fba_in_transit
        + row.fba_pending_transfer
        + row.fba_transferring
    )


def _weighted_daily_sales(row: StoreMskuRow) -> Decimal:
    return (
        row.sales_7d / Decimal("7") * Decimal("0.6")
        + row.sales_14d / Decimal("14") * Decimal("0.3")
        + row.sales_30d / Decimal("30") * Decimal("0.1")
    )


def _sales_days(fba_total_inventory: Decimal, weighted_daily_sales: Decimal) -> Decimal | None:
    if weighted_daily_sales == 0:
        return None
    return fba_total_inventory / weighted_daily_sales


def calculate_inventory_rows(
    msku_rows: list[StoreMskuRow],
    *,
    combo_map: dict[str, ComboSku],
    stock_quantities: dict[str, Decimal],
) -> tuple[list[ActualInventoryRow], list[str]]:
    missing: OrderedDict[str, str] = OrderedDict()
    result_rows: list[ActualInventoryRow] = []
    for row in msku_rows:
        local_key = normalize_sku_key(row.local_sku)
        combo = combo_map.get(local_key)
        is_combo_sku = combo is not None
        fba_total_inventory = _fba_total_inventory(row)
        weighted_daily_sales = _weighted_daily_sales(row)
        actual_inventory: Decimal | None
        child_skus = ""
        if not local_key:
            actual_inventory = None
        elif combo is None:
            actual_inventory = stock_quantities.get(local_key)
            if actual_inventory is None:
                missing.setdefault(local_key, row.local_sku)
        else:
            child_skus = _combo_child_text(combo)
            possible_counts: list[Decimal] = []
            for component in combo.components:
                component_key = normalize_sku_key(component.stock_sku)
                quantity = stock_quantities.get(component_key)
                if quantity is None:
                    missing.setdefault(component_key, component.stock_sku)
                    continue
                possible_counts.append((quantity / component.quantity).to_integral_value(rounding=ROUND_FLOOR))
            actual_inventory = min(possible_counts) if len(possible_counts) == len(combo.components) else None

        result_rows.append(
            ActualInventoryRow(
                msku=row.msku,
                parent_asin=row.parent_asin,
                asin=row.asin,
                local_sku=row.local_sku,
                product_link=row.product_link,
                actual_inventory=actual_inventory,
                child_skus=child_skus,
                is_combo_sku=is_combo_sku,
                fba_total_inventory=fba_total_inventory,
                weighted_daily_sales=weighted_daily_sales,
                sales_days=_sales_days(fba_total_inventory, weighted_daily_sales),
            )
        )
    return result_rows, list(missing.values())


def split_inventory_rows(rows: list[ActualInventoryRow]) -> ActualInventoryRowGroups:
    combo_inventory_rows: list[ActualInventoryRow] = []
    stock_inventory_rows: list[ActualInventoryRow] = []
    no_local_sku_rows: list[ActualInventoryRow] = []
    no_inventory_rows: list[ActualInventoryRow] = []
    for row in rows:
        if not normalize_sku_key(row.local_sku):
            no_local_sku_rows.append(row)
        elif row.actual_inventory is None:
            no_inventory_rows.append(row)
        elif row.is_combo_sku:
            combo_inventory_rows.append(row)
        else:
            stock_inventory_rows.append(row)
    return ActualInventoryRowGroups(
        combo_inventory_rows=combo_inventory_rows,
        stock_inventory_rows=stock_inventory_rows,
        no_local_sku_rows=no_local_sku_rows,
        no_inventory_rows=no_inventory_rows,
    )


def _base_row_values(row: ActualInventoryRow) -> list[Any]:
    return [
        row.msku,
        row.parent_asin,
        row.asin,
        row.local_sku,
        row.product_link,
        _display_decimal(row.actual_inventory),
        row.child_skus,
    ]


def _inventory_row_values(row: ActualInventoryRow) -> list[Any]:
    return [
        row.msku,
        row.parent_asin,
        row.asin,
        row.local_sku,
        row.product_link,
        _display_decimal(row.fba_total_inventory),
        _display_two_decimal(row.weighted_daily_sales),
        _display_two_decimal(row.sales_days),
        _display_decimal(row.actual_inventory),
        row.child_skus,
    ]


def _sorted_inventory_rows(rows: list[ActualInventoryRow]) -> list[ActualInventoryRow]:
    return sorted(rows, key=lambda row: row.weighted_daily_sales, reverse=True)


def _append_inventory_sheet(
    workbook: Any,
    title: str,
    columns: tuple[str, ...],
    rows: list[ActualInventoryRow],
    row_values: Any,
    *,
    active: bool = False,
    highlight_actual_inventory: bool = False,
) -> None:
    from openpyxl.styles import PatternFill

    worksheet = workbook.active if active else workbook.create_sheet(title)
    worksheet.title = title
    worksheet.append(list(columns))
    for row in rows:
        worksheet.append(row_values(row))
    for index, header in enumerate(columns, start=1):
        if header not in TWO_DECIMAL_COLUMNS:
            continue
        for cells in worksheet.iter_rows(min_row=2, min_col=index, max_col=index):
            cells[0].number_format = "0.00"
    if highlight_actual_inventory:
        fill = PatternFill(fill_type="solid", fgColor=ACTUAL_INVENTORY_HIGHLIGHT_COLOR)
        for column_index, header in enumerate(columns, start=1):
            if header not in INVENTORY_HIGHLIGHT_COLUMNS:
                continue
            for cells in worksheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index):
                cells[0].fill = fill
    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_format.defaultRowHeight = EXCEL_ROW_HEIGHT
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = EXCEL_ROW_HEIGHT
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = EXCEL_COLUMN_WIDTH


def write_actual_inventory_xlsx(rows: list[ActualInventoryRow], output_path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入真实库存xlsx") from exc

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        groups = split_inventory_rows(rows)
        _append_inventory_sheet(
            workbook,
            "真实库存-组合sku",
            INVENTORY_OUTPUT_COLUMNS,
            _sorted_inventory_rows(groups.combo_inventory_rows),
            _inventory_row_values,
            active=True,
            highlight_actual_inventory=True,
        )
        _append_inventory_sheet(
            workbook,
            "真实库存-库存sku",
            INVENTORY_OUTPUT_COLUMNS,
            _sorted_inventory_rows(groups.stock_inventory_rows),
            _inventory_row_values,
            highlight_actual_inventory=True,
        )
        _append_inventory_sheet(workbook, "无本地SKU", BASE_OUTPUT_COLUMNS, groups.no_local_sku_rows, _base_row_values)
        _append_inventory_sheet(workbook, "无库存数据", BASE_OUTPUT_COLUMNS, groups.no_inventory_rows, _base_row_values)
        workbook.save(target_path)
    finally:
        workbook.close()
    return target_path


async def export_store_msku_actual_inventory(
    store_name: str,
    *,
    input_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    timeout_sec: float = 180,
    poll_interval_sec: float = 3,
) -> ActualInventoryResult:
    clean_store_name = normalize_store_name(store_name)
    source = find_latest_store_msku_file(clean_store_name, input_dir=input_dir)
    msku_rows = load_store_msku_rows(source.path)
    local_skus = _unique_text([row.local_sku for row in msku_rows])

    output_directory = _resolve_output_dir(output_dir)
    private_cookie, memcache_key = await _resolve_private_auth()
    private_amz_cookie = await _resolve_private_amz_cookie()
    await prewarm_combo_sku_export(
        local_skus,
        private_cookie_header=private_cookie,
        private_amz_cookie_header=private_amz_cookie,
    )
    combo_xlsx_path = await export_combo_sku_xlsx(
        local_skus,
        store_name=clean_store_name,
        cookie_header=private_cookie,
        memcache_key=memcache_key,
        output_dir=output_directory,
        timeout_sec=timeout_sec,
        poll_interval_sec=poll_interval_sec,
    )
    combo_map = filter_combo_map_for_source(parse_combo_sku_xlsx(combo_xlsx_path), source_local_skus=local_skus)
    stock_skus = stock_skus_for_inventory(local_skus, combo_map)

    await search_warehouse_stock(stock_skus, cookie_header=private_amz_cookie)
    stock_xlsx_path = await download_warehouse_stock_xlsx(
        cookie_header=private_amz_cookie,
        store_name=clean_store_name,
        output_dir=output_directory,
    )
    stock_quantities = parse_stock_inventory_xlsx(stock_xlsx_path)
    inventory_rows, missing_stock_skus = calculate_inventory_rows(
        msku_rows,
        combo_map=combo_map,
        stock_quantities=stock_quantities,
    )
    inventory_groups = split_inventory_rows(inventory_rows)

    final_xlsx_path = output_directory / f"{source.source_data_time}-{_safe_file_part(clean_store_name)}_actual_inventory.xlsx"
    write_actual_inventory_xlsx(inventory_rows, final_xlsx_path)
    return ActualInventoryResult(
        store_name=clean_store_name,
        source_xlsx_path=str(source.path),
        source_data_time=source.source_data_time,
        local_sku_count=len(local_skus),
        combo_sku_count=len(combo_map),
        stock_sku_count=len(stock_skus),
        missing_stock_skus=missing_stock_skus,
        xlsx_path=str(final_xlsx_path),
        inventory_row_count=len(inventory_groups.inventory_rows),
        no_local_sku_count=len(inventory_groups.no_local_sku_rows),
        no_inventory_row_count=len(inventory_groups.no_inventory_rows),
    )


__all__ = [
    "AVAILABLE_STOCK_COLUMN",
    "BASE_OUTPUT_COLUMNS",
    "COMBO_EXPORT_FIELDS",
    "COMBO_SKU_COLUMN",
    "FBA_STOCK_COLUMNS",
    "INVENTORY_OUTPUT_COLUMNS",
    "OUTPUT_COLUMNS",
    "SALES_COLUMNS",
    "SENTINEL_COMBO_SKU",
    "SOURCE",
    "STOCK_SKU_COLUMN",
    "WAREHOUSE_ID",
    "ActualInventoryResult",
    "ActualInventoryRow",
    "ActualInventoryRowGroups",
    "ComboComponent",
    "ComboSku",
    "StoreMskuActualInventoryAuthError",
    "StoreMskuActualInventoryError",
    "StoreMskuActualInventoryTimeoutError",
    "StoreMskuRow",
    "calculate_inventory_rows",
    "combo_query_skus",
    "export_combo_sku_xlsx",
    "export_store_msku_actual_inventory",
    "filter_combo_map_for_source",
    "find_latest_store_msku_file",
    "load_store_msku_rows",
    "normalize_sku_key",
    "parse_combo_sku_xlsx",
    "parse_stock_inventory_xlsx",
    "prewarm_combo_sku_export",
    "search_warehouse_stock",
    "split_inventory_rows",
    "stock_skus_for_inventory",
    "write_actual_inventory_xlsx",
    "_combo_step1_form_data",
    "_timestamp_text",
    "_warehouse_search_form_data",
]
