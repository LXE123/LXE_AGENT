from __future__ import annotations

import copy
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

DEFAULT_TEMPLATE_NAME = "默认"
US_GROUP_1_TEMPLATE_NAME = "US-一组"
UK_GROUP_1_TEMPLATE_NAME = "UK-一组"
DE_GROUP_1_TEMPLATE_NAME = "DE-一组"
US_LIN_MEIQI_GROUP_2_TEMPLATE_NAME = "2组-US站点-林美淇"
DEFAULT_TEMPLATE_VERSION = 1
DEFAULT_TEMPLATE_DIR = Path(__file__).resolve().parent / "replenishment_templates"
DEFAULT_TEMPLATE_PATH = DEFAULT_TEMPLATE_DIR / "default_template.json"
BUILTIN_TEMPLATE_PATHS = (
    DEFAULT_TEMPLATE_PATH,
    DEFAULT_TEMPLATE_DIR / "us_group_1_template.json",
    DEFAULT_TEMPLATE_DIR / "uk_group_1_template.json",
    DEFAULT_TEMPLATE_DIR / "de_group_1_template.json",
    DEFAULT_TEMPLATE_DIR / "us_lin_meiqi_group_2_template.json",
)
BUILTIN_TEMPLATE_NAMES = (
    DEFAULT_TEMPLATE_NAME,
    US_GROUP_1_TEMPLATE_NAME,
    UK_GROUP_1_TEMPLATE_NAME,
    DE_GROUP_1_TEMPLATE_NAME,
    US_LIN_MEIQI_GROUP_2_TEMPLATE_NAME,
)
DEFAULT_CUSTOM_TEMPLATE_STORE = Path("artifacts") / "mabang_replenishment_templates" / "templates.json"
DEFAULT_EDITABLE_OUTPUT_DIR = Path("artifacts") / "mabang_replenishment_templates" / "editable"

SOURCE_DEFAULT = "default"
SOURCE_CUSTOM = "custom"
SOURCE = "mabang_replenishment_template"
REPLENISHMENT_TEMPLATE_FILE_SUFFIX = "备货算法配置表"

TEMPLATE_INFO_SHEET = "参数方案信息"
WEIGHTED_SALES_SHEET = "日销计算"
REPLENISHMENT_DAYS_SHEET = "空运补货天数"
SHIPPING_SHEET = "空运判断"
SEA_ENTRY_SHEET = "海运进入条件"
SEA_DAYS_SHEET = "海运补货天数"
SEA_COMPANION_AIR_SHEET = "海运同时空运"
SPECIAL_RULES_SHEET = "特殊MSKU规则"
SHEET_NAMES = (
    TEMPLATE_INFO_SHEET,
    WEIGHTED_SALES_SHEET,
    REPLENISHMENT_DAYS_SHEET,
    SHIPPING_SHEET,
    SEA_ENTRY_SHEET,
    SEA_DAYS_SHEET,
    SEA_COMPANION_AIR_SHEET,
    SPECIAL_RULES_SHEET,
)
OLD_TEMPLATE_SHEET_NAMES = ("模板信息", "加权日销", "运输方式", "海运规则", "补货天数", "海运备货天数")
OLD_TEMPLATE_ERROR = "旧版备货算法配置表不再支持，请重新导出新版备货算法配置表后修改"

EXCEL_ROW_HEIGHT = 15
EXCEL_COLUMN_WIDTH = 15
FLOAT_TOLERANCE = 1e-9
HEADER_FILL_COLOR = "FF1F4E78"
EDITABLE_FILL_COLOR = "FFFFF2CC"
READONLY_FILL_COLOR = "FFE7E6E6"
WHITE_FONT_COLOR = "FFFFFFFF"

TREND_KEYS = ("growth", "stable", "decline")
MSKU_SPLIT_RE = re.compile(r"[\s,，;；、]+")
RANGE_INTERVAL_RE = re.compile(r"^([\(\[])(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)([\]\)])$")
RANGE_LOWER_RE = re.compile(r"^(>=|>)(-?\d+(?:\.\d+)?)$")
RANGE_UPPER_RE = re.compile(r"^(<=|<)(-?\d+(?:\.\d+)?)$")
TRUE_TEXTS = {"1", "true", "yes", "y", "on", "是", "启用", "开启", "开"}
FALSE_TEXTS = {"0", "false", "no", "n", "off", "否", "禁用", "关闭", "关"}


class ReplenishmentTemplateError(ValueError):
    pass


@dataclass(frozen=True)
class ReplenishmentTemplate:
    name: str
    version: int
    description: str
    params: dict[str, Any]
    source: str = SOURCE_CUSTOM

    def to_payload(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "version": self.version,
            "description": self.description,
            "source": self.source,
            "params": copy.deepcopy(self.params),
        }

    def to_store_payload(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "version": self.version,
            "description": self.description,
            "params": copy.deepcopy(self.params),
        }


@dataclass(frozen=True)
class DailySalesRange:
    lower: float | None
    upper: float | None
    lower_inclusive: bool = False
    is_fallback: bool = False


@dataclass(frozen=True)
class TemplateValidationResult:
    template: ReplenishmentTemplate
    warnings: tuple[str, ...] = ()

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "template": self.template.to_payload(),
            "warnings": list(self.warnings),
            "source": SOURCE,
        }


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _number(value: Any, *, default: float | None = None) -> float | None:
    if value is None:
        return default
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        number = float(value)
        return default if math.isnan(number) else number
    text = _clean_text(value).replace(",", "")
    if not text or text.lower() == "nan":
        return default
    try:
        number = float(text)
    except (TypeError, ValueError):
        return default
    return default if math.isnan(number) else number


def _int_value(value: Any, *, default: int | None = None) -> int | None:
    number = _number(value, default=None)
    if number is None:
        return default
    if not float(number).is_integer():
        raise ReplenishmentTemplateError(f"必须是整数: {value}")
    return int(number)


def _bool_value(value: Any, *, field_name: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if math.isnan(number):
            raise ReplenishmentTemplateError(f"{field_name}必须是布尔值: {value}")
        if number == 1:
            return True
        if number == 0:
            return False
        raise ReplenishmentTemplateError(f"{field_name}必须是布尔值: {value}")
    text = _clean_text(value).casefold()
    if text in TRUE_TEXTS:
        return True
    if text in FALSE_TEXTS:
        return False
    raise ReplenishmentTemplateError(f"{field_name}必须是布尔值: {value}")


def _timestamp_text() -> str:
    return datetime.now().strftime("%Y%m%d%H%M")


def _safe_file_part(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"[^A-Za-z0-9_.\-\u4e00-\u9fff]+", "_", text)
    return text.strip("._-") or "template"


def _display_number_text(value: float | int | None) -> str:
    if value is None:
        return ""
    number = float(value)
    return str(int(number)) if number.is_integer() else f"{number:g}"


def _normalize_range_text(value: Any) -> str:
    return (
        _clean_text(value)
        .replace(" ", "")
        .replace("，", ",")
        .replace("（", "(")
        .replace("）", ")")
        .replace("【", "[")
        .replace("】", "]")
        .replace("＞", ">")
        .replace("＜", "<")
        .replace("≥", ">=")
        .replace("≤", "<=")
    )


def _range_error(sheet_name: str, row_number: int, value: Any) -> ReplenishmentTemplateError:
    return ReplenishmentTemplateError(
        f"{sheet_name}第{row_number}行日销范围格式无效: {value}；"
        "支持 >10、>=1、(5,10]、[1,5]、≤2、<=2、低销量兜底"
    )


def _parse_daily_sales_range(value: Any, *, sheet_name: str, row_number: int) -> DailySalesRange:
    text = _normalize_range_text(value)
    if text in {"低销量兜底", "兜底", "低销量"}:
        return DailySalesRange(lower=None, upper=None, is_fallback=True)
    if not text:
        raise _range_error(sheet_name, row_number, value)

    lower_match = RANGE_LOWER_RE.match(text)
    if lower_match:
        lower = _number(lower_match.group(2), default=None)
        if lower is None:
            raise _range_error(sheet_name, row_number, value)
        return DailySalesRange(lower=lower, upper=None, lower_inclusive=lower_match.group(1) == ">=")

    upper_match = RANGE_UPPER_RE.match(text)
    if upper_match:
        upper = _number(upper_match.group(2), default=None)
        if upper is None:
            raise _range_error(sheet_name, row_number, value)
        return DailySalesRange(lower=None, upper=upper, is_fallback=True)

    interval_match = RANGE_INTERVAL_RE.match(text)
    if interval_match:
        lower = _number(interval_match.group(2), default=None)
        upper = _number(interval_match.group(3), default=None)
        if lower is None or upper is None or lower >= upper:
            raise _range_error(sheet_name, row_number, value)
        return DailySalesRange(
            lower=lower,
            upper=upper,
            lower_inclusive=interval_match.group(1) == "[",
        )

    raise _range_error(sheet_name, row_number, value)


def _replenishment_range_label(current_gt: Any, previous_gt: float | None) -> str:
    gt = _number(current_gt, default=None)
    if gt is None:
        return "低销量兜底"
    if previous_gt is None:
        return f">{_display_number_text(gt)}"
    return f"({_display_number_text(gt)},{_display_number_text(previous_gt)}]"


def _sea_range_label(item: dict[str, Any], params: dict[str, Any]) -> str:
    gt = _number(item.get("daily_sales_gt"), default=None)
    lte = _number(item.get("daily_sales_lte"), default=None)
    inclusive_min = False
    if gt is not None:
        min_daily_sales = _number(params.get("sea", {}).get("min_daily_sales"), default=None)
        inclusive_min = (
            min_daily_sales is not None
            and sea_min_daily_sales_inclusive_from_template(params)
            and abs(float(gt) - float(min_daily_sales)) <= 0.0000001
        )
    if gt is None and lte is None:
        return "全部"
    if gt is None:
        return f"≤{_display_number_text(lte)}"
    if lte is None:
        return f">={_display_number_text(gt)}" if inclusive_min else f">{_display_number_text(gt)}"
    left = "[" if inclusive_min else "("
    return f"{left}{_display_number_text(gt)},{_display_number_text(lte)}]"


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _custom_store_path(path: str | Path | None = None) -> Path:
    return Path(path) if path is not None else DEFAULT_CUSTOM_TEMPLATE_STORE


def _editable_output_dir(path: str | Path | None = None) -> Path:
    directory = Path(path) if path is not None else DEFAULT_EDITABLE_OUTPUT_DIR
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def _default_template_path(path: str | Path | None = None) -> Path:
    return Path(path) if path is not None else DEFAULT_TEMPLATE_PATH


def is_builtin_template_name(value: Any) -> bool:
    return _clean_text(value) in set(BUILTIN_TEMPLATE_NAMES)


def _normalize_template(raw: dict[str, Any], *, source: str) -> ReplenishmentTemplate:
    if not isinstance(raw, dict):
        raise ReplenishmentTemplateError("参数方案内容必须是JSON对象")
    name = _clean_text(raw.get("name"))
    if not name:
        raise ReplenishmentTemplateError("参数方案名不能为空")
    version = _int_value(raw.get("version"), default=DEFAULT_TEMPLATE_VERSION)
    if version is None or version <= 0:
        raise ReplenishmentTemplateError("参数方案版本必须是正整数")
    params = raw.get("params")
    if not isinstance(params, dict):
        raise ReplenishmentTemplateError(f"参数方案缺少params: {name}")
    return ReplenishmentTemplate(
        name=name,
        version=version,
        description=_clean_text(raw.get("description")),
        params=copy.deepcopy(params),
        source=source,
    )


def load_default_template(*, path: str | Path | None = None) -> ReplenishmentTemplate:
    default_path = _default_template_path(path)
    if not default_path.is_file():
        raise FileNotFoundError(f"内置参数方案文件不存在: {default_path}")
    template = _normalize_template(_read_json(default_path), source=SOURCE_DEFAULT)
    if template.name != DEFAULT_TEMPLATE_NAME:
        raise ReplenishmentTemplateError(f"内置参数方案名必须是{DEFAULT_TEMPLATE_NAME}: {template.name}")
    return template


def load_builtin_templates() -> list[ReplenishmentTemplate]:
    templates: list[ReplenishmentTemplate] = []
    seen_names: set[str] = set()
    for template_path, expected_name in zip(BUILTIN_TEMPLATE_PATHS, BUILTIN_TEMPLATE_NAMES, strict=True):
        if not template_path.is_file():
            raise FileNotFoundError(f"内置参数方案文件不存在: {template_path}")
        template = _normalize_template(_read_json(template_path), source=SOURCE_DEFAULT)
        if template.name != expected_name:
            raise ReplenishmentTemplateError(f"内置参数方案名必须是{expected_name}: {template.name}")
        if template.name in seen_names:
            raise ReplenishmentTemplateError(f"内置参数方案名重复: {template.name}")
        seen_names.add(template.name)
        templates.append(template)
    return templates


def _load_custom_store(path: str | Path | None = None) -> list[dict[str, Any]]:
    store_path = _custom_store_path(path)
    if not store_path.is_file():
        return []
    data = _read_json(store_path)
    templates = data.get("templates") if isinstance(data, dict) else None
    if not isinstance(templates, list):
        raise ReplenishmentTemplateError(f"自定义参数方案库格式错误: {store_path}")
    return [item for item in templates if isinstance(item, dict)]


def load_custom_templates(*, store_path: str | Path | None = None) -> list[ReplenishmentTemplate]:
    return [_normalize_template(item, source=SOURCE_CUSTOM) for item in _load_custom_store(store_path)]


def list_templates(*, store_path: str | Path | None = None) -> list[ReplenishmentTemplate]:
    templates = load_builtin_templates()
    custom_templates = load_custom_templates(store_path=store_path)
    builtin_names = {template.name for template in templates}
    for template in custom_templates:
        if template.name in builtin_names:
            raise ReplenishmentTemplateError(f"自定义参数方案不允许覆盖系统参数方案: {template.name}")
    names = set(builtin_names)
    for template in custom_templates:
        if template.name in names:
            raise ReplenishmentTemplateError(f"参数方案名重复: {template.name}")
        names.add(template.name)
        templates.append(template)
    return templates


def get_template(template_name: str | None = None, *, store_path: str | Path | None = None) -> ReplenishmentTemplate:
    clean_name = _clean_text(template_name) or DEFAULT_TEMPLATE_NAME
    for template in list_templates(store_path=store_path):
        if template.name == clean_name:
            return template
    raise ReplenishmentTemplateError(f"未知备货算法参数方案: {clean_name}")


def _next_custom_template_name(existing_names: set[str]) -> str:
    index = 1
    while True:
        name = f"自定义参数方案{index}"
        if name not in existing_names:
            return name
        index += 1


def list_parameter_groups() -> list[dict[str, Any]]:
    return [
        {
            "group": "日销计算",
            "params": [
                {"key": "weighted_sales.7d_weight", "name": "7天销量权重", "value_type": "number"},
                {"key": "weighted_sales.14d_weight", "name": "14天销量权重", "value_type": "number"},
                {"key": "weighted_sales.30d_weight", "name": "30天销量权重", "value_type": "number"},
            ],
        },
        {
            "group": "空运补货天数",
            "params": [
                {"key": "replenishment_days", "name": "日销分档下增长/平稳/下降空运补货天数", "value_type": "matrix"},
            ],
        },
        {
            "group": "空运判断",
            "params": [
                {"key": "shipping.air_urgent_sales_days_lte", "name": "空运急发可销售天数阈值", "value_type": "number"},
                {"key": "shipping.air_sales_days_lte", "name": "空运可销售天数阈值", "value_type": "number"},
            ],
        },
        {
            "group": "海运进入条件",
            "params": [
                {"key": "sea.enabled", "name": "是否启用海运", "value_type": "boolean"},
                {"key": "sea.min_daily_sales", "name": "海运最低加权日销", "value_type": "number"},
                {"key": "sea.min_daily_sales_inclusive", "name": "海运最低加权日销是否包含等于", "value_type": "boolean"},
                {"key": "sea.min_weight_kg", "name": "海运最低总重量kg", "value_type": "number"},
                {"key": "sea.min_net_quantity", "name": "海运建议量最小件数", "value_type": "number"},
            ],
        },
        {
            "group": "海运补货天数",
            "params": [
                {"key": "sea.tiers", "name": "海运日销分档和补货天数", "value_type": "matrix"},
            ],
        },
        {
            "group": "海运同时空运",
            "params": [
                {"key": "sea.companion_air_enabled", "name": "是否启用海运同时备空运", "value_type": "boolean"},
                {"key": "sea.companion_air_tiers", "name": "海运同时空运日销分档和天数", "value_type": "matrix"},
            ],
        },
        {
            "group": "特殊MSKU规则",
            "params": [
                {"key": "special_rules", "name": "指定MSKU命中特殊参数覆盖", "value_type": "table"},
            ],
        },
    ]


def _require_mapping(value: Any, name: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ReplenishmentTemplateError(f"参数方案参数缺少{name}")
    return value


def _require_list(value: Any, name: str) -> list[Any]:
    if not isinstance(value, list):
        raise ReplenishmentTemplateError(f"参数方案参数{name}必须是列表")
    return value


def validate_template(template: ReplenishmentTemplate) -> TemplateValidationResult:
    params = template.params
    warnings: list[str] = []

    weighted = _require_mapping(params.get("weighted_sales"), "weighted_sales")
    weights = []
    for key in ("7d_weight", "14d_weight", "30d_weight"):
        number = _number(weighted.get(key), default=None)
        if number is None or number < 0:
            raise ReplenishmentTemplateError(f"加权日销权重必须是非负数: {key}")
        weights.append(number)
    if abs(sum(weights) - 1.0) > 0.0001:
        warnings.append(f"加权日销权重合计为{sum(weights):.4f}，建议等于1")

    trend_groups = _require_mapping(params.get("trend_groups"), "trend_groups")
    for key in ("growth", "decline", "stable", "skip"):
        values = trend_groups.get(key)
        if not isinstance(values, list) or not all(_clean_text(item) for item in values):
            raise ReplenishmentTemplateError(f"趋势分组必须是非空文本列表: {key}")

    replenishment_days = _require_list(params.get("replenishment_days"), "replenishment_days")
    if not replenishment_days:
        raise ReplenishmentTemplateError("空运补货天数分档不能为空")
    previous_threshold: float | None = None
    has_fallback = False
    for item in replenishment_days:
        if not isinstance(item, dict):
            raise ReplenishmentTemplateError("空运补货天数分档必须是对象")
        threshold = _number(item.get("daily_sales_gt"), default=None)
        if threshold is None:
            has_fallback = True
        elif threshold < 0:
            raise ReplenishmentTemplateError("空运补货天数日销分档不能为负数")
        if previous_threshold is not None and threshold is not None and threshold >= previous_threshold:
            raise ReplenishmentTemplateError("空运补货天数日销分档必须按从高到低排列")
        if threshold is not None:
            previous_threshold = threshold
        for key in TREND_KEYS:
            days = _int_value(item.get(key), default=None)
            if days is None or days < 0:
                raise ReplenishmentTemplateError(f"空运补货天数必须是非负整数: {item.get('label') or threshold}, {key}")
    if not has_fallback:
        raise ReplenishmentTemplateError("空运补货天数必须包含低销量兜底分档")

    shipping = _require_mapping(params.get("shipping"), "shipping")
    urgent_days = _number(shipping.get("air_urgent_sales_days_lte"), default=None)
    air_days = _number(shipping.get("air_sales_days_lte"), default=None)
    if urgent_days is None or urgent_days < 0:
        raise ReplenishmentTemplateError("空运急发阈值必须大于等于0")
    if air_days is None or air_days < 0:
        raise ReplenishmentTemplateError("空运阈值必须大于等于0")
    if urgent_days > air_days:
        raise ReplenishmentTemplateError("空运急发阈值必须小于等于空运阈值")

    sea = _require_mapping(params.get("sea"), "sea")
    _bool_value(sea.get("enabled"), field_name="sea.enabled")
    if "min_daily_sales_inclusive" in sea:
        _bool_value(sea.get("min_daily_sales_inclusive"), field_name="sea.min_daily_sales_inclusive")
    min_daily_sales = _number(sea.get("min_daily_sales"), default=None)
    min_weight_kg = _number(sea.get("min_weight_kg"), default=None)
    if min_daily_sales is None or min_daily_sales < 0:
        raise ReplenishmentTemplateError("海运最低日销必须大于等于0")
    if min_weight_kg is None or min_weight_kg < 0:
        raise ReplenishmentTemplateError("海运最低重量kg必须大于等于0")
    _validate_sea_tiers(_require_list(sea.get("tiers"), "sea.tiers"), name="sea.tiers")
    companion_enabled = sea_companion_air_enabled_from_template(params)
    if companion_enabled:
        companion_tiers = _require_list(sea.get("companion_air_tiers"), "sea.companion_air_tiers")
        if not companion_tiers:
            raise ReplenishmentTemplateError("海运同时空运分档不能为空")
        _validate_sea_tiers(companion_tiers, name="sea.companion_air_tiers")
        min_net_quantity = _number(sea.get("min_net_quantity"), default=None)
        if min_net_quantity is None or min_net_quantity < 0:
            raise ReplenishmentTemplateError("海运建议量最小件数必须大于等于0")
    elif "companion_air_tiers" in sea and sea.get("companion_air_tiers") not in (None, ""):
        _validate_sea_tiers(
            _require_list(sea.get("companion_air_tiers"), "sea.companion_air_tiers"),
            name="sea.companion_air_tiers",
        )
    if "min_net_quantity" in sea and sea.get("min_net_quantity") not in (None, ""):
        min_net_quantity = _number(sea.get("min_net_quantity"), default=None)
        if min_net_quantity is None or min_net_quantity < 0:
            raise ReplenishmentTemplateError("海运建议量最小件数必须大于等于0")

    for index, rule in enumerate(_require_list(params.get("special_rules", []), "special_rules"), start=1):
        if not isinstance(rule, dict):
            raise ReplenishmentTemplateError("特殊MSKU规则必须是对象")
        if not _clean_text(rule.get("rule_name")):
            rule["rule_name"] = f"特殊规则{index}"
        msku_values = rule.get("msku_list")
        if not isinstance(msku_values, list) or not [_clean_text(msku) for msku in msku_values]:
            raise ReplenishmentTemplateError(f"特殊MSKU规则缺少MSKU列表: {rule.get('rule_name')}")
        overrides = rule.get("overrides")
        if not isinstance(overrides, dict):
            raise ReplenishmentTemplateError(f"特殊MSKU规则缺少overrides: {rule.get('rule_name')}")
        merged_params = apply_overrides(params, overrides)
        merged_params["special_rules"] = []
        validate_template(ReplenishmentTemplate(template.name, template.version, template.description, merged_params, template.source))

    return TemplateValidationResult(template=template, warnings=tuple(warnings))


def _validate_sea_tiers(tiers: list[Any], *, name: str) -> None:
    previous_lte: float | None = None
    for item in tiers:
        if not isinstance(item, dict):
            raise ReplenishmentTemplateError(f"{name}分档必须是对象")
        gt = _number(item.get("daily_sales_gt"), default=None)
        lte = _number(item.get("daily_sales_lte"), default=None)
        if gt is not None and gt < 0:
            raise ReplenishmentTemplateError(f"{name}分档下限不能为负数")
        if lte is not None and lte < 0:
            raise ReplenishmentTemplateError(f"{name}分档上限不能为负数")
        if gt is not None and lte is not None and gt >= lte:
            raise ReplenishmentTemplateError(f"{name}分档下限必须小于上限")
        if previous_lte is not None and gt is not None and gt < previous_lte:
            raise ReplenishmentTemplateError(f"{name}分档不能重叠")
        if lte is not None:
            previous_lte = lte
        days_values = item.get("days")
        if not isinstance(days_values, list) or not days_values:
            raise ReplenishmentTemplateError(f"{name}补货天数不能为空")
        if len(days_values) != 1:
            raise ReplenishmentTemplateError(f"{name}补货天数只能配置一个正整数")
        for day in days_values:
            parsed_day = _int_value(day, default=None)
            if parsed_day is None or parsed_day <= 0:
                raise ReplenishmentTemplateError(f"{name}补货天数必须是正整数: {day}")


def _days_for_trend(item: dict[str, Any], trend: str) -> int:
    key = {"增长": "growth", "平稳": "stable", "下降": "decline"}[trend]
    days = _int_value(item.get(key), default=None)
    if days is None:
        raise ReplenishmentTemplateError(f"空运补货天数缺少趋势: {trend}")
    return days


def trend_group_from_template(trend: str, template: ReplenishmentTemplate) -> str | None:
    clean_trend = _clean_text(trend)
    groups = template.params.get("trend_groups", {})
    if clean_trend in set(groups.get("growth", [])):
        return "增长"
    if clean_trend in set(groups.get("decline", [])):
        return "下降"
    if clean_trend in set(groups.get("stable", [])):
        return "平稳"
    if clean_trend in set(groups.get("skip", [])):
        return None
    raise ReplenishmentTemplateError(f"未知销量趋势: {clean_trend}")


def calculate_weighted_daily_sales(
    *,
    sales_7d: float,
    sales_14d: float,
    sales_30d: float,
    params: dict[str, Any],
) -> float:
    weighted = params["weighted_sales"]
    return (
        sales_7d / 7 * float(weighted["7d_weight"])
        + sales_14d / 14 * float(weighted["14d_weight"])
        + sales_30d / 30 * float(weighted["30d_weight"])
    )


def replenishment_days_from_template(weighted_daily_sales: float, mapped_trend: str, params: dict[str, Any]) -> int:
    for item in params["replenishment_days"]:
        threshold = _number(item.get("daily_sales_gt"), default=None)
        if threshold is None or weighted_daily_sales > threshold:
            return _days_for_trend(item, mapped_trend)
    raise ReplenishmentTemplateError("空运补货天数缺少兜底分档")


def sea_day_candidates_from_template(weighted_daily_sales: float, params: dict[str, Any]) -> list[int]:
    if not sea_enabled_from_template(params):
        return []
    if not sea_daily_sales_meets_min_from_template(weighted_daily_sales, params):
        return []
    return _sea_days_from_tiers(weighted_daily_sales, params, params["sea"]["tiers"])


def sea_companion_air_day_candidates_from_template(weighted_daily_sales: float, params: dict[str, Any]) -> list[int]:
    if not sea_companion_air_enabled_from_template(params):
        return []
    if not sea_daily_sales_meets_min_from_template(weighted_daily_sales, params):
        return []
    return _sea_days_from_tiers(weighted_daily_sales, params, params["sea"].get("companion_air_tiers", []))


def _sea_days_from_tiers(weighted_daily_sales: float, params: dict[str, Any], tiers: list[dict[str, Any]]) -> list[int]:
    for item in tiers:
        gt = _number(item.get("daily_sales_gt"), default=None)
        lte = _number(item.get("daily_sales_lte"), default=None)
        if not _sea_tier_lower_bound_matches(weighted_daily_sales, gt, params):
            continue
        if lte is not None and weighted_daily_sales > float(lte) + FLOAT_TOLERANCE:
            continue
        return [int(day) for day in item["days"]]
    return []


def sea_min_daily_sales_inclusive_from_template(params: dict[str, Any]) -> bool:
    sea = _require_mapping(params.get("sea") if isinstance(params, dict) else None, "sea")
    if "min_daily_sales_inclusive" not in sea:
        return False
    return _bool_value(sea.get("min_daily_sales_inclusive"), field_name="sea.min_daily_sales_inclusive")


def sea_companion_air_enabled_from_template(params: dict[str, Any]) -> bool:
    sea = _require_mapping(params.get("sea") if isinstance(params, dict) else None, "sea")
    if "companion_air_enabled" not in sea:
        return False
    return _bool_value(sea.get("companion_air_enabled"), field_name="sea.companion_air_enabled")


def sea_min_net_quantity_from_template(params: dict[str, Any]) -> float:
    sea = _require_mapping(params.get("sea") if isinstance(params, dict) else None, "sea")
    return float(_number(sea.get("min_net_quantity"), default=0) or 0)


def sea_daily_sales_meets_min_from_template(weighted_daily_sales: float, params: dict[str, Any]) -> bool:
    min_daily_sales = float(params["sea"]["min_daily_sales"])
    if sea_min_daily_sales_inclusive_from_template(params):
        return weighted_daily_sales + FLOAT_TOLERANCE >= min_daily_sales
    return weighted_daily_sales > min_daily_sales + FLOAT_TOLERANCE


def _sea_tier_lower_bound_matches(weighted_daily_sales: float, gt: float | None, params: dict[str, Any]) -> bool:
    if gt is None:
        return True
    min_daily_sales = float(params["sea"]["min_daily_sales"])
    if sea_min_daily_sales_inclusive_from_template(params) and abs(float(gt) - min_daily_sales) <= 0.0000001:
        return weighted_daily_sales + FLOAT_TOLERANCE >= float(gt)
    return weighted_daily_sales > float(gt) + FLOAT_TOLERANCE


def sea_enabled_from_template(params: dict[str, Any]) -> bool:
    sea = _require_mapping(params.get("sea") if isinstance(params, dict) else None, "sea")
    return _bool_value(sea.get("enabled"), field_name="sea.enabled")


def apply_overrides(base_params: dict[str, Any], overrides: dict[str, Any]) -> dict[str, Any]:
    result = copy.deepcopy(base_params)
    for section_name, section_values in (overrides or {}).items():
        if not isinstance(section_values, dict):
            continue
        target = result.setdefault(section_name, {})
        if not isinstance(target, dict):
            continue
        for key, value in section_values.items():
            if value is not None and _clean_text(value) != "":
                target[key] = value
    return result


def effective_params_for_msku(template: ReplenishmentTemplate, msku: str) -> tuple[dict[str, Any], str]:
    clean_msku = _clean_text(msku)
    for rule in template.params.get("special_rules", []):
        msku_set = {_clean_text(item) for item in rule.get("msku_list", []) if _clean_text(item)}
        if clean_msku in msku_set:
            return apply_overrides(template.params, rule.get("overrides", {})), _clean_text(rule.get("rule_name")) or "特殊规则"
    return copy.deepcopy(template.params), ""


def _set_standard_dimensions(workbook: Any) -> None:
    from openpyxl.utils import get_column_letter

    for worksheet in workbook.worksheets:
        worksheet.sheet_format.defaultRowHeight = EXCEL_ROW_HEIGHT
        worksheet.freeze_panes = "A2"
        for row_index in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = EXCEL_ROW_HEIGHT
        for column_index in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = EXCEL_COLUMN_WIDTH


def _append_note_block(worksheet: Any, notes: list[str]) -> None:
    if not notes:
        return
    worksheet.append([])
    worksheet.append(["修改说明"])
    for note in notes:
        worksheet.append([note])


def _style_template_workbook(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    editable_fill = PatternFill("solid", fgColor=EDITABLE_FILL_COLOR)
    readonly_fill = PatternFill("solid", fgColor=READONLY_FILL_COLOR)
    header_font = Font(color=WHITE_FONT_COLOR, bold=True)
    center = Alignment(vertical="center")

    def style_header(worksheet: Any, row_index: int = 1) -> None:
        for cell in worksheet[row_index]:
            if _clean_text(cell.value):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

    def style_columns(worksheet: Any, columns: tuple[int, ...], fill: Any, *, start_row: int = 2) -> None:
        for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row):
            for column_index in columns:
                if column_index <= len(row):
                    row[column_index - 1].fill = fill
                    row[column_index - 1].alignment = center

    def style_note_blocks(worksheet: Any) -> None:
        in_note_block = False
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            if _clean_text(row[0].value) == "修改说明":
                in_note_block = True
                row[0].font = Font(bold=True)
            if in_note_block:
                for cell in row[: worksheet.max_column]:
                    cell.fill = readonly_fill
                    cell.alignment = center

    for worksheet in workbook.worksheets:
        style_header(worksheet)

    style_columns(workbook[TEMPLATE_INFO_SHEET], (1,), readonly_fill)
    style_columns(workbook[TEMPLATE_INFO_SHEET], (2,), editable_fill)

    style_columns(workbook[WEIGHTED_SALES_SHEET], (1, 3), readonly_fill)
    style_columns(workbook[WEIGHTED_SALES_SHEET], (2,), editable_fill)

    style_columns(workbook[REPLENISHMENT_DAYS_SHEET], (1,), readonly_fill)
    style_columns(workbook[REPLENISHMENT_DAYS_SHEET], (2, 3, 4, 5), editable_fill)

    style_columns(workbook[SHIPPING_SHEET], (1, 3), readonly_fill)
    style_columns(workbook[SHIPPING_SHEET], (2,), editable_fill)

    style_columns(workbook[SEA_ENTRY_SHEET], (1, 3), readonly_fill)
    style_columns(workbook[SEA_ENTRY_SHEET], (2,), editable_fill)

    style_columns(workbook[SEA_DAYS_SHEET], (1, 2), editable_fill)

    companion = workbook[SEA_COMPANION_AIR_SHEET]
    companion.freeze_panes = "A2"
    style_columns(companion, (1,), readonly_fill, start_row=2)
    style_columns(companion, (2,), editable_fill, start_row=2)
    style_header(companion, row_index=4)
    style_columns(companion, (1, 2), editable_fill, start_row=5)

    style_columns(workbook[SPECIAL_RULES_SHEET], tuple(range(1, workbook[SPECIAL_RULES_SHEET].max_column + 1)), editable_fill)

    column_widths = {
        TEMPLATE_INFO_SHEET: {1: 18, 2: 34},
        WEIGHTED_SALES_SHEET: {1: 22, 2: 14, 3: 34},
        REPLENISHMENT_DAYS_SHEET: {1: 16, 2: 18, 3: 12, 4: 12, 5: 12},
        SHIPPING_SHEET: {1: 24, 2: 14, 3: 38},
        SEA_ENTRY_SHEET: {1: 28, 2: 14, 3: 46},
        SEA_DAYS_SHEET: {1: 18, 2: 18},
        SEA_COMPANION_AIR_SHEET: {1: 28, 2: 18},
        SPECIAL_RULES_SHEET: {1: 18, 2: 28, 3: 14, 4: 14, 5: 14, 6: 16, 7: 14, 8: 16, 9: 18, 10: 24},
    }
    for sheet_name, widths in column_widths.items():
        worksheet = workbook[sheet_name]
        for column_index, width in widths.items():
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

    for worksheet in workbook.worksheets:
        style_note_blocks(worksheet)


def export_template_xlsx(
    template_name: str | None = None,
    *,
    output_dir: str | Path | None = None,
    store_path: str | Path | None = None,
) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入备货算法配置表") from exc

    template = get_template(template_name, store_path=store_path)
    params = template.params
    target_path = _editable_output_dir(output_dir) / f"{_timestamp_text()}-{_safe_file_part(template.name)}_{REPLENISHMENT_TEMPLATE_FILE_SUFFIX}.xlsx"

    workbook = Workbook()
    try:
        info = workbook.active
        info.title = TEMPLATE_INFO_SHEET
        info.append(["字段", "值"])
        info.append(["方案名称", template.name])
        info.append(["版本", template.version])
        info.append(["方案说明", template.description])
        _append_note_block(
            info,
            [
                "1. 方案名称和方案说明用于识别参数方案；导入时也可以通过 --name 指定正式方案名。",
                "2. 版本用于系统记录参数方案替换次数，通常不需要手动修改。",
            ],
        )

        weighted = workbook.create_sheet(WEIGHTED_SALES_SHEET)
        weighted.append(["参数", "值", "说明"])
        weighted.append(["7天销量权重", params["weighted_sales"]["7d_weight"], "7天销量 / 7 的权重"])
        weighted.append(["14天销量权重", params["weighted_sales"]["14d_weight"], "14天销量 / 14 的权重"])
        weighted.append(["30天销量权重", params["weighted_sales"]["30d_weight"], "30天销量 / 30 的权重"])
        _append_note_block(
            weighted,
            [
                "1. 只修改“值”列；参数名和说明不作为算法输入。",
                "2. 三个权重建议合计为 1，用于计算加权日销。",
            ],
        )

        replenishment = workbook.create_sheet(REPLENISHMENT_DAYS_SHEET)
        replenishment.append(["日销层级", "日销范围", "增长", "平稳", "下降"])
        previous_threshold: float | None = None
        for item in params["replenishment_days"]:
            threshold = _number(item.get("daily_sales_gt"), default=None)
            replenishment.append([
                item.get("label", ""),
                _replenishment_range_label(threshold, previous_threshold),
                item.get("growth", ""),
                item.get("stable", ""),
                item.get("decline", ""),
            ])
            if threshold is not None:
                previous_threshold = threshold
        _append_note_block(
            replenishment,
            [
                "1. 日销层级仅用于阅读，不影响计算。",
                "2. 日销范围是权威输入，决定该行命中的日销分档；建议按从高到低填写。",
                "3. 增长/平稳/下降是命中该分档后的理论空运补货天数。",
                "4. 该 sheet 不影响海运补货天数；海运请修改“海运补货天数”sheet。",
                "5. 最后一行建议保留“低销量兜底”。",
            ],
        )

        shipping = workbook.create_sheet(SHIPPING_SHEET)
        shipping.append(["参数", "值", "说明"])
        shipping.append(["空运急发可销售天数<=", params["shipping"]["air_urgent_sales_days_lte"], "小于等于该值时进入空运（急发）"])
        shipping.append(["空运可销售天数<=", params["shipping"]["air_sales_days_lte"], "小于等于该值时进入空运"])
        _append_note_block(
            shipping,
            [
                "1. 只修改“值”列。",
                "2. 可销售天数小于等于急发阈值进入空运（急发）。",
                "3. 可销售天数大于急发阈值且小于等于空运阈值进入普通空运。",
            ],
        )

        sea_entry = workbook.create_sheet(SEA_ENTRY_SHEET)
        sea_entry.append(["参数", "值", "说明"])
        sea_entry.append(["是否启用海运", "是" if sea_enabled_from_template(params) else "否", "否表示超过空运阈值后不计算海运"])
        sea_entry.append(["海运最低日销", params["sea"]["min_daily_sales"], "加权日销不满足该门槛时不建议海运"])
        sea_entry.append([
            "海运最低日销是否包含等于",
            "是" if sea_min_daily_sales_inclusive_from_template(params) else "否",
            "是表示加权日销等于最低日销时也可进入海运判断",
        ])
        sea_entry.append(["海运最低重量kg", params["sea"]["min_weight_kg"], "扣减库存后的海运部分重量大于该值才建议海运"])
        sea_entry.append([
            "海运建议量最小件数",
            params["sea"].get("min_net_quantity", ""),
            "扣减库存后的海运建议量小于该值时不建议海运",
        ])
        _append_note_block(
            sea_entry,
            [
                "1. 只修改“值”列。",
                "2. 是否启用海运支持 是/否、true/false、1/0、启用/关闭。",
                "3. 最低日销、最低重量和最小件数是进入海运的前置门槛。",
            ],
        )

        sea_days = workbook.create_sheet(SEA_DAYS_SHEET)
        sea_days.append(["日销范围", "海运补货天数"])
        for item in params["sea"]["tiers"]:
            sea_days.append([
                _sea_range_label(item, params),
                ",".join(str(day) for day in item.get("days", [])),
            ])
        _append_note_block(
            sea_days,
            [
                "1. 日销范围决定命中哪一档海运补货天数。",
                "2. 海运补货天数是命中海运后的理论目标天数。",
                "3. 海运补货天数只能填写一个正整数，不支持 100,110 这种多值写法。",
                "4. 支持 >20、>=1、(5,20]、[1,5]、≤2、低销量兜底 等范围写法。",
            ],
        )

        companion = workbook.create_sheet(SEA_COMPANION_AIR_SHEET)
        companion.append(["项目", "值"])
        companion.append(["是否启用海运同时备空运", "是" if sea_companion_air_enabled_from_template(params) else "否"])
        companion.append([])
        companion.append(["日销范围", "同时空运天数"])
        for item in params["sea"].get("companion_air_tiers", []):
            companion.append([
                _sea_range_label(item, params),
                ",".join(str(day) for day in item.get("days", [])),
            ])
        _append_note_block(
            companion,
            [
                "1. 开关值支持 是/否、true/false、1/0、启用/关闭。",
                "2. 开启后，命中海运的 MSKU 会同时预留一段空运建议量。",
                "3. 日销范围和同时空运天数共同决定同时空运部分的理论天数。",
            ],
        )

        special = workbook.create_sheet(SPECIAL_RULES_SHEET)
        special.append([
            "规则名称",
            "MSKU列表",
            "7天权重",
            "14天权重",
            "30天权重",
            "空运急发阈值",
            "空运阈值",
            "海运最低日销",
            "海运最低重量kg",
            "备注",
        ])
        for rule in params.get("special_rules", []):
            overrides = rule.get("overrides", {})
            weighted_overrides = overrides.get("weighted_sales", {})
            shipping_overrides = overrides.get("shipping", {})
            sea_overrides = overrides.get("sea", {})
            special.append([
                rule.get("rule_name", ""),
                ",".join(str(item) for item in rule.get("msku_list", [])),
                weighted_overrides.get("7d_weight", ""),
                weighted_overrides.get("14d_weight", ""),
                weighted_overrides.get("30d_weight", ""),
                shipping_overrides.get("air_urgent_sales_days_lte", ""),
                shipping_overrides.get("air_sales_days_lte", ""),
                sea_overrides.get("min_daily_sales", ""),
                sea_overrides.get("min_weight_kg", ""),
                rule.get("remark", ""),
            ])
        _append_note_block(
            special,
            [
                "1. 特殊MSKU规则用于给指定 MSKU 覆盖部分算法参数。",
                "2. MSKU列表可用逗号、空格、顿号、分号分隔。",
                "3. 留空的覆盖字段不会覆盖默认参数。",
            ],
        )

        _set_standard_dimensions(workbook)
        _style_template_workbook(workbook)
        workbook.save(target_path)
    finally:
        workbook.close()
    return target_path


def _sheet_records(xlsx_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取备货算法配置表") from exc

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ReplenishmentTemplateError(f"备货算法配置表缺少sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        records: list[dict[str, Any]] = []
        for values in rows:
            row_values = list(values or [])
            if _clean_text(row_values[0] if row_values else "") == "修改说明":
                break
            if any(_clean_text(value) for value in row_values):
                records.append(dict(zip(headers, row_values, strict=False)))
        return records
    finally:
        workbook.close()


def _sea_companion_air_sheet_data(xlsx_path: Path) -> tuple[Any, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取备货算法配置表") from exc

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SEA_COMPANION_AIR_SHEET not in workbook.sheetnames:
            raise ReplenishmentTemplateError(f"备货算法配置表缺少sheet: {SEA_COMPANION_AIR_SHEET}")
        worksheet = workbook[SEA_COMPANION_AIR_SHEET]
        enabled_value = worksheet.cell(row=2, column=2).value
        records: list[dict[str, Any]] = []
        for row_index in range(5, worksheet.max_row + 1):
            range_value = worksheet.cell(row=row_index, column=1).value
            days_value = worksheet.cell(row=row_index, column=2).value
            if _clean_text(range_value) == "修改说明":
                break
            if not _clean_text(range_value) and not _clean_text(days_value):
                continue
            records.append(
                {
                    "日销范围": range_value,
                    "同时空运天数": days_value,
                    "_row_number": row_index,
                }
            )
        return enabled_value, records
    finally:
        workbook.close()


def _ensure_new_template_xlsx(xlsx_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取备货算法配置表") from exc

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet_names = set(workbook.sheetnames)
        missing = [sheet_name for sheet_name in SHEET_NAMES if sheet_name not in sheet_names]
        if missing:
            if any(sheet_name in sheet_names for sheet_name in OLD_TEMPLATE_SHEET_NAMES):
                raise ReplenishmentTemplateError(OLD_TEMPLATE_ERROR)
            raise ReplenishmentTemplateError(f"备货算法配置表缺少sheet: {missing[0]}")

        expected_headers = {
            TEMPLATE_INFO_SHEET: ["字段", "值"],
            REPLENISHMENT_DAYS_SHEET: ["日销层级", "日销范围", "增长", "平稳", "下降"],
            SEA_DAYS_SHEET: ["日销范围", "海运补货天数"],
            SEA_COMPANION_AIR_SHEET: ["项目", "值"],
        }
        for sheet_name, headers in expected_headers.items():
            worksheet = workbook[sheet_name]
            actual = [_clean_text(worksheet.cell(row=1, column=index).value) for index in range(1, len(headers) + 1)]
            if actual != headers:
                raise ReplenishmentTemplateError(OLD_TEMPLATE_ERROR)

        info = workbook[TEMPLATE_INFO_SHEET]
        info_labels = [
            _clean_text(info.cell(row=2, column=1).value),
            _clean_text(info.cell(row=3, column=1).value),
            _clean_text(info.cell(row=4, column=1).value),
        ]
        if info_labels != ["方案名称", "版本", "方案说明"]:
            raise ReplenishmentTemplateError(OLD_TEMPLATE_ERROR)

        companion = workbook[SEA_COMPANION_AIR_SHEET]
        companion_tier_headers = [
            _clean_text(companion.cell(row=4, column=1).value),
            _clean_text(companion.cell(row=4, column=2).value),
        ]
        if companion_tier_headers != ["日销范围", "同时空运天数"]:
            raise ReplenishmentTemplateError(OLD_TEMPLATE_ERROR)
    finally:
        workbook.close()


def _first_record_value(records: list[dict[str, Any]], label: str) -> Any:
    for record in records:
        if _clean_text(record.get("字段")) == label:
            return record.get("值")
    return ""


def _param_value(records: list[dict[str, Any]], label: str) -> Any:
    for record in records:
        if _clean_text(record.get("参数")) == label:
            return record.get("值")
    return ""


def _parse_single_day_list(value: Any, *, sheet_name: str, row_number: int, field_name: str) -> list[int]:
    text = _clean_text(value)
    if not text:
        raise ReplenishmentTemplateError(f"{sheet_name}第{row_number}行{field_name}不能为空")
    days: list[int] = []
    for part in re.split(r"[,，;；、\s]+", text):
        clean = _clean_text(part)
        if not clean:
            continue
        day = _int_value(clean, default=None)
        if day is None:
            raise ReplenishmentTemplateError(f"{sheet_name}第{row_number}行{field_name}无效: {clean}")
        days.append(day)
    if len(days) != 1:
        raise ReplenishmentTemplateError(f"{sheet_name}第{row_number}行{field_name}只能填写一个正整数: {value}")
    if days[0] <= 0:
        raise ReplenishmentTemplateError(f"{sheet_name}第{row_number}行{field_name}必须是正整数: {value}")
    return days


def _ensure_supported_sea_range(
    range_spec: DailySalesRange,
    *,
    params: dict[str, Any],
    sheet_name: str,
    row_number: int,
    value: Any,
) -> None:
    if not range_spec.lower_inclusive or range_spec.lower is None:
        return
    min_daily_sales = _number(params.get("sea", {}).get("min_daily_sales"), default=None)
    if (
        min_daily_sales is not None
        and sea_min_daily_sales_inclusive_from_template(params)
        and abs(float(range_spec.lower) - float(min_daily_sales)) <= 0.0000001
    ):
        return
    raise ReplenishmentTemplateError(
        f"{sheet_name}第{row_number}行日销范围暂不支持这个包含等于下限: {value}；"
        "海运分档只有最低日销这一档可以写成包含等于"
    )


def _split_msku_list(value: Any) -> list[str]:
    return [item for item in (_clean_text(part) for part in MSKU_SPLIT_RE.split(_clean_text(value))) if item]


def _optional_override(value: Any) -> float | None:
    if _clean_text(value) == "":
        return None
    number = _number(value, default=None)
    if number is None:
        raise ReplenishmentTemplateError(f"参数覆盖值必须是数字: {value}")
    return number


def _special_rule_overrides(record: dict[str, Any]) -> dict[str, Any]:
    overrides: dict[str, Any] = {}
    weighted: dict[str, Any] = {}
    for label, key in (("7天权重", "7d_weight"), ("14天权重", "14d_weight"), ("30天权重", "30d_weight")):
        value = _optional_override(record.get(label))
        if value is not None:
            weighted[key] = value
    if weighted:
        overrides["weighted_sales"] = weighted

    shipping: dict[str, Any] = {}
    for label, key in (("空运急发阈值", "air_urgent_sales_days_lte"), ("空运阈值", "air_sales_days_lte")):
        value = _optional_override(record.get(label))
        if value is not None:
            shipping[key] = value
    if shipping:
        overrides["shipping"] = shipping

    sea: dict[str, Any] = {}
    for label, key in (("海运最低日销", "min_daily_sales"), ("海运最低重量kg", "min_weight_kg")):
        value = _optional_override(record.get(label))
        if value is not None:
            sea[key] = value
    if sea:
        overrides["sea"] = sea
    return overrides


def parse_template_xlsx(xlsx_path: str | Path, *, import_name: str | None = None) -> ReplenishmentTemplate:
    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"备货算法配置表不存在: {source_path}")

    _ensure_new_template_xlsx(source_path)
    info_records = _sheet_records(source_path, TEMPLATE_INFO_SHEET)
    weighted_records = _sheet_records(source_path, WEIGHTED_SALES_SHEET)
    replenishment_records = _sheet_records(source_path, REPLENISHMENT_DAYS_SHEET)
    shipping_records = _sheet_records(source_path, SHIPPING_SHEET)
    sea_entry_records = _sheet_records(source_path, SEA_ENTRY_SHEET)
    sea_day_records = _sheet_records(source_path, SEA_DAYS_SHEET)
    companion_air_enabled_value, sea_companion_records = _sea_companion_air_sheet_data(source_path)
    special_records = _sheet_records(source_path, SPECIAL_RULES_SHEET)

    name = _clean_text(import_name) or _clean_text(_first_record_value(info_records, "方案名称"))
    version = _int_value(_first_record_value(info_records, "版本"), default=DEFAULT_TEMPLATE_VERSION)
    description = _clean_text(_first_record_value(info_records, "方案说明"))
    params: dict[str, Any] = {
        "weighted_sales": {
            "7d_weight": _number(_param_value(weighted_records, "7天销量权重"), default=None),
            "14d_weight": _number(_param_value(weighted_records, "14天销量权重"), default=None),
            "30d_weight": _number(_param_value(weighted_records, "30天销量权重"), default=None),
        },
        "trend_groups": copy.deepcopy(load_default_template().params["trend_groups"]),
        "replenishment_days": [],
        "shipping": {
            "air_urgent_sales_days_lte": _number(_param_value(shipping_records, "空运急发可销售天数<="), default=None),
            "air_sales_days_lte": _number(_param_value(shipping_records, "空运可销售天数<="), default=None),
        },
        "sea": {
            "enabled": _bool_value(_param_value(sea_entry_records, "是否启用海运"), field_name="是否启用海运"),
            "min_daily_sales": _number(_param_value(sea_entry_records, "海运最低日销"), default=None),
            "min_daily_sales_inclusive": _bool_value(
                _param_value(sea_entry_records, "海运最低日销是否包含等于"),
                field_name="海运最低日销是否包含等于",
            ),
            "min_weight_kg": _number(_param_value(sea_entry_records, "海运最低重量kg"), default=None),
            "companion_air_enabled": _bool_value(companion_air_enabled_value, field_name="是否启用海运同时备空运"),
            "companion_air_tiers": [],
            "min_net_quantity": _number(_param_value(sea_entry_records, "海运建议量最小件数"), default=0),
            "tiers": [],
        },
        "special_rules": [],
    }

    for row_number, record in enumerate(replenishment_records, start=2):
        range_spec = _parse_daily_sales_range(
            record.get("日销范围"),
            sheet_name=REPLENISHMENT_DAYS_SHEET,
            row_number=row_number,
        )
        threshold = None if range_spec.is_fallback else range_spec.lower
        params["replenishment_days"].append(
            {
                "label": _clean_text(record.get("日销层级")),
                "daily_sales_gt": threshold,
                "growth": _int_value(record.get("增长"), default=None),
                "stable": _int_value(record.get("平稳"), default=None),
                "decline": _int_value(record.get("下降"), default=None),
            }
        )

    for row_number, record in enumerate(sea_day_records, start=2):
        if not _clean_text(record.get("海运补货天数")) and not _clean_text(record.get("日销范围")):
            continue
        range_spec = _parse_daily_sales_range(
            record.get("日销范围"),
            sheet_name=SEA_DAYS_SHEET,
            row_number=row_number,
        )
        _ensure_supported_sea_range(
            range_spec,
            params=params,
            sheet_name=SEA_DAYS_SHEET,
            row_number=row_number,
            value=record.get("日销范围"),
        )
        params["sea"]["tiers"].append(
            {
                "daily_sales_gt": range_spec.lower,
                "daily_sales_lte": range_spec.upper,
                "days": _parse_single_day_list(
                    record.get("海运补货天数"),
                    sheet_name=SEA_DAYS_SHEET,
                    row_number=row_number,
                    field_name="海运补货天数",
                ),
            }
        )

    for record in sea_companion_records:
        row_number = int(record.get("_row_number") or 0)
        range_spec = _parse_daily_sales_range(
            record.get("日销范围"),
            sheet_name=SEA_COMPANION_AIR_SHEET,
            row_number=row_number,
        )
        _ensure_supported_sea_range(
            range_spec,
            params=params,
            sheet_name=SEA_COMPANION_AIR_SHEET,
            row_number=row_number,
            value=record.get("日销范围"),
        )
        params["sea"]["companion_air_tiers"].append(
            {
                "daily_sales_gt": range_spec.lower,
                "daily_sales_lte": range_spec.upper,
                "days": _parse_single_day_list(
                    record.get("同时空运天数"),
                    sheet_name=SEA_COMPANION_AIR_SHEET,
                    row_number=row_number,
                    field_name="同时空运天数",
                ),
            }
        )

    for index, record in enumerate(special_records, start=1):
        msku_list = _split_msku_list(record.get("MSKU列表"))
        if not msku_list and not _clean_text(record.get("规则名称")):
            continue
        params["special_rules"].append(
            {
                "rule_name": _clean_text(record.get("规则名称")) or f"特殊规则{index}",
                "msku_list": msku_list,
                "overrides": _special_rule_overrides(record),
                "remark": _clean_text(record.get("备注")),
            }
        )

    template = ReplenishmentTemplate(
        name=name,
        version=version or DEFAULT_TEMPLATE_VERSION,
        description=description,
        params=params,
        source=SOURCE_CUSTOM,
    )
    return validate_template(template).template


def validate_template_xlsx(xlsx_path: str | Path, *, import_name: str | None = None) -> TemplateValidationResult:
    template = parse_template_xlsx(xlsx_path, import_name=import_name)
    return validate_template(template)


def import_template_xlsx(
    xlsx_path: str | Path,
    *,
    name: str | None = None,
    store_path: str | Path | None = None,
) -> TemplateValidationResult:
    existing = list_templates(store_path=store_path)
    existing_names = {template.name for template in existing}
    import_name = _clean_text(name)
    if not import_name:
        import_name = _next_custom_template_name(existing_names)
    if is_builtin_template_name(import_name):
        raise ReplenishmentTemplateError(f"系统参数方案不允许导入覆盖: {import_name}")
    if import_name in existing_names:
        raise ReplenishmentTemplateError(f"参数方案名已存在: {import_name}")

    template = parse_template_xlsx(xlsx_path, import_name=import_name)
    template = ReplenishmentTemplate(
        name=import_name,
        version=max(1, int(template.version or 1)),
        description=template.description,
        params=template.params,
        source=SOURCE_CUSTOM,
    )
    result = validate_template(template)

    custom_templates = [item.to_store_payload() for item in load_custom_templates(store_path=store_path)]
    custom_templates.append(result.template.to_store_payload())
    _write_json(_custom_store_path(store_path), {"templates": custom_templates})
    return result


def replace_template_xlsx(
    xlsx_path: str | Path,
    *,
    template_name: str,
    store_path: str | Path | None = None,
) -> tuple[TemplateValidationResult, int]:
    clean_name = _clean_text(template_name)
    if not clean_name:
        raise ReplenishmentTemplateError("参数方案名不能为空")
    if is_builtin_template_name(clean_name):
        raise ReplenishmentTemplateError(f"系统参数方案不允许替换: {clean_name}")

    custom_templates = load_custom_templates(store_path=store_path)
    index = next((idx for idx, item in enumerate(custom_templates) if item.name == clean_name), None)
    if index is None:
        raise ReplenishmentTemplateError(f"只能替换已存在的自定义参数方案: {clean_name}")

    existing_template = custom_templates[index]
    parsed_template = parse_template_xlsx(xlsx_path, import_name=clean_name)
    replacement = ReplenishmentTemplate(
        name=clean_name,
        version=existing_template.version + 1,
        description=parsed_template.description,
        params=parsed_template.params,
        source=SOURCE_CUSTOM,
    )
    result = validate_template(replacement)

    custom_templates[index] = result.template
    _write_json(_custom_store_path(store_path), {"templates": [item.to_store_payload() for item in custom_templates]})
    return result, existing_template.version


def rename_template(
    template_name: str,
    *,
    new_name: str,
    store_path: str | Path | None = None,
) -> ReplenishmentTemplate:
    clean_old_name = _clean_text(template_name)
    clean_new_name = _clean_text(new_name)
    if not clean_old_name:
        raise ReplenishmentTemplateError("参数方案名不能为空")
    if not clean_new_name:
        raise ReplenishmentTemplateError("新参数方案名不能为空")
    if is_builtin_template_name(clean_old_name):
        raise ReplenishmentTemplateError(f"系统参数方案不允许重命名: {clean_old_name}")
    if is_builtin_template_name(clean_new_name):
        raise ReplenishmentTemplateError(f"新参数方案名不能是系统参数方案: {clean_new_name}")

    custom_templates = load_custom_templates(store_path=store_path)
    if any(template.name == clean_new_name for template in custom_templates):
        raise ReplenishmentTemplateError(f"参数方案名已存在: {clean_new_name}")

    index = next((idx for idx, item in enumerate(custom_templates) if item.name == clean_old_name), None)
    if index is None:
        raise ReplenishmentTemplateError(f"只能重命名已存在的自定义参数方案: {clean_old_name}")

    old_template = custom_templates[index]
    renamed = ReplenishmentTemplate(
        name=clean_new_name,
        version=old_template.version,
        description=old_template.description,
        params=old_template.params,
        source=SOURCE_CUSTOM,
    )
    validate_template(renamed)
    custom_templates[index] = renamed
    _write_json(_custom_store_path(store_path), {"templates": [item.to_store_payload() for item in custom_templates]})
    return renamed


def templates_payload(*, store_path: str | Path | None = None) -> dict[str, Any]:
    templates = list_templates(store_path=store_path)
    return {
        "success": True,
        "template_count": len(templates),
        "templates": [
            {
                "name": template.name,
                "version": template.version,
                "description": template.description,
                "source_type": template.source,
            }
            for template in templates
        ],
        "source": SOURCE,
    }


__all__ = [
    "DEFAULT_TEMPLATE_NAME",
    "DE_GROUP_1_TEMPLATE_NAME",
    "REPLENISHMENT_TEMPLATE_FILE_SUFFIX",
    "ReplenishmentTemplate",
    "ReplenishmentTemplateError",
    "TemplateValidationResult",
    "UK_GROUP_1_TEMPLATE_NAME",
    "US_GROUP_1_TEMPLATE_NAME",
    "US_LIN_MEIQI_GROUP_2_TEMPLATE_NAME",
    "apply_overrides",
    "calculate_weighted_daily_sales",
    "effective_params_for_msku",
    "export_template_xlsx",
    "get_template",
    "import_template_xlsx",
    "is_builtin_template_name",
    "list_parameter_groups",
    "list_templates",
    "load_builtin_templates",
    "parse_template_xlsx",
    "replenishment_days_from_template",
    "rename_template",
    "replace_template_xlsx",
    "sea_companion_air_day_candidates_from_template",
    "sea_companion_air_enabled_from_template",
    "sea_daily_sales_meets_min_from_template",
    "sea_day_candidates_from_template",
    "sea_enabled_from_template",
    "sea_min_daily_sales_inclusive_from_template",
    "sea_min_net_quantity_from_template",
    "templates_payload",
    "trend_group_from_template",
    "validate_template",
    "validate_template_xlsx",
]
