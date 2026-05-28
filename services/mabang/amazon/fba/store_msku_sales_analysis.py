from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from shared.config import config

DEFAULT_INPUT_DIR = Path("artifacts") / "mabang_store_msku"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "mabang_store_msku_analysis"
SOURCE = "mabang_store_msku_sales_analysis"
EXCEL_ROW_HEIGHT = 15
EXCEL_COLUMN_WIDTH = 15
SOURCE_FILE_RE = re.compile(r"^(?P<source_time>\d{12})-(?P<store>.+)_msku_data\.xlsx$", re.IGNORECASE)
PRODUCT_LINK_COLUMN = "商品链接"
URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)
REQUIRED_COLUMNS = ("父ASIN", "ASIN", "MSKU", PRODUCT_LINK_COLUMN, "7天销量", "14天销量", "30天销量", "90天销量")
SALES_COLUMNS = ("7天销量", "14天销量", "30天销量", "90天销量")
METRIC_COLUMNS = ("加权日销", "销量趋势速率", "销量趋势")
LINK_TOP_SHEET = "链接销量前10"
LINK_OTHER_SHEET = "其他链接"
ASIN_TOP_SHEET = "ASIN销量前50"
ASIN_OTHER_SHEET = "其他ASIN"
DETAIL_SHEET = "MSKU明细"
REPORT_SHEETS = (LINK_TOP_SHEET, LINK_OTHER_SHEET, ASIN_TOP_SHEET, ASIN_OTHER_SHEET, DETAIL_SHEET)


class StoreMskuSalesAnalysisError(ValueError):
    pass


@dataclass(frozen=True)
class SourceMskuFile:
    path: Path
    source_data_time: str
    source_datetime: datetime


@dataclass(frozen=True)
class SalesMetrics:
    weighted_daily_sales: float
    trend_ratio: float | None
    trend: str


@dataclass(frozen=True)
class StoreMskuSalesAnalysisResult:
    store_name: str
    source_xlsx_path: str
    source_data_time: str
    data_is_stale: bool
    link_count: int
    asin_count: int
    msku_count: int
    report_xlsx_path: str
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "source_xlsx_path": self.source_xlsx_path,
            "source_data_time": self.source_data_time,
            "data_is_stale": self.data_is_stale,
            "link_count": self.link_count,
            "asin_count": self.asin_count,
            "msku_count": self.msku_count,
            "report_xlsx_path": self.report_xlsx_path,
            "source": self.source,
        }


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _safe_file_part(value: Any) -> str:
    text = _clean_text(value)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("._-") or "store_msku"


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _configured_path(name: str, default: Path) -> Path:
    configured = str(getattr(config, name, "") or "").strip()
    return Path(configured) if configured else default


def _input_dir(input_dir: str | Path | None = None) -> Path:
    return Path(input_dir) if input_dir is not None else _configured_path("MABANG_STORE_MSKU_OUTPUT_DIR", DEFAULT_INPUT_DIR)


def _output_dir(output_dir: str | Path | None = None) -> Path:
    path = Path(output_dir) if output_dir is not None else _configured_path(
        "MABANG_STORE_MSKU_ANALYSIS_OUTPUT_DIR",
        DEFAULT_OUTPUT_DIR,
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _parse_source_file(path: Path) -> SourceMskuFile | None:
    match = SOURCE_FILE_RE.match(path.name)
    if not match:
        return None
    source_data_time = match.group("source_time")
    try:
        source_datetime = datetime.strptime(source_data_time, "%Y%m%d%H%M")
    except ValueError:
        return None
    return SourceMskuFile(path=path, source_data_time=source_data_time, source_datetime=source_datetime)


def find_latest_store_msku_file(store_name: str, *, input_dir: str | Path | None = None) -> SourceMskuFile:
    clean_store_name = normalize_store_name(store_name)
    directory = _input_dir(input_dir)
    safe_store_name = _safe_file_part(clean_store_name)
    if not directory.is_dir():
        raise StoreMskuSalesAnalysisError(f"未找到本地店铺MSKU数据文件: {clean_store_name}")

    candidates: list[SourceMskuFile] = []
    for path in directory.glob(f"*-{safe_store_name}_msku_data.xlsx"):
        parsed = _parse_source_file(path)
        if parsed is not None:
            candidates.append(parsed)
    if not candidates:
        raise StoreMskuSalesAnalysisError(f"未找到本地店铺MSKU数据文件: {clean_store_name}")
    return max(candidates, key=lambda item: (item.source_datetime, item.path.name))


def _source_is_stale(source_datetime: datetime, *, today: date | None = None) -> bool:
    check_date = today or date.today()
    return source_datetime.date() < check_date


def _number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        number = float(value)
        return 0.0 if math.isnan(number) else number
    text = _clean_text(value).replace(",", "")
    if not text or text.lower() == "nan":
        return 0.0
    try:
        number = float(text)
    except (TypeError, ValueError):
        return 0.0
    return 0.0 if math.isnan(number) else number


def _display_number(value: float) -> float | int:
    rounded = round(float(value), 4)
    if rounded.is_integer():
        return int(rounded)
    return rounded


def compute_sales_metrics(sales_7d: Any, sales_14d: Any, sales_30d: Any) -> SalesMetrics:
    sales_7 = _number(sales_7d)
    sales_14 = _number(sales_14d)
    sales_30 = _number(sales_30d)

    weighted_daily_sales = sales_7 / 7 * 0.6 + sales_14 / 14 * 0.3 + sales_30 / 30 * 0.1
    current_speed = sales_7 / 7
    recent_speed = (sales_14 - sales_7) / 7
    older_speed = (sales_30 - sales_14) / 16
    baseline = recent_speed * 0.6 + older_speed * 0.4

    trend_ratio: float | None
    baseline_is_zero = math.isclose(baseline, 0.0, abs_tol=1e-12)

    if baseline_is_zero:
        trend_ratio = None
    else:
        trend_ratio = current_speed / baseline

    if sales_30 < 10:
        trend = "样本不足"
    elif baseline_is_zero:
        if current_speed > 0:
            trend = "新增出单/恢复出单"
        else:
            trend = "无销量"
    elif trend_ratio is not None and trend_ratio >= 1.5:
        trend = "快速增长"
    elif trend_ratio is not None and trend_ratio >= 1.2:
        trend = "增长"
    elif trend_ratio is not None and trend_ratio <= 0.6:
        trend = "快速下降"
    elif trend_ratio is not None and trend_ratio <= 0.85:
        trend = "下降"
    else:
        trend = "平稳"

    return SalesMetrics(
        weighted_daily_sales=weighted_daily_sales,
        trend_ratio=trend_ratio,
        trend=trend,
    )


def _load_rows(xlsx_path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取店铺MSKU数据Excel") from exc

    source_path = Path(xlsx_path)
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        header_values = next(rows, None)
        headers = [_clean_text(cell) for cell in list(header_values or [])]
        if not any(headers):
            raise StoreMskuSalesAnalysisError("店铺MSKU数据Excel表头为空")
        missing = [column for column in REQUIRED_COLUMNS if column not in headers]
        if missing:
            raise StoreMskuSalesAnalysisError(f"店铺MSKU数据缺少列: {', '.join(missing)}")

        records: list[dict[str, Any]] = []
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            if any(_clean_text(value) for value in row.values()):
                records.append(row)
    except StoreMskuSalesAnalysisError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取店铺MSKU数据Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    return headers, records


def _sales_totals(row: dict[str, Any]) -> dict[str, float]:
    return {column: _number(row.get(column)) for column in SALES_COLUMNS}


def _append_metrics(row: dict[str, Any], metrics: SalesMetrics) -> dict[str, Any]:
    row["加权日销"] = _display_number(metrics.weighted_daily_sales)
    row["销量趋势速率"] = "" if metrics.trend_ratio is None else _display_number(metrics.trend_ratio)
    row["销量趋势"] = metrics.trend
    return row


def _analyze_msku_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    analyzed: list[dict[str, Any]] = []
    for record in records:
        row = dict(record)
        totals = _sales_totals(row)
        for column, value in totals.items():
            row[column] = _display_number(value)
        metrics = compute_sales_metrics(totals["7天销量"], totals["14天销量"], totals["30天销量"])
        analyzed.append(_append_metrics(row, metrics))
    return analyzed


def _group_key(value: Any, fallback: str) -> str:
    return _clean_text(value) or fallback


def _extract_product_url(value: Any) -> str:
    match = URL_RE.search(_clean_text(value))
    if not match:
        return ""
    return match.group(0).rstrip("),，;；。")


def _product_url_prefix(value: Any, asin: Any) -> str:
    url = _extract_product_url(value)
    if not url:
        return ""

    clean_asin = _clean_text(asin)
    if clean_asin and clean_asin in url:
        prefix = url.split(clean_asin, 1)[0]
        if prefix:
            return prefix

    slash_index = url.rfind("/")
    if slash_index < 0:
        return ""
    return url[: slash_index + 1]


def _summarize_links(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for row in records:
        parent_asin = _group_key(row.get("父ASIN"), "未填写父ASIN")
        group = groups.setdefault(
            parent_asin,
            {
                "父ASIN": parent_asin,
                "MSKU数": 0,
                "ASIN": set(),
                "商品链接前缀": "",
                **{column: 0.0 for column in SALES_COLUMNS},
            },
        )
        group["MSKU数"] += 1
        asin = _clean_text(row.get("ASIN"))
        if asin:
            group["ASIN"].add(asin)
        link_prefix = _product_url_prefix(row.get(PRODUCT_LINK_COLUMN), asin)
        if link_prefix and not group["商品链接前缀"]:
            group["商品链接前缀"] = link_prefix
        totals = _sales_totals(row)
        for column, value in totals.items():
            group[column] += value

    summaries: list[dict[str, Any]] = []
    for group in groups.values():
        metrics = compute_sales_metrics(group["7天销量"], group["14天销量"], group["30天销量"])
        product_link = ""
        if group["商品链接前缀"] and group["父ASIN"] != "未填写父ASIN":
            product_link = f"{group['商品链接前缀']}{group['父ASIN']}"
        summary = {
            "父ASIN": group["父ASIN"],
            "MSKU数": group["MSKU数"],
            "ASIN数": len(group["ASIN"]),
            PRODUCT_LINK_COLUMN: product_link,
        }
        summary.update({column: _display_number(group[column]) for column in SALES_COLUMNS})
        summaries.append(_append_metrics(summary, metrics))
    return sorted(summaries, key=lambda item: (-_number(item["30天销量"]), _clean_text(item["父ASIN"])))


def _summarize_asins(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str, str]] = set()
    summaries: list[dict[str, Any]] = []
    for row in records:
        asin = _group_key(row.get("ASIN"), "未填写ASIN")
        parent_asin = _group_key(row.get("父ASIN"), "未填写父ASIN")
        msku = _group_key(row.get("MSKU"), "未填写MSKU")
        key = (asin, parent_asin, msku)
        if key in seen:
            raise StoreMskuSalesAnalysisError(
                f"ASIN表存在重复售卖项: ASIN={asin}, 父ASIN={parent_asin}, MSKU={msku}"
            )
        seen.add(key)

        totals = _sales_totals(row)
        metrics = compute_sales_metrics(totals["7天销量"], totals["14天销量"], totals["30天销量"])
        summary = {
            "ASIN": asin,
            "父ASIN": parent_asin,
            "MSKU": msku,
            PRODUCT_LINK_COLUMN: _clean_text(row.get(PRODUCT_LINK_COLUMN)),
        }
        summary.update({column: _display_number(totals[column]) for column in SALES_COLUMNS})
        summaries.append(_append_metrics(summary, metrics))
    return sorted(
        summaries,
        key=lambda item: (
            -_number(item["30天销量"]),
            _clean_text(item["ASIN"]),
            _clean_text(item["父ASIN"]),
            _clean_text(item["MSKU"]),
        ),
    )


def _write_table(worksheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_format.defaultRowHeight = EXCEL_ROW_HEIGHT
    for row_index in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = EXCEL_ROW_HEIGHT
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = EXCEL_COLUMN_WIDTH


def _write_report(
    *,
    report_path: Path,
    original_headers: list[str],
    msku_rows: list[dict[str, Any]],
    link_rows: list[dict[str, Any]],
    asin_rows: list[dict[str, Any]],
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入销量分析报告") from exc

    workbook = Workbook()
    try:
        sheet_specs = [
            (LINK_TOP_SHEET, ["父ASIN", "MSKU数", "ASIN数", *SALES_COLUMNS, PRODUCT_LINK_COLUMN, *METRIC_COLUMNS], link_rows[:10]),
            (LINK_OTHER_SHEET, ["父ASIN", "MSKU数", "ASIN数", *SALES_COLUMNS, PRODUCT_LINK_COLUMN, *METRIC_COLUMNS], link_rows[10:]),
            (ASIN_TOP_SHEET, ["ASIN", "父ASIN", "MSKU", *SALES_COLUMNS, PRODUCT_LINK_COLUMN, *METRIC_COLUMNS], asin_rows[:50]),
            (ASIN_OTHER_SHEET, ["ASIN", "父ASIN", "MSKU", *SALES_COLUMNS, PRODUCT_LINK_COLUMN, *METRIC_COLUMNS], asin_rows[50:]),
            (DETAIL_SHEET, [*original_headers, *METRIC_COLUMNS], msku_rows),
        ]
        for index, (title, headers, rows) in enumerate(sheet_specs):
            worksheet = workbook.active if index == 0 else workbook.create_sheet()
            worksheet.title = title
            _write_table(worksheet, list(headers), rows)
        workbook.save(report_path)
    finally:
        workbook.close()


def analyze_store_msku_sales(
    store_name: str,
    *,
    input_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    today: date | None = None,
) -> StoreMskuSalesAnalysisResult:
    clean_store_name = normalize_store_name(store_name)
    source = find_latest_store_msku_file(clean_store_name, input_dir=input_dir)
    original_headers, records = _load_rows(source.path)
    msku_rows = _analyze_msku_rows(records)
    link_rows = _summarize_links(records)
    asin_rows = _summarize_asins(records)

    report_dir = _output_dir(output_dir)
    report_path = report_dir / f"{source.source_data_time}-{_safe_file_part(clean_store_name)}_sales_analysis.xlsx"
    _write_report(
        report_path=report_path,
        original_headers=original_headers,
        msku_rows=msku_rows,
        link_rows=link_rows,
        asin_rows=asin_rows,
    )

    return StoreMskuSalesAnalysisResult(
        store_name=clean_store_name,
        source_xlsx_path=str(source.path),
        source_data_time=source.source_data_time,
        data_is_stale=_source_is_stale(source.source_datetime, today=today),
        link_count=len(link_rows),
        asin_count=len(asin_rows),
        msku_count=len(records),
        report_xlsx_path=str(report_path),
    )


__all__ = [
    "ASIN_OTHER_SHEET",
    "ASIN_TOP_SHEET",
    "DETAIL_SHEET",
    "LINK_OTHER_SHEET",
    "LINK_TOP_SHEET",
    "REPORT_SHEETS",
    "PRODUCT_LINK_COLUMN",
    "REQUIRED_COLUMNS",
    "SOURCE",
    "SALES_COLUMNS",
    "StoreMskuSalesAnalysisError",
    "StoreMskuSalesAnalysisResult",
    "compute_sales_metrics",
    "find_latest_store_msku_file",
    "analyze_store_msku_sales",
    "normalize_store_name",
]
