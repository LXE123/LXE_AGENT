from __future__ import annotations

import math
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from shared.infra.net import erp_http_session, external_http_session
from shared.logging import logger
from services.mabang import config as mabang_settings

from ...auth import get_fba_free_token
from ...errors import MabangBusinessError, MabangRequestError
from .batch_delivery import (
    BatchDeliveryApiAuthError,
    DEFAULT_BATCH_DELIVERY_LIST_URL,
    DEFAULT_TASK_PUSH_URL,
    SIMPLE_TASK_CONFIG_ID,
    _configured_text,
    _int_value,
    _read_api_json,
    _request_headers,
    request_download_info,
    wait_for_delivery_task,
)

DEFAULT_SHOP_COUNTRY_URL = "https://api-private.mabangerp.com/fba/api/v1/shop/shopCountry"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "mabang_fba_unlinked_shipments"
DEFAULT_SNAPSHOT_DIR = Path("artifacts") / "mabang_fba_unlinked_shipments_snapshots"
SOURCE = "mabang_fba_unlinked_shipments"
SNAPSHOT_SOURCE = "mabang_fba_unlinked_shipments_snapshot"
UNLINKED_SHIPMENTS_SNAPSHOT_FILE_SUFFIX = "未关联货件快照"
SNAPSHOT_SUMMARY_SHEET = "未关联货件汇总"
SNAPSHOT_DETAIL_SHEET = "未关联货件明细"
SNAPSHOT_SUMMARY_COLUMNS = (
    "店铺",
    "MSKU",
    "未关联数量",
    "明细行数",
    "涉及状态",
    "涉及运输方式",
    "涉及发货单号",
    "source_files",
)
SNAPSHOT_DETAIL_COLUMNS = (
    "店铺",
    "MSKU",
    "未关联数量",
    "状态",
    "发货单号",
    "货件单号",
    "物流方式",
    "物流渠道",
    "创建时间",
    "source_file",
)
SNAPSHOT_REQUIRED_COLUMNS = ("店铺", "MSKU", "未关联数量")


class UnlinkedShipmentError(MabangBusinessError):
    pass


@dataclass(frozen=True)
class ShopOption:
    store_id: int
    name: str
    raw: dict[str, Any]


@dataclass(frozen=True)
class UnlinkedShipmentStatusSpec:
    status_name: str
    params: dict[str, Any]


@dataclass(frozen=True)
class UnlinkedShipmentStatusResult:
    status_name: str
    total: int
    task_id: int | None = None
    file_hash: str = ""
    file_name: str = ""
    raw_file_path: str = ""

    def to_payload(self) -> dict[str, Any]:
        return {
            "status_name": self.status_name,
            "total": self.total,
            "task_id": self.task_id,
            "file_hash": self.file_hash,
            "file_name": self.file_name,
            "raw_file_path": self.raw_file_path,
        }


@dataclass(frozen=True)
class StoreUnlinkedShipmentDownloadResult:
    store_name: str
    store_id: int
    download_time: str
    status_results: list[UnlinkedShipmentStatusResult]
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "store_id": self.store_id,
            "download_time": self.download_time,
            "status_results": [row.to_payload() for row in self.status_results],
            "source": self.source,
        }


@dataclass(frozen=True)
class UnlinkedShipmentSnapshotResult:
    store_name: str
    snapshot_time: str
    snapshot_xlsx_path: str
    raw_file_count: int
    detail_count: int
    msku_count: int
    total_unlinked_quantity: float
    source: str = SNAPSHOT_SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "snapshot_time": self.snapshot_time,
            "snapshot_xlsx_path": self.snapshot_xlsx_path,
            "raw_file_count": self.raw_file_count,
            "detail_count": self.detail_count,
            "msku_count": self.msku_count,
            "total_unlinked_quantity": _display_quantity(self.total_unlinked_quantity),
            "source": self.source,
        }


UNLINKED_SHIPMENT_STATUS_SPECS = (
    UnlinkedShipmentStatusSpec(
        status_name="WMS待配货",
        params={"status": 6, "is_batch_create": 1, "delivery_type": 2},
    ),
    UnlinkedShipmentStatusSpec(
        status_name="WMS待装箱",
        params={"status": 9, "is_batch_create": 1},
    ),
    UnlinkedShipmentStatusSpec(
        status_name="待关联货件",
        params={"status": 10, "is_batch_create": 1},
    ),
)


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _safe_path_part(value: Any, *, fallback: str) -> str:
    text = _clean_text(value)
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip(" ._-") or fallback


def _configured_path(name: str, default: Path) -> Path:
    return mabang_settings.configured_path(name, default)


def _output_dir(output_dir: str | Path | None = None) -> Path:
    path = Path(output_dir) if output_dir is not None else _configured_path(
        "MABANG_FBA_UNLINKED_SHIPMENTS_OUTPUT_DIR",
        DEFAULT_OUTPUT_DIR,
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _snapshot_dir(output_dir: str | Path | None = None) -> Path:
    path = Path(output_dir) if output_dir is not None else _configured_path(
        "MABANG_FBA_UNLINKED_SHIPMENTS_SNAPSHOT_DIR",
        DEFAULT_SNAPSHOT_DIR,
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _today_text(report_date: str | date | None = None) -> str:
    if isinstance(report_date, date):
        return report_date.isoformat()
    text = _clean_text(report_date)
    if text:
        return text
    return date.today().isoformat()


def _timestamp_text(value: str | None = None) -> str:
    text = _clean_text(value)
    if text:
        return text
    return datetime.now().strftime("%Y%m%d%H%M")


def safe_poll_interval_sec(value: float | int | str | None) -> float:
    try:
        number = float(10 if value is None else value)
    except (TypeError, ValueError):
        return 10.0
    if math.isnan(number):
        return 10.0
    return max(10.0, number)


def safe_timeout_sec(value: float | int | str | None) -> float:
    try:
        number = float(180 if value is None else value)
    except (TypeError, ValueError):
        return 180.0
    if math.isnan(number):
        return 180.0
    return max(0.0, number)


def _number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        number = float(value)
        return 0.0 if math.isnan(number) else number
    text = _clean_text(value).replace(",", "")
    if not text or text.lower() == "nan":
        return 0.0
    try:
        number = float(text)
    except (TypeError, ValueError):
        return 0.0
    return 0.0 if math.isnan(number) else number


def _display_quantity(value: float) -> int | float:
    number = float(value or 0)
    if math.isnan(number):
        return 0
    rounded = round(number, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _decode_csv_bytes(raw: bytes, path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnlinkedShipmentError(f"未关联货件CSV无法识别编码: {path}")


def _csv_records(path: Path) -> list[dict[str, Any]]:
    text = _decode_csv_bytes(path.read_bytes(), path)
    return [dict(row) for row in csv.DictReader(text.splitlines())]


def _xlsx_records(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取未关联货件xlsx") from exc

    workbook = None
    records: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
            if not any(headers):
                continue
            for values in rows:
                row = dict(zip(headers, list(values or []), strict=False))
                if any(_clean_text(value) for value in row.values()):
                    records.append(row)
    except UnlinkedShipmentError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取未关联货件xlsx失败: {path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass
    return records


def _table_records(path: str | Path) -> list[dict[str, Any]]:
    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(f"未关联货件文件不存在: {source_path}")
    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        return _csv_records(source_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _xlsx_records(source_path)
    raise UnlinkedShipmentError(f"未关联货件文件只支持csv/xlsx: {source_path}")


def _infer_status_from_path(path: Path) -> str:
    name = path.name
    for spec in UNLINKED_SHIPMENT_STATUS_SPECS:
        if spec.status_name in name:
            return spec.status_name
    return ""


def _raw_detail_rows(raw_file_paths: list[str | Path], *, store_name: str | None = None) -> list[dict[str, Any]]:
    clean_store_name = _clean_text(store_name)
    detail_rows: list[dict[str, Any]] = []
    for raw_file_path in raw_file_paths:
        source_path = Path(raw_file_path)
        inferred_status = _infer_status_from_path(source_path)
        records = _table_records(source_path)
        for record in records:
            if not any(_clean_text(value) for value in record.values()):
                continue
            store = _clean_text(record.get("店铺")) or clean_store_name
            msku = _clean_text(record.get("MSKU"))
            quantity_text = _clean_text(record.get("MSKU发货量"))
            if not store or not msku:
                continue
            if clean_store_name and store != clean_store_name:
                raise UnlinkedShipmentError(f"原生文件店铺不匹配: expected={clean_store_name}, actual={store}, file={source_path}")
            if not quantity_text:
                raise UnlinkedShipmentError(f"原生文件缺少MSKU发货量: MSKU={msku}, file={source_path}")
            quantity = _number(quantity_text)
            if quantity < 0:
                raise UnlinkedShipmentError(f"原生文件MSKU发货量不能为负数: MSKU={msku}, quantity={quantity_text}, file={source_path}")
            if quantity <= 0:
                continue
            detail_rows.append(
                {
                    "店铺": store,
                    "MSKU": msku,
                    "未关联数量": quantity,
                    "状态": _clean_text(record.get("发货单状态")) or inferred_status,
                    "发货单号": _clean_text(record.get("发货单号")),
                    "货件单号": _clean_text(record.get("货件单号")),
                    "物流方式": _clean_text(record.get("物流方式")),
                    "物流渠道": _clean_text(record.get("物流渠道")),
                    "创建时间": _clean_text(record.get("创建时间")),
                    "source_file": str(source_path),
                }
            )
    return detail_rows


def _join_unique(values: list[str]) -> str:
    seen: dict[str, None] = {}
    for value in values:
        text = _clean_text(value)
        if text:
            seen.setdefault(text, None)
    return "、".join(seen)


def summarize_unlinked_shipment_details(detail_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    for row in detail_rows:
        key = (_clean_text(row.get("店铺")), _clean_text(row.get("MSKU")))
        if not key[0] or not key[1]:
            continue
        group = groups.setdefault(
            key,
            {
                "店铺": key[0],
                "MSKU": key[1],
                "未关联数量": 0.0,
                "明细行数": 0,
                "_statuses": [],
                "_routes": [],
                "_delivery_nos": [],
                "_source_files": [],
            },
        )
        group["未关联数量"] += _number(row.get("未关联数量"))
        group["明细行数"] += 1
        group["_statuses"].append(_clean_text(row.get("状态")))
        route = _clean_text(row.get("物流渠道")) or _clean_text(row.get("物流方式"))
        group["_routes"].append(route)
        group["_delivery_nos"].append(_clean_text(row.get("发货单号")))
        group["_source_files"].append(_clean_text(row.get("source_file")))

    summaries: list[dict[str, Any]] = []
    for group in groups.values():
        summaries.append(
            {
                "店铺": group["店铺"],
                "MSKU": group["MSKU"],
                "未关联数量": _display_quantity(group["未关联数量"]),
                "明细行数": group["明细行数"],
                "涉及状态": _join_unique(group["_statuses"]),
                "涉及运输方式": _join_unique(group["_routes"]),
                "涉及发货单号": _join_unique(group["_delivery_nos"]),
                "source_files": _join_unique(group["_source_files"]),
            }
        )
    return sorted(summaries, key=lambda item: (item["店铺"], item["MSKU"]))


def write_unlinked_shipments_snapshot(
    summary_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    target_path: str | Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入未关联货件快照") from exc

    path = Path(target_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        for index, (sheet_name, headers, rows) in enumerate(
            (
                (SNAPSHOT_SUMMARY_SHEET, SNAPSHOT_SUMMARY_COLUMNS, summary_rows),
                (SNAPSHOT_DETAIL_SHEET, SNAPSHOT_DETAIL_COLUMNS, detail_rows),
            )
        ):
            worksheet = workbook.active if index == 0 else workbook.create_sheet()
            worksheet.title = sheet_name
            worksheet.append(list(headers))
            for row in rows:
                worksheet.append([row.get(header, "") for header in headers])
            worksheet.freeze_panes = "A2"
            if rows:
                worksheet.auto_filter.ref = worksheet.dimensions
        workbook.save(path)
    finally:
        workbook.close()
    return path


def build_store_unlinked_shipments_snapshot(
    raw_file_paths: list[str | Path] | tuple[str | Path, ...],
    *,
    store_name: str | None = None,
    output_dir: str | Path | None = None,
    snapshot_time: str | None = None,
) -> UnlinkedShipmentSnapshotResult:
    paths = [Path(path) for path in raw_file_paths]
    if not paths:
        raise ValueError("raw_file_paths 不能为空")
    detail_rows = _raw_detail_rows(paths, store_name=store_name)
    summary_rows = summarize_unlinked_shipment_details(detail_rows)
    stores = sorted({_clean_text(row.get("店铺")) for row in summary_rows if _clean_text(row.get("店铺"))})
    clean_store_name = _clean_text(store_name) or (stores[0] if len(stores) == 1 else "")
    if not clean_store_name:
        raise UnlinkedShipmentError(f"未关联货件快照必须限定单个店铺: stores={', '.join(stores) or '无'}")
    if any(store != clean_store_name for store in stores):
        raise UnlinkedShipmentError(f"未关联货件快照包含多个店铺: expected={clean_store_name}, stores={', '.join(stores)}")

    timestamp = _timestamp_text(snapshot_time)
    target_path = _snapshot_dir(output_dir) / f"{timestamp}-{_safe_path_part(clean_store_name, fallback='store')}_{UNLINKED_SHIPMENTS_SNAPSHOT_FILE_SUFFIX}.xlsx"
    write_unlinked_shipments_snapshot(summary_rows, detail_rows, target_path)
    return UnlinkedShipmentSnapshotResult(
        store_name=clean_store_name,
        snapshot_time=timestamp,
        snapshot_xlsx_path=str(target_path),
        raw_file_count=len(paths),
        detail_count=len(detail_rows),
        msku_count=len(summary_rows),
        total_unlinked_quantity=sum(_number(row.get("未关联数量")) for row in summary_rows),
    )


def _snapshot_records(path: str | Path) -> list[dict[str, Any]]:
    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(f"未关联货件快照不存在: {source_path}")
    if source_path.suffix.lower() == ".csv":
        records = _csv_records(source_path)
    else:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法读取未关联货件快照") from exc

        workbook = None
        try:
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            sheet_name = SNAPSHOT_SUMMARY_SHEET if SNAPSHOT_SUMMARY_SHEET in workbook.sheetnames else workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
            records = []
            for values in rows:
                row = dict(zip(headers, list(values or []), strict=False))
                if any(_clean_text(value) for value in row.values()):
                    records.append(row)
        except Exception as exc:
            raise RuntimeError(f"读取未关联货件快照失败: {source_path}, error={exc}") from exc
        finally:
            try:
                if workbook is not None:
                    workbook.close()
            except Exception:
                pass

    headers = set(records[0].keys()) if records else set(SNAPSHOT_REQUIRED_COLUMNS)
    missing = [column for column in SNAPSHOT_REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise UnlinkedShipmentError(f"未关联货件快照缺少列: {', '.join(missing)}, path={source_path}")
    return records


def load_unlinked_shipment_quantities(path: str | Path, *, store_name: str | None = None) -> dict[str, float]:
    clean_store_name = _clean_text(store_name)
    records = _snapshot_records(path)
    stores = {_clean_text(row.get("店铺")) for row in records if _clean_text(row.get("店铺"))}
    if clean_store_name and stores and clean_store_name not in stores:
        raise UnlinkedShipmentError(f"未关联货件快照中未找到店铺: {clean_store_name}")

    quantities: dict[str, float] = {}
    for record in records:
        store = _clean_text(record.get("店铺"))
        if clean_store_name and store != clean_store_name:
            continue
        msku = _clean_text(record.get("MSKU"))
        if not msku:
            continue
        quantity = _number(record.get("未关联数量"))
        if quantity < 0:
            raise UnlinkedShipmentError(f"未关联货件快照数量不能为负数: MSKU={msku}")
        quantities[msku] = quantities.get(msku, 0.0) + quantity
    return quantities


def _status_payload(
    spec: UnlinkedShipmentStatusSpec,
    store_id: int,
    *,
    page: int,
    pre_page: int,
    include_export_fields: bool = False,
) -> dict[str, Any]:
    payload = {
        **spec.params,
        "store": [int(store_id)],
        "page": int(page),
        "prePage": int(pre_page),
    }
    if include_export_fields:
        payload.update(
            {
                "ids": [],
                "export_type": "1",
                "currency_type": "1",
                "entry_type": "",
            }
        )
    return payload


def _shop_options(payload: dict[str, Any]) -> list[ShopOption]:
    data = payload.get("data")
    shops = data.get("shop") if isinstance(data, dict) else None
    if not isinstance(shops, list):
        raise UnlinkedShipmentError("店铺列表数据格式异常")
    result: list[ShopOption] = []
    for item in shops:
        if not isinstance(item, dict):
            continue
        store_id = _int_value(item.get("id"))
        name = _clean_text(item.get("name"))
        if store_id is None or store_id <= 0 or not name:
            continue
        result.append(ShopOption(store_id=store_id, name=name, raw=dict(item)))
    return result


def _similar_shop_names(store_name: str, shops: list[ShopOption], *, limit: int = 5) -> list[str]:
    import difflib

    target = store_name.casefold()
    names = [shop.name for shop in shops]
    candidates: list[str] = []
    for name in names:
        folded = name.casefold()
        if target and (target in folded or folded in target):
            candidates.append(name)
    for name in difflib.get_close_matches(store_name, names, n=limit, cutoff=0.45):
        if name not in candidates:
            candidates.append(name)
    return candidates[:limit]


def pick_shop_option(store_name: str, shops: list[ShopOption]) -> ShopOption:
    clean_store_name = normalize_store_name(store_name)
    matches = [shop for shop in shops if shop.name == clean_store_name]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        ids = ", ".join(str(shop.store_id) for shop in matches)
        raise UnlinkedShipmentError(f"店铺名匹配到多个店铺: {clean_store_name}, ids={ids}")

    candidates = _similar_shop_names(clean_store_name, shops)
    suffix = f"。是否指: {', '.join(candidates)}?" if candidates else ""
    raise UnlinkedShipmentError(f"未找到店铺: {clean_store_name}{suffix}")


async def fetch_shop_options(*, token: str | None = None) -> list[ShopOption]:
    active_token = _clean_text(token) or await get_fba_free_token()
    api_url = _configured_text("FBA_SHOP_COUNTRY_API_URL", DEFAULT_SHOP_COUNTRY_URL)
    async with erp_http_session.get(
        api_url,
        params={"warehouse": "1"},
        headers=_request_headers(active_token, json_content=False),
    ) as resp:
        payload = await _read_api_json(resp, action="查询FBA店铺列表")
    return _shop_options(payload)


async def resolve_shop_option(store_name: str, *, token: str | None = None) -> ShopOption:
    return pick_shop_option(store_name, await fetch_shop_options(token=token))


def _list_total(payload: dict[str, Any]) -> int:
    data = payload.get("data")
    if not isinstance(data, dict):
        raise UnlinkedShipmentError("未关联货件列表数据格式异常")
    total = _int_value(data.get("total"))
    if total is None or total < 0:
        raise UnlinkedShipmentError("未关联货件列表缺少有效 total")
    return total


async def fetch_status_total(
    spec: UnlinkedShipmentStatusSpec,
    store_id: int,
    *,
    token: str | None = None,
) -> int:
    active_token = _clean_text(token) or await get_fba_free_token()
    api_url = _configured_text("FBA_DELIVERY_LIST_API_URL", DEFAULT_BATCH_DELIVERY_LIST_URL)
    payload = _status_payload(spec, store_id, page=1, pre_page=1)
    async with erp_http_session.post(api_url, json=payload, headers=_request_headers(active_token)) as resp:
        response = await _read_api_json(resp, action=f"查询{spec.status_name}发货单")
    return _list_total(response)


async def create_unlinked_export_task(
    spec: UnlinkedShipmentStatusSpec,
    store_id: int,
    *,
    token: str | None = None,
    report_date: str | date | None = None,
) -> int:
    active_token = _clean_text(token) or await get_fba_free_token()
    date_text = _today_text(report_date)
    api_url = _configured_text("FBA_DELIVERY_TASK_PUSH_URL", DEFAULT_TASK_PUSH_URL)
    payload = {
        "reportEndDate": date_text,
        "reportStartDate": date_text,
        "simpleTaskConfigId": SIMPLE_TASK_CONFIG_ID,
        "reportParams": _status_payload(spec, store_id, page=1, pre_page=20, include_export_fields=True),
    }
    async with erp_http_session.post(api_url, json=payload, headers=_request_headers(active_token)) as resp:
        response = await _read_api_json(resp, action=f"创建{spec.status_name}导出任务")

    data = response.get("data")
    task_id = _int_value(data.get("taskId") if isinstance(data, dict) else None)
    if task_id is None or task_id <= 0:
        raise UnlinkedShipmentError(f"创建{spec.status_name}导出任务返回缺少 taskId")
    return task_id


def _file_suffix(file_name: str) -> str:
    suffix = Path(_clean_text(file_name)).suffix
    return suffix if suffix else ".bin"


async def download_raw_file_from_url(
    download_url: str,
    *,
    store_name: str,
    status_name: str,
    task_id: int,
    file_name: str,
    output_dir: str | Path | None = None,
    download_time: str | None = None,
) -> Path:
    url = _clean_text(download_url)
    if not url:
        raise ValueError("download_url 不能为空")

    directory = _output_dir(output_dir)
    timestamp = _timestamp_text(download_time)
    store_part = _safe_path_part(store_name, fallback="store")
    status_part = _safe_path_part(status_name, fallback="status")
    target_path = directory / f"{timestamp}-{store_part}-{status_part}-{int(task_id)}{_file_suffix(file_name)}"
    headers = {
        "Accept": "text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*"
    }
    async with external_http_session.get(url, headers=headers) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"下载未关联货件原生文件失败(status={status_code}): {msg}")
        if not body:
            raise UnlinkedShipmentError("下载未关联货件原生文件返回空文件")

    target_path.write_bytes(body)
    return target_path


async def _download_status_file(
    spec: UnlinkedShipmentStatusSpec,
    *,
    store_name: str,
    store_id: int,
    token: str,
    output_dir: str | Path | None,
    timeout_sec: float,
    poll_interval_sec: float,
    report_date: str | date | None,
    download_time: str,
) -> UnlinkedShipmentStatusResult:
    logger.info("[UnlinkedShipments] %s 开始查询", spec.status_name)
    total = await fetch_status_total(spec, store_id, token=token)
    logger.info("[UnlinkedShipments] %s total=%d", spec.status_name, total)
    if total <= 0:
        logger.info("[UnlinkedShipments] %s total=0，跳过导出", spec.status_name)
        return UnlinkedShipmentStatusResult(status_name=spec.status_name, total=0)

    task_id = await create_unlinked_export_task(spec, store_id, token=token, report_date=report_date)
    logger.info("[UnlinkedShipments] %s 创建导出任务: taskId=%d", spec.status_name, task_id)
    task = await wait_for_delivery_task(
        task_id,
        token=token,
        timeout_sec=timeout_sec,
        poll_interval_sec=poll_interval_sec,
        progress_label=f"[UnlinkedShipments] {spec.status_name}",
    )
    logger.info(
        "[UnlinkedShipments] %s 任务完成: taskId=%d fileHash=%s",
        spec.status_name,
        task.task_id,
        task.file_hash,
    )
    download_info = await request_download_info(task.task_id, task.file_hash, token=token)
    logger.info(
        "[UnlinkedShipments] %s 获取下载地址成功: taskId=%d fileName=%s",
        spec.status_name,
        download_info.task_id,
        download_info.file_name,
    )
    raw_path = await download_raw_file_from_url(
        download_info.download_url,
        store_name=store_name,
        status_name=spec.status_name,
        task_id=download_info.task_id,
        file_name=download_info.file_name,
        output_dir=output_dir,
        download_time=download_time,
    )
    logger.info("[UnlinkedShipments] %s 下载完成: %s", spec.status_name, raw_path)
    return UnlinkedShipmentStatusResult(
        status_name=spec.status_name,
        total=total,
        task_id=download_info.task_id,
        file_hash=download_info.file_hash,
        file_name=download_info.file_name,
        raw_file_path=str(raw_path),
    )


async def _download_store_unlinked_shipments_with_token(
    clean_store_name: str,
    token: str,
    *,
    timeout_sec: float,
    poll_interval_sec: float,
    output_dir: str | Path | None,
    report_date: str | date | None,
    download_time: str | None,
) -> StoreUnlinkedShipmentDownloadResult:
    shop = await resolve_shop_option(clean_store_name, token=token)
    safe_timeout = safe_timeout_sec(timeout_sec)
    safe_poll_interval = safe_poll_interval_sec(poll_interval_sec)
    timestamp = _timestamp_text(download_time)
    logger.info(
        "[UnlinkedShipments] 店铺匹配成功: store_name=%s store_id=%d",
        clean_store_name,
        shop.store_id,
    )

    status_results: list[UnlinkedShipmentStatusResult] = []
    for spec in UNLINKED_SHIPMENT_STATUS_SPECS:
        status_results.append(
            await _download_status_file(
                spec,
                store_name=clean_store_name,
                store_id=shop.store_id,
                token=token,
                output_dir=output_dir,
                timeout_sec=safe_timeout,
                poll_interval_sec=safe_poll_interval,
                report_date=report_date,
                download_time=timestamp,
            )
        )
    raw_count = sum(1 for row in status_results if str(row.raw_file_path or "").strip())
    logger.info("[UnlinkedShipments] 三个状态处理完成: raw_file_count=%d", raw_count)

    return StoreUnlinkedShipmentDownloadResult(
        store_name=clean_store_name,
        store_id=shop.store_id,
        download_time=timestamp,
        status_results=status_results,
    )


async def download_store_unlinked_shipments(
    store_name: str,
    *,
    timeout_sec: float = 180,
    poll_interval_sec: float = 10,
    output_dir: str | Path | None = None,
    report_date: str | date | None = None,
    download_time: str | None = None,
) -> StoreUnlinkedShipmentDownloadResult:
    clean_store_name = normalize_store_name(store_name)
    logger.info(
        "[UnlinkedShipments] 开始下载: store_name=%s timeout=%gs poll_interval=%gs",
        clean_store_name,
        safe_timeout_sec(timeout_sec),
        safe_poll_interval_sec(poll_interval_sec),
    )
    token = await get_fba_free_token()

    try:
        return await _download_store_unlinked_shipments_with_token(
            clean_store_name,
            token,
            timeout_sec=timeout_sec,
            poll_interval_sec=poll_interval_sec,
            output_dir=output_dir,
            report_date=report_date,
            download_time=download_time,
        )
    except BatchDeliveryApiAuthError:
        logger.warning("[FBAAuthRetry] 未关联货件鉴权失败，准备强制刷新 freeToken: store_name=%s", clean_store_name)

    retry_token = await get_fba_free_token(force_refresh=True)
    try:
        result = await _download_store_unlinked_shipments_with_token(
            clean_store_name,
            retry_token,
            timeout_sec=timeout_sec,
            poll_interval_sec=poll_interval_sec,
            output_dir=output_dir,
            report_date=report_date,
            download_time=download_time,
        )
    except BatchDeliveryApiAuthError as exc:
        raise BatchDeliveryApiAuthError(f"{exc}，已强制刷新后重试仍失败") from exc

    logger.info("[FBAAuthRetry] 未关联货件强制刷新后重试成功: store_name=%s", clean_store_name)
    return result


__all__ = [
    "DEFAULT_OUTPUT_DIR",
    "DEFAULT_SNAPSHOT_DIR",
    "SNAPSHOT_DETAIL_COLUMNS",
    "SNAPSHOT_DETAIL_SHEET",
    "SNAPSHOT_REQUIRED_COLUMNS",
    "SNAPSHOT_SOURCE",
    "SNAPSHOT_SUMMARY_COLUMNS",
    "SNAPSHOT_SUMMARY_SHEET",
    "SOURCE",
    "UNLINKED_SHIPMENTS_SNAPSHOT_FILE_SUFFIX",
    "UNLINKED_SHIPMENT_STATUS_SPECS",
    "ShopOption",
    "StoreUnlinkedShipmentDownloadResult",
    "UnlinkedShipmentError",
    "UnlinkedShipmentSnapshotResult",
    "UnlinkedShipmentStatusResult",
    "UnlinkedShipmentStatusSpec",
    "build_store_unlinked_shipments_snapshot",
    "create_unlinked_export_task",
    "download_raw_file_from_url",
    "download_store_unlinked_shipments",
    "fetch_shop_options",
    "fetch_status_total",
    "load_unlinked_shipment_quantities",
    "normalize_store_name",
    "pick_shop_option",
    "resolve_shop_option",
    "safe_poll_interval_sec",
    "safe_timeout_sec",
    "summarize_unlinked_shipment_details",
    "write_unlinked_shipments_snapshot",
]
