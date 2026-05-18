from __future__ import annotations

import json
import re
import csv
import shutil
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit

from shared.config import config
from shared.infra.net import erp_http_session, external_http_session

from ...auth import get_auth_context
from ...cookies import build_cookie_header, extract_named_cookies, list_cookie_names
from ...errors import MabangAuthError, MabangBusinessError, MabangRequestError
from .batch_delivery import download_fba_delivery_csv

DEFAULT_LISTSEARCH_URL = "https://private-amz.mabangerp.com/index.php?mod=fbanew.listsearch"
DEFAULT_FBA_EXPORT_URL = "https://private.mabangerp.com/index.php?mod=export.doFbaExportFile"
DEFAULT_PRIVATE_AMZ_ORIGIN = "https://private-amz.mabangerp.com"
DEFAULT_PRIVATE_AMZ_REFERER = "https://private-amz.mabangerp.com/"
DEFAULT_PRIVATE_ORIGIN = "https://private.mabangerp.com"
DEFAULT_PRIVATE_REFERER = "https://private.mabangerp.com/"
DEFAULT_OUTPUT_DIR = Path("artifacts") / "mabang_msku_detail"
DEFAULT_DELIVERY_FILE_DIR = Path("artifacts") / "mabang_fba_delivery"
PRIVATE_AMZ_HOST = "private-amz.mabangerp.com"
PRIVATE_HOST = "private.mabangerp.com"
MEMCACHE_COOKIE_NAME = "MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE"
PRIVATE_AMZ_REQUIRED_COOKIE_NAMES = (
    "PHPSESSID",
    "MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE",
    "MABANG_ERP_PRO_MEMBERINFO_LOGIN_PLUS",
    "signed",
    "route",
)
AUTH_FAIL_STATUS = {401, 403}
SOURCE = "mabang_msku_detail"
MSKU_COLUMN = "MSKU"
DELIVERY_SHOP_COLUMN = "店铺"
DETAIL_SHOP_COLUMN = "店铺名称"
SHOP_MISMATCH_SHEET_NAME = "店铺不一致"
EXPECTED_DETAIL_HEADERS = (
    "店铺名称",
    "MSKU",
    "图片链接",
    "本地SKU",
    "父ASIN",
    "ASIN",
    "本地SKU名称",
    "售价",
    "产品名称",
    "商品链接",
)


class MskuDetailDownloadError(MabangBusinessError):
    pass


class MskuDetailDownloadAuthError(MskuDetailDownloadError, MabangAuthError):
    pass


@dataclass(frozen=True)
class MskuDetailAuth:
    private_amz_cookie_header: str
    private_cookie_header: str
    memcache_key: str


@dataclass(frozen=True)
class DeliveryMskuSource:
    mskus: list[str]
    msku_shop_pairs: frozenset[tuple[str, str]]


@dataclass(frozen=True)
class MskuDetailShopSplit:
    matched_detail_count: int
    shop_mismatch_count: int
    shop_mismatch_sheet: str = SHOP_MISMATCH_SHEET_NAME


@dataclass(frozen=True)
class MskuDetailExcelResult:
    ship_no: str
    delivery_file_path: str
    delivery_file_source: str
    msku_count: int
    id_count: int
    excel_path: str
    xlsx_path: str
    converted: bool
    raw_excel_deleted: bool
    matched_detail_count: int
    shop_mismatch_count: int
    shop_mismatch_sheet: str = SHOP_MISMATCH_SHEET_NAME
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "ship_no": self.ship_no,
            "delivery_file_path": self.delivery_file_path,
            "delivery_file_source": self.delivery_file_source,
            "msku_count": self.msku_count,
            "id_count": self.id_count,
            "excel_path": self.excel_path,
            "xlsx_path": self.xlsx_path,
            "converted": self.converted,
            "raw_excel_deleted": self.raw_excel_deleted,
            "matched_detail_count": self.matched_detail_count,
            "shop_mismatch_count": self.shop_mismatch_count,
            "shop_mismatch_sheet": self.shop_mismatch_sheet,
            "source": self.source,
        }


def _configured_text(name: str, default: str) -> str:
    return str(getattr(config, name, default) or default).strip()


def _clean_cell(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() == "nan":
        return ""
    return text


def normalize_ship_no(value: Any) -> str:
    return str(value or "").strip().upper()


def _require_ship_no(value: Any) -> str:
    ship_no = normalize_ship_no(value)
    if not ship_no:
        raise ValueError("ship_no 不能为空")
    if not ship_no.startswith("SP"):
        raise ValueError(f"ship_no 格式无效: {ship_no}")
    return ship_no


def _safe_file_part(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text)
    return text.strip("._-") or "msku_detail"


def _resolve_output_dir(output_dir: str | Path | None = None) -> Path:
    if output_dir is not None:
        path = Path(output_dir)
    else:
        configured = str(getattr(config, "MABANG_MSKU_DETAIL_OUTPUT_DIR", "") or "").strip()
        path = Path(configured) if configured else DEFAULT_OUTPUT_DIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def _resolve_delivery_file_dir(delivery_file_dir: str | Path | None = None) -> Path:
    if delivery_file_dir is not None:
        return Path(delivery_file_dir)
    configured = str(getattr(config, "FBA_DELIVERY_CSV_DIR", "") or "").strip()
    return Path(configured) if configured else DEFAULT_DELIVERY_FILE_DIR


def _excel_suffix_from_url(file_url: str) -> str:
    suffix = Path(urlsplit(str(file_url or "")).path).suffix.lower()
    if suffix in {".xls", ".xlsx"}:
        return suffix
    return ".xls"


def _delivery_msku_source_from_rows(headers: list[str], rows: list[dict[str, Any]]) -> DeliveryMskuSource:
    if MSKU_COLUMN not in headers:
        raise MskuDetailDownloadError(f"发货单文件缺少列: {MSKU_COLUMN}")
    if DELIVERY_SHOP_COLUMN not in headers:
        raise MskuDetailDownloadError(f"发货单文件缺少列: {DELIVERY_SHOP_COLUMN}")

    unique_mskus: OrderedDict[str, str] = OrderedDict()
    pairs: set[tuple[str, str]] = set()
    for row in rows:
        msku = _clean_cell(row.get(MSKU_COLUMN, ""))
        shop = _clean_cell(row.get(DELIVERY_SHOP_COLUMN, ""))
        if not msku or not shop:
            continue
        unique_mskus.setdefault(msku, msku)
        pairs.add((msku, shop))

    if not unique_mskus or not pairs:
        raise MskuDetailDownloadError("发货单文件未解析到有效 MSKU 和店铺")
    return DeliveryMskuSource(mskus=list(unique_mskus.values()), msku_shop_pairs=frozenset(pairs))


def _read_csv_rows(source_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with source_path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                headers = [_clean_cell(name) for name in list(reader.fieldnames or [])]
                rows = [
                    {_clean_cell(key): _clean_cell(value) for key, value in row.items()}
                    for row in reader
                ]
            return headers, rows
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise RuntimeError(f"读取发货单 CSV 失败: {source_path.name}, error={last_error}") from last_error
    return [], []


def load_msku_source_from_delivery_file(delivery_file_path: str | Path) -> DeliveryMskuSource:
    source_path = Path(delivery_file_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"发货单文件不存在: {source_path}")

    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        headers, rows = _read_csv_rows(source_path)
        return _delivery_msku_source_from_rows(headers, rows)

    if suffix not in {".xls", ".xlsx"}:
        raise MskuDetailDownloadError(f"不支持的发货单文件格式: {source_path.suffix}")

    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError("缺少 pandas 依赖，无法读取发货单 Excel") from exc

    try:
        df = pd.read_excel(source_path, sheet_name=0, dtype=str)
    except Exception as exc:
        raise RuntimeError(f"读取发货单 Excel 失败: {source_path.name}, error={exc}") from exc

    columns = [_clean_cell(column) for column in list(df.columns)]
    df.columns = columns
    rows = [
        {_clean_cell(key): _clean_cell(value) for key, value in row.items()}
        for row in df.to_dict(orient="records")
    ]
    return _delivery_msku_source_from_rows(columns, rows)


def load_mskus_from_delivery_file(delivery_file_path: str | Path) -> list[str]:
    return load_msku_source_from_delivery_file(delivery_file_path).mskus


def load_mskus_from_consignment_excel(excel_path: str | Path) -> list[str]:
    return load_mskus_from_delivery_file(excel_path)


def find_latest_delivery_file(
    ship_no: str,
    *,
    delivery_file_dir: str | Path | None = None,
) -> Path | None:
    target = _require_ship_no(ship_no)
    directory = _resolve_delivery_file_dir(delivery_file_dir)
    if not directory.is_dir():
        return None
    candidates = [
        path
        for path in directory.glob(f"{target}_*")
        if path.is_file() and path.suffix.lower() in {".csv", ".xls", ".xlsx"}
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: (path.stat().st_mtime, path.name)).resolve()


async def resolve_delivery_file(
    ship_no: str,
    *,
    delivery_file_dir: str | Path | None = None,
) -> tuple[Path, str]:
    local_path = find_latest_delivery_file(ship_no, delivery_file_dir=delivery_file_dir)
    if local_path is not None:
        return local_path, "local"

    result = await download_fba_delivery_csv(ship_no)
    csv_path = Path(str(result.csv_path or "")).expanduser()
    if not csv_path.is_file():
        raise MskuDetailDownloadError(f"下载完成但找不到发货单文件: {csv_path}")
    return csv_path.resolve(), "downloaded"


def _request_headers(cookie_header: str, *, origin: str, referer: str) -> dict[str, str]:
    return {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": origin,
        "Referer": referer,
        "Cookie": cookie_header,
    }


def _listsearch_url() -> str:
    return _configured_text("MABANG_MSKU_LISTSEARCH_URL", DEFAULT_LISTSEARCH_URL)


def _fba_export_url() -> str:
    return _configured_text("MABANG_MSKU_FBA_EXPORT_URL", DEFAULT_FBA_EXPORT_URL)


async def _read_msku_json(resp: Any, *, action: str) -> dict[str, Any]:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise MskuDetailDownloadAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code >= 400:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")

    data: Any = None
    if text:
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            data = None
    if data is None:
        try:
            data = await resp.json(content_type=None)
        except Exception:
            data = None
    if not isinstance(data, dict):
        raise MskuDetailDownloadError(f"{action}返回非JSON对象")
    if data.get("success") is False:
        message = str(data.get("msg") or data.get("message") or data.get("error") or "unknown").strip()
        raise MskuDetailDownloadError(f"{action}业务异常: {message}")
    return data


async def resolve_msku_detail_auth() -> MskuDetailAuth:
    context = await get_auth_context(scope="private_amz")
    private_amz_cookie_header = build_cookie_header(
        context.cookies_by_domain,
        request_host=PRIVATE_AMZ_HOST,
        extra_cookies={"mabang_lite_rowsPerPage": "100"},
    )
    if not private_amz_cookie_header:
        raise MskuDetailDownloadAuthError("未获取到 private-amz.mabangerp.com Cookie")
    private_amz_cookie_names = set(
        list_cookie_names(
            context.cookies_by_domain,
            request_host=PRIVATE_AMZ_HOST,
            extra_cookies={"mabang_lite_rowsPerPage": "100"},
        )
    )
    missing_private_amz = [
        name for name in PRIVATE_AMZ_REQUIRED_COOKIE_NAMES if name not in private_amz_cookie_names
    ]
    if missing_private_amz:
        raise MskuDetailDownloadAuthError(
            f"缺少 private-amz 关键 Cookie: {', '.join(missing_private_amz)}"
        )

    private_cookie_header = build_cookie_header(
        context.cookies_by_domain,
        request_host=PRIVATE_HOST,
        extra_cookies={"exportv2": "2"},
    )
    if not private_cookie_header:
        raise MskuDetailDownloadAuthError("未获取到 private.mabangerp.com Cookie")

    values = extract_named_cookies(context.cookies_by_domain, (MEMCACHE_COOKIE_NAME,))
    memcache_key = str(values.get(MEMCACHE_COOKIE_NAME) or "").strip()
    if not memcache_key:
        raise MskuDetailDownloadAuthError(f"缺少关键 Cookie: {MEMCACHE_COOKIE_NAME}")

    return MskuDetailAuth(
        private_amz_cookie_header=private_amz_cookie_header,
        private_cookie_header=private_cookie_header,
        memcache_key=memcache_key,
    )


def _msku_lines(values: list[str]) -> str:
    return "\r\n".join(values) + "\r\n"


def _step1_form_data(mskus: list[str]) -> list[tuple[str, str]]:
    return [
        ("shopId", ""),
        ("status", "1"),
        ("ispair", ""),
        ("isChange", "1"),
        ("amazonsite", ""),
        ("stockStatus", ""),
        ("stockStatusAmz", ""),
        ("developerId", ""),
        ("saleId", ""),
        ("setdataflag", ""),
        ("searchtexttype", "platformSkuLike"),
        ("Orderby", ""),
        ("highsearch", ""),
        ("atn", "list"),
        ("searchtext", ""),
        ("searchtype", "4"),
        ("selecttype", "platformSkuIn"),
        ("platformSkuData", _msku_lines(mskus)),
    ]


def parse_listsearch_ids(payload: dict[str, Any]) -> list[str]:
    raw_id = str(payload.get("id") or "").strip()
    if not raw_id:
        raise MskuDetailDownloadError("MSKU列表查询返回缺少 id")
    ids = [item.strip() for item in raw_id.split(",") if item.strip()]
    if not ids:
        raise MskuDetailDownloadError("MSKU列表查询返回 id 为空")
    return ids


async def fetch_msku_detail_ids(mskus: list[str], *, cookie_header: str) -> list[str]:
    unique = list(OrderedDict((msku, msku) for msku in mskus if _clean_cell(msku)).values())
    if not unique:
        raise ValueError("msku 不能为空")
    headers = _request_headers(
        cookie_header,
        origin=_configured_text("MABANG_MSKU_LISTSEARCH_ORIGIN", DEFAULT_PRIVATE_AMZ_ORIGIN),
        referer=_configured_text("MABANG_MSKU_LISTSEARCH_REFERER", DEFAULT_PRIVATE_AMZ_REFERER),
    )
    async with erp_http_session.post(
        _listsearch_url(),
        data=_step1_form_data(unique),
        headers=headers,
    ) as resp:
        payload = await _read_msku_json(resp, action="MSKU列表查询")
    return parse_listsearch_ids(payload)


def _step2_form_data(ids: list[str], *, memcache_key: str) -> list[tuple[str, str]]:
    order_ids = _msku_lines(ids)
    fields = [
        ("fieldlabel", "uq101"),
        ("fieldlabel", "uq102"),
        ("fieldlabel", "uq181"),
        ("fieldlabel", "uq165"),
        ("fieldlabel", "uq103"),
        ("fieldlabel", "uq104"),
        ("fieldlabel", "uq194"),
        ("fieldlabel", "uq105"),
        ("fieldlabel", "uq141"),
        ("fieldlabel", "uq164"),
    ]
    maps = [
        ("店铺名称", "uq101"),
        ("MSKU", "uq102"),
        ("图片链接", "uq181"),
        ("本地SKU", "uq103"),
        ("父ASIN", "uq194"),
        ("ASIN", "uq105"),
        ("本地SKU名称", "uq141"),
        ("售价", "uq164"),
        ("产品名称", "uq104"),
        ("商品链接", "uq165"),
    ]
    form: list[tuple[str, str]] = [
        ("backUrl", ""),
        ("orderIds", order_ids),
        *fields,
    ]
    for name, uq in maps:
        form.extend(
            [
                ("map-name[]", name),
                ("map-uq[]", uq),
                ("map-text[]", ""),
            ]
        )
    form.extend(
        [
            ("templateName", ""),
            ("templateId", "1045273"),
            ("datasOpen", "2"),
            ("memcacheKey", memcache_key),
            ("showRmbColumn", "2"),
            ("pageSave", "1"),
            ("operateType", "5"),
            ("params", ""),
            ("InterfaceUrl", ""),
            ("mainMenu", ""),
            ("hiddenPage", ""),
            ("hiddenPageSize", ""),
            ("tableBase", ""),
            ("isMerage", "2"),
            ("showRmbColumn", "2"),
        ]
    )
    return form


def parse_export_gourl(payload: dict[str, Any]) -> str:
    url = str(payload.get("gourl") or "").strip()
    if not url:
        raise MskuDetailDownloadError("MSKU明细导出返回缺少 gourl")
    return url


async def export_msku_detail_file_url(
    ids: list[str],
    *,
    cookie_header: str,
    memcache_key: str,
) -> str:
    clean_ids = [str(item or "").strip() for item in ids if str(item or "").strip()]
    if not clean_ids:
        raise ValueError("id 不能为空")
    headers = _request_headers(
        cookie_header,
        origin=_configured_text("MABANG_MSKU_FBA_EXPORT_ORIGIN", DEFAULT_PRIVATE_ORIGIN),
        referer=_configured_text("MABANG_MSKU_FBA_EXPORT_REFERER", DEFAULT_PRIVATE_REFERER),
    )
    async with erp_http_session.post(
        _fba_export_url(),
        data=_step2_form_data(clean_ids, memcache_key=memcache_key),
        headers=headers,
    ) as resp:
        payload = await _read_msku_json(resp, action="MSKU明细导出")
    return parse_export_gourl(payload)


async def download_msku_detail_excel_from_url(
    file_url: str,
    *,
    ship_no: str,
    output_dir: str | Path | None = None,
) -> Path:
    url = str(file_url or "").strip()
    if not url:
        raise ValueError("file_url 不能为空")
    directory = _resolve_output_dir(output_dir)
    target_path = directory / f"{_safe_file_part(ship_no)}_msku_detail{_excel_suffix_from_url(url)}"
    headers = {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream,*/*"
    }
    async with external_http_session.get(url, headers=headers) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"下载MSKU明细Excel失败(status={status_code}): {msg}")
        if not body:
            raise MskuDetailDownloadError("下载MSKU明细Excel返回空文件")

    target_path.write_bytes(body)
    return target_path


def validate_msku_detail_excel_headers(excel_path: str | Path) -> None:
    source_path = Path(excel_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"MSKU明细Excel不存在: {source_path}")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取MSKU明细Excel") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        columns = [_clean_cell(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except Exception as exc:
        raise RuntimeError(f"读取MSKU明细Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    missing = [column for column in EXPECTED_DETAIL_HEADERS if column not in columns]
    if missing:
        raise MskuDetailDownloadError(f"MSKU明细Excel缺少列: {', '.join(missing)}")


def convert_msku_detail_xls_to_xlsx(excel_path: str | Path) -> Path:
    source_path = Path(excel_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"MSKU明细Excel不存在: {source_path}")

    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("缺少 xlrd 依赖，无法转换MSKU明细xls") from exc
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入MSKU明细xlsx") from exc

    target_path = source_path.with_suffix(".xlsx")
    try:
        workbook = xlrd.open_workbook(str(source_path), ignore_workbook_corruption=True)
        sheet = workbook.sheet_by_index(0)
        output = Workbook()
        worksheet = output.active
        worksheet.title = str(sheet.name or "Sheet1")[:31] or "Sheet1"
        for row_index in range(sheet.nrows):
            worksheet.append([sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)])
        output.save(target_path)
    except Exception as exc:
        raise RuntimeError(f"转换MSKU明细xls为xlsx失败: {source_path}, error={exc}") from exc
    return target_path


def normalize_msku_detail_excel(excel_path: str | Path) -> tuple[Path, bool]:
    source_path = Path(excel_path).expanduser()
    if source_path.suffix.lower() == ".xls":
        return convert_msku_detail_xls_to_xlsx(source_path), True
    if source_path.suffix.lower() == ".xlsx":
        target_path = source_path.with_name(f"{source_path.stem}_normalized.xlsx")
        shutil.copy2(source_path, target_path)
        return target_path, False
    return source_path, False


def split_msku_detail_by_delivery_shop(
    xlsx_path: str | Path,
    msku_shop_pairs: frozenset[tuple[str, str]] | set[tuple[str, str]],
) -> MskuDetailShopSplit:
    source_path = Path(xlsx_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"MSKU明细Excel不存在: {source_path}")
    allowed_pairs = {
        (_clean_cell(msku), _clean_cell(shop))
        for msku, shop in msku_shop_pairs
        if _clean_cell(msku) and _clean_cell(shop)
    }
    if not allowed_pairs:
        raise MskuDetailDownloadError("发货单文件未解析到有效 MSKU 和店铺")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法处理MSKU明细Excel") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path)
        worksheet = workbook.worksheets[0]
        headers = [_clean_cell(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if MSKU_COLUMN not in headers:
            raise MskuDetailDownloadError(f"MSKU明细Excel缺少列: {MSKU_COLUMN}")
        if DETAIL_SHOP_COLUMN not in headers:
            raise MskuDetailDownloadError(f"MSKU明细Excel缺少列: {DETAIL_SHOP_COLUMN}")

        msku_index = headers.index(MSKU_COLUMN)
        shop_index = headers.index(DETAIL_SHOP_COLUMN)
        matched_rows: list[list[Any]] = []
        mismatch_rows: list[list[Any]] = []
        max_column = len(headers)
        for row in worksheet.iter_rows(min_row=2, max_col=max_column, values_only=True):
            values = list(row)
            if not any(_clean_cell(value) for value in values):
                continue
            pair = (_clean_cell(values[msku_index]), _clean_cell(values[shop_index]))
            if pair in allowed_pairs:
                matched_rows.append(values)
            else:
                mismatch_rows.append(values)

        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row - 1)
        for row in matched_rows:
            worksheet.append(row)

        if SHOP_MISMATCH_SHEET_NAME in workbook.sheetnames:
            workbook.remove(workbook[SHOP_MISMATCH_SHEET_NAME])
        mismatch_sheet = workbook.create_sheet(SHOP_MISMATCH_SHEET_NAME, index=1)
        mismatch_sheet.append(headers)
        for row in mismatch_rows:
            mismatch_sheet.append(row)

        workbook.save(source_path)
        return MskuDetailShopSplit(
            matched_detail_count=len(matched_rows),
            shop_mismatch_count=len(mismatch_rows),
        )
    except MskuDetailDownloadError:
        raise
    except Exception as exc:
        raise RuntimeError(f"按店铺拆分MSKU明细Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def delete_raw_msku_detail_xls(excel_path: str | Path) -> bool:
    source_path = Path(excel_path).expanduser()
    if source_path.suffix.lower() != ".xls":
        return False
    if not source_path.exists():
        return False
    try:
        source_path.unlink()
    except Exception as exc:
        raise RuntimeError(f"删除MSKU明细xls失败: {source_path}, error={exc}") from exc
    return True


async def download_msku_detail_excel(
    ship_no: str,
    *,
    output_dir: str | Path | None = None,
) -> MskuDetailExcelResult:
    normalized = _require_ship_no(ship_no)
    delivery_file_path, delivery_file_source = await resolve_delivery_file(normalized)
    delivery_source = load_msku_source_from_delivery_file(delivery_file_path)
    auth = await resolve_msku_detail_auth()
    ids = await fetch_msku_detail_ids(delivery_source.mskus, cookie_header=auth.private_amz_cookie_header)
    file_url = await export_msku_detail_file_url(
        ids,
        cookie_header=auth.private_cookie_header,
        memcache_key=auth.memcache_key,
    )
    excel_path = await download_msku_detail_excel_from_url(
        file_url,
        ship_no=normalized,
        output_dir=output_dir,
    )
    xlsx_path, converted = normalize_msku_detail_excel(excel_path)
    validate_msku_detail_excel_headers(xlsx_path)
    shop_split = split_msku_detail_by_delivery_shop(xlsx_path, delivery_source.msku_shop_pairs)
    raw_excel_deleted = delete_raw_msku_detail_xls(excel_path)
    return MskuDetailExcelResult(
        ship_no=normalized,
        delivery_file_path=str(delivery_file_path),
        delivery_file_source=delivery_file_source,
        msku_count=len(delivery_source.mskus),
        id_count=len(ids),
        excel_path=str(excel_path),
        xlsx_path=str(xlsx_path),
        converted=converted,
        raw_excel_deleted=raw_excel_deleted,
        matched_detail_count=shop_split.matched_detail_count,
        shop_mismatch_count=shop_split.shop_mismatch_count,
        shop_mismatch_sheet=shop_split.shop_mismatch_sheet,
    )


__all__ = [
    "EXPECTED_DETAIL_HEADERS",
    "MSKU_COLUMN",
    "DELIVERY_SHOP_COLUMN",
    "DETAIL_SHOP_COLUMN",
    "SHOP_MISMATCH_SHEET_NAME",
    "DeliveryMskuSource",
    "MskuDetailAuth",
    "MskuDetailDownloadAuthError",
    "MskuDetailDownloadError",
    "MskuDetailExcelResult",
    "MskuDetailShopSplit",
    "convert_msku_detail_xls_to_xlsx",
    "delete_raw_msku_detail_xls",
    "download_msku_detail_excel",
    "download_msku_detail_excel_from_url",
    "export_msku_detail_file_url",
    "fetch_msku_detail_ids",
    "find_latest_delivery_file",
    "load_msku_source_from_delivery_file",
    "load_mskus_from_delivery_file",
    "load_mskus_from_consignment_excel",
    "normalize_ship_no",
    "parse_export_gourl",
    "parse_listsearch_ids",
    "normalize_msku_detail_excel",
    "resolve_delivery_file",
    "resolve_msku_detail_auth",
    "split_msku_detail_by_delivery_shop",
    "validate_msku_detail_excel_headers",
]
