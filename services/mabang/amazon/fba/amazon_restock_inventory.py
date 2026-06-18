from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from services.mabang import config as mabang_settings

from .store_msku_actual_inventory import find_latest_store_msku_file

DEFAULT_SNAPSHOT_DIR = Path("artifacts") / "amazon_restock_inventory_snapshots"
SOURCE = "amazon_restock_inventory_snapshot"
AMAZON_RESTOCK_INVENTORY_SNAPSHOT_FILE_SUFFIX = "亚马逊补充库存快照"
SUMMARY_SHEET = "亚马逊补充库存汇总"
DETAIL_SHEET = "亚马逊补充库存明细"
VALIDATION_SHEET = "校验摘要"
AMAZON_RESTOCK_TOTAL_COLUMN = "FBA 总库存（亚马逊补充库存）"
REQUIRED_CSV_COLUMNS = (
    "Merchant SKU",
    "Total Units",
    "Inbound",
    "Available",
    "FC transfer",
    "FC Processing",
    "Customer Order",
    "Working",
    "Shipped",
    "Receiving",
)
REQUIRED_MABANG_COLUMNS = ("MSKU",)
SNAPSHOT_REQUIRED_COLUMNS = ("店铺", "MSKU", "快照日期", AMAZON_RESTOCK_TOTAL_COLUMN)
SKU_MATCH_RATIO_THRESHOLD = 0.7
TOP_INVENTORY_MATCH_RATIO_THRESHOLD = 0.7
TOP_INVENTORY_LIMIT = 10
NUMBER_TOLERANCE = 1e-6

NUMERIC_COLUMNS = {
    AMAZON_RESTOCK_TOTAL_COLUMN,
    "Available",
    "Inbound",
    "Working",
    "Shipped",
    "Receiving",
    "FC transfer",
    "FC Processing",
    "Customer Order",
    "Unfulfillable",
    "Recommended replenishment qty",
    "Units Sold Last 30 Days",
    "Sales last 30 days",
    "Price",
}

SUMMARY_COLUMNS = (
    "店铺",
    "MSKU",
    "快照日期",
    "站点",
    AMAZON_RESTOCK_TOTAL_COLUMN,
    "Available",
    "Inbound",
    "Working",
    "Shipped",
    "Receiving",
    "FC transfer",
    "FC Processing",
    "Customer Order",
    "Unfulfillable",
    "Recommended replenishment qty",
    "Recommended ship date",
    "Recommended action",
    "Alert",
    "Units Sold Last 30 Days",
    "ASIN",
    "FNSKU",
    "Product Name",
    "Fulfilled by",
    "明细行数",
)
DETAIL_COLUMNS = (
    "店铺",
    "MSKU",
    "快照日期",
    "站点",
    AMAZON_RESTOCK_TOTAL_COLUMN,
    "Available",
    "Inbound",
    "Working",
    "Shipped",
    "Receiving",
    "FC transfer",
    "FC Processing",
    "Customer Order",
    "Unfulfillable",
    "Recommended replenishment qty",
    "Recommended ship date",
    "Recommended action",
    "Alert",
    "Units Sold Last 30 Days",
    "ASIN",
    "FNSKU",
    "Product Name",
    "Fulfilled by",
)


class AmazonRestockInventorySnapshotError(ValueError):
    pass


@dataclass(frozen=True)
class AmazonRestockInventoryValidationSummary:
    country: str
    mabang_site: str
    amazon_sku_count: int
    matched_amazon_sku_count: int
    amazon_sku_match_ratio: float
    top_inventory_sku_count: int
    top_inventory_matched_count: int

    def to_payload(self) -> dict[str, Any]:
        return {
            "country": self.country,
            "mabang_site": self.mabang_site,
            "amazon_sku_count": self.amazon_sku_count,
            "matched_amazon_sku_count": self.matched_amazon_sku_count,
            "amazon_sku_match_ratio": round(self.amazon_sku_match_ratio, 4),
            "top_inventory_sku_count": self.top_inventory_sku_count,
            "top_inventory_matched_count": self.top_inventory_matched_count,
        }


@dataclass(frozen=True)
class AmazonRestockInventorySnapshotResult:
    store_name: str
    snapshot_time: str
    snapshot_date: str
    snapshot_xlsx_path: str
    source_csv_path: str
    source_msku_xlsx_path: str
    row_count: int
    msku_count: int
    total_amazon_restock_inventory: float
    validation: AmazonRestockInventoryValidationSummary
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "snapshot_time": self.snapshot_time,
            "snapshot_date": self.snapshot_date,
            "snapshot_xlsx_path": self.snapshot_xlsx_path,
            "source_csv_path": self.source_csv_path,
            "source_msku_xlsx_path": self.source_msku_xlsx_path,
            "row_count": self.row_count,
            "msku_count": self.msku_count,
            "total_amazon_restock_inventory": _display_quantity(self.total_amazon_restock_inventory),
            "amazon_restock_inventory_validation": self.validation.to_payload(),
            "source": self.source,
        }


@dataclass(frozen=True)
class AmazonRestockInventorySnapshotData:
    store_name: str
    snapshot_date: str
    quantities_by_msku: dict[str, float]
    validation: AmazonRestockInventoryValidationSummary | None


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _safe_path_part(value: Any, *, fallback: str = "store") -> str:
    text = _clean_text(value)
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip(" ._-") or fallback


def _configured_path(name: str, default: Path) -> Path:
    return mabang_settings.configured_path(name, default)


def _snapshot_dir(output_dir: str | Path | None = None) -> Path:
    path = Path(output_dir) if output_dir is not None else _configured_path(
        "AMAZON_RESTOCK_INVENTORY_SNAPSHOT_DIR",
        DEFAULT_SNAPSHOT_DIR,
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _number(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        number = float(value)
        return 0.0 if math.isnan(number) else number
    text = _clean_text(value)
    if not text or text in {"-", "—"}:
        return 0.0
    text = text.replace(",", "").replace("$", "").replace("£", "").replace("€", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return 0.0 if math.isnan(number) else number


def _display_quantity(value: Any) -> int | float:
    number = float(value or 0)
    if number.is_integer():
        return int(number)
    return round(number, 2)


def _csv_records(source_path: Path) -> list[dict[str, Any]]:
    if not source_path.is_file():
        raise FileNotFoundError(f"亚马逊补充库存 CSV 不存在: {source_path}")
    raw = source_path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    records = list(csv.DictReader(text.splitlines()))
    if not records:
        raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存 CSV 没有数据: {source_path}")
    headers = set(records[0].keys())
    missing = [column for column in REQUIRED_CSV_COLUMNS if column not in headers]
    if missing:
        raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存 CSV 缺少列: {', '.join(missing)}")
    return records


def _snapshot_date_text(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    raise AmazonRestockInventorySnapshotError(f"无法解析亚马逊补充库存快照日期: {text}")


def _snapshot_date_from_records(records: list[dict[str, Any]], snapshot_date: str | None = None) -> str:
    explicit = _snapshot_date_text(snapshot_date)
    if explicit:
        return explicit

    date_fields = (
        "snapshot-date",
        "Snapshot Date",
        "snapshot_date",
        "Report Date",
        "report-date",
        "Date",
        "Request Date",
    )
    dates: set[str] = set()
    for record in records:
        for field in date_fields:
            if field in record and _clean_text(record.get(field)):
                dates.add(_snapshot_date_text(record.get(field)))
    dates.discard("")
    if len(dates) > 1:
        raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存快照日期不唯一: {', '.join(sorted(dates))}")
    if dates:
        return next(iter(dates))
    return datetime.now().strftime("%Y%m%d")


def _snapshot_time_text(snapshot_date: str, value: str | None = None) -> str:
    text = _clean_text(value)
    if text:
        if not re.fullmatch(r"\d{12}", text):
            raise AmazonRestockInventorySnapshotError(f"snapshot_time 必须是 12 位时间: {text}")
        return text
    return f"{snapshot_date}{datetime.now().strftime('%H%M')}"


def _mabang_msku_snapshot(
    store_name: str,
    *,
    msku_xlsx_path: str | Path | None = None,
    msku_dir: str | Path | None = None,
) -> tuple[Path, set[str], str]:
    source_path = Path(msku_xlsx_path) if msku_xlsx_path else find_latest_store_msku_file(store_name, input_dir=msku_dir).path
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取店铺MSKU数据Excel") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(values, None) or [])]
        missing = [column for column in REQUIRED_MABANG_COLUMNS if column not in headers]
        if missing:
            raise AmazonRestockInventorySnapshotError(f"店铺MSKU数据缺少列: {', '.join(missing)}")
        indexes = {header: index for index, header in enumerate(headers)}
        msku_values: set[str] = set()
        site_values: set[str] = set()
        site_index = indexes.get("站点")
        for row_values in values:
            if not any(_clean_text(value) for value in row_values or []):
                continue
            msku = _clean_text(row_values[indexes["MSKU"]] if indexes["MSKU"] < len(row_values) else "")
            site = _clean_text(row_values[site_index] if site_index is not None and site_index < len(row_values) else "")
            if msku:
                msku_values.add(msku)
            if site:
                site_values.add(site)
    except AmazonRestockInventorySnapshotError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取店铺MSKU数据Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    if not msku_values:
        raise AmazonRestockInventorySnapshotError(f"店铺MSKU数据没有可用于校验的 MSKU: {source_path}")
    return source_path, msku_values, ", ".join(sorted(site_values))


def _country_text(records: list[dict[str, Any]]) -> str:
    countries = {_clean_text(record.get("Country")).upper() for record in records if _clean_text(record.get("Country"))}
    return ", ".join(sorted(countries))


def _assert_inventory_formulas(record: dict[str, Any], *, msku: str) -> None:
    inbound = _number(record.get("Inbound"))
    inbound_expected = _number(record.get("Working")) + _number(record.get("Shipped")) + _number(record.get("Receiving"))
    if abs(inbound - inbound_expected) > NUMBER_TOLERANCE:
        raise AmazonRestockInventorySnapshotError(
            "亚马逊补充库存 Inbound 公式不一致: "
            f"MSKU={msku}, Inbound={_display_quantity(inbound)}, "
            f"Working+Shipped+Receiving={_display_quantity(inbound_expected)}"
        )

    total_units = _number(record.get("Total Units"))
    total_expected = (
        _number(record.get("Available"))
        + _number(record.get("FC transfer"))
        + _number(record.get("FC Processing"))
        + _number(record.get("Customer Order"))
        + inbound
    )
    if abs(total_units - total_expected) > NUMBER_TOLERANCE:
        raise AmazonRestockInventorySnapshotError(
            "亚马逊补充库存 Total Units 公式不一致: "
            f"MSKU={msku}, Total Units={_display_quantity(total_units)}, "
            f"Available+FC transfer+FC Processing+Customer Order+Inbound={_display_quantity(total_expected)}"
        )


def _detail_row(record: dict[str, Any], *, store_name: str, snapshot_date: str, country: str, msku: str) -> dict[str, Any]:
    return {
        "店铺": store_name,
        "MSKU": msku,
        "快照日期": snapshot_date,
        "站点": country,
        AMAZON_RESTOCK_TOTAL_COLUMN: _number(record.get("Total Units")),
        "Available": _number(record.get("Available")),
        "Inbound": _number(record.get("Inbound")),
        "Working": _number(record.get("Working")),
        "Shipped": _number(record.get("Shipped")),
        "Receiving": _number(record.get("Receiving")),
        "FC transfer": _number(record.get("FC transfer")),
        "FC Processing": _number(record.get("FC Processing")),
        "Customer Order": _number(record.get("Customer Order")),
        "Unfulfillable": _number(record.get("Unfulfillable")),
        "Recommended replenishment qty": _number(record.get("Recommended replenishment qty")),
        "Recommended ship date": _clean_text(record.get("Recommended ship date")),
        "Recommended action": _clean_text(record.get("Recommended action")),
        "Alert": _clean_text(record.get("Alert")),
        "Units Sold Last 30 Days": _number(record.get("Units Sold Last 30 Days")),
        "ASIN": _clean_text(record.get("ASIN")),
        "FNSKU": _clean_text(record.get("FNSKU")),
        "Product Name": _clean_text(record.get("Product Name")),
        "Fulfilled by": _clean_text(record.get("Fulfilled by")),
    }


def _merge_summary_text(current: str, value: Any) -> str:
    clean_value = _clean_text(value)
    if not clean_value:
        return current
    values = {_clean_text(item) for item in str(current).split("、") if _clean_text(item)}
    values.add(clean_value)
    return "、".join(sorted(values))


def _aggregate_amazon_rows(
    records: list[dict[str, Any]],
    *,
    store_name: str,
    snapshot_date: str | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    resolved_snapshot_date = _snapshot_date_from_records(records, snapshot_date)
    country = _country_text(records)
    summary_by_sku: dict[str, dict[str, Any]] = {}
    detail_rows: list[dict[str, Any]] = []
    for record in records:
        msku = _clean_text(record.get("Merchant SKU"))
        if not msku:
            continue
        _assert_inventory_formulas(record, msku=msku)
        detail_row = _detail_row(record, store_name=store_name, snapshot_date=resolved_snapshot_date, country=country, msku=msku)
        detail_rows.append(detail_row)
        summary = summary_by_sku.setdefault(
            msku,
            {
                "店铺": store_name,
                "MSKU": msku,
                "快照日期": resolved_snapshot_date,
                "站点": country,
                AMAZON_RESTOCK_TOTAL_COLUMN: 0.0,
                "Available": 0.0,
                "Inbound": 0.0,
                "Working": 0.0,
                "Shipped": 0.0,
                "Receiving": 0.0,
                "FC transfer": 0.0,
                "FC Processing": 0.0,
                "Customer Order": 0.0,
                "Unfulfillable": 0.0,
                "Recommended replenishment qty": 0.0,
                "Recommended ship date": "",
                "Recommended action": "",
                "Alert": "",
                "Units Sold Last 30 Days": 0.0,
                "ASIN": "",
                "FNSKU": "",
                "Product Name": "",
                "Fulfilled by": "",
                "明细行数": 0,
            },
        )
        for column in NUMERIC_COLUMNS | {"明细行数"}:
            if column == "明细行数":
                continue
            if column in summary:
                summary[column] += float(detail_row.get(column) or 0)
        for column in ("Recommended ship date", "Recommended action", "Alert"):
            summary[column] = _merge_summary_text(str(summary.get(column) or ""), detail_row.get(column))
        for column in ("ASIN", "FNSKU", "Product Name", "Fulfilled by"):
            summary[column] = summary[column] or detail_row.get(column, "")
        summary["明细行数"] += 1
    if not summary_by_sku:
        raise AmazonRestockInventorySnapshotError("亚马逊补充库存 CSV 未解析到有效 Merchant SKU")
    return list(summary_by_sku.values()), detail_rows, resolved_snapshot_date, country


def _validate_snapshot(
    *,
    country: str,
    mabang_site: str,
    mabang_mskus: set[str],
    summary_rows: list[dict[str, Any]],
) -> AmazonRestockInventoryValidationSummary:
    amazon_skus = {_clean_text(row.get("MSKU")) for row in summary_rows if _clean_text(row.get("MSKU"))}
    matched_skus = amazon_skus & mabang_mskus
    match_ratio = len(matched_skus) / len(amazon_skus) if amazon_skus else 0.0
    if match_ratio < SKU_MATCH_RATIO_THRESHOLD:
        raise AmazonRestockInventorySnapshotError(
            "亚马逊补充库存文件疑似不是当前店铺数据: "
            f"sku_match_ratio={match_ratio:.4f}, matched={len(matched_skus)}, amazon_sku_count={len(amazon_skus)}"
        )

    top_rows = sorted(
        summary_rows,
        key=lambda row: (-float(row.get(AMAZON_RESTOCK_TOTAL_COLUMN) or 0), _clean_text(row.get("MSKU"))),
    )[:TOP_INVENTORY_LIMIT]
    top_count = len(top_rows)
    top_matched_count = sum(1 for row in top_rows if _clean_text(row.get("MSKU")) in mabang_mskus)
    top_required = math.ceil(top_count * TOP_INVENTORY_MATCH_RATIO_THRESHOLD)
    if top_matched_count < top_required:
        raise AmazonRestockInventorySnapshotError(
            "亚马逊补充库存文件疑似不是当前店铺数据: "
            f"top_inventory_matched_count={top_matched_count}, top_inventory_sku_count={top_count}, required={top_required}"
        )

    return AmazonRestockInventoryValidationSummary(
        country=country,
        mabang_site=mabang_site,
        amazon_sku_count=len(amazon_skus),
        matched_amazon_sku_count=len(matched_skus),
        amazon_sku_match_ratio=match_ratio,
        top_inventory_sku_count=top_count,
        top_inventory_matched_count=top_matched_count,
    )


def _write_sheet(worksheet: Any, headers: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(
            [
                _display_quantity(row.get(header)) if header in NUMERIC_COLUMNS or header == "明细行数" else row.get(header, "")
                for header in headers
            ]
        )
    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = 18


def write_amazon_restock_inventory_snapshot(
    summary_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    validation: AmazonRestockInventoryValidationSummary,
    output_path: str | Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入亚马逊补充库存快照") from exc

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        summary_ws = workbook.active
        summary_ws.title = SUMMARY_SHEET
        _write_sheet(summary_ws, SUMMARY_COLUMNS, sorted(summary_rows, key=lambda row: _clean_text(row.get("MSKU"))))

        detail_ws = workbook.create_sheet(DETAIL_SHEET)
        _write_sheet(detail_ws, DETAIL_COLUMNS, detail_rows)

        validation_ws = workbook.create_sheet(VALIDATION_SHEET)
        validation_ws.append(["字段", "值"])
        for key, value in validation.to_payload().items():
            validation_ws.append([key, value])
        validation_ws.column_dimensions["A"].width = 32
        validation_ws.column_dimensions["B"].width = 24
        workbook.save(target_path)
    finally:
        workbook.close()
    return target_path


def build_amazon_restock_inventory_snapshot(
    csv_path: str | Path,
    *,
    store_name: str,
    output_dir: str | Path | None = None,
    msku_xlsx_path: str | Path | None = None,
    msku_dir: str | Path | None = None,
    snapshot_time: str | None = None,
    snapshot_date: str | None = None,
) -> AmazonRestockInventorySnapshotResult:
    clean_store_name = normalize_store_name(store_name)
    source_csv_path = Path(csv_path)
    records = _csv_records(source_csv_path)
    source_msku_path, mabang_mskus, mabang_site = _mabang_msku_snapshot(
        clean_store_name,
        msku_xlsx_path=msku_xlsx_path,
        msku_dir=msku_dir,
    )
    summary_rows, detail_rows, resolved_snapshot_date, country = _aggregate_amazon_rows(
        records,
        store_name=clean_store_name,
        snapshot_date=snapshot_date,
    )
    validation = _validate_snapshot(
        country=country,
        mabang_site=mabang_site,
        mabang_mskus=mabang_mskus,
        summary_rows=summary_rows,
    )
    timestamp = _snapshot_time_text(resolved_snapshot_date, snapshot_time)
    target_path = _snapshot_dir(output_dir) / f"{timestamp}-{_safe_path_part(clean_store_name)}_{AMAZON_RESTOCK_INVENTORY_SNAPSHOT_FILE_SUFFIX}.xlsx"
    write_amazon_restock_inventory_snapshot(summary_rows, detail_rows, validation, target_path)
    return AmazonRestockInventorySnapshotResult(
        store_name=clean_store_name,
        snapshot_time=timestamp,
        snapshot_date=resolved_snapshot_date,
        snapshot_xlsx_path=str(target_path),
        source_csv_path=str(source_csv_path),
        source_msku_xlsx_path=str(source_msku_path),
        row_count=len(detail_rows),
        msku_count=len(summary_rows),
        total_amazon_restock_inventory=sum(_number(row.get(AMAZON_RESTOCK_TOTAL_COLUMN)) for row in summary_rows),
        validation=validation,
    )


def _records_from_xlsx(path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(f"亚马逊补充库存快照不存在: {source_path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取亚马逊补充库存快照") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存快照缺少 sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        records = []
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            if any(_clean_text(value) for value in row.values()):
                records.append(row)
        return records
    except AmazonRestockInventorySnapshotError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取亚马逊补充库存快照失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def _validation_from_records(records: list[dict[str, Any]]) -> AmazonRestockInventoryValidationSummary | None:
    if not records:
        return None
    values = {_clean_text(row.get("字段")): row.get("值") for row in records}
    required = {
        "country",
        "mabang_site",
        "amazon_sku_count",
        "matched_amazon_sku_count",
        "amazon_sku_match_ratio",
        "top_inventory_sku_count",
        "top_inventory_matched_count",
    }
    if not required.issubset(values):
        return None
    return AmazonRestockInventoryValidationSummary(
        country=_clean_text(values["country"]),
        mabang_site=_clean_text(values["mabang_site"]),
        amazon_sku_count=int(_number(values["amazon_sku_count"])),
        matched_amazon_sku_count=int(_number(values["matched_amazon_sku_count"])),
        amazon_sku_match_ratio=float(_number(values["amazon_sku_match_ratio"])),
        top_inventory_sku_count=int(_number(values["top_inventory_sku_count"])),
        top_inventory_matched_count=int(_number(values["top_inventory_matched_count"])),
    )


def load_amazon_restock_inventory_snapshot(path: str | Path, *, store_name: str | None = None) -> AmazonRestockInventorySnapshotData:
    records = _records_from_xlsx(path, SUMMARY_SHEET)
    headers = set(records[0].keys()) if records else set(SNAPSHOT_REQUIRED_COLUMNS)
    missing = [column for column in SNAPSHOT_REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存快照缺少列: {', '.join(missing)}, path={path}")
    clean_store_name = _clean_text(store_name)
    stores = {_clean_text(row.get("店铺")) for row in records if _clean_text(row.get("店铺"))}
    if clean_store_name and stores and clean_store_name not in stores:
        raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存快照中未找到店铺: {clean_store_name}")
    dates = {_clean_text(row.get("快照日期")) for row in records if _clean_text(row.get("快照日期"))}
    if len(dates) != 1:
        raise AmazonRestockInventorySnapshotError(f"亚马逊补充库存快照日期不唯一: {', '.join(sorted(dates)) or '空'}")

    quantities: dict[str, float] = {}
    for record in records:
        if clean_store_name and _clean_text(record.get("店铺")) != clean_store_name:
            continue
        msku = _clean_text(record.get("MSKU"))
        if not msku:
            continue
        quantities[msku] = quantities.get(msku, 0.0) + _number(record.get(AMAZON_RESTOCK_TOTAL_COLUMN))

    validation = _validation_from_records(_records_from_xlsx(path, VALIDATION_SHEET))
    return AmazonRestockInventorySnapshotData(
        store_name=clean_store_name or (next(iter(stores)) if len(stores) == 1 else ""),
        snapshot_date=next(iter(dates)),
        quantities_by_msku=quantities,
        validation=validation,
    )


__all__ = [
    "AMAZON_RESTOCK_INVENTORY_SNAPSHOT_FILE_SUFFIX",
    "AMAZON_RESTOCK_TOTAL_COLUMN",
    "AmazonRestockInventorySnapshotData",
    "AmazonRestockInventorySnapshotError",
    "AmazonRestockInventorySnapshotResult",
    "AmazonRestockInventoryValidationSummary",
    "DEFAULT_SNAPSHOT_DIR",
    "SOURCE",
    "build_amazon_restock_inventory_snapshot",
    "load_amazon_restock_inventory_snapshot",
    "normalize_store_name",
    "write_amazon_restock_inventory_snapshot",
]
