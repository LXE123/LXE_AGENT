from __future__ import annotations

import csv
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from services.mabang import config as mabang_settings

from .store_msku_actual_inventory import find_latest_store_msku_file

DEFAULT_SNAPSHOT_DIR = Path("artifacts") / "amazon_fba_inventory_snapshots"
SOURCE = "amazon_fba_inventory_snapshot"
AMAZON_INVENTORY_SNAPSHOT_FILE_SUFFIX = "亚马逊后台库存快照"
SUMMARY_SHEET = "Amazon库存汇总"
DETAIL_SHEET = "Amazon库存明细"
VALIDATION_SHEET = "校验摘要"
AMAZON_FBA_TOTAL_COLUMN = "FBA 总库存（亚马逊后台数据）"
REQUIRED_CSV_COLUMNS = ("snapshot-date", "sku", "marketplace", "Inventory Supply at FBA")
REQUIRED_MABANG_COLUMNS = ("MSKU",)
SNAPSHOT_REQUIRED_COLUMNS = ("店铺", "MSKU", "快照日期", AMAZON_FBA_TOTAL_COLUMN)
SKU_MATCH_RATIO_THRESHOLD = 0.7
TOP_INVENTORY_MATCH_RATIO_THRESHOLD = 0.7
TOP_INVENTORY_LIMIT = 10

SUMMARY_COLUMNS = (
    "店铺",
    "MSKU",
    "快照日期",
    "站点",
    AMAZON_FBA_TOTAL_COLUMN,
    "available",
    "inbound-quantity",
    "Total Reserved Quantity",
    "unfulfillable-quantity",
    "estimated-excess-quantity",
    "recommended-action",
    "asin",
    "product-name",
    "明细行数",
)
DETAIL_COLUMNS = (
    "店铺",
    "MSKU",
    "快照日期",
    "站点",
    AMAZON_FBA_TOTAL_COLUMN,
    "available",
    "inbound-quantity",
    "Total Reserved Quantity",
    "unfulfillable-quantity",
    "estimated-excess-quantity",
    "recommended-action",
    "asin",
    "product-name",
)

SITE_TO_MARKETPLACE = {
    "美国站": "US",
    "美国": "US",
    "US": "US",
    "英国站": "UK",
    "英国": "UK",
    "UK": "UK",
    "GB": "UK",
    "德国站": "DE",
    "德国": "DE",
    "DE": "DE",
    "法国站": "FR",
    "法国": "FR",
    "FR": "FR",
    "意大利站": "IT",
    "意大利": "IT",
    "IT": "IT",
    "西班牙站": "ES",
    "西班牙": "ES",
    "ES": "ES",
    "加拿大站": "CA",
    "加拿大": "CA",
    "CA": "CA",
    "日本站": "JP",
    "日本": "JP",
    "JP": "JP",
    "澳大利亚站": "AU",
    "澳大利亚": "AU",
    "AU": "AU",
    "墨西哥站": "MX",
    "墨西哥": "MX",
    "MX": "MX",
    "荷兰站": "NL",
    "荷兰": "NL",
    "NL": "NL",
    "瑞典站": "SE",
    "瑞典": "SE",
    "SE": "SE",
    "波兰站": "PL",
    "波兰": "PL",
    "PL": "PL",
    "比利时站": "BE",
    "比利时": "BE",
    "BE": "BE",
}


class AmazonInventorySnapshotError(ValueError):
    pass


@dataclass(frozen=True)
class AmazonInventoryValidationSummary:
    marketplace: str
    mabang_site: str
    amazon_sku_count: int
    matched_amazon_sku_count: int
    amazon_sku_match_ratio: float
    top_inventory_sku_count: int
    top_inventory_matched_count: int

    def to_payload(self) -> dict[str, Any]:
        return {
            "marketplace": self.marketplace,
            "mabang_site": self.mabang_site,
            "amazon_sku_count": self.amazon_sku_count,
            "matched_amazon_sku_count": self.matched_amazon_sku_count,
            "amazon_sku_match_ratio": round(self.amazon_sku_match_ratio, 4),
            "top_inventory_sku_count": self.top_inventory_sku_count,
            "top_inventory_matched_count": self.top_inventory_matched_count,
        }


@dataclass(frozen=True)
class AmazonInventorySnapshotResult:
    store_name: str
    snapshot_time: str
    snapshot_date: str
    snapshot_xlsx_path: str
    source_csv_path: str
    source_msku_xlsx_path: str
    row_count: int
    msku_count: int
    total_amazon_fba_inventory: float
    validation: AmazonInventoryValidationSummary
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "snapshot_time": self.snapshot_time,
            "snapshot_date": self.snapshot_date,
            "snapshot_xlsx_path": self.snapshot_xlsx_path,
            "source_csv_path": self.source_csv_path,
            "source_msku_xlsx_path": self.source_msku_xlsx_path,
            "row_count": self.row_count,
            "msku_count": self.msku_count,
            "total_amazon_fba_inventory": _display_quantity(self.total_amazon_fba_inventory),
            "amazon_inventory_validation": self.validation.to_payload(),
            "source": self.source,
        }


@dataclass(frozen=True)
class AmazonInventorySnapshotData:
    store_name: str
    snapshot_date: str
    quantities_by_msku: dict[str, float]
    validation: AmazonInventoryValidationSummary | None


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def _safe_path_part(value: Any, *, fallback: str = "store") -> str:
    text = _clean_text(value)
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip(" ._-") or fallback


def _configured_path(name: str, default: Path) -> Path:
    return mabang_settings.configured_path(name, default)


def _snapshot_dir(output_dir: str | Path | None = None) -> Path:
    path = Path(output_dir) if output_dir is not None else _configured_path(
        "AMAZON_FBA_INVENTORY_SNAPSHOT_DIR",
        DEFAULT_SNAPSHOT_DIR,
    )
    path.mkdir(parents=True, exist_ok=True)
    return path


def _number(value: Any) -> float:
    if isinstance(value, bool) or value is None:
        return 0.0
    if isinstance(value, (int, float)):
        number = float(value)
        return 0.0 if math.isnan(number) else number
    text = _clean_text(value)
    if not text or text in {"-", "—"}:
        return 0.0
    text = text.replace(",", "").replace("$", "").replace("£", "").replace("€", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return 0.0 if math.isnan(number) else number


def _display_quantity(value: Any) -> int | float:
    number = float(value or 0)
    if number.is_integer():
        return int(number)
    return round(number, 2)


def _csv_records(source_path: Path) -> list[dict[str, Any]]:
    if not source_path.is_file():
        raise FileNotFoundError(f"Amazon 后台库存 CSV 不存在: {source_path}")
    raw = source_path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    records = list(csv.DictReader(text.splitlines()))
    if not records:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存 CSV 没有数据: {source_path}")
    headers = set(records[0].keys())
    missing = [column for column in REQUIRED_CSV_COLUMNS if column not in headers]
    if missing:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存 CSV 缺少列: {', '.join(missing)}")
    return records


def _snapshot_date_text(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    raise AmazonInventorySnapshotError(f"无法解析 Amazon 后台库存 snapshot-date: {text}")


def _snapshot_time_text(snapshot_date: str, value: str | None = None) -> str:
    text = _clean_text(value)
    if text:
        if not re.fullmatch(r"\d{12}", text):
            raise AmazonInventorySnapshotError(f"snapshot_time 必须是 12 位时间: {text}")
        return text
    return f"{snapshot_date}{datetime.now().strftime('%H%M')}"


def _marketplace_code(value: Any) -> str:
    text = _clean_text(value).upper()
    return SITE_TO_MARKETPLACE.get(text, text)


def _mabang_msku_snapshot(
    store_name: str,
    *,
    msku_xlsx_path: str | Path | None = None,
    msku_dir: str | Path | None = None,
) -> tuple[Path, set[str], str]:
    source_path = Path(msku_xlsx_path) if msku_xlsx_path else find_latest_store_msku_file(store_name, input_dir=msku_dir).path
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取店铺MSKU数据Excel") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(values, None) or [])]
        missing = [column for column in REQUIRED_MABANG_COLUMNS if column not in headers]
        if missing:
            raise AmazonInventorySnapshotError(f"店铺MSKU数据缺少列: {', '.join(missing)}")
        indexes = {header: index for index, header in enumerate(headers)}
        msku_values: set[str] = set()
        site_values: set[str] = set()
        site_index = indexes.get("站点")
        for row_values in values:
            if not any(_clean_text(value) for value in row_values or []):
                continue
            msku = _clean_text(row_values[indexes["MSKU"]] if indexes["MSKU"] < len(row_values) else "")
            site = _clean_text(row_values[site_index] if site_index is not None and site_index < len(row_values) else "")
            if msku:
                msku_values.add(msku)
            if site:
                site_values.add(site)
    except AmazonInventorySnapshotError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取店铺MSKU数据Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    if not msku_values:
        raise AmazonInventorySnapshotError(f"店铺MSKU数据没有可用于校验的 MSKU: {source_path}")
    return source_path, msku_values, ", ".join(sorted(site_values))


def _aggregate_amazon_rows(records: list[dict[str, Any]], *, store_name: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    snapshot_dates = {_snapshot_date_text(record.get("snapshot-date")) for record in records}
    snapshot_dates.discard("")
    if len(snapshot_dates) != 1:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存 snapshot-date 不唯一: {', '.join(sorted(snapshot_dates)) or '空'}")
    marketplaces = {_marketplace_code(record.get("marketplace")) for record in records if _clean_text(record.get("marketplace"))}
    if len(marketplaces) != 1:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存 marketplace 不唯一: {', '.join(sorted(marketplaces)) or '空'}")
    snapshot_date = next(iter(snapshot_dates))
    marketplace = next(iter(marketplaces))

    summary_by_sku: dict[str, dict[str, Any]] = {}
    detail_rows: list[dict[str, Any]] = []
    for record in records:
        msku = _clean_text(record.get("sku"))
        if not msku:
            continue
        detail_row = {
            "店铺": store_name,
            "MSKU": msku,
            "快照日期": snapshot_date,
            "站点": marketplace,
            AMAZON_FBA_TOTAL_COLUMN: _number(record.get("Inventory Supply at FBA")),
            "available": _number(record.get("available")),
            "inbound-quantity": _number(record.get("inbound-quantity")),
            "Total Reserved Quantity": _number(record.get("Total Reserved Quantity")),
            "unfulfillable-quantity": _number(record.get("unfulfillable-quantity")),
            "estimated-excess-quantity": _number(record.get("estimated-excess-quantity")),
            "recommended-action": _clean_text(record.get("recommended-action")),
            "asin": _clean_text(record.get("asin")),
            "product-name": _clean_text(record.get("product-name")),
        }
        detail_rows.append(detail_row)
        summary = summary_by_sku.setdefault(
            msku,
            {
                "店铺": store_name,
                "MSKU": msku,
                "快照日期": snapshot_date,
                "站点": marketplace,
                AMAZON_FBA_TOTAL_COLUMN: 0.0,
                "available": 0.0,
                "inbound-quantity": 0.0,
                "Total Reserved Quantity": 0.0,
                "unfulfillable-quantity": 0.0,
                "estimated-excess-quantity": 0.0,
                "recommended-action": "",
                "asin": "",
                "product-name": "",
                "明细行数": 0,
            },
        )
        for column in (
            AMAZON_FBA_TOTAL_COLUMN,
            "available",
            "inbound-quantity",
            "Total Reserved Quantity",
            "unfulfillable-quantity",
            "estimated-excess-quantity",
        ):
            summary[column] += float(detail_row[column] or 0)
        actions = {_clean_text(action) for action in str(summary["recommended-action"]).split("、") if _clean_text(action)}
        if detail_row["recommended-action"]:
            actions.add(detail_row["recommended-action"])
        summary["recommended-action"] = "、".join(sorted(actions))
        summary["asin"] = summary["asin"] or detail_row["asin"]
        summary["product-name"] = summary["product-name"] or detail_row["product-name"]
        summary["明细行数"] += 1
    if not summary_by_sku:
        raise AmazonInventorySnapshotError("Amazon 后台库存 CSV 未解析到有效 sku")
    return list(summary_by_sku.values()), detail_rows, snapshot_date, marketplace


def _validate_snapshot(
    *,
    store_name: str,
    marketplace: str,
    mabang_site: str,
    mabang_mskus: set[str],
    summary_rows: list[dict[str, Any]],
) -> AmazonInventoryValidationSummary:
    amazon_skus = {_clean_text(row.get("MSKU")) for row in summary_rows if _clean_text(row.get("MSKU"))}
    matched_skus = amazon_skus & mabang_mskus
    match_ratio = len(matched_skus) / len(amazon_skus) if amazon_skus else 0.0
    if match_ratio < SKU_MATCH_RATIO_THRESHOLD:
        raise AmazonInventorySnapshotError(
            "Amazon 后台库存文件疑似不是当前店铺/站点数据: "
            f"sku_match_ratio={match_ratio:.4f}, matched={len(matched_skus)}, amazon_sku_count={len(amazon_skus)}"
        )

    top_rows = sorted(
        summary_rows,
        key=lambda row: (-float(row.get(AMAZON_FBA_TOTAL_COLUMN) or 0), _clean_text(row.get("MSKU"))),
    )[:TOP_INVENTORY_LIMIT]
    top_count = len(top_rows)
    top_matched_count = sum(1 for row in top_rows if _clean_text(row.get("MSKU")) in mabang_mskus)
    top_required = math.ceil(top_count * TOP_INVENTORY_MATCH_RATIO_THRESHOLD)
    if top_matched_count < top_required:
        raise AmazonInventorySnapshotError(
            "Amazon 后台库存文件疑似不是当前店铺/站点数据: "
            f"top_inventory_matched_count={top_matched_count}, top_inventory_sku_count={top_count}, required={top_required}"
        )

    return AmazonInventoryValidationSummary(
        marketplace=marketplace,
        mabang_site=mabang_site,
        amazon_sku_count=len(amazon_skus),
        matched_amazon_sku_count=len(matched_skus),
        amazon_sku_match_ratio=match_ratio,
        top_inventory_sku_count=top_count,
        top_inventory_matched_count=top_matched_count,
    )


def _write_sheet(worksheet: Any, headers: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([_display_quantity(row.get(header)) if header in {
            AMAZON_FBA_TOTAL_COLUMN,
            "available",
            "inbound-quantity",
            "Total Reserved Quantity",
            "unfulfillable-quantity",
            "estimated-excess-quantity",
            "明细行数",
        } else row.get(header, "") for header in headers])
    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        worksheet.column_dimensions[column_cells[0].column_letter].width = 18


def write_amazon_inventory_snapshot(
    summary_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    validation: AmazonInventoryValidationSummary,
    output_path: str | Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入 Amazon 后台库存快照") from exc

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        summary_ws = workbook.active
        summary_ws.title = SUMMARY_SHEET
        _write_sheet(summary_ws, SUMMARY_COLUMNS, sorted(summary_rows, key=lambda row: _clean_text(row.get("MSKU"))))

        detail_ws = workbook.create_sheet(DETAIL_SHEET)
        _write_sheet(detail_ws, DETAIL_COLUMNS, detail_rows)

        validation_ws = workbook.create_sheet(VALIDATION_SHEET)
        validation_ws.append(["字段", "值"])
        for key, value in validation.to_payload().items():
            validation_ws.append([key, value])
        validation_ws.column_dimensions["A"].width = 32
        validation_ws.column_dimensions["B"].width = 24
        workbook.save(target_path)
    finally:
        workbook.close()
    return target_path


def build_amazon_inventory_snapshot(
    csv_path: str | Path,
    *,
    store_name: str,
    output_dir: str | Path | None = None,
    msku_xlsx_path: str | Path | None = None,
    msku_dir: str | Path | None = None,
    snapshot_time: str | None = None,
) -> AmazonInventorySnapshotResult:
    clean_store_name = normalize_store_name(store_name)
    source_csv_path = Path(csv_path)
    records = _csv_records(source_csv_path)
    source_msku_path, mabang_mskus, mabang_site = _mabang_msku_snapshot(
        clean_store_name,
        msku_xlsx_path=msku_xlsx_path,
        msku_dir=msku_dir,
    )
    summary_rows, detail_rows, snapshot_date, marketplace = _aggregate_amazon_rows(records, store_name=clean_store_name)
    validation = _validate_snapshot(
        store_name=clean_store_name,
        marketplace=marketplace,
        mabang_site=mabang_site,
        mabang_mskus=mabang_mskus,
        summary_rows=summary_rows,
    )
    timestamp = _snapshot_time_text(snapshot_date, snapshot_time)
    target_path = _snapshot_dir(output_dir) / f"{timestamp}-{_safe_path_part(clean_store_name)}_{AMAZON_INVENTORY_SNAPSHOT_FILE_SUFFIX}.xlsx"
    write_amazon_inventory_snapshot(summary_rows, detail_rows, validation, target_path)
    return AmazonInventorySnapshotResult(
        store_name=clean_store_name,
        snapshot_time=timestamp,
        snapshot_date=snapshot_date,
        snapshot_xlsx_path=str(target_path),
        source_csv_path=str(source_csv_path),
        source_msku_xlsx_path=str(source_msku_path),
        row_count=len(detail_rows),
        msku_count=len(summary_rows),
        total_amazon_fba_inventory=sum(_number(row.get(AMAZON_FBA_TOTAL_COLUMN)) for row in summary_rows),
        validation=validation,
    )


def _records_from_xlsx(path: str | Path, sheet_name: str) -> list[dict[str, Any]]:
    source_path = Path(path)
    if not source_path.is_file():
        raise FileNotFoundError(f"Amazon 后台库存快照不存在: {source_path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取 Amazon 后台库存快照") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise AmazonInventorySnapshotError(f"Amazon 后台库存快照缺少 sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = [_clean_text(cell) for cell in list(next(rows, None) or [])]
        records = []
        for values in rows:
            row = dict(zip(headers, list(values or []), strict=False))
            if any(_clean_text(value) for value in row.values()):
                records.append(row)
        return records
    except AmazonInventorySnapshotError:
        raise
    except Exception as exc:
        raise RuntimeError(f"读取 Amazon 后台库存快照失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass


def _validation_from_records(records: list[dict[str, Any]]) -> AmazonInventoryValidationSummary | None:
    if not records:
        return None
    values = {_clean_text(row.get("字段")): row.get("值") for row in records}
    required = {
        "marketplace",
        "mabang_site",
        "amazon_sku_count",
        "matched_amazon_sku_count",
        "amazon_sku_match_ratio",
        "top_inventory_sku_count",
        "top_inventory_matched_count",
    }
    if not required.issubset(values):
        return None
    return AmazonInventoryValidationSummary(
        marketplace=_clean_text(values["marketplace"]),
        mabang_site=_clean_text(values["mabang_site"]),
        amazon_sku_count=int(_number(values["amazon_sku_count"])),
        matched_amazon_sku_count=int(_number(values["matched_amazon_sku_count"])),
        amazon_sku_match_ratio=float(_number(values["amazon_sku_match_ratio"])),
        top_inventory_sku_count=int(_number(values["top_inventory_sku_count"])),
        top_inventory_matched_count=int(_number(values["top_inventory_matched_count"])),
    )


def load_amazon_inventory_snapshot(path: str | Path, *, store_name: str | None = None) -> AmazonInventorySnapshotData:
    records = _records_from_xlsx(path, SUMMARY_SHEET)
    headers = set(records[0].keys()) if records else set(SNAPSHOT_REQUIRED_COLUMNS)
    missing = [column for column in SNAPSHOT_REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存快照缺少列: {', '.join(missing)}, path={path}")
    clean_store_name = _clean_text(store_name)
    stores = {_clean_text(row.get("店铺")) for row in records if _clean_text(row.get("店铺"))}
    if clean_store_name and stores and clean_store_name not in stores:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存快照中未找到店铺: {clean_store_name}")
    dates = {_clean_text(row.get("快照日期")) for row in records if _clean_text(row.get("快照日期"))}
    if len(dates) != 1:
        raise AmazonInventorySnapshotError(f"Amazon 后台库存快照日期不唯一: {', '.join(sorted(dates)) or '空'}")

    quantities: dict[str, float] = {}
    for record in records:
        if clean_store_name and _clean_text(record.get("店铺")) != clean_store_name:
            continue
        msku = _clean_text(record.get("MSKU"))
        if not msku:
            continue
        quantities[msku] = quantities.get(msku, 0.0) + _number(record.get(AMAZON_FBA_TOTAL_COLUMN))

    validation = _validation_from_records(_records_from_xlsx(path, VALIDATION_SHEET))
    return AmazonInventorySnapshotData(
        store_name=clean_store_name or (next(iter(stores)) if len(stores) == 1 else ""),
        snapshot_date=next(iter(dates)),
        quantities_by_msku=quantities,
        validation=validation,
    )


__all__ = [
    "AMAZON_FBA_TOTAL_COLUMN",
    "AMAZON_INVENTORY_SNAPSHOT_FILE_SUFFIX",
    "AmazonInventorySnapshotData",
    "AmazonInventorySnapshotError",
    "AmazonInventorySnapshotResult",
    "AmazonInventoryValidationSummary",
    "DEFAULT_SNAPSHOT_DIR",
    "SOURCE",
    "build_amazon_inventory_snapshot",
    "load_amazon_inventory_snapshot",
    "normalize_store_name",
    "write_amazon_inventory_snapshot",
]
