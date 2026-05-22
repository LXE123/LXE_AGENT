from __future__ import annotations

import asyncio
from datetime import datetime
from pathlib import Path
from typing import Any

from services.amazon.amazon_logistic.input.validator import normalize_consignment_no
from services.amazon.amazon_logistic.sources.consignment_excel import (
    ensure_consignment_excel_ready,
    find_consignment_excel,
    resolve_column,
    resolve_consignment_excel_dir,
)


def _resolve_any_path(raw: str) -> Path:
    text = str(raw or "").strip()
    if not text:
        raise ValueError("路径不能为空")
    path = Path(text).expanduser()
    if not path.is_absolute():
        path = (Path.cwd() / path).resolve()
    else:
        path = path.resolve()
    return path


def _detect_consignment_source(excel_path: Path) -> str:
    try:
        local_dir = resolve_consignment_excel_dir().resolve()
        if excel_path.resolve().is_relative_to(local_dir):
            return "local"
    except Exception:
        pass
    return "wms"


def _normalize_site(site: str) -> str:
    text = str(site or "").strip().upper()
    aliases = {
        "美国": "US",
        "美国站": "US",
        "US": "US",
        "USA": "US",
        "英国": "UK",
        "英国站": "UK",
        "UK": "UK",
        "GB": "UK",
        "德国": "DE",
        "德国站": "DE",
        "DE": "DE",
        "法国": "FR",
        "法国站": "FR",
        "FR": "FR",
        "意大利": "IT",
        "IT": "IT",
        "西班牙": "ES",
        "ES": "ES",
        "加拿大": "CA",
        "CA": "CA",
        "日本": "JP",
        "JP": "JP",
    }
    return aliases.get(text, text)


def _normalize_quantity(value: Any) -> int:
    try:
        numeric = float(value)
    except Exception as exc:
        raise RuntimeError(f"数量列包含不可解析值: {value}") from exc
    rounded = round(numeric)
    if abs(numeric - rounded) > 1e-6:
        raise RuntimeError(f"数量列包含非整数值: {value}")
    return int(rounded)


def _normalize_template_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower()


def _select_template_sheet_name(sheet_names: list[str]) -> str:
    preferred = "Create workflow – template"
    if preferred in sheet_names:
        return preferred
    fallback = "Create workflow - template"
    if fallback in sheet_names:
        return fallback
    for name in sheet_names:
        lower = str(name or "").strip().lower()
        if "workflow" in lower or "template" in lower:
            return str(name)
    raise RuntimeError(f"找不到模板工作表，可用工作表: {sheet_names}")


def _find_header_row(
    worksheet: Any,
    *,
    required_headers: tuple[str, ...],
    search_row_limit: int = 40,
    search_col_limit: int = 20,
) -> tuple[int, dict[str, int]] | None:
    required = tuple(_normalize_template_text(header) for header in required_headers if str(header or "").strip())
    max_row = min(max(1, int(getattr(worksheet, "max_row", 0) or 0)), max(1, int(search_row_limit)))
    max_col = min(max(1, int(getattr(worksheet, "max_column", 0) or 0)), max(1, int(search_col_limit)))

    for row_idx in range(1, max_row + 1):
        found: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            normalized = _normalize_template_text(worksheet.cell(row=row_idx, column=col_idx).value)
            if not normalized:
                continue
            for header in required:
                if normalized == header and header not in found:
                    found[header] = col_idx
        if all(header in found for header in required):
            return row_idx, found
    return None


def _find_header_column(
    worksheet: Any,
    *,
    header_row: int,
    header_columns: dict[str, int],
    header: str,
) -> int:
    normalized_header = _normalize_template_text(header)
    if normalized_header in header_columns:
        return int(header_columns[normalized_header])

    max_col = max(1, int(getattr(worksheet, "max_column", 0) or 0))
    for col_idx in range(1, max_col + 1):
        normalized = _normalize_template_text(worksheet.cell(row=int(header_row), column=col_idx).value)
        if normalized == normalized_header:
            return col_idx
    return 0


def _select_template_sheet(
    workbook: Any,
    *,
    normalized_site: str,
) -> tuple[str, Any, int, dict[str, int]]:
    required_headers = ("Merchant SKU", "Quantity")
    for worksheet in workbook.worksheets:
        matched = _find_header_row(worksheet, required_headers=required_headers)
        if matched is None:
            continue
        header_row, header_columns = matched
        return str(worksheet.title), worksheet, header_row, header_columns

    fallback_sheet_name = _select_template_sheet_name([str(name) for name in workbook.sheetnames])
    fallback_worksheet = workbook[fallback_sheet_name]
    matched = _find_header_row(fallback_worksheet, required_headers=required_headers)
    if matched is None:
        raise RuntimeError(
            "Amazon 模板中未找到所需标题: Merchant SKU, Quantity "
            f"(工作表: {fallback_sheet_name})"
        )
    header_row, header_columns = matched
    return fallback_sheet_name, fallback_worksheet, header_row, header_columns


def _build_alt_template_path(template_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return template_path.with_name(f"{template_path.stem}_filled_{timestamp}{template_path.suffix}")


def _load_consignment_summary(consignment_excel_path: Path) -> tuple[Any, list[dict[str, Any]], int]:
    try:
        import pandas as pd
    except Exception as exc:
        raise RuntimeError("缺少 pandas 依赖，无法读取托运单 Excel") from exc

    try:
        df = pd.read_excel(consignment_excel_path)
    except FileNotFoundError as exc:
        raise RuntimeError(f"找不到托运单 Excel: {consignment_excel_path}") from exc
    except Exception as exc:
        raise RuntimeError(f"读取托运单 Excel 失败: {exc}") from exc

    if df.empty:
        raise RuntimeError(f"托运单 Excel 没有数据: {consignment_excel_path.name}")

    columns = [str(column) for column in list(df.columns)]
    msku_col = resolve_column(columns, ("MSKU", "Merchant SKU", "merchant sku"))
    quantity_col = resolve_column(columns, ("装箱数量", "Quantity", "数量"))

    if not msku_col or not quantity_col:
        raise RuntimeError(
            "托运单 Excel 缺少必需列: "
            f"MSKU={msku_col or 'missing'}, Quantity={quantity_col or 'missing'}"
        )

    working = df[[msku_col, quantity_col]].copy()
    working[msku_col] = working[msku_col].map(lambda value: str(value or "").strip())
    working = working[working[msku_col] != ""]
    if working.empty:
        raise RuntimeError(f"托运单 Excel 未解析到有效 MSKU: {consignment_excel_path.name}")

    try:
        working[quantity_col] = working[quantity_col].map(_normalize_quantity)
    except Exception as exc:
        raise RuntimeError(f"托运单 Excel 数量列无效: {exc}") from exc

    grouped = working.groupby(msku_col, as_index=False)[quantity_col].sum()
    grouped.columns = ["Merchant SKU", "Quantity"]
    grouped = grouped.sort_values(by="Merchant SKU", kind="stable").reset_index(drop=True)

    sku_summary: list[dict[str, Any]] = []
    for row in grouped.to_dict(orient="records"):
        sku_summary.append({
            "msku": str(row.get("Merchant SKU") or "").strip(),
            "quantity": int(row.get("Quantity") or 0),
        })
    total_quantity = sum(int(item["quantity"]) for item in sku_summary)
    return grouped, sku_summary, total_quantity


def _clear_template_rows(
    worksheet: Any,
    *,
    start_row: int,
    columns: tuple[int, ...],
) -> None:
    max_row = int(getattr(worksheet, "max_row", 0) or 0)
    for row_idx in range(max(1, int(start_row)), max_row + 1):
        for col_idx in columns:
            worksheet.cell(row=row_idx, column=int(col_idx)).value = None


def _merged_range_for_cell(worksheet: Any, *, row: int, column: int):
    coordinate = worksheet.cell(row=int(row), column=int(column)).coordinate
    for cell_range in getattr(getattr(worksheet, "merged_cells", None), "ranges", []) or []:
        if coordinate in cell_range:
            return cell_range
    return None


def _writable_cell(worksheet: Any, *, row: int, column: int):
    cell_range = _merged_range_for_cell(worksheet, row=row, column=column)
    if cell_range is not None:
        return worksheet.cell(row=cell_range.min_row, column=cell_range.min_col)
    return worksheet.cell(row=int(row), column=int(column))


def _find_label_value_cell(
    worksheet: Any,
    label: str,
    *,
    search_row_limit: int = 40,
    search_col_limit: int = 20,
):
    normalized_label = _normalize_template_text(label)
    max_row = min(max(1, int(getattr(worksheet, "max_row", 0) or 0)), max(1, int(search_row_limit)))
    max_col = min(max(1, int(getattr(worksheet, "max_column", 0) or 0)), max(1, int(search_col_limit)))

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            if _normalize_template_text(worksheet.cell(row=row_idx, column=col_idx).value) != normalized_label:
                continue
            value_col = col_idx + 1
            label_range = _merged_range_for_cell(worksheet, row=row_idx, column=col_idx)
            if label_range is not None and value_col <= int(label_range.max_col):
                value_col = int(label_range.max_col) + 1
            return _writable_cell(worksheet, row=row_idx, column=value_col)

    raise RuntimeError(f"Amazon 非美国站模板中未找到字段: {label}")


def _write_notus_default_owner_fields(worksheet: Any) -> None:
    _find_label_value_cell(worksheet, "Default prep owner").value = "Seller"
    _find_label_value_cell(worksheet, "Default labeling owner").value = "Seller"


def _write_us_template(
    worksheet: Any,
    rows: list[dict[str, Any]],
    *,
    header_row: int,
    header_columns: dict[str, int],
) -> int:
    sku_col = int(header_columns[_normalize_template_text("Merchant SKU")])
    quantity_col = int(header_columns[_normalize_template_text("Quantity")])
    current_row = int(header_row) + 1
    _clear_template_rows(worksheet, start_row=current_row, columns=(sku_col, quantity_col))
    for item in rows:
        worksheet.cell(row=current_row, column=sku_col, value=item["msku"])
        worksheet.cell(row=current_row, column=quantity_col, value=item["quantity"])
        current_row += 1
    return len(rows)


def _write_notus_template(
    worksheet: Any,
    rows: list[dict[str, Any]],
    *,
    header_row: int,
    header_columns: dict[str, int],
) -> int:
    sku_col = int(header_columns[_normalize_template_text("Merchant SKU")])
    quantity_col = int(header_columns[_normalize_template_text("Quantity")])
    prep_owner_col = _find_header_column(
        worksheet,
        header_row=header_row,
        header_columns=header_columns,
        header="Prep owner",
    )
    label_owner_col = _find_header_column(
        worksheet,
        header_row=header_row,
        header_columns=header_columns,
        header="Labeling owner",
    )
    clear_columns = tuple(
        dict.fromkeys(
            col
            for col in (sku_col, quantity_col, prep_owner_col, label_owner_col)
            if int(col or 0) > 0
        )
    )
    current_row = int(header_row) + 1
    _write_notus_default_owner_fields(worksheet)
    _clear_template_rows(worksheet, start_row=current_row, columns=clear_columns)
    for item in rows:
        worksheet.cell(row=current_row, column=sku_col, value=item["msku"])
        worksheet.cell(row=current_row, column=quantity_col, value=item["quantity"])
        current_row += 1
    return len(rows)


def _build_download_consignment_payload(consignment_no: str, excel_path: Path) -> dict[str, Any]:
    normalized = normalize_consignment_no(consignment_no)
    if not normalized:
        raise RuntimeError("consignment_no 不能为空")
    return {
        "consignment_no": normalized,
        "excel_path": str(excel_path),
        "source": _detect_consignment_source(excel_path),
        "message": "托运单 Excel 已准备好。",
    }


def _sync_fill_shipment_template(template_path: str, consignment_excel_path: str, site: str) -> dict[str, Any]:
    template_file = _resolve_any_path(template_path)
    consignment_file = _resolve_any_path(consignment_excel_path)
    normalized_site = _normalize_site(site)
    if not template_file.is_file():
        raise RuntimeError(f"找不到 Amazon 模板文件: {template_file}")
    if not consignment_file.is_file():
        raise RuntimeError(f"找不到托运单 Excel: {consignment_file}")

    grouped_df, sku_summary, total_quantity = _load_consignment_summary(consignment_file)

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入 Amazon 模板") from exc

    try:
        workbook = load_workbook(template_file)
    except Exception as exc:
        raise RuntimeError(f"读取 Amazon 模板失败: {exc}") from exc

    sheet_name, worksheet, header_row, header_columns = _select_template_sheet(
        workbook,
        normalized_site=normalized_site,
    )
    written_rows = (
        _write_us_template(
            worksheet,
            sku_summary,
            header_row=header_row,
            header_columns=header_columns,
        )
        if normalized_site == "US"
        else _write_notus_template(
            worksheet,
            sku_summary,
            header_row=header_row,
            header_columns=header_columns,
        )
    )

    saved_path = template_file
    try:
        workbook.save(template_file)
    except PermissionError:
        saved_path = _build_alt_template_path(template_file)
        workbook.save(saved_path)
    except Exception as exc:
        raise RuntimeError(f"写入 Amazon 模板失败: {exc}") from exc

    preview = sku_summary[:20]
    payload = {
        "site": normalized_site or str(site or "").strip(),
        "sheet_name": sheet_name,
        "filled_template_path": str(saved_path),
        "written_rows": int(written_rows),
        "total_quantity": int(total_quantity),
        "sku_summary": preview,
        "sku_summary_truncated": len(sku_summary) > len(preview),
        "message": "Amazon 模板已填写完成。",
    }
    return payload


def prepare_consignment_excel_payload(consignment_no: str) -> dict[str, Any]:
    normalized = normalize_consignment_no(consignment_no)
    if not normalized:
        raise RuntimeError("consignment_no 不能为空")
    try:
        excel_path = asyncio.run(ensure_consignment_excel_ready(normalized))
    except Exception as exc:
        raise RuntimeError(f"托运单数据准备失败: {exc}") from exc
    return _build_download_consignment_payload(normalized, excel_path)


def prepare_local_consignment_excel_payload(consignment_no: str) -> dict[str, Any]:
    normalized = normalize_consignment_no(consignment_no)
    if not normalized:
        raise RuntimeError("consignment_no 不能为空")
    try:
        excel_path = find_consignment_excel(normalized)
    except Exception as exc:
        raise RuntimeError(f"托运单数据准备失败: {exc}") from exc
    return _build_download_consignment_payload(normalized, excel_path)


def prepare_upload_local_consignment_excel_payload(consignment_no: str) -> dict[str, Any]:
    normalized = normalize_consignment_no(consignment_no)
    if not normalized:
        raise RuntimeError("consignment_no 不能为空")
    try:
        excel_path = find_consignment_excel(normalized)
    except Exception as exc:
        raise RuntimeError(f"托运单数据准备失败: {exc}") from exc
    return _build_download_consignment_payload(normalized, excel_path)


def fill_shipment_template_payload(
    template_path: str,
    consignment_excel_path: str,
    site: str,
) -> dict[str, Any]:
    try:
        return _sync_fill_shipment_template(template_path, consignment_excel_path, site)
    except Exception as exc:
        raise RuntimeError(f"填写 Amazon 模板失败: {exc}") from exc


__all__ = [
    "fill_shipment_template_payload",
    "prepare_consignment_excel_payload",
    "prepare_local_consignment_excel_payload",
    "prepare_upload_local_consignment_excel_payload",
]
