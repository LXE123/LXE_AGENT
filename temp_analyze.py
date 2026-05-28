import openpyxl

# 读取销量分析报告
wb1 = openpyxl.load_workbook('artifacts/mabang_store_msku_analysis/202605271823-Amazon-HSP-US_sales_analysis.xlsx')

print('=== 销量分析报告 - 父ASIN B0CHMFVC5R ===')

# MSKU明细
ws_msku = wb1['MSKU明细']
headers_msku = [c.value for c in ws_msku[1]]
print('\n--- MSKU明细 ---')
for row in ws_msku.iter_rows(min_row=2, values_only=True):
    if row[4] == 'B0CHMFVC5R':
        d = dict(zip(headers_msku, row))
        print(f"MSKU={d['MSKU']}, ASIN={d['ASIN']}, 本地SKU={d['本地SKU']}, 本地SKU名称={d['本地SKU名称']}")
        print(f"  销量: 7天={d['7天销量']}, 14天={d['14天销量']}, 30天={d['30天销量']}, 90天={d['90天销量']}, 日均={d['日均销量']}, 加权日销={d['加权日销']}, 趋势={d['销量趋势']}")
        print(f"  库存: 可售={d['可售']}, 待入库={d['待入库']}, 在途={d['在途']}, 预留={d['预留']}, 计划入库={d['计划入库']}, 采购在途={d['采购在途']}")
        print(f"  其他: 售价={d['售价']}, 重量={d['单品重量(g)(cm)']}, 上架时间={d['上架时间']}, 7天退货率={d['7天退货率']}, 30天退货率={d['30天退货率']}")

# 链接销量
ws_link = wb1['链接销量前10']
headers_link = [c.value for c in ws_link[1]]
print('\n--- 链接销量前10 ---')
for row in ws_link.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_link, row))
    if 'B0CHMFVC5R' in str(row):
        print(f"商品链接={d.get('商品链接', '')}, 父ASIN={d.get('父ASIN', '')}, 加权日销={d.get('加权日销', '')}, 销量趋势={d.get('销量趋势', '')}")

ws_link2 = wb1['其他链接']
headers_link2 = [c.value for c in ws_link2[1]]
print('\n--- 其他链接 ---')
for row in ws_link2.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_link2, row))
    if 'B0CHMFVC5R' in str(row):
        print(f"商品链接={d.get('商品链接', '')}, 父ASIN={d.get('父ASIN', '')}, 加权日销={d.get('加权日销', '')}, 销量趋势={d.get('销量趋势', '')}")

# ASIN销量
ws_asin = wb1['ASIN销量前50']
headers_asin = [c.value for c in ws_asin[1]]
print('\n--- ASIN销量前50 ---')
for row in ws_asin.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_asin, row))
    if 'B0CHMFVC5R' in str(row):
        print(f"ASIN={d.get('ASIN', '')}, 父ASIN={d.get('父ASIN', '')}, MSKU={d.get('MSKU', '')}, 加权日销={d.get('加权日销', '')}, 销量趋势={d.get('销量趋势', '')}")

ws_asin2 = wb1['其他ASIN']
headers_asin2 = [c.value for c in ws_asin2[1]]
print('\n--- 其他ASIN ---')
for row in ws_asin2.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_asin2, row))
    if 'B0CHMFVC5R' in str(row):
        print(f"ASIN={d.get('ASIN', '')}, 父ASIN={d.get('父ASIN', '')}, MSKU={d.get('MSKU', '')}, 加权日销={d.get('加权日销', '')}, 销量趋势={d.get('销量趋势', '')}")

# 真实库存报告
print('\n\n=== 真实库存报告 - 父ASIN B0CHMFVC5R ===')
wb2 = openpyxl.load_workbook('artifacts/mabang_store_msku_inventory/202605271823-Amazon-HSP-US_actual_inventory.xlsx')
for ws_name in wb2.sheetnames:
    ws = wb2[ws_name]
    headers = [c.value for c in ws[1]]
    print(f'\n--- {ws_name} (列: {headers}) ---')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'B0CHMFVC5R' in str(row):
            d = dict(zip(headers, row))
            print(d)

# 备货计算表
print('\n\n=== 备货计算表 - 父ASIN B0CHMFVC5R ===')
wb3 = openpyxl.load_workbook('artifacts/mabang_store_msku_replenishment/202605271823-Amazon-HSP-US_replenishment.xlsx')
for ws_name in wb3.sheetnames:
    ws = wb3[ws_name]
    headers = [c.value for c in ws[1]]
    print(f'\n--- {ws_name} ---')
    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'B0CHMFVC5R' in str(row):
            d = dict(zip(headers, row))
            print(d)