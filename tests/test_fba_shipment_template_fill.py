from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from agent_runtime.tools.fba_shipment_tools import fill_shipment_template_payload


def _write_consignment_excel(path: Path) -> None:
    pd.DataFrame(
        [
            {"MSKU": "SKU-B", "Quantity": 2},
            {"MSKU": "SKU-A", "Quantity": 3},
            {"MSKU": "SKU-B", "Quantity": 4},
        ]
    ).to_excel(path, index=False)


def _write_template(path: Path, *, include_default_owner: bool = True, include_owner_columns: bool = True) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Create workflow - template"

    if include_default_owner:
        worksheet.cell(row=2, column=1, value="Default prep owner")
        worksheet.cell(row=3, column=1, value="Default labeling owner")

    worksheet.cell(row=8, column=1, value="Merchant SKU")
    worksheet.cell(row=8, column=2, value="Quantity")
    if include_owner_columns:
        worksheet.cell(row=8, column=3, value="Prep owner")
        worksheet.cell(row=8, column=4, value="Labeling owner")
        worksheet.cell(row=9, column=3, value="stale-prep-owner")
        worksheet.cell(row=9, column=4, value="stale-label-owner")
        worksheet.cell(row=10, column=3, value="stale-prep-owner")
        worksheet.cell(row=10, column=4, value="stale-label-owner")
    worksheet.cell(row=8, column=5, value="Expiration date (MM/DD/YYYY)")
    workbook.save(path)


def test_notus_template_writes_default_owner_and_leaves_row_owner_blank(tmp_path: Path) -> None:
    consignment_path = tmp_path / "consignment.xlsx"
    template_path = tmp_path / "template.xlsx"
    _write_consignment_excel(consignment_path)
    _write_template(template_path)

    payload = fill_shipment_template_payload(str(template_path), str(consignment_path), "DE")

    assert payload["written_rows"] == 2
    workbook = load_workbook(template_path)
    worksheet = workbook["Create workflow - template"]
    assert worksheet.cell(row=2, column=2).value == "Seller"
    assert worksheet.cell(row=3, column=2).value == "Seller"
    assert worksheet.cell(row=9, column=1).value == "SKU-A"
    assert worksheet.cell(row=9, column=2).value == 3
    assert worksheet.cell(row=10, column=1).value == "SKU-B"
    assert worksheet.cell(row=10, column=2).value == 6
    assert worksheet.cell(row=9, column=3).value is None
    assert worksheet.cell(row=9, column=4).value is None
    assert worksheet.cell(row=10, column=3).value is None
    assert worksheet.cell(row=10, column=4).value is None


def test_us_template_behavior_stays_sku_and_quantity_only(tmp_path: Path) -> None:
    consignment_path = tmp_path / "consignment.xlsx"
    template_path = tmp_path / "template.xlsx"
    _write_consignment_excel(consignment_path)
    _write_template(template_path, include_default_owner=False, include_owner_columns=False)

    payload = fill_shipment_template_payload(str(template_path), str(consignment_path), "US")

    assert payload["written_rows"] == 2
    workbook = load_workbook(template_path)
    worksheet = workbook["Create workflow - template"]
    assert worksheet.cell(row=9, column=1).value == "SKU-A"
    assert worksheet.cell(row=9, column=2).value == 3
    assert worksheet.cell(row=10, column=1).value == "SKU-B"
    assert worksheet.cell(row=10, column=2).value == 6


def test_notus_template_requires_default_owner_fields(tmp_path: Path) -> None:
    consignment_path = tmp_path / "consignment.xlsx"
    template_path = tmp_path / "template.xlsx"
    _write_consignment_excel(consignment_path)
    _write_template(template_path, include_default_owner=False)

    with pytest.raises(RuntimeError, match="Default prep owner"):
        fill_shipment_template_payload(str(template_path), str(consignment_path), "DE")
