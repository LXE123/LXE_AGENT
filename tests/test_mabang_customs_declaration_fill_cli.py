from __future__ import annotations

import json
import csv
from decimal import Decimal
from pathlib import Path

import pytest

from services.agent_cli.mabang import fill_customs_declaration as cli


def _default_msku_for_sku(sku: object) -> str:
    text = str(sku or "").strip()
    return f"MSKU-{text}" if text else "MSKU-ACTUAL"


def _write_merge_detail_sheet(worksheet, rows: list[dict[str, object]]) -> None:
    worksheet.append(["SKU", "产品名称", "发货量", "规则型号", "单价"])
    for row in rows:
        sku = row.get("SKU", "")
        if not sku:
            continue
        worksheet.append(
            [
                sku,
                row.get("产品名称", row.get("品名", "")),
                row.get("发货量", ""),
                row.get("规则型号", row.get("规格型号", "")),
                row.get("单价", ""),
            ]
        )


def _write_default_delivery_csv_for_rows(path: Path, rows: list[dict[str, object]]) -> None:
    try:
        sp_no = cli.extract_sp_no_from_filename(path)
    except ValueError:
        return
    delivery_rows = [
        {
            cli.quantity_validation.DELIVERY_MSKU_COLUMN: _default_msku_for_sku(row.get("SKU")),
            cli.quantity_validation.SKU_SHIP_QTY_COLUMN: f"{row.get('SKU')} × 1",
        }
        for row in rows
        if row.get("SKU") and row.get("发货量") not in (None, "")
    ]
    if not delivery_rows:
        return
    delivery_path = path.parent / f"{sp_no}_1.csv"
    if delivery_path.is_file():
        return
    _write_delivery_csv(
        delivery_path,
        delivery_rows,
    )


def _sync_default_consignment_for_rows(path: Path, rows: list[dict[str, object]]) -> None:
    try:
        sp_no = cli.extract_sp_no_from_filename(path)
    except ValueError:
        return
    consignment_path = path.parent / f"{sp_no}.xlsx"
    if not consignment_path.is_file():
        return

    import pandas as pd

    existing = pd.read_excel(consignment_path, dtype=str)
    if "MSKU" in existing.columns and existing["MSKU"].fillna("").astype(str).str.strip().any():
        return
    box_rows: list[dict[str, object]] = []
    for record in existing.to_dict(orient="records"):
        box_rows.append(
            {
                "箱序号": record.get("箱序号", ""),
                "毛重": record.get("毛重", ""),
            }
        )
    if not box_rows:
        box_rows = [{"箱序号": 1, "毛重": 1}]

    actual_rows: list[dict[str, object]] = []
    first_box = box_rows[0]
    for row in rows:
        sku = row.get("SKU")
        quantity = row.get("发货量")
        if not sku or quantity in (None, ""):
            continue
        actual_rows.append(
            {
                "箱序号": first_box.get("箱序号", 1),
                "毛重": first_box.get("毛重", 1),
                "MSKU": _default_msku_for_sku(sku),
                "装箱数量": quantity,
            }
        )
    if not actual_rows:
        return
    pd.DataFrame(box_rows + actual_rows).to_excel(consignment_path, sheet_name="FBA装箱任务", index=False)


@pytest.fixture(autouse=True)
def _default_delivery_lookup(monkeypatch, tmp_path):
    def fake_find_latest_delivery_csv(sp_no: str, *, csv_dir=None) -> Path | None:
        path = tmp_path / f"{str(sp_no).upper()}_1.csv"
        return path if path.is_file() else None

    monkeypatch.setattr(cli.quantity_validation, "find_latest_delivery_csv", fake_find_latest_delivery_csv)


def _write_input_workbook(
    path: Path,
    rows: list[dict[str, object]],
    *,
    headers: tuple[str, ...] = cli.INPUT_HEADERS,
    sheet_count: int = 3,
    source_sheet_title: str = "第三个sheet",
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    if sheet_count >= 3:
        workbook.active.title = "可出口退税"
        _write_merge_detail_sheet(workbook.active, rows)
    while len(workbook.worksheets) < sheet_count:
        workbook.create_sheet(f"Sheet{len(workbook.worksheets) + 1}")
    worksheet = workbook.worksheets[2] if sheet_count >= 3 else workbook.worksheets[-1]
    worksheet.title = source_sheet_title
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    workbook.save(path)
    _write_default_delivery_csv_for_rows(path, rows)
    _sync_default_consignment_for_rows(path, rows)


def _write_input_workbook_with_extra_sheet_before_summary(path: Path, rows: list[dict[str, object]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "可出口退税"
    _write_merge_detail_sheet(workbook.active, rows)
    workbook.create_sheet("Sheet1")
    extra = workbook.create_sheet("Sheet3")
    extra.append(["库存SKU", "库存SKU中文名称", "库存SKU英文名称", "库存SKU"])
    summary = workbook.create_sheet(cli.SOURCE_WORKSHEET_NAME)
    summary.append(list(cli.INPUT_HEADERS))
    for row in rows:
        summary.append([row.get(header, "") for header in cli.INPUT_HEADERS])
    workbook.save(path)
    _write_default_delivery_csv_for_rows(path, rows)
    _sync_default_consignment_for_rows(path, rows)


def _write_input_workbook_with_expected_stock_sheet(path: Path, rows: list[dict[str, object]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    expected = workbook.active
    expected.title = "可出口退税"
    _write_merge_detail_sheet(expected, rows)
    workbook.create_sheet("Sheet2")
    summary = workbook.create_sheet(cli.SOURCE_WORKSHEET_NAME)
    summary.append(list(cli.INPUT_HEADERS))
    for row in rows:
        summary.append([row.get(header, "") for header in cli.INPUT_HEADERS])
    workbook.save(path)
    _write_default_delivery_csv_for_rows(path, rows)
    _sync_default_consignment_for_rows(path, rows)


def _write_input_workbook_with_expected_stock_and_summary_rows(
    path: Path,
    *,
    expected_rows: list[dict[str, object]],
    summary_rows: list[dict[str, object]],
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    expected = workbook.active
    expected.title = "可出口退税"
    expected.append(["SKU", "产品名称", "发货量", "规则型号", "单价"])
    for row in expected_rows:
        expected.append(
            [
                row.get("SKU", ""),
                row.get("产品名称", row.get("品名", "")),
                row.get("发货量", ""),
                row.get("规则型号", row.get("规格型号", "")),
                row.get("单价", ""),
            ]
        )
    workbook.create_sheet("Sheet2")
    summary = workbook.create_sheet(cli.SOURCE_WORKSHEET_NAME)
    summary.append(list(cli.INPUT_HEADERS))
    for row in summary_rows:
        summary.append([row.get(header, "") for header in cli.INPUT_HEADERS])
    workbook.save(path)
    _write_default_delivery_csv_for_rows(path, summary_rows)
    _sync_default_consignment_for_rows(path, summary_rows)


def _write_workbook_without_summary(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active.append(["不是", "汇总表"])
    workbook.create_sheet("Sheet2")
    workbook["Sheet2"].append(["也不是", "汇总表"])
    workbook.save(path)


def _write_customs_detail_template(worksheet, *, block_count: int = 3) -> None:
    worksheet.cell(row=19, column=1, value="项号")
    worksheet.cell(row=19, column=2, value="商品编号")
    worksheet.merge_cells(start_row=19, start_column=2, end_row=19, end_column=3)
    worksheet.cell(row=19, column=4, value="商品名称及规格型号")
    worksheet.merge_cells(start_row=19, start_column=4, end_row=19, end_column=6)
    worksheet.cell(row=19, column=7, value="数量及单位")
    worksheet.merge_cells(start_row=19, start_column=7, end_row=19, end_column=8)
    worksheet.cell(row=19, column=9, value="单价/总价/币制")
    worksheet.merge_cells(start_row=19, start_column=9, end_row=19, end_column=10)
    worksheet.cell(row=19, column=11, value="原产国（地区）")
    worksheet.merge_cells(start_row=19, start_column=11, end_row=19, end_column=12)
    worksheet.cell(row=19, column=13, value="最终目的国（地区）")
    worksheet.merge_cells(start_row=19, start_column=13, end_row=19, end_column=15)
    worksheet.cell(row=19, column=16, value="境内货源地")
    worksheet.merge_cells(start_row=19, start_column=16, end_row=19, end_column=18)
    worksheet.cell(row=19, column=19, value="征免")
    worksheet.cell(row=19, column=20, value="净重")
    worksheet.cell(row=19, column=21, value="毛重")
    worksheet.cell(row=19, column=22, value="件数")

    for offset in range(block_count):
        start_row = 20 + offset * 3
        worksheet.cell(row=start_row, column=1, value=offset + 1)
        worksheet.cell(row=start_row, column=2, value=f"=VLOOKUP(A{start_row},申报要素!A:F,3,0)")
        worksheet.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
        worksheet.cell(row=start_row, column=4, value=f"=VLOOKUP(A{start_row},申报要素!A:F,2,0)")
        worksheet.merge_cells(start_row=start_row, start_column=4, end_row=start_row, end_column=6)
        worksheet.cell(row=start_row + 1, column=4, value=f"=VLOOKUP(A{start_row},申报要素!A:F,6,0)")
        worksheet.merge_cells(start_row=start_row + 1, start_column=4, end_row=start_row + 2, end_column=6)
        worksheet.cell(row=start_row, column=9, value=99)
        worksheet.merge_cells(start_row=start_row, start_column=9, end_row=start_row, end_column=10)
        worksheet.cell(row=start_row + 1, column=9, value=f"=G{start_row + 2}*I{start_row}")
        worksheet.merge_cells(start_row=start_row + 1, start_column=9, end_row=start_row + 1, end_column=10)
        worksheet.cell(row=start_row + 2, column=9, value="人民币")
        worksheet.merge_cells(start_row=start_row + 2, start_column=9, end_row=start_row + 2, end_column=10)
        worksheet.cell(row=start_row, column=11, value="中国")
        worksheet.merge_cells(start_row=start_row, start_column=11, end_row=start_row, end_column=12)
        worksheet.cell(row=start_row, column=13, value="英国")
        worksheet.merge_cells(start_row=start_row, start_column=13, end_row=start_row, end_column=15)
        worksheet.cell(row=start_row, column=16, value="深圳特区")
        worksheet.merge_cells(start_row=start_row, start_column=16, end_row=start_row, end_column=18)
        worksheet.cell(row=start_row, column=19, value="照章征税")
        worksheet.cell(row=start_row, column=20, value=0.1)
        worksheet.merge_cells(start_row=start_row, start_column=20, end_row=start_row + 2, end_column=20)
        worksheet.cell(row=start_row, column=21, value=0.1)
        worksheet.merge_cells(start_row=start_row, start_column=21, end_row=start_row + 2, end_column=21)
        worksheet.cell(row=start_row + 2, column=7, value=999)
        worksheet.cell(row=start_row + 2, column=8, value="旧单位")
    if block_count > 0:
        worksheet.cell(row=20, column=22, value=1)
        worksheet.merge_cells(start_row=20, start_column=22, end_row=19 + block_count * 3, end_column=22)
    worksheet.print_area = f"A1:V{19 + block_count * 3}"


def _write_invoice_formula_template(worksheet, *, detail_rows: int = 3) -> None:
    worksheet.cell(row=7, column=1, value="標記號碼\nMark & No")
    worksheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    worksheet.cell(row=7, column=3, value="貨物名稱\nDescription")
    worksheet.cell(row=7, column=4, value="型号\nModel")
    worksheet.cell(row=7, column=5, value="數量\nQuantity")
    worksheet.merge_cells(start_row=7, start_column=5, end_row=7, end_column=6)
    worksheet.cell(row=7, column=7, value="單價\nUnit price")
    worksheet.cell(row=7, column=8, value="總金額\nAmount")
    worksheet.merge_cells(start_row=7, start_column=8, end_row=7, end_column=9)

    for offset in range(detail_rows):
        row_index = 8 + offset
        source_index = offset + 1
        worksheet.cell(row=row_index, column=3, value=f"=OFFSET(报关单!$D$1,ROW(报关单!D{source_index})*3+16,0)")
        worksheet.cell(row=row_index, column=4, value=f"=OFFSET(报关单!$D$1,ROW(报关单!E{source_index})*3+17,0)")
        worksheet.cell(row=row_index, column=5, value=f"=OFFSET(报关单!$G$1,ROW(报关单!G{source_index})*3+18,0)")
        worksheet.cell(row=row_index, column=6, value=f"=OFFSET(报关单!$H$1,ROW(报关单!H{source_index})*3+18,0)")
        worksheet.cell(row=row_index, column=7, value=f"=IFERROR(I{row_index}/E{row_index},0)")
        worksheet.cell(row=row_index, column=8, value=f"=OFFSET(报关单!$I$1,ROW(报关单!I{source_index})*3+18,0)")
        worksheet.cell(row=row_index, column=9, value=f"=OFFSET(报关单!$I$1,ROW(报关单!I{source_index})*3+17,0)")

    summary_row = 8 + detail_rows
    worksheet.cell(row=summary_row, column=3, value="旧中文金额")
    worksheet.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=6)
    worksheet.cell(row=summary_row, column=7, value="TOTAL:")
    worksheet.cell(row=summary_row, column=8, value="=H8")
    worksheet.cell(row=summary_row, column=9, value=f"=SUM(I8:I{summary_row - 1})")
    worksheet.print_area = f"A1:I{summary_row + 3}"


def _write_packing_formula_template(worksheet, *, detail_rows: int = 3) -> None:
    worksheet.cell(row=9, column=1, value="箱号\nCtn.No.")
    worksheet.cell(row=9, column=2, value="货物名称及规格\nDescription")
    worksheet.cell(row=9, column=3, value="型号\nModel")
    worksheet.cell(row=9, column=4, value="箱数：\nPkg：")
    worksheet.cell(row=9, column=5, value="数量：\nGe.Quantity")
    worksheet.merge_cells(start_row=9, start_column=5, end_row=9, end_column=6)
    worksheet.cell(row=9, column=7, value="毛重(KG)：\nG.W.(KG):")
    worksheet.cell(row=9, column=8, value="净重(KG)：\nN.W.(KG):")

    for offset in range(detail_rows):
        row_index = 10 + offset
        source_index = offset + 1
        invoice_row = 8 + offset
        worksheet.cell(row=row_index, column=1, value=f"=发票!A{invoice_row}")
        worksheet.cell(row=row_index, column=2, value=f"=发票!C{invoice_row}")
        worksheet.cell(row=row_index, column=3, value=f"=发票!D{invoice_row}")
        worksheet.cell(row=row_index, column=4, value=2)
        worksheet.cell(row=row_index, column=5, value=f"=发票!E{invoice_row}")
        worksheet.cell(row=row_index, column=6, value=f"=发票!F{invoice_row}")
        worksheet.cell(row=row_index, column=7, value=f"=OFFSET(报关单!$U$1,ROW(报关单!U{source_index})*3+16,0)")
        worksheet.cell(row=row_index, column=8, value=f"=OFFSET(报关单!$T$1,ROW(报关单!T{source_index})*3+16,0)")

    summary_row = 10 + detail_rows
    worksheet.cell(row=summary_row, column=1, value="合计\nTotal")
    worksheet.cell(row=summary_row, column=4, value=f"=SUM(D10:D{summary_row - 1})")
    worksheet.cell(row=summary_row, column=7, value=f"=SUM(G10:G{summary_row - 1})")
    worksheet.cell(row=summary_row, column=8, value=f"=SUM(H10:H{summary_row - 1})")
    worksheet.print_area = f"A1:H{summary_row + 2}"


def _write_contract_formula_template(worksheet, *, detail_rows: int = 3) -> None:
    worksheet.cell(row=17, column=2, value="Name of commodity")
    worksheet.cell(row=17, column=4, value="Model")
    worksheet.cell(row=17, column=5, value="Quantity")
    worksheet.cell(row=17, column=6, value="Unit")
    worksheet.cell(row=17, column=7, value="Unit Price")
    worksheet.cell(row=17, column=8, value="Amount")

    for offset in range(detail_rows):
        row_index = 18 + offset
        invoice_row = 8 + offset
        worksheet.cell(row=row_index, column=2, value=f"=发票!C{invoice_row}")
        worksheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=3)
        worksheet.cell(row=row_index, column=4, value=f"=发票!D{invoice_row}")
        worksheet.cell(row=row_index, column=5, value=f"=发票!E{invoice_row}")
        worksheet.cell(row=row_index, column=6, value=f"=发票!F{invoice_row}")
        worksheet.cell(row=row_index, column=7, value=f"=发票!G{invoice_row}")
        worksheet.cell(row=row_index, column=8, value=f"=发票!H{invoice_row}")
        worksheet.cell(row=row_index, column=9, value=f"=发票!I{invoice_row}")
        worksheet.merge_cells(start_row=row_index, start_column=9, end_row=row_index, end_column=11)

    summary_row = 18 + detail_rows
    worksheet.cell(row=summary_row, column=7, value="   总      值")
    worksheet.cell(row=summary_row, column=8, value="=H18")
    worksheet.merge_cells(start_row=summary_row, start_column=8, end_row=summary_row + 1, end_column=8)
    worksheet.cell(row=summary_row, column=9, value=f"=SUM(I18:K{summary_row - 1})")
    worksheet.merge_cells(start_row=summary_row, start_column=9, end_row=summary_row + 1, end_column=11)
    worksheet.cell(row=summary_row + 1, column=7, value="Total Amount:")
    worksheet.cell(row=summary_row + 2, column=5, value=f'=IF(H{summary_row}="CNY","总计人民币:",IF(H{summary_row}="USD","总计美元: ","  "))')
    worksheet.cell(row=summary_row + 2, column=6, value="人民币壹圆整")
    worksheet.merge_cells(start_row=summary_row + 2, start_column=6, end_row=summary_row + 2, end_column=9)
    worksheet.print_area = f"A1:K{summary_row + 5}"


def _write_formula_template_sheets(workbook, *, detail_rows: int = 3) -> None:
    invoice_sheet = workbook.create_sheet("发票")
    _write_invoice_formula_template(invoice_sheet, detail_rows=detail_rows)
    packing_sheet = workbook.create_sheet("箱单")
    _write_packing_formula_template(packing_sheet, detail_rows=detail_rows)
    contract_sheet = workbook.create_sheet("合同")
    _write_contract_formula_template(contract_sheet, detail_rows=detail_rows)


def _write_template(
    path: Path,
    *,
    old_row_count: int = 3,
    sheet_name: str = "申报要素",
    customs_detail_blocks: int = 3,
    formula_detail_rows: int = 3,
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "报关单"
    _write_customs_detail_template(workbook.active, block_count=customs_detail_blocks)
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.cell(row=8, column=1, value="申报要素")
    for column_index, header in enumerate(cli.TARGET_HEADERS, start=1):
        worksheet.cell(row=9, column=column_index, value=header)
    for offset in range(old_row_count):
        row_index = 10 + offset
        for column_index in range(1, len(cli.TARGET_HEADERS) + 1):
            worksheet.cell(row=row_index, column=column_index, value=f"旧数据{offset}-{column_index}")
    _write_formula_template_sheets(workbook, detail_rows=formula_detail_rows)
    workbook.save(path)


def _valid_source_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "日期": "2026-04-26",
        "SKU": "SKU-1",
        "品名": "米兰尼斯表带",
        "规格型号": "38MM",
        "发货量": 2,
        "单价": 1.5,
        "供货商": "供应商",
        "采购订单号": "PO-1",
        "采购总价": 3,
        "商品名称": "智能手表表带",
        "售价": 5,
        "总价": 10,
        "单位": "个",
    }
    row.update(overrides)
    if "SKU" in overrides and "规格型号" not in overrides and row.get("SKU") != "SKU-1":
        row["规格型号"] = str(row.get("SKU") or "")
        row["规则型号"] = row["规格型号"]
    return row


def _write_consignment_excel(path: Path, rows: list[dict[str, object]]) -> Path:
    import pandas as pd

    pd.DataFrame(rows).to_excel(path, sheet_name="FBA装箱任务", index=False)
    return path


def _write_delivery_csv(path: Path, rows: list[dict[str, object]], *, headers: list[str] | None = None) -> Path:
    if headers is None:
        headers = [
            cli.quantity_validation.DELIVERY_MSKU_COLUMN,
            cli.quantity_validation.SKU_SHIP_QTY_COLUMN,
        ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})
    return path


def _delivery_headers_with_msku_ship_quantity() -> list[str]:
    return [
        cli.quantity_validation.DELIVERY_MSKU_COLUMN,
        cli.quantity_validation.MSKU_SHIP_QTY_COLUMN,
        cli.quantity_validation.SKU_SHIP_QTY_COLUMN,
    ]


def _patch_delivery_lookup(monkeypatch, mapping: dict[str, Path | None]) -> None:
    def fake_find_latest_delivery_csv(sp_no: str, *, csv_dir=None) -> Path | None:
        return mapping.get(sp_no)

    monkeypatch.setattr(cli.quantity_validation, "find_latest_delivery_csv", fake_find_latest_delivery_csv)


def _read_validation_report_rows(path: str | Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["数量校验"]
    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]
    rows: list[dict[str, object]] = []
    for row_index in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row=row_index, column=column).value for column in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(dict(zip(headers, values)))
    workbook.close()
    return rows


def _read_summary_comparison_rows(path: str | Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook["汇总表计算前后对比"]
    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]
    rows: list[dict[str, object]] = []
    for row_index in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row=row_index, column=column).value for column in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(dict(zip(headers, values)))
    workbook.close()
    return rows


def _write_default_consignment_excel(tmp_path: Path, *, total_gross_weight: object = 10) -> Path:
    return _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "毛重": total_gross_weight},
        ],
    )


def _patch_consignment_lookup(monkeypatch, mapping: dict[str, Path]) -> None:
    def fake_find_consignment_excel(sp_no: str) -> Path:
        return mapping[sp_no]

    monkeypatch.setattr(cli, "find_consignment_excel", fake_find_consignment_excel)


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


def test_extract_sp_no_missing_fails():
    with pytest.raises(ValueError, match="文件名中缺少 SP 单号"):
        cli.extract_sp_no_from_filename("4.26-新棱镜备货.xlsx")


@pytest.mark.parametrize("country", ["日本", "澳大利亚", "德国", "英国", "美国", "加拿大"])
def test_extract_destination_country_from_filename_supports_known_countries(country):
    assert cli.extract_destination_country_from_filename(f"4.26-SP260414001-新棱镜备货-{country}（4.28）-2.xlsx") == country


def test_extract_destination_country_from_filename_requires_supported_country():
    with pytest.raises(ValueError, match="文件名中缺少目的国"):
        cli.extract_destination_country_from_filename("4.26-SP260414001-新棱镜备货-法国.xlsx")


def test_input_workbook_without_summary_sheet_or_headers_fails(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    _write_workbook_without_summary(input_path)

    with pytest.raises(ValueError, match="缺少 汇总表.*Sheet1.*Sheet2"):
        cli.read_source_rows(input_path)


def test_read_source_rows_accepts_single_named_summary_sheet(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-SUMMARY")],
        sheet_count=1,
        source_sheet_title=cli.SOURCE_WORKSHEET_NAME,
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].source_name == "米兰尼斯表带"


def test_read_source_rows_accepts_single_sheet_with_summary_headers(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-SUMMARY")],
        sheet_count=1,
        source_sheet_title="任意名称",
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].source_name == "米兰尼斯表带"


def test_read_source_rows_uses_third_sheet_when_it_is_summary(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-SUMMARY")],
        source_sheet_title=cli.SOURCE_WORKSHEET_NAME,
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].source_name == "米兰尼斯表带"


def test_read_source_rows_prefers_named_summary_sheet_over_third_sheet(tmp_path):
    input_path = tmp_path / "5.8-SP260424015-新棱镜备货-美国.xlsx"
    _write_input_workbook_with_extra_sheet_before_summary(
        input_path,
        [_valid_source_row(SKU="SKU-SUMMARY")],
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].source_name == "米兰尼斯表带"


def test_read_source_rows_does_not_require_supplier_or_purchase_order(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    assert "供货商" not in cli.INPUT_HEADERS
    assert "采购订单号" not in cli.INPUT_HEADERS
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-SUMMARY")],
        sheet_count=1,
        source_sheet_title=cli.SOURCE_WORKSHEET_NAME,
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].source_name == "米兰尼斯表带"


def test_read_source_rows_accepts_unordered_summary_headers(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    headers = (
        "单位",
        "总价",
        "售价",
        "商品名称",
        "采购总价",
        "单价",
        "发货量",
        "规格型号",
        "品名",
        "SKU",
        "日期",
    )
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-SUMMARY", 品名="硅胶表带", 规格型号="42MM", 发货量=7, 单价=2.5)],
        headers=headers,
        sheet_count=1,
        source_sheet_title=cli.SOURCE_WORKSHEET_NAME,
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].source_name == "硅胶表带"
    assert rows[0].model == "42MM"
    assert rows[0].quantity == 7
    assert rows[0].sale_price == 5
    assert rows[0].total_price == 10


def test_input_header_mismatch_fails(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    headers = ("日期", "SKU", "错误列") + cli.INPUT_HEADERS[3:]
    _write_input_workbook(
        input_path,
        [_valid_source_row()],
        headers=headers,
        source_sheet_title=cli.SOURCE_WORKSHEET_NAME,
    )

    with pytest.raises(ValueError, match="4.26-SP260414001-备货.xlsx.*汇总表.*缺少必需表头: 品名"):
        cli.read_source_rows(input_path)


def test_read_source_rows_rejects_duplicate_required_summary_header(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    headers = ("日期", "SKU", "SKU") + tuple(
        header for header in cli.INPUT_HEADERS if header not in {"日期", "SKU"}
    )
    _write_input_workbook(
        input_path,
        [_valid_source_row()],
        headers=headers,
        sheet_count=1,
        source_sheet_title=cli.SOURCE_WORKSHEET_NAME,
    )

    with pytest.raises(ValueError, match="表头重复: SKU"):
        cli.read_source_rows(input_path)


def test_read_source_rows_stops_when_date_is_empty(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-备货.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-1"),
            _valid_source_row(日期="", SKU="SKU-STOP"),
            _valid_source_row(SKU="SKU-SHOULD-NOT-READ"),
        ],
    )

    rows = cli.read_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].row_number == 2
    assert rows[0].source_name == "米兰尼斯表带"
    assert rows[0].model == "38MM"
    assert rows[0].commodity_name == "智能手表表带"
    assert rows[0].quantity == 2
    assert rows[0].unit == "个"


def test_read_consignment_weight_info_dedupes_by_box_sequence(tmp_path):
    excel_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "毛重": 12.5, "SKU": "A"},
            {"箱序号": 1, "毛重": 12.5, "SKU": "B"},
            {"箱序号": 3, "毛重": 7.3, "SKU": "C"},
        ],
    )

    payload = cli.read_consignment_weight_info(excel_path)

    assert payload.excel_path == str(excel_path.resolve())
    assert payload.box_count == 3
    assert payload.total_gross_weight == Decimal("19.8")


def test_read_consignment_weight_info_rejects_conflicting_gross_weight(tmp_path):
    excel_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "毛重": 12.5},
            {"箱序号": 1, "毛重": 12.6},
        ],
    )

    with pytest.raises(RuntimeError, match="同一箱序号存在不同毛重"):
        cli.read_consignment_weight_info(excel_path)


def test_read_consignment_weight_info_requires_columns(tmp_path):
    excel_path = _write_consignment_excel(tmp_path / "SP260414001.xlsx", [{"箱序号": 1}])

    with pytest.raises(RuntimeError, match="缺少必需列: 毛重"):
        cli.read_consignment_weight_info(excel_path)


def test_read_consignment_weight_info_rejects_invalid_gross_weight(tmp_path):
    excel_path = _write_consignment_excel(tmp_path / "SP260414001.xlsx", [{"箱序号": 1, "毛重": "abc"}])

    with pytest.raises(RuntimeError, match="毛重 无法解析为数字"):
        cli.read_consignment_weight_info(excel_path)


def test_resolve_consignment_excel_path_uses_local_lookup_when_not_specified(monkeypatch, tmp_path):
    excel_path = _write_default_consignment_excel(tmp_path)

    def fake_find_consignment_excel(sp_no: str) -> Path:
        assert sp_no == "SP260414001"
        return excel_path

    monkeypatch.setattr(cli, "find_consignment_excel", fake_find_consignment_excel)

    assert cli._resolve_consignment_excel_path("SP260414001") == excel_path.resolve()


def test_resolve_consignment_excel_path_fails_when_local_file_is_missing(monkeypatch):
    def fake_find_consignment_excel(sp_no: str) -> Path:
        raise FileNotFoundError(f"未找到托运单Excel: {sp_no}")

    monkeypatch.setattr(cli, "find_consignment_excel", fake_find_consignment_excel)

    with pytest.raises(FileNotFoundError, match="未找到托运单Excel: SP260414001"):
        cli._resolve_consignment_excel_path("SP260414001")


def test_allocate_weights_by_quantity_rounds_and_adjusts_largest_item():
    rows = [
        cli.SourceDeclarationRow(2, "", "", 1, "", 1, 1, "个"),
        cli.SourceDeclarationRow(3, "", "", 1, "", 1, 1, "个"),
        cli.SourceDeclarationRow(4, "", "", 1, "", 1, 1, "个"),
    ]

    allocation = cli.allocate_weights_by_quantity(rows, total_gross_weight=Decimal("1.0"))

    assert allocation.gross_weights == [Decimal("0.4"), Decimal("0.3"), Decimal("0.3")]
    assert allocation.net_weights == [Decimal("0.3"), Decimal("0.2"), Decimal("0.2")]


@pytest.mark.parametrize(
    ("gross_weight", "expected_net_weight"),
    [
        (Decimal("1.2"), Decimal("1.1")),
        (Decimal("0.1"), Decimal("0.09")),
        (Decimal("0.00"), Decimal("0.00")),
    ],
)
def test_calculate_net_weight_uses_fallback_when_needed(gross_weight, expected_net_weight):
    assert cli._calculate_net_weight(gross_weight) == expected_net_weight


@pytest.mark.parametrize(
    ("amount", "expected"),
    [
        (Decimal("20025"), "人民币贰万零贰拾伍圆整"),
        (Decimal("28882.155"), "人民币贰万捌仟捌佰捌拾贰圆壹角伍分伍厘"),
        (Decimal("0"), "人民币零圆整"),
        (Decimal("10.10"), "人民币壹拾圆壹角"),
    ],
)
def test_amount_to_chinese_upper_rmb(amount, expected):
    assert cli.amount_to_chinese_upper_rmb(amount) == expected


def test_calculate_total_amount_sums_total_price():
    rows = [
        cli.SourceDeclarationRow(2, "", "", 1, "", 1, "10.125", "个"),
        cli.SourceDeclarationRow(3, "", "", 1, "", 1, Decimal("20.005"), "个"),
    ]

    assert cli.calculate_total_amount(rows) == Decimal("30.130")


@pytest.mark.parametrize(
    ("total_price", "expected_message"),
    [
        ("", "第2行总价 不能为空"),
        ("abc", "第2行总价 无法解析为数字"),
        ("-1", "第2行总价 不能小于 0"),
    ],
)
def test_calculate_total_amount_rejects_invalid_total_price(total_price, expected_message):
    rows = [cli.SourceDeclarationRow(2, "", "", 1, "", 1, total_price, "个")]

    with pytest.raises(ValueError, match=expected_message):
        cli.calculate_total_amount(rows)


@pytest.mark.parametrize(
    ("commodity_name", "source_name", "expected_hs", "expected_element"),
    [
        ("智能手表表带", "米兰尼斯表带", "9113200000", "贱金属"),
        ("智能手表表带", "米兰表带", "9113200000", "贱金属"),
        ("智能手表表带", "编织表带", "9113900090", "尼龙"),
        ("智能手表表带", "尼龙表带", "9113900090", "尼龙"),
        ("编织表带", "未知品名", "9113900090", "尼龙"),
        ("尼龙表带", "未知品名", "9113900090", "尼龙"),
        ("智能手表表带", "真皮表带", "9113900090", "皮革"),
        ("智能手表表带", "尖尾皮带", "9113900090", "皮革"),
        ("智能手表表带", "硅胶表带", "9113900090", "硅胶"),
        ("智能手表表壳", "金属表壳", "9111800000", "贱金属"),
        ("智能手表表壳", "pc表壳", "9111800000", "PC"),
        ("智能手表表壳", "tpu表壳", "9111800000", "TPU"),
        ("智能手表表壳", "透明保护壳", "9111800000", "PC"),
        ("产品包装盒", "包装盒", "4819200000", "纸质+塑料"),
        ("手表保护套", "pc保护套", "3926909090", "PC"),
        ("手表保护套", "tpu保护套", "3926909090", "TPU"),
        ("手表保护套", "透明保护套", "3926909090", "PC"),
        ("其他产品", "其他品名", "", ""),
    ],
)
def test_classification_rules(commodity_name, source_name, expected_hs, expected_element):
    row = cli.SourceDeclarationRow(
        row_number=2,
        source_name=source_name,
        model="38MM",
        quantity=1,
        commodity_name=commodity_name,
        sale_price=1,
        total_price=1,
        unit="个",
    )

    result = cli.classify_declaration(row)

    assert result.hs_code == expected_hs
    assert result.declaration_element == expected_element


def test_fill_writes_template_copy_and_clears_old_rows(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path, [_valid_source_row()])
    _write_template(template_path, old_row_count=3)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    assert payload["success"] is True
    assert payload["sp_no"] == "SP260414001"
    assert payload["destination_country"] == "美国"
    assert payload["consignment_excel_path"] == str(consignment_path.resolve())
    assert payload["box_count"] == 1
    assert payload["total_gross_weight"] == 10
    assert payload["row_count"] == 1
    assert payload["unmatched_count"] == 0
    output_path = Path(payload["output_xlsx"])
    assert output_path == output_dir / "SP260414001_custom_declaration_documents.xlsx"

    output_workbook = load_workbook(output_path)
    output_sheet = output_workbook["申报要素"]
    assert [output_sheet.cell(row=10, column=column).value for column in range(1, 10)] == [
        1,
        "智能手表表带",
        "9113200000",
        None,
        "0无品牌",
        "38MM",
        "0无品牌",
        "0不享惠",
        "贱金属",
    ]
    assert [output_sheet.cell(row=11, column=column).value for column in range(1, 10)] == [
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ]

    template_workbook = load_workbook(template_path)
    template_sheet = template_workbook["申报要素"]
    assert template_sheet.cell(row=10, column=1).value == "旧数据0-1"


def test_fill_generates_quantity_validation_report_when_quantities_match(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-COMBO", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-COMBO",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1, SKU-B × 2",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 发货量=2, 总价=10),
            _valid_source_row(SKU="SKU-B", 发货量=4, 总价=20),
        ],
    )
    _write_template(template_path, customs_detail_blocks=2)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    assert payload["success"] is True
    assert payload["quantity_validation_status"] == "passed"
    assert payload["quantity_validation_summary"] == {
        "total_sku_count": 2,
        "matched_count": 2,
        "mismatch_count": 0,
        "unresolved_count": 0,
        "not_shipped_msku_count": 0,
    }
    assert Path(payload["validation_report_xlsx"]).is_file()
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert {row["SKU"]: row["状态"] for row in rows} == {"SKU-A": "一致", "SKU-B": "一致"}


def test_fill_quantity_validation_ignores_expected_stock_total_row(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    detail_row = _valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)
    total_row = _valid_source_row(SKU="", 产品名称="", 品名="", 规格型号="", 规则型号="", 发货量=3230, 单价="", 总价="")
    _write_input_workbook_with_expected_stock_and_summary_rows(
        input_path,
        expected_rows=[detail_row, total_row],
        summary_rows=[detail_row],
    )
    _write_template(template_path)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    assert payload["success"] is True
    assert payload["quantity_validation_status"] == "passed"
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert [row["SKU"] for row in rows] == ["SKU-A"]
    assert rows[0]["状态"] == "一致"


def test_fill_quantity_validation_still_reports_detail_row_missing_sku(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_and_summary_rows(
        input_path,
        expected_rows=[
            _valid_source_row(SKU="SKU-A", 发货量=2, 总价=10),
            _valid_source_row(SKU="", 产品名称="漏 SKU 产品", 品名="漏 SKU 产品", 规格型号="MISSING", 规则型号="MISSING", 发货量=1, 单价=1, 总价=1),
        ],
        summary_rows=[_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    assert payload["success"] is True
    assert payload["quantity_validation_status"] == "incomplete"
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert rows[0]["状态"] == "无法校验"
    assert "SKU 不能为空" in rows[0]["问题说明"]


def test_fill_keeps_customs_output_when_quantity_validation_mismatches(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 3, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    assert payload["success"] is True
    assert Path(payload["output_xlsx"]).is_file()
    assert payload["quantity_validation_status"] == "mismatch"
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert rows[0]["SKU"] == "SKU-A"
    assert rows[0]["预期发货量"] == 2
    assert rows[0]["实际发货量"] == 3
    assert rows[0]["差异"] == 1
    assert rows[0]["状态"] == "差异"


def test_fill_uses_actual_quantities_and_writes_summary_comparison_sheet(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
            {"箱序号": 1, "MSKU": "MSKU-A2", "装箱数量": 1, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 2,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A2",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 1,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-B",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 0,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-B × 0",
            },
        ],
        headers=_delivery_headers_with_msku_ship_quantity(),
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 发货量=5, 售价=5, 总价=25),
            _valid_source_row(SKU="SKU-B", 发货量=4, 售价=6, 总价=24),
        ],
    )
    _write_template(template_path, customs_detail_blocks=2)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    assert payload["success"] is True
    assert payload["quantity_basis"] == "actual"
    assert payload["row_count"] == 1
    assert payload["total_amount"] == 15
    assert payload["total_amount_upper"] == "人民币壹拾伍圆整"

    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    declaration_sheet = workbook["申报要素"]
    customs_sheet = workbook["报关单"]
    assert declaration_sheet.cell(row=10, column=1).value == 1
    assert declaration_sheet.cell(row=11, column=1).value is None
    assert customs_sheet.cell(row=22, column=7).value == 3
    assert customs_sheet.cell(row=23, column=1).value is None

    report_workbook = load_workbook(payload["validation_report_xlsx"], data_only=True)
    assert report_workbook.sheetnames[:3] == ["数量校验", "汇总表计算前后对比", "数据来源"]
    for sheet_name in report_workbook.sheetnames[:3]:
        sheet = report_workbook[sheet_name]
        for column_index in range(1, sheet.max_column + 1):
            assert sheet.column_dimensions[get_column_letter(column_index)].width == 15
        for row_index in range(1, sheet.max_row + 1):
            assert sheet.row_dimensions[row_index].height == 15
    assert report_workbook["数量校验"].cell(row=2, column=2).alignment.wrap_text is True
    assert report_workbook["汇总表计算前后对比"].cell(row=4, column=15).alignment.wrap_text is True
    report_workbook.close()

    validation_rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    validation_by_sku = {row["SKU"]: row for row in validation_rows if row["SKU"]}
    assert validation_by_sku["SKU-A"]["MSKU"] == "MSKU-A\nMSKU-A2"
    assert validation_by_sku["SKU-B"]["MSKU"] == "MSKU-B"

    comparison_rows = _read_summary_comparison_rows(payload["validation_report_xlsx"])
    detail_rows = [row for row in comparison_rows if row["状态"] != "合计"]
    total_row = next(row for row in comparison_rows if row["状态"] == "合计")
    by_sku = {row["汇总表SKU"]: row for row in detail_rows}
    assert by_sku["SKU-A"]["原发货量"] == 5
    assert by_sku["SKU-A"]["实际发货量"] == 3
    assert by_sku["SKU-A"]["数量差异"] == -2
    assert by_sku["SKU-A"]["原总价"] == 25
    assert by_sku["SKU-A"]["重算总价"] == 15
    assert by_sku["SKU-A"]["金额差异"] == -10
    assert by_sku["SKU-A"]["状态"] == "数量变化"
    assert by_sku["SKU-B"]["实际发货量"] == 0
    assert by_sku["SKU-B"]["状态"] == "未发货不写入"
    assert total_row["数量差异"] == -6
    assert total_row["金额差异"] == -34
    assert total_row["问题说明"] == "\n".join(
        [
            "SKU-A SKU-A 2",
            "SKU-B SKU-B 4",
            "留下来6条当库存",
            "期望发货数量9 实际发货数量3",
        ]
    )


def test_fill_maps_actual_quantities_by_summary_sku_representative_model(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-COMBO", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-COMBO",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A1 × 1, SKU-A2 × 2",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_and_summary_rows(
        input_path,
        expected_rows=[
            _valid_source_row(SKU="SKU-A1", 规格型号="MODEL-FIRST", 规则型号="MODEL-FIRST", 发货量=2, 单价=1.5, 总价=3),
            _valid_source_row(SKU="SKU-A2", 规格型号="MODEL-FIRST", 规则型号="MODEL-FIRST", 发货量=4, 单价=2, 总价=8),
        ],
        summary_rows=[
            _valid_source_row(SKU="SKU-A1", 规格型号="汇总表乱填型号", 发货量=6, 售价=4, 总价=24),
        ],
    )
    _write_template(template_path)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    assert payload["success"] is True
    assert payload["row_count"] == 1
    assert payload["total_amount"] == 24
    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    declaration_sheet = workbook["申报要素"]
    customs_sheet = workbook["报关单"]
    assert declaration_sheet.cell(row=10, column=6).value == "MODEL-FIRST"
    assert customs_sheet.cell(row=22, column=7).value == 6

    comparison_rows = _read_summary_comparison_rows(payload["validation_report_xlsx"])
    detail_rows = [row for row in comparison_rows if row["状态"] != "合计"]
    total_row = next(row for row in comparison_rows if row["状态"] == "合计")
    assert len(detail_rows) == 1
    assert detail_rows[0]["汇总表SKU"] == "SKU-A1"
    assert detail_rows[0]["规格型号"] == "MODEL-FIRST"
    assert detail_rows[0]["实际发货量"] == 6
    assert "汇总表 SKU 命中第一个表格型号组" in detail_rows[0]["问题说明"]
    assert total_row["数量差异"] == 0
    assert total_row["金额差异"] == 0


def test_fill_rejects_when_summary_sku_is_not_in_expected_stock_sheet(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook_with_expected_stock_and_summary_rows(
        input_path,
        expected_rows=[_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
        summary_rows=[_valid_source_row(SKU="SKU-MISSING", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    with pytest.raises(ValueError, match="汇总表第2行 SKU=SKU-MISSING 不在备货单第一个表格"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_rejects_duplicate_summary_representatives_for_same_model(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook_with_expected_stock_and_summary_rows(
        input_path,
        expected_rows=[
            _valid_source_row(SKU="SKU-A", 规格型号="MODEL-DUP", 规则型号="MODEL-DUP", 发货量=2, 总价=10),
            _valid_source_row(SKU="SKU-B", 规格型号="MODEL-DUP", 规则型号="MODEL-DUP", 发货量=3, 总价=15),
        ],
        summary_rows=[
            _valid_source_row(SKU="SKU-A", 发货量=2, 总价=10),
            _valid_source_row(SKU="SKU-B", 发货量=3, 总价=15),
        ],
    )
    _write_template(template_path)

    with pytest.raises(ValueError, match="汇总表同一个规则型号存在多个代表 SKU"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_rejects_actual_model_without_summary_representative(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-B", "装箱数量": 3, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-B",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-B × 1",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_and_summary_rows(
        input_path,
        expected_rows=[
            _valid_source_row(SKU="SKU-A", 规格型号="MODEL-A", 规则型号="MODEL-A", 发货量=2, 总价=10),
            _valid_source_row(SKU="SKU-B", 规格型号="MODEL-B", 规则型号="MODEL-B", 发货量=3, 总价=15),
        ],
        summary_rows=[_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    with pytest.raises(ValueError, match="型号组在汇总表中没有代表 SKU.*SKU=SKU-B.*规则型号=MODEL-B"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_rejects_when_delivery_csv_is_missing(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": None})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    with pytest.raises(FileNotFoundError, match="本地未找到发货单 CSV"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_quantity_validation_reports_unknown_wms_msku(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-WMS", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-CSV",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
        ],
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    with pytest.raises(ValueError, match="无法按实际发货量生成报关资料.*MSKU-WMS"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_quantity_validation_marks_zero_delivery_msku_as_not_shipped(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 2,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-ZERO",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 0,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-ZERO × 0",
            },
        ],
        headers=_delivery_headers_with_msku_ship_quantity(),
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    assert payload["success"] is True
    assert payload["quantity_validation_status"] == "passed"
    assert payload["quantity_validation_summary"] == {
        "total_sku_count": 1,
        "matched_count": 1,
        "mismatch_count": 0,
        "unresolved_count": 0,
        "not_shipped_msku_count": 1,
    }
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert any(row["SKU"] == "SKU-A" and row["状态"] == "一致" for row in rows)
    not_shipped_row = next(row for row in rows if row["状态"] == "未发货")
    assert not_shipped_row["MSKU"] == "MSKU-ZERO"
    assert "MSKU发货量 为 0" in not_shipped_row["问题说明"]
    assert not any(row["状态"] == "无法校验" and "MSKU-ZERO" in row["问题说明"] for row in rows)


def test_fill_quantity_validation_reports_positive_delivery_msku_missing_from_wms(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 2,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-MISSING",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 1,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-MISSING × 1",
            },
        ],
        headers=_delivery_headers_with_msku_ship_quantity(),
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_template(template_path)

    with pytest.raises(ValueError, match="无法按实际发货量生成报关资料.*MSKU-MISSING"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_quantity_validation_uses_wms_when_delivery_msku_quantity_is_zero(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 3, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    delivery_path = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.MSKU_SHIP_QTY_COLUMN: 0,
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
        ],
        headers=_delivery_headers_with_msku_ship_quantity(),
    )
    _patch_delivery_lookup(monkeypatch, {"SP260414001": delivery_path})
    _write_input_workbook_with_expected_stock_sheet(
        input_path,
        [_valid_source_row(SKU="SKU-A", 发货量=3, 总价=10)],
    )
    _write_template(template_path)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    assert payload["success"] is True
    assert payload["quantity_validation_status"] == "passed"
    assert payload["quantity_validation_summary"]["not_shipped_msku_count"] == 0
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert len(rows) == 1
    assert rows[0]["SKU"] == "SKU-A"
    assert rows[0]["实际发货量"] == 3
    assert rows[0]["状态"] == "一致"


def test_fill_rejects_invalid_total_price(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path, [_valid_source_row(总价="abc")])
    _write_template(template_path)

    with pytest.raises(ValueError, match="第2行总价 无法解析为数字"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_unmatched_row_is_written_blank_and_added_to_notice(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(
        input_path,
        [_valid_source_row(品名="未知品名", 商品名称="未知商品")],
    )
    _write_template(template_path, old_row_count=1)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    output_sheet = load_workbook(payload["output_xlsx"])["申报要素"]
    assert output_sheet.cell(row=10, column=3).value is None
    assert output_sheet.cell(row=10, column=9).value is None
    assert payload["unmatched_count"] == 1
    assert payload["notice"] == ["第2行未匹配申报规则: 商品名称=未知商品, 品名=未知品名"]


def test_find_customs_detail_layout_from_report_sheet(tmp_path):
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    _write_template(template_path, customs_detail_blocks=4)

    from openpyxl import load_workbook

    worksheet = load_workbook(template_path)["报关单"]
    layout = cli._find_customs_detail_layout(worksheet)

    assert layout.header_row == 19
    assert layout.item_no_col == 1
    assert layout.quantity_col == 7
    assert layout.unit_col == 8
    assert layout.price_col == 9
    assert layout.block_count == 4


def test_write_merged_safe_redirects_to_top_left(tmp_path):
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    _write_template(template_path, customs_detail_blocks=1)

    from openpyxl import load_workbook

    workbook = load_workbook(template_path)
    worksheet = workbook["报关单"]
    cli._write_merged_safe(worksheet, row=20, column=10, value=12.5)

    assert worksheet.cell(row=20, column=9).value == 12.5


def test_delete_rows_preserving_template_rejects_crossing_merge():
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试"
    worksheet.cell(row=2, column=1, value="跨界合并")
    worksheet.merge_cells(start_row=2, start_column=1, end_row=6, end_column=1)

    with pytest.raises(ValueError, match="合并单元格跨越删除区域"):
        cli._delete_rows_preserving_template(worksheet, start_row=3, amount=2)


def test_delete_rows_preserving_template_shifts_merged_ranges_and_formulas():
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试"
    worksheet.cell(row=6, column=1, value="下方合并")
    worksheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    worksheet.cell(row=7, column=3, value="=A6")

    cli._delete_rows_preserving_template(worksheet, start_row=3, amount=2)

    assert "A4:B4" in [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
    assert worksheet.cell(row=5, column=3).value == "=A4"


def test_delete_rows_preserving_template_shrinks_merge_ending_inside_deleted_rows():
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试"
    worksheet.cell(row=2, column=1, value="保留上半段")
    worksheet.merge_cells(start_row=2, start_column=1, end_row=5, end_column=2)

    cli._delete_rows_preserving_template(worksheet, start_row=4, amount=2)

    assert "A2:B3" in [str(merged_range) for merged_range in worksheet.merged_cells.ranges]


def test_delete_rows_preserving_template_updates_print_area_and_page_breaks():
    from openpyxl import Workbook
    from openpyxl.worksheet.pagebreak import Break

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试"
    worksheet.print_area = "A1:D20"
    worksheet.row_breaks.append(Break(id=5))
    worksheet.row_breaks.append(Break(id=10))
    worksheet.row_breaks.append(Break(id=15))
    worksheet.row_breaks.append(Break(id=20))
    for row_index in range(1, 21):
        worksheet.cell(row=row_index, column=1, value=row_index)

    cli._delete_rows_preserving_template(worksheet, start_row=8, amount=4)

    assert cli._print_area_bounds(worksheet) == (1, 1, 4, 16)
    assert [row_break.id for row_break in worksheet.row_breaks.brk] == [5, 11, 16]


def test_formula_translation_preserves_chinese_sheet_names():
    formula = "=OFFSET(报关单!$D$1,ROW(报关单!D1)*3+16,0)+发票!C8"

    translated = cli._translate_formula_row_references(formula, row_delta=2)

    assert translated == "=OFFSET(报关单!$D$1,ROW(报关单!D3)*3+16,0)+发票!C10"


def test_fill_writes_customs_detail_rows_and_preserves_formulas(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path, total_gross_weight=18)
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-1", 发货量=2, 售价=5, 单位="个"),
            _valid_source_row(SKU="SKU-2", 发货量=7, 售价=6.5, 单位="套"),
            _valid_source_row(SKU="SKU-3", 发货量=9, 售价=8, 单位="件"),
        ],
    )
    _write_template(template_path, customs_detail_blocks=3)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    worksheet = load_workbook(payload["output_xlsx"], data_only=False)["报关单"]
    assert payload["customs_detail_row_count"] == 3
    assert worksheet.cell(row=20, column=1).value == 1
    assert worksheet.cell(row=20, column=9).value == 5
    assert worksheet.cell(row=20, column=13).value == "美国"
    assert worksheet.cell(row=20, column=20).value == 1.9
    assert worksheet.cell(row=20, column=21).value == 2
    assert worksheet.cell(row=20, column=22).value == 1
    assert worksheet.cell(row=22, column=7).value == 2
    assert worksheet.cell(row=22, column=8).value == "个"
    assert worksheet.cell(row=23, column=1).value == 2
    assert worksheet.cell(row=23, column=9).value == 6.5
    assert worksheet.cell(row=23, column=13).value == "美国"
    assert worksheet.cell(row=23, column=20).value == 6.9
    assert worksheet.cell(row=23, column=21).value == 7
    assert worksheet.cell(row=25, column=7).value == 7
    assert worksheet.cell(row=25, column=8).value == "套"
    assert worksheet.cell(row=26, column=1).value == 3
    assert worksheet.cell(row=26, column=9).value == 8
    assert worksheet.cell(row=26, column=13).value == "美国"
    assert worksheet.cell(row=26, column=20).value == 8.9
    assert worksheet.cell(row=26, column=21).value == 9
    assert worksheet.cell(row=28, column=7).value == 9
    assert worksheet.cell(row=28, column=8).value == "件"

    assert worksheet.cell(row=20, column=2).value == "=VLOOKUP(A20,申报要素!A:F,3,0)"
    assert worksheet.cell(row=20, column=4).value == "=VLOOKUP(A20,申报要素!A:F,2,0)"
    assert worksheet.cell(row=21, column=4).value == "=VLOOKUP(A20,申报要素!A:F,6,0)"
    assert worksheet.cell(row=21, column=9).value == "=G22*I20"
    assert worksheet.cell(row=22, column=9).value == "人民币"


def test_fill_populates_formula_sheets_and_extends_formula_rows(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path, total_gross_weight=18)
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-1", 发货量=2, 售价=5, 单位="个"),
            _valid_source_row(SKU="SKU-2", 发货量=7, 售价=6.5, 单位="套"),
            _valid_source_row(SKU="SKU-3", 发货量=9, 售价=8, 单位="件"),
        ],
    )
    _write_template(template_path, customs_detail_blocks=3, formula_detail_rows=1)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    assert payload["formula_sheet_row_count"] == 3
    assert payload["formula_sheets"] == {"发票": 3, "箱单": 3, "合同": 3}
    assert payload["total_amount"] == 127.5
    assert payload["total_amount_upper"] == "人民币壹佰贰拾柒圆伍角"

    invoice = workbook["发票"]
    assert invoice.cell(row=8, column=3).value == "=OFFSET(报关单!$D$1,ROW(报关单!D1)*3+16,0)"
    assert invoice.cell(row=9, column=3).value == "=OFFSET(报关单!$D$1,ROW(报关单!D2)*3+16,0)"
    assert invoice.cell(row=10, column=7).value == "=IFERROR(I10/E10,0)"
    assert invoice.cell(row=11, column=3).value == "人民币壹佰贰拾柒圆伍角"
    assert invoice.cell(row=11, column=7).value == "TOTAL:"
    assert invoice.cell(row=11, column=9).value == "=SUM(I8:I10)"

    packing = workbook["箱单"]
    assert packing.cell(row=10, column=2).value == "=发票!C8"
    assert packing.cell(row=11, column=2).value == "=发票!C9"
    assert packing.cell(row=12, column=7).value == "=OFFSET(报关单!$U$1,ROW(报关单!U3)*3+16,0)"
    assert packing.cell(row=13, column=1).value == "合计\nTotal"
    assert packing.cell(row=13, column=7).value == "=SUM(G10:G12)"

    contract = workbook["合同"]
    assert contract.cell(row=18, column=2).value == "=发票!C8"
    assert contract.cell(row=20, column=2).value == "=发票!C10"
    assert contract.cell(row=20, column=9).value == "=发票!I10"
    assert "B20:C20" in [str(merged_range) for merged_range in contract.merged_cells.ranges]
    assert "I20:K20" in [str(merged_range) for merged_range in contract.merged_cells.ranges]
    assert contract.cell(row=21, column=9).value == "=SUM(I18:K20)"
    assert contract.cell(row=23, column=6).value == "人民币壹佰贰拾柒圆伍角"


def test_fill_keeps_and_clears_extra_customs_detail_blocks(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path, [_valid_source_row(发货量=2, 售价=5, 单位="个")])
    _write_template(template_path, customs_detail_blocks=3)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    worksheet = load_workbook(payload["output_xlsx"], data_only=False)["报关单"]
    assert worksheet.max_row == 28
    assert worksheet.cell(row=20, column=1).value == 1
    assert worksheet.cell(row=20, column=2).value == "=VLOOKUP(A20,申报要素!A:F,3,0)"
    assert worksheet.cell(row=20, column=4).value == "=VLOOKUP(A20,申报要素!A:F,2,0)"
    assert worksheet.cell(row=21, column=4).value == "=VLOOKUP(A20,申报要素!A:F,6,0)"
    assert worksheet.cell(row=21, column=9).value == "=G22*I20"
    assert worksheet.cell(row=22, column=9).value == "人民币"
    assert worksheet.cell(row=23, column=1).value is None
    assert worksheet.cell(row=23, column=9).value is None
    assert worksheet.cell(row=23, column=13).value is None
    assert worksheet.cell(row=23, column=20).value is None
    assert worksheet.cell(row=23, column=21).value is None
    assert worksheet.cell(row=25, column=7).value is None
    assert worksheet.cell(row=25, column=8).value is None
    assert worksheet.cell(row=26, column=1).value is None
    assert worksheet.cell(row=26, column=9).value is None
    assert worksheet.cell(row=26, column=13).value is None
    assert worksheet.cell(row=26, column=20).value is None
    assert worksheet.cell(row=26, column=21).value is None
    assert worksheet.cell(row=28, column=7).value is None
    assert worksheet.cell(row=28, column=8).value is None
    assert worksheet.cell(row=23, column=2).value is None
    assert worksheet.cell(row=23, column=4).value is None
    assert worksheet.cell(row=24, column=4).value is None
    assert worksheet.cell(row=24, column=9).value is None
    assert worksheet.cell(row=25, column=9).value is None
    assert worksheet.cell(row=23, column=11).value is None
    assert worksheet.cell(row=23, column=19).value is None


def test_fill_keeps_and_clears_extra_formula_rows(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path, [_valid_source_row(发货量=2, 售价=5, 单位="个")])
    _write_template(template_path, customs_detail_blocks=3, formula_detail_rows=3)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    invoice = workbook["发票"]
    assert invoice.cell(row=9, column=3).value is None
    assert invoice.cell(row=9, column=7).value is None
    assert invoice.cell(row=11, column=3).value == "人民币壹拾圆整"
    assert invoice.cell(row=11, column=7).value == "TOTAL:"
    assert invoice.cell(row=11, column=9).value == "=SUM(I8:I8)"

    packing = workbook["箱单"]
    assert packing.cell(row=11, column=1).value == "=发票!A9"
    assert packing.cell(row=11, column=2).value is None
    assert packing.cell(row=11, column=4).value == 2
    assert packing.cell(row=11, column=7).value is None
    assert packing.cell(row=13, column=7).value == "=SUM(G10:G10)"

    contract = workbook["合同"]
    assert contract.cell(row=19, column=2).value is None
    assert contract.cell(row=19, column=9).value is None
    assert contract.cell(row=21, column=9).value == "=SUM(I18:K18)"
    assert contract.cell(row=22, column=7).value == "Total Amount:"
    assert contract.cell(row=23, column=6).value == "人民币壹拾圆整"


def test_fill_deletes_extra_rows_beyond_kept_blanks(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path, [_valid_source_row(发货量=2, 售价=5, 单位="个")])
    _write_template(template_path, customs_detail_blocks=6, formula_detail_rows=8)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    customs_sheet = workbook["报关单"]
    assert customs_sheet.max_row == 28
    assert cli._print_area_bounds(customs_sheet) == (1, 1, 22, 28)
    assert customs_sheet.cell(row=23, column=1).value is None
    assert customs_sheet.cell(row=26, column=1).value is None

    invoice = workbook["发票"]
    assert cli._print_area_bounds(invoice) == (1, 1, 9, 17)
    assert invoice.cell(row=13, column=3).value is None
    assert invoice.cell(row=13, column=7).value is None
    assert invoice.cell(row=14, column=7).value == "TOTAL:"
    assert invoice.cell(row=14, column=9).value == "=SUM(I8:I8)"

    packing = workbook["箱单"]
    assert cli._print_area_bounds(packing) == (1, 1, 8, 18)
    assert packing.cell(row=15, column=2).value is None
    assert packing.cell(row=16, column=1).value == "合计\nTotal"
    assert packing.cell(row=16, column=7).value == "=SUM(G10:G10)"

    contract = workbook["合同"]
    assert cli._print_area_bounds(contract) == (1, 1, 11, 29)
    assert contract.cell(row=23, column=2).value is None
    assert contract.cell(row=24, column=7).value == "   总      值"
    assert contract.cell(row=24, column=9).value == "=SUM(I18:K18)"
    assert contract.cell(row=25, column=7).value == "Total Amount:"


def test_write_uppercase_amount_requires_invoice_total_marker(tmp_path):
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    _write_template(template_path)

    from openpyxl import load_workbook

    workbook = load_workbook(template_path)
    workbook["发票"].cell(row=11, column=7).value = None

    with pytest.raises(ValueError, match="发票 sheet 找不到单元格: TOTAL:"):
        cli._write_uppercase_amount(workbook, "人民币壹圆整")


def test_write_uppercase_amount_requires_contract_total_marker(tmp_path):
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    _write_template(template_path)

    from openpyxl import load_workbook

    workbook = load_workbook(template_path)
    workbook["合同"].cell(row=22, column=7).value = None

    with pytest.raises(ValueError, match="合同 sheet 找不到单元格: Total Amount:"):
        cli._write_uppercase_amount(workbook, "人民币壹圆整")


def test_fill_allows_fifty_product_rows(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path, total_gross_weight=50)
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU=f"SKU-{index}", 发货量=1, 总价=1)
            for index in range(1, 51)
        ],
    )
    _write_template(template_path, customs_detail_blocks=50, formula_detail_rows=50)

    payload = cli.fill_customs_declaration(
        input_path,
        template_xlsx=template_path,
        output_dir=output_dir,
        consignment_excel=consignment_path,
    )

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    assert payload["row_count"] == 50
    assert payload["customs_detail_row_count"] == 50
    assert payload["formula_sheets"] == {"发票": 50, "箱单": 50, "合同": 50}
    assert workbook["报关单"].cell(row=167, column=1).value == 50
    assert workbook["发票"].cell(row=58, column=7).value == "TOTAL:"
    assert workbook["箱单"].cell(row=60, column=1).value == "合计\nTotal"
    assert workbook["合同"].cell(row=68, column=9).value == "=SUM(I18:K67)"


def test_fill_rejects_more_than_fifty_product_rows(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path, total_gross_weight=51)
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU=f"SKU-{index}", 发货量=1, 总价=1)
            for index in range(1, 52)
        ],
    )
    _write_template(template_path, customs_detail_blocks=50, formula_detail_rows=50)

    with pytest.raises(ValueError, match="商品数超过报关资料模板容量"):
        cli.fill_customs_declaration(
            input_path,
            template_xlsx=template_path,
            output_dir=output_dir,
            consignment_excel=consignment_path,
        )


def test_fill_multiple_input_workbooks_combines_rows_without_cross_sp_weight_allocation(monkeypatch, tmp_path):
    input_path_1 = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    input_path_2 = tmp_path / "4.26-SP260422010-新棱镜备货-美国（4.28）-1.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path_1 = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [
            {"箱序号": 1, "毛重": 1.5},
            {"箱序号": 2, "毛重": 2.5},
        ],
    )
    consignment_path_2 = _write_consignment_excel(
        tmp_path / "SP260422010.xlsx",
        [
            {"箱序号": 1, "毛重": 4},
            {"箱序号": 2, "毛重": 3},
            {"箱序号": 3, "毛重": 3},
        ],
    )
    _patch_consignment_lookup(
        monkeypatch,
        {
            "SP260414001": consignment_path_1,
            "SP260422010": consignment_path_2,
        },
    )
    _write_input_workbook(
        input_path_1,
        [
            _valid_source_row(SKU="SKU-SP1-A", 品名="编织表带", 发货量=1, 售价=2, 总价=10, 单位="条"),
            _valid_source_row(SKU="SKU-SP1-B", 品名="硅胶表带", 发货量=3, 售价=3, 总价=10, 单位="条"),
        ],
    )
    _write_input_workbook(
        input_path_2,
        [
            _valid_source_row(SKU="SKU-SP2-A", 品名="PC保护套", 商品名称="手表保护套", 发货量=5, 售价=4, 总价=10.5, 单位="个"),
        ],
    )
    _write_template(template_path, customs_detail_blocks=3, formula_detail_rows=1)

    payload = cli.fill_customs_declaration(
        [input_path_1, input_path_2],
        template_xlsx=template_path,
        output_dir=output_dir,
    )

    from openpyxl import load_workbook

    assert payload["success"] is True
    assert payload["sp_nos"] == ["SP260414001", "SP260422010"]
    assert payload["sp_no"] == "SP260414001_SP260422010"
    assert payload["destination_country"] == "美国"
    assert payload["input_xlsx_paths"] == [str(input_path_1), str(input_path_2)]
    assert payload["consignment_excel_paths"] == {
        "SP260414001": str(consignment_path_1.resolve()),
        "SP260422010": str(consignment_path_2.resolve()),
    }
    assert payload["box_count"] == 5
    assert payload["total_gross_weight"] == 14
    assert payload["total_amount"] == 31
    assert payload["total_amount_upper"] == "人民币叁拾壹圆整"
    assert payload["row_count"] == 3
    assert payload["formula_sheets"] == {"发票": 3, "箱单": 3, "合同": 3}
    assert Path(payload["output_xlsx"]) == output_dir / "SP260414001_SP260422010_custom_declaration_documents.xlsx"

    workbook = load_workbook(payload["output_xlsx"], data_only=False)
    declaration_sheet = workbook["申报要素"]
    assert [declaration_sheet.cell(row=row, column=1).value for row in (10, 11, 12)] == [1, 2, 3]
    assert [declaration_sheet.cell(row=row, column=2).value for row in (10, 11, 12)] == [
        "智能手表表带",
        "智能手表表带",
        "手表保护套",
    ]

    customs_sheet = workbook["报关单"]
    assert customs_sheet.cell(row=20, column=1).value == 1
    assert customs_sheet.cell(row=20, column=21).value == 1
    assert customs_sheet.cell(row=20, column=20).value == 0.9
    assert customs_sheet.cell(row=20, column=22).value == 5
    assert customs_sheet.cell(row=22, column=7).value == 1
    assert customs_sheet.cell(row=23, column=1).value == 2
    assert customs_sheet.cell(row=23, column=21).value == 3
    assert customs_sheet.cell(row=23, column=20).value == 2.9
    assert customs_sheet.cell(row=25, column=7).value == 3
    assert customs_sheet.cell(row=26, column=1).value == 3
    assert customs_sheet.cell(row=26, column=21).value == 10
    assert customs_sheet.cell(row=26, column=20).value == 9.9
    assert customs_sheet.cell(row=28, column=7).value == 5


def test_fill_multiple_input_workbooks_writes_combined_quantity_validation_report(monkeypatch, tmp_path):
    input_path_1 = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    input_path_2 = tmp_path / "4.26-SP260422010-新棱镜备货-美国（4.28）-1.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path_1 = _write_consignment_excel(
        tmp_path / "SP260414001.xlsx",
        [{"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 1}],
    )
    consignment_path_2 = _write_consignment_excel(
        tmp_path / "SP260422010.xlsx",
        [{"箱序号": 1, "MSKU": "MSKU-B", "装箱数量": 3, "长": 40, "宽": 30, "高": 20, "毛重": 2}],
    )
    delivery_path_1 = _write_delivery_csv(
        tmp_path / "SP260414001_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-A",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-A × 1",
            },
        ],
    )
    delivery_path_2 = _write_delivery_csv(
        tmp_path / "SP260422010_1.csv",
        [
            {
                cli.quantity_validation.DELIVERY_MSKU_COLUMN: "MSKU-B",
                cli.quantity_validation.SKU_SHIP_QTY_COLUMN: "SKU-B × 1",
            },
        ],
    )
    _patch_consignment_lookup(
        monkeypatch,
        {
            "SP260414001": consignment_path_1,
            "SP260422010": consignment_path_2,
        },
    )
    _patch_delivery_lookup(
        monkeypatch,
        {
            "SP260414001": delivery_path_1,
            "SP260422010": delivery_path_2,
        },
    )
    _write_input_workbook_with_expected_stock_sheet(
        input_path_1,
        [_valid_source_row(SKU="SKU-A", 发货量=2, 总价=10)],
    )
    _write_input_workbook_with_expected_stock_sheet(
        input_path_2,
        [_valid_source_row(SKU="SKU-B", 发货量=3, 总价=10)],
    )
    _write_template(template_path, customs_detail_blocks=2)

    payload = cli.fill_customs_declaration(
        [input_path_1, input_path_2],
        template_xlsx=template_path,
        output_dir=output_dir,
    )

    assert payload["success"] is True
    assert payload["quantity_validation_status"] == "passed"
    assert payload["quantity_validation_summary"] == {
        "total_sku_count": 2,
        "matched_count": 2,
        "mismatch_count": 0,
        "unresolved_count": 0,
        "not_shipped_msku_count": 0,
    }
    assert Path(payload["validation_report_xlsx"]).name == "SP260414001_SP260422010_quantity_validation_report.xlsx"
    rows = _read_validation_report_rows(payload["validation_report_xlsx"])
    assert {(row["SP单号"], row["SKU"], row["状态"]) for row in rows} == {
        ("SP260414001", "SKU-A", "一致"),
        ("SP260422010", "SKU-B", "一致"),
    }


def test_fill_multiple_input_workbooks_requires_same_destination(monkeypatch, tmp_path):
    input_path_1 = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    input_path_2 = tmp_path / "4.26-SP260422010-新棱镜备货-英国（4.28）-1.xlsx"
    consignment_path_1 = _write_default_consignment_excel(tmp_path)
    consignment_path_2 = _write_consignment_excel(tmp_path / "SP260422010.xlsx", [{"箱序号": 1, "毛重": 1}])
    _patch_consignment_lookup(
        monkeypatch,
        {
            "SP260414001": consignment_path_1,
            "SP260422010": consignment_path_2,
        },
    )
    _write_input_workbook(input_path_1, [_valid_source_row()])
    _write_input_workbook(input_path_2, [_valid_source_row()])

    with pytest.raises(ValueError, match="多个备货单目的国不一致"):
        cli.fill_customs_declaration([input_path_1, input_path_2])


def test_fill_multiple_input_workbooks_rejects_single_consignment_override(tmp_path):
    input_path_1 = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    input_path_2 = tmp_path / "4.26-SP260422010-新棱镜备货-美国（4.28）-1.xlsx"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path_1, [_valid_source_row()])
    _write_input_workbook(input_path_2, [_valid_source_row()])

    with pytest.raises(ValueError, match="多备货单模式不支持 --consignment-excel"):
        cli.fill_customs_declaration([input_path_1, input_path_2], consignment_excel=consignment_path)


def test_main_accepts_repeated_input_xlsx(monkeypatch, tmp_path, capsys):
    input_path_1 = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    input_path_2 = tmp_path / "4.26-SP260422010-新棱镜备货-美国（4.28）-1.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path_1 = _write_consignment_excel(tmp_path / "SP260414001.xlsx", [{"箱序号": 1, "毛重": 1}])
    consignment_path_2 = _write_consignment_excel(tmp_path / "SP260422010.xlsx", [{"箱序号": 1, "毛重": 2}])
    _patch_consignment_lookup(
        monkeypatch,
        {
            "SP260414001": consignment_path_1,
            "SP260422010": consignment_path_2,
        },
    )
    _write_input_workbook(input_path_1, [_valid_source_row(SKU="SKU-1")])
    _write_input_workbook(input_path_2, [_valid_source_row(SKU="SKU-2")])
    _write_template(template_path, customs_detail_blocks=2)

    exit_code = cli.main(
        [
            "--input-xlsx",
            str(input_path_1),
            "--input-xlsx",
            str(input_path_2),
            "--template-xlsx",
            str(template_path),
            "--output-dir",
            str(output_dir),
        ]
    )

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["sp_nos"] == ["SP260414001", "SP260422010"]
    assert payload["row_count"] == 2
    assert payload["box_count"] == 2


def test_main_returns_failure_json_when_customs_detail_capacity_is_insufficient(tmp_path, capsys):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-1"),
            _valid_source_row(SKU="SKU-2"),
        ],
    )
    _write_template(template_path, customs_detail_blocks=1)

    exit_code = cli.main(
        [
            "--input-xlsx",
            str(input_path),
            "--template-xlsx",
            str(template_path),
            "--output-dir",
            str(output_dir),
            "--consignment-excel",
            str(consignment_path),
        ]
    )

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload["success"] is False
    assert "报关单明细区容量不足" in payload["exception"]


def test_main_success_outputs_json(tmp_path, capsys):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "custom_declaration_documents.xlsx"
    output_dir = tmp_path / "artifacts"
    consignment_path = _write_default_consignment_excel(tmp_path)
    _write_input_workbook(input_path, [_valid_source_row(品名="TPU保护壳", 商品名称="手表保护套")])
    _write_template(template_path)

    exit_code = cli.main(
        [
            "--input-xlsx",
            str(input_path),
            "--template-xlsx",
            str(template_path),
            "--output-dir",
            str(output_dir),
            "--consignment-excel",
            str(consignment_path),
        ]
    )

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["sp_no"] == "SP260414001"
    assert payload["destination_country"] == "美国"
    assert payload["box_count"] == 1
    assert payload["total_gross_weight"] == 10
    assert payload["source"] == "customs_declaration_fill"
    assert Path(payload["output_xlsx"]).is_file()


def test_main_failure_outputs_json(capsys):
    exit_code = cli.main(["--input-xlsx", "missing-sp.xlsx"])

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload["success"] is False
    assert "文件名中缺少 SP 单号" in payload["exception"]
