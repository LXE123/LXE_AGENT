from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from services.agent_cli.browser.amazon_fba.step2_excel_fill import _read_source_data, fill_multi_box_step2_template
from services.browser.workflows.amazon_fba_prepare_multi_box_excel import (
    _weight_capped_notice,
    extract_box_count_from_consignment_excel,
)


def _write_step2_consignment(
    path: Path,
    *,
    length: int = 30,
    width: int = 20,
    height: int = 10,
    weight: int = 30,
) -> None:
    columns = [
        "箱子编号",
        "箱序号",
        "MSKU",
        "FBA产品名称",
        "库存sku",
        "库存sku中文名称",
        "FNSKU",
        "装箱数量",
        "长",
        "宽",
        "高",
        "毛重",
    ]
    pd.DataFrame(
        [[1, 1, "SKU-1", "", "", "", "", 10, length, width, height, weight]],
        columns=columns,
    ).to_excel(path, sheet_name="FBA装箱任务", index=False)


def _write_step2_template(
    path: Path,
    *,
    length_label: str,
    width_label: str,
    height_label: str,
    weight_label: str,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Step2"

    worksheet.cell(row=1, column=1, value="SKU 总数: 1")
    worksheet.cell(row=2, column=1, value="包装箱总数:")
    worksheet.cell(row=2, column=2, value=1)
    worksheet.cell(row=4, column=1, value="SKU")
    worksheet.cell(row=4, column=4, value="包装箱 1 数量")
    worksheet.cell(row=5, column=1, value="SKU-1")
    worksheet.cell(row=8, column=3, value=length_label)
    worksheet.cell(row=9, column=3, value=width_label)
    worksheet.cell(row=10, column=3, value=height_label)
    worksheet.cell(row=11, column=3, value=weight_label)
    workbook.save(path)


def test_extract_box_count_uses_max_box_sequence_column(tmp_path: Path) -> None:
    excel_path = tmp_path / "consignment.xlsx"
    pd.DataFrame(
        [
            {"箱子编号": "", "箱序号": 1, "MSKU": "SKU-1", "装箱数量": 10},
            {"箱子编号": "", "箱序号": 1, "MSKU": "SKU-2", "装箱数量": 20},
            {"箱子编号": 2, "箱序号": 2, "MSKU": "SKU-1", "装箱数量": 10},
            {"箱子编号": 2, "箱序号": 2, "MSKU": "SKU-2", "装箱数量": 20},
        ]
    ).to_excel(excel_path, sheet_name="FBA装箱任务", index=False)

    assert extract_box_count_from_consignment_excel(excel_path) == 2


def test_extract_box_count_supports_legacy_box_number_column(tmp_path: Path) -> None:
    excel_path = tmp_path / "consignment.xlsx"
    pd.DataFrame(
        [
            {"箱编号": "", "箱号2": 1, "MSKU": "SKU-1", "装箱数量": 10},
            {"箱编号": 1, "箱号2": 1, "MSKU": "SKU-2", "装箱数量": 20},
            {"箱编号": "", "箱号2": 2, "MSKU": "SKU-1", "装箱数量": 10},
            {"箱编号": 2, "箱号2": 2, "MSKU": "SKU-2", "装箱数量": 20},
        ]
    ).to_excel(excel_path, sheet_name="FBA装箱任务", index=False)

    assert extract_box_count_from_consignment_excel(excel_path) == 2


def test_extract_box_count_requires_box_sequence_column(tmp_path: Path) -> None:
    excel_path = tmp_path / "consignment.xlsx"
    pd.DataFrame([{"箱子编号": 1, "MSKU": "SKU-1", "装箱数量": 10}]).to_excel(
        excel_path,
        sheet_name="FBA装箱任务",
        index=False,
    )

    with pytest.raises(RuntimeError, match="缺少 箱序号 列"):
        extract_box_count_from_consignment_excel(excel_path)


def test_step2_source_data_uses_box_sequence_column(tmp_path: Path) -> None:
    excel_path = tmp_path / "consignment.xlsx"
    columns = [
        "箱子编号",
        "箱序号",
        "MSKU",
        "FBA产品名称",
        "库存sku",
        "库存sku中文名称",
        "FNSKU",
        "装箱数量",
        "长",
        "宽",
        "高",
        "毛重",
    ]
    pd.DataFrame(
        [
            ["", 1, "SKU-1", "", "", "", "", 10, "", "", "", ""],
            [1, 1, "SKU-2", "", "", "", "", 20, 35, 35, 35, 14],
            [2, 2, "SKU-1", "", "", "", "", 30, 40, 40, 40, 15],
        ],
        columns=columns,
    ).to_excel(excel_path, sheet_name="FBA装箱任务", index=False)

    pivot_table, box_info = _read_source_data(excel_path)

    assert list(pivot_table.columns) == [1, 2]
    assert int(pivot_table.loc["SKU-1", 1]) == 10
    assert int(pivot_table.loc["SKU-1", 2]) == 30
    assert int(pivot_table.loc["SKU-2", 1]) == 20
    assert int(box_info.loc[1, "长"]) == 35
    assert int(box_info.loc[2, "毛重"]) == 15


def test_step2_template_uses_imperial_units_and_40_lb_cap(tmp_path: Path) -> None:
    consignment_path = tmp_path / "consignment.xlsx"
    template_path = tmp_path / "template.xlsx"
    _write_step2_consignment(consignment_path, length=30, width=20, height=10, weight=30)
    _write_step2_template(
        template_path,
        length_label="包装箱长度（英寸）：",
        width_label="包装箱宽度（英寸）：",
        height_label="包装箱高度（英寸）：",
        weight_label="包装箱重量（磅）：",
    )

    payload = fill_multi_box_step2_template(consignment_path, template_path)

    workbook = load_workbook(payload["filled_template_path"])
    worksheet = workbook["Step2"]
    assert worksheet.cell(row=5, column=4).value == 10
    assert worksheet.cell(row=8, column=4).value == 12
    assert worksheet.cell(row=9, column=4).value == 8
    assert worksheet.cell(row=10, column=4).value == 4
    assert worksheet.cell(row=11, column=4).value == 40
    assert payload["weight_capped_boxes"] == [
        {
            "box_no": 1,
            "weight_unit": "lb",
            "max_weight": 40,
            "original_weight": 66,
            "capped_weight": 40,
            "original_weight_lb": 66,
            "capped_weight_lb": 40,
        }
    ]
    assert "40 lb" in _weight_capped_notice(payload["weight_capped_boxes"])


def test_step2_template_uses_metric_units_and_23_kg_cap(tmp_path: Path) -> None:
    consignment_path = tmp_path / "consignment.xlsx"
    template_path = tmp_path / "template.xlsx"
    _write_step2_consignment(consignment_path, length=30, width=20, height=10, weight=30)
    _write_step2_template(
        template_path,
        length_label="包装箱长度（厘米）：",
        width_label="包装箱宽度（厘米）：",
        height_label="包装箱高度（厘米）：",
        weight_label="包装箱重量（千克）：",
    )

    payload = fill_multi_box_step2_template(consignment_path, template_path)

    workbook = load_workbook(payload["filled_template_path"])
    worksheet = workbook["Step2"]
    assert worksheet.cell(row=8, column=4).value == 30
    assert worksheet.cell(row=9, column=4).value == 20
    assert worksheet.cell(row=10, column=4).value == 10
    assert worksheet.cell(row=11, column=4).value == 23
    assert payload["weight_capped_boxes"] == [
        {
            "box_no": 1,
            "weight_unit": "kg",
            "max_weight": 23,
            "original_weight": 30,
            "capped_weight": 23,
            "original_weight_kg": 30,
            "capped_weight_kg": 23,
        }
    ]
    assert "23 kg" in _weight_capped_notice(payload["weight_capped_boxes"])


def test_step2_template_missing_units_defaults_to_source_units(tmp_path: Path) -> None:
    consignment_path = tmp_path / "consignment.xlsx"
    template_path = tmp_path / "template.xlsx"
    _write_step2_consignment(consignment_path, length=30, width=20, height=10, weight=20)
    _write_step2_template(
        template_path,
        length_label="包装箱长度：",
        width_label="包装箱宽度：",
        height_label="包装箱高度：",
        weight_label="包装箱重量：",
    )

    payload = fill_multi_box_step2_template(consignment_path, template_path)

    workbook = load_workbook(payload["filled_template_path"])
    worksheet = workbook["Step2"]
    assert worksheet.cell(row=8, column=4).value == 30
    assert worksheet.cell(row=9, column=4).value == 20
    assert worksheet.cell(row=10, column=4).value == 10
    assert worksheet.cell(row=11, column=4).value == 20
    assert payload["weight_capped_boxes"] == []
