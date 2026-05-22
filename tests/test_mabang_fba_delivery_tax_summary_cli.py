from __future__ import annotations

import asyncio
import json
import os
from collections import OrderedDict
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

import pytest

from services.agent_cli.mabang import summarize_fba_delivery_tax_sku as cli
from services.mabang.amazon.fba.batch_delivery import BatchDeliveryCsvResult


def _write_delivery_csv(path: Path, rows: list[str], *, include_column: bool = True) -> None:
    headers = ["发货单号"]
    if include_column:
        headers.append("SKU发货量")
    headers.append("备注")
    lines = [",".join(f'"{header}"' for header in headers)]
    for value in rows:
        fields = ["SP260508022"]
        if include_column:
            fields.append(value)
        fields.append("")
        lines.append(",".join(f'"{field}"' for field in fields))
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


def _write_tax_products(path: Path, rows: list[dict[str, str]], *, columns: list[str] | None = None) -> None:
    import pandas as pd

    if columns is None:
        columns = ["sku", "产品名称"]
    frame_rows = []
    for row in rows:
        frame_rows.append({column: row.get(column, "") for column in columns})
    pd.DataFrame(frame_rows, columns=columns).to_excel(path, sheet_name="Sheet1", index=False)


async def _noop_close_all_network_clients() -> None:
    return None


def test_missing_delivery_no_returns_failure_json(monkeypatch, capsys):
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main([])

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload == {
        "success": False,
        "delivery_no": "",
        "exception": "delivery_no 不能为空",
    }


def test_invalid_delivery_no_returns_failure_json(monkeypatch, capsys):
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main(["--delivery-no", "FBA123"])

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload == {
        "success": False,
        "delivery_no": "FBA123",
        "exception": "delivery_no 格式无效: FBA123",
    }


def test_find_latest_delivery_csv_picks_newest_file(tmp_path):
    older = tmp_path / "SP260508022_1.csv"
    newer = tmp_path / "SP260508022_2.csv"
    _write_delivery_csv(older, ["SKU-A × 1"])
    _write_delivery_csv(newer, ["SKU-B × 1"])
    os.utime(older, (100, 100))
    os.utime(newer, (200, 200))

    assert cli.find_latest_delivery_csv("sp260508022", csv_dir=tmp_path) == newer


def test_resolve_delivery_csv_downloads_when_local_missing(monkeypatch, tmp_path):
    downloaded = tmp_path / "downloaded.csv"
    _write_delivery_csv(downloaded, ["SKU-A × 1"])

    async def fake_download(delivery_no: str):
        assert delivery_no == "SP260508022"
        return BatchDeliveryCsvResult(
            delivery_no=delivery_no,
            delivery_id=1,
            task_id=2,
            file_hash="hash",
            file_name="downloaded.csv",
            csv_path=str(downloaded),
        )

    monkeypatch.setattr(cli, "download_fba_delivery_csv", fake_download)

    assert asyncio.run(cli.resolve_delivery_csv("SP260508022", csv_dir=tmp_path / "missing")) == downloaded


def test_missing_sku_column_raises(tmp_path):
    csv_path = tmp_path / "SP260508022_1.csv"
    _write_delivery_csv(csv_path, ["SKU-A × 1"], include_column=False)

    with pytest.raises(RuntimeError, match="缺少列: SKU发货量"):
        cli.summarize_tax_sku_quantities(csv_path)


def test_summarize_tax_sku_quantities_parses_and_sums_rows(tmp_path):
    csv_path = tmp_path / "SP260508022_1.csv"
    _write_delivery_csv(
        csv_path,
        [
            "DP230828103 × 2，DP230828106 × 2",
            "DP230828103 × 1, DP230828106 x 3",
            "DP230828107 X 4；DP230828108 * 5",
        ],
    )

    summary = cli.summarize_tax_sku_quantities(csv_path)

    assert summary == {
        "DP230828103": 3,
        "DP230828106": 5,
        "DP230828107": 4,
        "DP230828108": 5,
    }


def test_load_export_tax_products_reads_sheet1_columns(tmp_path):
    products_path = tmp_path / "products.xlsx"
    _write_tax_products(
        products_path,
        [
            {"sku": " DP230828103 ", "产品名称": "产品A"},
            {"sku": "DP230828106", "产品名称": "产品B"},
            {"sku": "DP230828103", "产品名称": "重复产品A"},
        ],
    )

    products = cli.load_export_tax_products(products_path)

    assert products == {
        "DP230828103": {"sku": "DP230828103", "product_name": "产品A"},
        "DP230828106": {"sku": "DP230828106", "product_name": "产品B"},
    }


def test_export_tax_products_default_path_is_xlsx():
    assert cli.EXPORT_TAX_PRODUCTS_PATH.name == "export_tax_products.xlsx"


def test_load_export_tax_products_missing_file_raises(tmp_path):
    with pytest.raises(RuntimeError, match="找不到出口退税产品表"):
        cli.load_export_tax_products(tmp_path / "missing.xlsx")


def test_load_export_tax_products_missing_columns_raises(tmp_path):
    products_path = tmp_path / "products.xlsx"
    _write_tax_products(products_path, [{"sku": "DP230828103"}], columns=["sku"])

    with pytest.raises(RuntimeError, match="缺少列: 产品名称"):
        cli.load_export_tax_products(products_path)


def test_split_tax_sku_summary_routes_matched_and_unmatched():
    summary = OrderedDict(
        [
            ("DP230828103", Decimal("3")),
            (" DP230828106 ", Decimal("2")),
            ("DP230828108", Decimal("5")),
        ]
    )
    products = OrderedDict(
        [
            ("DP230828103", {"sku": "DP230828103", "product_name": "产品A"}),
            ("DP230828106", {"sku": "DP230828106", "product_name": "产品B"}),
        ]
    )

    matched_rows, unmatched_rows = cli.split_tax_sku_summary(summary, products)

    assert matched_rows == [
        ["DP230828103", "产品A", 3],
        ["DP230828106", "产品B", 2],
    ]
    assert unmatched_rows == [["DP230828108", "", 5]]


def test_fill_unmatched_product_names_writes_names_and_counts_missing():
    rows = [["DP230828108", "", 5], ["DP230828109", "", 2]]

    filled, matched_count, missing_count = cli.fill_unmatched_product_names(
        rows,
        {"DP230828108": "产品C"},
    )

    assert filled == [["DP230828108", "产品C", 5], ["DP230828109", "", 2]]
    assert matched_count == 1
    assert missing_count == 1


def test_write_summary_xlsx_uses_two_fixed_sheets(tmp_path):
    csv_path = tmp_path / "SP260508022_1.csv"
    _write_delivery_csv(csv_path, ["DP230828103 × 2，DP230828106 × 2", "DP230828103 × 1"])
    summary = cli.summarize_tax_sku_quantities(csv_path)
    products = OrderedDict(
        [
            ("DP230828103", {"sku": "DP230828103", "product_name": "产品A"}),
        ]
    )
    matched_rows, unmatched_rows = cli.split_tax_sku_summary(summary, products)

    xlsx_path = cli.write_summary_xlsx(
        matched_rows,
        unmatched_rows,
        delivery_no="SP260508022",
        output_dir=tmp_path,
    )

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path)
    assert workbook.sheetnames == ["可出口退税", "不可出口退税"]
    matched_sheet = workbook["可出口退税"]
    unmatched_sheet = workbook["不可出口退税"]
    assert list(matched_sheet.iter_rows(values_only=True)) == [
        ("SKU", "产品名称", "发货量"),
        ("DP230828103", "产品A", 3),
    ]
    assert list(unmatched_sheet.iter_rows(values_only=True)) == [
        ("SKU", "产品名称", "发货量"),
        ("DP230828106", None, 2),
    ]


def test_write_summary_xlsx_keeps_empty_sheet_headers(tmp_path):
    xlsx_path = cli.write_summary_xlsx([], [["DP230828106", "产品B", 2]], delivery_no="SP260508022", output_dir=tmp_path)

    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path)
    assert list(workbook["可出口退税"].iter_rows(values_only=True)) == [
        ("SKU", "产品名称", "发货量"),
    ]
    assert list(workbook["不可出口退税"].iter_rows(values_only=True)) == [
        ("SKU", "产品名称", "发货量"),
        ("DP230828106", "产品B", 2),
    ]


def test_main_success_generates_xlsx_with_unmatched_product_names(monkeypatch, tmp_path, capsys):
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)
    csv_dir = tmp_path / "csv"
    output_dir = tmp_path / "out"
    csv_dir.mkdir()
    _write_delivery_csv(
        csv_dir / "SP260508022_1.csv",
        ["DP230828103 × 2，DP230828106 × 2", "DP230828103 × 1"],
    )
    products_path = tmp_path / "products.xlsx"
    _write_tax_products(products_path, [{"sku": "DP230828103", "产品名称": "产品A"}])
    monkeypatch.setattr(cli, "DELIVERY_CSV_DIR", csv_dir)
    monkeypatch.setattr(cli, "OUTPUT_DIR", output_dir)
    monkeypatch.setattr(cli, "EXPORT_TAX_PRODUCTS_PATH", products_path)

    async def fake_export_stock_sku_names(skus, *, delivery_no="", output_dir=None, **kwargs):
        assert skus == ["DP230828106"]
        assert delivery_no == "SP260508022"
        assert output_dir == cli.STOCK_SKU_OUTPUT_DIR
        return SimpleNamespace(
            names_by_key=OrderedDict([("DP230828106", "产品B")]),
            xlsx_paths=[str(tmp_path / "stock.xlsx")],
        )

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    exit_code = cli.main(["--delivery-no", "SP260508022"])

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["delivery_no"] == "SP260508022"
    assert payload["sku_count"] == 2
    assert payload["matched_sku_count"] == 1
    assert payload["unmatched_sku_count"] == 1
    assert payload["stock_sku_xlsx_paths"] == [str(tmp_path / "stock.xlsx")]
    assert payload["stock_name_matched_count"] == 1
    assert payload["stock_name_missing_count"] == 0
    assert payload["source"] == "fba_delivery_tax_summary"
    assert Path(payload["xlsx_path"]).is_file()


def test_summarize_delivery_tax_sku_skips_stock_export_when_no_unmatched(monkeypatch, tmp_path):
    csv_dir = tmp_path / "csv"
    output_dir = tmp_path / "out"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["DP230828103 × 2"])
    products_path = tmp_path / "products.xlsx"
    _write_tax_products(products_path, [{"sku": "DP230828103", "产品名称": "产品A"}])

    async def fail_export_stock_sku_names(*args, **kwargs):
        raise AssertionError("stock export should not be called")

    monkeypatch.setattr(cli, "export_stock_sku_names", fail_export_stock_sku_names)

    payload = asyncio.run(
        cli.summarize_delivery_tax_sku(
            "SP260508022",
            csv_dir=csv_dir,
            output_dir=output_dir,
            products_path=products_path,
        )
    )

    assert payload["unmatched_sku_count"] == 0
    assert payload["stock_sku_xlsx_paths"] == []
    assert payload["stock_name_matched_count"] == 0
    assert payload["stock_name_missing_count"] == 0
