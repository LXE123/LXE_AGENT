from __future__ import annotations

import json
from pathlib import Path

import pytest

from services.mabang.amazon.fba import amazon_fba_inventory as amazon_inv
from services.mabang.amazon.fba import amazon_restock_inventory as restock_inv
from services.mabang.amazon.fba import store_msku_replenishment as repl
from services.mabang.amazon.fba import replenishment_template as tmpl


def _write_workbook(path: Path, sheets: dict[str, tuple[list[str], list[dict]]]) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        for index, (sheet_name, (headers, rows)) in enumerate(sheets.items()):
            worksheet = workbook.active if index == 0 else workbook.create_sheet()
            worksheet.title = sheet_name
            worksheet.append(headers)
            for row in rows:
                worksheet.append([row.get(header, "") for header in headers])
        workbook.save(path)
    finally:
        workbook.close()
    return path


def _load_records(path: Path, sheet_name: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        return [dict(zip(headers, values, strict=False)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
    finally:
        workbook.close()


def _headers(path: Path, sheet_name: str) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        return [cell.value for cell in worksheet[1]]
    finally:
        workbook.close()


def _column_number_formats(path: Path, sheet_name: str, header: str) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        column_index = headers.index(header) + 1
        return [
            worksheet.cell(row=row_index, column=column_index).number_format
            for row_index in range(2, worksheet.max_row + 1)
        ]
    finally:
        workbook.close()


def _assert_standard_dimensions(path: Path, sheet_names: list[str]) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    try:
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            assert worksheet.sheet_format.defaultRowHeight == 15
            for row_index in range(1, worksheet.max_row + 1):
                assert worksheet.row_dimensions[row_index].height == 15
            for column_index in range(1, worksheet.max_column + 1):
                assert worksheet.column_dimensions[get_column_letter(column_index)].width == 15
    finally:
        workbook.close()


def _sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _inventory_headers() -> list[str]:
    return ["MSKU", "父ASIN", "ASIN", "本地SKU", "本地SKU名称", "产品名称", "备注", "商品链接", "FBA总库存", "加权日销", "可销售天数", "真实库存数量", "子SKU"]


def _sales_headers() -> list[str]:
    return ["MSKU", "父ASIN", "ASIN", "本地SKU", "7天销量", "14天销量", "30天销量", "销量趋势速率", "销量趋势", "单品重量(g)(cm)"]


def _replenishment_row(
    msku: str,
    sheet_name: str,
    *,
    replenish_quantity: int | None,
    actual_inventory: float | None,
) -> repl.ReplenishmentRow:
    return repl.ReplenishmentRow(
        msku=msku,
        parent_asin=f"PARENT-{msku}",
        asin=f"ASIN-{msku}",
        local_sku=f"SKU-{msku}",
        local_sku_name=f"本地名-{msku}",
        product_name=f"产品名-{msku}",
        remark="",
        product_link=f"https://example.test/{msku}",
        sku_type="库存sku",
        template_name="默认模板",
        matched_rule="默认规则",
        sales_trend="平稳",
        trend_group="平稳",
        weighted_daily_sales=1,
        sales_days=10,
        fba_total_inventory=10,
        unlinked_quantity=0,
        actual_inventory=actual_inventory,
        weight_grams=100,
        replenish_days=10 if replenish_quantity is not None else None,
        replenish_quantity=replenish_quantity,
        original_replenish_quantity=replenish_quantity,
        sea_days=None,
        sea_quantity=None,
        estimated_weight_kg=None,
        decision_reason="测试",
        child_skus="",
        sheet_name=sheet_name,
    )


def _inventory_input_row(msku: str, *, fba_total_inventory: float) -> repl.InventoryInputRow:
    return repl.InventoryInputRow(
        msku=msku,
        parent_asin=f"PARENT-{msku}",
        asin=f"ASIN-{msku}",
        local_sku=f"SKU-{msku}",
        local_sku_name=f"本地名-{msku}",
        product_name=f"产品名-{msku}",
        remark="",
        product_link=f"https://example.test/{msku}",
        sku_type="库存sku",
        weighted_daily_sales=0,
        sales_days=None,
        fba_total_inventory=fba_total_inventory,
        actual_inventory=999,
        child_skus="",
    )


def _sales_detail_for_daily_sales(daily_sales: float, *, weight_grams: float) -> repl.SalesDetail:
    return repl.SalesDetail(
        trend="平稳",
        weight_grams=weight_grams,
        sales_7d=daily_sales * 7,
        sales_14d=daily_sales * 14,
        sales_30d=daily_sales * 30,
    )


def _write_inventory_report(path: Path, *, remarks: dict[str, str] | None = None) -> Path:
    headers = _inventory_headers()
    remarks = dict(remarks or {})
    return _write_workbook(
        path,
        {
            "真实库存-组合sku": (
                headers,
                [
                    {
                        "MSKU": "SEA-1",
                        "父ASIN": "PARENT-SEA",
                        "ASIN": "ASIN-SEA",
                        "本地SKU": "COMBO-SEA",
                        "本地SKU名称": "海运组合本地名",
                        "产品名称": "Sea Combo Product",
                        "备注": remarks.get("SEA-1", ""),
                        "商品链接": "美国 http://www.amazon.com/gp/product/ASIN-SEA",
                        "FBA总库存": 480,
                        "加权日销": 6,
                        "可销售天数": 80,
                        "真实库存数量": 20,
                        "子SKU": "STOCK-A * 1",
                    }
                ],
            ),
            "真实库存-库存sku": (
                headers,
                [
                    {
                        "MSKU": "URGENT-1",
                        "父ASIN": "PARENT-AIR",
                        "ASIN": "ASIN-U",
                        "本地SKU": "SKU-U",
                        "本地SKU名称": "急发本地名",
                        "产品名称": "Urgent Product",
                        "备注": remarks.get("URGENT-1", ""),
                        "商品链接": "美国 http://www.amazon.com/gp/product/ASIN-U",
                        "FBA总库存": 90,
                        "加权日销": 3,
                        "可销售天数": 30,
                        "真实库存数量": 10,
                    },
                    {
                        "MSKU": "AIR-1",
                        "父ASIN": "PARENT-AIR",
                        "ASIN": "ASIN-A",
                        "本地SKU": "SKU-A",
                        "本地SKU名称": "空运本地名",
                        "产品名称": "Air Product",
                        "备注": remarks.get("AIR-1", ""),
                        "商品链接": "德国 https://www.amazon.de/dp/ASIN-A",
                        "FBA总库存": 300,
                        "加权日销": 6,
                        "可销售天数": 50,
                        "真实库存数量": 30,
                    },
                    {
                        "MSKU": "NO-1",
                        "父ASIN": "PARENT-NO",
                        "ASIN": "ASIN-N",
                        "本地SKU": "SKU-N",
                        "本地SKU名称": "暂不发货本地名",
                        "产品名称": "No Ship Product",
                        "备注": remarks.get("NO-1", ""),
                        "商品链接": "https://example.test/no",
                        "FBA总库存": 360,
                        "加权日销": 4,
                        "可销售天数": 90,
                        "真实库存数量": 40,
                    },
                    {
                        "MSKU": "SAMPLE-1",
                        "父ASIN": "PARENT-SAMPLE",
                        "ASIN": "ASIN-S",
                        "本地SKU": "SKU-S",
                        "本地SKU名称": "样本不足本地名",
                        "产品名称": "Sample Product",
                        "备注": remarks.get("SAMPLE-1", ""),
                        "商品链接": "https://example.test/sample",
                        "FBA总库存": 80,
                        "加权日销": 8,
                        "可销售天数": 10,
                        "真实库存数量": 8,
                    },
                ],
            ),
        },
    )


def _write_sales_report(path: Path) -> Path:
    headers = _sales_headers()
    rows = [
        {"MSKU": "SEA-1", "父ASIN": "PARENT-SEA", "ASIN": "ASIN-SEA", "本地SKU": "COMBO-SEA", "7天销量": 42, "14天销量": 84, "30天销量": 180, "销量趋势速率": 1.5, "销量趋势": "增长", "单品重量(g)(cm)": "1200g 10*10*10"},
        {"MSKU": "URGENT-1", "父ASIN": "PARENT-AIR", "ASIN": "ASIN-U", "本地SKU": "SKU-U", "7天销量": 21, "14天销量": 42, "30天销量": 90, "销量趋势速率": 1, "销量趋势": "平稳", "单品重量(g)(cm)": "10"},
        {"MSKU": "AIR-1", "父ASIN": "PARENT-AIR", "ASIN": "ASIN-A", "本地SKU": "SKU-A", "7天销量": 42, "14天销量": 84, "30天销量": 180, "销量趋势速率": 0.8, "销量趋势": "下降", "单品重量(g)(cm)": "20"},
        {"MSKU": "NO-1", "父ASIN": "PARENT-NO", "ASIN": "ASIN-N", "本地SKU": "SKU-N", "7天销量": 28, "14天销量": 56, "30天销量": 120, "销量趋势速率": 1, "销量趋势": "平稳", "单品重量(g)(cm)": "100"},
        {"MSKU": "SAMPLE-1", "父ASIN": "PARENT-SAMPLE", "ASIN": "ASIN-S", "本地SKU": "SKU-S", "7天销量": 56, "14天销量": 112, "30天销量": 240, "销量趋势速率": "", "销量趋势": "样本不足", "单品重量(g)(cm)": "50"},
    ]
    return _write_workbook(path, {"MSKU明细": (headers, rows)})


def _write_amazon_fba_inventory_snapshot(
    path: Path,
    *,
    snapshot_date: str = "20260525",
    rows: list[dict] | None = None,
) -> Path:
    summary_rows = rows or [
        {"店铺": "Amazon-Test", "MSKU": "AIR-1", "快照日期": snapshot_date, amazon_inv.AMAZON_FBA_TOTAL_COLUMN: 320},
        {"店铺": "Amazon-Test", "MSKU": "SEA-1", "快照日期": snapshot_date, amazon_inv.AMAZON_FBA_TOTAL_COLUMN: 500},
    ]
    return _write_workbook(
        path,
        {
            amazon_inv.SUMMARY_SHEET: (
                ["店铺", "MSKU", "快照日期", amazon_inv.AMAZON_FBA_TOTAL_COLUMN],
                summary_rows,
            ),
            amazon_inv.VALIDATION_SHEET: (
                ["字段", "值"],
                [
                    {"字段": "marketplace", "值": "US"},
                    {"字段": "mabang_site", "值": "美国站"},
                    {"字段": "amazon_sku_count", "值": 10},
                    {"字段": "matched_amazon_sku_count", "值": 8},
                    {"字段": "amazon_sku_match_ratio", "值": 0.8},
                    {"字段": "top_inventory_sku_count", "值": 10},
                    {"字段": "top_inventory_matched_count", "值": 8},
                ],
            ),
        },
    )


def _write_amazon_restock_inventory_snapshot(
    path: Path,
    *,
    snapshot_date: str = "20260525",
    rows: list[dict] | None = None,
) -> Path:
    summary_rows = rows or [
        {"店铺": "Amazon-Test", "MSKU": "AIR-1", "快照日期": snapshot_date, restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN: 330},
        {"店铺": "Amazon-Test", "MSKU": "SEA-1", "快照日期": snapshot_date, restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN: 510},
    ]
    return _write_workbook(
        path,
        {
            restock_inv.SUMMARY_SHEET: (
                ["店铺", "MSKU", "快照日期", restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN],
                summary_rows,
            ),
            restock_inv.VALIDATION_SHEET: (
                ["字段", "值"],
                [
                    {"字段": "country", "值": "US"},
                    {"字段": "mabang_site", "值": "美国站"},
                    {"字段": "amazon_sku_count", "值": 10},
                    {"字段": "matched_amazon_sku_count", "值": 8},
                    {"字段": "amazon_sku_match_ratio", "值": 0.8},
                    {"字段": "top_inventory_sku_count", "值": 10},
                    {"字段": "top_inventory_matched_count", "值": 8},
                ],
            ),
        },
    )


def test_find_matching_report_files_requires_common_source_time(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    _write_sales_report(sales_dir / "202605261530-Amazon-Test_销量分析.xlsx")

    result = repl.find_matching_report_files(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
    )

    assert result.source_data_time == "202605251530"
    assert result.sales_analysis_path.name == "202605251530-Amazon-Test_销量分析.xlsx"
    assert result.actual_inventory_path.name == "202605251530-Amazon-Test_真实库存.xlsx"


def test_find_matching_report_files_accepts_legacy_english_file_names(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_sales_analysis.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_actual_inventory.xlsx")

    result = repl.find_matching_report_files(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
    )

    assert result.source_data_time == "202605251530"
    assert result.sales_analysis_path.name == "202605251530-Amazon-Test_sales_analysis.xlsx"
    assert result.actual_inventory_path.name == "202605251530-Amazon-Test_actual_inventory.xlsx"


def test_find_matching_report_files_fails_without_common_time(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605261530-Amazon-Test_真实库存.xlsx")

    with pytest.raises(repl.StoreMskuReplenishmentError, match="未找到同源时间"):
        repl.find_matching_report_files("Amazon-Test", sales_analysis_dir=sales_dir, actual_inventory_dir=inventory_dir)


def test_find_same_day_unlinked_shipments_snapshot_picks_latest_same_day(tmp_path) -> None:
    snapshot_dir = tmp_path / "snapshots"
    snapshot_dir.mkdir()
    old_same_day = snapshot_dir / "202605250900-Amazon-Test_未关联货件快照.xlsx"
    latest_same_day = snapshot_dir / "202605251700-Amazon-Test_未关联货件快照.xlsx"
    other_day = snapshot_dir / "202605261700-Amazon-Test_未关联货件快照.xlsx"
    other_store = snapshot_dir / "202605251900-Amazon-Other_未关联货件快照.xlsx"
    for path in [old_same_day, latest_same_day, other_day, other_store]:
        path.touch()

    path, warning = repl.find_same_day_unlinked_shipments_snapshot(
        "Amazon-Test",
        "202605251530",
        snapshot_dir=snapshot_dir,
    )

    assert path == latest_same_day
    assert warning == ""


def test_find_same_day_unlinked_shipments_snapshot_accepts_legacy_english_file_names(tmp_path) -> None:
    snapshot_dir = tmp_path / "snapshots"
    snapshot_dir.mkdir()
    legacy_snapshot = snapshot_dir / "202605251700-Amazon-Test_unlinked_shipments_snapshot.xlsx"
    legacy_snapshot.touch()

    path, warning = repl.find_same_day_unlinked_shipments_snapshot(
        "Amazon-Test",
        "202605251530",
        snapshot_dir=snapshot_dir,
    )

    assert path == legacy_snapshot
    assert warning == ""


def test_load_inventory_rows_allows_missing_optional_name_columns(tmp_path) -> None:
    path = tmp_path / "actual_inventory.xlsx"
    old_headers = [
        "MSKU",
        "父ASIN",
        "ASIN",
        "本地SKU",
        "商品链接",
        "FBA总库存",
        "加权日销",
        "可销售天数",
        "真实库存数量",
        "子SKU",
    ]
    _write_workbook(
        path,
        {
            "真实库存-组合sku": (
                old_headers,
                [
                    {
                        "MSKU": "MSKU-C",
                        "父ASIN": "PARENT-C",
                        "ASIN": "ASIN-C",
                        "本地SKU": "COMBO-C",
                        "商品链接": "https://example.test/c",
                        "FBA总库存": 10,
                        "加权日销": 1,
                        "可销售天数": 10,
                        "真实库存数量": 5,
                        "子SKU": "SKU-A * 1",
                    }
                ],
            ),
            "真实库存-库存sku": (old_headers, []),
        },
    )

    rows = repl.load_inventory_rows(path)

    assert len(rows) == 1
    assert rows[0].local_sku_name == ""
    assert rows[0].product_name == ""


def test_replenishment_rules_and_report_output(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_dir = tmp_path / "snapshots"
    sales_path = _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    inventory_path = _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_dir=snapshot_dir,
    )

    report_path = Path(result.report_xlsx_path)
    assert result.to_payload() == {
        "success": True,
        "store_name": "Amazon-Test",
        "source_data_time": "202605251530",
        "sales_analysis_xlsx_path": str(sales_path),
        "actual_inventory_xlsx_path": str(inventory_path),
        "template_name": "默认模板",
        "template_version": 1,
        "row_count": 5,
        "link_count": 4,
        "air_urgent_count": 1,
        "air_count": 1,
        "sea_count": 1,
        "clearance_count": 0,
        "no_ship_count": 1,
        "sample_insufficient_count": 1,
        "report_xlsx_path": str(output_dir / "202605251530-Amazon-Test_备货建议.xlsx"),
        "source": "mabang_store_msku_replenishment",
        "unlinked_shipments_snapshot_warning": repl.UNLINKED_SNAPSHOT_MISSING_WARNING,
    }
    assert report_path.is_file()
    assert _sheet_names(report_path) == ["空运（急发）", "空运", "海运", "真实库存不足", "清货", "暂不建议发货", "链接备货汇总", "样本不足"]
    _assert_standard_dimensions(report_path, _sheet_names(report_path))
    assert _headers(report_path, "链接备货汇总") == list(repl.SUMMARY_COLUMNS)
    assert _headers(report_path, "真实库存不足") == list(repl.INVENTORY_SHORTAGE_COLUMNS)
    assert _headers(report_path, "清货") == list(repl.CLEARANCE_COLUMNS)
    assert _headers(report_path, "空运（急发）") == list(repl.AIR_DETAIL_COLUMNS)
    assert _headers(report_path, "空运") == list(repl.AIR_DETAIL_COLUMNS)
    for sheet_name in ["空运（急发）", "空运", "海运", "真实库存不足", "清货", "暂不建议发货", "样本不足"]:
        headers = _headers(report_path, sheet_name)
        assert headers.index(restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN) == headers.index(repl.MABANG_FBA_TOTAL_COLUMN) + 1
        assert headers.index(amazon_inv.AMAZON_FBA_TOTAL_COLUMN) == headers.index(restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN) + 1
        assert "销量趋势速率" in headers
        assert "销量趋势" not in headers
    for sheet_name in _sheet_names(report_path):
        headers = _headers(report_path, sheet_name)
        for header in (repl.MABANG_FBA_TOTAL_COLUMN, "真实库存数量"):
            if header in headers:
                formats = _column_number_formats(report_path, sheet_name, header)
                if formats:
                    assert set(formats) == {"0"}
    assert set(_column_number_formats(report_path, "真实库存不足", "库存缺口")) == {"0"}
    assert set(_column_number_formats(report_path, "链接备货汇总", "链接真实本地库存汇总")) == {"0"}
    assert set(_column_number_formats(report_path, "海运", "加权日销")) == {"0.00"}
    assert set(_column_number_formats(report_path, "海运", "预计总重量kg")) == {"0.00"}
    for header in ["海运天数", "海运建议量", "同时空运天数", "同时空运建议量"]:
        assert header not in _headers(report_path, "空运（急发）")
        assert header not in _headers(report_path, "空运")
        assert header in _headers(report_path, "海运")
        assert header in _headers(report_path, "真实库存不足")
    for sheet_name in ["海运", "暂不建议发货", "样本不足"]:
        headers = _headers(report_path, sheet_name)
        assert headers == list(repl.DETAIL_COLUMNS)
        assert "未关联抵扣前建议量" not in headers
        assert "补货量（减去未关联货件）" not in headers
        assert "补货量（减去 FBA 总库存）" not in headers
        assert "海运实际补货量" not in headers

    sea_rows = _load_records(report_path, "海运")
    assert sea_rows[0]["MSKU"] == "SEA-1"
    assert sea_rows[0]["本地SKU名称"] == "海运组合本地名"
    assert sea_rows[0]["产品名称"] == "Sea Combo Product"
    assert sea_rows[0]["销量趋势速率"] == 1.5
    assert sea_rows[0]["补货天数"] == 90
    assert sea_rows[0]["补货量"] == 540
    assert sea_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert sea_rows[0]["海运天数"] == 90
    assert sea_rows[0]["海运建议量"] == 60
    assert sea_rows[0]["预计总重量kg"] == 72
    assert "ceil(6.00*90)=540" in sea_rows[0]["决策原因"]
    assert "扣FBA后=60" in sea_rows[0]["决策原因"]
    assert sea_rows[0]["真实库存数量"] == 20
    assert sea_rows[0]["模板名称"] == "默认模板"
    assert sea_rows[0]["命中规则"] == "默认规则"

    urgent_rows = _load_records(report_path, "空运（急发）")
    assert urgent_rows[0]["MSKU"] == "URGENT-1"
    assert urgent_rows[0]["本地SKU名称"] == "急发本地名"
    assert urgent_rows[0]["产品名称"] == "Urgent Product"
    assert urgent_rows[0]["补货天数"] == 60
    assert urgent_rows[0]["补货量"] == 180
    assert urgent_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 90
    assert urgent_rows[0]["预计总重量kg"] == 0.9
    assert urgent_rows[0]["真实库存数量"] == 10

    air_rows = _load_records(report_path, "空运")
    assert air_rows[0]["MSKU"] == "AIR-1"
    assert air_rows[0]["本地SKU名称"] == "空运本地名"
    assert air_rows[0]["产品名称"] == "Air Product"
    assert air_rows[0]["补货天数"] == 60
    assert air_rows[0]["补货量"] == 360
    assert air_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert air_rows[0]["预计总重量kg"] == 1.2
    assert air_rows[0]["真实库存数量"] == 30

    no_ship_rows = _load_records(report_path, "暂不建议发货")
    assert no_ship_rows[0]["MSKU"] == "NO-1"
    assert no_ship_rows[0]["本地SKU名称"] == "暂不发货本地名"
    assert no_ship_rows[0]["产品名称"] == "No Ship Product"
    assert no_ship_rows[0]["补货量"] == 240
    assert no_ship_rows[0]["预计总重量kg"] in (None, "")
    assert "加权日销=4.00 <= 5" in no_ship_rows[0]["决策原因"]

    sample_rows = _load_records(report_path, "样本不足")
    assert sample_rows[0]["MSKU"] == "SAMPLE-1"
    assert sample_rows[0]["本地SKU名称"] == "样本不足本地名"
    assert sample_rows[0]["产品名称"] == "Sample Product"
    assert sample_rows[0]["补货量"] in (None, "")
    assert sample_rows[0]["预计总重量kg"] in (None, "")
    assert sample_rows[0]["决策原因"] == "销量趋势为样本不足，不计算备货量"

    shortage_rows = _load_records(report_path, "真实库存不足")
    assert {row["MSKU"] for row in shortage_rows} == {"SEA-1", "AIR-1", "URGENT-1"}
    assert {row["运输渠道"] for row in shortage_rows} == {"海运", "空运", "空运（急发）"}
    shortage_by_msku = {row["MSKU"]: row for row in shortage_rows}
    assert shortage_by_msku["SEA-1"]["本地SKU名称"] == "海运组合本地名"
    assert shortage_by_msku["SEA-1"]["产品名称"] == "Sea Combo Product"
    assert shortage_by_msku["SEA-1"]["库存缺口"] == 40
    assert shortage_by_msku["AIR-1"]["库存缺口"] == 30
    assert shortage_by_msku["URGENT-1"]["库存缺口"] == 80
    assert shortage_by_msku["SEA-1"]["预计总重量kg"] == 72
    assert shortage_by_msku["AIR-1"]["预计总重量kg"] == 1.2
    assert shortage_by_msku["URGENT-1"]["预计总重量kg"] == 0.9
    assert all(row["MSKU"] not in {"NO-1", "SAMPLE-1"} for row in shortage_rows)
    assert any(row["MSKU"] == "SEA-1" for row in sea_rows)
    assert any(row["MSKU"] == "AIR-1" for row in air_rows)
    assert any(row["MSKU"] == "URGENT-1" for row in urgent_rows)

    summary_rows = _load_records(report_path, "链接备货汇总")
    summary_by_parent = {row["父ASIN"]: row for row in summary_rows}
    assert summary_by_parent["PARENT-SEA"]["商品链接"] == "http://www.amazon.com/gp/product/PARENT-SEA"
    assert summary_by_parent["PARENT-SEA"]["总补货量"] == 60
    assert summary_by_parent["PARENT-SEA"]["海运建议量"] == 60
    assert summary_by_parent["PARENT-SEA"]["涉及运输方式"] == "海运"
    assert summary_by_parent["PARENT-SEA"]["链接真实本地库存汇总"] == 20
    assert summary_by_parent["PARENT-AIR"]["商品链接"] == "http://www.amazon.com/gp/product/PARENT-AIR"
    assert summary_by_parent["PARENT-AIR"]["总补货量"] == 150
    assert summary_by_parent["PARENT-AIR"]["空运（急发）补货量"] == 90
    assert summary_by_parent["PARENT-AIR"]["空运补货量"] == 60
    assert summary_by_parent["PARENT-AIR"]["涉及运输方式"] == "空运（急发）、空运"
    assert summary_by_parent["PARENT-AIR"]["链接真实本地库存汇总"] == 40
    assert summary_rows[-1]["父ASIN"] == "PARENT-SAMPLE"
    assert summary_rows[-1]["总补货量"] == 0
    assert summary_rows[-1]["涉及运输方式"] == "样本不足"


def test_amazon_fba_inventory_snapshot_adds_comparison_fields_without_changing_execution(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_dir = tmp_path / "unlinked"
    sales_path = _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    inventory_path = _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_fba_inventory_snapshot(tmp_path / "amazon_snapshot.xlsx")

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_dir=snapshot_dir,
        amazon_fba_inventory_snapshot_path=amazon_snapshot_path,
    )

    payload = result.to_payload()
    assert payload["sales_analysis_xlsx_path"] == str(sales_path)
    assert payload["actual_inventory_xlsx_path"] == str(inventory_path)
    assert payload["amazon_fba_inventory_snapshot_path"] == str(amazon_snapshot_path)
    assert payload["amazon_fba_inventory_validation"] == {
        "marketplace": "US",
        "mabang_site": "美国站",
        "amazon_sku_count": 10,
        "matched_amazon_sku_count": 8,
        "amazon_sku_match_ratio": 0.8,
        "top_inventory_sku_count": 10,
        "top_inventory_matched_count": 8,
    }

    report_path = Path(result.report_xlsx_path)
    air_rows = _load_records(report_path, "空运")
    assert air_rows[0]["MSKU"] == "AIR-1"
    assert air_rows[0]["补货量"] == 360
    assert air_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert air_rows[0][amazon_inv.AMAZON_FBA_TOTAL_COLUMN] == 320
    assert air_rows[0][repl.AMAZON_DEDUCTED_REPLENISH_COLUMN] == 40

    sea_rows = _load_records(report_path, "海运")
    assert sea_rows[0]["MSKU"] == "SEA-1"
    assert sea_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert sea_rows[0][amazon_inv.AMAZON_FBA_TOTAL_COLUMN] == 500
    assert sea_rows[0][repl.AMAZON_DEDUCTED_REPLENISH_COLUMN] == 40

    urgent_rows = _load_records(report_path, "空运（急发）")
    assert urgent_rows[0]["MSKU"] == "URGENT-1"
    assert urgent_rows[0][amazon_inv.AMAZON_FBA_TOTAL_COLUMN] in (None, "")
    assert urgent_rows[0][repl.AMAZON_DEDUCTED_REPLENISH_COLUMN] in (None, "")
    assert "未匹配亚马逊物流库存" in urgent_rows[0]["决策原因"]


def test_amazon_restock_inventory_snapshot_adds_comparison_fields_without_changing_execution(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_dir = tmp_path / "unlinked"
    sales_path = _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    inventory_path = _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_restock_inventory_snapshot(tmp_path / "restock_snapshot.xlsx")

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_dir=snapshot_dir,
        amazon_restock_inventory_snapshot_path=amazon_snapshot_path,
    )

    payload = result.to_payload()
    assert payload["sales_analysis_xlsx_path"] == str(sales_path)
    assert payload["actual_inventory_xlsx_path"] == str(inventory_path)
    assert payload["amazon_restock_inventory_snapshot_path"] == str(amazon_snapshot_path)
    assert payload["amazon_restock_inventory_validation"] == {
        "country": "US",
        "mabang_site": "美国站",
        "amazon_sku_count": 10,
        "matched_amazon_sku_count": 8,
        "amazon_sku_match_ratio": 0.8,
        "top_inventory_sku_count": 10,
        "top_inventory_matched_count": 8,
    }

    report_path = Path(result.report_xlsx_path)
    air_rows = _load_records(report_path, "空运")
    assert air_rows[0]["MSKU"] == "AIR-1"
    assert air_rows[0]["补货量"] == 360
    assert air_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert air_rows[0][restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN] == 330
    assert air_rows[0][repl.AMAZON_RESTOCK_DEDUCTED_REPLENISH_COLUMN] == 30

    sea_rows = _load_records(report_path, "海运")
    assert sea_rows[0]["MSKU"] == "SEA-1"
    assert sea_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert sea_rows[0][restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN] == 510
    assert sea_rows[0][repl.AMAZON_RESTOCK_DEDUCTED_REPLENISH_COLUMN] == 30

    urgent_rows = _load_records(report_path, "空运（急发）")
    assert urgent_rows[0]["MSKU"] == "URGENT-1"
    assert urgent_rows[0][restock_inv.AMAZON_RESTOCK_TOTAL_COLUMN] in (None, "")
    assert urgent_rows[0][repl.AMAZON_RESTOCK_DEDUCTED_REPLENISH_COLUMN] in (None, "")
    assert "未匹配亚马逊补充库存" in urgent_rows[0]["决策原因"]


@pytest.mark.parametrize("snapshot_date", ["20260524", "20260525", "20260526"])
def test_amazon_restock_inventory_snapshot_allows_same_or_adjacent_day(tmp_path, snapshot_date: str) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_restock_inventory_snapshot(tmp_path / "restock_snapshot.xlsx", snapshot_date=snapshot_date)

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        amazon_restock_inventory_snapshot_path=amazon_snapshot_path,
    )

    assert result.amazon_restock_inventory_snapshot_path == str(amazon_snapshot_path)


def test_amazon_restock_inventory_snapshot_rejects_date_more_than_one_day_apart(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_restock_inventory_snapshot(tmp_path / "restock_snapshot.xlsx", snapshot_date="20260527")

    with pytest.raises(repl.StoreMskuReplenishmentError, match="相差超过1天"):
        repl.calculate_store_msku_replenishment(
            "Amazon-Test",
            sales_analysis_dir=sales_dir,
            actual_inventory_dir=inventory_dir,
            output_dir=output_dir,
            amazon_restock_inventory_snapshot_path=amazon_snapshot_path,
        )


@pytest.mark.parametrize("snapshot_date", ["20260524", "20260525", "20260526"])
def test_amazon_fba_inventory_snapshot_allows_same_or_adjacent_day(tmp_path, snapshot_date: str) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_fba_inventory_snapshot(tmp_path / "amazon_snapshot.xlsx", snapshot_date=snapshot_date)

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        amazon_fba_inventory_snapshot_path=amazon_snapshot_path,
    )

    assert result.amazon_fba_inventory_snapshot_path == str(amazon_snapshot_path)


def test_amazon_fba_inventory_snapshot_rejects_date_more_than_one_day_apart(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_fba_inventory_snapshot(tmp_path / "amazon_snapshot.xlsx", snapshot_date="20260527")

    with pytest.raises(repl.StoreMskuReplenishmentError, match="亚马逊物流库存快照日期与备货数据日期相差超过1天"):
        repl.calculate_store_msku_replenishment(
            "Amazon-Test",
            sales_analysis_dir=sales_dir,
            actual_inventory_dir=inventory_dir,
            output_dir=output_dir,
            amazon_fba_inventory_snapshot_path=amazon_snapshot_path,
        )


def test_amazon_fba_inventory_snapshot_rejects_invalid_date_format(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    amazon_snapshot_path = _write_amazon_fba_inventory_snapshot(tmp_path / "amazon_snapshot.xlsx", snapshot_date="2026-05-25")

    with pytest.raises(repl.StoreMskuReplenishmentError, match="亚马逊物流库存快照日期格式无效"):
        repl.calculate_store_msku_replenishment(
            "Amazon-Test",
            sales_analysis_dir=sales_dir,
            actual_inventory_dir=inventory_dir,
            output_dir=output_dir,
            amazon_fba_inventory_snapshot_path=amazon_snapshot_path,
        )


def test_clearance_rows_move_out_of_transport_and_summary_sheets(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_dir = tmp_path / "snapshots"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(
        inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx",
        remarks={
            "SEA-1": "清货-海运",
            "URGENT-1": "清货急发",
            "AIR-1": "空运清货",
            "NO-1": "清货但本来不发",
            "SAMPLE-1": "清货样本不足",
        },
    )

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_dir=snapshot_dir,
    )

    assert result.air_urgent_count == 0
    assert result.air_count == 0
    assert result.sea_count == 0
    assert result.clearance_count == 3
    assert result.no_ship_count == 1
    assert result.sample_insufficient_count == 1

    report_path = Path(result.report_xlsx_path)
    assert _load_records(report_path, "空运（急发）") == []
    assert _load_records(report_path, "空运") == []
    assert _load_records(report_path, "海运") == []
    assert _load_records(report_path, "真实库存不足") == []

    clearance_rows = _load_records(report_path, "清货")
    assert {row["MSKU"] for row in clearance_rows} == {"SEA-1", "URGENT-1", "AIR-1"}
    clearance_by_msku = {row["MSKU"]: row for row in clearance_rows}
    assert clearance_by_msku["SEA-1"]["运输渠道"] == "海运"
    assert clearance_by_msku["URGENT-1"]["运输渠道"] == "空运（急发）"
    assert clearance_by_msku["AIR-1"]["运输渠道"] == "空运"
    assert clearance_by_msku["SEA-1"]["备注"] == "清货-海运"
    assert clearance_by_msku["URGENT-1"]["备注"] == "清货急发"
    assert clearance_by_msku["AIR-1"]["备注"] == "空运清货"
    assert clearance_by_msku["SEA-1"]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert clearance_by_msku["URGENT-1"]["补货量（减去 FBA 总库存和未关联货件）"] == 90
    assert clearance_by_msku["AIR-1"]["补货量（减去 FBA 总库存和未关联货件）"] == 60
    assert "备注包含清货" in clearance_by_msku["SEA-1"]["决策原因"]

    no_ship_rows = _load_records(report_path, "暂不建议发货")
    sample_rows = _load_records(report_path, "样本不足")
    assert [row["MSKU"] for row in no_ship_rows] == ["NO-1"]
    assert no_ship_rows[0]["备注"] == "清货但本来不发"
    assert [row["MSKU"] for row in sample_rows] == ["SAMPLE-1"]
    assert sample_rows[0]["备注"] == "清货样本不足"

    summary_rows = _load_records(report_path, "链接备货汇总")
    summary_parents = {row["父ASIN"] for row in summary_rows}
    assert "PARENT-SEA" not in summary_parents
    assert "PARENT-AIR" not in summary_parents
    assert "PARENT-NO" in summary_parents
    assert "PARENT-SAMPLE" in summary_parents


def test_unlinked_shipments_snapshot_deducts_final_replenishment_quantity(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_path = tmp_path / "202605251600-Amazon-Test_未关联货件快照.xlsx"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    _write_workbook(
        snapshot_path,
        {
            "未关联货件汇总": (
                ["店铺", "MSKU", "未关联数量"],
                [
                    {"店铺": "Amazon-Test", "MSKU": "AIR-1", "未关联数量": 20},
                    {"店铺": "Amazon-Test", "MSKU": "URGENT-1", "未关联数量": 250},
                    {"店铺": "Amazon-Test", "MSKU": "SEA-1", "未关联数量": 40},
                    {"店铺": "Amazon-Test", "MSKU": "UNMATCHED-1", "未关联数量": 999},
                ],
            )
        },
    )

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_path=snapshot_path,
    )

    assert result.air_urgent_count == 0
    assert result.air_count == 1
    assert result.sea_count == 0
    assert result.no_ship_count == 3
    assert result.to_payload()["unlinked_shipments_snapshot_path"] == str(snapshot_path)
    report_path = Path(result.report_xlsx_path)

    air_rows = _load_records(report_path, "空运")
    assert air_rows[0]["MSKU"] == "AIR-1"
    assert air_rows[0][repl.MABANG_FBA_TOTAL_COLUMN] == 300
    assert air_rows[0]["可销售天数"] == 50
    assert air_rows[0]["未关联数量"] == 20
    assert air_rows[0]["补货量"] == 360
    assert air_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 40
    assert air_rows[0]["预计总重量kg"] == 0.8
    assert "补货量（减去 FBA 总库存和未关联货件）=40" in air_rows[0]["决策原因"]

    no_ship_rows = _load_records(report_path, "暂不建议发货")
    urgent_row = next(row for row in no_ship_rows if row["MSKU"] == "URGENT-1")
    assert urgent_row["补货量"] == 180
    assert urgent_row["补货量（减去 FBA 总库存和未关联货件）"] == 0
    assert urgent_row["未关联数量"] == 250
    assert "FBA 总库存（马帮数据）和未关联数量已覆盖本次建议量" in urgent_row["决策原因"]

    sea_row = next(row for row in no_ship_rows if row["MSKU"] == "SEA-1")
    assert sea_row["补货量"] == 540
    assert sea_row["补货量（减去 FBA 总库存和未关联货件）"] == 20
    assert sea_row["海运建议量"] == 20
    assert sea_row["预计总重量kg"] in (None, "")
    assert "扣减FBA 总库存（马帮数据）和未关联货件后，海运重量不足60kg" in sea_row["决策原因"]
    assert _load_records(report_path, "海运") == []

    shortage_rows = _load_records(report_path, "真实库存不足")
    assert [row["MSKU"] for row in shortage_rows] == ["AIR-1"]
    assert shortage_rows[0]["库存缺口"] == 10

    summary_rows = _load_records(report_path, "链接备货汇总")
    air_summary = next(row for row in summary_rows if row["父ASIN"] == "PARENT-AIR")
    assert air_summary["总补货量"] == 40
    assert air_summary["空运补货量"] == 40
    assert air_summary["链接未关联数量汇总"] == 270
    sea_summary = next(row for row in summary_rows if row["父ASIN"] == "PARENT-SEA")
    assert sea_summary["总补货量"] == 20
    assert sea_summary["海运建议量"] == 0
    assert sea_summary["链接未关联数量汇总"] == 40


def test_same_day_unlinked_shipments_snapshot_auto_deducts_latest_snapshot(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_dir = tmp_path / "snapshots"
    old_snapshot = snapshot_dir / "202605251200-Amazon-Test_未关联货件快照.xlsx"
    latest_snapshot = snapshot_dir / "202605251800-Amazon-Test_未关联货件快照.xlsx"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    _write_workbook(
        old_snapshot,
        {
            "未关联货件汇总": (
                ["店铺", "MSKU", "未关联数量"],
                [{"店铺": "Amazon-Test", "MSKU": "AIR-1", "未关联数量": 1}],
            )
        },
    )
    _write_workbook(
        latest_snapshot,
        {
            "未关联货件汇总": (
                ["店铺", "MSKU", "未关联数量"],
                [{"店铺": "Amazon-Test", "MSKU": "AIR-1", "未关联数量": 20}],
            )
        },
    )

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_dir=snapshot_dir,
    )

    payload = result.to_payload()
    assert payload["unlinked_shipments_snapshot_path"] == str(latest_snapshot)
    assert "unlinked_shipments_snapshot_warning" not in payload
    air_rows = _load_records(Path(result.report_xlsx_path), "空运")
    assert air_rows[0]["MSKU"] == "AIR-1"
    assert air_rows[0]["未关联数量"] == 20
    assert air_rows[0]["补货量"] == 360
    assert air_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 40


def test_non_same_day_unlinked_shipments_snapshot_warns_without_deduction(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_dir = tmp_path / "snapshots"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    _write_workbook(
        snapshot_dir / "202605261200-Amazon-Test_未关联货件快照.xlsx",
        {
            "未关联货件汇总": (
                ["店铺", "MSKU", "未关联数量"],
                [{"店铺": "Amazon-Test", "MSKU": "AIR-1", "未关联数量": 100}],
            )
        },
    )

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
        unlinked_shipments_snapshot_dir=snapshot_dir,
    )

    payload = result.to_payload()
    assert "unlinked_shipments_snapshot_path" not in payload
    assert payload["unlinked_shipments_snapshot_warning"] == repl.UNLINKED_SNAPSHOT_IGNORED_NON_SAME_DAY_WARNING
    air_rows = _load_records(Path(result.report_xlsx_path), "空运")
    assert air_rows[0]["MSKU"] == "AIR-1"
    assert air_rows[0]["未关联数量"] == 0
    assert air_rows[0]["补货量"] == 360
    assert air_rows[0]["补货量（减去 FBA 总库存和未关联货件）"] == 60


def test_manual_unlinked_shipments_snapshot_must_be_same_day(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    snapshot_path = tmp_path / "202605261200-Amazon-Test_未关联货件快照.xlsx"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")
    _write_workbook(
        snapshot_path,
        {
            "未关联货件汇总": (
                ["店铺", "MSKU", "未关联数量"],
                [{"店铺": "Amazon-Test", "MSKU": "AIR-1", "未关联数量": 100}],
            )
        },
    )

    with pytest.raises(repl.StoreMskuReplenishmentError, match="日期与备货数据日期不一致"):
        repl.calculate_store_msku_replenishment(
            "Amazon-Test",
            sales_analysis_dir=sales_dir,
            actual_inventory_dir=inventory_dir,
            output_dir=output_dir,
            unlinked_shipments_snapshot_path=snapshot_path,
        )


def test_inventory_shortage_rows_only_include_suggested_rows_with_known_shortage() -> None:
    rows = [
        _replenishment_row("URGENT", repl.AIR_URGENT_SHEET, replenish_quantity=100, actual_inventory=20),
        _replenishment_row("AIR", repl.AIR_SHEET, replenish_quantity=100, actual_inventory=99.5),
        _replenishment_row("SEA", repl.SEA_SHEET, replenish_quantity=100, actual_inventory=80),
        _replenishment_row("ENOUGH", repl.AIR_SHEET, replenish_quantity=100, actual_inventory=100),
        _replenishment_row("UNKNOWN", repl.AIR_SHEET, replenish_quantity=100, actual_inventory=None),
        _replenishment_row("NO-SHIP", repl.NO_SHIP_SHEET, replenish_quantity=100, actual_inventory=0),
        _replenishment_row("SAMPLE", repl.SAMPLE_INSUFFICIENT_SHEET, replenish_quantity=None, actual_inventory=0),
    ]

    shortage_rows = repl.inventory_shortage_rows(rows)

    assert [row["MSKU"] for row in shortage_rows] == ["URGENT", "AIR", "SEA"]
    assert [row["运输渠道"] for row in shortage_rows] == ["空运（急发）", "空运", "海运"]
    assert shortage_rows[0]["库存缺口"] == 80
    assert shortage_rows[1]["库存缺口"] == 0.5
    assert shortage_rows[2]["库存缺口"] == 20


def test_parse_weight_grams_uses_first_number() -> None:
    assert repl.parse_weight_grams("120g 10*20*30") == 120
    assert repl.parse_weight_grams("1,200.5g") == 1200.5
    assert repl.parse_weight_grams("") is None


def test_detail_payload_estimated_weight_uses_final_replenishment_quantity() -> None:
    air_row = _replenishment_row("AIR-WEIGHT", repl.AIR_SHEET, replenish_quantity=80, actual_inventory=10)
    no_ship_row = _replenishment_row("NO-SHIP-WEIGHT", repl.NO_SHIP_SHEET, replenish_quantity=80, actual_inventory=10)

    assert air_row.to_detail_payload()["预计总重量kg"] == 8
    assert no_ship_row.to_detail_payload()["预计总重量kg"] == ""


def test_custom_template_changes_replenishment_result(tmp_path, monkeypatch) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")

    custom_params = tmpl.load_default_template().to_store_payload()
    custom_params["name"] = "保守海运模板"
    custom_params["version"] = 2
    custom_params["params"]["sea"]["min_weight_kg"] = 100
    store_path = tmp_path / "artifacts" / "mabang_replenishment_templates" / "templates.json"
    store_path.parent.mkdir(parents=True, exist_ok=True)
    store_path.write_text(
        json.dumps({"templates": [custom_params]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.chdir(tmp_path)

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        template_name="保守海运模板",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
    )

    assert result.template_name == "保守海运模板"
    assert result.template_version == 2
    report_path = Path(result.report_xlsx_path)
    sea_rows = _load_records(report_path, "海运")
    no_ship_rows = _load_records(report_path, "暂不建议发货")
    assert sea_rows == []
    assert any(row["MSKU"] == "SEA-1" and "海运重量不足100kg" in row["决策原因"] for row in no_ship_rows)


def test_us_group_1_builtin_template_uses_110_day_sea(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        template_name="US模板-一组",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
    )

    assert result.template_name == "US模板-一组"
    assert result.sea_count == 1
    report_path = Path(result.report_xlsx_path)
    sea_rows = _load_records(report_path, "海运")
    sea_row = next(row for row in sea_rows if row["MSKU"] == "SEA-1")
    assert sea_row["补货天数"] == 110
    assert sea_row["补货量"] == 660
    assert sea_row["补货量（减去 FBA 总库存和未关联货件）"] == 180
    assert sea_row["海运天数"] == 110
    assert sea_row["海运建议量"] == 180
    assert sea_row["预计总重量kg"] == 216
    assert "按海运备货天数110天计算补货量" in sea_row["决策原因"]
    no_ship_rows = _load_records(report_path, "暂不建议发货")
    no_ship_row = next(row for row in no_ship_rows if row["MSKU"] == "NO-1")
    assert "加权日销=4.00 <= 5" in no_ship_row["决策原因"]
    summary_rows = _load_records(report_path, "链接备货汇总")
    sea_summary = next(row for row in summary_rows if row["父ASIN"] == "PARENT-SEA")
    assert sea_summary["总补货量"] == 180
    assert sea_summary["海运建议量"] == 180


def test_lin_meiqi_group_2_template_uses_inclusive_one_day_sea_threshold() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-1", fba_total_inventory=100),
        _sales_detail_for_daily_sales(1, weight_grams=2100),
        template,
    )

    assert row.sheet_name == repl.NO_SHIP_SHEET
    assert row.weighted_daily_sales == pytest.approx(1)
    assert row.sales_days == pytest.approx(100)
    assert row.replenish_days == 100
    assert row.replenish_quantity == 0
    assert row.original_replenish_quantity == 100
    assert row.fba_deducted_replenish_quantity == 0
    assert row.sea_days == 100
    assert row.sea_quantity == 0
    assert row.companion_air_days == 70
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 0
    assert row.estimated_weight_kg is None
    assert "FBA 总库存（马帮数据）已覆盖本次建议量" in row.decision_reason


def test_lin_meiqi_group_2_template_keeps_eighty_sales_days_in_air() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-6-AIR", fba_total_inventory=480),
        _sales_detail_for_daily_sales(6, weight_grams=610),
        template,
    )

    assert row.sheet_name == repl.NO_SHIP_SHEET
    assert row.sales_days == pytest.approx(80)
    assert row.replenish_days == 80
    assert row.original_replenish_quantity == 480
    assert row.fba_deducted_replenish_quantity == 0
    assert row.replenish_quantity == 0
    assert "FBA 总库存（马帮数据）已覆盖本次建议量" in row.decision_reason


def test_lin_meiqi_group_2_template_rejects_below_one_day_sea() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-BELOW-1", fba_total_inventory=100),
        _sales_detail_for_daily_sales(0.99, weight_grams=1000),
        template,
    )

    assert row.sheet_name == repl.NO_SHIP_SHEET
    assert row.weighted_daily_sales == pytest.approx(0.99)
    assert "加权日销=0.99 < 1" in row.decision_reason


def test_lin_meiqi_group_2_template_uses_120_day_sea_above_twenty_daily_sales() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-21", fba_total_inventory=2000),
        _sales_detail_for_daily_sales(21, weight_grams=200),
        template,
    )

    assert row.sheet_name == repl.SEA_SHEET
    assert row.weighted_daily_sales == pytest.approx(21)
    assert row.sales_days == pytest.approx(2000 / 21)
    assert row.replenish_days == 120
    assert row.replenish_quantity == 520
    assert row.original_replenish_quantity == 2520
    assert row.fba_deducted_replenish_quantity == 520
    assert row.sea_days == 120
    assert row.sea_quantity == 520
    assert row.companion_air_days == 80
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 520
    assert row.estimated_weight_kg == pytest.approx(104)


def test_lin_meiqi_group_2_template_treats_above_three_hundred_as_above_twenty() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-301", fba_total_inventory=30000),
        _sales_detail_for_daily_sales(301, weight_grams=10),
        template,
    )

    assert row.sheet_name == repl.SEA_SHEET
    assert row.replenish_days == 120
    assert row.replenish_quantity == 6120
    assert row.original_replenish_quantity == 36120
    assert row.fba_deducted_replenish_quantity == 6120
    assert row.sea_days == 120
    assert row.sea_quantity == 6120
    assert row.companion_air_days == 80
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 6120
    assert row.estimated_weight_kg == pytest.approx(61.2)


def test_lin_meiqi_group_2_template_deducts_unlinked_after_companion_air() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-6-UNLINKED", fba_total_inventory=600),
        _sales_detail_for_daily_sales(6, weight_grams=1300),
        template,
        unlinked_quantity=10,
    )

    assert row.sheet_name == repl.SEA_SHEET
    assert row.original_replenish_quantity == 660
    assert row.fba_deducted_replenish_quantity == 60
    assert row.replenish_quantity == 50
    assert row.sea_quantity == 50
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 50
    assert row.estimated_weight_kg == pytest.approx(65)
    assert "补货量=660，FBA 总库存（马帮数据）=600，扣FBA后=60" in row.decision_reason
    assert "补货量（减去 FBA 总库存和未关联货件）=50" in row.decision_reason


def test_lin_meiqi_group_2_template_deducts_unlinked_from_sea_after_companion_air() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-6-UNLINKED-SEA", fba_total_inventory=500),
        _sales_detail_for_daily_sales(6, weight_grams=600),
        template,
        unlinked_quantity=50,
    )

    assert row.sheet_name == repl.SEA_SHEET
    assert row.original_replenish_quantity == 660
    assert row.fba_deducted_replenish_quantity == 160
    assert row.replenish_quantity == 110
    assert row.sea_quantity == 110
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 110
    assert row.estimated_weight_kg == pytest.approx(66)
    assert "补货量（减去 FBA 总库存和未关联货件）=110" in row.decision_reason


def test_lin_meiqi_group_2_template_rejects_small_net_sea_quantity() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")
    params = json.loads(json.dumps(template.params, ensure_ascii=False))
    params["sea"]["min_net_quantity"] = 61
    custom_template = tmpl.ReplenishmentTemplate(
        name="小净量测试模板",
        version=1,
        description="",
        params=params,
    )

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-6-SMALL-NET", fba_total_inventory=600),
        _sales_detail_for_daily_sales(6, weight_grams=2100),
        custom_template,
    )

    assert row.sheet_name == repl.NO_SHIP_SHEET
    assert row.original_replenish_quantity == 660
    assert row.fba_deducted_replenish_quantity == 60
    assert row.replenish_quantity == 60
    assert row.sea_quantity == 60
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 60
    assert "海运数量不足61件" in row.decision_reason


def test_lin_meiqi_group_2_template_rejects_small_net_sea_weight() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-6-LIGHT-NET", fba_total_inventory=600),
        _sales_detail_for_daily_sales(6, weight_grams=100),
        template,
    )

    assert row.sheet_name == repl.NO_SHIP_SHEET
    assert row.replenish_days == 110
    assert row.original_replenish_quantity == 660
    assert row.fba_deducted_replenish_quantity == 60
    assert row.replenish_quantity == 60
    assert row.sea_quantity == 60
    assert row.companion_air_days == 75
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 60
    assert row.estimated_weight_kg is None
    assert "海运重量不足60kg" in row.decision_reason


def test_lin_meiqi_group_2_template_rejects_unlinked_deduction_below_min_sea_quantity() -> None:
    template = tmpl.get_template("2组-US站点-林美淇")

    row = repl.calculate_replenishment_row(
        _inventory_input_row("DAILY-6-UNLINKED-SMALL-SEA", fba_total_inventory=500),
        _sales_detail_for_daily_sales(6, weight_grams=400),
        template,
        unlinked_quantity=131,
    )

    assert row.sheet_name == repl.NO_SHIP_SHEET
    assert row.original_replenish_quantity == 660
    assert row.fba_deducted_replenish_quantity == 160
    assert row.replenish_quantity == 29
    assert row.sea_quantity == 29
    assert row.companion_air_quantity == 0
    assert row.sea_net_quantity == 29
    assert row.estimated_weight_kg is None
    assert "扣减FBA 总库存（马帮数据）和未关联货件后，海运数量不足30件" in row.decision_reason


def test_uk_group_1_builtin_template_disables_sea(tmp_path) -> None:
    sales_dir = tmp_path / "sales"
    inventory_dir = tmp_path / "inventory"
    output_dir = tmp_path / "output"
    _write_sales_report(sales_dir / "202605251530-Amazon-Test_销量分析.xlsx")
    _write_inventory_report(inventory_dir / "202605251530-Amazon-Test_真实库存.xlsx")

    result = repl.calculate_store_msku_replenishment(
        "Amazon-Test",
        template_name="UK模板-一组",
        sales_analysis_dir=sales_dir,
        actual_inventory_dir=inventory_dir,
        output_dir=output_dir,
    )

    assert result.template_name == "UK模板-一组"
    assert result.sea_count == 0
    assert result.no_ship_count == 2
    report_path = Path(result.report_xlsx_path)
    assert _load_records(report_path, "海运") == []
    no_ship_rows = _load_records(report_path, "暂不建议发货")
    sea_candidate = next(row for row in no_ship_rows if row["MSKU"] == "SEA-1")
    assert sea_candidate["补货天数"] == 85
    assert sea_candidate["补货量"] == 510
    assert "模板已关闭海运" in sea_candidate["决策原因"]
