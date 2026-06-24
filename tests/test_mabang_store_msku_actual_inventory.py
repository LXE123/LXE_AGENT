from __future__ import annotations

import asyncio
import json
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest

from services.mabang.amazon.fba import store_msku_actual_inventory as inv
from services.mabang.auth import MabangAuthContext


def _xlsx_bytes(rows: list[dict], *, columns: list[str]) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _write_xlsx(path: Path, rows: list[dict], *, columns: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_xlsx_bytes(rows, columns=columns))
    return path


def _combo_xlsx_bytes(rows: list[dict], *, max_components: int = 2) -> bytes:
    columns = ["组合sku编码", "关联sku个数"]
    for index in range(1, max_components + 1):
        columns.extend([f"关联sku编号{index}", f"关联sku捆绑数量{index}"])
    return _xlsx_bytes(rows, columns=columns)


def _stock_xlsx_bytes(rows: list[dict]) -> bytes:
    return _xlsx_bytes(rows, columns=["库存SKU编号", "可用库存量"])


class _FakeResponse:
    def __init__(
        self,
        payload: dict | None = None,
        *,
        status: int = 200,
        body: bytes = b"",
        text_body: str | None = None,
    ) -> None:
        self.status = status
        self._payload = dict(payload or {})
        self._body = body
        self._text_body = text_body

    async def text(self) -> str:
        if self._text_body is not None:
            return self._text_body
        return json.dumps(self._payload, ensure_ascii=False)

    async def json(self, content_type=None) -> dict:
        if self._text_body is not None:
            raise ValueError("not json")
        return dict(self._payload)

    async def read(self) -> bytes:
        return self._body


class _FakeRequest:
    def __init__(self, response: _FakeResponse) -> None:
        self._response = response

    async def __aenter__(self) -> _FakeResponse:
        return self._response

    async def __aexit__(self, exc_type, exc, tb) -> None:
        return None


class _FakeSession:
    def __init__(self, responses: list[_FakeResponse]) -> None:
        self.responses = list(responses)
        self.calls: list[dict] = []

    def post(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "POST", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))

    def get(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "GET", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))


async def _fake_auth_context(*args, **kwargs) -> MabangAuthContext:
    return MabangAuthContext(
        scope=str(kwargs.get("scope") or ""),
        account="",
        source="test",
        cookies_by_domain={
            ".mabangerp.com": [
                {"name": "PHPSESSID", "value": "sid", "domain": ".mabangerp.com"},
                {
                    "name": "MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE",
                    "value": "memcache-key",
                    "domain": ".mabangerp.com",
                },
                {
                    "name": "MABANG_ERP_PRO_MEMBERINFO_LOGIN_PLUS",
                    "value": "login-plus",
                    "domain": ".mabangerp.com",
                },
                {"name": "signed", "value": "signed-value", "domain": ".mabangerp.com"},
                {"name": "route", "value": "route-value", "domain": ".mabangerp.com"},
            ]
        },
        free_token="",
        wms_cookie_header="",
        raw={},
    )


def _form_value(call: dict, name: str) -> str:
    for key, value in call.get("data", []):
        if key == name:
            return value
    raise AssertionError(f"missing form field: {name}")


def _form_values(call: dict, name: str) -> list[str]:
    return [value for key, value in call.get("data", []) if key == name]


def _load_records(path: Path, sheet_name: str = "真实库存（深圳仓库）-库存sku") -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        return [dict(zip(headers, values, strict=False)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
    finally:
        workbook.close()


def _sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _column_number_formats(path: Path, sheet_name: str, column_name: str) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        column_index = headers.index(column_name) + 1
        return [
            cell.number_format
            for cells in worksheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index)
            for cell in cells
        ]
    finally:
        workbook.close()


def _column_fill_colors(path: Path, sheet_name: str, column_name: str) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        column_index = headers.index(column_name) + 1
        return [
            str(cell.fill.fgColor.rgb or "")
            for cells in worksheet.iter_rows(min_row=1, min_col=column_index, max_col=column_index)
            for cell in cells
        ]
    finally:
        workbook.close()


def _assert_standard_dimensions(path: Path, sheet_names: list[str]) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    try:
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            assert worksheet.sheet_format.defaultRowHeight == 15
            for row_index in range(1, worksheet.max_row + 1):
                assert worksheet.row_dimensions[row_index].height == 15
            for column_index in range(1, worksheet.max_column + 1):
                assert worksheet.column_dimensions[get_column_letter(column_index)].width == 15
    finally:
        workbook.close()


def test_combo_step1_form_appends_hsp022_once_and_uses_export_fields() -> None:
    form = inv._combo_step1_form_data(["SKU-A", "HSP022"], memcache_key="memcache-key")

    assert _form_value({"data": form}, "orderIds") == "SKU-A\r\nHSP022\r\n"
    assert _form_values({"data": form}, "fieldlabel") == ["uq101", "uq136", "uq138"]
    assert _form_values({"data": form}, "map-name[]") == ["组合sku编码", "关联sku个数", "关联sku信息"]
    assert _form_value({"data": form}, "memcacheKey") == "memcache-key"
    assert _form_value({"data": form}, "operateType") == "19"
    assert _form_value({"data": form}, "hiddenPage") == "1"


def test_combo_prewarm_forms_append_hsp022_once() -> None:
    list_form = inv._combo_list_prewarm_form_data(["SKU-A", "HSP022"])
    template_form = inv._combo_export_template_prewarm_form_data(["SKU-A", "HSP022"])

    assert _form_value({"data": list_form}, "searchLike") == "comboSku"
    assert _form_value({"data": list_form}, "operate") == "Like"
    assert _form_value({"data": list_form}, "isBatchSearch") == "1"
    assert _form_value({"data": list_form}, "selecttype") == "comboSku"
    assert _form_value({"data": list_form}, "stockData") == "SKU-A\r\nHSP022"
    assert _form_value({"data": template_form}, "mod") == "export.exportTemplate"
    assert _form_value({"data": template_form}, "data") == "SKU-A\r\nHSP022"
    assert _form_value({"data": template_form}, "type") == "1"
    assert _form_value({"data": template_form}, "menu") == "combosku"
    assert _form_value({"data": template_form}, "exportUrl") == inv._combo_export_url()


def test_combo_prewarm_failure_blocks_export(monkeypatch) -> None:
    fake_session = _FakeSession([_FakeResponse(status=500, text_body="server error")])
    monkeypatch.setattr(inv, "erp_http_session", fake_session)

    with pytest.raises(Exception, match="组合SKU预热 1请求失败"):
        asyncio.run(
            inv.prewarm_combo_sku_export(
                ["SKU-A"],
                private_cookie_header="private-cookie",
                private_amz_cookie_header="private-amz-cookie",
                delay_sec=0,
            )
        )

    assert len(fake_session.calls) == 1
    assert fake_session.calls[0]["url"] == inv._combo_list_url()


def test_warehouse_search_form_uses_fixed_warehouse_and_crlf_skus() -> None:
    form = inv._warehouse_search_form_data(["SKU-A", "SKU-B"])

    assert _form_value({"data": form}, "warehouseIds[]") == "1014318"
    assert _form_value({"data": form}, "stockSkuStr") == "SKU-A\r\nSKU-B\r\n"
    assert _form_value({"data": form}, "statusIN[]") == "3"


def test_parse_combo_sku_xlsx_and_hsp022_filter_rules(tmp_path) -> None:
    xlsx_path = tmp_path / "combo.xlsx"
    xlsx_path.write_bytes(
        _combo_xlsx_bytes(
            [
                {
                    "组合sku编码": "COMBO-A",
                    "关联sku个数": 2,
                    "关联sku编号1": "STOCK-A",
                    "关联sku捆绑数量1": 1,
                    "关联sku编号2": "STOCK-B",
                    "关联sku捆绑数量2": 5,
                },
                {
                    "组合sku编码": "HSP022",
                    "关联sku个数": 1,
                    "关联sku编号1": "STOCK-H",
                    "关联sku捆绑数量1": 1,
                },
            ]
        )
    )

    combo_map = inv.parse_combo_sku_xlsx(xlsx_path)
    filtered_without_source_hsp = inv.filter_combo_map_for_source(combo_map, source_local_skus=["COMBO-A"])
    filtered_with_source_hsp = inv.filter_combo_map_for_source(combo_map, source_local_skus=["COMBO-A", "HSP022"])

    assert sorted(combo.combo_sku for combo in filtered_without_source_hsp.values()) == ["COMBO-A"]
    assert sorted(combo.combo_sku for combo in filtered_with_source_hsp.values()) == ["COMBO-A", "HSP022"]
    combo = filtered_without_source_hsp[inv.normalize_sku_key("COMBO-A")]
    assert [(item.stock_sku, item.quantity) for item in combo.components] == [
        ("STOCK-A", Decimal("1")),
        ("STOCK-B", Decimal("5")),
    ]


def test_filter_combo_map_requires_hsp022_sentinel() -> None:
    combo_map = {
        inv.normalize_sku_key("COMBO-A"): inv.ComboSku(
            combo_sku="COMBO-A",
            components=(inv.ComboComponent("STOCK-A", Decimal("1")),),
        )
    }

    with pytest.raises(inv.StoreMskuActualInventoryError, match="缺少哨兵SKU: HSP022"):
        inv.filter_combo_map_for_source(combo_map, source_local_skus=["COMBO-A"])


def test_parse_stock_inventory_sums_duplicate_stock_sku_rows(tmp_path) -> None:
    xlsx_path = tmp_path / "stock.xlsx"
    xlsx_path.write_bytes(
        _stock_xlsx_bytes(
            [
                {"库存SKU编号": "STOCK-A", "可用库存量": "2"},
                {"库存SKU编号": "STOCK-A", "可用库存量": "3"},
                {"库存SKU编号": "STOCK-B", "可用库存量": "12"},
            ]
        )
    )

    result = inv.parse_stock_inventory_xlsx(xlsx_path)

    assert result == {
        "STOCK-A": Decimal("5"),
        "STOCK-B": Decimal("12"),
    }


def test_load_store_msku_rows_requires_product_link_column(tmp_path) -> None:
    source_path = tmp_path / "202605251530-Amazon-Test_店铺MSKU数据.xlsx"
    _write_xlsx(
        source_path,
        [{"MSKU": "MSKU-A", "父ASIN": "PARENT-A", "ASIN": "ASIN-A", "本地SKU": "SKU-A"}],
        columns=["MSKU", "父ASIN", "ASIN", "本地SKU"],
    )

    with pytest.raises(inv.StoreMskuActualInventoryError, match="店铺MSKU数据缺少列: 商品链接"):
        inv.load_store_msku_rows(source_path)


def test_find_latest_store_msku_file_accepts_legacy_english_file_name(tmp_path) -> None:
    source_path = tmp_path / "202605251530-Amazon-Test_msku_data.xlsx"
    _write_xlsx(
        source_path,
        [{"MSKU": "MSKU-A", "父ASIN": "PARENT-A", "ASIN": "ASIN-A", "本地SKU": "SKU-A", "商品链接": "https://example.test/a"}],
        columns=["MSKU", "父ASIN", "ASIN", "本地SKU", "商品链接"],
    )

    result = inv.find_latest_store_msku_file("Amazon-Test", input_dir=tmp_path)

    assert result.path == source_path
    assert result.source_data_time == "202605251530"


def test_load_store_msku_rows_requires_sales_and_fba_stock_columns(tmp_path) -> None:
    source_path = tmp_path / "202605251530-Amazon-Test_店铺MSKU数据.xlsx"
    _write_xlsx(
        source_path,
        [
            {
                "MSKU": "MSKU-A",
                "父ASIN": "PARENT-A",
                "ASIN": "ASIN-A",
                "本地SKU": "SKU-A",
                "商品链接": "https://example.test/a",
            }
        ],
        columns=["MSKU", "父ASIN", "ASIN", "本地SKU", "商品链接"],
    )

    with pytest.raises(inv.StoreMskuActualInventoryError, match="7天销量.*调仓中"):
        inv.load_store_msku_rows(source_path)


def test_load_store_msku_rows_allows_missing_optional_name_columns(tmp_path) -> None:
    source_path = tmp_path / "202605251530-Amazon-Test_店铺MSKU数据.xlsx"
    _write_xlsx(
        source_path,
        [
            {
                "MSKU": "MSKU-A",
                "父ASIN": "PARENT-A",
                "ASIN": "ASIN-A",
                "本地SKU": "SKU-A",
                "商品链接": "https://example.test/a",
                "7天销量": 7,
                "14天销量": 14,
                "30天销量": 30,
                "可售": 4,
                "待入库": 3,
                "预留": 2,
                "在途": 1,
                "待调仓": 6,
                "调仓中": 7,
            }
        ],
        columns=[
            "MSKU",
            "父ASIN",
            "ASIN",
            "本地SKU",
            "商品链接",
            "7天销量",
            "14天销量",
            "30天销量",
            "可售",
            "待入库",
            "预留",
            "在途",
            "待调仓",
            "调仓中",
        ],
    )

    rows = inv.load_store_msku_rows(source_path)

    assert len(rows) == 1
    assert rows[0].local_sku_name == ""
    assert rows[0].product_name == ""
    assert rows[0].remark == ""


def test_load_store_msku_rows_parses_invalid_sales_and_fba_stock_as_zero(tmp_path) -> None:
    source_path = tmp_path / "202605251530-Amazon-Test_店铺MSKU数据.xlsx"
    _write_xlsx(
        source_path,
        [
            {
                "MSKU": "MSKU-A",
                "父ASIN": "PARENT-A",
                "ASIN": "ASIN-A",
                "本地SKU": "SKU-A",
                "商品链接": "https://example.test/a",
                "本地SKU名称": "本地产品A",
                "产品名称": "Product A",
                "备注": "清货观察",
                "7天销量": "nan",
                "14天销量": "1,400",
                "30天销量": "not-a-number",
                "可售": "1,000",
                "待入库": "",
                "预留": "nan",
                "在途": "not-a-number",
                "待调仓": "2",
                "调仓中": "bad",
            }
        ],
        columns=[
            "MSKU",
            "父ASIN",
            "ASIN",
            "本地SKU",
            "商品链接",
            "本地SKU名称",
            "产品名称",
            "备注",
            "7天销量",
            "14天销量",
            "30天销量",
            "可售",
            "待入库",
            "预留",
            "在途",
            "待调仓",
            "调仓中",
        ],
    )

    rows = inv.load_store_msku_rows(source_path)

    assert len(rows) == 1
    assert rows[0].local_sku_name == "本地产品A"
    assert rows[0].product_name == "Product A"
    assert rows[0].sales_7d == Decimal("0")
    assert rows[0].sales_14d == Decimal("1400")
    assert rows[0].sales_30d == Decimal("0")
    assert rows[0].fba_sellable == Decimal("1000")
    assert rows[0].fba_inbound == Decimal("0")
    assert rows[0].fba_reserved == Decimal("0")
    assert rows[0].fba_in_transit == Decimal("0")
    assert rows[0].fba_pending_transfer == Decimal("2")
    assert rows[0].fba_transferring == Decimal("0")


def test_calculate_inventory_rows_handles_normal_combo_and_missing_stock() -> None:
    rows = [
        inv.StoreMskuRow(
            "MSKU-A",
            "PARENT-A",
            "ASIN-A",
            "SKU-A",
            "https://example.test/a",
            local_sku_name="单品本地名",
            product_name="Single Product",
            remark="清货单品",
            sales_7d=Decimal("7"),
            sales_14d=Decimal("14"),
            sales_30d=Decimal("30"),
            fba_sellable=Decimal("4"),
            fba_inbound=Decimal("3"),
            fba_reserved=Decimal("2"),
            fba_in_transit=Decimal("1"),
            fba_pending_transfer=Decimal("6"),
            fba_transferring=Decimal("7"),
        ),
        inv.StoreMskuRow(
            "MSKU-C",
            "PARENT-C",
            "ASIN-C",
            "COMBO-A",
            "https://example.test/c",
            local_sku_name="组合本地名",
            product_name="Combo Product",
            remark="组合备注",
            sales_7d=Decimal("70"),
            sales_14d=Decimal("140"),
            sales_30d=Decimal("300"),
            fba_sellable=Decimal("20"),
            fba_inbound=Decimal("15"),
            fba_reserved=Decimal("10"),
            fba_in_transit=Decimal("5"),
            fba_pending_transfer=Decimal("4"),
            fba_transferring=Decimal("6"),
        ),
        inv.StoreMskuRow("MSKU-M", "PARENT-M", "ASIN-M", "SKU-MISSING", "https://example.test/m"),
        inv.StoreMskuRow("MSKU-CM", "PARENT-CM", "ASIN-CM", "COMBO-MISSING", "https://example.test/cm"),
    ]
    combo_map = {
        inv.normalize_sku_key("COMBO-A"): inv.ComboSku(
            combo_sku="COMBO-A",
            components=(
                inv.ComboComponent("STOCK-A", Decimal("1")),
                inv.ComboComponent("STOCK-B", Decimal("2")),
            ),
        ),
        inv.normalize_sku_key("COMBO-MISSING"): inv.ComboSku(
            combo_sku="COMBO-MISSING",
            components=(inv.ComboComponent("STOCK-MISSING", Decimal("1")),),
        ),
    }

    result_rows, missing = inv.calculate_inventory_rows(
        rows,
        combo_map=combo_map,
        stock_quantities={
            "SKU-A": Decimal("9"),
            "STOCK-A": Decimal("10"),
            "STOCK-B": Decimal("12"),
        },
    )

    assert [row.actual_inventory for row in result_rows] == [Decimal("9"), Decimal("6"), None, None]
    assert [row.local_sku_name for row in result_rows] == ["单品本地名", "组合本地名", "", ""]
    assert [row.product_name for row in result_rows] == ["Single Product", "Combo Product", "", ""]
    assert [row.remark for row in result_rows] == ["清货单品", "组合备注", "", ""]
    assert [row.product_link for row in result_rows] == [
        "https://example.test/a",
        "https://example.test/c",
        "https://example.test/m",
        "https://example.test/cm",
    ]
    assert [row.is_combo_sku for row in result_rows] == [False, True, False, True]
    assert result_rows[0].fba_total_inventory == Decimal("23")
    assert result_rows[0].weighted_daily_sales == Decimal("1.0")
    assert result_rows[0].sales_days == Decimal("23")
    assert result_rows[1].fba_total_inventory == Decimal("60")
    assert result_rows[1].weighted_daily_sales == Decimal("10.0")
    assert result_rows[1].sales_days == Decimal("6")
    assert result_rows[1].child_skus == "STOCK-A * 1, STOCK-B * 2"
    assert missing == ["SKU-MISSING", "STOCK-MISSING"]


def test_write_actual_inventory_xlsx_splits_rows_by_inventory_state(tmp_path) -> None:
    output_path = tmp_path / "actual_inventory.xlsx"
    rows = [
        inv.ActualInventoryRow(
            "MSKU-S1",
            "PARENT-S1",
            "ASIN-S1",
            "SKU-S1",
            "https://example.test/stock-1",
            Decimal("10"),
            "",
            fba_total_inventory=Decimal("30"),
            weighted_daily_sales=Decimal("3.456"),
            sales_days=Decimal("8.666"),
            local_sku_name="单品本地1",
            product_name="Stock Product 1",
            remark="清货单品1",
        ),
        inv.ActualInventoryRow(
            "MSKU-S0",
            "PARENT-S0",
            "ASIN-S0",
            "SKU-ZERO",
            "https://example.test/zero",
            Decimal("0"),
            "",
            fba_total_inventory=Decimal("0"),
            weighted_daily_sales=Decimal("0"),
            sales_days=None,
        ),
        inv.ActualInventoryRow(
            "MSKU-C",
            "PARENT-C",
            "ASIN-C",
            "COMBO-A",
            "https://example.test/combo",
            Decimal("6"),
            "STOCK-A * 1",
            is_combo_sku=True,
            fba_total_inventory=Decimal("20"),
            weighted_daily_sales=Decimal("5"),
            sales_days=Decimal("4"),
            local_sku_name="组合本地商品",
            product_name="Combo Product",
            remark="组合备注",
        ),
        inv.ActualInventoryRow("MSKU-N", "PARENT-N", "ASIN-N", "", "https://example.test/no-local", None, ""),
        inv.ActualInventoryRow(
            "MSKU-M",
            "PARENT-M",
            "ASIN-M",
            "SKU-MISSING",
            "https://example.test/missing",
            None,
            "",
        ),
        inv.ActualInventoryRow(
            "MSKU-CM",
            "PARENT-CM",
            "ASIN-CM",
            "COMBO-MISSING",
            "https://example.test/combo-missing",
            None,
            "STOCK-MISSING * 1",
        ),
    ]

    groups = inv.split_inventory_rows(rows)
    inv.write_actual_inventory_xlsx(rows, output_path)

    assert len(groups.combo_inventory_rows) == 1
    assert len(groups.stock_inventory_rows) == 2
    assert len(groups.inventory_rows) == 3
    assert len(groups.no_local_sku_rows) == 1
    assert len(groups.no_inventory_rows) == 2
    assert _sheet_names(output_path) == ["真实库存（深圳仓库）-组合sku", "真实库存（深圳仓库）-库存sku", "无本地SKU", "无库存数据"]
    _assert_standard_dimensions(output_path, _sheet_names(output_path))
    for sheet_name in _sheet_names(output_path):
        records = _load_records(output_path, sheet_name)
        if records:
            expected_columns = (
                list(inv.INVENTORY_OUTPUT_COLUMNS)
                if sheet_name in {"真实库存（深圳仓库）-组合sku", "真实库存（深圳仓库）-库存sku"}
                else list(inv.OUTPUT_COLUMNS)
            )
            assert list(records[0]) == expected_columns

    combo_records = _load_records(output_path, "真实库存（深圳仓库）-组合sku")
    stock_records = _load_records(output_path, "真实库存（深圳仓库）-库存sku")
    no_local_records = _load_records(output_path, "无本地SKU")
    no_inventory_records = _load_records(output_path, "无库存数据")

    assert [record["MSKU"] for record in combo_records] == ["MSKU-C"]
    assert list(combo_records[0]) == list(inv.INVENTORY_OUTPUT_COLUMNS)
    assert combo_records[0]["本地SKU名称"] == "组合本地商品"
    assert combo_records[0]["产品名称"] == "Combo Product"
    assert combo_records[0]["备注"] == "组合备注"
    assert combo_records[0]["商品链接"] == "https://example.test/combo"
    assert combo_records[0]["FBA总库存"] == 20
    assert combo_records[0]["加权日销"] == 5
    assert combo_records[0]["可销售天数"] == 4
    assert combo_records[0]["真实库存（深圳仓库）数量"] == 6
    assert [record["MSKU"] for record in stock_records] == ["MSKU-S1", "MSKU-S0"]
    assert stock_records[0]["本地SKU名称"] == "单品本地1"
    assert stock_records[0]["产品名称"] == "Stock Product 1"
    assert stock_records[0]["备注"] == "清货单品1"
    assert stock_records[0]["商品链接"] == "https://example.test/stock-1"
    assert stock_records[0]["FBA总库存"] == 30
    assert stock_records[0]["加权日销"] == 3.46
    assert stock_records[0]["可销售天数"] == 8.67
    assert stock_records[1]["商品链接"] == "https://example.test/zero"
    assert stock_records[1]["加权日销"] == 0
    assert stock_records[1]["可销售天数"] in (None, "")
    assert stock_records[1]["真实库存（深圳仓库）数量"] == 0
    assert [record["MSKU"] for record in no_local_records] == ["MSKU-N"]
    assert list(no_local_records[0]) == list(inv.OUTPUT_COLUMNS)
    assert no_local_records[0]["商品链接"] == "https://example.test/no-local"
    assert [record["MSKU"] for record in no_inventory_records] == ["MSKU-M", "MSKU-CM"]
    assert list(no_inventory_records[0]) == list(inv.OUTPUT_COLUMNS)
    assert no_inventory_records[0]["商品链接"] == "https://example.test/missing"
    assert no_inventory_records[1]["商品链接"] == "https://example.test/combo-missing"
    assert no_inventory_records[1]["子SKU"] == "STOCK-MISSING * 1"
    assert _column_number_formats(output_path, "真实库存（深圳仓库）-组合sku", "加权日销") == ["0.00"]
    assert _column_number_formats(output_path, "真实库存（深圳仓库）-组合sku", "可销售天数") == ["0.00"]
    assert _column_number_formats(output_path, "真实库存（深圳仓库）-库存sku", "加权日销") == ["0.00", "0.00"]
    assert _column_number_formats(output_path, "真实库存（深圳仓库）-库存sku", "可销售天数") == ["0.00", "0.00"]
    assert _column_fill_colors(output_path, "真实库存（深圳仓库）-组合sku", "MSKU") == ["00FFF2CC", "00FFF2CC"]
    assert _column_fill_colors(output_path, "真实库存（深圳仓库）-组合sku", "真实库存（深圳仓库）数量") == ["00FFF2CC", "00FFF2CC"]
    assert _column_fill_colors(output_path, "真实库存（深圳仓库）-库存sku", "MSKU") == [
        "00FFF2CC",
        "00FFF2CC",
        "00FFF2CC",
    ]
    assert _column_fill_colors(output_path, "真实库存（深圳仓库）-库存sku", "真实库存（深圳仓库）数量") == [
        "00FFF2CC",
        "00FFF2CC",
        "00FFF2CC",
    ]
    assert _column_fill_colors(output_path, "无本地SKU", "MSKU") == ["00000000", "00000000"]
    assert _column_fill_colors(output_path, "无本地SKU", "真实库存（深圳仓库）数量") == ["00000000", "00000000"]
    assert _column_fill_colors(output_path, "无库存数据", "MSKU") == ["00000000", "00000000", "00000000"]
    assert _column_fill_colors(output_path, "无库存数据", "真实库存（深圳仓库）数量") == ["00000000", "00000000", "00000000"]


def test_export_store_msku_actual_inventory_success_with_fake_network(monkeypatch, tmp_path) -> None:
    source_path = tmp_path / "input" / "202605251530-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx"
    _write_xlsx(
        source_path,
        [
            {
                "MSKU": "MSKU-A",
                "父ASIN": "PARENT-A",
                "ASIN": "ASIN-A",
                "本地SKU": "SKU-A",
                "商品链接": "https://example.test/a",
                "7天销量": 7,
                "14天销量": 14,
                "30天销量": 30,
                "可售": 4,
                "待入库": 3,
                "预留": 2,
                "在途": 1,
                "待调仓": 6,
                "调仓中": 7,
            },
            {
                "MSKU": "MSKU-C",
                "父ASIN": "PARENT-C",
                "ASIN": "ASIN-C",
                "本地SKU": "COMBO-A",
                "商品链接": "https://example.test/c",
                "7天销量": 70,
                "14天销量": 140,
                "30天销量": 300,
                "可售": 20,
                "待入库": 15,
                "预留": 10,
                "在途": 5,
                "待调仓": 4,
                "调仓中": 6,
            },
            {
                "MSKU": "MSKU-N",
                "父ASIN": "PARENT-N",
                "ASIN": "ASIN-N",
                "本地SKU": "",
                "商品链接": "https://example.test/no-local",
                "7天销量": 1,
                "14天销量": 2,
                "30天销量": 3,
                "可售": 4,
                "待入库": 5,
                "预留": 6,
                "在途": 7,
                "待调仓": 8,
                "调仓中": 9,
            },
        ],
        columns=[
            "MSKU",
            "父ASIN",
            "ASIN",
            "本地SKU",
            "商品链接",
            "7天销量",
            "14天销量",
            "30天销量",
            "可售",
            "待入库",
            "预留",
            "在途",
            "待调仓",
            "调仓中",
        ],
    )
    fake_session = _FakeSession(
        [
            _FakeResponse({"ignored": "prewarm-list"}),
            _FakeResponse({"ignored": "prewarm-template"}),
            _FakeResponse({"success_type": 2, "sn": "sn-1", "subtask_num": 1, "chunkNum": 10000, "success": True}),
            _FakeResponse({"updateR": True, "subO": [{"id": "1", "success": "1"}], "success": True}),
            _FakeResponse({"async": True, "taskId": "task-1", "success": True}),
            _FakeResponse({"success": True, "file_url": "https://cos.example.test/combo.xlsx", "state": 1}),
            _FakeResponse(
                body=_combo_xlsx_bytes(
                    [
                        {
                            "组合sku编码": "COMBO-A",
                            "关联sku个数": 2,
                            "关联sku编号1": "STOCK-A",
                            "关联sku捆绑数量1": 1,
                            "关联sku编号2": "STOCK-B",
                            "关联sku捆绑数量2": 2,
                        },
                        {
                            "组合sku编码": "HSP022",
                            "关联sku个数": 1,
                            "关联sku编号1": "STOCK-H",
                            "关联sku捆绑数量1": 1,
                        },
                    ]
                )
            ),
            _FakeResponse({"success": True}),
            _FakeResponse(
                body=_stock_xlsx_bytes(
                    [
                        {"库存SKU编号": "SKU-A", "可用库存量": 9},
                        {"库存SKU编号": "STOCK-A", "可用库存量": 10},
                        {"库存SKU编号": "STOCK-B", "可用库存量": 12},
                    ]
                )
            ),
        ]
    )
    monkeypatch.setattr(inv, "erp_http_session", fake_session)
    monkeypatch.setattr(inv, "external_http_session", fake_session)
    monkeypatch.setattr(inv, "get_auth_context", _fake_auth_context)
    monkeypatch.setattr(inv, "_timestamp_text", lambda *_args, **_kwargs: "202605271530")
    sleep_calls: list[float] = []

    async def fake_sleep(delay: float) -> None:
        sleep_calls.append(delay)

    monkeypatch.setattr(inv.asyncio, "sleep", fake_sleep)

    result = asyncio.run(
        inv.export_store_msku_actual_inventory(
            "Amazon-Lerxiuer-FR",
            input_dir=tmp_path / "input",
            output_dir=tmp_path / "output",
            timeout_sec=0,
            poll_interval_sec=0.1,
        )
    )

    payload = result.to_payload()
    assert payload == {
        "success": True,
        "store_name": "Amazon-Lerxiuer-FR",
        "warehouse_id": "1014318",
        "warehouse_name": "深圳仓库",
        "source_msku_xlsx_path": str(source_path),
        "source_msku_data_time": "202605251530",
        "unique_local_sku_count": 2,
        "detected_combo_sku_count": 1,
        "queried_warehouse_stock_sku_count": 3,
        "matched_warehouse_inventory_msku_row_count": 2,
        "missing_local_sku_msku_row_count": 1,
        "missing_warehouse_inventory_msku_row_count": 0,
        "missing_warehouse_stock_sku_count": 0,
        "missing_warehouse_stock_skus": [],
        "shenzhen_warehouse_inventory_report_xlsx_path": str(tmp_path / "output" / "202605251530-Amazon-Lerxiuer-FR_真实库存（深圳仓库）.xlsx"),
        "result_source": "mabang_store_msku_shenzhen_warehouse_inventory",
    }
    report_path = Path(result.shenzhen_warehouse_inventory_report_xlsx_path)
    assert _sheet_names(report_path) == ["真实库存（深圳仓库）-组合sku", "真实库存（深圳仓库）-库存sku", "无本地SKU", "无库存数据"]
    combo_records = _load_records(report_path, "真实库存（深圳仓库）-组合sku")
    stock_records = _load_records(report_path, "真实库存（深圳仓库）-库存sku")
    assert combo_records[0]["真实库存（深圳仓库）数量"] == 6
    assert combo_records[0]["商品链接"] == "https://example.test/c"
    assert combo_records[0]["FBA总库存"] == 60
    assert combo_records[0]["加权日销"] == 10
    assert combo_records[0]["可销售天数"] == 6
    assert combo_records[0]["子SKU"] == "STOCK-A * 1, STOCK-B * 2"
    assert stock_records[0]["真实库存（深圳仓库）数量"] == 9
    assert stock_records[0]["商品链接"] == "https://example.test/a"
    assert stock_records[0]["FBA总库存"] == 23
    assert stock_records[0]["加权日销"] == 1
    assert stock_records[0]["可销售天数"] == 23
    no_local_records = _load_records(report_path, "无本地SKU")
    assert [record["MSKU"] for record in no_local_records] == ["MSKU-N"]
    assert no_local_records[0]["商品链接"] == "https://example.test/no-local"
    assert _load_records(report_path, "无库存数据") == []

    post_calls = [call for call in fake_session.calls if call["method"] == "POST"]
    assert [call["url"] for call in post_calls[:6]] == [
        inv._combo_list_url(),
        inv._combo_export_template_url(),
        inv._combo_export_url(),
        inv._combo_export_url(),
        inv._combo_export_url(),
        inv._combo_export_url(),
    ]
    assert sleep_calls == [1.0]
    assert _form_value(post_calls[0], "stockData") == "SKU-A\r\nCOMBO-A\r\nHSP022"
    assert _form_value(post_calls[1], "data") == "SKU-A\r\nCOMBO-A\r\nHSP022"
    assert [_form_value(call, "step") for call in post_calls[2:6]] == ["1", "2", "3", "4"]
    assert _form_value(post_calls[2], "orderIds") == "SKU-A\r\nCOMBO-A\r\nHSP022\r\n"
    assert _form_value(post_calls[6], "stockSkuStr") == "SKU-A\r\nSTOCK-A\r\nSTOCK-B\r\n"
    assert _form_value(post_calls[6], "warehouseIds[]") == "1014318"
