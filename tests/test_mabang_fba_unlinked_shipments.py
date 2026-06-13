from __future__ import annotations

import asyncio
import json
from pathlib import Path

import pytest

from services.mabang.amazon.fba import unlinked_shipments as ship
from services.mabang.amazon.fba.batch_delivery import BatchDeliveryDownloadInfo, BatchDeliveryTask


class _FakeResponse:
    def __init__(self, payload: dict | None = None, *, status: int = 200, body: bytes = b"") -> None:
        self.status = status
        self._payload = dict(payload or {})
        self._body = body

    async def text(self) -> str:
        return json.dumps(self._payload, ensure_ascii=False)

    async def json(self, content_type=None) -> dict:
        return dict(self._payload)

    async def read(self) -> bytes:
        return self._body


class _FakeRequest:
    def __init__(self, response: _FakeResponse) -> None:
        self._response = response

    async def __aenter__(self) -> _FakeResponse:
        return self._response

    async def __aexit__(self, exc_type, exc, tb) -> None:
        return None


class _FakeSession:
    def __init__(self, responses: list[_FakeResponse]) -> None:
        self.responses = list(responses)
        self.calls: list[dict] = []

    def get(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "GET", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))

    def post(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "POST", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))


def _write_raw_csv(path: Path, rows: list[dict]) -> Path:
    headers = [
        "发货单号",
        "发货单状态",
        "物流方式",
        "物流渠道",
        "创建时间",
        "店铺",
        "MSKU",
        "MSKU发货量",
        "货件单号",
    ]
    lines = [",".join(headers)]
    for row in rows:
        lines.append(",".join(str(row.get(header, "")) for header in headers))
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    return path


def _load_snapshot_records(path: Path, sheet_name: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        return [dict(zip(headers, values, strict=False)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
    finally:
        workbook.close()


def test_build_store_unlinked_shipments_snapshot_aggregates_raw_csv(tmp_path) -> None:
    wms_path = _write_raw_csv(
        tmp_path / "202606130900-Amazon-Test-US-WMS待配货-1.csv",
        [
            {
                "发货单号": "SP1",
                "物流方式": "空运",
                "物流渠道": "出口退税空运",
                "创建时间": "2026-06-13 09:00:00",
                "店铺": "Amazon-Test-US",
                "MSKU": "MSKU-A",
                "MSKU发货量": "10",
                "货件单号": "m1",
            },
            {
                "发货单号": "SP1",
                "物流方式": "空运",
                "物流渠道": "出口退税空运",
                "创建时间": "2026-06-13 09:00:00",
                "店铺": "Amazon-Test-US",
                "MSKU": "MSKU-B",
                "MSKU发货量": "3",
                "货件单号": "m1",
            },
        ],
    )
    linked_path = _write_raw_csv(
        tmp_path / "202606130901-Amazon-Test-US-待关联货件-2.csv",
        [
            {
                "发货单号": "SP2",
                "发货单状态": "待关联货件",
                "物流方式": "海运",
                "物流渠道": "",
                "创建时间": "2026-06-13 09:01:00",
                "店铺": "Amazon-Test-US",
                "MSKU": "MSKU-A",
                "MSKU发货量": "5",
                "货件单号": "m2",
            }
        ],
    )

    result = ship.build_store_unlinked_shipments_snapshot(
        [wms_path, linked_path],
        store_name="Amazon-Test-US",
        output_dir=tmp_path,
        snapshot_time="202606130902",
    )

    assert result.to_payload() == {
        "success": True,
        "store_name": "Amazon-Test-US",
        "snapshot_time": "202606130902",
        "snapshot_xlsx_path": str(tmp_path / "202606130902-Amazon-Test-US_unlinked_shipments_snapshot.xlsx"),
        "raw_file_count": 2,
        "detail_count": 3,
        "msku_count": 2,
        "total_unlinked_quantity": 18,
        "source": "mabang_fba_unlinked_shipments_snapshot",
    }
    snapshot_path = Path(result.snapshot_xlsx_path)
    assert snapshot_path.is_file()
    summary_rows = _load_snapshot_records(snapshot_path, ship.SNAPSHOT_SUMMARY_SHEET)
    assert [(row["MSKU"], row["未关联数量"], row["明细行数"]) for row in summary_rows] == [
        ("MSKU-A", 15, 2),
        ("MSKU-B", 3, 1),
    ]
    assert summary_rows[0]["涉及状态"] == "WMS待配货、待关联货件"
    assert summary_rows[0]["涉及运输方式"] == "出口退税空运、海运"
    detail_rows = _load_snapshot_records(snapshot_path, ship.SNAPSHOT_DETAIL_SHEET)
    assert detail_rows[0]["状态"] == "WMS待配货"
    assert ship.load_unlinked_shipment_quantities(snapshot_path, store_name="Amazon-Test-US") == {
        "MSKU-A": 15,
        "MSKU-B": 3,
    }


def test_pick_shop_option_requires_exact_unique_match() -> None:
    shops = [
        ship.ShopOption(store_id=1, name="Amazon-Test-US", raw={}),
        ship.ShopOption(store_id=2, name="Amazon-Test-UK", raw={}),
    ]

    selected = ship.pick_shop_option("Amazon-Test-US", shops)

    assert selected.store_id == 1


def test_pick_shop_option_missing_returns_similar_candidates() -> None:
    shops = [ship.ShopOption(store_id=1, name="Amazon-Yihaiqian-US", raw={})]

    with pytest.raises(ship.UnlinkedShipmentError, match="是否指: Amazon-Yihaiqian-US"):
        ship.pick_shop_option("Amazon-Yihaiqian-U", shops)


def test_pick_shop_option_rejects_duplicate_exact_matches() -> None:
    shops = [
        ship.ShopOption(store_id=1, name="Amazon-Test-US", raw={}),
        ship.ShopOption(store_id=2, name="Amazon-Test-US", raw={}),
    ]

    with pytest.raises(ship.UnlinkedShipmentError, match="匹配到多个店铺"):
        ship.pick_shop_option("Amazon-Test-US", shops)


def test_status_payloads_match_delivery_order_tabs() -> None:
    payloads = {
        spec.status_name: ship._status_payload(spec, 697476809, page=1, pre_page=1)
        for spec in ship.UNLINKED_SHIPMENT_STATUS_SPECS
    }

    assert payloads["WMS待配货"] == {
        "status": 6,
        "is_batch_create": 1,
        "delivery_type": 2,
        "store": [697476809],
        "page": 1,
        "prePage": 1,
    }
    assert payloads["WMS待装箱"] == {
        "status": 9,
        "is_batch_create": 1,
        "store": [697476809],
        "page": 1,
        "prePage": 1,
    }
    assert payloads["待关联货件"] == {
        "status": 10,
        "is_batch_create": 1,
        "store": [697476809],
        "page": 1,
        "prePage": 1,
    }


def test_create_unlinked_export_task_sends_taskreport_payload(monkeypatch) -> None:
    fake_session = _FakeSession(
        [
            _FakeResponse(
                {
                    "code": 200,
                    "msg": "success",
                    "data": {"taskId": 370502},
                }
            )
        ]
    )
    monkeypatch.setattr(ship, "erp_http_session", fake_session)

    task_id = asyncio.run(
        ship.create_unlinked_export_task(
            ship.UNLINKED_SHIPMENT_STATUS_SPECS[1],
            697476809,
            token="token",
            report_date="2026-06-12",
        )
    )

    assert task_id == 370502
    call = fake_session.calls[0]
    assert call["method"] == "POST"
    assert call["json"] == {
        "reportEndDate": "2026-06-12",
        "reportStartDate": "2026-06-12",
        "simpleTaskConfigId": "amz-fba-batch-delivery",
        "reportParams": {
            "status": 9,
            "is_batch_create": 1,
            "store": [697476809],
            "page": 1,
            "prePage": 20,
            "ids": [],
            "export_type": "1",
            "currency_type": "1",
            "entry_type": "",
        },
    }


def test_download_raw_file_from_url_saves_original_binary(monkeypatch, tmp_path) -> None:
    fake_session = _FakeSession([_FakeResponse(body=b"raw-bytes")])
    monkeypatch.setattr(ship, "external_http_session", fake_session)

    path = asyncio.run(
        ship.download_raw_file_from_url(
            "https://cos.example.test/file.xlsx",
            store_name="Amazon-Test-US",
            status_name="WMS待装箱",
            task_id=370502,
            file_name="fba报表-发货单.xlsx",
            output_dir=tmp_path,
            download_time="202606121730",
        )
    )

    assert path == tmp_path / "202606121730-Amazon-Test-US-WMS待装箱-370502.xlsx"
    assert path.read_bytes() == b"raw-bytes"
    assert "Authorization" not in fake_session.calls[0].get("headers", {})


def test_download_store_unlinked_shipments_skips_zero_totals_and_polls_with_min_interval(monkeypatch, tmp_path) -> None:
    async def fake_get_token() -> str:
        return "token"

    async def fake_resolve(store_name: str, *, token: str | None = None) -> ship.ShopOption:
        assert store_name == "Amazon-Test-US"
        assert token == "token"
        return ship.ShopOption(store_id=697476809, name=store_name, raw={})

    totals = {"WMS待配货": 0, "WMS待装箱": 3, "待关联货件": 2}
    created: list[str] = []
    wait_calls: list[dict] = []

    async def fake_fetch_total(spec: ship.UnlinkedShipmentStatusSpec, store_id: int, *, token: str | None = None) -> int:
        assert store_id == 697476809
        assert token == "token"
        return totals[spec.status_name]

    async def fake_create_task(
        spec: ship.UnlinkedShipmentStatusSpec,
        store_id: int,
        *,
        token: str | None = None,
        report_date=None,
    ) -> int:
        created.append(spec.status_name)
        return 370500 + len(created)

    async def fake_wait_task(task_id: int, *, token: str | None = None, timeout_sec: float, poll_interval_sec: float):
        wait_calls.append(
            {
                "task_id": task_id,
                "token": token,
                "timeout_sec": timeout_sec,
                "poll_interval_sec": poll_interval_sec,
            }
        )
        return BatchDeliveryTask(
            task_id=task_id,
            file_hash=f"hash-{task_id}",
            file_name=f"task-{task_id}.csv",
            task_status=2,
            task_status_text="处理完成",
            raw={},
        )

    async def fake_download_info(task_id: int, file_hash: str, *, token: str | None = None):
        return BatchDeliveryDownloadInfo(
            task_id=task_id,
            file_hash=file_hash,
            file_name=f"download-{task_id}.csv",
            download_url=f"https://cos.example.test/{task_id}.csv",
            raw={},
        )

    async def fake_download_raw(download_url: str, **kwargs):
        return Path(tmp_path) / f"{kwargs['status_name']}-{kwargs['task_id']}.csv"

    monkeypatch.setattr(ship, "get_fba_free_token", fake_get_token)
    monkeypatch.setattr(ship, "resolve_shop_option", fake_resolve)
    monkeypatch.setattr(ship, "fetch_status_total", fake_fetch_total)
    monkeypatch.setattr(ship, "create_unlinked_export_task", fake_create_task)
    monkeypatch.setattr(ship, "wait_for_delivery_task", fake_wait_task)
    monkeypatch.setattr(ship, "request_download_info", fake_download_info)
    monkeypatch.setattr(ship, "download_raw_file_from_url", fake_download_raw)

    result = asyncio.run(
        ship.download_store_unlinked_shipments(
            "Amazon-Test-US",
            timeout_sec=30,
            poll_interval_sec=1,
            output_dir=tmp_path,
            download_time="202606121730",
        )
    )

    assert result.store_id == 697476809
    assert result.download_time == "202606121730"
    assert created == ["WMS待装箱", "待关联货件"]
    assert [row.to_payload() for row in result.status_results] == [
        {
            "status_name": "WMS待配货",
            "total": 0,
            "task_id": None,
            "file_hash": "",
            "file_name": "",
            "raw_file_path": "",
        },
        {
            "status_name": "WMS待装箱",
            "total": 3,
            "task_id": 370501,
            "file_hash": "hash-370501",
            "file_name": "download-370501.csv",
            "raw_file_path": str(tmp_path / "WMS待装箱-370501.csv"),
        },
        {
            "status_name": "待关联货件",
            "total": 2,
            "task_id": 370502,
            "file_hash": "hash-370502",
            "file_name": "download-370502.csv",
            "raw_file_path": str(tmp_path / "待关联货件-370502.csv"),
        },
    ]
    assert wait_calls == [
        {"task_id": 370501, "token": "token", "timeout_sec": 30.0, "poll_interval_sec": 10.0},
        {"task_id": 370502, "token": "token", "timeout_sec": 30.0, "poll_interval_sec": 10.0},
    ]
