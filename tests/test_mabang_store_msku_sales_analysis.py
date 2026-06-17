from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from services.mabang.amazon.fba import store_msku_sales_analysis as analysis


SOURCE_COLUMNS = [
    "父ASIN",
    "ASIN",
    "MSKU",
    "商品链接",
    "7天销量",
    "14天销量",
    "30天销量",
    "90天销量",
    "店铺名称",
]


def _write_source_xlsx(path: Path, rows: list[dict], *, columns: list[str] | None = None) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    headers = columns or SOURCE_COLUMNS
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(column, "") for column in headers])
    workbook.save(path)
    workbook.close()
    return path


def _load_sheet_records(path: Path, sheet_name: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        return [
            dict(zip(headers, values, strict=False))
            for values in worksheet.iter_rows(min_row=2, values_only=True)
        ]
    finally:
        workbook.close()


def _load_sheet_headers(path: Path, sheet_name: str) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return [cell.value for cell in worksheet[1]]
    finally:
        workbook.close()


def _assert_standard_dimensions(path: Path, sheet_names: tuple[str, ...]) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    try:
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            assert worksheet.sheet_format.defaultRowHeight == 15
            for row_index in range(1, worksheet.max_row + 1):
                assert worksheet.row_dimensions[row_index].height == 15
            for column_index in range(1, worksheet.max_column + 1):
                assert worksheet.column_dimensions[get_column_letter(column_index)].width == 15
    finally:
        workbook.close()


def test_find_latest_store_msku_file_uses_timestamp_prefix_and_ignores_malformed(tmp_path) -> None:
    store_name = "Amazon-Lerxiuer-FR"
    older = _write_source_xlsx(
        tmp_path / "202605241200-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx",
        [{"父ASIN": "P1", "ASIN": "A1", "MSKU": "M1"}],
    )
    latest = _write_source_xlsx(
        tmp_path / "202605251530-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx",
        [{"父ASIN": "P2", "ASIN": "A2", "MSKU": "M2"}],
    )
    _write_source_xlsx(
        tmp_path / "not-a-date-Amazon-Lerxiuer-FR_msku_data.xlsx",
        [{"父ASIN": "P3", "ASIN": "A3", "MSKU": "M3"}],
    )

    result = analysis.find_latest_store_msku_file(store_name, input_dir=tmp_path)

    assert result.path == latest
    assert result.path != older
    assert result.source_data_time == "202605251530"


def test_find_latest_store_msku_file_errors_when_missing(tmp_path) -> None:
    with pytest.raises(analysis.StoreMskuSalesAnalysisError, match="未找到本地店铺MSKU数据文件"):
        analysis.find_latest_store_msku_file("Amazon-Lerxiuer-FR", input_dir=tmp_path)


def test_find_latest_store_msku_file_accepts_legacy_english_file_name(tmp_path) -> None:
    legacy = _write_source_xlsx(
        tmp_path / "202605251530-Amazon-Lerxiuer-FR_msku_data.xlsx",
        [{"父ASIN": "P1", "ASIN": "A1", "MSKU": "M1"}],
    )

    result = analysis.find_latest_store_msku_file("Amazon-Lerxiuer-FR", input_dir=tmp_path)

    assert result.path == legacy
    assert result.source_data_time == "202605251530"


def test_compute_sales_metrics_handles_weighted_speed_and_trend_branches() -> None:
    metrics = analysis.compute_sales_metrics("70", "112", "240")

    assert metrics.weighted_daily_sales == pytest.approx(9.2)
    assert metrics.trend_ratio == pytest.approx(1.470588, rel=1e-5)
    assert metrics.trend == "增长"
    assert analysis.compute_sales_metrics(7, 8, 9).trend == "样本不足"
    assert analysis.compute_sales_metrics(14, 14, 14).trend == "新增出单/恢复出单"
    assert analysis.compute_sales_metrics(0, -7, 17).trend == "无销量"
    assert analysis.compute_sales_metrics(7, 21, 100).trend == "快速下降"


def test_analyze_generates_report_with_aggregations_detail_columns_and_stale_flag(tmp_path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    source_path = _write_source_xlsx(
        input_dir / "202605251530-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx",
        [
            {
                "父ASIN": "PARENT-1",
                "ASIN": "ASIN-1",
                "MSKU": "MSKU-1",
                "商品链接": "美国 http://www.amazon.com/gp/product/ASIN-1",
                "7天销量": "70",
                "14天销量": "112",
                "30天销量": "240",
                "90天销量": "500",
                "店铺名称": "Amazon-Lerxiuer-FR",
            },
            {
                "父ASIN": "PARENT-1",
                "ASIN": "ASIN-2",
                "MSKU": "MSKU-2",
                "商品链接": "美国 https://www.amazon.com/dp/ASIN-2",
                "7天销量": "7",
                "14天销量": "14",
                "30天销量": "30",
                "90天销量": "80",
                "店铺名称": "Amazon-Lerxiuer-FR",
            },
            {
                "父ASIN": "PARENT-2",
                "ASIN": "ASIN-1",
                "MSKU": "MSKU-3",
                "商品链接": "德国 https://www.amazon.de/dp/ASIN-1",
                "7天销量": "1,400",
                "14天销量": "",
                "30天销量": "bad",
                "90天销量": "10",
                "店铺名称": "Amazon-Lerxiuer-FR",
            },
        ],
    )

    result = analysis.analyze_store_msku_sales(
        "Amazon-Lerxiuer-FR",
        input_dir=input_dir,
        output_dir=output_dir,
        today=date(2026, 5, 26),
    )
    payload = result.to_payload()

    assert payload["success"] is True
    assert payload["store_name"] == "Amazon-Lerxiuer-FR"
    assert payload["source_xlsx_path"] == str(source_path)
    assert payload["source_data_time"] == "202605251530"
    assert payload["data_is_stale"] is True
    assert payload["link_count"] == 2
    assert payload["asin_count"] == 3
    assert payload["msku_count"] == 3
    assert payload["report_xlsx_path"] == str(output_dir / "202605251530-Amazon-Lerxiuer-FR_销量分析.xlsx")

    report_path = Path(result.report_xlsx_path)
    assert report_path.is_file()
    _assert_standard_dimensions(report_path, analysis.REPORT_SHEETS)
    link_records = _load_sheet_records(report_path, analysis.LINK_TOP_SHEET)
    parent_1 = next(row for row in link_records if row["父ASIN"] == "PARENT-1")
    assert parent_1["MSKU数"] == 2
    assert parent_1["ASIN数"] == 2
    assert parent_1["30天销量"] == 270
    assert parent_1["商品链接"] == "http://www.amazon.com/gp/product/PARENT-1"

    asin_records = _load_sheet_records(report_path, analysis.ASIN_TOP_SHEET)
    asin_1_rows = [row for row in asin_records if row["ASIN"] == "ASIN-1"]
    assert len(asin_1_rows) == 2
    asin_1 = next(row for row in asin_1_rows if row["父ASIN"] == "PARENT-1")
    assert asin_1["MSKU"] == "MSKU-1"
    assert asin_1["30天销量"] == 240
    assert asin_1["商品链接"] == "美国 http://www.amazon.com/gp/product/ASIN-1"
    asin_headers = _load_sheet_headers(report_path, analysis.ASIN_TOP_SHEET)
    assert "MSKU" in asin_headers
    assert "MSKU数" not in asin_headers

    for sheet_name in [
        analysis.LINK_TOP_SHEET,
        analysis.LINK_OTHER_SHEET,
        analysis.ASIN_TOP_SHEET,
        analysis.ASIN_OTHER_SHEET,
    ]:
        headers = _load_sheet_headers(report_path, sheet_name)
        assert "商品链接" in headers
        assert headers.index("商品链接") == headers.index("加权日销") - 1

    detail_records = _load_sheet_records(report_path, analysis.DETAIL_SHEET)
    assert {"加权日销", "销量趋势速率", "销量趋势"}.issubset(detail_records[0])
    assert detail_records[0]["加权日销"] == pytest.approx(9.2)
    assert detail_records[0]["销量趋势"] == "增长"
    assert detail_records[2]["30天销量"] == 0


def test_report_splits_top_links_and_asins_by_30_day_sales(tmp_path) -> None:
    rows = [
        {
            "父ASIN": f"PARENT-{index:02d}",
            "ASIN": f"ASIN-{index:02d}",
            "MSKU": f"MSKU-{index:02d}",
            "商品链接": f"美国 http://www.amazon.com/gp/product/ASIN-{index:02d}",
            "7天销量": index,
            "14天销量": index * 2,
            "30天销量": index * 10,
            "90天销量": index * 30,
        }
        for index in range(1, 52)
    ]
    _write_source_xlsx(tmp_path / "input" / "202605261010-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx", rows)

    result = analysis.analyze_store_msku_sales(
        "Amazon-Lerxiuer-FR",
        input_dir=tmp_path / "input",
        output_dir=tmp_path / "output",
        today=date(2026, 5, 26),
    )

    report_path = Path(result.report_xlsx_path)
    assert analysis.LINK_TOP_SHEET == "链接销量前10"
    assert _load_sheet_records(report_path, analysis.LINK_TOP_SHEET)[0]["父ASIN"] == "PARENT-51"
    assert len(_load_sheet_records(report_path, analysis.LINK_TOP_SHEET)) == 10
    assert len(_load_sheet_records(report_path, analysis.LINK_OTHER_SHEET)) == 41
    assert _load_sheet_records(report_path, analysis.LINK_OTHER_SHEET)[0]["父ASIN"] == "PARENT-41"
    assert len(_load_sheet_records(report_path, analysis.ASIN_TOP_SHEET)) == 50
    assert _load_sheet_records(report_path, analysis.ASIN_OTHER_SHEET)[0]["ASIN"] == "ASIN-01"
    assert result.data_is_stale is False


def test_missing_required_columns_error(tmp_path) -> None:
    _write_source_xlsx(
        tmp_path / "202605251530-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx",
        [{"父ASIN": "P1", "ASIN": "A1", "MSKU": "M1"}],
        columns=["父ASIN", "ASIN", "MSKU", "7天销量", "14天销量", "30天销量", "90天销量"],
    )

    with pytest.raises(analysis.StoreMskuSalesAnalysisError, match="缺少列: 商品链接"):
        analysis.analyze_store_msku_sales("Amazon-Lerxiuer-FR", input_dir=tmp_path, output_dir=tmp_path / "out")


def test_duplicate_asin_parent_msku_errors(tmp_path) -> None:
    _write_source_xlsx(
        tmp_path / "202605251530-Amazon-Lerxiuer-FR_店铺MSKU数据.xlsx",
        [
            {
                "父ASIN": "PARENT-1",
                "ASIN": "ASIN-1",
                "MSKU": "MSKU-1",
                "商品链接": "美国 http://www.amazon.com/gp/product/ASIN-1",
                "7天销量": 10,
                "14天销量": 20,
                "30天销量": 30,
                "90天销量": 90,
            },
            {
                "父ASIN": "PARENT-1",
                "ASIN": "ASIN-1",
                "MSKU": "MSKU-1",
                "商品链接": "美国 http://www.amazon.com/gp/product/ASIN-1",
                "7天销量": 1,
                "14天销量": 2,
                "30天销量": 3,
                "90天销量": 9,
            },
        ],
    )

    with pytest.raises(
        analysis.StoreMskuSalesAnalysisError,
        match="ASIN表存在重复售卖项: ASIN=ASIN-1, 父ASIN=PARENT-1, MSKU=MSKU-1",
    ):
        analysis.analyze_store_msku_sales(
            "Amazon-Lerxiuer-FR",
            input_dir=tmp_path,
            output_dir=tmp_path / "out",
        )
