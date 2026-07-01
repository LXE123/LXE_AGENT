from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill, Side

from services.agent_cli.mabang import fill_purchase_contracts as cli
from services.agent_cli.mabang import generate_restock_workbook as purchase_summary


def _purchase_row(
    *,
    manufacturer: str,
    model: str,
    quantity: object,
    unit_price: object,
    total_price: object,
    product_name: str = "合同产品A",
    unit: str = "条",
    tax_rate: str = "13%",
) -> list[object]:
    values = {
        "库存sku": "SKU-A",
        "产品名称": "产品A",
        "来源SP单号": "SP260508022",
        "库存sku（第一行）": "SKU-A",
        "产品名称（第一行）": "产品A",
        "型号": model,
        "原价": unit_price,
        "厂家": manufacturer,
        "单位": unit,
        "合同产品名称": product_name,
        "合同编号前缀": "JY",
        "税率": tax_rate,
        "数量": quantity,
        "总价": total_price,
    }
    return [values.get(column) for column in purchase_summary.MANUFACTURER_COLUMNS]


def _write_purchase_summary(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = purchase_summary.SUMMARY_SHEET_NAME
    worksheet.append(list(purchase_summary.MANUFACTURER_COLUMNS))
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def _style_detail_row(worksheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="FFE2F0D9")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in worksheet[5]:
        cell.fill = fill
        cell.border = border
    worksheet.row_dimensions[5].height = 33


def _write_contract_sheet(worksheet, *, contract_no: str = "合同编号：KEEP", include_model: bool = True) -> None:
    worksheet["A1"] = "采购合同"
    worksheet["E2"] = f"{contract_no}\nDate: 2000年1月1日"
    worksheet["E3"] = "交货日期：2000年1月4日\n付款期限：发货验收付款\n币种：人民币 税率：0%"
    headers = ["序号", "产品名称", "规格型号", "单位", "数量", "含税单价", "含税金额（元）", "备注"]
    detail_row = ["旧序号", "旧产品", "旧型号", "旧单位", 0, 0, 0, "旧备注"]
    summary_row = ["合计", None, None, None, None, None, 0, None]
    if not include_model:
        headers.pop(2)
        detail_row.pop(2)
        summary_row.pop(2)
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=4, column=column_index, value=header)
    worksheet.append(detail_row)
    worksheet.append(summary_row)
    worksheet.column_dimensions["B"].width = 24
    _style_detail_row(worksheet)


def _write_addendum_template_sheet(worksheet) -> None:
    worksheet["A1"] = "补充协议附加件明细"
    worksheet["A2"] = "采购合同编号：KEEP-ADDENDUM"
    headers = ["序号", "产品名称", "规格型号", "单位", "数量", "含税单价", "含税金额（元）", "备注"]
    detail_row = ["旧序号", "旧产品", "旧型号", "旧单位", 0, 0, 0, "旧备注"]
    summary_row = ["合计", None, None, None, None, None, 0, None]
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=4, column=column_index, value=header)
    worksheet.append(detail_row)
    worksheet.append(summary_row)
    worksheet.merge_cells("A1:H1")
    worksheet.column_dimensions["B"].width = 31
    _style_detail_row(worksheet)


def _merge_tax_amount_header_and_add_blank_detail_rows(worksheet) -> None:
    worksheet.insert_cols(8)
    worksheet.merge_cells("G4:H4")
    worksheet.insert_rows(6, amount=2)


def _write_contract_template(path: Path, sheet_names: list[str], *, include_addendum: bool = True) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in sheet_names:
        worksheet = workbook.create_sheet(sheet_name)
        _write_contract_sheet(worksheet, contract_no=f"合同编号：KEEP-{sheet_name}")
    if include_addendum:
        addendum_worksheet = workbook.create_sheet(cli.ADDENDUM_TEMPLATE_SHEET)
        _write_addendum_template_sheet(addendum_worksheet)
    workbook.save(path)


def _write_contract_template_without_model(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _write_contract_sheet(worksheet, contract_no=f"合同编号：KEEP-{sheet_name}", include_model=False)
    addendum_worksheet = workbook.create_sheet(cli.ADDENDUM_TEMPLATE_SHEET)
    _write_addendum_template_sheet(addendum_worksheet)
    workbook.save(path)


def _write_contract_template_with_malformed_detail_rows(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _write_contract_sheet(worksheet, contract_no=f"合同编号：KEEP-{sheet_name}")
    worksheet.insert_rows(6, amount=3)
    for row_index in range(6, 9):
        worksheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=5)
        worksheet.cell(row=row_index, column=2, value=f"错误合并{row_index}")
        worksheet.cell(row=row_index, column=7, value="旧金额")
        worksheet.row_dimensions[row_index].height = 55
    addendum_worksheet = workbook.create_sheet(cli.ADDENDUM_TEMPLATE_SHEET)
    _write_addendum_template_sheet(addendum_worksheet)
    workbook.save(path)


def _write_contract_template_with_empty_tax_value(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _write_contract_sheet(worksheet, contract_no=f"合同编号：KEEP-{sheet_name}")
    worksheet["E3"] = "交货日期：2000年1月4日\n付款期限：发货验收付款\n币种：人民币 税率："
    addendum_worksheet = workbook.create_sheet(cli.ADDENDUM_TEMPLATE_SHEET)
    _write_addendum_template_sheet(addendum_worksheet)
    workbook.save(path)


def _write_contract_template_with_merged_summary_and_terms(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _write_contract_sheet(worksheet, contract_no=f"合同编号：KEEP-{sheet_name}")
    worksheet.merge_cells("A6:F6")
    worksheet["A7"] = "销售条款：保留格式"
    worksheet.merge_cells("A7:H7")
    addendum_worksheet = workbook.create_sheet(cli.ADDENDUM_TEMPLATE_SHEET)
    _write_addendum_template_sheet(addendum_worksheet)
    workbook.save(path)


def _write_contract_template_with_merged_tax_amount_header(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    _write_contract_sheet(worksheet, contract_no=f"合同编号：KEEP-{sheet_name}")
    _merge_tax_amount_header_and_add_blank_detail_rows(worksheet)
    addendum_worksheet = workbook.create_sheet(cli.ADDENDUM_TEMPLATE_SHEET)
    _write_addendum_template_sheet(addendum_worksheet)
    _merge_tax_amount_header_and_add_blank_detail_rows(addendum_worksheet)
    workbook.save(path)


def _sheet_values(path: Path, sheet_name: str, cell_range: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return [
            tuple(cell.value for cell in row)
            for row in worksheet[cell_range]
        ]
    finally:
        workbook.close()


def _workbook_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _cell_value(path: Path, sheet_name: str, cell: str) -> object:
    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet_name][cell].value
    finally:
        workbook.close()


def _cell_fill(path: Path, sheet_name: str, cell: str) -> str:
    workbook = load_workbook(path, data_only=True)
    try:
        return str(workbook[sheet_name][cell].fill.fgColor.rgb)
    finally:
        workbook.close()


def _row_height(path: Path, sheet_name: str, row_index: int) -> float | None:
    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet_name].row_dimensions[row_index].height
    finally:
        workbook.close()


def _column_width(path: Path, sheet_name: str, column: str) -> float | None:
    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet_name].column_dimensions[column].width
    finally:
        workbook.close()


def _single_row_merged_ranges(path: Path, sheet_name: str, row_index: int) -> list[str]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return sorted(
            str(merged_range)
            for merged_range in worksheet.merged_cells.ranges
            if merged_range.min_row == row_index and merged_range.max_row == row_index
        )
    finally:
        workbook.close()


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


async def _noop_close_all_network_clients() -> None:
    return None


def test_fill_purchase_contracts_generates_one_file_per_manufacturer(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [
            _purchase_row(manufacturer="厂家A", model="JY-1", quantity=300, unit_price=6.8, total_price=2040),
            _purchase_row(
                manufacturer="厂家B",
                model="JZ-19",
                quantity=10,
                unit_price=2,
                total_price=20,
                product_name="合同产品B",
                unit="个",
                tax_rate="9%",
            ),
        ],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template(template_path, ["厂家A", "深圳厂家B模板", "无关模板"])

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    assert payload["success"] is True
    assert payload["generated_count"] == 2
    assert payload["warnings"] == []
    output_by_manufacturer = {
        item["manufacturer"]: Path(item["output_xlsx"])
        for item in payload["output_files"]
    }
    assert _workbook_sheet_names(output_by_manufacturer["厂家A"]) == ["厂家A", cli.ADDENDUM_OUTPUT_SHEET]
    assert _workbook_sheet_names(output_by_manufacturer["厂家B"]) == ["深圳厂家B模板", cli.ADDENDUM_OUTPUT_SHEET]
    assert _sheet_values(output_by_manufacturer["厂家A"], "厂家A", "A5:G6") == [
        (1, "合同产品A", "JY-1", "条", 300, 6.8, 2040),
        ("合计", None, None, None, None, None, 2040),
    ]
    assert _sheet_values(output_by_manufacturer["厂家A"], cli.ADDENDUM_OUTPUT_SHEET, "A5:G6") == [
        (1, "合同产品A", "JY-1", "条", 300, 6.8, 2040),
        ("合计", None, None, None, None, None, 2040),
    ]
    assert _cell_value(output_by_manufacturer["厂家A"], cli.ADDENDUM_OUTPUT_SHEET, "A2") == "采购合同编号：KEEP-ADDENDUM"
    assert _single_row_merged_ranges(output_by_manufacturer["厂家A"], cli.ADDENDUM_OUTPUT_SHEET, 1) == ["A1:H1"]
    assert _row_height(output_by_manufacturer["厂家A"], cli.ADDENDUM_OUTPUT_SHEET, 5) == 33
    assert _column_width(output_by_manufacturer["厂家A"], cli.ADDENDUM_OUTPUT_SHEET, "B") == 31
    assert cli.ADDENDUM_TEMPLATE_SHEET not in _workbook_sheet_names(output_by_manufacturer["厂家A"])
    assert "Date: 2026年7月1日" in str(_cell_value(output_by_manufacturer["厂家A"], "厂家A", "E2"))
    assert "交货日期：2026年7月4日" in str(_cell_value(output_by_manufacturer["厂家A"], "厂家A", "E3"))
    assert "税率：13%" in str(_cell_value(output_by_manufacturer["厂家A"], "厂家A", "E3"))
    assert "税率：9%" in str(_cell_value(output_by_manufacturer["厂家B"], "深圳厂家B模板", "E3"))


def test_fill_purchase_contracts_skips_missing_and_ambiguous_template_sheets(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [
            _purchase_row(manufacturer="厂家A", model="A-1", quantity=1, unit_price=1, total_price=1),
            _purchase_row(manufacturer="厂家B", model="B-1", quantity=1, unit_price=1, total_price=1),
            _purchase_row(manufacturer="厂家C", model="C-1", quantity=1, unit_price=1, total_price=1),
        ],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template(template_path, ["厂家A", "厂家B模板1", "厂家B模板2"])

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    assert payload["generated_count"] == 1
    assert payload["skipped_manufacturer_count"] == 2
    assert payload["skipped_manufacturers"] == ["厂家B", "厂家C"]
    assert payload["warnings"] == [
        "合同模板中厂家 `厂家B` 匹配到多个 sheet: 厂家B模板1, 厂家B模板2，已跳过",
        "合同模板中未找到厂家 `厂家C` 对应的 sheet，已跳过",
    ]


def test_fill_purchase_contracts_rejects_missing_addendum_template(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [_purchase_row(manufacturer="厂家A", model="A-1", quantity=1, unit_price=1, total_price=1)],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template(template_path, ["厂家A"], include_addendum=False)

    with pytest.raises(RuntimeError, match="合同汇总模板缺少 sheet: 附加件明细模板"):
        cli.fill_purchase_contracts(
            purchase_summary_xlsx=purchase_path,
            contract_template_xlsx=template_path,
            output_dir=tmp_path / "out",
            today=date(2026, 7, 1),
        )


def test_fill_purchase_contracts_applies_header_column_spans_to_detail_rows(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [_purchase_row(manufacturer="厂家A", model="A-1", quantity=2, unit_price=3, total_price=6)],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template_with_merged_tax_amount_header(template_path, "厂家A")

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    output_path = Path(payload["output_files"][0]["output_xlsx"])
    for sheet_name in ("厂家A", cli.ADDENDUM_OUTPUT_SHEET):
        assert "G4:H4" in _single_row_merged_ranges(output_path, sheet_name, 4)
        assert _sheet_values(output_path, sheet_name, "A5:I8") == [
            (1, "合同产品A", "A-1", "条", 2, 3, 6, None, None),
            (None, None, None, None, None, None, None, None, None),
            (None, None, None, None, None, None, None, None, None),
            ("合计", None, None, None, None, None, 6, None, None),
        ]
        for row_index in range(5, 9):
            assert f"G{row_index}:H{row_index}" in _single_row_merged_ranges(
                output_path,
                sheet_name,
                row_index,
            )


def test_fill_purchase_contracts_inserts_rows_preserves_style_and_keeps_contract_no(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [
            _purchase_row(manufacturer="厂家A", model="A-1", quantity=2, unit_price=3, total_price=6),
            _purchase_row(manufacturer="厂家A", model="A-2", quantity=4, unit_price=5, total_price=20),
        ],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template(template_path, ["厂家A"])

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    output_path = Path(payload["output_files"][0]["output_xlsx"])
    assert "合同编号：KEEP-厂家A" in str(_cell_value(output_path, "厂家A", "E2"))
    assert _sheet_values(output_path, "厂家A", "A5:G7") == [
        (1, "合同产品A", "A-1", "条", 2, 3, 6),
        (2, "合同产品A", "A-2", "条", 4, 5, 20),
        ("合计", None, None, None, None, None, 26),
    ]
    assert _cell_fill(output_path, "厂家A", "B6") == _cell_fill(output_path, "厂家A", "B5")
    assert _row_height(output_path, "厂家A", 6) == 33


def test_fill_purchase_contracts_resets_malformed_detail_row_merges(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [
            _purchase_row(manufacturer="厂家A", model="A-1", quantity=2, unit_price=3, total_price=6),
            _purchase_row(manufacturer="厂家A", model="A-2", quantity=4, unit_price=5, total_price=20),
        ],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template_with_malformed_detail_rows(template_path, "厂家A")

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    output_path = Path(payload["output_files"][0]["output_xlsx"])
    assert _sheet_values(output_path, "厂家A", "A5:G9") == [
        (1, "合同产品A", "A-1", "条", 2, 3, 6),
        (2, "合同产品A", "A-2", "条", 4, 5, 20),
        (None, None, None, None, None, None, None),
        (None, None, None, None, None, None, None),
        ("合计", None, None, None, None, None, 26),
    ]
    assert _single_row_merged_ranges(output_path, "厂家A", 6) == _single_row_merged_ranges(
        output_path,
        "厂家A",
        5,
    )
    assert _single_row_merged_ranges(output_path, "厂家A", 7) == _single_row_merged_ranges(
        output_path,
        "厂家A",
        5,
    )
    assert _cell_fill(output_path, "厂家A", "B7") == _cell_fill(output_path, "厂家A", "B5")
    assert _row_height(output_path, "厂家A", 7) == 33


def test_fill_purchase_contracts_preserves_summary_and_terms_merges_when_inserting_rows(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [
            _purchase_row(manufacturer="厂家A", model="A-1", quantity=2, unit_price=3, total_price=6),
            _purchase_row(manufacturer="厂家A", model="A-2", quantity=4, unit_price=5, total_price=20),
            _purchase_row(manufacturer="厂家A", model="A-3", quantity=6, unit_price=5, total_price=30),
            _purchase_row(manufacturer="厂家A", model="A-4", quantity=8, unit_price=5, total_price=40),
        ],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template_with_merged_summary_and_terms(template_path, "厂家A")

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    output_path = Path(payload["output_files"][0]["output_xlsx"])
    assert _sheet_values(output_path, "厂家A", "A5:H10") == [
        (1, "合同产品A", "A-1", "条", 2, 3, 6, None),
        (2, "合同产品A", "A-2", "条", 4, 5, 20, None),
        (3, "合同产品A", "A-3", "条", 6, 5, 30, None),
        (4, "合同产品A", "A-4", "条", 8, 5, 40, None),
        ("合计", None, None, None, None, None, 96, None),
        ("销售条款：保留格式", None, None, None, None, None, None, None),
    ]
    assert _single_row_merged_ranges(output_path, "厂家A", 9) == ["A9:F9"]
    assert _single_row_merged_ranges(output_path, "厂家A", 10) == ["A10:H10"]
    assert _single_row_merged_ranges(output_path, "厂家A", 6) == []
    assert _single_row_merged_ranges(output_path, "厂家A", 7) == []


def test_fill_detail_rows_tolerates_dirty_merged_cell_records():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "厂家A"
    _write_contract_sheet(worksheet)
    worksheet.insert_rows(6, amount=3)
    for row_index in range(6, 9):
        worksheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=5)
        worksheet.cell(row=row_index, column=2, value=f"错误合并{row_index}")
        worksheet.row_dimensions[row_index].height = 55
        worksheet._cells.pop((row_index, 3), None)

    row_count = cli._fill_detail_rows(
        worksheet,
        [
            cli.PurchaseContractLine(
                manufacturer="厂家A",
                product_name="合同产品A",
                model="A-1",
                unit="条",
                quantity=Decimal("2"),
                tax_unit_price=Decimal("3"),
                tax_amount=Decimal("6"),
                tax_rate="13%",
            ),
            cli.PurchaseContractLine(
                manufacturer="厂家A",
                product_name="合同产品A",
                model="A-2",
                unit="条",
                quantity=Decimal("4"),
                tax_unit_price=Decimal("5"),
                tax_amount=Decimal("20"),
                tax_rate="13%",
            ),
        ],
    )

    assert row_count == 2
    assert [
        tuple(worksheet.cell(row=row_index, column=column_index).value for column_index in range(1, 8))
        for row_index in range(5, 10)
    ] == [
        (1, "合同产品A", "A-1", "条", 2, 3, 6),
        (2, "合同产品A", "A-2", "条", 4, 5, 20),
        (None, None, None, None, None, None, None),
        (None, None, None, None, None, None, None),
        ("合计", None, None, None, None, None, 26),
    ]
    assert sorted(
        str(merged_range)
        for merged_range in worksheet.merged_cells.ranges
        if merged_range.min_row == 6 and merged_range.max_row == 6
    ) == sorted(
        str(merged_range)
        for merged_range in worksheet.merged_cells.ranges
        if merged_range.min_row == 5 and merged_range.max_row == 5
    )


def test_fill_purchase_contracts_does_not_pollute_delivery_date_when_tax_value_is_missing(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [_purchase_row(manufacturer="厂家A", model="A-1", quantity=2, unit_price=3, total_price=6)],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template_with_empty_tax_value(template_path, "厂家A")

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    output_path = Path(payload["output_files"][0]["output_xlsx"])
    text = str(_cell_value(output_path, "厂家A", "E3"))
    assert "交货日期：2026年7月4日" in text
    assert "交货日期：13%" not in text
    assert "税率：13%" not in text
    assert payload["warnings"] == ["厂家 `厂家A` 合同模板未找到税率位置"]


def test_fill_purchase_contracts_allows_missing_model_column(tmp_path):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [_purchase_row(manufacturer="厂家A", model="A-1", quantity=2, unit_price=3, total_price=6)],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template_without_model(template_path, "厂家A")

    payload = cli.fill_purchase_contracts(
        purchase_summary_xlsx=purchase_path,
        contract_template_xlsx=template_path,
        output_dir=tmp_path / "out",
        today=date(2026, 7, 1),
    )

    assert payload["generated_count"] == 1
    output_path = Path(payload["output_files"][0]["output_xlsx"])
    assert _sheet_values(output_path, "厂家A", "A5:F6") == [
        (1, "合同产品A", "条", 2, 3, 6),
        ("合计", None, None, None, None, 6),
    ]


def test_fill_purchase_contracts_main_outputs_success_json(monkeypatch, tmp_path, capsys):
    purchase_path = tmp_path / "purchase_summary.xlsx"
    _write_purchase_summary(
        purchase_path,
        [_purchase_row(manufacturer="厂家A", model="A-1", quantity=1, unit_price=1, total_price=1)],
    )
    template_path = tmp_path / "contract_template.xlsx"
    _write_contract_template(template_path, ["厂家A"])
    monkeypatch.setattr(cli, "OUTPUT_DIR", tmp_path / "out")
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main(
        [
            "--purchase-summary-xlsx",
            str(purchase_path),
            "--contract-template-xlsx",
            str(template_path),
        ]
    )

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["source"] == "fba_purchase_contract_fill"
    assert payload["generated_count"] == 1
