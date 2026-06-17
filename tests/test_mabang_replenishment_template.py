from __future__ import annotations

import json
from pathlib import Path

import pytest

from agent_runtime.skill_index import load_skill_index
from services.agent_cli.mabang import replenishment_template as cli
from services.mabang.amazon.fba import replenishment_template as tmpl


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


def _cell_value(path: Path, sheet_name: str, row: int, column: int):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet_name].cell(row=row, column=column).value
    finally:
        workbook.close()


def test_default_template_loads_current_algorithm() -> None:
    template = tmpl.load_default_template()

    assert template.name == "默认模板"
    assert template.version == 1
    assert template.params["weighted_sales"] == {
        "7d_weight": 0.6,
        "14d_weight": 0.3,
        "30d_weight": 0.1,
    }
    assert template.params["sea"]["enabled"] is True
    assert tmpl.sea_companion_air_enabled_from_template(template.params) is False
    assert tmpl.replenishment_days_from_template(6, "平稳", template.params) == 75
    assert tmpl.sea_day_candidates_from_template(55, template.params) == [100, 110]
    assert tmpl.validate_template(template).warnings == ()


def test_builtin_templates_include_us_uk_de_and_lin_meiqi_rules() -> None:
    templates = {template.name: template for template in tmpl.load_builtin_templates()}

    assert list(templates) == ["默认模板", "US模板-一组", "UK模板-一组", "DE模板-一组", "2组-US站点-林美淇"]

    us_params = templates["US模板-一组"].params
    assert tmpl.sea_enabled_from_template(us_params) is True
    assert tmpl.replenishment_days_from_template(11, "增长", us_params) == 80
    assert tmpl.replenishment_days_from_template(6, "平稳", us_params) == 75
    assert tmpl.replenishment_days_from_template(2, "下降", us_params) == 70
    assert tmpl.replenishment_days_from_template(0.5, "增长", us_params) == 60
    assert us_params["sea"]["min_daily_sales"] == 5
    assert us_params["sea"]["min_weight_kg"] == 60
    assert tmpl.sea_day_candidates_from_template(6, us_params) == [110]
    assert tmpl.sea_day_candidates_from_template(4, us_params) == []

    uk_params = templates["UK模板-一组"].params
    assert tmpl.sea_enabled_from_template(uk_params) is False
    assert tmpl.replenishment_days_from_template(6, "增长", uk_params) == 85
    assert tmpl.replenishment_days_from_template(2, "平稳", uk_params) == 75
    assert tmpl.replenishment_days_from_template(0.5, "下降", uk_params) == 65
    assert tmpl.sea_day_candidates_from_template(6, uk_params) == []

    de_params = templates["DE模板-一组"].params
    assert tmpl.sea_enabled_from_template(de_params) is False
    assert tmpl.replenishment_days_from_template(11, "增长", de_params) == 90
    assert tmpl.replenishment_days_from_template(6, "平稳", de_params) == 85
    assert tmpl.replenishment_days_from_template(2, "下降", de_params) == 75
    assert tmpl.replenishment_days_from_template(0.5, "增长", de_params) == 65

    lin_params = templates["2组-US站点-林美淇"].params
    assert lin_params["weighted_sales"] == {"7d_weight": 0.7, "14d_weight": 0.2, "30d_weight": 0.1}
    assert tmpl.validate_template(templates["2组-US站点-林美淇"]).warnings == ()
    assert tmpl.replenishment_days_from_template(51, "增长", lin_params) == 90
    assert tmpl.replenishment_days_from_template(11, "平稳", lin_params) == 80
    assert tmpl.replenishment_days_from_template(2, "下降", lin_params) == 75
    assert tmpl.replenishment_days_from_template(1, "增长", lin_params) == 70
    assert tmpl.sea_enabled_from_template(lin_params) is True
    assert tmpl.sea_min_daily_sales_inclusive_from_template(lin_params) is True
    assert tmpl.sea_companion_air_enabled_from_template(lin_params) is True
    assert lin_params["sea"]["min_daily_sales"] == 1
    assert lin_params["sea"]["min_weight_kg"] == 60
    assert lin_params["sea"]["min_net_quantity"] == 30
    assert tmpl.sea_day_candidates_from_template(0.99, lin_params) == []
    assert tmpl.sea_day_candidates_from_template(1, lin_params) == [100]
    assert tmpl.sea_day_candidates_from_template(5, lin_params) == [100]
    assert tmpl.sea_day_candidates_from_template(6, lin_params) == [110]
    assert tmpl.sea_day_candidates_from_template(21, lin_params) == [120]
    assert tmpl.sea_day_candidates_from_template(301, lin_params) == [120]
    assert tmpl.sea_companion_air_day_candidates_from_template(0.99, lin_params) == []
    assert tmpl.sea_companion_air_day_candidates_from_template(1, lin_params) == [70]
    assert tmpl.sea_companion_air_day_candidates_from_template(5, lin_params) == [70]
    assert tmpl.sea_companion_air_day_candidates_from_template(6, lin_params) == [75]
    assert tmpl.sea_companion_air_day_candidates_from_template(21, lin_params) == [80]
    assert tmpl.sea_companion_air_day_candidates_from_template(301, lin_params) == [80]


def test_export_validate_and_import_template_xlsx(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")

    assert exported_path.is_file()
    assert exported_path.name.endswith("_备货模板.xlsx")
    assert _cell_value(exported_path, "模板信息", 2, 2) == "默认模板"
    assert _cell_value(exported_path, "加权日销", 2, 2) == 0.6
    assert _cell_value(exported_path, "海运规则", 2, 5) == "是"

    result = tmpl.validate_template_xlsx(exported_path)
    assert result.template.name == "默认模板"
    assert result.template.params["shipping"]["air_urgent_sales_days_lte"] == 40
    assert result.template.params["sea"]["enabled"] is True

    imported = tmpl.import_template_xlsx(exported_path, store_path=tmp_path / "templates.json")
    assert imported.template.name == "自定义模板1"

    second_imported = tmpl.import_template_xlsx(exported_path, store_path=tmp_path / "templates.json")
    assert second_imported.template.name == "自定义模板2"

    templates = tmpl.list_templates(store_path=tmp_path / "templates.json")
    assert [template.name for template in templates] == [
        "默认模板",
        "US模板-一组",
        "UK模板-一组",
        "DE模板-一组",
        "2组-US站点-林美淇",
        "自定义模板1",
        "自定义模板2",
    ]


def test_import_rejects_duplicate_and_default_names(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")
    tmpl.import_template_xlsx(exported_path, name="老王模板", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="模板名已存在"):
        tmpl.import_template_xlsx(exported_path, name="老王模板", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="不允许导入覆盖"):
        tmpl.import_template_xlsx(exported_path, name="默认模板", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="系统模板不允许导入覆盖"):
        tmpl.import_template_xlsx(exported_path, name="UK模板-一组", store_path=tmp_path / "templates.json")


def test_replace_template_updates_existing_custom_template(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")
    imported = tmpl.import_template_xlsx(exported_path, name="老王模板", store_path=tmp_path / "templates.json")
    assert imported.template.version == 1

    from openpyxl import load_workbook

    workbook = load_workbook(exported_path)
    try:
        workbook["模板信息"].cell(row=2, column=2).value = "xlsx里的新名字"
        workbook["模板信息"].cell(row=4, column=2).value = "替换后的说明"
        workbook["运输方式"].cell(row=2, column=2).value = 35
        workbook.save(exported_path)
    finally:
        workbook.close()

    replaced, old_version = tmpl.replace_template_xlsx(
        exported_path,
        template_name="老王模板",
        store_path=tmp_path / "templates.json",
    )

    assert old_version == 1
    assert replaced.template.name == "老王模板"
    assert replaced.template.version == 2
    assert replaced.template.description == "替换后的说明"
    assert replaced.template.params["shipping"]["air_urgent_sales_days_lte"] == 35


def test_replace_rejects_default_and_missing_template(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="不允许替换"):
        tmpl.replace_template_xlsx(exported_path, template_name="默认模板", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="系统模板不允许替换"):
        tmpl.replace_template_xlsx(exported_path, template_name="US模板-一组", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="只能替换已存在"):
        tmpl.replace_template_xlsx(exported_path, template_name="不存在", store_path=tmp_path / "templates.json")


def test_rename_template_updates_name_without_version_change(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")
    tmpl.import_template_xlsx(exported_path, name="自定义模板A", store_path=tmp_path / "templates.json")

    renamed = tmpl.rename_template("自定义模板A", new_name="夏季备货模板", store_path=tmp_path / "templates.json")

    assert renamed.name == "夏季备货模板"
    assert renamed.version == 1
    templates = tmpl.list_templates(store_path=tmp_path / "templates.json")
    assert [template.name for template in templates] == [
        "默认模板",
        "US模板-一组",
        "UK模板-一组",
        "DE模板-一组",
        "2组-US站点-林美淇",
        "夏季备货模板",
    ]


def test_rename_rejects_default_duplicate_and_missing_template(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")
    tmpl.import_template_xlsx(exported_path, name="模板A", store_path=tmp_path / "templates.json")
    tmpl.import_template_xlsx(exported_path, name="模板B", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="不允许重命名"):
        tmpl.rename_template("默认模板", new_name="新默认", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="系统模板不允许重命名"):
        tmpl.rename_template("DE模板-一组", new_name="新DE", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="模板名已存在"):
        tmpl.rename_template("模板A", new_name="模板B", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="新模板名不能是系统模板"):
        tmpl.rename_template("模板A", new_name="US模板-一组", store_path=tmp_path / "templates.json")

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="只能重命名已存在"):
        tmpl.rename_template("不存在", new_name="模板C", store_path=tmp_path / "templates.json")


def test_validate_template_detects_invalid_threshold() -> None:
    template = tmpl.load_default_template()
    broken = tmpl.ReplenishmentTemplate(
        name="坏模板",
        version=1,
        description="",
        params={
            **template.params,
            "shipping": {
                "air_urgent_sales_days_lte": 80,
                "air_sales_days_lte": 70,
            },
        },
    )

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="空运急发阈值"):
        tmpl.validate_template(broken)


def test_validate_template_requires_sea_enabled() -> None:
    template = tmpl.load_default_template()
    params = json.loads(json.dumps(template.params, ensure_ascii=False))
    params["sea"].pop("enabled")
    broken = tmpl.ReplenishmentTemplate(
        name="旧模板",
        version=1,
        description="",
        params=params,
    )

    with pytest.raises(tmpl.ReplenishmentTemplateError, match="sea.enabled必须是布尔值"):
        tmpl.validate_template(broken)


def test_template_xlsx_parses_disabled_sea_switch(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("默认模板", output_dir=tmp_path / "editable")

    from openpyxl import load_workbook

    workbook = load_workbook(exported_path)
    try:
        workbook["海运规则"].cell(row=2, column=5).value = "否"
        workbook.save(exported_path)
    finally:
        workbook.close()

    result = tmpl.validate_template_xlsx(exported_path)

    assert result.template.params["sea"]["enabled"] is False


def test_template_xlsx_round_trips_inclusive_sea_min_daily_sales(tmp_path) -> None:
    exported_path = tmpl.export_template_xlsx("2组-US站点-林美淇", output_dir=tmp_path / "editable")

    assert _cell_value(exported_path, "海运规则", 4, 5) == "是"

    result = tmpl.validate_template_xlsx(exported_path)

    assert result.template.name == "2组-US站点-林美淇"
    assert result.template.params["sea"]["min_daily_sales_inclusive"] is True
    assert _cell_value(exported_path, "海运规则", 6, 5) == "是"
    assert _cell_value(exported_path, "海运规则", 7, 5) == 30
    assert _cell_value(exported_path, "海运规则", 11, 4) == "70"
    assert result.template.params["sea"]["companion_air_enabled"] is True
    assert result.template.params["sea"]["min_net_quantity"] == 30
    assert tmpl.sea_day_candidates_from_template(1, result.template.params) == [100]
    assert tmpl.sea_companion_air_day_candidates_from_template(1, result.template.params) == [70]


def test_special_rule_applies_msku_overrides() -> None:
    template = tmpl.load_default_template()
    custom = tmpl.ReplenishmentTemplate(
        name="特殊模板",
        version=1,
        description="",
        params={
            **template.params,
            "special_rules": [
                {
                    "rule_name": "大件规则",
                    "msku_list": ["MSKU-A"],
                    "overrides": {
                        "sea": {"min_weight_kg": 100},
                        "shipping": {"air_urgent_sales_days_lte": 35},
                    },
                }
            ],
        },
    )

    params, rule_name = tmpl.effective_params_for_msku(custom, "MSKU-A")
    assert rule_name == "大件规则"
    assert params["sea"]["min_weight_kg"] == 100
    assert params["shipping"]["air_urgent_sales_days_lte"] == 35

    params, rule_name = tmpl.effective_params_for_msku(custom, "MSKU-B")
    assert rule_name == ""
    assert params["sea"]["min_weight_kg"] == 60


def test_cli_list_and_list_params(capsys) -> None:
    assert cli.main(["list"]) == 0
    payload = _read_payload(capsys)
    assert payload["templates"][0]["name"] == "默认模板"
    assert [item["name"] for item in payload["templates"][:4]] == [
        "默认模板",
        "US模板-一组",
        "UK模板-一组",
        "DE模板-一组",
    ]

    assert cli.main(["list-params"]) == 0
    payload = _read_payload(capsys)
    assert payload["groups"][0]["group"] == "加权日销"
    assert any(param["key"] == "sea.enabled" for group in payload["groups"] for param in group["params"])
    assert any(param["key"] == "sea.min_daily_sales_inclusive" for group in payload["groups"] for param in group["params"])
    assert any(param["key"] == "sea.companion_air_enabled" for group in payload["groups"] for param in group["params"])
    assert any(param["key"] == "sea.companion_air_tiers" for group in payload["groups"] for param in group["params"])
    assert any(param["key"] == "sea.min_net_quantity" for group in payload["groups"] for param in group["params"])


def test_cli_show_export_validate_and_import(monkeypatch, tmp_path, capsys) -> None:
    monkeypatch.chdir(tmp_path)

    assert cli.main(["show", "--template", "默认模板"]) == 0
    payload = _read_payload(capsys)
    assert payload["template"]["name"] == "默认模板"

    assert cli.main(["export", "--template", "默认模板"]) == 0
    payload = _read_payload(capsys)
    exported_path = Path(payload["xlsx_path"])
    assert exported_path.is_file()

    assert cli.main(["validate-file", "--xlsx", str(exported_path)]) == 0
    payload = _read_payload(capsys)
    assert payload["template_name"] == "默认模板"

    assert cli.main(["import", "--xlsx", str(exported_path), "--name", "CLI模板"]) == 0
    payload = _read_payload(capsys)
    assert payload["template_name"] == "CLI模板"

    assert cli.main(["replace", "--template", "CLI模板", "--xlsx", str(exported_path)]) == 0
    payload = _read_payload(capsys)
    assert payload["template_name"] == "CLI模板"
    assert payload["old_version"] == 1
    assert payload["new_version"] == 2

    assert cli.main(["rename", "--template", "CLI模板", "--name", "CLI重命名模板"]) == 0
    payload = _read_payload(capsys)
    assert payload["old_name"] == "CLI模板"
    assert payload["new_name"] == "CLI重命名模板"
    assert payload["template_version"] == 2


def test_skill_index_loads_replenishment_template_manage() -> None:
    manifest = load_skill_index(force_reload=True).get("replenishment-template-manage")

    assert manifest is not None
    assert manifest.name == "replenishment-template-manage"
    assert manifest.type == "amazon_replenish"
