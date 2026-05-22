from __future__ import annotations

import asyncio
import csv
import json
import sys
import types
from io import BytesIO
from pathlib import Path

import pytest

from services.mabang.amazon.fba import msku_detail as msku
from services.mabang.auth import MabangAuthContext


def _xlsx_bytes(rows: list[dict[str, str]], *, columns: list[str] | None = None) -> bytes:
    from openpyxl import Workbook

    if columns is None:
        columns = list(msku.EXPECTED_DETAIL_HEADERS)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_xlsx(path: Path, rows: list[dict[str, str]], *, columns: list[str] | None = None) -> Path:
    path.write_bytes(_xlsx_bytes(rows, columns=columns))
    return path


def _write_csv(path: Path, rows: list[dict[str, str]], *, columns: list[str]) -> Path:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)
    return path


class _FakeResponse:
    def __init__(
        self,
        payload: dict | None = None,
        *,
        status: int = 200,
        body: bytes = b"",
        text_body: str | None = None,
    ) -> None:
        self.status = status
        self._payload = dict(payload or {})
        self._body = body
        self._text_body = text_body

    async def text(self) -> str:
        if self._text_body is not None:
            return self._text_body
        return json.dumps(self._payload, ensure_ascii=False)

    async def json(self, content_type=None) -> dict:
        if self._text_body is not None:
            raise ValueError("not json")
        return dict(self._payload)

    async def read(self) -> bytes:
        return self._body


class _FakeRequest:
    def __init__(self, response: _FakeResponse) -> None:
        self._response = response

    async def __aenter__(self) -> _FakeResponse:
        return self._response

    async def __aexit__(self, exc_type, exc, tb) -> None:
        return None


class _FakeSession:
    def __init__(self, responses: list[_FakeResponse]) -> None:
        self.responses = list(responses)
        self.calls: list[dict] = []

    def post(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "POST", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))

    def get(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "GET", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))


async def _fake_auth_context(*args, **kwargs) -> MabangAuthContext:
    return MabangAuthContext(
        scope="erp",
        account="",
        source="test",
        cookies_by_domain={
            ".mabangerp.com": [
                {"name": "PHPSESSID", "value": "sid", "domain": ".mabangerp.com"},
                {"name": "signed", "value": "signed-value", "domain": ".mabangerp.com"},
                {"name": "route", "value": "route-value", "domain": ".mabangerp.com"},
                {
                    "name": "MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE",
                    "value": "memcache-key",
                    "domain": ".mabangerp.com",
                },
                {
                    "name": "MABANG_ERP_PRO_MEMBERINFO_LOGIN_PLUS",
                    "value": "login-plus",
                    "domain": ".mabangerp.com",
                },
            ]
        },
        free_token="",
        wms_cookie_header="",
        raw={},
    )


async def _fake_auth_without_memcache(*args, **kwargs) -> MabangAuthContext:
    return MabangAuthContext(
        scope="erp",
        account="",
        source="test",
        cookies_by_domain={".mabangerp.com": [{"name": "PHPSESSID", "value": "sid", "domain": ".mabangerp.com"}]},
        free_token="",
        wms_cookie_header="",
        raw={},
    )


def _form_value(call: dict, name: str) -> str:
    for key, value in call.get("data", []):
        if key == name:
            return value
    raise AssertionError(f"missing form field: {name}")


def _form_values(call: dict, name: str) -> list[str]:
    return [value for key, value in call.get("data", []) if key == name]


def test_load_mskus_from_delivery_csv_dedupes_and_preserves_order(tmp_path):
    delivery_path = _write_csv(
        tmp_path / "SP260414001_1001.csv",
        [
            {"MSKU": "MSKU-A", "店铺": "Shop-A", "SKU": "LOCAL-A"},
            {"MSKU": "MSKU-B", "店铺": "Shop-A", "SKU": "LOCAL-B"},
            {"MSKU": "MSKU-A", "店铺": "Shop-A", "SKU": "LOCAL-C"},
            {"MSKU": "", "店铺": "Shop-A", "SKU": "LOCAL-D"},
        ],
        columns=["MSKU", "店铺", "SKU"],
    )

    assert msku.load_mskus_from_delivery_file(delivery_path) == ["MSKU-A", "MSKU-B"]
    source = msku.load_msku_source_from_delivery_file(delivery_path)
    assert source.msku_shop_pairs == frozenset({("MSKU-A", "Shop-A"), ("MSKU-B", "Shop-A")})


def test_load_mskus_from_delivery_excel_dedupes_and_preserves_order(tmp_path):
    delivery_path = _write_xlsx(
        tmp_path / "SP260414001.xlsx",
        [
            {"MSKU": "MSKU-A", "店铺": "Shop-A", "SKU": "LOCAL-A"},
            {"MSKU": "MSKU-B", "店铺": "Shop-B", "SKU": "LOCAL-B"},
            {"MSKU": "MSKU-A", "店铺": "Shop-A", "SKU": "LOCAL-C"},
        ],
        columns=["MSKU", "店铺", "SKU"],
    )

    assert msku.load_mskus_from_delivery_file(delivery_path) == ["MSKU-A", "MSKU-B"]
    source = msku.load_msku_source_from_delivery_file(delivery_path)
    assert source.msku_shop_pairs == frozenset({("MSKU-A", "Shop-A"), ("MSKU-B", "Shop-B")})


def test_load_mskus_from_delivery_file_requires_msku_column(tmp_path):
    delivery_path = _write_csv(
        tmp_path / "SP260414001_1001.csv",
        [{"店铺": "Shop-A", "SKU": "MSKU-A"}],
        columns=["店铺", "SKU"],
    )

    with pytest.raises(msku.MskuDetailDownloadError, match="缺少列: MSKU"):
        msku.load_mskus_from_delivery_file(delivery_path)


def test_load_mskus_from_delivery_file_requires_shop_column(tmp_path):
    delivery_path = _write_csv(
        tmp_path / "SP260414001_1001.csv",
        [{"MSKU": "MSKU-A", "SKU": "LOCAL-A"}],
        columns=["MSKU", "SKU"],
    )

    with pytest.raises(msku.MskuDetailDownloadError, match="缺少列: 店铺"):
        msku.load_mskus_from_delivery_file(delivery_path)


def test_download_msku_detail_excel_posts_forms_and_validates_headers(monkeypatch, tmp_path):
    delivery_path = _write_csv(
        tmp_path / "SP260414001_1001.csv",
        [
            {"MSKU": "MSKU-A", "店铺": "Shop-A"},
            {"MSKU": "MSKU-B", "店铺": "Shop-A"},
            {"MSKU": "MSKU-A", "店铺": "Shop-A"},
        ],
        columns=["MSKU", "店铺"],
    )
    fake_session = _FakeSession(
        [
            _FakeResponse({"success": True, "id": "1001,1002,1003"}),
            _FakeResponse({"success": True, "gourl": "https://upload.example.test/detail.xlsx"}),
            _FakeResponse(
                body=_xlsx_bytes(
                    [
                        {"MSKU": "MSKU-A", "店铺名称": "Shop-A"},
                        {"MSKU": "MSKU-A", "店铺名称": "Shop-B"},
                    ]
                )
            ),
        ]
    )

    async def fake_resolve_delivery_file(ship_no: str):
        assert ship_no == "SP260414001"
        return delivery_path.resolve(), "local"

    monkeypatch.setattr(msku, "resolve_delivery_file", fake_resolve_delivery_file)
    monkeypatch.setattr(msku, "get_auth_context", _fake_auth_context)
    monkeypatch.setattr(msku, "erp_http_session", fake_session)
    monkeypatch.setattr(msku, "external_http_session", fake_session)

    result = asyncio.run(msku.download_msku_detail_excel("sp260414001", output_dir=tmp_path / "out"))

    post_calls = [call for call in fake_session.calls if call["method"] == "POST"]
    get_calls = [call for call in fake_session.calls if call["method"] == "GET"]
    assert len(post_calls) == 2
    assert len(get_calls) == 1
    assert _form_value(post_calls[0], "platformSkuData") == "MSKU-A\r\nMSKU-B\r\n"
    assert _form_value(post_calls[0], "selecttype") == "platformSkuIn"
    assert _form_value(post_calls[1], "orderIds") == "1001\r\n1002\r\n1003\r\n"
    assert _form_value(post_calls[1], "memcacheKey") == "memcache-key"
    assert _form_values(post_calls[1], "map-name[]") == list(msku.EXPECTED_DETAIL_HEADERS)
    assert result.ship_no == "SP260414001"
    assert result.delivery_file_path == str(delivery_path.resolve())
    assert result.delivery_file_source == "local"
    assert result.msku_count == 2
    assert result.id_count == 3
    assert result.excel_path == str(tmp_path / "out" / "SP260414001_msku_detail.xlsx")
    assert result.xlsx_path == str(tmp_path / "out" / "SP260414001_msku_detail_normalized.xlsx")
    assert result.converted is False
    assert result.raw_excel_deleted is False
    assert result.matched_detail_count == 1
    assert result.shop_mismatch_count == 1
    assert result.shop_mismatch_sheet == "店铺不一致"
    assert Path(result.excel_path).is_file()
    assert Path(result.xlsx_path).is_file()


def test_download_msku_detail_excel_fails_when_delivery_file_missing(monkeypatch):
    async def fake_resolve_delivery_file(ship_no: str):
        raise FileNotFoundError(f"未找到FBA发货单文件: {ship_no}")

    monkeypatch.setattr(msku, "resolve_delivery_file", fake_resolve_delivery_file)

    with pytest.raises(FileNotFoundError, match="未找到FBA发货单文件"):
        asyncio.run(msku.download_msku_detail_excel("SP260414001"))


def test_find_latest_delivery_file_selects_latest_matching_file(tmp_path):
    old_path = _write_csv(tmp_path / "SP260414001_1001.csv", [{"MSKU": "MSKU-A"}], columns=["MSKU"])
    latest_path = _write_csv(tmp_path / "SP260414001_1002.csv", [{"MSKU": "MSKU-B"}], columns=["MSKU"])
    other_path = _write_csv(tmp_path / "SP260414002_1003.csv", [{"MSKU": "MSKU-C"}], columns=["MSKU"])
    old_path.touch()
    other_path.touch()
    latest_path.touch()

    assert msku.find_latest_delivery_file("SP260414001", delivery_file_dir=tmp_path) == latest_path.resolve()


def test_resolve_delivery_file_downloads_when_local_file_missing(monkeypatch, tmp_path):
    delivery_path = _write_csv(tmp_path / "SP260414001_1001.csv", [{"MSKU": "MSKU-A"}], columns=["MSKU"])
    calls: list[str] = []

    async def fake_download_fba_delivery_csv(delivery_no: str):
        calls.append(delivery_no)
        return types.SimpleNamespace(csv_path=str(delivery_path))

    monkeypatch.setattr(msku, "download_fba_delivery_csv", fake_download_fba_delivery_csv)

    resolved_path, source = asyncio.run(
        msku.resolve_delivery_file("SP260414001", delivery_file_dir=tmp_path / "missing")
    )

    assert calls == ["SP260414001"]
    assert resolved_path == delivery_path.resolve()
    assert source == "downloaded"


def test_resolve_auth_requires_memcache_cookie(monkeypatch):
    monkeypatch.setattr(msku, "get_auth_context", _fake_auth_without_memcache)

    with pytest.raises(msku.MskuDetailDownloadAuthError, match="MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE"):
        asyncio.run(msku.resolve_msku_detail_auth())


def test_fetch_ids_rejects_auth_failure(monkeypatch):
    fake_session = _FakeSession([_FakeResponse(status=403, text_body="forbidden")])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadAuthError, match="鉴权失败"):
        asyncio.run(msku.fetch_msku_detail_ids(["MSKU-A"], cookie_header="PHPSESSID=sid"))


def test_fetch_ids_rejects_non_json(monkeypatch):
    fake_session = _FakeSession([_FakeResponse(text_body="not-json")])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadError, match="返回非JSON对象"):
        asyncio.run(msku.fetch_msku_detail_ids(["MSKU-A"], cookie_header="PHPSESSID=sid"))


def test_fetch_ids_rejects_business_failure(monkeypatch):
    fake_session = _FakeSession([_FakeResponse({"success": False, "msg": "bad request"})])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadError, match="业务异常: bad request"):
        asyncio.run(msku.fetch_msku_detail_ids(["MSKU-A"], cookie_header="PHPSESSID=sid"))


def test_export_file_url_rejects_auth_failure(monkeypatch):
    fake_session = _FakeSession([_FakeResponse(status=403, text_body="forbidden")])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadAuthError, match="鉴权失败"):
        asyncio.run(
            msku.export_msku_detail_file_url(
                ["1001"],
                cookie_header="PHPSESSID=sid",
                memcache_key="memcache-key",
            )
        )


def test_export_file_url_rejects_non_json(monkeypatch):
    fake_session = _FakeSession([_FakeResponse(text_body="not-json")])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadError, match="返回非JSON对象"):
        asyncio.run(
            msku.export_msku_detail_file_url(
                ["1001"],
                cookie_header="PHPSESSID=sid",
                memcache_key="memcache-key",
            )
        )


def test_export_file_url_rejects_business_failure(monkeypatch):
    fake_session = _FakeSession([_FakeResponse({"success": False, "msg": "export failed"})])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadError, match="业务异常: export failed"):
        asyncio.run(
            msku.export_msku_detail_file_url(
                ["1001"],
                cookie_header="PHPSESSID=sid",
                memcache_key="memcache-key",
            )
        )


def test_export_file_url_requires_gourl(monkeypatch):
    fake_session = _FakeSession([_FakeResponse({"success": True})])
    monkeypatch.setattr(msku, "erp_http_session", fake_session)

    with pytest.raises(msku.MskuDetailDownloadError, match="缺少 gourl"):
        asyncio.run(
            msku.export_msku_detail_file_url(
                ["1001"],
                cookie_header="PHPSESSID=sid",
                memcache_key="memcache-key",
            )
        )


def test_validate_msku_detail_excel_headers_requires_expected_columns(tmp_path):
    excel_path = _write_xlsx(tmp_path / "detail.xlsx", [{"MSKU": "MSKU-A"}], columns=["MSKU", "售价"])

    with pytest.raises(msku.MskuDetailDownloadError, match="缺少列: 店铺名称"):
        msku.validate_msku_detail_excel_headers(excel_path)


def _worksheet_rows(path: Path, sheet_name: str) -> list[list[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_split_msku_detail_by_delivery_shop_moves_mismatched_rows(tmp_path):
    detail_path = _write_xlsx(
        tmp_path / "detail.xlsx",
        [
            {"MSKU": "MSKU-A", "店铺名称": "Shop-A", "售价": "10"},
            {"MSKU": "MSKU-A", "店铺名称": "Shop-B", "售价": "11"},
            {"MSKU": "MSKU-B", "店铺名称": "Shop-B", "售价": "12"},
        ],
    )

    result = msku.split_msku_detail_by_delivery_shop(
        detail_path,
        frozenset({("MSKU-A", "Shop-A"), ("MSKU-B", "Shop-B")}),
    )

    main_rows = _worksheet_rows(detail_path, "Sheet")
    mismatch_rows = _worksheet_rows(detail_path, "店铺不一致")
    assert result.matched_detail_count == 2
    assert result.shop_mismatch_count == 1
    assert main_rows[0] == list(msku.EXPECTED_DETAIL_HEADERS)
    assert [row[1] for row in main_rows[1:]] == ["MSKU-A", "MSKU-B"]
    assert [row[1] for row in mismatch_rows[1:]] == ["MSKU-A"]
    assert [row[0] for row in mismatch_rows[1:]] == ["Shop-B"]


def test_split_msku_detail_by_delivery_shop_keeps_empty_mismatch_sheet(tmp_path):
    detail_path = _write_xlsx(
        tmp_path / "detail.xlsx",
        [{"MSKU": "MSKU-A", "店铺名称": "Shop-A", "售价": "10"}],
    )

    result = msku.split_msku_detail_by_delivery_shop(detail_path, frozenset({("MSKU-A", "Shop-A")}))

    mismatch_rows = _worksheet_rows(detail_path, "店铺不一致")
    assert result.matched_detail_count == 1
    assert result.shop_mismatch_count == 0
    assert mismatch_rows == [list(msku.EXPECTED_DETAIL_HEADERS)]


def test_split_msku_detail_by_delivery_shop_allows_all_rows_mismatched(tmp_path):
    detail_path = _write_xlsx(
        tmp_path / "detail.xlsx",
        [{"MSKU": "MSKU-A", "店铺名称": "Shop-B", "售价": "10"}],
    )

    result = msku.split_msku_detail_by_delivery_shop(detail_path, frozenset({("MSKU-A", "Shop-A")}))

    main_rows = _worksheet_rows(detail_path, "Sheet")
    mismatch_rows = _worksheet_rows(detail_path, "店铺不一致")
    assert result.matched_detail_count == 0
    assert result.shop_mismatch_count == 1
    assert main_rows == [list(msku.EXPECTED_DETAIL_HEADERS)]
    assert [row[1] for row in mismatch_rows[1:]] == ["MSKU-A"]


def test_split_msku_detail_by_delivery_shop_requires_shop_column(tmp_path):
    detail_path = _write_xlsx(tmp_path / "detail.xlsx", [{"MSKU": "MSKU-A"}], columns=["MSKU", "售价"])

    with pytest.raises(msku.MskuDetailDownloadError, match="缺少列: 店铺名称"):
        msku.split_msku_detail_by_delivery_shop(detail_path, frozenset({("MSKU-A", "Shop-A")}))


def test_convert_msku_detail_xls_to_xlsx_uses_ignore_workbook_corruption(monkeypatch, tmp_path):
    source_path = tmp_path / "detail.xls"
    source_path.write_bytes(b"legacy-xls")
    calls: list[dict] = []

    class FakeSheet:
        name = "Recovered_Sheet1"
        nrows = 2
        ncols = len(msku.EXPECTED_DETAIL_HEADERS)

        def cell_value(self, row_index: int, column_index: int) -> str:
            if row_index == 0:
                return msku.EXPECTED_DETAIL_HEADERS[column_index]
            if msku.EXPECTED_DETAIL_HEADERS[column_index] == "MSKU":
                return "MSKU-A"
            return f"value-{column_index}"

    class FakeBook:
        def sheet_by_index(self, index: int) -> FakeSheet:
            assert index == 0
            return FakeSheet()

    def fake_open_workbook(filename: str, **kwargs):
        calls.append({"filename": filename, **kwargs})
        if not kwargs.get("ignore_workbook_corruption"):
            raise RuntimeError("Workbook corruption: seen[2] == 4")
        return FakeBook()

    monkeypatch.setitem(sys.modules, "xlrd", types.SimpleNamespace(open_workbook=fake_open_workbook))

    xlsx_path = msku.convert_msku_detail_xls_to_xlsx(source_path)

    assert xlsx_path == tmp_path / "detail.xlsx"
    assert calls == [{"filename": str(source_path), "ignore_workbook_corruption": True}]
    msku.validate_msku_detail_excel_headers(xlsx_path)


def test_normalize_msku_detail_excel_copies_xlsx_to_preserve_raw_file(tmp_path):
    source_path = _write_xlsx(tmp_path / "detail.xlsx", [{"MSKU": "MSKU-A"}])

    xlsx_path, converted = msku.normalize_msku_detail_excel(source_path)

    assert xlsx_path == tmp_path / "detail_normalized.xlsx"
    assert converted is False
    assert source_path.is_file()
    assert xlsx_path.is_file()


def test_delete_raw_msku_detail_xls_removes_xls_only(tmp_path):
    xls_path = tmp_path / "detail.xls"
    xlsx_path = tmp_path / "detail.xlsx"
    xls_path.write_bytes(b"legacy")
    xlsx_path.write_bytes(b"modern")

    assert msku.delete_raw_msku_detail_xls(xls_path) is True
    assert not xls_path.exists()
    assert msku.delete_raw_msku_detail_xls(xlsx_path) is False
    assert xlsx_path.exists()
