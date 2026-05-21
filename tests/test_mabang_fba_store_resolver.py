from __future__ import annotations

import asyncio
from pathlib import Path

import pytest

from services.mabang.amazon.fba import store_resolver as stores
from services.mabang.auth import MabangAuthContext


class _FakeResponse:
    def __init__(self, text_body: str, *, status: int = 200) -> None:
        self.status = status
        self._text_body = text_body

    async def text(self) -> str:
        return self._text_body


class _FakeRequest:
    def __init__(self, response: _FakeResponse) -> None:
        self._response = response

    async def __aenter__(self) -> _FakeResponse:
        return self._response

    async def __aexit__(self, exc_type, exc, tb) -> None:
        return None


class _FakeSession:
    def __init__(self, responses: list[_FakeResponse]) -> None:
        self.responses = list(responses)
        self.calls: list[dict] = []

    def get(self, url: str, **kwargs) -> _FakeRequest:
        self.calls.append({"method": "GET", "url": url, **kwargs})
        return _FakeRequest(self.responses.pop(0))


async def _fake_auth_context(*args, **kwargs) -> MabangAuthContext:
    return MabangAuthContext(
        scope="private_amz",
        account="",
        source="test",
        cookies_by_domain={
            ".mabangerp.com": [
                {"name": "PHPSESSID", "value": "sid", "domain": ".mabangerp.com"},
                {"name": "signed", "value": "signed-value", "domain": ".mabangerp.com"},
                {"name": "route", "value": "route-value", "domain": ".mabangerp.com"},
                {
                    "name": "MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE",
                    "value": "memcache-key",
                    "domain": ".mabangerp.com",
                },
                {
                    "name": "MABANG_ERP_PRO_MEMBERINFO_LOGIN_PLUS",
                    "value": "login-plus",
                    "domain": ".mabangerp.com",
                },
            ]
        },
        free_token="",
        wms_cookie_header="",
        raw={},
    )


def _form_value(call: dict, name: str) -> str:
    for key, value in call.get("params", []):
        if key == name:
            return value
    raise AssertionError(f"missing param: {name}")


def test_parse_fba_store_options_skips_incomplete_items_and_dedupes() -> None:
    html = """
    <ul>
      <li><input name="fbaWarehouseIds[]" value="101"><span class="texts"> JP Store </span></li>
      <li><input name="fbaWarehouseIds[]" value="101"><span class="texts">JP Store</span></li>
      <li><input name="fbaWarehouseIds[]" value=""><span class="texts">Empty ID</span></li>
      <li><input name="other" value="102"><span class="texts">Wrong Input</span></li>
      <li><input name="fbaWarehouseIds[]" value="103"></li>
      <li><span class="texts">No Input</span></li>
      <li><input name="fbaWarehouseIds[]" value="104"><span class="texts">UK Store</span></li>
    </ul>
    """

    result = stores.parse_fba_store_options(html)

    assert [store.to_payload() for store in result] == [
        {
            "store_name": "JP Store",
            "store_id": "101",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
        {
            "store_name": "UK Store",
            "store_id": "104",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
    ]


def test_parse_fba_store_options_includes_dropdown_shop_id_children() -> None:
    html = """
    <li class="">
      <label class="checkbox-inline">
        <input name="fbaWarehouseIds[]" value="1024109" type="checkbox">
        <span class="texts">Amazon-Liansheng-BR</span>
      </label>
    </li>
    <li class="dropdown-submenu">
      <label class="checkbox-inline">
        <input name="fbaWarehouseIds[]" value="1014510" type="checkbox">
        <span class="texts">Amazon-区 <i class="ico-arrow-right3"></i></span>
      </label>
      <ul class="dropdown-menu" style="width:220px;">
        <li><a data-type="shopId" data-val="697618612">Amazon-Lerxiuer-SE</a></li>
        <li><a data-type="shopId" data-val="697456820">Amazon-Lerxiuer-DE</a></li>
      </ul>
    </li>
    """

    result = stores.parse_fba_store_options(html)

    assert [store.to_payload() for store in result] == [
        {
            "store_name": "Amazon-Liansheng-BR",
            "store_id": "1024109",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
        {
            "store_name": "Amazon-区",
            "store_id": "1014510",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
        {
            "store_name": "Amazon-Lerxiuer-SE",
            "store_id": "697618612",
            "id_type": "shopId",
            "parent_store_name": "Amazon-区",
        },
        {
            "store_name": "Amazon-Lerxiuer-DE",
            "store_id": "697456820",
            "id_type": "shopId",
            "parent_store_name": "Amazon-区",
        },
    ]


def test_parse_fba_store_options_raises_when_empty() -> None:
    with pytest.raises(stores.FbaStoreResolverParseError, match="未解析到FBA店铺名称和ID"):
        stores.parse_fba_store_options("<ul><li><span>empty</span></li></ul>")


def test_match_fba_store_prefers_exact_match() -> None:
    result = stores.match_fba_store(
        " jp store ",
        [
            stores.FbaStore(store_name="JP Store", store_id="101"),
            stores.FbaStore(store_name="JP Store Outlet", store_id="102"),
        ],
    )

    assert result.match_status == "exact"
    assert result.store.store_name == "JP Store"
    assert result.store.store_id == "101"
    assert result.store.id_type == "fbaWarehouseIds[]"


def test_match_fba_store_accepts_unique_contains_match() -> None:
    result = stores.match_fba_store(
        "Outlet",
        [
            stores.FbaStore(store_name="JP Store", store_id="101"),
            stores.FbaStore(store_name="JP Store Outlet", store_id="102"),
        ],
    )

    assert result.match_status == "contains"
    assert result.store.store_id == "102"
    assert result.store.id_type == "fbaWarehouseIds[]"


def test_match_fba_store_rejects_ambiguous_contains_match() -> None:
    with pytest.raises(stores.FbaStoreAmbiguousError) as exc_info:
        stores.match_fba_store(
            "Store",
            [
                stores.FbaStore(store_name="JP Store", store_id="101"),
                stores.FbaStore(store_name="UK Store", store_id="102"),
            ],
        )

    assert "店铺名不唯一" in str(exc_info.value)
    assert exc_info.value.candidates == [
        {
            "store_name": "JP Store",
            "store_id": "101",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
        {
            "store_name": "UK Store",
            "store_id": "102",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
    ]


def test_match_fba_store_rejects_not_found() -> None:
    with pytest.raises(stores.FbaStoreNotFoundError) as exc_info:
        stores.match_fba_store(
            "US",
            [stores.FbaStore(store_name="JP Store", store_id="101")],
        )

    assert "未找到FBA店铺" in str(exc_info.value)
    assert exc_info.value.candidates == []


def _worksheet_rows(path: Path) -> list[list[object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook["stores"]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_write_fba_stores_xlsx_writes_expected_columns_and_rows(tmp_path) -> None:
    path = stores.write_fba_stores_xlsx(
        [
            stores.FbaStore(store_name="Amazon-区", store_id="1014510"),
            stores.FbaStore(
                store_name="Amazon-Lerxiuer-SE",
                store_id="697618612",
                id_type=stores.ID_TYPE_SHOP,
                parent_store_name="Amazon-区",
                parent_store_id="1014510",
                parent_id_type=stores.ID_TYPE_FBA_WAREHOUSE,
            ),
        ],
        output_dir=tmp_path,
        filename_prefix="fba_stores",
    )

    rows = _worksheet_rows(path)
    assert rows[0] == list(stores.STORE_XLSX_HEADERS)
    assert rows[1] == [
        "Amazon-区",
        "1014510",
        "fbaWarehouseIds[]",
        None,
    ]
    assert rows[2] == [
        "Amazon-Lerxiuer-SE",
        "697618612",
        "shopId",
        "Amazon-区",
    ]


def test_list_fba_stores_writes_xlsx_and_returns_counts(monkeypatch, tmp_path) -> None:
    async def fake_fetch_fba_stores():
        return [
            stores.FbaStore(store_name="Amazon-区", store_id="1014510"),
            stores.FbaStore(
                store_name="Amazon-Lerxiuer-SE",
                store_id="697618612",
                id_type=stores.ID_TYPE_SHOP,
                parent_store_name="Amazon-区",
                parent_store_id="1014510",
                parent_id_type=stores.ID_TYPE_FBA_WAREHOUSE,
            ),
        ]

    monkeypatch.setattr(stores, "fetch_fba_stores", fake_fetch_fba_stores)

    result = asyncio.run(stores.list_fba_stores(output_dir=tmp_path))
    payload = result.to_payload()

    assert payload["success"] is True
    assert payload["store_count"] == 2
    assert payload["fba_warehouse_count"] == 1
    assert payload["shop_count"] == 1
    assert "stores" not in payload
    assert Path(payload["xlsx_path"]).is_file()
    assert len(_worksheet_rows(Path(payload["xlsx_path"]))) == 3


def test_fetch_fba_stores_uses_auth_cookie_and_c_m_key(monkeypatch) -> None:
    html = """
    <ul>
      <li><input name="fbaWarehouseIds[]" value="101"><span class="texts">JP Store</span></li>
    </ul>
    """
    fake_session = _FakeSession([_FakeResponse(html)])
    monkeypatch.setattr(stores, "get_auth_context", _fake_auth_context)
    monkeypatch.setattr(stores, "erp_http_session", fake_session)

    result = asyncio.run(stores.fetch_fba_stores())

    assert [store.to_payload() for store in result] == [
        {
            "store_name": "JP Store",
            "store_id": "101",
            "id_type": "fbaWarehouseIds[]",
            "parent_store_name": "",
        },
    ]
    assert len(fake_session.calls) == 1
    call = fake_session.calls[0]
    assert call["url"] == stores.DEFAULT_STORE_LIST_URL
    assert _form_value(call, "mod") == "fbanew.list"
    assert _form_value(call, "cMKey") == "memcache-key"
    assert "MABANG_ERP_PRO_MEMBERINFO_LOGIN_COOKIE=memcache-key" in call["headers"]["Cookie"]
    assert "mabang_lite_rowsPerPage=100" in call["headers"]["Cookie"]


def test_resolve_fba_store_returns_payload(monkeypatch) -> None:
    async def fake_fetch_fba_stores():
        return [stores.FbaStore(store_name="JP Store", store_id="101")]

    monkeypatch.setattr(stores, "fetch_fba_stores", fake_fetch_fba_stores)

    result = asyncio.run(stores.resolve_fba_store("JP"))

    assert result.to_payload() == {
        "success": True,
        "query": "JP",
        "match_status": "contains",
        "store_name": "JP Store",
        "store_id": "101",
        "id_type": "fbaWarehouseIds[]",
        "parent_store_name": "",
        "source": "mabang_fba_store_resolver",
    }
