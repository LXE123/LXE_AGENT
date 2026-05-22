from __future__ import annotations

import asyncio
import json
from collections import OrderedDict
from pathlib import Path
from types import SimpleNamespace

import pytest

from services.agent_cli.mabang import import_export_tax_products as cli


def _write_products(
    path: Path,
    rows: list[dict[str, str]],
    *,
    columns: list[str] | None = None,
    sheet_name: str = "Sheet1",
) -> None:
    from openpyxl import Workbook

    if columns is None:
        columns = ["sku", "产品名称"]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    workbook.save(path)


def _read_rows(path: Path) -> list[tuple]:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    worksheet = workbook["Sheet1"]
    return list(worksheet.iter_rows(values_only=True))


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


async def _noop_close_all_network_clients() -> None:
    return None


def test_missing_sku_returns_failure_json(monkeypatch, capsys):
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main([])

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload == {"success": False, "exception": "sku 不能为空"}


def test_normalize_input_skus_dedupes_and_ignores_empty_values():
    assert cli.normalize_input_skus([" SKU-A ", "", "SKU-A", "SKU-B\nSKU-C，SKU-D"]) == [
        "SKU-A",
        "SKU-B",
        "SKU-C",
        "SKU-D",
    ]


def test_import_skips_existing_and_not_found_then_appends_found(monkeypatch, tmp_path):
    products_path = tmp_path / "products.xlsx"
    backup_dir = tmp_path / "backup"
    _write_products(
        products_path,
        [
            {"sku": "SKU-EXIST", "产品名称": "已有产品"},
        ],
    )

    async def fake_export_stock_sku_names(skus, **kwargs):
        assert skus == ["SKU-NEW", "SKU-MISSING"]
        return SimpleNamespace(
            names_by_key=OrderedDict([("SKU-NEW", "新产品")]),
            xlsx_paths=[],
        )

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)
    monkeypatch.setattr(cli, "_timestamp", lambda: "20260512_120000")

    payload = asyncio.run(
        cli.import_export_tax_products(
            ["SKU-EXIST", "SKU-NEW", "SKU-MISSING"],
            products_path=products_path,
            backup_dir=backup_dir,
        )
    )

    assert payload["success"] is True
    assert payload["requested_sku_count"] == 3
    assert payload["imported_count"] == 1
    assert payload["skipped_duplicate_count"] == 1
    assert payload["skipped_not_found_count"] == 1
    assert payload["imported_skus"] == ["SKU-NEW"]
    assert payload["skipped_duplicate_skus"] == ["SKU-EXIST"]
    assert payload["skipped_not_found_skus"] == ["SKU-MISSING"]
    assert payload["backup_path"] == str(backup_dir / "products_20260512_120000.xlsx")
    assert Path(payload["backup_path"]).is_file()
    assert _read_rows(products_path) == [
        ("sku", "产品名称"),
        ("SKU-EXIST", "已有产品"),
        ("SKU-NEW", "新产品"),
    ]
    assert _read_rows(Path(payload["backup_path"])) == [
        ("sku", "产品名称"),
        ("SKU-EXIST", "已有产品"),
    ]


def test_import_does_not_write_or_call_api_when_all_duplicate(monkeypatch, tmp_path):
    products_path = tmp_path / "products.xlsx"
    backup_dir = tmp_path / "backup"
    _write_products(products_path, [{"sku": "SKU-EXIST", "产品名称": "已有产品"}])

    async def fail_export_stock_sku_names(*args, **kwargs):
        raise AssertionError("stock export should not be called")

    monkeypatch.setattr(cli, "export_stock_sku_names", fail_export_stock_sku_names)

    payload = asyncio.run(
        cli.import_export_tax_products(
            ["SKU-EXIST"],
            products_path=products_path,
            backup_dir=backup_dir,
        )
    )

    assert payload["imported_count"] == 0
    assert payload["skipped_duplicate_skus"] == ["SKU-EXIST"]
    assert payload["backup_path"] == ""
    assert not backup_dir.exists()
    assert _read_rows(products_path) == [
        ("sku", "产品名称"),
        ("SKU-EXIST", "已有产品"),
    ]


def test_import_does_not_write_when_api_finds_no_skus(monkeypatch, tmp_path):
    products_path = tmp_path / "products.xlsx"
    backup_dir = tmp_path / "backup"
    _write_products(products_path, [{"sku": "SKU-EXIST", "产品名称": "已有产品"}])

    async def fake_export_stock_sku_names(skus, **kwargs):
        assert skus == ["SKU-MISSING"]
        return SimpleNamespace(names_by_key=OrderedDict(), xlsx_paths=[])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.import_export_tax_products(
            ["SKU-MISSING"],
            products_path=products_path,
            backup_dir=backup_dir,
        )
    )

    assert payload["imported_count"] == 0
    assert payload["skipped_not_found_skus"] == ["SKU-MISSING"]
    assert payload["backup_path"] == ""
    assert not backup_dir.exists()


def test_main_success_outputs_json(monkeypatch, tmp_path, capsys):
    products_path = tmp_path / "products.xlsx"
    backup_dir = tmp_path / "backup"
    _write_products(products_path, [{"sku": "SKU-EXIST", "产品名称": "已有产品"}])

    async def fake_export_stock_sku_names(skus, **kwargs):
        return SimpleNamespace(names_by_key=OrderedDict([("SKU-NEW", "新产品")]), xlsx_paths=[])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)
    monkeypatch.setattr(cli, "_timestamp", lambda: "20260512_120000")

    exit_code = cli.main(
        [
            "--sku",
            "SKU-NEW",
            "--products-path",
            str(products_path),
            "--backup-dir",
            str(backup_dir),
        ]
    )

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["products_path"] == str(products_path)
    assert payload["backup_path"] == str(backup_dir / "products_20260512_120000.xlsx")
    assert payload["source"] == "export_tax_products_import"


def test_main_validation_failure_returns_failure_json(monkeypatch, tmp_path, capsys):
    products_path = tmp_path / "products.xlsx"
    _write_products(products_path, [{"sku": "SKU-A"}], columns=["sku"])
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main(["--sku", "SKU-NEW", "--products-path", str(products_path)])

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload["success"] is False
    assert "缺少列: 产品名称" in payload["exception"]
