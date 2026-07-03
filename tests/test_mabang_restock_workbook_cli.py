from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pytest

from services.agent_cli.mabang import generate_restock_workbook as cli
from services.agent_cli.mabang import generate_fba_restock_workbook as restock_cli
from services.agent_cli.mabang import generate_purchase_summary_workbook as purchase_cli
from services.agent_cli.mabang import generate_purchase_batch_workbooks as batch_cli

PURCHASE_COLUMNS = (
    "库存sku",
    "产品名称",
    "来源SP单号",
    "库存sku（第一行）",
    "产品名称（第一行）",
    "型号",
    "原价",
    "均价",
    "厂家",
    "单位",
    "合同产品名称",
    "合同编号前缀",
    "税率",
    "数量",
    "总价",
    "总价（均价）",
)
PURCHASE_UNMATCHED_COLUMNS = ("库存sku", "来源SP单号", "数量", "问题说明")
RESTOCK_COLUMNS = (
    "库存sku",
    "产品名称",
    "库存sku（第一行）",
    "产品名称（第一行）",
    "型号",
    "原价",
    "均价",
    "售价",
    "售价(均价)",
    "毛利率",
    "厂家",
    "单位",
    "合同产品名称",
    "数量",
    "总价",
    "总价（均价）",
    "总价（售价）",
    "总价（售价(均价)）",
)
RESTOCK_UNMATCHED_COLUMNS = ("库存sku", "数量", "问题说明")
MISSING_CONTRACT_SHEET_WARNING = "出口退税总表缺少 sheet: 供应商合同信息，单位和合同产品名称将留空"


def _first_line(value: object) -> object:
    if not isinstance(value, str):
        return value
    return value.split("\n", 1)[0]


def _purchase_row(
    stock_skus: object,
    product_names: object,
    source_delivery_nos: object,
    model: object,
    original_price: object,
    manufacturer: object,
    quantity: object,
    total_price: object,
    unit: object = None,
    contract_product_name: object = None,
    contract_no_prefix: object = None,
    tax_rate: object = None,
    average_price: object = None,
    average_total_price: object = None,
) -> tuple[object, ...]:
    return (
        stock_skus,
        product_names,
        source_delivery_nos,
        _first_line(stock_skus),
        _first_line(product_names),
        model,
        original_price,
        average_price,
        manufacturer,
        unit,
        contract_product_name,
        contract_no_prefix,
        tax_rate,
        quantity,
        total_price,
        average_total_price,
    )


def _purchase_total_row(
    quantity: object,
    total_price: object,
    average_total_price: object = None,
) -> tuple[object, ...]:
    return (
        "合计",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        quantity,
        total_price,
        average_total_price,
    )


def _restock_total_row(
    quantity: object,
    total_price: object,
    sale_total_price: object,
    average_total_price: object = None,
    average_sale_total_price: object = None,
) -> tuple[object, ...]:
    return (
        "合计",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        quantity,
        total_price,
        average_total_price,
        sale_total_price,
        average_sale_total_price,
    )


def _write_delivery_csv(
    path: Path,
    rows: list[str],
    *,
    countries: list[str] | None = None,
    include_country: bool = False,
) -> None:
    headers = ["发货单号", "SKU发货量"]
    if include_country or countries is not None:
        headers.append("国家")
    headers.append("备注")
    lines = [",".join(f'"{header}"' for header in headers)]
    country_values = list(countries or [])
    for index, value in enumerate(rows):
        fields = ["SP260508022", value]
        if include_country or countries is not None:
            fields.append(country_values[index] if index < len(country_values) else "")
        fields.append("")
        lines.append(",".join(f'"{field}"' for field in fields))
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def _write_master_xlsx(
    path: Path,
    rows: list[dict[str, object]],
    *,
    columns: list[str] | None = None,
    contract_rows: list[dict[str, object]] | None = None,
    contract_columns: list[str] | None = None,
) -> None:
    from openpyxl import Workbook

    if columns is None:
        columns = ["库存sku", "产品名称", "型号", "原价", "厂家", "备用厂家"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SKU表"
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    if contract_rows is not None:
        if contract_columns is None:
            contract_columns = ["供货方", "单位", "合同产品名称", "合同编号前缀", "税率"]
        contract_sheet = workbook.create_sheet("供应商合同信息")
        contract_sheet.append(contract_columns)
        for row in contract_rows:
            contract_sheet.append([row.get(column, "") for column in contract_columns])
    workbook.save(path)


def _sheet_values(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        return list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()


def _sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _cell_wrap_text(path: Path, sheet_name: str, cell: str) -> bool | None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet_name][cell].alignment.wrap_text
    finally:
        workbook.close()


def _cell_number_format(path: Path, sheet_name: str, cell: str) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        return str(workbook[sheet_name][cell].number_format)
    finally:
        workbook.close()


def _cell_fill_rgb(path: Path, sheet_name: str, cell: str) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        return str(workbook[sheet_name][cell].fill.fgColor.rgb)
    finally:
        workbook.close()


def _sheet_dimensions(path: Path, sheet_name: str) -> tuple[list[float | None], list[float | None]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        widths = [
            worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width
            for column_index in range(1, worksheet.max_column + 1)
        ]
        heights = [
            worksheet.row_dimensions[row_index].height
            for row_index in range(1, worksheet.max_row + 1)
        ]
        return widths, heights
    finally:
        workbook.close()


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


async def _noop_close_all_network_clients() -> None:
    return None


def test_generate_restock_workbook_groups_by_manufacturer(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    csv_path = csv_dir / "SP260508022_1.csv"
    _write_delivery_csv(csv_path, ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 1.5, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "M-B", "原价": 2, "厂家": "厂家B"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert payload["success"] is True
    assert payload["delivery_nos"] == ["SP260508022"]
    assert payload["csv_paths"] == [str(csv_path)]
    assert payload["sku_count"] == 2
    assert payload["sku_source_count"] == 2
    assert payload["matched_sku_count"] == 2
    assert payload["unmatched_sku_count"] == 0
    assert payload["manufacturer_count"] == 2
    assert _sheet_names(output_path) == ["采购汇总", "未匹配", "厂家A", "厂家B"]
    assert _sheet_values(output_path, "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 1.5, "厂家A", 2, 3),
        _purchase_row("SKU-B", "产品B", "SP260508022", "M-B", 2, "厂家B", 3, 6),
        _purchase_total_row(5, 9),
    ]
    assert _sheet_values(output_path, "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 1.5, "厂家A", 2, 3),
        _purchase_total_row(2, 3),
    ]
    assert _sheet_values(output_path, "厂家B") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-B", "产品B", "SP260508022", "M-B", 2, "厂家B", 3, 6),
        _purchase_total_row(3, 6),
    ]
    assert _cell_fill_rgb(output_path, "采购汇总", "A4") == cli.TOTAL_ROW_FILL_COLOR
    assert _cell_fill_rgb(output_path, "厂家A", "A3") == cli.TOTAL_ROW_FILL_COLOR
    assert _cell_fill_rgb(output_path, "厂家B", "A3") == cli.TOTAL_ROW_FILL_COLOR
    assert _sheet_values(output_path, "未匹配") == [
        PURCHASE_UNMATCHED_COLUMNS,
    ]


def test_generate_restock_workbook_fills_contract_fields_from_second_sheet(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-A", "税率": "13%"},
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-A", "税率": "13%"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert payload["warnings"] == []
    assert payload["contract_mapping_count"] == 1
    assert payload["contract_unmapped_manufacturer_count"] == 0
    assert payload["contract_conflict_manufacturer_count"] == 0
    assert payload["contract_prefix_conflict_manufacturer_count"] == 0
    assert payload["contract_tax_rate_conflict_manufacturer_count"] == 0
    assert _sheet_values(output_path, "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row(
            "SKU-A",
            "产品A",
            "SP260508022",
            "M-A",
            2,
            "厂家A",
            2,
            4,
            "个",
            "合同产品A",
            "HT-A",
            "13%",
        ),
        _purchase_total_row(2, 4),
    ]
    assert _sheet_values(output_path, "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row(
            "SKU-A",
            "产品A",
            "SP260508022",
            "M-A",
            2,
            "厂家A",
            2,
            4,
            "个",
            "合同产品A",
            "HT-A",
            "13%",
        ),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_contract_prefix_missing_header(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
        contract_columns=["供货方", "单位", "合同产品名称", "税率"],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 1
    assert payload["contract_prefix_conflict_manufacturer_count"] == 0
    assert payload["warnings"] == [
        "出口退税总表 export_tax.xlsx 供应商合同信息 sheet 缺少列: 合同编号前缀，合同编号前缀将留空"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4, "个", "合同产品A", None, "13%"),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_contract_prefix_conflict(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-A"},
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-B"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 1
    assert payload["contract_conflict_manufacturer_count"] == 0
    assert payload["contract_prefix_conflict_manufacturer_count"] == 1
    assert payload["contract_prefix_conflict_manufacturer_examples"] == ["厂家A"]
    assert payload["warnings"] == [
        "出口退税总表 供应商合同信息 sheet 存在同一供货方对应不同合同编号前缀，"
        "合同编号前缀已留空: count=1, examples=厂家A"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4, "个", "合同产品A"),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_contract_tax_rate_missing_header(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-A"}],
        contract_columns=["供货方", "单位", "合同产品名称", "合同编号前缀"],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 1
    assert payload["contract_tax_rate_conflict_manufacturer_count"] == 0
    assert payload["warnings"] == [
        "出口退税总表 export_tax.xlsx 供应商合同信息 sheet 缺少列: 税率，税率将留空"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4, "个", "合同产品A", "HT-A"),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_contract_tax_rate_conflict(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-A", "税率": "13%"},
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "合同编号前缀": "HT-A", "税率": "9%"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 1
    assert payload["contract_conflict_manufacturer_count"] == 0
    assert payload["contract_tax_rate_conflict_manufacturer_count"] == 1
    assert payload["contract_tax_rate_conflict_manufacturer_examples"] == ["厂家A"]
    assert payload["warnings"] == [
        "出口退税总表 供应商合同信息 sheet 存在同一供货方对应不同税率，"
        "税率已留空: count=1, examples=厂家A"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4, "个", "合同产品A", "HT-A"),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_contract_mapping_conflict(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A"},
            {"供货方": "厂家A", "单位": "套", "合同产品名称": "合同产品A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 0
    assert payload["contract_conflict_manufacturer_count"] == 1
    assert payload["contract_conflict_manufacturer_examples"] == ["厂家A"]
    assert payload["warnings"] == [
        "出口退税总表 供应商合同信息 sheet 存在同一供货方对应不同单位或合同产品名称，"
        "相关厂家字段已留空: count=1, examples=厂家A"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_contract_mapping_missing_required_header(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个"}],
        contract_columns=["供货方", "单位"],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 0
    assert payload["warnings"] == [
        "出口退税总表 export_tax.xlsx 供应商合同信息 sheet 缺少必需列: 合同产品名称，单位和合同产品名称将留空"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_warns_unmapped_contract_manufacturer(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家B", "单位": "个", "合同产品名称": "合同产品B"}],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["contract_mapping_count"] == 1
    assert payload["contract_unmapped_manufacturer_count"] == 1
    assert payload["contract_unmapped_manufacturer_examples"] == ["厂家A"]
    assert payload["warnings"] == [
        "出口退税总表 供应商合同信息 sheet 未找到部分厂家对应的供货方映射，"
        "单位和合同产品名称已留空: count=1, examples=厂家A"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_sums_multiple_delivery_nos(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    _write_delivery_csv(csv_dir / "SP260508023_1.csv", ["SKU-A × 3，SKU-B × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 4, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "M-B", "原价": 2, "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022", "SP260508023"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["sku_source_count"] == 2
    assert Path(payload["output_xlsx"]).name == "SP260508022_SP260508023_purchase_summary.xlsx"
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022\nSP260508023", "M-A", 4, "厂家A", 5, 20),
        _purchase_row("SKU-B", "产品B", "SP260508023", "M-B", 2, "厂家A", 1, 2),
        _purchase_total_row(6, 22),
    ]


def test_generate_restock_workbook_merges_rows_by_model_with_multiline_skus(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert payload["matched_sku_count"] == 2
    assert payload["manufacturer_count"] == 1
    assert _sheet_values(output_path, "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A\nSKU-B", "产品A\n产品A", "SP260508022", "JZ-19", 2, "厂家A", 5, 10),
        _purchase_total_row(5, 10),
    ]
    assert _cell_wrap_text(output_path, "厂家A", "A2") is True
    assert _cell_wrap_text(output_path, "厂家A", "B2") is True
    assert _cell_wrap_text(output_path, "厂家A", "C2") is True
    widths, heights = _sheet_dimensions(output_path, "厂家A")
    assert widths == [15] * 16
    assert heights == [15] * 3
    widths, heights = _sheet_dimensions(output_path, "采购汇总")
    assert widths == [15] * 16
    assert heights == [15] * 3


def test_generate_restock_workbook_ignores_same_model_product_name_conflict(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A\nSKU-B", "产品A\n产品B", "SP260508022", "JZ-19", 2, "厂家A", 5, 10),
        _purchase_total_row(5, 10),
    ]


def test_generate_restock_workbook_does_not_record_zero_quantity_source(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 0"])
    _write_delivery_csv(csv_dir / "SP260508023_1.csv", ["SKU-A × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022", "SP260508023"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["sku_source_count"] == 1
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508023", "M-A", 2, "厂家A", 3, 6),
        _purchase_total_row(3, 6),
    ]


def test_generate_restock_workbook_rejects_same_model_with_different_price(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品A", "型号": "JZ-19", "原价": 3, "厂家": "厂家A"},
        ],
    )

    with pytest.raises(RuntimeError, match="同一厂家同一型号的原价不一致"):
        cli.generate_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_restock_workbook_writes_zhengfei_average_price(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1，SKU-B × 2，SKU-C × 4"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "深圳正飞科技"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-20", "原价": 3, "厂家": "深圳正飞科技"},
            {"库存sku": "SKU-C", "产品名称": "产品C", "型号": "M-C", "原价": 3, "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert _sheet_values(output_path, "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row(
            "SKU-A",
            "产品A",
            "SP260508022",
            "JZ-19",
            2,
            "深圳正飞科技",
            1,
            2,
            average_price=2.67,
            average_total_price=2.67,
        ),
        _purchase_row(
            "SKU-B",
            "产品B",
            "SP260508022",
            "JZ-20",
            3,
            "深圳正飞科技",
            2,
            6,
            average_price=2.67,
            average_total_price=5.34,
        ),
        _purchase_row("SKU-C", "产品C", "SP260508022", "M-C", 3, "厂家A", 4, 12),
        _purchase_total_row(7, 20, 8.01),
    ]
    assert _sheet_values(output_path, "深圳正飞科技") == [
        PURCHASE_COLUMNS,
        _purchase_row(
            "SKU-A",
            "产品A",
            "SP260508022",
            "JZ-19",
            2,
            "深圳正飞科技",
            1,
            2,
            average_price=2.67,
            average_total_price=2.67,
        ),
        _purchase_row(
            "SKU-B",
            "产品B",
            "SP260508022",
            "JZ-20",
            3,
            "深圳正飞科技",
            2,
            6,
            average_price=2.67,
            average_total_price=5.34,
        ),
        _purchase_total_row(3, 8, 8.01),
    ]
    assert _cell_number_format(output_path, "采购汇总", "P2") == "0.00"
    assert _cell_number_format(output_path, "深圳正飞科技", "P2") == "0.00"


def test_generate_restock_workbook_rejects_zhengfei_same_model_with_different_price(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "深圳正飞科技"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-19", "原价": 3, "厂家": "深圳正飞科技"},
        ],
    )

    with pytest.raises(RuntimeError, match="同一厂家同一型号的原价不一致"):
        cli.generate_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_restock_workbook_does_not_merge_same_model_across_manufacturers(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家B"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert payload["manufacturer_count"] == 2
    assert _sheet_values(output_path, "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "JZ-19", 2, "厂家A", 2, 4),
        _purchase_total_row(2, 4),
    ]
    assert _sheet_values(output_path, "厂家B") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-B", "产品A", "SP260508022", "JZ-19", 2, "厂家B", 3, 6),
        _purchase_total_row(3, 6),
    ]


def test_generate_restock_workbook_keeps_empty_model_rows_unmerged_with_warning(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品A", "型号": "", "原价": 2, "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["unmerged_empty_model_sku_count"] == 2
    assert payload["unmerged_empty_model_skus"] == ["SKU-A", "SKU-B"]
    assert payload["warnings"] == [
        MISSING_CONTRACT_SHEET_WARNING,
        "出口退税总表存在型号为空的库存sku，已按 SKU 粒度保留不合并: count=2, examples=SKU-A, SKU-B"
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", None, 2, "厂家A", 2, 4),
        _purchase_row("SKU-B", "产品A", "SP260508022", None, 2, "厂家A", 3, 6),
        _purchase_total_row(5, 10),
    ]


def test_generate_restock_workbook_writes_unmatched_sheet(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1，SKU-X × 4"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": ""}],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert payload["matched_sku_count"] == 1
    assert payload["unmatched_sku_count"] == 1
    assert payload["manufacturer_count"] == 1
    assert _sheet_names(output_path) == ["采购汇总", "未匹配", "未填写厂家"]
    assert _sheet_values(output_path, "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "未填写厂家", 1, 2),
        _purchase_total_row(1, 2),
    ]
    assert _sheet_values(output_path, "未填写厂家") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "未填写厂家", 1, 2),
        _purchase_total_row(1, 2),
    ]
    assert _sheet_values(output_path, "未匹配") == [
        PURCHASE_UNMATCHED_COLUMNS,
        ("SKU-X", "SP260508022", 4, "出口退税总表未找到库存sku"),
    ]
    widths, heights = _sheet_dimensions(output_path, "未匹配")
    assert widths == [15] * 4
    assert heights == [15] * 2


def test_generate_restock_workbook_unmatched_sources_use_line_breaks(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-X × 1"])
    _write_delivery_csv(csv_dir / "SP260508023_1.csv", ["SKU-X × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022", "SP260508023"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert _sheet_values(Path(payload["output_xlsx"]), "未匹配") == [
        PURCHASE_UNMATCHED_COLUMNS,
        ("SKU-X", "SP260508022\nSP260508023", 3, "出口退税总表未找到库存sku"),
    ]


def test_generate_restock_workbook_purchase_summary_preserves_delivery_order(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-B × 3，SKU-A × 2"])
    _write_delivery_csv(csv_dir / "SP260508023_1.csv", ["SKU-C × 4，SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 1, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "M-B", "原价": 1, "厂家": "厂家A"},
            {"库存sku": "SKU-C", "产品名称": "产品C", "型号": "M-C", "原价": 1, "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022", "SP260508023"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert _sheet_values(Path(payload["output_xlsx"]), "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-B", "产品B", "SP260508022", "M-B", 1, "厂家A", 3, 3),
        _purchase_row("SKU-A", "产品A", "SP260508022\nSP260508023", "M-A", 1, "厂家A", 3, 3),
        _purchase_row("SKU-C", "产品C", "SP260508023", "M-C", 1, "厂家A", 4, 4),
        _purchase_total_row(10, 10),
    ]


def test_generate_restock_workbook_total_price_number_format(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": "1.5", "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "M-B", "原价": "0.125", "厂家": "厂家A"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert _cell_number_format(output_path, "采购汇总", "O2") == "0.00"
    assert _cell_number_format(output_path, "采购汇总", "O3") == "0.000"
    assert _cell_number_format(output_path, "采购汇总", "P2") == "General"
    assert _cell_number_format(output_path, "厂家A", "O2") == "0.00"
    assert _cell_number_format(output_path, "厂家A", "O3") == "0.000"
    assert _cell_number_format(output_path, "厂家A", "P2") == "General"


def test_generate_restock_workbook_total_price_format_ignores_float_noise(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 195，SKU-B × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {
                "库存sku": "SKU-A",
                "产品名称": "产品A",
                "型号": "M-A",
                "原价": "0.8500000000000001",
                "厂家": "厂家A",
            },
            {
                "库存sku": "SKU-B",
                "产品名称": "产品B",
                "型号": "M-B",
                "原价": "0.125",
                "厂家": "厂家A",
            },
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    output_path = Path(payload["output_xlsx"])
    assert _cell_number_format(output_path, "采购汇总", "O2") == "0.00"
    assert _cell_number_format(output_path, "采购汇总", "O3") == "0.000"


def test_generate_restock_workbook_dedupes_identical_master_stock_sku(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {
                "库存sku": "SKU-A",
                "产品名称": "产品A",
                "型号": "M-A",
                "原价": 2,
                "厂家": "厂家A",
                "备用厂家": "备用A",
            },
            {
                "库存sku": "SKU-A",
                "产品名称": "产品A",
                "型号": "M-A",
                "原价": 2.0,
                "厂家": "厂家A",
                "备用厂家": "备用A",
            },
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["success"] is True
    assert payload["matched_sku_count"] == 1
    assert payload["manufacturer_count"] == 1
    assert payload["deduped_duplicate_sku_count"] == 1
    assert payload["deduped_duplicate_row_count"] == 1
    assert payload["deduped_duplicate_sku_examples"] == ["SKU-A"]
    assert payload["warnings"] == [
        "出口退税总表存在完全相同的重复库存sku，已自动去重: "
        "sku_count=1, row_count=1, examples=SKU-A",
        MISSING_CONTRACT_SHEET_WARNING,
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_skips_master_rows_with_empty_stock_sku(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "", "产品名称": "无SKU产品", "型号": "M-X", "原价": "bad", "厂家": "厂家X"},
        ],
    )

    payload = cli.generate_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["success"] is True
    assert payload["matched_sku_count"] == 1
    assert payload["skipped_empty_sku_row_count"] == 1
    assert payload["skipped_empty_sku_rows"] == [3]
    assert payload["warnings"] == [
        "出口退税总表存在库存sku为空的行，已忽略: count=1, rows=3",
        MISSING_CONTRACT_SHEET_WARNING,
    ]
    assert _sheet_values(Path(payload["output_xlsx"]), "厂家A") == [
        PURCHASE_COLUMNS,
        _purchase_row("SKU-A", "产品A", "SP260508022", "M-A", 2, "厂家A", 2, 4),
        _purchase_total_row(2, 4),
    ]


def test_generate_restock_workbook_missing_local_delivery_csv_fails(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )

    with pytest.raises(RuntimeError, match="本地未找到发货单 CSV"):
        cli.generate_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            csv_dir=tmp_path / "missing",
            output_dir=tmp_path,
        )


def test_generate_restock_workbook_missing_master_fails(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])

    with pytest.raises(RuntimeError, match="找不到出口退税总表"):
        cli.generate_restock_workbook(
            ["SP260508022"],
            master_xlsx=tmp_path / "missing.xlsx",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_load_master_products_requires_sku_sheet_name(tmp_path):
    from openpyxl import Workbook

    master_path = tmp_path / "export_tax.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "出口退税总表"
    worksheet.append(["库存sku", "产品名称", "型号", "原价", "厂家", "备用厂家"])
    worksheet.append(["SKU-A", "产品A", "M-A", 2, "厂家A", ""])
    workbook.save(master_path)

    with pytest.raises(RuntimeError, match="缺少 sheet: SKU表"):
        cli.load_master_products(master_path)


def test_load_master_products_missing_required_header_fails(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        columns=["库存sku", "产品名称", "型号", "原价", "厂家"],
    )

    with pytest.raises(RuntimeError, match="缺少必需列: 备用厂家"):
        cli.load_master_products(master_path)


def test_load_master_products_accepts_uppercase_sku_header_alias(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存SKU": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        columns=["库存SKU", "产品名称", "型号", "原价", "厂家", "备用厂家"],
    )

    products = cli.load_master_products(master_path)

    assert list(products) == ["SKU-A"]
    assert products["SKU-A"]["stock_sku"] == "SKU-A"


def test_load_master_products_rejects_duplicate_sku_header_aliases(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {
                "库存sku": "SKU-A",
                "库存SKU": "SKU-A",
                "产品名称": "产品A",
                "型号": "M-A",
                "原价": 2,
                "厂家": "厂家A",
            }
        ],
        columns=["库存sku", "库存SKU", "产品名称", "型号", "原价", "厂家", "备用厂家"],
    )

    with pytest.raises(RuntimeError, match="第1行表头重复: 库存sku"):
        cli.load_master_products(master_path)


@pytest.mark.parametrize(
    ("changed_field", "changed_value"),
    [
        ("产品名称", "产品A2"),
        ("型号", "M-A2"),
        ("原价", 3),
        ("厂家", "厂家B"),
        ("备用厂家", "备用B"),
    ],
)
def test_load_master_products_duplicate_stock_sku_with_different_fields_fails(
    tmp_path,
    changed_field,
    changed_value,
):
    master_path = tmp_path / "export_tax.xlsx"
    duplicate_row = {
        "库存sku": "SKU-A",
        "产品名称": "产品A",
        "型号": "M-A",
        "原价": 2,
        "厂家": "厂家A",
        "备用厂家": "备用A",
    }
    duplicate_row[changed_field] = changed_value
    _write_master_xlsx(
        master_path,
        [
            {
                "库存sku": "SKU-A",
                "产品名称": "产品A",
                "型号": "M-A",
                "原价": 2,
                "厂家": "厂家A",
                "备用厂家": "备用A",
            },
            duplicate_row,
        ],
    )

    with pytest.raises(RuntimeError, match="库存sku重复且字段不一致: SKU-A, 首次行=2, 冲突行=3"):
        cli.load_master_products(master_path)


def test_load_master_products_invalid_price_fails(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": "abc", "厂家": "厂家A"}],
    )

    with pytest.raises(RuntimeError, match="原价非数字: abc"):
        cli.load_master_products(master_path)


def test_load_master_products_reads_large_master_with_streaming_rows(tmp_path):
    from openpyxl import Workbook

    master_path = tmp_path / "large_export_tax.xlsx"
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("SKU表")
    worksheet.append(["库存sku", "产品名称", "型号", "原价", "厂家", "备用厂家"])
    for index in range(1, 15001):
        worksheet.append([f"SKU-{index}", f"产品{index}", f"M-{index}", 1.23, f"厂家{index % 7}", ""])
    workbook.save(master_path)

    products = cli.load_master_products(master_path)

    assert len(products) == 15000
    assert products["SKU-1"]["stock_sku"] == "SKU-1"
    assert products["SKU-15000"]["manufacturer"] == "厂家6"


def test_main_outputs_success_json(monkeypatch, tmp_path, capsys):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )
    monkeypatch.setattr(cli, "DELIVERY_CSV_DIR", csv_dir)
    monkeypatch.setattr(cli, "OUTPUT_DIR", tmp_path / "out")
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main(["--delivery-no", "SP260508022", "--master-xlsx", str(master_path)])

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["source"] == "fba_purchase_summary"
    assert Path(payload["output_xlsx"]).is_file()


def test_purchase_summary_main_outputs_success_json(monkeypatch, tmp_path, capsys):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )
    monkeypatch.setattr(purchase_cli, "DELIVERY_CSV_DIR", csv_dir)
    monkeypatch.setattr(purchase_cli, "OUTPUT_DIR", tmp_path / "out")
    monkeypatch.setattr(purchase_cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = purchase_cli.main(["--delivery-no", "SP260508022", "--master-xlsx", str(master_path)])

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["source"] == "fba_purchase_summary"
    assert Path(payload["output_xlsx"]).name == "SP260508022_purchase_summary.xlsx"
    assert Path(payload["output_xlsx"]).is_file()


def test_generate_fba_restock_workbook_writes_single_sp_restock_sheet(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    csv_path = csv_dir / "SP260605003_1.csv"
    _write_delivery_csv(csv_path, ["SKU-B × 3，SKU-A × 2，SKU-X × 4"], countries=["德国"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
        ],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260605003"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
        today=date(2026, 6, 8),
    )

    output_path = Path(payload["output_xlsx"])
    assert payload["success"] is True
    assert payload["source"] == "fba_restock_workbook"
    assert payload["delivery_no"] == "SP260605003"
    assert payload["csv_path"] == str(csv_path)
    assert payload["country"] == "德国"
    assert payload["matched_sku_count"] == 2
    assert payload["unmatched_sku_count"] == 1
    assert payload["contract_mapping_count"] == 1
    assert payload["gross_margin"] == "0.3"
    assert payload["pricing_basis"] == "tax_exclusive_cost"
    assert Path(payload["output_xlsx"]).name == "6.8-SP260605003-新棱镜备货-德国.xlsx"
    assert _sheet_names(output_path) == ["备货单", "未匹配"]
    restock_values = _sheet_values(output_path, "备货单")
    assert restock_values[0] == RESTOCK_COLUMNS
    assert restock_values[1] == (
        "SKU-B\nSKU-A",
        "产品B\n产品A",
        "SKU-B",
        "产品B",
        "JZ-19",
        2,
        None,
        2.53,
        None,
        0.3,
        "厂家A",
        "个",
        "合同产品A",
        5,
        10,
        None,
        12.65,
        None,
    )
    assert restock_values[2] == _restock_total_row(5, 10, 12.65)
    assert _cell_fill_rgb(output_path, "备货单", "A3") == cli.TOTAL_ROW_FILL_COLOR
    assert _sheet_values(output_path, "未匹配") == [
        RESTOCK_UNMATCHED_COLUMNS,
        ("SKU-X", 4, "出口退税总表未找到库存sku"),
    ]
    widths, heights = _sheet_dimensions(output_path, "备货单")
    assert widths == [15] * 18
    assert heights == [15] * 3
    assert _cell_wrap_text(output_path, "备货单", "A2") is True
    assert _cell_number_format(output_path, "备货单", "H2") == "0.00"
    assert _cell_number_format(output_path, "备货单", "O2") == "0.00"
    assert _cell_number_format(output_path, "备货单", "P2") == "General"
    assert _cell_number_format(output_path, "备货单", "Q2") == "0.00"
    assert _cell_number_format(output_path, "备货单", "R2") == "General"


def test_generate_fba_restock_workbook_uses_unknown_country_when_country_column_is_missing(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260605003_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260605003"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
        today=date(2026, 7, 2),
    )

    assert payload["country"] == "未知国家"
    assert Path(payload["output_xlsx"]).name == "7.2-SP260605003-新棱镜备货-未知国家.xlsx"
    assert "缺少 `国家` 字段" in payload["warnings"][-1]


def test_generate_fba_restock_workbook_uses_unknown_country_when_country_is_empty(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260605003_1.csv", ["SKU-A × 1"], include_country=True)
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260605003"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
        today=date(2026, 12, 5),
    )

    assert payload["country"] == "未知国家"
    assert Path(payload["output_xlsx"]).name == "12.5-SP260605003-新棱镜备货-未知国家.xlsx"
    assert "`国家` 字段为空" in payload["warnings"][-1]


def test_generate_fba_restock_workbook_uses_first_country_and_sanitizes_file_name(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(
        csv_dir / "SP260605003_1.csv",
        ["SKU-A × 1", "SKU-B × 1"],
        countries=["德/国", "法国"],
    )
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-20", "原价": 2, "厂家": "厂家A"},
        ],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260605003"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
        today=date(2026, 6, 8),
    )

    assert payload["country"] == "德/国"
    assert Path(payload["output_xlsx"]).name == "6.8-SP260605003-新棱镜备货-德_国.xlsx"
    assert "存在多个不同国家" in payload["warnings"][-1]


def test_generate_fba_restock_workbook_writes_zhengfei_average_sale_price(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1，SKU-B × 2"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "深圳正飞科技"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-20", "原价": 3, "厂家": "深圳正飞科技"},
        ],
        contract_rows=[{"供货方": "深圳正飞科技", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    rows = _sheet_values(Path(payload["output_xlsx"]), "备货单")
    assert rows[0] == RESTOCK_COLUMNS
    assert rows[1] == (
        "SKU-A",
        "产品A",
        "SKU-A",
        "产品A",
        "JZ-19",
        2,
        2.67,
        2.53,
        3.38,
        0.3,
        "深圳正飞科技",
        "个",
        "合同产品A",
        1,
        2,
        2.67,
        2.53,
        3.38,
    )
    assert rows[2] == (
        "SKU-B",
        "产品B",
        "SKU-B",
        "产品B",
        "JZ-20",
        3,
        2.67,
        3.79,
        3.38,
        0.3,
        "深圳正飞科技",
        "个",
        "合同产品A",
        2,
        6,
        5.34,
        7.58,
        6.76,
    )
    assert rows[3] == _restock_total_row(
        3,
        8,
        10.11,
        average_total_price=8.01,
        average_sale_total_price=10.14,
    )
    output_path = Path(payload["output_xlsx"])
    assert _cell_number_format(output_path, "备货单", "G2") == "0.00"
    assert _cell_number_format(output_path, "备货单", "I2") == "0.00"
    assert _cell_number_format(output_path, "备货单", "P2") == "0.00"
    assert _cell_number_format(output_path, "备货单", "R2") == "0.00"


def test_generate_purchase_batch_workbooks_uses_batch_zhengfei_average_for_each_restock(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1，SKU-C × 4"], countries=["德国"])
    _write_delivery_csv(csv_dir / "SP260508023_1.csv", ["SKU-B × 2"], countries=["德国"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "深圳正飞科技"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-20", "原价": 3, "厂家": "深圳正飞科技"},
            {"库存sku": "SKU-C", "产品名称": "产品C", "型号": "M-C", "原价": 3, "厂家": "厂家A"},
        ],
        contract_rows=[
            {"供货方": "深圳正飞科技", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"},
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品C", "税率": "13%"},
        ],
    )

    payload = batch_cli.generate_purchase_batch_workbooks(
        ["SP260508022", "SP260508023"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        purchase_output_dir=tmp_path / "purchase",
        restock_output_dir=tmp_path / "restock",
        today=date(2026, 6, 8),
    )

    assert payload["success"] is True
    assert payload["source"] == "fba_purchase_batch_workbooks"
    assert payload["delivery_nos"] == ["SP260508022", "SP260508023"]
    assert len(payload["restock_xlsx_paths"]) == 2
    assert payload["restock_outputs"] == [
        {"delivery_no": "SP260508022", "output_xlsx": payload["restock_xlsx_paths"][0]},
        {"delivery_no": "SP260508023", "output_xlsx": payload["restock_xlsx_paths"][1]},
    ]
    assert payload["gross_margin"] == "0.3"
    assert payload["matched_sku_count"] == 3
    assert payload["unmatched_sku_count"] == 0
    assert payload["restock_matched_sku_count"] == 3
    assert payload["restock_unmatched_sku_count"] == 0

    purchase_path = Path(payload["purchase_summary_xlsx"])
    assert _sheet_values(purchase_path, "采购汇总") == [
        PURCHASE_COLUMNS,
        _purchase_row(
            "SKU-A",
            "产品A",
            "SP260508022",
            "JZ-19",
            2,
            "深圳正飞科技",
            1,
            2,
            "个",
            "合同产品A",
            None,
            "13%",
            average_price=2.67,
            average_total_price=2.67,
        ),
        _purchase_row("SKU-C", "产品C", "SP260508022", "M-C", 3, "厂家A", 4, 12, "个", "合同产品C", None, "13%"),
        _purchase_row(
            "SKU-B",
            "产品B",
            "SP260508023",
            "JZ-20",
            3,
            "深圳正飞科技",
            2,
            6,
            "个",
            "合同产品A",
            None,
            "13%",
            average_price=2.67,
            average_total_price=5.34,
        ),
        _purchase_total_row(7, 20, 8.01),
    ]

    first_restock_path = Path(payload["restock_xlsx_paths"][0])
    assert first_restock_path.name == "6.8-SP260508022-新棱镜备货-德国.xlsx"
    assert _sheet_values(first_restock_path, "备货单") == [
        RESTOCK_COLUMNS,
        (
            "SKU-A",
            "产品A",
            "SKU-A",
            "产品A",
            "JZ-19",
            2,
            2.67,
            2.53,
            3.38,
            0.3,
            "深圳正飞科技",
            "个",
            "合同产品A",
            1,
            2,
            2.67,
            2.53,
            3.38,
        ),
        (
            "SKU-C",
            "产品C",
            "SKU-C",
            "产品C",
            "M-C",
            3,
            None,
            3.79,
            None,
            0.3,
            "厂家A",
            "个",
            "合同产品C",
            4,
            12,
            None,
            15.16,
            None,
        ),
        _restock_total_row(5, 14, 17.69, average_total_price=2.67, average_sale_total_price=3.38),
    ]

    second_restock_path = Path(payload["restock_xlsx_paths"][1])
    assert second_restock_path.name == "6.8-SP260508023-新棱镜备货-德国.xlsx"
    assert _sheet_values(second_restock_path, "备货单") == [
        RESTOCK_COLUMNS,
        (
            "SKU-B",
            "产品B",
            "SKU-B",
            "产品B",
            "JZ-20",
            3,
            2.67,
            3.79,
            3.38,
            0.3,
            "深圳正飞科技",
            "个",
            "合同产品A",
            2,
            6,
            5.34,
            7.58,
            6.76,
        ),
        _restock_total_row(2, 6, 7.58, average_total_price=5.34, average_sale_total_price=6.76),
    ]


def test_generate_purchase_batch_workbooks_missing_local_delivery_csv_fails(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )

    with pytest.raises(RuntimeError, match="SP260508023_\\*.csv"):
        batch_cli.generate_purchase_batch_workbooks(
            ["SP260508022", "SP260508023"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=csv_dir,
            purchase_output_dir=tmp_path,
            restock_output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_accepts_tax_rate_forms_for_sale_price(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1，SKU-B × 1，SKU-C × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "M-B", "原价": 2, "厂家": "厂家B"},
            {"库存sku": "SKU-C", "产品名称": "产品C", "型号": "M-C", "原价": 2, "厂家": "厂家C"},
        ],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"},
            {"供货方": "厂家B", "单位": "个", "合同产品名称": "合同产品B", "税率": "9%"},
            {"供货方": "厂家C", "单位": "个", "合同产品名称": "合同产品C", "税率": "1.13"},
        ],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    rows = _sheet_values(Path(payload["output_xlsx"]), "备货单")
    assert rows[1][7] == 2.53
    assert rows[2][7] == 2.62
    assert rows[3][7] == 2.53


@pytest.mark.parametrize("gross_margin", ["0.19", "0.51", "abc"])
def test_generate_fba_restock_workbook_rejects_invalid_gross_margin(gross_margin):
    with pytest.raises(ValueError, match="毛利率必须是 0.2～0.5 之间的数字"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx="missing.xlsx",
            gross_margin=gross_margin,
        )


def test_generate_fba_restock_workbook_rejects_missing_tax_rate(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A"}],
        contract_columns=["供货方", "单位", "合同产品名称"],
    )

    with pytest.raises(RuntimeError, match="备货单售价计算缺少税率"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_rejects_unmapped_tax_rate(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家B", "单位": "个", "合同产品名称": "合同产品B", "税率": "13%"}],
    )

    with pytest.raises(RuntimeError, match="备货单售价计算缺少税率"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_rejects_conflicting_tax_rate(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"},
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "9%"},
        ],
    )

    with pytest.raises(RuntimeError, match="备货单售价计算缺少税率"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_rejects_invalid_tax_rate(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "abc"}],
    )

    with pytest.raises(RuntimeError, match="备货单售价计算税率无法解析"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_rejects_multiple_delivery_nos(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )

    with pytest.raises(ValueError, match="一次只能处理一个 --delivery-no"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022", "SP260508023"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=tmp_path,
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_missing_local_delivery_csv_fails(tmp_path):
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
    )

    with pytest.raises(RuntimeError, match="本地未找到发货单 CSV"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=tmp_path / "missing",
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_missing_master_fails(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"])

    with pytest.raises(RuntimeError, match="找不到出口退税总表"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=tmp_path / "missing.xlsx",
            gross_margin="0.3",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_generate_fba_restock_workbook_warns_same_model_across_manufacturers(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"], countries=["德国"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-19", "原价": 2, "厂家": "厂家B"},
        ],
        contract_rows=[
            {"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"},
            {"供货方": "厂家B", "单位": "个", "合同产品名称": "合同产品B", "税率": "13%"},
        ],
    )

    payload = restock_cli.generate_fba_restock_workbook(
        ["SP260508022"],
        master_xlsx=master_path,
        gross_margin="0.3",
        csv_dir=csv_dir,
        output_dir=tmp_path,
    )

    assert payload["cross_manufacturer_model_count"] == 1
    assert payload["warnings"] == [
        "不同厂家有相同型号，已保留为不同行，请业务人员核查: count=1, examples=JZ-19: 厂家A, 厂家B"
    ]
    restock_values = _sheet_values(Path(payload["output_xlsx"]), "备货单")
    assert restock_values[0] == RESTOCK_COLUMNS
    assert restock_values[1][:6] == ("SKU-A", "产品A", "SKU-A", "产品A", "JZ-19", 2)
    assert restock_values[1][7] == 2.53
    assert restock_values[1][9:] == (0.3, "厂家A", "个", "合同产品A", 2, 4, None, 5.06, None)
    assert restock_values[2][:6] == ("SKU-B", "产品B", "SKU-B", "产品B", "JZ-19", 2)
    assert restock_values[2][7] == 2.53
    assert restock_values[2][9:] == (0.3, "厂家B", "个", "合同产品B", 3, 6, None, 7.59, None)
    assert restock_values[3] == _restock_total_row(5, 10, 12.65)


def test_generate_fba_restock_workbook_rejects_same_manufacturer_model_with_different_price(tmp_path):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 2，SKU-B × 3"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [
            {"库存sku": "SKU-A", "产品名称": "产品A", "型号": "JZ-19", "原价": 2, "厂家": "厂家A"},
            {"库存sku": "SKU-B", "产品名称": "产品B", "型号": "JZ-19", "原价": 3, "厂家": "厂家A"},
        ],
    )

    with pytest.raises(RuntimeError, match="同一厂家同一型号的原价不一致"):
        restock_cli.generate_fba_restock_workbook(
            ["SP260508022"],
            master_xlsx=master_path,
            gross_margin="0.3",
            csv_dir=csv_dir,
            output_dir=tmp_path,
        )


def test_fba_restock_main_outputs_success_json(monkeypatch, tmp_path, capsys):
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(csv_dir / "SP260508022_1.csv", ["SKU-A × 1"], countries=["德国"])
    master_path = tmp_path / "export_tax.xlsx"
    _write_master_xlsx(
        master_path,
        [{"库存sku": "SKU-A", "产品名称": "产品A", "型号": "M-A", "原价": 2, "厂家": "厂家A"}],
        contract_rows=[{"供货方": "厂家A", "单位": "个", "合同产品名称": "合同产品A", "税率": "13%"}],
    )
    monkeypatch.setattr(restock_cli, "DELIVERY_CSV_DIR", csv_dir)
    monkeypatch.setattr(restock_cli, "OUTPUT_DIR", tmp_path / "out")
    monkeypatch.setattr(restock_cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = restock_cli.main(
        ["--delivery-no", "SP260508022", "--master-xlsx", str(master_path), "--gross-margin", "0.3"]
    )

    payload = _read_payload(capsys)
    assert exit_code == 0
    assert payload["success"] is True
    assert payload["source"] == "fba_restock_workbook"
    assert payload["gross_margin"] == "0.3"
    today = date.today()
    assert Path(payload["output_xlsx"]).name == f"{today.month}.{today.day}-SP260508022-新棱镜备货-德国.xlsx"
    assert Path(payload["output_xlsx"]).is_file()


def test_fba_restock_main_outputs_failure_json_for_invalid_gross_margin(monkeypatch, tmp_path, capsys):
    monkeypatch.setattr(restock_cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = restock_cli.main(
        ["--delivery-no", "SP260508022", "--master-xlsx", str(tmp_path / "missing.xlsx"), "--gross-margin", "0.9"]
    )

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload["success"] is False
    assert payload["gross_margin"] == "0.9"
    assert payload["exception"] == "毛利率必须是 0.2～0.5 之间的数字"
