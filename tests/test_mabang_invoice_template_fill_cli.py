from __future__ import annotations

import asyncio
import csv
import json
from collections import OrderedDict
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest

from services.agent_cli.mabang import fill_invoice_template as cli


def _read_payload(capsys) -> dict:
    output = capsys.readouterr().out.strip().splitlines()
    assert output
    return json.loads(output[-1])


async def _noop_close_all_network_clients() -> None:
    return None


def _write_input_workbook(
    path: Path,
    rows: list[dict[str, object]],
    *,
    sheet_count: int = 3,
    merge_rows: list[dict[str, object]] | None = None,
    summary_headers: tuple[str, ...] = cli.INPUT_HEADERS,
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.active.title = "Sheet2"
    while len(workbook.worksheets) < sheet_count:
        workbook.create_sheet(f"Sheet{len(workbook.worksheets) + 1}")
    if sheet_count >= 2:
        merge_worksheet = workbook.worksheets[1]
        merge_worksheet.title = "Sheet1"
        merge_worksheet.append(list(cli.MERGE_DETAIL_HEADERS))
        for row in list(merge_rows if merge_rows is not None else rows):
            merge_worksheet.append(
                [
                    row.get("SKU", ""),
                    row.get("产品名称", row.get("品名", "")),
                    row.get("发货量", ""),
                    row.get("规则型号", row.get("规格型号", "")),
                    row.get("单价", ""),
                ]
            )
    if sheet_count >= 3:
        worksheet = workbook.worksheets[2]
        worksheet.title = "汇总表"
        worksheet.append(list(summary_headers))
        for row in rows:
            worksheet.append([row.get(header, "") for header in summary_headers])
    workbook.save(path)


def _append_summary_sheet(workbook, title: str, rows: list[dict[str, object]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(cli.INPUT_HEADERS))
    for row in rows:
        worksheet.append([row.get(header, "") for header in cli.INPUT_HEADERS])


def _append_merge_detail_sheet(workbook, title: str, rows: list[dict[str, object]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(cli.MERGE_DETAIL_HEADERS))
    for row in rows:
        worksheet.append(
            [
                row.get("SKU", ""),
                row.get("产品名称", row.get("品名", "")),
                row.get("发货量", ""),
                row.get("规则型号", row.get("规格型号", "")),
                row.get("单价", ""),
            ]
        )


def _write_invoice_input_workbook_with_order(
    path: Path,
    rows: list[dict[str, object]],
    *,
    merge_rows: list[dict[str, object]] | None = None,
    order: tuple[str, ...] = ("other", "merge", "summary"),
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, item in enumerate(order, start=1):
        if item == "summary":
            _append_summary_sheet(workbook, "汇总表" if "汇总表" not in workbook.sheetnames else f"汇总表{index}", rows)
        elif item == "merge":
            _append_merge_detail_sheet(workbook, f"财务明细{index}", list(merge_rows if merge_rows is not None else rows))
        else:
            worksheet = workbook.create_sheet(f"Sheet{index}")
            worksheet.append(["无关", "sheet"])
    workbook.save(path)


def _valid_source_row(**overrides: object) -> dict[str, object]:
    row: dict[str, object] = {
        "日期": "2026-04-26",
        "SKU": "SKU-A",
        "品名": "硅胶表带",
        "规格型号": "38MM",
        "发货量": 2,
        "单价": 1.5,
        "供货商": "供应商",
        "采购订单号": "PO-1",
        "采购总价": 3,
        "商品名称": "智能手表表带",
        "售价": 5,
        "总价": 10,
        "单位": "个",
    }
    row.update(overrides)
    return row


def _write_invoice_template(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = cli.INVOICE_TEMPLATE_SHEET
    for _ in range(14):
        worksheet.append([])
    worksheet.append(list(cli.INVOICE_TEMPLATE_HEADERS))
    worksheet.append(["旧数据"] * len(cli.INVOICE_TEMPLATE_HEADERS))
    workbook.save(path)


def _png_bytes(color: str = "red") -> BytesIO:
    from PIL import Image as PILImage

    stream = BytesIO()
    image = PILImage.new("RGB", (20, 20), color)
    image.save(stream, format="PNG")
    stream.seek(0)
    return stream


def _write_stock_sku_xlsx(path: Path, rows: list[dict[str, str]], *, image_skus: set[str] | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image

    image_skus = set(image_skus or set())
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append([cli.STOCK_SKU_COLUMN, "库存SKU中文名称", cli.STOCK_SKU_IMAGE_COLUMN])
    image_streams: list[BytesIO] = []
    for row in rows:
        worksheet.append([row.get(cli.STOCK_SKU_COLUMN, ""), row.get("库存SKU中文名称", ""), ""])
        sku = str(row.get(cli.STOCK_SKU_COLUMN, "") or "").strip()
        if sku in image_skus:
            stream = _png_bytes()
            image_streams.append(stream)
            worksheet.add_image(Image(stream), f"C{worksheet.max_row}")
    workbook.save(path)


def _write_delivery_csv(path: Path, rows: list[dict[str, object]]) -> None:
    headers = [cli.DELIVERY_MSKU_COLUMN, cli.SKU_SHIP_QTY_COLUMN]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _write_consignment_excel(path: Path, rows: list[dict[str, object]]) -> None:
    from openpyxl import Workbook

    headers = ["箱序号", "MSKU", "装箱数量", "长", "宽", "高", "毛重"]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FBA装箱任务"
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def test_main_missing_input_returns_failure_json(monkeypatch, capsys):
    monkeypatch.setattr(cli, "close_all_network_clients", _noop_close_all_network_clients)

    exit_code = cli.main([])

    payload = _read_payload(capsys)
    assert exit_code == 1
    assert payload == {
        "success": False,
        "exception": "input_xlsx 不能为空",
    }


def test_read_invoice_source_rows_requires_summary_sheet(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    _write_input_workbook(input_path, [_valid_source_row()], sheet_count=2)

    with pytest.raises(ValueError, match="缺少 汇总表"):
        cli.read_invoice_source_rows(input_path)


def test_read_invoice_source_rows_accepts_summary_sheet_in_any_position(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    _write_invoice_input_workbook_with_order(
        input_path,
        [_valid_source_row(SKU="SKU-A")],
        order=("summary",),
    )

    rows = cli.read_invoice_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].sku == "SKU-A"


def test_read_invoice_source_rows_does_not_require_supplier_or_purchase_order(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    assert "供货商" not in cli.INPUT_HEADERS
    assert "采购订单号" not in cli.INPUT_HEADERS
    _write_input_workbook(input_path, [_valid_source_row(SKU="SKU-A")])

    rows = cli.read_invoice_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].sku == "SKU-A"


def test_read_invoice_source_rows_accepts_unordered_summary_headers(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    headers = (
        "总价",
        "单位",
        "日期",
        "SKU",
        "商品名称",
        "售价",
        "采购总价",
        "单价",
        "品名",
        "发货量",
        "规格型号",
    )
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 品名="硅胶表带", 规格型号="42MM", 发货量=7, 单价=2.5)],
        summary_headers=headers,
    )

    rows = cli.read_invoice_source_rows(input_path)

    assert len(rows) == 1
    assert rows[0].sku == "SKU-A"
    assert rows[0].source_name == "硅胶表带"
    assert rows[0].model == "42MM"
    assert rows[0].quantity == 7
    assert rows[0].purchase_price == 2.5


def test_read_invoice_source_rows_requires_non_ignored_summary_headers(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    headers = tuple(header for header in cli.INPUT_HEADERS if header != "SKU")
    _write_input_workbook(input_path, [_valid_source_row()], summary_headers=headers)

    with pytest.raises(ValueError, match="缺少必需表头: SKU"):
        cli.read_invoice_source_rows(input_path)


def test_read_invoice_source_rows_rejects_duplicate_required_summary_header(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    headers = ("日期", "SKU", "SKU") + tuple(
        header for header in cli.INPUT_HEADERS if header not in {"日期", "SKU"}
    )
    _write_input_workbook(input_path, [_valid_source_row()], summary_headers=headers)

    with pytest.raises(ValueError, match="表头重复: SKU"):
        cli.read_invoice_source_rows(input_path)


def test_read_stock_sku_merge_infos_accepts_merge_detail_sheet_in_any_position(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    _write_invoice_input_workbook_with_order(
        input_path,
        [_valid_source_row(SKU="SKU-A")],
        merge_rows=[_valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.23, 发货量=5)],
        order=("summary", "other", "merge"),
    )

    merge_infos = cli.read_stock_sku_merge_infos(input_path)

    info = merge_infos[cli.normalize_sku_key("SKU-A")]
    assert info.merge_key == ("M-1", cli.Decimal("1.23"))
    assert info.quantity == cli.Decimal("5")


def test_read_stock_sku_merge_infos_fails_when_multiple_candidates_exist(tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    _write_invoice_input_workbook_with_order(
        input_path,
        [_valid_source_row(SKU="SKU-A")],
        order=("merge", "summary", "merge"),
    )

    with pytest.raises(ValueError, match="多个财务合并明细表候选 sheet.*财务明细1.*财务明细3"):
        cli.read_stock_sku_merge_infos(input_path)


def test_fill_invoice_template_writes_fields_and_copies_images(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国（4.28）-2.xlsx"
    template_path = tmp_path / "invoice_Template.xlsx"
    stock_path = tmp_path / "stock.xlsx"
    delivery_path = tmp_path / "SP260414001_1.csv"
    consignment_path = tmp_path / "SP260414001.xlsx"
    output_dir = tmp_path / "out"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 规格型号="M-A", 单价=1.1, 商品名称="智能手表表带", 品名="硅胶表带", 发货量=2),
            _valid_source_row(SKU="SKU-C", 规格型号="M-C", 单价=1.2, 商品名称="编织表带", 品名="尼龙编织表带", 发货量=3),
            _valid_source_row(SKU="SKU-B", 规格型号="M-B", 单价=1.3, 商品名称="智能手表表带", 品名="真皮表带", 发货量=4),
        ],
    )
    _write_delivery_csv(
        delivery_path,
        [
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-A", cli.SKU_SHIP_QTY_COLUMN: "SKU-A × 2"},
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-C", cli.SKU_SHIP_QTY_COLUMN: "SKU-C × 3"},
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-B", cli.SKU_SHIP_QTY_COLUMN: "SKU-B × 4"},
        ],
    )
    _write_consignment_excel(
        consignment_path,
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 10.5},
            {"箱序号": 1, "MSKU": "MSKU-C", "装箱数量": 3, "长": 40, "宽": 30, "高": 20, "毛重": 10.5},
            {"箱序号": 2, "MSKU": "MSKU-B", "装箱数量": 4, "长": 50, "宽": 35, "高": 25, "毛重": 8.2},
        ],
    )
    _write_invoice_template(template_path)
    _write_stock_sku_xlsx(
        stock_path,
        [
            {cli.STOCK_SKU_COLUMN: "SKU-A", "库存SKU中文名称": "产品A"},
            {cli.STOCK_SKU_COLUMN: "SKU-C", "库存SKU中文名称": "产品C"},
            {cli.STOCK_SKU_COLUMN: "SKU-B", "库存SKU中文名称": "产品B"},
        ],
        image_skus={"SKU-A", "SKU-C", "SKU-B"},
    )

    async def fake_export_stock_sku_names(skus, *, delivery_no="", output_dir=None, **kwargs):
        assert skus == ["SKU-A", "SKU-C", "SKU-B"]
        assert delivery_no == "SP260414001_invoice"
        return SimpleNamespace(xlsx_paths=[str(stock_path)])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.fill_invoice_template(
            input_path,
            template_path=template_path,
            output_dir=output_dir,
            stock_sku_output_dir=tmp_path / "stock-out",
            delivery_csv=delivery_path,
            consignment_excel=consignment_path,
        )
    )

    assert payload["success"] is True
    assert payload["sp_no"] == "SP260414001"
    assert payload["destination_country"] == "美国"
    assert payload["row_count"] == 3
    assert payload["source_row_count"] == 3
    assert payload["invoice_row_count"] == 3
    assert payload["box_count"] == 2
    assert payload["delivery_csv_path"] == str(delivery_path.resolve())
    assert payload["consignment_excel_path"] == str(consignment_path.resolve())
    assert payload["image_matched_count"] == 3
    assert payload["image_missing_count"] == 0
    assert payload["stock_sku_xlsx_paths"] == [str(stock_path)]

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"])
    worksheet = workbook[cli.INVOICE_TEMPLATE_SHEET]
    assert worksheet.cell(row=16, column=1).value == "1"
    assert worksheet.cell(row=16, column=3).value == "10.5"
    assert worksheet.cell(row=16, column=4).value == "40"
    assert worksheet.cell(row=16, column=5).value == "30"
    assert worksheet.cell(row=16, column=6).value == "20"
    assert worksheet.cell(row=16, column=7).value == "Watch Band"
    assert worksheet.cell(row=16, column=8).value == "智能手表表带"
    assert worksheet.cell(row=16, column=9).value == 0.35
    assert worksheet.cell(row=16, column=10).value == 2
    assert worksheet.cell(row=16, column=11).value == "个"
    assert worksheet.cell(row=16, column=12).value == "硅胶/silicone"
    assert worksheet.cell(row=16, column=13).value == "9113900090"
    assert worksheet.cell(row=16, column=14).value == "装饰/decorate"
    assert worksheet.cell(row=16, column=15).value == "无"
    assert worksheet.cell(row=16, column=16).value == "无"
    assert worksheet.cell(row=16, column=17).value == "SKU-A"
    assert worksheet.cell(row=17, column=7).value == "Braided Watch Band"
    assert worksheet.cell(row=17, column=9).value == 0.5
    assert worksheet.cell(row=17, column=12).value == "尼龙/nylon"
    assert worksheet.cell(row=17, column=17).value == "SKU-C"
    assert worksheet.cell(row=18, column=1).value == "2"
    assert worksheet.cell(row=18, column=10).value == 4
    assert worksheet.cell(row=18, column=9).value == 0.5
    assert worksheet.cell(row=18, column=12).value == "皮革/leather"
    assert len(getattr(worksheet, "_images", [])) == 3
    workbook.close()


def test_invoice_rules_for_case_box_and_protector(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    template_path = tmp_path / "invoice_Template.xlsx"
    stock_path = tmp_path / "stock.xlsx"
    delivery_path = tmp_path / "SP260414001_1.csv"
    consignment_path = tmp_path / "SP260414001.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-CASE", 规格型号="M-CASE", 单价=1.1, 商品名称="智能手表表壳", 品名="PC表壳", 发货量=10),
            _valid_source_row(SKU="SKU-BOX", 规格型号="M-BOX", 单价=1.2, 商品名称="包装盒", 品名="纸质包装盒", 发货量=5),
            _valid_source_row(SKU="SKU-COVER", 规格型号="M-COVER", 单价=1.3, 商品名称="手表保护套", 品名="TPU手表保护套", 发货量=4),
        ],
    )
    _write_delivery_csv(
        delivery_path,
        [
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-CASE", cli.SKU_SHIP_QTY_COLUMN: "SKU-CASE × 10"},
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-BOX", cli.SKU_SHIP_QTY_COLUMN: "SKU-BOX × 5"},
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-COVER", cli.SKU_SHIP_QTY_COLUMN: "SKU-COVER × 4"},
        ],
    )
    _write_consignment_excel(
        consignment_path,
        [
            {"箱序号": 1, "MSKU": "MSKU-CASE", "装箱数量": 10, "长": 40, "宽": 30, "高": 20, "毛重": 10},
            {"箱序号": 1, "MSKU": "MSKU-BOX", "装箱数量": 5, "长": 40, "宽": 30, "高": 20, "毛重": 10},
            {"箱序号": 1, "MSKU": "MSKU-COVER", "装箱数量": 4, "长": 40, "宽": 30, "高": 20, "毛重": 10},
        ],
    )
    _write_invoice_template(template_path)
    _write_stock_sku_xlsx(
        stock_path,
        [
            {cli.STOCK_SKU_COLUMN: "SKU-CASE", "库存SKU中文名称": "表壳"},
            {cli.STOCK_SKU_COLUMN: "SKU-BOX", "库存SKU中文名称": "包装盒"},
            {cli.STOCK_SKU_COLUMN: "SKU-COVER", "库存SKU中文名称": "保护套"},
        ],
    )

    async def fake_export_stock_sku_names(*args, **kwargs):
        return SimpleNamespace(xlsx_paths=[str(stock_path)])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.fill_invoice_template(
            input_path,
            template_path=template_path,
            output_dir=tmp_path / "out",
            delivery_csv=delivery_path,
            consignment_excel=consignment_path,
        )
    )

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"])
    worksheet = workbook[cli.INVOICE_TEMPLATE_SHEET]
    assert worksheet.cell(row=16, column=9).value == 0.35
    assert worksheet.cell(row=16, column=12).value == "PC/PC"
    assert worksheet.cell(row=16, column=14).value == "装饰/decorate"
    assert worksheet.cell(row=17, column=9).value == 0.32
    assert worksheet.cell(row=17, column=12).value == "纸质+塑料/paper+plastic"
    assert worksheet.cell(row=17, column=14).value == "装表带表壳/Watch strap and case"
    assert worksheet.cell(row=18, column=9).value == 0.35
    assert worksheet.cell(row=18, column=12).value == "TPU/TPU"
    assert worksheet.cell(row=18, column=14).value == "保护手表用/Used for protecting watches"
    workbook.close()


def test_unknown_price_rule_and_missing_image_are_notices(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    template_path = tmp_path / "invoice_Template.xlsx"
    stock_path = tmp_path / "stock.xlsx"
    delivery_path = tmp_path / "SP260414001_1.csv"
    consignment_path = tmp_path / "SP260414001.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-X", 商品名称="未知产品", 品名="未知材质"),
        ],
    )
    _write_delivery_csv(
        delivery_path,
        [{cli.DELIVERY_MSKU_COLUMN: "MSKU-X", cli.SKU_SHIP_QTY_COLUMN: "SKU-X × 2"}],
    )
    _write_consignment_excel(
        consignment_path,
        [{"箱序号": 1, "MSKU": "MSKU-X", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 1}],
    )
    _write_invoice_template(template_path)
    _write_stock_sku_xlsx(stock_path, [{cli.STOCK_SKU_COLUMN: "SKU-X", "库存SKU中文名称": "产品X"}])

    async def fake_export_stock_sku_names(*args, **kwargs):
        return SimpleNamespace(xlsx_paths=[str(stock_path)])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.fill_invoice_template(
            input_path,
            template_path=template_path,
            output_dir=tmp_path / "out",
            delivery_csv=delivery_path,
            consignment_excel=consignment_path,
        )
    )

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"])
    worksheet = workbook[cli.INVOICE_TEMPLATE_SHEET]
    assert worksheet.cell(row=16, column=9).value == cli.UNKNOWN_DECLARED_PRICE_TEXT
    assert worksheet.cell(row=16, column=12).value in (None, "")
    assert worksheet.cell(row=16, column=14).value in (None, "")
    workbook.close()
    assert payload["image_missing_count"] == 1
    assert any("没有该材质的计算价格方式" in item for item in payload["notice"])
    assert any("缺少产品用途规则" in item for item in payload["notice"])
    assert any("缺少产品图片" in item for item in payload["notice"])


def test_box_split_expands_combo_sku_and_merges_same_box_stock_sku(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    template_path = tmp_path / "invoice_Template.xlsx"
    stock_path = tmp_path / "stock.xlsx"
    delivery_path = tmp_path / "SP260414001_1.csv"
    consignment_path = tmp_path / "SP260414001.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 规格型号="M-A", 单价=1.1, 商品名称="智能手表表带", 品名="硅胶表带", 发货量=3),
            _valid_source_row(SKU="SKU-B", 规格型号="M-B", 单价=1.2, 商品名称="编织表带", 品名="尼龙编织表带", 发货量=5),
        ],
    )
    _write_delivery_csv(
        delivery_path,
        [
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-1", cli.SKU_SHIP_QTY_COLUMN: "SKU-A × 2，SKU-B × 4"},
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-2", cli.SKU_SHIP_QTY_COLUMN: "SKU-A × 1，SKU-B × 1"},
        ],
    )
    _write_consignment_excel(
        consignment_path,
        [
            {"箱序号": 1, "MSKU": "MSKU-1", "装箱数量": 1, "长": 40, "宽": 30, "高": 20, "毛重": 6},
            {"箱序号": 1, "MSKU": "MSKU-2", "装箱数量": 1, "长": 40, "宽": 30, "高": 20, "毛重": 6},
            {"箱序号": 2, "MSKU": "MSKU-1", "装箱数量": 1, "长": 50, "宽": 35, "高": 25, "毛重": 7},
        ],
    )
    _write_invoice_template(template_path)
    _write_stock_sku_xlsx(
        stock_path,
        [
            {cli.STOCK_SKU_COLUMN: "SKU-A", "库存SKU中文名称": "产品A"},
            {cli.STOCK_SKU_COLUMN: "SKU-B", "库存SKU中文名称": "产品B"},
        ],
    )

    async def fake_export_stock_sku_names(skus, **kwargs):
        assert skus == ["SKU-A", "SKU-B"]
        return SimpleNamespace(xlsx_paths=[str(stock_path)])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.fill_invoice_template(
            input_path,
            template_path=template_path,
            output_dir=tmp_path / "out",
            delivery_csv=delivery_path,
            consignment_excel=consignment_path,
        )
    )

    from openpyxl import load_workbook

    assert payload["source_row_count"] == 2
    assert payload["invoice_row_count"] == 4
    assert payload["box_count"] == 2
    workbook = load_workbook(payload["output_xlsx"])
    worksheet = workbook[cli.INVOICE_TEMPLATE_SHEET]
    assert [worksheet.cell(row=row, column=1).value for row in range(16, 20)] == ["1", "1", "2", "2"]
    assert [worksheet.cell(row=row, column=17).value for row in range(16, 20)] == ["SKU-A", "SKU-B", "SKU-A", "SKU-B"]
    assert [worksheet.cell(row=row, column=10).value for row in range(16, 20)] == [2, 3, 1, 2]
    assert worksheet.cell(row=16, column=3).value == "6"
    assert worksheet.cell(row=18, column=3).value == "7"
    workbook.close()


def test_box_split_merges_finance_skus_by_model_and_price_per_box(monkeypatch, tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    template_path = tmp_path / "invoice_Template.xlsx"
    stock_path = tmp_path / "stock.xlsx"
    delivery_path = tmp_path / "SP260416020_1.csv"
    consignment_path = tmp_path / "SP260416020.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 商品名称="硅胶表带", 品名="硅胶表带", 发货量=10),
        ],
        merge_rows=[
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.616, 发货量=6),
            _valid_source_row(SKU="SKU-B", 规格型号="M-1", 单价=3.616, 发货量=4),
        ],
    )
    _write_delivery_csv(
        delivery_path,
        [
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-A", cli.SKU_SHIP_QTY_COLUMN: "SKU-A × 6"},
            {cli.DELIVERY_MSKU_COLUMN: "MSKU-B", cli.SKU_SHIP_QTY_COLUMN: "SKU-B × 4"},
        ],
    )
    _write_consignment_excel(
        consignment_path,
        [
            {"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 3, "长": 40, "宽": 30, "高": 20, "毛重": 6},
            {"箱序号": 1, "MSKU": "MSKU-B", "装箱数量": 1, "长": 40, "宽": 30, "高": 20, "毛重": 6},
            {"箱序号": 2, "MSKU": "MSKU-A", "装箱数量": 3, "长": 50, "宽": 35, "高": 25, "毛重": 7},
            {"箱序号": 2, "MSKU": "MSKU-B", "装箱数量": 3, "长": 50, "宽": 35, "高": 25, "毛重": 7},
        ],
    )
    _write_invoice_template(template_path)
    _write_stock_sku_xlsx(stock_path, [{cli.STOCK_SKU_COLUMN: "SKU-A", "库存SKU中文名称": "产品A"}])

    async def fake_export_stock_sku_names(skus, **kwargs):
        assert skus == ["SKU-A"]
        return SimpleNamespace(xlsx_paths=[str(stock_path)])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.fill_invoice_template(
            input_path,
            template_path=template_path,
            output_dir=tmp_path / "out",
            delivery_csv=delivery_path,
            consignment_excel=consignment_path,
        )
    )

    from openpyxl import load_workbook

    assert payload["source_row_count"] == 1
    assert payload["invoice_row_count"] == 2
    assert payload["merge_group_count"] == 1
    assert payload["merged_sku_count"] == 1
    workbook = load_workbook(payload["output_xlsx"])
    worksheet = workbook[cli.INVOICE_TEMPLATE_SHEET]
    assert [worksheet.cell(row=row, column=1).value for row in range(16, 18)] == ["1", "2"]
    assert [worksheet.cell(row=row, column=17).value for row in range(16, 18)] == ["SKU-A", "SKU-A"]
    assert [worksheet.cell(row=row, column=10).value for row in range(16, 18)] == [4, 6]
    workbook.close()


def test_fill_invoice_template_infers_blank_summary_model_from_merge_detail(monkeypatch, tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    template_path = tmp_path / "invoice_Template.xlsx"
    stock_path = tmp_path / "stock.xlsx"
    delivery_path = tmp_path / "SP260416020_1.csv"
    consignment_path = tmp_path / "SP260416020.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 规格型号="", 单价=2.5, 商品名称="硅胶表带", 品名="硅胶表带", 发货量=5)],
        merge_rows=[_valid_source_row(SKU="SKU-A", 规格型号="ZF-5", 单价=2.5, 发货量=5)],
    )
    _write_delivery_csv(
        delivery_path,
        [{cli.DELIVERY_MSKU_COLUMN: "MSKU-A", cli.SKU_SHIP_QTY_COLUMN: "SKU-A × 5"}],
    )
    _write_consignment_excel(
        consignment_path,
        [{"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 5, "长": 40, "宽": 30, "高": 20, "毛重": 6}],
    )
    _write_invoice_template(template_path)
    _write_stock_sku_xlsx(stock_path, [{cli.STOCK_SKU_COLUMN: "SKU-A", "库存SKU中文名称": "产品A"}])

    async def fake_export_stock_sku_names(skus, **kwargs):
        assert skus == ["SKU-A"]
        return SimpleNamespace(xlsx_paths=[str(stock_path)])

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    payload = asyncio.run(
        cli.fill_invoice_template(
            input_path,
            template_path=template_path,
            output_dir=tmp_path / "out",
            delivery_csv=delivery_path,
            consignment_excel=consignment_path,
        )
    )

    assert payload["invoice_row_count"] == 1
    assert payload["merge_group_count"] == 1
    assert any("规格型号为空，已按财务合并明细表补为 ZF-5" in item for item in payload["notice"])

    from openpyxl import load_workbook

    workbook = load_workbook(payload["output_xlsx"])
    worksheet = workbook[cli.INVOICE_TEMPLATE_SHEET]
    assert worksheet.cell(row=16, column=17).value == "SKU-A"
    assert worksheet.cell(row=16, column=10).value == 5
    workbook.close()


def test_blank_summary_model_fails_when_merge_detail_has_no_same_sku(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 规格型号="", 单价=2.5, 发货量=5)],
        merge_rows=[_valid_source_row(SKU="SKU-B", 规格型号="ZF-5", 单价=2.5, 发货量=5)],
    )
    source_rows = cli.read_invoice_source_rows(input_path)
    merge_infos = cli.read_stock_sku_merge_infos(input_path)

    with pytest.raises(ValueError, match="财务合并明细表找不到同 SKU"):
        cli.infer_missing_summary_models(source_rows, merge_infos)


def test_blank_summary_model_fails_when_merge_detail_price_mismatches(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 规格型号="", 单价=2.5, 发货量=5)],
        merge_rows=[_valid_source_row(SKU="SKU-A", 规格型号="ZF-5", 单价=2.6, 发货量=5)],
    )
    source_rows = cli.read_invoice_source_rows(input_path)
    merge_infos = cli.read_stock_sku_merge_infos(input_path)

    with pytest.raises(ValueError, match="同 SKU 单价不一致"):
        cli.infer_missing_summary_models(source_rows, merge_infos)


def test_summary_merge_key_duplicate_fails(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 发货量=5),
            _valid_source_row(SKU="SKU-B", 规格型号="M-1", 单价=3.62, 发货量=5),
        ],
    )

    source_rows = cli.read_invoice_source_rows(input_path)
    merge_infos = cli.read_stock_sku_merge_infos(input_path)
    with pytest.raises(ValueError, match="汇总表同一个规则型号\\+单价存在多个保留 SKU"):
        cli.build_invoice_box_rows(source_rows, merge_infos, OrderedDict(), [])


def test_summary_same_merge_key_same_sku_is_summed(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 发货量=2),
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 发货量=3),
        ],
        merge_rows=[_valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 发货量=5)],
    )
    source_rows = cli.read_invoice_source_rows(input_path)
    merge_infos = cli.read_stock_sku_merge_infos(input_path)
    delivery_components = OrderedDict([("MSKU-A", OrderedDict([("SKU-A", cli.Decimal("5"))]))])
    consignment_rows = [
        cli.ConsignmentMskuRow(
            row_number=2,
            box_info=cli.ConsignmentBoxInfo(box_no="1", gross_weight="1", length="1", width="1", height="1"),
            msku="MSKU-A",
            quantity=cli.Decimal("5"),
        )
    ]

    invoice_rows = cli.build_invoice_box_rows(source_rows, merge_infos, delivery_components, consignment_rows)

    assert len(invoice_rows) == 1
    assert invoice_rows[0].source.sku == "SKU-A"
    assert invoice_rows[0].quantity == cli.Decimal("5")


def test_merge_detail_duplicate_same_sku_model_price_is_summed(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.36, 发货量=500)],
        merge_rows=[
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.356, 发货量=25),
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.356, 发货量=215),
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.356, 发货量=260),
        ],
    )

    merge_infos = cli.read_stock_sku_merge_infos(input_path)

    info = merge_infos[cli.normalize_sku_key("SKU-A")]
    assert info.merge_key == ("M-1", cli.Decimal("1.36"))
    assert info.quantity == cli.Decimal("500")


def test_merge_detail_same_sku_different_model_or_price_fails(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.36, 发货量=5)],
        merge_rows=[
            _valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=1.356, 发货量=2),
            _valid_source_row(SKU="SKU-A", 规格型号="M-2", 单价=1.356, 发货量=3),
        ],
    )

    with pytest.raises(ValueError, match="同一 SKU 存在不同规则型号或单价"):
        cli.read_stock_sku_merge_infos(input_path)


def test_split_sku_missing_merge_info_fails(tmp_path):
    input_path = tmp_path / "4.27-SP260416020-新棱镜备货-美国.xlsx"
    _write_input_workbook(
        input_path,
        [_valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 发货量=2)],
        merge_rows=[_valid_source_row(SKU="SKU-A", 规格型号="M-1", 单价=3.62, 发货量=2)],
    )
    source_rows = cli.read_invoice_source_rows(input_path)
    merge_infos = cli.read_stock_sku_merge_infos(input_path)
    delivery_components = OrderedDict([("MSKU-X", OrderedDict([("SKU-X", cli.Decimal("2"))]))])
    consignment_rows = [
        cli.ConsignmentMskuRow(
            row_number=2,
            box_info=cli.ConsignmentBoxInfo(box_no="1", gross_weight="1", length="1", width="1", height="1"),
            msku="MSKU-X",
            quantity=cli.Decimal("2"),
        )
    ]
    with pytest.raises(ValueError, match="不在财务合并明细表中"):
        cli.build_invoice_box_rows(source_rows, merge_infos, delivery_components, consignment_rows)


def test_box_split_quantity_mismatch_fails(monkeypatch, tmp_path):
    input_path = tmp_path / "4.26-SP260414001-新棱镜备货-美国.xlsx"
    delivery_path = tmp_path / "SP260414001_1.csv"
    consignment_path = tmp_path / "SP260414001.xlsx"
    _write_input_workbook(input_path, [_valid_source_row(SKU="SKU-A", 发货量=3)])
    _write_delivery_csv(
        delivery_path,
        [{cli.DELIVERY_MSKU_COLUMN: "MSKU-A", cli.SKU_SHIP_QTY_COLUMN: "SKU-A × 2"}],
    )
    _write_consignment_excel(
        consignment_path,
        [{"箱序号": 1, "MSKU": "MSKU-A", "装箱数量": 2, "长": 40, "宽": 30, "高": 20, "毛重": 1}],
    )

    async def fake_export_stock_sku_names(*args, **kwargs):
        raise AssertionError("数量校验失败前不应调用库存 SKU 导出")

    monkeypatch.setattr(cli, "export_stock_sku_names", fake_export_stock_sku_names)

    with pytest.raises(ValueError, match="拆分归并后库存SKU数量与汇总表不一致"):
        asyncio.run(
            cli.fill_invoice_template(
                input_path,
                template_path=tmp_path / "invoice_Template.xlsx",
                output_dir=tmp_path / "out",
                delivery_csv=delivery_path,
                consignment_excel=consignment_path,
            )
        )


def test_fill_invoice_template_requires_local_delivery_csv(tmp_path):
    input_path = tmp_path / "4.26-SP269999999-新棱镜备货-美国.xlsx"
    _write_input_workbook(input_path, [_valid_source_row()])

    with pytest.raises(FileNotFoundError, match="本地未找到发货单 CSV"):
        asyncio.run(cli.fill_invoice_template(input_path, template_path=tmp_path / "invoice_Template.xlsx"))
