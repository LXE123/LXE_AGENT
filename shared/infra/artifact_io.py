import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from shared.logging import logger


REPORT_IO_PREFIX = "[ReportIO]"
EXCEL_COLUMN_HIGHLIGHTS = {
    "真实本地库存": "FCE4D6",
    "海运补货": "DDEBF7",
    "决策原因": "EAD1DC",
    "实际补货": "E2F0D9",
    "建议补货": "FFF2CC",
}


def build_timestamped_path(base_dir: str, prefix: str, ext: str) -> str:
    os.makedirs(base_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(base_dir, f"{prefix}_{timestamp}.{ext}")


def write_json(
    data: Any,
    path: str,
    ensure_ascii: bool = False,
    indent: int = 2,
) -> Optional[str]:
    try:
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(data, handle, ensure_ascii=ensure_ascii, indent=indent)
        return path
    except Exception as error:
        logger.error(f"{REPORT_IO_PREFIX} 写入 JSON 失败: {error}")
        return None


def write_text(text: str, path: str) -> Optional[str]:
    try:
        with open(path, "w", encoding="utf-8") as handle:
            handle.write(text)
        return path
    except Exception as error:
        logger.error(f"{REPORT_IO_PREFIX} 写入文本失败: {error}")
        return None


def write_markdown(markdown_text: str, path: str) -> Optional[str]:
    try:
        with open(path, "w", encoding="utf-8") as handle:
            handle.write(markdown_text)
        return path
    except Exception as error:
        logger.error(f"{REPORT_IO_PREFIX} 写入 Markdown 失败: {error}")
        return None


def _alloc_sheet_name(raw_name: str, used_names: set) -> str:
    base = (raw_name or "Sheet")[:31]
    if base not in used_names:
        used_names.add(base)
        return base

    idx = 1
    while True:
        suffix = f"_{idx}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        idx += 1


def _apply_highlighted_columns(worksheet) -> None:
    try:
        from openpyxl.styles import PatternFill
    except ImportError:
        return

    header_to_col = {}
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        header_to_col[str(cell.value).strip()] = cell.column

    for header, color in EXCEL_COLUMN_HIGHLIGHTS.items():
        col_idx = header_to_col.get(header)
        if not col_idx:
            continue

        fill = PatternFill(fill_type="solid", fgColor=color)
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            row[0].fill = fill


def write_excel_sheets(path: str, sheets: Dict[str, Any]) -> Optional[str]:
    try:
        import pandas as pd
    except ImportError:
        logger.error(f"{REPORT_IO_PREFIX} 未安装 pandas，请先安装: pip install pandas openpyxl")
        return None

    try:
        used_names = set()
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name, value in sheets.items():
                if isinstance(value, pd.DataFrame):
                    df = value
                elif isinstance(value, list):
                    df = pd.DataFrame(value)
                else:
                    df = pd.DataFrame(value)
                sheet_name = _alloc_sheet_name(str(name), used_names)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        return path
    except Exception as error:
        logger.error(f"{REPORT_IO_PREFIX} 写入 Excel 失败: {error}")
        return None


def write_excel_single(path: str, rows: List[Dict[str, Any]]) -> Optional[str]:
    try:
        import pandas as pd
    except ImportError:
        logger.error(f"{REPORT_IO_PREFIX} 未安装 pandas，请先安装: pip install pandas openpyxl")
        return None

    try:
        df = pd.DataFrame(rows)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets.get(writer.book.sheetnames[0])
            if worksheet is not None:
                _apply_highlighted_columns(worksheet)
        return path
    except Exception as error:
        logger.error(f"{REPORT_IO_PREFIX} 写入 Excel 失败: {error}")
        return None


__all__ = [
    "build_timestamped_path",
    "write_excel_sheets",
    "write_excel_single",
    "write_json",
    "write_markdown",
    "write_text",
]
