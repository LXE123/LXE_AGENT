from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from functools import partial
from pathlib import Path
from typing import Any

from services.mabang import config as mabang_settings
from services.mabang.auth_constants import (
    MABANG_MEMCACHE_COOKIE_NAME,
    PRIVATE_AMZ_REQUIRED_COOKIE_NAMES,
)
from services.mabang.export_common import (
    ExportPipelineSpec,
    PrivateAmzExportAuth,
    build_private_amz_headers,
    run_export_pipeline,
)
from services.mabang.export_common import (
    clean_text as _clean_text,
)
from services.mabang.export_common import (
    configured_text as _configured_text,
)
from services.mabang.export_common import (
    excel_suffix_from_url as _excel_suffix_from_url,
)
from services.mabang.export_common import (
    request_headers as _request_headers,
)
from services.mabang.export_common import (
    safe_store_msku_file_part as _safe_file_part,
)
from shared.infra.net import erp_http_session, external_http_session
from shared.datasets import dataset_dir

from ...auth import get_auth_context
from ...errors import MabangAuthError, MabangBusinessError, MabangRequestError
from .store_resolver import ID_TYPE_FBA_WAREHOUSE, ID_TYPE_SHOP

DEFAULT_LISTSEARCH_URL = "https://private-amz.mabangerp.com/index.php?mod=fbanew.listsearch"
DEFAULT_FBA_EXPORT_URL = "https://private.mabangerp.com/index.php?mod=export.doFbaExportFile"
DEFAULT_PRIVATE_AMZ_ORIGIN = "https://private-amz.mabangerp.com"
DEFAULT_PRIVATE_AMZ_REFERER = "https://private-amz.mabangerp.com/"
DEFAULT_PRIVATE_ORIGIN = "https://private.mabangerp.com"
DEFAULT_PRIVATE_REFERER = "https://private.mabangerp.com/"
DEFAULT_OUTPUT_DIR = dataset_dir("replenish_store_msku")
AUTH_FAIL_STATUS = {401, 403}
SOURCE = "mabang_store_msku_download"
STORE_MSKU_FILE_SUFFIX = "店铺MSKU数据"
CORE_STORE_MSKU_HEADERS = ("店铺名称", "MSKU", "ASIN", "本地SKU")
STORE_MSKU_FIELDLABELS = (
    "uq101",
    "uq102",
    "uq165",
    "uq154",
    "uq103",
    "uq104",
    "uq194",
    "uq105",
    "uq106",
    "uq107",
    "uq108",
    "uq109",
    "uq110",
    "uq114",
    "uq118",
    "uq119",
    "uq172",
    "uq173",
    "uq120",
    "uq124",
    "uq125",
    "uq126",
    "uq127",
    "uq128",
    "uq131",
    "uq136",
    "uq138",
    "uq141",
    "uq143",
    "uq164",
    "uq205",
    "uq167",
    "uq178",
    "uq186",
    "uq201",
    "uq202",
)
STORE_MSKU_EXPORT_FIELD_DEFS = (
    ("店铺名称", "uq101", ""),
    ("站点", "uq143", ""),
    ("商品链接", "uq165", ""),
    ("MSKU", "uq102", ""),
    ("父ASIN", "uq194", ""),
    ("ASIN", "uq105", ""),
    ("本地SKU", "uq103", ""),
    ("FNSKU", "uq154", ""),
    ("本地SKU名称", "uq141", ""),
    ("产品名称", "uq104", ""),
    ("7天销量", "uq107", ""),
    ("14天销量", "uq108", ""),
    ("30天销量", "uq109", ""),
    ("90天销量", "uq110", ""),
    ("日均销量", "uq114", ""),
    ("排名", "uq178", ""),
    ("利润（原始货币）", "uq205", ""),
    ("售价", "uq164", ""),
    ("7天退货量", "uq118", ""),
    ("7天退货率", "uq119", ""),
    ("30天退货量", "uq172", ""),
    ("30天退货率", "uq173", ""),
    ("上架时间", "uq167", ""),
    ("库存状态", "uq106", ""),
    ("可售", "uq124", ""),
    ("待入库", "uq125", ""),
    ("在途", "uq128", ""),
    ("预留", "uq126", ""),
    ("待调仓", "uq201", ""),
    ("调仓中", "uq202", ""),
    ("单品重量(g)(cm)", "uq131", "1"),
    ("计划入库", "uq127", ""),
    ("采购在途", "uq120", ""),
    ("总在途量(默认设置)", "uq186", ""),
    ("申请补货量", "uq138", ""),
    ("备注", "uq136", ""),
)
STORE_MSKU_EXPORT_FIELDS = tuple((name, uq) for name, uq, _map_text in STORE_MSKU_EXPORT_FIELD_DEFS)


class StoreMskuDownloadError(MabangBusinessError):
    pass


class StoreMskuDownloadAuthError(StoreMskuDownloadError, MabangAuthError):
    pass


@dataclass(frozen=True)
class StoreMskuExcelResult:
    store_name: str
    store_id: str
    id_type: str
    id_count: int
    xlsx_path: str
    converted: bool
    raw_excel_deleted: bool
    source: str = SOURCE

    def to_payload(self) -> dict[str, Any]:
        return {
            "success": True,
            "store_name": self.store_name,
            "store_id": self.store_id,
            "id_type": self.id_type,
            "id_count": self.id_count,
            "xlsx_path": self.xlsx_path,
            "converted": self.converted,
            "raw_excel_deleted": self.raw_excel_deleted,
            "source": self.source,
        }


def _timestamp_text(value: datetime | None = None) -> str:
    return (value or datetime.now()).strftime("%Y%m%d%H%M")


def _resolve_output_dir(output_dir: str | Path | None = None) -> Path:
    if output_dir is not None:
        path = Path(output_dir)
    else:
        configured = str(mabang_settings.MABANG_STORE_MSKU_OUTPUT_DIR or "").strip()
        path = Path(configured) if configured else DEFAULT_OUTPUT_DIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def normalize_store_id(value: Any) -> str:
    store_id = _clean_text(value)
    if not store_id:
        raise ValueError("store_id 不能为空")
    return store_id


def normalize_store_name(value: Any) -> str:
    store_name = _clean_text(value)
    if not store_name:
        raise ValueError("store_name 不能为空")
    return store_name


def normalize_id_type(value: Any) -> str:
    id_type = _clean_text(value)
    if not id_type:
        raise ValueError("id_type 不能为空")
    if id_type not in {ID_TYPE_FBA_WAREHOUSE, ID_TYPE_SHOP}:
        raise ValueError(f"id_type 只支持 {ID_TYPE_FBA_WAREHOUSE} 或 {ID_TYPE_SHOP}: {id_type}")
    return id_type


def _listsearch_url() -> str:
    return _configured_text("MABANG_STORE_MSKU_LISTSEARCH_URL", DEFAULT_LISTSEARCH_URL)


def _fba_export_url() -> str:
    return _configured_text("MABANG_STORE_MSKU_FBA_EXPORT_URL", DEFAULT_FBA_EXPORT_URL)


async def _read_store_msku_json(resp: Any, *, action: str) -> dict[str, Any]:
    status_code = int(getattr(resp, "status", 0) or 0)
    text = await resp.text()
    if status_code in AUTH_FAIL_STATUS:
        raise StoreMskuDownloadAuthError(f"{action}鉴权失败(status={status_code})")
    if status_code >= 400:
        msg = text[:300] if text else "empty response"
        raise MabangRequestError(f"{action}请求失败(status={status_code}): {msg}")

    data: Any = None
    if text:
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            data = None
    if data is None:
        try:
            data = await resp.json(content_type=None)
        except Exception:
            data = None
    if not isinstance(data, dict):
        raise StoreMskuDownloadError(f"{action}返回非JSON对象")
    if data.get("success") is False:
        message = str(data.get("msg") or data.get("message") or data.get("error") or "unknown").strip()
        raise StoreMskuDownloadError(f"{action}业务异常: {message}")
    return data


async def resolve_store_msku_auth() -> PrivateAmzExportAuth:
    context = await get_auth_context(purpose="store_msku_download")
    return build_private_amz_headers(
        context,
        required_names=PRIVATE_AMZ_REQUIRED_COOKIE_NAMES,
        extra_cookies={"mabang_lite_rowsPerPage": "100"},
        private_extra_cookies={"exportv2": "2"},
        memcache_cookie_name=MABANG_MEMCACHE_COOKIE_NAME,
        error_type=StoreMskuDownloadAuthError,
    )


def _step1_form_data(store_id: str, id_type: str) -> list[tuple[str, str]]:
    clean_id = normalize_store_id(store_id)
    clean_type = normalize_id_type(id_type)
    if clean_type == ID_TYPE_FBA_WAREHOUSE:
        form: list[tuple[str, str]] = [
            (ID_TYPE_FBA_WAREHOUSE, clean_id),
            (ID_TYPE_SHOP, ""),
        ]
    else:
        form = [(ID_TYPE_SHOP, clean_id)]
    form.extend(
        [
            ("status", "1"),
            ("ispair", ""),
            ("isChange", "1"),
            ("amazonsite", ""),
            ("stockStatus", ""),
            ("stockStatusAmz", ""),
            ("developerId", ""),
            ("saleId", ""),
            ("setdataflag", ""),
            ("searchtexttype", "platformSkuLike"),
            ("Orderby", ""),
            ("highsearch", ""),
            ("atn", "list"),
            ("searchtext", ""),
            ("searchtype", "4"),
            ("selecttype", "platformSkuIn"),
            ("platformSkuData", ""),
        ]
    )
    return form


def parse_store_msku_ids(payload: dict[str, Any]) -> list[str]:
    raw_id = _clean_text(payload.get("id"))
    if not raw_id:
        raise StoreMskuDownloadError("店铺MSKU列表查询返回缺少 id")
    ids = [item.strip() for item in raw_id.split(",") if item.strip()]
    if not ids:
        raise StoreMskuDownloadError("店铺MSKU列表查询返回 id 为空")
    return ids


async def fetch_store_msku_ids(
    store_id: str,
    id_type: str,
) -> list[str]:
    auth = await resolve_store_msku_auth()
    headers = _request_headers(
        auth.private_amz_cookie_header,
        origin=_configured_text("MABANG_STORE_MSKU_LISTSEARCH_ORIGIN", DEFAULT_PRIVATE_AMZ_ORIGIN),
        referer=_configured_text("MABANG_STORE_MSKU_LISTSEARCH_REFERER", DEFAULT_PRIVATE_AMZ_REFERER),
    )
    async with erp_http_session.post(
        _listsearch_url(),
        data=_step1_form_data(store_id, id_type),
        headers=headers,
    ) as resp:
        payload = await _read_store_msku_json(resp, action="店铺MSKU列表查询")
    return parse_store_msku_ids(payload)


def _ids_lines(ids: list[str]) -> str:
    return "\r\n".join(ids)


def _step2_form_data(ids: list[str], *, memcache_key: str) -> list[tuple[str, str]]:
    clean_ids = [_clean_text(item) for item in ids if _clean_text(item)]
    if not clean_ids:
        raise ValueError("id 不能为空")

    form: list[tuple[str, str]] = [
        ("backUrl", ""),
        ("orderIds", _ids_lines(clean_ids)),
    ]
    form.extend(("fieldlabel", uq) for uq in STORE_MSKU_FIELDLABELS)
    for name, uq, map_text in STORE_MSKU_EXPORT_FIELD_DEFS:
        form.extend(
            [
                ("map-name[]", name),
                ("map-uq[]", uq),
                ("map-text[]", map_text),
            ]
        )
    form.extend(
        [
            ("templateName", ""),
            ("templateId", "1052958"),
            ("datasOpen", "2"),
            ("memcacheKey", memcache_key),
            ("showRmbColumn", "2"),
            ("pageSave", "1"),
            ("operateType", "5"),
            ("params", ""),
            ("InterfaceUrl", ""),
            ("mainMenu", ""),
            ("hiddenPage", ""),
            ("hiddenPageSize", ""),
            ("tableBase", ""),
            ("isMerage", "2"),
            ("showRmbColumn", "2"),
        ]
    )
    return form


def parse_store_msku_export_gourl(payload: dict[str, Any]) -> str:
    url = _clean_text(payload.get("gourl"))
    if not url:
        raise StoreMskuDownloadError("店铺MSKU导出返回缺少 gourl")
    return url


async def export_store_msku_file_url(
    ids: list[str],
) -> str:
    auth = await resolve_store_msku_auth()
    headers = _request_headers(
        auth.private_cookie_header,
        origin=_configured_text("MABANG_STORE_MSKU_FBA_EXPORT_ORIGIN", DEFAULT_PRIVATE_ORIGIN),
        referer=_configured_text("MABANG_STORE_MSKU_FBA_EXPORT_REFERER", DEFAULT_PRIVATE_REFERER),
    )
    async with erp_http_session.post(
        _fba_export_url(),
        data=_step2_form_data(ids, memcache_key=auth.memcache_key),
        headers=headers,
    ) as resp:
        payload = await _read_store_msku_json(resp, action="店铺MSKU导出")
    return parse_store_msku_export_gourl(payload)


async def download_store_msku_excel_from_url(
    file_url: str,
    *,
    store_id: str,
    store_name: str = "",
    output_dir: str | Path | None = None,
) -> Path:
    url = _clean_text(file_url)
    if not url:
        raise ValueError("file_url 不能为空")

    directory = _resolve_output_dir(output_dir)
    prefix = _safe_file_part(normalize_store_name(store_name))
    target_path = directory / f"{_timestamp_text()}-{prefix}_{STORE_MSKU_FILE_SUFFIX}{_excel_suffix_from_url(url)}"
    headers = {
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream,*/*"
    }
    async with external_http_session.get(url, headers=headers) as resp:
        status_code = int(getattr(resp, "status", 0) or 0)
        body = await resp.read()
        if status_code >= 400:
            msg = body.decode("utf-8", errors="replace")[:300] if body else "empty response"
            raise MabangRequestError(f"下载店铺MSKU数据Excel失败(status={status_code}): {msg}")
        if not body:
            raise StoreMskuDownloadError("下载店铺MSKU数据Excel返回空文件")

    target_path.write_bytes(body)
    return target_path


def convert_store_msku_xls_to_xlsx(excel_path: str | Path) -> Path:
    source_path = Path(excel_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")

    try:
        import xlrd
    except Exception as exc:
        raise RuntimeError("缺少 xlrd 依赖，无法转换店铺MSKU数据xls") from exc
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法写入店铺MSKU数据xlsx") from exc

    target_path = source_path.with_suffix(".xlsx")
    output = None
    try:
        workbook = xlrd.open_workbook(str(source_path), ignore_workbook_corruption=True)
        sheet = workbook.sheet_by_index(0)
        output = Workbook()
        worksheet = output.active
        worksheet.title = str(sheet.name or "Sheet1")[:31] or "Sheet1"
        for row_index in range(sheet.nrows):
            worksheet.append([sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)])
        output.save(target_path)
    except Exception as exc:
        raise RuntimeError(f"转换店铺MSKU数据xls为xlsx失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if output is not None:
                output.close()
        except Exception:
            pass
    return target_path


def delete_raw_store_msku_xls(excel_path: str | Path) -> bool:
    source_path = Path(excel_path).expanduser()
    if source_path.suffix.lower() != ".xls":
        return False
    if not source_path.exists():
        return False
    try:
        source_path.unlink()
    except Exception as exc:
        raise RuntimeError(f"删除店铺MSKU数据xls失败: {source_path}, error={exc}") from exc
    return True


def normalize_store_msku_excel(excel_path: str | Path) -> tuple[Path, bool, bool]:
    source_path = Path(excel_path).expanduser()
    suffix = source_path.suffix.lower()
    if suffix == ".xls":
        xlsx_path = convert_store_msku_xls_to_xlsx(source_path)
        return xlsx_path, True, delete_raw_store_msku_xls(source_path)
    if suffix == ".xlsx":
        return source_path, False, False
    raise StoreMskuDownloadError(f"不支持的店铺MSKU数据文件格式: {source_path.suffix}")


def validate_store_msku_excel_headers(excel_path: str | Path) -> None:
    source_path = Path(excel_path).expanduser()
    if not source_path.is_file():
        raise FileNotFoundError(f"店铺MSKU数据Excel不存在: {source_path}")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取店铺MSKU数据Excel") from exc

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        columns = [_clean_text(cell) for cell in list(first_row or [])]
    except Exception as exc:
        raise RuntimeError(f"读取店铺MSKU数据Excel失败: {source_path}, error={exc}") from exc
    finally:
        try:
            if workbook is not None:
                workbook.close()
        except Exception:
            pass

    missing = [column for column in CORE_STORE_MSKU_HEADERS if column not in columns]
    if missing:
        raise StoreMskuDownloadError(f"店铺MSKU数据Excel缺少列: {', '.join(missing)}")


async def download_store_msku_excel(
    store_id: str,
    id_type: str,
    *,
    store_name: str = "",
    output_dir: str | Path | None = None,
) -> StoreMskuExcelResult:
    clean_store_id = normalize_store_id(store_id)
    clean_id_type = normalize_id_type(id_type)
    clean_store_name = normalize_store_name(store_name)

    def transform_result(ids: list[str], excel_path: Path) -> StoreMskuExcelResult:
        xlsx_path, converted, raw_excel_deleted = normalize_store_msku_excel(excel_path)
        validate_store_msku_excel_headers(xlsx_path)
        return StoreMskuExcelResult(
            store_name=clean_store_name,
            store_id=clean_store_id,
            id_type=clean_id_type,
            id_count=len(ids),
            xlsx_path=str(xlsx_path),
            converted=converted,
            raw_excel_deleted=raw_excel_deleted,
        )

    return await run_export_pipeline(
        ExportPipelineSpec(
            fetch_ids=fetch_store_msku_ids,
            fetch_args=(clean_store_id, clean_id_type),
            request_file_url=export_store_msku_file_url,
            download_file=partial(
                download_store_msku_excel_from_url,
                store_id=clean_store_id,
                store_name=clean_store_name,
                output_dir=output_dir,
            ),
            transform_result=transform_result,
        )
    )


__all__ = [
    "CORE_STORE_MSKU_HEADERS",
    "DEFAULT_FBA_EXPORT_URL",
    "DEFAULT_LISTSEARCH_URL",
    "SOURCE",
    "STORE_MSKU_FILE_SUFFIX",
    "STORE_MSKU_EXPORT_FIELDS",
    "STORE_MSKU_EXPORT_FIELD_DEFS",
    "STORE_MSKU_FIELDLABELS",
    "StoreMskuDownloadAuthError",
    "StoreMskuDownloadError",
    "StoreMskuExcelResult",
    "convert_store_msku_xls_to_xlsx",
    "delete_raw_store_msku_xls",
    "download_store_msku_excel",
    "download_store_msku_excel_from_url",
    "export_store_msku_file_url",
    "fetch_store_msku_ids",
    "normalize_id_type",
    "normalize_store_id",
    "normalize_store_name",
    "normalize_store_msku_excel",
    "parse_store_msku_export_gourl",
    "parse_store_msku_ids",
    "resolve_store_msku_auth",
    "_timestamp_text",
    "validate_store_msku_excel_headers",
]
