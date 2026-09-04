"""海鹰关键词搜索量 Excel 报表的生成与解析查询。

布局与原版保持一致：每个国家一个 Sheet；Sheet 内每个查询关键词占
2 列（搜索词/搜索量），关键词块之间空 1 列；第 1 行关键词标题，
第 2 行表头，第 3 行起为数据。
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384
REPORT_FILE_NAME = "海鹰Shopee关键词搜索量_全量.xlsx"


def safe_sheet_title(value: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", value).strip() or "国家"
    base = base[:31]
    title = base
    counter = 2
    while title.casefold() in used:
        suffix = f"_{counter}"
        title = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(title.casefold())
    return title


def _make_cell(
    sheet: Any,
    value: Any,
    *,
    fill: PatternFill,
    font: Font,
    border: Border,
    alignment: Alignment,
    number_format: str | None = None,
) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell.fill = fill
    cell.font = font
    cell.border = border
    cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    return cell


def build_excel(
    output_path: Path,
    combined: dict[str, dict[str, dict[str, int | float | None]]],
) -> None:
    """combined: { "国家名(code)": { 查询关键词: {搜索词: 搜索量} } }。"""
    workbook = Workbook(write_only=True)
    used_titles: set[str] = set()
    thin_green = Side(style="thin", color="A9D18E")
    border = Border(left=thin_green, right=thin_green, top=thin_green, bottom=thin_green)
    title_fill = PatternFill("solid", fgColor="70AD47")
    header_fill = PatternFill("solid", fgColor="C6E0B4")
    body_fill = PatternFill("solid", fgColor="E2F0D9")
    title_font = Font(bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="203864")
    body_font = Font(color="000000")
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for country_key, query_map in combined.items():
        country_name = country_key.rsplit("(", 1)[0]
        sheet = workbook.create_sheet(safe_sheet_title(country_name, used_titles))
        sheet.freeze_panes = "A3"
        sheet.sheet_view.showGridLines = False
        query_entries = list(query_map.items())
        required_columns = len(query_entries) * 3 - 1
        if required_columns > EXCEL_MAX_COLUMNS:
            raise RuntimeError(f"{country_name} 需要 {required_columns} 列，超过 Excel 上限")

        maximum_rows = max(
            (len(keyword_map) for _, keyword_map in query_entries),
            default=0,
        )
        if maximum_rows + 2 > EXCEL_MAX_ROWS:
            raise RuntimeError(f"{country_name} 需要 {maximum_rows + 2} 行，超过 Excel 上限")

        row1: list[Any] = []
        row2: list[Any] = []
        cached_items: list[list[tuple[str, int | float | None]]] = []

        for index, (query_keyword, keyword_map) in enumerate(query_entries):
            cached_items.append(list(keyword_map.items()))
            row1.extend(
                [
                    _make_cell(sheet, query_keyword, fill=title_fill, font=title_font, border=border, alignment=left),
                    _make_cell(sheet, None, fill=title_fill, font=title_font, border=border, alignment=left),
                ]
            )
            row2.extend(
                [
                    _make_cell(sheet, "搜索词", fill=header_fill, font=header_font, border=border, alignment=center),
                    _make_cell(sheet, "搜索量", fill=header_fill, font=header_font, border=border, alignment=center),
                ]
            )
            start_col = index * 3 + 1
            longest = max(
                [len(query_keyword), *[len(item[0]) for item in cached_items[-1]]],
                default=18,
            )
            sheet.column_dimensions[get_column_letter(start_col)].width = min(38, max(18, longest + 2))
            sheet.column_dimensions[get_column_letter(start_col + 1)].width = 12
            if index < len(query_entries) - 1:
                row1.append(None)
                row2.append(None)
                sheet.column_dimensions[get_column_letter(start_col + 2)].width = 3

        sheet.append(row1)
        sheet.append(row2)

        for data_index in range(maximum_rows):
            row: list[Any] = []
            for query_index, items in enumerate(cached_items):
                if data_index < len(items):
                    keyword, search_volume = items[data_index]
                    row.extend(
                        [
                            _make_cell(sheet, keyword, fill=body_fill, font=body_font, border=border, alignment=left),
                            _make_cell(
                                sheet,
                                search_volume,
                                fill=body_fill,
                                font=body_font,
                                border=border,
                                alignment=right,
                                number_format="#,##0",
                            ),
                        ]
                    )
                else:
                    row.extend([None, None])
                if query_index < len(cached_items) - 1:
                    row.append(None)
            sheet.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.stem}.tmp{output_path.suffix}")
    try:
        workbook.save(temporary_path)
        temporary_path.replace(output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def verify_workbook(output_path: Path, expected_sheets: int) -> None:
    if not output_path.is_file():
        raise RuntimeError(f"Excel 未生成：{output_path}")
    workbook = load_workbook(output_path, read_only=True)
    try:
        actual_sheets = len(workbook.sheetnames)
    finally:
        workbook.close()
    if actual_sheets != expected_sheets:
        raise RuntimeError(
            f"Excel Sheet 数量不正确：预期 {expected_sheets}，实际 {actual_sheets}"
        )


def _parse_sheet(ws: Any) -> dict[str, list[tuple[str, int]]]:
    """把 Sheet 解析为 {关键词标题: [(搜索词, 搜索量), ...]}，按搜索量降序。"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {}
    title_row = rows[0]
    blocks: dict[str, list[tuple[str, int]]] = {}
    col = 0
    while col < len(title_row):
        title = title_row[col]
        if isinstance(title, str) and title.strip():
            blocks[title.strip()] = []
        col += 3
    for data_row in rows[2:]:
        col = 0
        block_index = 0
        titles = list(blocks.keys())
        while col < len(data_row) and block_index < len(titles):
            keyword = titles[block_index]
            word = data_row[col]
            volume = data_row[col + 1] if col + 1 < len(data_row) else None
            if word is not None and str(word).strip():
                try:
                    blocks[keyword].append((str(word).strip(), int(volume) if volume is not None else 0))
                except (TypeError, ValueError):
                    blocks[keyword].append((str(word).strip(), 0))
            col += 3
            block_index += 1
    for key in blocks:
        blocks[key].sort(key=lambda item: item[1], reverse=True)
    return blocks


def list_countries(report_path: Path) -> list[str]:
    workbook = load_workbook(report_path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def query_report(
    report_path: Path,
    country: str | None,
    keyword: str | None,
    top: int,
) -> dict[str, Any]:
    workbook = load_workbook(report_path, read_only=True)
    try:
        target_sheets = [country] if country else list(workbook.sheetnames)
        result: dict[str, Any] = {"file": str(report_path), "countries": {}}
        total_rows = 0
        for sheet_name in target_sheets:
            if sheet_name not in workbook.sheetnames:
                result["countries"][sheet_name] = {"error": "Sheet 不存在"}
                continue
            blocks = _parse_sheet(workbook[sheet_name])
            if keyword:
                norm = keyword.strip().lower()
                matched = {k: v for k, v in blocks.items() if k.strip().lower() == norm}
                if not matched:
                    matched = {k: v for k, v in blocks.items() if norm in k.lower()}
                country_data: dict[str, Any] = {
                    "keywords": [
                        {
                            "keyword": k,
                            "rows": len(v),
                            "top": [{"search_word": w, "volume": vol} for w, vol in v[:top]],
                        }
                        for k, v in matched.items()
                    ]
                }
            else:
                country_data = {
                    "keywords": [
                        {
                            "keyword": k,
                            "rows": len(v),
                            "top": [{"search_word": w, "volume": vol} for w, vol in v[:top]],
                        }
                        for k, v in blocks.items()
                    ]
                }
                total_rows += sum(len(v) for v in blocks.values())
            result["countries"][sheet_name] = country_data
        result["total_rows"] = total_rows
        return result
    finally:
        workbook.close()
