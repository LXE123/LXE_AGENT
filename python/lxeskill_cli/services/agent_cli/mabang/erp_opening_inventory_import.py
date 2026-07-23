from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_cli import exception_text
from services.agent_cli.mabang.erp_http import (
    ErpHttpError,
    error_payload,
    request_json,
)


SOURCE = "fba_erp_opening_inventory_import"
EXPECTED_HEADERS = (
    "供应商",
    "采购订单号",
    "订单号",
    "型号",
    "含税单价",
    "数量",
)
CONTRACT_DATE_PATTERN = re.compile(r"(\d{8})")
SP_PATTERN = re.compile(r"(?<![A-Z0-9])(SP\d+)(?!\d)", re.IGNORECASE)


class OpeningInventoryError(RuntimeError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal_text(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    return format(value.normalize(), "f").rstrip("0").rstrip(".")


def _decimal_cell(
    value: Any,
    *,
    field_name: str,
    row_no: int,
    positive: bool,
) -> Decimal:
    text = _clean_cell(value)
    if not text:
        raise OpeningInventoryError(
            "opening_inventory_workbook_invalid",
            f"期初库存表第{row_no}行 {field_name} 不能为空",
        )
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise OpeningInventoryError(
            "opening_inventory_workbook_invalid",
            f"期初库存表第{row_no}行 {field_name} 非数字: {text}",
        ) from exc
    if not result.is_finite() or result < 0 or (positive and result <= 0):
        rule = "必须大于 0" if positive else "不能小于 0"
        raise OpeningInventoryError(
            "opening_inventory_workbook_invalid",
            f"期初库存表第{row_no}行 {field_name} {rule}: {text}",
        )
    return result


def _acquired_on(contract_no: str) -> str | None:
    for match in CONTRACT_DATE_PATTERN.finditer(contract_no):
        try:
            return date.fromisoformat(
                f"{match.group(1)[:4]}-{match.group(1)[4:6]}-{match.group(1)[6:]}"
            ).isoformat()
        except ValueError:
            continue
    return None


def _source_sp_no(source_reference: str) -> str:
    matches = list(dict.fromkeys(match.group(1).upper() for match in SP_PATTERN.finditer(source_reference)))
    return matches[0] if len(matches) == 1 else ""


def _read_workbook(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取期初库存 xlsx") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = tuple(next(rows))
        except StopIteration:
            raw_headers = ()
        headers = tuple(_clean_cell(value) for value in raw_headers)
        while headers and not headers[-1]:
            headers = headers[:-1]
        if headers != EXPECTED_HEADERS:
            raise OpeningInventoryError(
                "opening_inventory_workbook_invalid",
                "期初库存表第1行表头必须严格为: " + ", ".join(EXPECTED_HEADERS),
            )

        parsed: list[dict[str, Any]] = []
        for row_no, raw_row in enumerate(rows, start=2):
            values = tuple(raw_row[: len(EXPECTED_HEADERS)])
            if not any(_clean_cell(value) for value in values):
                continue
            padded = values + (None,) * (len(EXPECTED_HEADERS) - len(values))
            supplier, contract_no, source_reference, model = (
                _clean_cell(value) for value in padded[:4]
            )
            for field_name, value in (
                ("供应商", supplier),
                ("采购订单号", contract_no),
                ("订单号", source_reference),
                ("型号", model),
            ):
                if not value:
                    raise OpeningInventoryError(
                        "opening_inventory_workbook_invalid",
                        f"期初库存表第{row_no}行 {field_name} 不能为空",
                    )
            unit_price = _decimal_cell(
                padded[4], field_name="含税单价", row_no=row_no, positive=False
            )
            quantity = _decimal_cell(
                padded[5], field_name="数量", row_no=row_no, positive=True
            )
            parsed.append(
                {
                    "row_no": row_no,
                    "supplier_name": supplier,
                    "contract_no": contract_no,
                    "source_reference": source_reference,
                    "source_sp_no": _source_sp_no(source_reference),
                    "model": model,
                    "historical_tax_unit_price": _decimal_text(unit_price),
                    "remaining_quantity": _decimal_text(quantity),
                    "acquired_on": _acquired_on(contract_no),
                }
            )
        if not parsed:
            raise OpeningInventoryError(
                "opening_inventory_workbook_invalid",
                "期初库存表没有可导入的数据行",
            )
        return parsed
    finally:
        workbook.close()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _request_id(payload: dict[str, Any]) -> str:
    canonical = json.dumps(
        {
            key: value
            for key, value in payload.items()
            if key not in {"request_id", "confirm_source_sha256"}
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return f"opening-{hashlib.sha256(canonical).hexdigest()[:32]}"


def _build_request(path: Path, *, confirm_source_sha256: str = "") -> dict[str, Any]:
    source_sha256 = _file_sha256(path)
    payload: dict[str, Any] = {
        "source_file_name": path.name,
        "source_sha256": source_sha256,
        "captured_at": datetime.fromtimestamp(
            path.stat().st_mtime, tz=timezone.utc
        ).isoformat(),
        "lines": _read_workbook(path),
    }
    if confirm_source_sha256:
        payload["confirm_source_sha256"] = confirm_source_sha256.lower()
    payload["request_id"] = _request_id(payload)
    return payload


def run(arguments: dict[str, Any]) -> dict[str, Any]:
    input_xlsx = str(arguments.get("input_xlsx") or "").strip()
    try:
        if not input_xlsx:
            raise OpeningInventoryError(
                "opening_inventory_input_missing", "input_xlsx 不能为空"
            )
        path = Path(input_xlsx).expanduser()
        if path.suffix.lower() != ".xlsx":
            raise OpeningInventoryError(
                "opening_inventory_input_invalid", "期初库存文件必须是 .xlsx"
            )
        if not path.is_file():
            raise OpeningInventoryError(
                "opening_inventory_input_missing", f"找不到期初库存表: {path}"
            )
        path = path.resolve()
        payload = _build_request(
            path,
            confirm_source_sha256=str(
                arguments.get("confirm_source_sha256") or ""
            ).strip(),
        )
        status_code, response = request_json(
            "POST",
            "/api/v1/erp/opening-inventory/import",
            operation="导入期初库存",
            json_payload=payload,
            accepted_error_codes=frozenset(
                {"opening_inventory_confirmation_required"}
            ),
        )
        if status_code == 409:
            detail = dict(response.get("detail") or {})
            return {
                "success": False,
                "status": "confirmation_required",
                "input_xlsx": str(path),
                "source_sha256": payload["source_sha256"],
                "request_id": payload["request_id"],
                "error": {
                    "code": str(detail.get("code") or "opening_inventory_confirmation_required"),
                    "message": str(detail.get("message") or "ERP 要求确认期初库存导入"),
                    "http_status": status_code,
                    "detail": detail,
                },
                "preview": detail,
                "source": SOURCE,
            }
        return {
            "success": True,
            "input_xlsx": str(path),
            "source_sha256": payload["source_sha256"],
            "request_id": payload["request_id"],
            "row_count": len(payload["lines"]),
            **response,
            "source": SOURCE,
        }
    except OpeningInventoryError as exc:
        return {
            "success": False,
            "input_xlsx": input_xlsx,
            "exception": str(exc),
            "error": {"code": exc.code, "message": str(exc)},
            "source": SOURCE,
        }
    except ErpHttpError as exc:
        return {
            "success": False,
            "input_xlsx": input_xlsx,
            "exception": str(exc),
            "error": error_payload(exc),
            "source": SOURCE,
        }
    except Exception as exc:  # noqa: BLE001 - preserve real workbook failures
        message = exception_text(exc)
        return {
            "success": False,
            "input_xlsx": input_xlsx,
            "exception": message,
            "error": {"code": "opening_inventory_import_failed", "message": message},
            "source": SOURCE,
        }


__all__ = ["EXPECTED_HEADERS", "_build_request", "_read_workbook", "run"]
