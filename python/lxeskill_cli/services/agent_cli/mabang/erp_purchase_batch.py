from __future__ import annotations

import hashlib
import json
import os
import re
from io import BytesIO
from collections import OrderedDict, defaultdict
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Mapping
from urllib.parse import unquote

from services.agent_cli.mabang import generate_fba_restock_workbook as restock_workbook
from services.agent_cli.mabang import generate_restock_workbook as purchase_summary
from services.agent_cli.mabang.erp_http import ErpHttpError, error_payload, request_json
from services.agent_cli.mabang.shipment_quantity_validation import read_delivery_msku_infos
from shared.datasets import dataset_dir


PURCHASE_CONFIRMATION_CODES = frozenset(
    {
        "purchase_inventory_confirmation_required",
        "purchase_inventory_quote_stale",
        "purchase_batch_replace_confirmation_required",
    }
)
PURCHASE_CONFIRMATION_RESPONSE_SCHEMA = "lxe.erp.purchase-confirmation.v2"
UNMATCHED_CONFIRMATION_RESPONSE_SCHEMA = (
    "lxe.fba.purchase-unmatched-confirmation.v1"
)
UNMATCHED_ISSUE_CODE = "export_tax_master_stock_sku_not_found"
FORMAL_QUANTITY_COLUMNS = ("计划发货量", "本次采购量", "留存库存抵扣量")
INVENTORY_ROW_FILL_COLOR = "FFFFFF00"
CONTRACT_OUTPUT_DIR = dataset_dir("fba_purchase_contracts")
_INVALID_FILE_NAME_CHARS = re.compile(r'[\\/:*?"<>|\x00-\x1f]')


class PurchaseBatchClientError(RuntimeError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        detail: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.detail = dict(detail or {})


def _decimal_text(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    return format(value.normalize(), "f").rstrip("0").rstrip(".")


def _decimal(value: Any) -> Decimal:
    if isinstance(value, bool) or value in (None, ""):
        raise InvalidOperation("missing decimal")
    result = Decimal(str(value))
    if not result.is_finite():
        raise InvalidOperation("non-finite decimal")
    return result


def _file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_sha256(value: Mapping[str, Any]) -> str:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _require_text(value: Any, *, label: str) -> str:
    text = purchase_summary._clean_cell(value)
    if not text:
        raise PurchaseBatchClientError(
            "purchase_intent_invalid", f"正式 ERP 采购数据缺少{label}"
        )
    return text


def _unit_components(
    csv_path: str | Path,
    *,
    tracked_skus: set[str],
) -> list[dict[str, Any]]:
    infos = read_delivery_msku_infos(csv_path)
    mskus: list[dict[str, Any]] = []
    for raw_msku, info in infos.items():
        msku = _require_text(raw_msku, label="MSKU")
        denominator = info.msku_ship_quantity
        if denominator is None or denominator <= 0 or denominator != denominator.to_integral_value():
            raise PurchaseBatchClientError(
                "purchase_msku_unit_quantity_invalid",
                f"发货单 {Path(csv_path).name} 的 MSKU={msku} 缺少正整数 MSKU发货量",
            )
        components: list[dict[str, Any]] = []
        for stock_sku, source_quantity in info.components.items():
            if source_quantity <= 0:
                raise PurchaseBatchClientError(
                    "purchase_msku_unit_quantity_invalid",
                    f"发货单 {Path(csv_path).name} 的 MSKU={msku} 包含非正数 SKU 数量",
                )
            quantity_per_msku = source_quantity / denominator
            if quantity_per_msku != quantity_per_msku.to_integral_value():
                raise PurchaseBatchClientError(
                    "purchase_msku_unit_quantity_invalid",
                    f"发货单 {Path(csv_path).name} 的 MSKU={msku}, SKU={stock_sku} "
                    f"无法推导整数 quantity_per_msku: {source_quantity}/{denominator}",
                )
            normalized_sku = purchase_summary._clean_cell(stock_sku).upper()
            components.append(
                {
                    "stock_sku": normalized_sku,
                    "tracking_mode": (
                        "tracked" if normalized_sku in tracked_skus else "unmatched"
                    ),
                    "quantity_per_msku": _decimal_text(quantity_per_msku),
                }
            )
        if not components:
            raise PurchaseBatchClientError(
                "purchase_msku_unit_quantity_invalid",
                f"发货单 {Path(csv_path).name} 的 MSKU={msku} 没有库存 SKU 组成",
            )
        mskus.append({"msku": msku.upper(), "components": components})
    return mskus


def build_purchase_intent(
    delivery_nos: list[str],
    *,
    master_xlsx: str | Path,
    csv_dir: str | Path | None = None,
    confirm_inventory_quote_id: str = "",
    replace_batch_id: str = "",
    expected_version_no: int | None = None,
    change_reason: str = "",
) -> tuple[dict[str, Any], dict[str, Any]]:
    _batch_summary, _batch_sources, normalized_delivery_nos, csv_paths = (
        purchase_summary.summarize_delivery_quantities(delivery_nos, csv_dir=csv_dir)
    )
    products = purchase_summary.load_master_products(master_xlsx)

    supplier_configs: OrderedDict[str, dict[str, Any]] = OrderedDict()
    contract_lines: OrderedDict[tuple[str, str], dict[str, Any]] = OrderedDict()
    sps: list[dict[str, Any]] = []
    csv_hashes: list[dict[str, str]] = []
    unmatched_items: list[dict[str, Any]] = []
    unmatched_component_count = 0
    tracked_line_count = 0
    for sp_no, csv_path_text in zip(normalized_delivery_nos, csv_paths, strict=True):
        csv_path = Path(csv_path_text)
        delivery_product_names = purchase_summary.summarize_delivery_product_names(
            [csv_path]
        )
        delivery_summary = OrderedDict(
            (sku, quantity)
            for sku, quantity in purchase_summary.summarize_tax_sku_quantities_in_delivery_order(csv_path).items()
            if quantity > 0
        )
        country, _country_warnings = restock_workbook._delivery_country_metadata(csv_path)
        planned_lines: list[dict[str, Any]] = []
        unmatched_quantities: OrderedDict[str, Decimal] = OrderedDict()
        for stock_sku, quantity in delivery_summary.items():
            normalized_sku = purchase_summary._clean_cell(stock_sku).upper()
            raw_product = products.get(purchase_summary._sku_match_key(stock_sku))
            if raw_product is None:
                unmatched_quantities[normalized_sku] = quantity
                continue
            product = dict(raw_product)
            manufacturer = _require_text(product.get("manufacturer"), label=f"厂家: SKU={stock_sku}")
            model = _require_text(product.get("model"), label=f"型号: SKU={stock_sku}")
            product_name = _require_text(
                product.get("product_name"), label=f"产品名称: SKU={stock_sku}"
            )
            unit, contract_product_name, contract_prefix, tax_rate = (
                purchase_summary._contract_fields_for_manufacturer(
                    products, manufacturer, collect_warnings=False
                )
            )
            unit = _require_text(unit, label=f"单位: 厂家={manufacturer}")
            contract_product_name = _require_text(
                contract_product_name, label=f"合同产品名称: 厂家={manufacturer}"
            )
            contract_prefix = _require_text(
                contract_prefix, label=f"合同编号前缀: 厂家={manufacturer}"
            ).upper()
            tax_rate = _require_text(tax_rate, label=f"税率: 厂家={manufacturer}")
            try:
                restock_workbook._tax_multiplier_from_rate(
                    tax_rate,
                    manufacturer=manufacturer,
                    model=model,
                )
            except RuntimeError as exc:
                raise PurchaseBatchClientError(
                    "purchase_intent_invalid", str(exc)
                ) from exc
            if not contract_prefix.isalnum() or len(contract_prefix) > 16:
                raise PurchaseBatchClientError(
                    "purchase_intent_invalid",
                    f"厂家={manufacturer} 的合同编号前缀必须是 1～16 位字母或数字",
                )
            existing_supplier = supplier_configs.get(manufacturer)
            supplier_config = {
                "name": manufacturer,
                "contract_prefix": contract_prefix,
                "tax_rate": tax_rate,
            }
            if existing_supplier is not None and existing_supplier != supplier_config:
                raise PurchaseBatchClientError(
                    "purchase_intent_invalid", f"厂家={manufacturer} 的合同配置不一致"
                )
            supplier_configs[manufacturer] = supplier_config
            planned_lines.append(
                {
                    "stock_sku": normalized_sku,
                    "product_name": product_name,
                    "model": model,
                    "supplier_name": manufacturer,
                    "planned_shipment_quantity": _decimal_text(quantity),
                }
            )
            contract_key = (manufacturer, model.casefold())
            contract_line = contract_lines.get(contract_key)
            try:
                source_price = _decimal(product["original_price"])
            except (InvalidOperation, ValueError) as exc:
                raise PurchaseBatchClientError(
                    "purchase_intent_invalid",
                    f"正式 ERP 采购数据原价无法解析: SKU={stock_sku}, value={product.get('original_price')}",
                ) from exc
            if source_price < 0:
                raise PurchaseBatchClientError(
                    "purchase_intent_invalid",
                    f"正式 ERP 采购数据原价不能为负数: SKU={stock_sku}, value={source_price}",
                )
            if contract_line is None:
                contract_line = {
                    "supplier_name": manufacturer,
                    "tax_rate": tax_rate,
                    "line_ref": f"L{len(contract_lines) + 1:04d}",
                    "contract_product_name": contract_product_name,
                    "model": model,
                    "unit": unit,
                    "source_tax_unit_price": _decimal_text(source_price),
                    "planned_shipment_quantity": Decimal("0"),
                    "allocations": [],
                }
                contract_lines[contract_key] = contract_line
            elif (
                contract_line["source_tax_unit_price"] != _decimal_text(source_price)
                or contract_line["unit"] != unit
                or contract_line["contract_product_name"] != contract_product_name
            ):
                raise PurchaseBatchClientError(
                    "purchase_intent_invalid",
                    f"同一厂家型号的价格或合同字段不一致: 厂家={manufacturer}, 型号={model}",
                )
            contract_line["planned_shipment_quantity"] += quantity
            contract_line["allocations"].append(
                {
                    "sp_no": sp_no,
                    "stock_sku": normalized_sku,
                    "planned_quantity": _decimal_text(quantity),
                }
            )

        tracked_skus = {line["stock_sku"] for line in planned_lines}
        mskus = _unit_components(csv_path, tracked_skus=tracked_skus)
        affected_mskus: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
        for msku in mskus:
            for component in msku["components"]:
                if component["tracking_mode"] != "unmatched":
                    continue
                unmatched_component_count += 1
                affected_mskus[component["stock_sku"]].append(
                    {
                        "msku": msku["msku"],
                        "quantity_per_msku": component["quantity_per_msku"],
                    }
                )
        unmatched_lines: list[dict[str, str]] = []
        for stock_sku, quantity in unmatched_quantities.items():
            product_name = "\n".join(
                delivery_product_names.get(
                    purchase_summary._sku_match_key(stock_sku), []
                )
            )
            quantity_text = _decimal_text(quantity)
            unmatched_lines.append(
                {
                    "stock_sku": stock_sku,
                    "product_name": product_name,
                    "planned_shipment_quantity": quantity_text,
                    "issue_code": UNMATCHED_ISSUE_CODE,
                }
            )
            unmatched_items.append(
                {
                    "sp_no": sp_no,
                    "stock_sku": stock_sku,
                    "product_name": product_name,
                    "planned_shipment_quantity": quantity_text,
                    "affected_mskus": affected_mskus.get(stock_sku, []),
                }
            )
        tracked_line_count += len(planned_lines)
        planned_by_sku = {
            line["stock_sku"]: Decimal(line["planned_shipment_quantity"])
            for line in planned_lines
        }
        mapped_by_sku: defaultdict[str, Decimal] = defaultdict(lambda: Decimal("0"))
        infos = read_delivery_msku_infos(csv_path)
        for info in infos.values():
            for stock_sku, quantity in info.components.items():
                normalized_sku = purchase_summary._clean_cell(stock_sku).upper()
                if normalized_sku in tracked_skus:
                    mapped_by_sku[normalized_sku] += quantity
        if dict(mapped_by_sku) != planned_by_sku:
            raise PurchaseBatchClientError(
                "purchase_msku_mapping_mismatch",
                f"发货单 {csv_path.name} 的 MSKU 组成数量与库存 SKU 计划量不一致: "
                f"planned={dict(planned_by_sku)}, mapped={dict(mapped_by_sku)}",
            )
        delivery_sha256 = _file_sha256(csv_path)
        csv_hashes.append({"sp_no": sp_no, "sha256": delivery_sha256})
        sps.append(
            {
                "sp_no": sp_no,
                "country": country,
                "delivery_file_name": csv_path.name,
                "delivery_sha256": delivery_sha256,
                "planned_lines": planned_lines,
                "unmatched_lines": unmatched_lines,
                "mskus": mskus,
            }
        )

    unmatched_summary = {
        "stock_sku_count": len({item["stock_sku"] for item in unmatched_items}),
        "sp_sku_count": len(unmatched_items),
        "component_count": unmatched_component_count,
        "planned_shipment_quantity": _decimal_text(
            sum(
                (
                    Decimal(item["planned_shipment_quantity"])
                    for item in unmatched_items
                ),
                Decimal("0"),
            )
        ),
        "items": unmatched_items,
    }
    if tracked_line_count == 0:
        raise PurchaseBatchClientError(
            "purchase_intent_no_tracked_stock_sku",
            "整批发货单没有任何可由出口退税总表跟踪的库存 SKU，未创建 ERP 采购批次",
            detail={"unmatched_summary": unmatched_summary},
        )

    contracts_by_supplier: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for line in contract_lines.values():
        serialized = dict(line)
        serialized.pop("supplier_name", None)
        serialized.pop("tax_rate", None)
        serialized["planned_shipment_quantity"] = _decimal_text(
            serialized["planned_shipment_quantity"]
        )
        contracts_by_supplier.setdefault(line["supplier_name"], []).append(serialized)
    contracts = [
        {
            "supplier_name": supplier,
            "tax_rate": supplier_configs[supplier]["tax_rate"],
            "lines": lines,
        }
        for supplier, lines in contracts_by_supplier.items()
    ]
    suppliers = [
        {"name": config["name"], "contract_prefix": config["contract_prefix"]}
        for config in supplier_configs.values()
    ]
    prefix_owners: dict[str, str] = {}
    for supplier in suppliers:
        previous = prefix_owners.setdefault(supplier["contract_prefix"], supplier["name"])
        if previous != supplier["name"]:
            raise PurchaseBatchClientError(
                "purchase_intent_invalid",
                f"供应商合同编号前缀重复: {supplier['contract_prefix']} ({previous}, {supplier['name']})",
            )

    intent: dict[str, Any] = {
        "created_by": "lxeskill:fba-purchase-summary-create",
        "suppliers": suppliers,
        "sps": sps,
        "contracts": contracts,
    }
    source_fingerprint = {
        "master_xlsx_sha256": _file_sha256(Path(master_xlsx).expanduser()),
        "delivery_csvs": csv_hashes,
        "intent": intent,
    }
    intent["source_sha256"] = _canonical_sha256(source_fingerprint)
    if replace_batch_id:
        intent["replace_batch_id"] = replace_batch_id
        intent["expected_version_no"] = expected_version_no
        intent["change_reason"] = change_reason
    if confirm_inventory_quote_id:
        intent["confirm_inventory_quote_id"] = confirm_inventory_quote_id
    request_basis = {
        key: value for key, value in intent.items() if key != "confirm_inventory_quote_id"
    }
    intent["request_id"] = f"purchase-{_canonical_sha256(request_basis)[:32]}"
    confirmation_token = ""
    if unmatched_items:
        token_basis = {
            "schema": UNMATCHED_CONFIRMATION_RESPONSE_SCHEMA,
            "source_sha256": intent["source_sha256"],
            "items": unmatched_items,
        }
        confirmation_token = f"unmatched-{_canonical_sha256(token_basis)}"
    context = {
        "delivery_nos": normalized_delivery_nos,
        "csv_paths": csv_paths,
        "master_xlsx": str(Path(master_xlsx).expanduser()),
        "product_warnings": list(products.warnings),
        "unmatched_summary": unmatched_summary,
        "confirm_unmatched_sku_token": confirmation_token,
    }
    return intent, context


def unmatched_confirmation_result(
    context: Mapping[str, Any],
    *,
    stale: bool,
) -> dict[str, Any]:
    summary = dict(context.get("unmatched_summary") or {})
    token = str(context.get("confirm_unmatched_sku_token") or "")
    code = (
        "purchase_unmatched_sku_confirmation_stale"
        if stale
        else "purchase_unmatched_sku_confirmation_required"
    )
    message = (
        "发货单或出口退税总表已变化，请按最新未匹配 SKU 清单重新确认"
        if stale
        else "正式采购包含出口退税总表未匹配的库存 SKU，确认后这些组件将不参与采购、库存和装箱对账"
    )
    return {
        "success": False,
        "status": "confirmation_required",
        "response_schema": UNMATCHED_CONFIRMATION_RESPONSE_SCHEMA,
        "error": {"code": code, "message": message},
        "confirmation": {
            "kind": "unmatched_sku_exclusion",
            "token": token,
            **summary,
        },
    }


def import_purchase_intent(payload: Mapping[str, Any]) -> tuple[int, dict[str, Any]]:
    return request_json(
        "POST",
        "/api/v1/erp/purchase-batches/import",
        operation="创建正式采购批次",
        json_payload=payload,
        accepted_error_codes=PURCHASE_CONFIRMATION_CODES,
    )


def _incomplete(message: str, *, field: str = "") -> PurchaseBatchClientError:
    detail = {"field": field} if field else None
    return PurchaseBatchClientError(
        "erp_purchase_result_incomplete",
        f"ERP 采购响应不完整或数量不守恒: {message}",
        detail=detail,
    )


def _required_result_text(
    value: Any,
    *,
    field: str,
) -> str:
    text = purchase_summary._clean_cell(value)
    if not text:
        raise _incomplete(f"缺少 {field}", field=field)
    return text


def _required_result_decimal(
    value: Any,
    *,
    field: str,
    positive: bool = False,
) -> Decimal:
    try:
        number = _decimal(value)
    except (InvalidOperation, ValueError) as exc:
        raise _incomplete(f"{field} 不是有限数字", field=field) from exc
    if number < 0 or (positive and number <= 0):
        qualifier = "正数" if positive else "非负数"
        raise _incomplete(f"{field} 必须是{qualifier}", field=field)
    return number


def _required_result_list(value: Any, *, field: str) -> list[Any]:
    if not isinstance(value, list):
        raise _incomplete(f"{field} 必须是数组", field=field)
    return value


def _expected_purchase_lines(
    request_payload: Mapping[str, Any],
) -> OrderedDict[str, dict[str, Any]]:
    expected: OrderedDict[str, dict[str, Any]] = OrderedDict()
    contracts = _required_result_list(request_payload.get("contracts"), field="request.contracts")
    for contract_index, raw_contract in enumerate(contracts):
        field = f"request.contracts[{contract_index}]"
        if not isinstance(raw_contract, Mapping):
            raise _incomplete(f"{field} 必须是对象", field=field)
        supplier_name = _required_result_text(
            raw_contract.get("supplier_name"), field=f"{field}.supplier_name"
        )
        lines = _required_result_list(raw_contract.get("lines"), field=f"{field}.lines")
        for line_index, raw_line in enumerate(lines):
            line_field = f"{field}.lines[{line_index}]"
            if not isinstance(raw_line, Mapping):
                raise _incomplete(f"{line_field} 必须是对象", field=line_field)
            line_ref = _required_result_text(raw_line.get("line_ref"), field=f"{line_field}.line_ref")
            if line_ref in expected:
                raise _incomplete(f"请求中 line_ref 重复: {line_ref}", field=f"{line_field}.line_ref")
            planned = _required_result_decimal(
                raw_line.get("planned_shipment_quantity"),
                field=f"{line_field}.planned_shipment_quantity",
                positive=True,
            )
            allocations = _required_result_list(
                raw_line.get("allocations"), field=f"{line_field}.allocations"
            )
            expected_allocations: defaultdict[tuple[str, str], Decimal] = defaultdict(
                lambda: Decimal("0")
            )
            for allocation_index, raw_allocation in enumerate(allocations):
                allocation_field = f"{line_field}.allocations[{allocation_index}]"
                if not isinstance(raw_allocation, Mapping):
                    raise _incomplete(
                        f"{allocation_field} 必须是对象", field=allocation_field
                    )
                sp_no = _required_result_text(
                    raw_allocation.get("sp_no"), field=f"{allocation_field}.sp_no"
                ).upper()
                stock_sku = _required_result_text(
                    raw_allocation.get("stock_sku"),
                    field=f"{allocation_field}.stock_sku",
                ).upper()
                quantity = _required_result_decimal(
                    raw_allocation.get("planned_quantity"),
                    field=f"{allocation_field}.planned_quantity",
                    positive=True,
                )
                expected_allocations[(sp_no, stock_sku)] += quantity
            if sum(expected_allocations.values(), Decimal("0")) != planned:
                raise _incomplete(
                    f"请求行 {line_ref} 的 SP/SKU 分配合计不等于计划量",
                    field=f"{line_field}.allocations",
                )
            expected[line_ref] = {
                "supplier_name": supplier_name,
                "model": _required_result_text(raw_line.get("model"), field=f"{line_field}.model"),
                "planned": planned,
                "allocations": dict(expected_allocations),
            }
    return expected


def _validate_allocation_details(
    raw_line: Mapping[str, Any],
    *,
    line_ref: str,
    expected: Mapping[str, Any],
    field_prefix: str,
) -> tuple[Decimal, Decimal]:
    details = _required_result_list(
        raw_line.get("allocation_details"), field=f"{field_prefix}.allocation_details"
    )
    actual_allocations: defaultdict[tuple[str, str], Decimal] = defaultdict(
        lambda: Decimal("0")
    )
    carryover_total = Decimal("0")
    purchase_total = Decimal("0")
    for detail_index, raw_detail in enumerate(details):
        detail_field = f"{field_prefix}.allocation_details[{detail_index}]"
        if not isinstance(raw_detail, Mapping):
            raise _incomplete(f"{detail_field} 必须是对象", field=detail_field)
        sp_no = _required_result_text(
            raw_detail.get("sp_no"), field=f"{detail_field}.sp_no"
        ).upper()
        stock_sku = _required_result_text(
            raw_detail.get("stock_sku"), field=f"{detail_field}.stock_sku"
        ).upper()
        source_kind = _required_result_text(
            raw_detail.get("source_kind"), field=f"{detail_field}.source_kind"
        )
        if source_kind not in {"carryover", "current_purchase"}:
            raise _incomplete(
                f"{detail_field}.source_kind 未知: {source_kind}",
                field=f"{detail_field}.source_kind",
            )
        quantity = _required_result_decimal(
            raw_detail.get("quantity"), field=f"{detail_field}.quantity", positive=True
        )
        actual_allocations[(sp_no, stock_sku)] += quantity
        if source_kind == "carryover":
            _required_result_text(
                raw_detail.get("carryover_entry_id"),
                field=f"{detail_field}.carryover_entry_id",
            )
            _required_result_text(
                raw_detail.get("source_contract_no"),
                field=f"{detail_field}.source_contract_no",
            )
            _required_result_decimal(
                raw_detail.get("historical_tax_unit_price"),
                field=f"{detail_field}.historical_tax_unit_price",
            )
            carryover_total += quantity
        else:
            purchase_total += quantity
    if dict(actual_allocations) != expected["allocations"]:
        raise _incomplete(
            f"行 {line_ref} 的 SP/SKU 来源分配不守恒: "
            f"expected={expected['allocations']}, actual={dict(actual_allocations)}",
            field=f"{field_prefix}.allocation_details",
        )
    return carryover_total, purchase_total


def _validate_success_response(
    response: Mapping[str, Any],
    *,
    request_payload: Mapping[str, Any],
) -> None:
    for field in ("status", "batch_id", "batch_no", "revision_id"):
        _required_result_text(response.get(field), field=field)
    version_no = response.get("version_no")
    if isinstance(version_no, bool):
        raise _incomplete("version_no 必须是正整数", field="version_no")
    try:
        parsed_version = int(version_no)
    except (TypeError, ValueError) as exc:
        raise _incomplete("version_no 必须是正整数", field="version_no") from exc
    if parsed_version <= 0:
        raise _incomplete("version_no 必须是正整数", field="version_no")
    expected_sp_nos = [
        _required_result_text(item.get("sp_no"), field="request.sps[].sp_no")
        for item in _required_result_list(request_payload.get("sps"), field="request.sps")
        if isinstance(item, Mapping)
    ]
    raw_sp_nos = _required_result_list(response.get("sp_nos"), field="sp_nos")
    response_sp_nos = [_required_result_text(value, field="sp_nos[]") for value in raw_sp_nos]
    if response_sp_nos != expected_sp_nos:
        raise _incomplete("sp_nos 与请求顺序或内容不一致", field="sp_nos")

    unmatched_components = [
        component
        for sp in _required_result_list(request_payload.get("sps"), field="request.sps")
        if isinstance(sp, Mapping)
        for msku in list(sp.get("mskus") or [])
        if isinstance(msku, Mapping)
        for component in list(msku.get("components") or [])
        if isinstance(component, Mapping)
        and component.get("tracking_mode") == "unmatched"
    ]
    expected_unmatched_skus = {
        _required_result_text(
            component.get("stock_sku"),
            field="request.sps[].mskus[].components[].stock_sku",
        ).upper()
        for component in unmatched_components
    }
    unmatched_stock_sku_count = _required_nonnegative_int(
        response.get("unmatched_stock_sku_count"),
        field="unmatched_stock_sku_count",
    )
    unmatched_component_count = _required_nonnegative_int(
        response.get("unmatched_component_count"),
        field="unmatched_component_count",
    )
    if (
        unmatched_stock_sku_count != len(expected_unmatched_skus)
        or unmatched_component_count != len(unmatched_components)
    ):
        raise _incomplete(
            "未匹配 SKU/组件计数与请求不一致",
            field="unmatched_component_count",
        )

    expected_lines = _expected_purchase_lines(request_payload)
    raw_lines = _required_result_list(response.get("purchase_lines"), field="purchase_lines")
    actual_refs: set[str] = set()
    positive_suppliers: set[str] = set()
    for line_index, raw_line in enumerate(raw_lines):
        line_field = f"purchase_lines[{line_index}]"
        if not isinstance(raw_line, Mapping):
            raise _incomplete(f"{line_field} 必须是对象", field=line_field)
        line_ref = _required_result_text(raw_line.get("line_ref"), field=f"{line_field}.line_ref")
        if line_ref in actual_refs or line_ref not in expected_lines:
            raise _incomplete(f"成功响应 line_ref 无法匹配或重复: {line_ref}", field=f"{line_field}.line_ref")
        actual_refs.add(line_ref)
        expected = expected_lines[line_ref]
        supplier_name = _required_result_text(
            raw_line.get("supplier_name"), field=f"{line_field}.supplier_name"
        )
        model = _required_result_text(raw_line.get("model"), field=f"{line_field}.model")
        for name in ("tax_rate", "contract_product_name", "unit"):
            _required_result_text(raw_line.get(name), field=f"{line_field}.{name}")
        _required_result_decimal(
            raw_line.get("source_tax_unit_price"),
            field=f"{line_field}.source_tax_unit_price",
        )
        if supplier_name != expected["supplier_name"] or model.casefold() != expected["model"].casefold():
            raise _incomplete(f"成功响应行 {line_ref} 的供应商或型号与请求不一致", field=line_field)
        planned = _required_result_decimal(
            raw_line.get("planned_shipment_quantity"),
            field=f"{line_field}.planned_shipment_quantity",
            positive=True,
        )
        carryover = _required_result_decimal(
            raw_line.get("carryover_applied_quantity"),
            field=f"{line_field}.carryover_applied_quantity",
        )
        purchased = _required_result_decimal(
            raw_line.get("purchase_quantity"), field=f"{line_field}.purchase_quantity"
        )
        if planned != expected["planned"] or planned != carryover + purchased:
            raise _incomplete(f"成功响应行 {line_ref} 不满足 计划量=采购量+库存抵扣量", field=line_field)
        detail_carryover, detail_purchase = _validate_allocation_details(
            raw_line,
            line_ref=line_ref,
            expected=expected,
            field_prefix=line_field,
        )
        if detail_carryover != carryover or detail_purchase != purchased:
            raise _incomplete(f"成功响应行 {line_ref} 的来源分配与采购/抵扣量不一致", field=line_field)
        inventory_sources = _required_result_list(
            raw_line.get("inventory_sources"), field=f"{line_field}.inventory_sources"
        )
        suggested_total = Decimal("0")
        for source_index, raw_source in enumerate(inventory_sources):
            source_field = f"{line_field}.inventory_sources[{source_index}]"
            if not isinstance(raw_source, Mapping):
                raise _incomplete(f"{source_field} 必须是对象", field=source_field)
            _required_result_text(
                raw_source.get("carryover_entry_id"),
                field=f"{source_field}.carryover_entry_id",
            )
            _required_result_text(
                raw_source.get("source_contract_no"),
                field=f"{source_field}.source_contract_no",
            )
            _required_result_decimal(
                raw_source.get("historical_tax_unit_price"),
                field=f"{source_field}.historical_tax_unit_price",
            )
            available = _required_result_decimal(
                raw_source.get("available_quantity"),
                field=f"{source_field}.available_quantity",
            )
            suggested = _required_result_decimal(
                raw_source.get("suggested_applied_quantity"),
                field=f"{source_field}.suggested_applied_quantity",
            )
            if suggested > available:
                raise _incomplete(f"{source_field} 建议抵扣量超过可用量", field=source_field)
            suggested_total += suggested
        if suggested_total != carryover:
            raise _incomplete(f"成功响应行 {line_ref} 的库存来源合计不守恒", field=line_field)
        applications = _required_result_list(
            raw_line.get("applications"), field=f"{line_field}.applications"
        )
        application_total = Decimal("0")
        for application_index, raw_application in enumerate(applications):
            application_field = f"{line_field}.applications[{application_index}]"
            if not isinstance(raw_application, Mapping):
                raise _incomplete(
                    f"{application_field} 必须是对象", field=application_field
                )
            _required_result_text(
                raw_application.get("carryover_entry_id"),
                field=f"{application_field}.carryover_entry_id",
            )
            _required_result_text(
                raw_application.get("source_contract_no"),
                field=f"{application_field}.source_contract_no",
            )
            _required_result_decimal(
                raw_application.get("historical_tax_unit_price"),
                field=f"{application_field}.historical_tax_unit_price",
            )
            application_total += _required_result_decimal(
                raw_application.get("applied_quantity"),
                field=f"{application_field}.applied_quantity",
                positive=True,
            )
        if application_total != carryover:
            raise _incomplete(f"成功响应行 {line_ref} 的库存应用合计不守恒", field=line_field)
        if purchased > 0:
            positive_suppliers.add(supplier_name)
            _required_result_text(raw_line.get("contract_id"), field=f"{line_field}.contract_id")
            _required_result_text(raw_line.get("contract_no"), field=f"{line_field}.contract_no")
            _required_result_decimal(raw_line.get("tax_unit_price"), field=f"{line_field}.tax_unit_price")
    if actual_refs != set(expected_lines):
        raise _incomplete(
            f"成功响应采购行集合与请求不一致: missing={sorted(set(expected_lines) - actual_refs)}",
            field="purchase_lines",
        )

    contracts = _required_result_list(response.get("contracts"), field="contracts")
    contract_suppliers: set[str] = set()
    for contract_index, raw_contract in enumerate(contracts):
        field = f"contracts[{contract_index}]"
        if not isinstance(raw_contract, Mapping):
            raise _incomplete(f"{field} 必须是对象", field=field)
        _required_result_text(raw_contract.get("contract_id"), field=f"{field}.contract_id")
        supplier_name = _required_result_text(
            raw_contract.get("supplier_name"), field=f"{field}.supplier_name"
        )
        _required_result_text(raw_contract.get("contract_no"), field=f"{field}.contract_no")
        daily_sequence = _required_nonnegative_int(
            raw_contract.get("daily_sequence"), field=f"{field}.daily_sequence"
        )
        supplier_sequence = _required_nonnegative_int(
            raw_contract.get("supplier_contract_sequence"),
            field=f"{field}.supplier_contract_sequence",
        )
        supplier_count = _required_nonnegative_int(
            raw_contract.get("supplier_contract_count"),
            field=f"{field}.supplier_contract_count",
        )
        if daily_sequence == 0 or supplier_sequence == 0 or supplier_count == 0:
            raise _incomplete(f"{field} 的合同序号必须是正整数", field=field)
        if supplier_sequence > supplier_count:
            raise _incomplete(
                f"{field}.supplier_contract_sequence 不得大于供应商累计合同数",
                field=field,
            )
        if supplier_name in contract_suppliers:
            raise _incomplete(f"供应商合同重复: {supplier_name}", field=field)
        contract_suppliers.add(supplier_name)
    if contract_suppliers != positive_suppliers:
        raise _incomplete(
            "contracts 与实际采购量大于零的供应商集合不一致", field="contracts"
        )


def _required_nonnegative_int(value: Any, *, field: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise _incomplete(f"{field} 必须是非负整数", field=field)
    return value


def _validate_inventory_confirmation(
    confirmation: Mapping[str, Any],
    *,
    request_payload: Mapping[str, Any],
) -> None:
    _required_result_text(confirmation.get("quote_id"), field="confirmation.quote_id")
    if confirmation.get("inventory_changes_committed") is not False:
        raise _incomplete(
            "库存报价必须明确声明尚未修改库存",
            field="confirmation.inventory_changes_committed",
        )
    planned_total = _required_result_decimal(
        confirmation.get("planned_shipment_quantity"),
        field="confirmation.planned_shipment_quantity",
    )
    proposed_deduction_total = _required_result_decimal(
        confirmation.get("proposed_inventory_deduction_quantity"),
        field="confirmation.proposed_inventory_deduction_quantity",
    )
    proposed_purchase_total = _required_result_decimal(
        confirmation.get("proposed_purchase_quantity"),
        field="confirmation.proposed_purchase_quantity",
    )
    if planned_total != proposed_deduction_total + proposed_purchase_total:
        raise _incomplete(
            "确认总量不满足 计划量=拟抵扣量+拟采购量",
            field="confirmation",
        )

    expected_lines = list(_expected_purchase_lines(request_payload).values())
    expected_by_key = {
        (line["supplier_name"], str(line["model"]).casefold()): line
        for line in expected_lines
    }
    all_line_count = _required_nonnegative_int(
        confirmation.get("all_line_count"), field="confirmation.all_line_count"
    )
    affected_line_count = _required_nonnegative_int(
        confirmation.get("affected_line_count"),
        field="confirmation.affected_line_count",
    )
    omitted_line_count = _required_nonnegative_int(
        confirmation.get("omitted_unaffected_line_count"),
        field="confirmation.omitted_unaffected_line_count",
    )
    _required_nonnegative_int(
        confirmation.get("omitted_unused_inventory_source_count"),
        field="confirmation.omitted_unused_inventory_source_count",
    )
    raw_lines = _required_result_list(
        confirmation.get("affected_lines"), field="confirmation.affected_lines"
    )
    if (
        all_line_count != len(expected_lines)
        or all_line_count != affected_line_count + omitted_line_count
        or affected_line_count != len(raw_lines)
    ):
        raise _incomplete("确认型号计数不守恒", field="confirmation")
    if planned_total != sum(
        (line["planned"] for line in expected_lines), Decimal("0")
    ):
        raise _incomplete(
            "确认计划总量与请求不一致",
            field="confirmation.planned_shipment_quantity",
        )

    seen_keys: set[tuple[str, str]] = set()
    affected_deduction_total = Decimal("0")
    seen_inventory_entry_ids: set[str] = set()
    for line_index, raw_line in enumerate(raw_lines):
        line_field = f"confirmation.affected_lines[{line_index}]"
        if not isinstance(raw_line, Mapping):
            raise _incomplete(f"{line_field} 必须是对象", field=line_field)
        supplier_name = _required_result_text(
            raw_line.get("supplier_name"), field=f"{line_field}.supplier_name"
        )
        model = _required_result_text(
            raw_line.get("model"), field=f"{line_field}.model"
        )
        _required_result_text(
            raw_line.get("contract_product_name"),
            field=f"{line_field}.contract_product_name",
        )
        key = (supplier_name, model.casefold())
        expected = expected_by_key.get(key)
        if expected is None or key in seen_keys:
            raise _incomplete(
                f"受影响型号无法匹配请求或重复: {supplier_name}/{model}",
                field=line_field,
            )
        seen_keys.add(key)
        planned = _required_result_decimal(
            raw_line.get("planned_shipment_quantity"),
            field=f"{line_field}.planned_shipment_quantity",
            positive=True,
        )
        proposed_deduction = _required_result_decimal(
            raw_line.get("proposed_inventory_deduction_quantity"),
            field=f"{line_field}.proposed_inventory_deduction_quantity",
            positive=True,
        )
        proposed_purchase = _required_result_decimal(
            raw_line.get("proposed_purchase_quantity"),
            field=f"{line_field}.proposed_purchase_quantity",
        )
        if (
            planned != expected["planned"]
            or planned != proposed_deduction + proposed_purchase
        ):
            raise _incomplete(
                f"受影响型号数量不守恒: {supplier_name}/{model}",
                field=line_field,
            )
        sources = _required_result_list(
            raw_line.get("inventory_sources"),
            field=f"{line_field}.inventory_sources",
        )
        source_total = Decimal("0")
        for source_index, raw_source in enumerate(sources):
            source_field = f"{line_field}.inventory_sources[{source_index}]"
            if not isinstance(raw_source, Mapping):
                raise _incomplete(f"{source_field} 必须是对象", field=source_field)
            carryover_entry_id = _required_result_text(
                raw_source.get("carryover_entry_id"),
                field=f"{source_field}.carryover_entry_id",
            )
            if carryover_entry_id in seen_inventory_entry_ids:
                raise _incomplete(
                    f"库存批次身份重复: {carryover_entry_id}",
                    field=f"{source_field}.carryover_entry_id",
                )
            seen_inventory_entry_ids.add(carryover_entry_id)
            source_kind = _required_result_text(
                raw_source.get("source_kind"),
                field=f"{source_field}.source_kind",
            )
            if source_kind not in {"opening_inventory", "reconciliation"}:
                raise _incomplete(
                    f"{source_field}.source_kind 未知: {source_kind}",
                    field=f"{source_field}.source_kind",
                )
            _required_result_text(
                raw_source.get("source_contract_no"),
                field=f"{source_field}.source_contract_no",
            )
            _required_result_text(
                raw_source.get("source_sp_no"),
                field=f"{source_field}.source_sp_no",
            )
            _required_result_text(
                raw_source.get("source_reference"),
                field=f"{source_field}.source_reference",
            )
            _required_result_decimal(
                raw_source.get("historical_tax_unit_price"),
                field=f"{source_field}.historical_tax_unit_price",
            )
            original = _required_result_decimal(
                raw_source.get("original_quantity"),
                field=f"{source_field}.original_quantity",
                positive=True,
            )
            current_remaining = _required_result_decimal(
                raw_source.get("current_remaining_quantity"),
                field=f"{source_field}.current_remaining_quantity",
            )
            replacement_released = _required_result_decimal(
                raw_source.get("replacement_released_quantity"),
                field=f"{source_field}.replacement_released_quantity",
            )
            available_after_release = _required_result_decimal(
                raw_source.get("available_after_release"),
                field=f"{source_field}.available_after_release",
            )
            proposed_applied = _required_result_decimal(
                raw_source.get("proposed_applied_quantity"),
                field=f"{source_field}.proposed_applied_quantity",
                positive=True,
            )
            if current_remaining + replacement_released != available_after_release:
                raise _incomplete(
                    f"{source_field} 不满足 当前剩余+替代返还=替代后可用",
                    field=source_field,
                )
            if current_remaining > original or available_after_release > original:
                raise _incomplete(
                    f"{source_field} 的当前或替代后可用量超过原始数量",
                    field=source_field,
                )
            if proposed_applied > available_after_release:
                raise _incomplete(
                    f"{source_field} 拟抵扣量超过替代后可用量",
                    field=source_field,
                )
            source_total += proposed_applied
        if source_total != proposed_deduction:
            raise _incomplete(
                f"受影响型号的库存来源合计不守恒: {supplier_name}/{model}",
                field=line_field,
            )
        affected_deduction_total += proposed_deduction
    if affected_deduction_total != proposed_deduction_total:
        raise _incomplete(
            "受影响型号拟抵扣合计与确认总量不一致",
            field="confirmation.proposed_inventory_deduction_quantity",
        )


def _validate_batch_replacement_confirmation(
    confirmation: Mapping[str, Any],
    *,
    request_payload: Mapping[str, Any],
) -> None:
    conflicts = _required_result_list(
        confirmation.get("conflicts"), field="confirmation.conflicts"
    )
    if not conflicts:
        raise _incomplete("批次替换确认没有冲突明细", field="confirmation.conflicts")
    requested_sps = {
        _required_result_text(item.get("sp_no"), field="request.sps[].sp_no").upper()
        for item in _required_result_list(request_payload.get("sps"), field="request.sps")
        if isinstance(item, Mapping)
    }
    for index, raw_conflict in enumerate(conflicts):
        field = f"confirmation.conflicts[{index}]"
        if not isinstance(raw_conflict, Mapping):
            raise _incomplete(f"{field} 必须是对象", field=field)
        sp_no = _required_result_text(raw_conflict.get("sp_no"), field=f"{field}.sp_no")
        if sp_no.upper() not in requested_sps:
            raise _incomplete(f"{field}.sp_no 不在当前请求中", field=f"{field}.sp_no")
        for name in ("batch_id", "batch_no"):
            _required_result_text(raw_conflict.get(name), field=f"{field}.{name}")
        version = raw_conflict.get("version_no")
        if isinstance(version, bool) or not isinstance(version, int) or version <= 0:
            raise _incomplete(f"{field}.version_no 必须是正整数", field=f"{field}.version_no")


def validate_purchase_response(
    *,
    status_code: int,
    response: Mapping[str, Any],
    request_payload: Mapping[str, Any],
) -> None:
    if status_code < 400:
        _validate_success_response(response, request_payload=request_payload)
        return
    schema = _required_result_text(response.get("response_schema"), field="response_schema")
    if schema != PURCHASE_CONFIRMATION_RESPONSE_SCHEMA:
        raise PurchaseBatchClientError(
            "erp_purchase_confirmation_schema_unsupported",
            f"ERP 采购确认响应版本不受支持: {schema}",
            detail={"response_schema": schema},
        )
    response_request_id = _required_result_text(
        response.get("request_id"), field="request_id"
    )
    if response_request_id != request_payload.get("request_id"):
        raise _incomplete("409 确认响应 request_id 与请求不一致", field="request_id")
    error = response.get("error")
    if not isinstance(error, Mapping):
        raise _incomplete("409 业务响应缺少 error", field="error")
    code = _required_result_text(error.get("code"), field="error.code")
    _required_result_text(error.get("message"), field="error.message")
    status = _required_result_text(response.get("status"), field="status")
    confirmation = response.get("confirmation")
    if not isinstance(confirmation, Mapping):
        raise _incomplete("409 业务响应缺少 confirmation", field="confirmation")
    kind = _required_result_text(confirmation.get("kind"), field="confirmation.kind")
    if code in {
        "purchase_inventory_confirmation_required",
        "purchase_inventory_quote_stale",
    }:
        expected_status = (
            "quote_stale"
            if code == "purchase_inventory_quote_stale"
            else "confirmation_required"
        )
        if status != expected_status or kind != "inventory_quote":
            raise _incomplete("库存确认的 status/kind 与 error.code 不一致", field="confirmation")
        _validate_inventory_confirmation(
            confirmation,
            request_payload=request_payload,
        )
    elif code == "purchase_batch_replace_confirmation_required":
        if status != "confirmation_required" or kind != "batch_replacement":
            raise _incomplete("批次替换的 status/kind 与 error.code 不一致", field="confirmation")
        _validate_batch_replacement_confirmation(
            confirmation,
            request_payload=request_payload,
        )
    else:
        raise _incomplete(f"未识别的确认响应错误码: {code}", field="error.code")


def confirmation_result(
    *,
    response: Mapping[str, Any],
    status_code: int,
    request_payload: Mapping[str, Any],
) -> dict[str, Any]:
    raw_error = response.get("error")
    error = dict(raw_error) if isinstance(raw_error, Mapping) else {}
    code = str(error.get("code") or "purchase_confirmation_required")
    message = str(error.get("message") or "ERP 要求用户确认后才能创建采购批次")
    return {
        "success": False,
        "status": str(response.get("status") or "confirmation_required"),
        "request_id": request_payload["request_id"],
        "source_sha256": request_payload["source_sha256"],
        "error": {
            "code": code,
            "message": message,
            "http_status": status_code,
        },
        "erp": dict(response),
    }


def _renamed_path(path: Path, suffix: str) -> Path:
    target = path.with_name(f"{path.stem}{suffix}{path.suffix}")
    if target != path:
        path.replace(target)
    return target


def _rebuild_draft_quantity_sheet(worksheet: Any) -> None:
    old_columns = [purchase_summary._clean_cell(cell.value) for cell in worksheet[1]]
    if "数量" not in old_columns:
        return
    new_columns = _replace_quantity_header(old_columns)
    rebuilt: list[list[Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if purchase_summary._clean_cell(row[0] if row else "") == "合计":
            continue
        original = _row_dict(old_columns, row)
        try:
            planned = _decimal(original.get("数量"))
        except (InvalidOperation, ValueError) as exc:
            raise PurchaseBatchClientError(
                "draft_artifact_quantity_invalid",
                f"草稿工作簿数量无法解析: sheet={worksheet.title}, value={original.get('数量')}",
            ) from exc
        rebuilt.append(
            _formal_row_values(
                original,
                planned=planned,
                purchased=planned,
                carryover=Decimal("0"),
                columns=new_columns,
            )
        )
    title = worksheet.title
    workbook = worksheet.parent
    index = workbook.index(worksheet)
    workbook.remove(worksheet)
    target = workbook.create_sheet(title, index)
    purchase_summary._write_rows(target, tuple(new_columns), rebuilt, append_total=True)


def mark_draft_workbooks(result: dict[str, Any]) -> dict[str, Any]:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    paths = [Path(result["purchase_summary_xlsx"]), *(Path(item) for item in result["restock_xlsx_paths"])]
    updated: list[Path] = []
    for path in paths:
        workbook = load_workbook(path)
        for sheet_name in list(workbook.sheetnames):
            if sheet_name == purchase_summary.UNMATCHED_SHEET_NAME:
                continue
            _rebuild_draft_quantity_sheet(workbook[sheet_name])
        note = workbook.create_sheet("草稿-未同步ERP", 0)
        note["A1"] = "草稿：未同步 ERP，未占用库存，不含正式合同编号"
        note["A1"].font = Font(bold=True)
        note["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFFC000")
        note.column_dimensions["A"].width = 72
        workbook.save(path)
        updated.append(_renamed_path(path, "-DRAFT"))

    result["purchase_summary_xlsx"] = str(updated[0])
    result["restock_xlsx_paths"] = [str(path) for path in updated[1:]]
    result["restock_outputs"] = [
        {"delivery_no": delivery_no, "output_xlsx": str(path)}
        for delivery_no, path in zip(result["delivery_nos"], updated[1:], strict=True)
    ]
    result["mode"] = "draft"
    result["erp_synced"] = False
    result["official_contract_numbers_created"] = False
    return result


def _line_map(erp_result: Mapping[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    raw_lines = erp_result.get("purchase_lines") or erp_result.get("lines") or []
    return {
        (
            purchase_summary._clean_cell(line.get("supplier_name")),
            purchase_summary._clean_cell(line.get("model")).casefold(),
        ): dict(line)
        for line in raw_lines
        if isinstance(line, Mapping)
    }


def _replace_quantity_header(columns: list[str]) -> list[str]:
    index = columns.index("数量")
    return [*columns[:index], *FORMAL_QUANTITY_COLUMNS, *columns[index + 1 :]]


def _row_dict(columns: list[str], values: list[Any] | tuple[Any, ...]) -> dict[str, Any]:
    return {
        column: values[index] if index < len(values) else ""
        for index, column in enumerate(columns)
    }


def _formal_row_values(
    original: Mapping[str, Any],
    *,
    planned: Decimal,
    purchased: Decimal,
    carryover: Decimal,
    columns: list[str],
) -> list[Any]:
    values = dict(original)
    values["计划发货量"] = purchase_summary._decimal_to_cell_value(planned)
    values["本次采购量"] = purchase_summary._decimal_to_cell_value(purchased)
    values["留存库存抵扣量"] = purchase_summary._decimal_to_cell_value(carryover)
    values.pop("数量", None)
    return [values.get(column, "") for column in columns]


def _set_purchase_source_fields(
    values: dict[str, Any],
    *,
    details: list[Mapping[str, Any]],
    product_names: Mapping[tuple[str, str], str],
) -> None:
    if not details:
        raise _incomplete("采购汇总来源行为空", field="purchase_line.allocation_details")
    source_sps: list[str] = []
    seen_sps: set[str] = set()
    sku_quantities: OrderedDict[str, Decimal] = OrderedDict()
    sku_product_names: OrderedDict[str, str] = OrderedDict()
    for index, detail in enumerate(details):
        field = f"purchase_line.allocation_details[{index}]"
        sp_no = _required_result_text(detail.get("sp_no"), field=f"{field}.sp_no").upper()
        stock_sku = _required_result_text(
            detail.get("stock_sku"), field=f"{field}.stock_sku"
        ).upper()
        quantity = _required_result_decimal(
            detail.get("quantity"), field=f"{field}.quantity", positive=True
        )
        product_name = product_names.get((sp_no, stock_sku))
        if not product_name:
            raise _incomplete(
                f"SP={sp_no}, SKU={stock_sku} 缺少冻结产品名称",
                field="request.sps[].planned_lines[].product_name",
            )
        if sp_no not in seen_sps:
            seen_sps.add(sp_no)
            source_sps.append(sp_no)
        existing_product_name = sku_product_names.get(stock_sku)
        if existing_product_name is not None and existing_product_name != product_name:
            raise _incomplete(
                f"SKU={stock_sku} 的冻结产品名称不一致",
                field="request.sps[].planned_lines[].product_name",
            )
        sku_quantities[stock_sku] = sku_quantities.get(stock_sku, Decimal("0")) + quantity
        sku_product_names.setdefault(stock_sku, product_name)

    values["库存sku"] = "\n".join(
        f"{stock_sku} × {_decimal_text(quantity)}"
        for stock_sku, quantity in sku_quantities.items()
    )
    values["产品名称"] = "\n".join(sku_product_names.values())
    values["来源SP单号"] = "\n".join(source_sps)
    values["库存sku（第一行）"] = next(iter(sku_quantities))
    values["产品名称（第一行）"] = next(iter(sku_product_names.values()))


def _purchase_source_rows(
    original: Mapping[str, Any],
    line: Mapping[str, Any],
    *,
    columns: list[str],
    product_names: Mapping[tuple[str, str], str],
) -> tuple[list[list[Any]], list[list[Any]]]:
    current_details: list[Mapping[str, Any]] = []
    current_quantity = Decimal("0")
    inventory_quantity = Decimal("0")
    inventory_groups: OrderedDict[tuple[str, Decimal], dict[str, Any]] = OrderedDict()
    for index, raw_detail in enumerate(
        _required_result_list(
            line.get("allocation_details"), field="purchase_line.allocation_details"
        )
    ):
        field = f"purchase_line.allocation_details[{index}]"
        if not isinstance(raw_detail, Mapping):
            raise _incomplete(f"{field} 必须是对象", field=field)
        source_kind = _required_result_text(
            raw_detail.get("source_kind"), field=f"{field}.source_kind"
        )
        quantity = _required_result_decimal(
            raw_detail.get("quantity"), field=f"{field}.quantity", positive=True
        )
        if source_kind == "current_purchase":
            current_details.append(raw_detail)
            current_quantity += quantity
            continue
        if source_kind != "carryover":
            raise _incomplete(
                f"{field}.source_kind 未知: {source_kind}",
                field=f"{field}.source_kind",
            )
        contract_no = _required_result_text(
            raw_detail.get("source_contract_no"),
            field=f"{field}.source_contract_no",
        )
        price = _required_result_decimal(
            raw_detail.get("historical_tax_unit_price"),
            field=f"{field}.historical_tax_unit_price",
        )
        bucket = inventory_groups.setdefault(
            (contract_no, price),
            {"quantity": Decimal("0"), "details": []},
        )
        bucket["quantity"] += quantity
        bucket["details"].append(raw_detail)
        inventory_quantity += quantity

    planned = _required_result_decimal(
        line.get("planned_shipment_quantity"),
        field="purchase_line.planned_shipment_quantity",
        positive=True,
    )
    purchased = _required_result_decimal(
        line.get("purchase_quantity"), field="purchase_line.purchase_quantity"
    )
    carryover = _required_result_decimal(
        line.get("carryover_applied_quantity"),
        field="purchase_line.carryover_applied_quantity",
    )
    if (
        planned != purchased + carryover
        or current_quantity != purchased
        or inventory_quantity != carryover
    ):
        raise _incomplete(
            "采购汇总来源分配与计划量、采购量或库存抵扣量不一致",
            field="purchase_line.allocation_details",
        )

    current_rows: list[list[Any]] = []
    inventory_rows: list[list[Any]] = []
    if current_quantity > 0:
        values = dict(original)
        _set_purchase_source_fields(
            values,
            details=current_details,
            product_names=product_names,
        )
        values[purchase_summary.CURRENT_PURCHASE_CONTRACT_HEADER] = _required_result_text(
            line.get("contract_no"), field="purchase_line.contract_no"
        )
        values[purchase_summary.HISTORICAL_INVENTORY_CONTRACT_HEADER] = ""
        original_price = _decimal(values.get("原价"))
        values["总价"] = purchase_summary._decimal_to_cell_value(
            original_price * current_quantity
        )
        tax_unit_price = line.get("tax_unit_price")
        if tax_unit_price is not None and purchase_summary._is_zhengfei_manufacturer(
            values.get("厂家")
        ):
            average = _decimal(tax_unit_price)
            values["均价"] = purchase_summary._decimal_to_cell_value(average)
            values["总价（均价）"] = purchase_summary._decimal_to_cell_value(
                average * current_quantity
            )
        current_rows.append(
            _formal_row_values(
                values,
                planned=current_quantity,
                purchased=current_quantity,
                carryover=Decimal("0"),
                columns=columns,
            )
        )

    for (contract_no, price), bucket in inventory_groups.items():
        quantity = bucket["quantity"]
        values = dict(original)
        _set_purchase_source_fields(
            values,
            details=bucket["details"],
            product_names=product_names,
        )
        values["原价"] = purchase_summary._decimal_to_cell_value(price)
        values["均价"] = ""
        values["总价"] = purchase_summary._decimal_to_cell_value(price * quantity)
        values["总价（均价）"] = ""
        values[purchase_summary.CURRENT_PURCHASE_CONTRACT_HEADER] = ""
        values[purchase_summary.HISTORICAL_INVENTORY_CONTRACT_HEADER] = (
            f"{contract_no} × {_decimal_text(quantity)}"
        )
        inventory_rows.append(
            _formal_row_values(
                values,
                planned=quantity,
                purchased=Decimal("0"),
                carryover=quantity,
                columns=columns,
            )
        )
    return current_rows, inventory_rows


def _sp_product_names(
    request_payload: Mapping[str, Any],
) -> dict[tuple[str, str], str]:
    names: dict[tuple[str, str], str] = {}
    for sp_index, raw_sp in enumerate(
        _required_result_list(request_payload.get("sps"), field="request.sps")
    ):
        field = f"request.sps[{sp_index}]"
        if not isinstance(raw_sp, Mapping):
            raise _incomplete(f"{field} 必须是对象", field=field)
        sp_no = _required_result_text(raw_sp.get("sp_no"), field=f"{field}.sp_no").upper()
        for line_index, raw_line in enumerate(
            _required_result_list(raw_sp.get("planned_lines"), field=f"{field}.planned_lines")
        ):
            line_field = f"{field}.planned_lines[{line_index}]"
            if not isinstance(raw_line, Mapping):
                raise _incomplete(f"{line_field} 必须是对象", field=line_field)
            stock_sku = _required_result_text(
                raw_line.get("stock_sku"), field=f"{line_field}.stock_sku"
            ).upper()
            product_name = _required_result_text(
                raw_line.get("product_name"), field=f"{line_field}.product_name"
            )
            key = (sp_no, stock_sku)
            previous = names.setdefault(key, product_name)
            if previous != product_name:
                raise _incomplete(
                    f"同一 SP/SKU 的产品名称不一致: {sp_no}/{stock_sku}",
                    field=line_field,
                )
    return names


def _set_stock_sku_fields(
    values: dict[str, Any],
    *,
    skus: Mapping[str, Decimal],
    sp_no: str,
    product_names: Mapping[tuple[str, str], str],
) -> None:
    if not skus:
        raise _incomplete(f"SP={sp_no} 的来源行没有库存 SKU", field="allocation_details")
    sku_lines: list[str] = []
    name_lines: list[str] = []
    for stock_sku, quantity in skus.items():
        normalized_sku = purchase_summary._clean_cell(stock_sku).upper()
        product_name = product_names.get((sp_no.upper(), normalized_sku))
        if not product_name:
            raise _incomplete(
                f"SP={sp_no}, SKU={normalized_sku} 缺少冻结产品名称",
                field="request.sps[].planned_lines[].product_name",
            )
        sku_lines.append(
            f"{normalized_sku} × {purchase_summary._decimal_to_cell_value(quantity)}"
        )
        name_lines.append(product_name)
    values["库存sku"] = "\n".join(sku_lines)
    values["产品名称"] = "\n".join(name_lines)
    values["库存sku（第一行）"] = next(iter(skus)).upper()
    values["产品名称（第一行）"] = name_lines[0]


def _rebuild_purchase_sheet(
    worksheet: Any,
    lines: Mapping[tuple[str, str], dict[str, Any]],
    *,
    product_names: Mapping[tuple[str, str], str],
) -> None:
    from openpyxl.styles import PatternFill

    old_columns = [purchase_summary._clean_cell(cell.value) for cell in worksheet[1]]
    if "数量" not in old_columns:
        return
    old_rows = [list(row) for row in worksheet.iter_rows(min_row=2, values_only=True)]
    data_rows = [row for row in old_rows if purchase_summary._clean_cell(row[0] if row else "") != "合计"]
    new_columns = _replace_quantity_header(old_columns)
    normal_rows: list[list[Any]] = []
    inventory_rows: list[list[Any]] = []
    for row in data_rows:
        original = _row_dict(old_columns, row)
        key = (
            purchase_summary._clean_cell(original.get("厂家")),
            purchase_summary._clean_cell(original.get("型号")).casefold(),
        )
        line = lines.get(key)
        if line is None:
            raise PurchaseBatchClientError(
                "erp_purchase_result_incomplete",
                f"ERP 结果缺少采购行: 厂家={key[0]}, 型号={original.get('型号')}",
            )
        current, inventory = _purchase_source_rows(
            original,
            line,
            columns=new_columns,
            product_names=product_names,
        )
        normal_rows.extend(current)
        inventory_rows.extend(inventory)

    title = worksheet.title
    workbook = worksheet.parent
    index = workbook.index(worksheet)
    workbook.remove(worksheet)
    target = workbook.create_sheet(title, index)
    all_rows = [*normal_rows, *inventory_rows]
    purchase_summary._write_rows(target, tuple(new_columns), all_rows, append_total=True)
    yellow = PatternFill(fill_type="solid", fgColor=INVENTORY_ROW_FILL_COLOR)
    first_inventory_row = 2 + len(normal_rows)
    for row_index in range(first_inventory_row, first_inventory_row + len(inventory_rows)):
        for cell in target[row_index]:
            cell.fill = copy(yellow)


def _restock_source_rows(
    original: Mapping[str, Any],
    line: Mapping[str, Any],
    *,
    sp_no: str,
    columns: list[str],
    product_names: Mapping[tuple[str, str], str],
) -> tuple[list[list[Any]], list[list[Any]]]:
    current: list[list[Any]] = []
    inventory: list[list[Any]] = []
    details = [
        dict(item)
        for item in line.get("allocation_details") or []
        if isinstance(item, Mapping) and purchase_summary._clean_cell(item.get("sp_no")).upper() == sp_no.upper()
    ]
    current_skus: OrderedDict[str, Decimal] = OrderedDict()
    for item in details:
        if item.get("source_kind") != "current_purchase":
            continue
        sku = purchase_summary._clean_cell(item.get("stock_sku")).upper()
        current_skus[sku] = current_skus.get(sku, Decimal("0")) + _decimal(
            item.get("quantity")
        )
    current_quantity = sum(current_skus.values(), Decimal("0"))

    def apply_price_totals(
        values: dict[str, Any],
        *,
        quantity: Decimal,
        source_price: Decimal,
        average_price: Decimal | None,
    ) -> None:
        gross_margin = _decimal(values.get("毛利率"))
        manufacturer = purchase_summary._clean_cell(values.get("厂家"))
        model = purchase_summary._clean_cell(values.get("型号"))
        multiplier = restock_workbook._tax_multiplier_from_rate(
            line.get("tax_rate") or "",
            manufacturer=manufacturer,
            model=model,
        )
        sale_price = (source_price / multiplier / (Decimal("1") - gross_margin)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        values["总价（原价）"] = purchase_summary._decimal_to_cell_value(
            source_price * quantity
        )
        values["售价"] = purchase_summary._decimal_to_cell_value(sale_price)
        values["总价（售价）"] = purchase_summary._decimal_to_cell_value(
            sale_price * quantity
        )
        if average_price is None:
            values["均价"] = ""
            values["总价（均价）"] = ""
            values["售价(均价)"] = ""
            values["总价（售价(均价)）"] = ""
            return
        average_sale_price = (
            average_price / multiplier / (Decimal("1") - gross_margin)
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        values["均价"] = purchase_summary._decimal_to_cell_value(average_price)
        values["总价（均价）"] = purchase_summary._decimal_to_cell_value(
            average_price * quantity
        )
        values["售价(均价)"] = purchase_summary._decimal_to_cell_value(
            average_sale_price
        )
        values["总价（售价(均价)）"] = purchase_summary._decimal_to_cell_value(
            average_sale_price * quantity
        )

    if current_quantity > 0:
        values = dict(original)
        values["采购订单号"] = line.get("contract_no") or ""
        _set_stock_sku_fields(
            values,
            skus=current_skus,
            sp_no=sp_no,
            product_names=product_names,
        )
        original_price = _decimal(values.get("原价"))
        average_price: Decimal | None = None
        if line.get("tax_unit_price") is not None and purchase_summary._is_zhengfei_manufacturer(
            values.get("厂家")
        ):
            average_price = _decimal(line["tax_unit_price"])
        apply_price_totals(
            values,
            quantity=current_quantity,
            source_price=original_price,
            average_price=average_price,
        )
        current.append(
            _formal_row_values(
                values,
                planned=current_quantity,
                purchased=current_quantity,
                carryover=Decimal("0"),
                columns=columns,
            )
        )

    grouped: OrderedDict[tuple[str, str], dict[str, Any]] = OrderedDict()
    for detail in details:
        if detail.get("source_kind") != "carryover":
            continue
        contract_no = purchase_summary._clean_cell(detail.get("source_contract_no"))
        price = _decimal(detail.get("historical_tax_unit_price"))
        key = (contract_no, _decimal_text(price))
        bucket = grouped.setdefault(key, {"quantity": Decimal("0"), "skus": OrderedDict()})
        quantity = _decimal(detail.get("quantity"))
        bucket["quantity"] += quantity
        sku = purchase_summary._clean_cell(detail.get("stock_sku"))
        bucket["skus"][sku] = bucket["skus"].get(sku, Decimal("0")) + quantity
    for (contract_no, price_text), bucket in grouped.items():
        quantity = bucket["quantity"]
        price = Decimal(price_text)
        values = dict(original)
        values["采购订单号"] = contract_no
        values["原价"] = purchase_summary._decimal_to_cell_value(price)
        apply_price_totals(
            values,
            quantity=quantity,
            source_price=price,
            average_price=None,
        )
        if bucket["skus"]:
            _set_stock_sku_fields(
                values,
                skus=bucket["skus"],
                sp_no=sp_no,
                product_names=product_names,
            )
        inventory.append(
            _formal_row_values(
                values,
                planned=quantity,
                purchased=Decimal("0"),
                carryover=quantity,
                columns=columns,
            )
        )
    return current, inventory


def _rebuild_restock_sheet(
    worksheet: Any,
    lines: Mapping[tuple[str, str], dict[str, Any]],
    *,
    sp_no: str,
    product_names: Mapping[tuple[str, str], str],
) -> None:
    from openpyxl.styles import PatternFill

    old_columns = [purchase_summary._clean_cell(cell.value) for cell in worksheet[1]]
    if "数量" not in old_columns:
        return
    data_rows = [
        list(row)
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if purchase_summary._clean_cell(row[0] if row else "") != "合计"
    ]
    new_columns = _replace_quantity_header(old_columns)
    normal_rows: list[list[Any]] = []
    inventory_rows: list[list[Any]] = []
    for row in data_rows:
        original = _row_dict(old_columns, row)
        key = (
            purchase_summary._clean_cell(original.get("厂家")),
            purchase_summary._clean_cell(original.get("型号")).casefold(),
        )
        line = lines.get(key)
        if line is None:
            raise PurchaseBatchClientError(
                "erp_purchase_result_incomplete",
                f"ERP 结果缺少备货行: SP={sp_no}, 厂家={key[0]}, 型号={original.get('型号')}",
            )
        current, inventory = _restock_source_rows(
            original,
            line,
            sp_no=sp_no,
            columns=new_columns,
            product_names=product_names,
        )
        normal_rows.extend(current)
        inventory_rows.extend(inventory)

    title = worksheet.title
    workbook = worksheet.parent
    index = workbook.index(worksheet)
    workbook.remove(worksheet)
    target = workbook.create_sheet(title, index)
    all_rows = [*normal_rows, *inventory_rows]
    purchase_summary._write_rows(target, tuple(new_columns), all_rows, append_total=True)
    yellow = PatternFill(fill_type="solid", fgColor=INVENTORY_ROW_FILL_COLOR)
    first_inventory_row = 2 + len(normal_rows)
    for row_index in range(first_inventory_row, first_inventory_row + len(inventory_rows)):
        for cell in target[row_index]:
            cell.fill = copy(yellow)


def apply_formal_erp_result(
    result: dict[str, Any],
    erp_result: Mapping[str, Any],
    *,
    request_payload: Mapping[str, Any],
) -> dict[str, Any]:
    from openpyxl import load_workbook

    lines = _line_map(erp_result)
    if not lines:
        raise PurchaseBatchClientError(
            "erp_purchase_result_incomplete", "ERP 成功响应未返回 purchase_lines"
        )
    product_names = _sp_product_names(request_payload)
    purchase_path = Path(result["purchase_summary_xlsx"])
    workbook = load_workbook(purchase_path)
    for sheet_name in list(workbook.sheetnames):
        if sheet_name == purchase_summary.UNMATCHED_SHEET_NAME:
            continue
        _rebuild_purchase_sheet(
            workbook[sheet_name],
            lines,
            product_names=product_names,
        )
    workbook.save(purchase_path)

    for output in result["restock_outputs"]:
        path = Path(output["output_xlsx"])
        workbook = load_workbook(path)
        _rebuild_restock_sheet(
            workbook[restock_workbook.RESTOCK_SHEET_NAME],
            lines,
            sp_no=output["delivery_no"],
            product_names=product_names,
        )
        workbook.save(path)
    result["mode"] = "formal"
    result["erp_synced"] = True
    result["erp"] = dict(erp_result)
    result["batch_id"] = erp_result.get("batch_id")
    result["batch_no"] = erp_result.get("batch_no")
    result["version_no"] = erp_result.get("version_no")
    result["contracts"] = list(erp_result.get("contracts") or [])
    result["purchase_lines"] = list(erp_result.get("purchase_lines") or [])
    return result


def _contract_download_filename(
    headers: Mapping[str, str],
    contract: Mapping[str, Any],
) -> str:
    content_disposition = next(
        (
            str(value)
            for key, value in headers.items()
            if str(key).casefold() == "content-disposition"
        ),
        "",
    )
    file_name = ""
    encoded_match = re.search(
        r"filename\*\s*=\s*(?:UTF-8'')?([^;]+)",
        content_disposition,
        flags=re.IGNORECASE,
    )
    if encoded_match:
        file_name = unquote(encoded_match.group(1).strip().strip('"'))
    if not file_name:
        plain_match = re.search(
            r"filename\s*=\s*(?:\"([^\"]+)\"|([^;]+))",
            content_disposition,
            flags=re.IGNORECASE,
        )
        if plain_match:
            file_name = (plain_match.group(1) or plain_match.group(2) or "").strip()
    if not file_name:
        file_name = (
            purchase_summary._clean_cell(contract.get("contract_no"))
            or purchase_summary._clean_cell(contract.get("contract_id"))
            or "ERP采购合同"
        )
    file_name = _INVALID_FILE_NAME_CHARS.sub("_", Path(file_name).name).strip(" .")
    if not file_name:
        file_name = "ERP采购合同"
    if not file_name.casefold().endswith(".xlsx"):
        file_name = f"{file_name}.xlsx"
    return file_name


def download_contract_workbooks(
    result: dict[str, Any],
    *,
    output_dir: str | Path | None = None,
) -> dict[str, Any]:
    from openpyxl import load_workbook
    from services.agent_cli.mabang.erp_http import request_bytes

    target_dir = Path(output_dir or CONTRACT_OUTPUT_DIR).expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)
    resolved_dir = target_dir.resolve()
    outputs: list[dict[str, Any]] = []
    used_names: set[str] = set()
    contracts = _required_result_list(result.get("contracts"), field="contracts")
    for index, raw_contract in enumerate(contracts):
        field = f"contracts[{index}]"
        if not isinstance(raw_contract, Mapping):
            raise _incomplete(f"{field} 必须是对象", field=field)
        contract = dict(raw_contract)
        contract_id = _required_result_text(
            contract.get("contract_id"), field=f"{field}.contract_id"
        )
        _status, content, headers = request_bytes(
            "GET",
            f"/api/v1/erp/contracts/{contract_id}/download",
            operation=f"下载 ERP 正式合同 {contract.get('contract_no') or contract_id}",
        )
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
            workbook.close()
        except Exception as exc:  # noqa: BLE001 - preserve malformed server artifact context
            raise PurchaseBatchClientError(
                "erp_contract_download_invalid",
                f"ERP 返回的合同不是有效 xlsx: contract_id={contract_id}, error={exc}",
            ) from exc
        file_name = _contract_download_filename(headers, contract)
        if file_name.casefold() in used_names:
            stem = Path(file_name).stem
            file_name = f"{stem}-{contract_id[:8]}.xlsx"
        used_names.add(file_name.casefold())
        target = (resolved_dir / file_name).resolve()
        if target.parent != resolved_dir:
            raise PurchaseBatchClientError(
                "erp_contract_download_path_invalid",
                f"ERP 合同文件名超出输出目录: {file_name}",
            )
        temporary = target.with_name(f".{target.name}.{os.getpid()}.part")
        try:
            temporary.write_bytes(content)
            os.replace(temporary, target)
        finally:
            if temporary.exists():
                temporary.unlink()
        outputs.append(
            {
                "contract_id": contract_id,
                "supplier_name": contract.get("supplier_name"),
                "contract_no": contract.get("contract_no"),
                "output_xlsx": str(target),
            }
        )
    result["contract_outputs"] = outputs
    result["contract_xlsx_paths"] = [item["output_xlsx"] for item in outputs]
    return result


def client_error_payload(exc: PurchaseBatchClientError | ErpHttpError) -> dict[str, Any]:
    if isinstance(exc, ErpHttpError):
        return error_payload(exc)
    payload: dict[str, Any] = {"code": exc.code, "message": str(exc)}
    if exc.detail:
        payload["detail"] = exc.detail
    return payload


__all__ = [
    "PurchaseBatchClientError",
    "apply_formal_erp_result",
    "build_purchase_intent",
    "client_error_payload",
    "confirmation_result",
    "download_contract_workbooks",
    "import_purchase_intent",
    "mark_draft_workbooks",
    "unmatched_confirmation_result",
    "validate_purchase_response",
]
