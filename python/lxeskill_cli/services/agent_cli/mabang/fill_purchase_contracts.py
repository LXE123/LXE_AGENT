from __future__ import annotations

import re
from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from services.agent_cli._shared.json_cli import exception_text as _exception_text
from services.agent_cli.mabang import generate_restock_workbook as purchase_summary
from shared.datasets import dataset_dir

OUTPUT_DIR = dataset_dir("fba_purchase_contracts")
SOURCE = "fba_purchase_contract_fill"

PURCHASE_SUMMARY_SHEET = purchase_summary.SUMMARY_SHEET_NAME
ADDENDUM_TEMPLATE_SHEET = "附加件明细模板"
ADDENDUM_OUTPUT_SHEET = "补充协议附加件明细"
REQUIRED_PURCHASE_COLUMNS = (
    "厂家",
    "合同产品名称",
    "型号",
    "单位",
    "原价",
    "总价",
    "税率",
)
PURCHASE_QUANTITY_COLUMNS = ("本次采购量", "数量")
DETAIL_HEADER_MATCHERS = {
    "sequence": ("序号",),
    "product_name": ("产品名称",),
    "model": ("规格型号",),
    "unit": ("单位",),
    "quantity": ("数量",),
    "tax_unit_price": ("含税单价",),
    "tax_amount": ("含税金额",),
    "remark": ("备注",),
}
REQUIRED_DETAIL_HEADER_KEYS = tuple(key for key in DETAIL_HEADER_MATCHERS if key != "model")
INVALID_FILE_NAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


@dataclass
class PurchaseContractLine:
    manufacturer: str
    product_name: str
    model: str
    unit: str
    quantity: Decimal
    tax_unit_price: Decimal
    tax_amount: Decimal
    tax_rate: str


@dataclass
class DetailTableLayout:
    header_row: int
    detail_start_row: int
    summary_row: int
    columns: dict[str, int]
    header_column_spans: dict[str, tuple[int, int]]


def _clean_cell(value: Any) -> str:
    return purchase_summary._clean_cell(value)


def _decimal_to_cell_value(value: Decimal) -> int | float:
    return purchase_summary._decimal_to_cell_value(value)


def _decimal_from_value(value: Any, *, field_name: str, row_number: int) -> Decimal:
    text = str(value).strip() if value is not None else ""
    if not text:
        raise RuntimeError(f"采购汇总表 第{row_number}行 {field_name} 不能为空")
    try:
        result = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise RuntimeError(f"采购汇总表 第{row_number}行 {field_name}非数字: {text}") from exc
    if not result.is_finite():
        raise RuntimeError(f"采购汇总表 第{row_number}行 {field_name}非数字: {text}")
    return result


def _row_value_by_header(row: tuple[Any, ...], indexes: dict[str, int], header: str) -> Any:
    column_index = indexes.get(header)
    if column_index is None:
        return None
    return row[column_index - 1] if column_index - 1 < len(row) else None


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", _clean_cell(value))


def _header_indexes(header_values: tuple[Any, ...]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for column_index, raw_header in enumerate(header_values, start=1):
        header = _clean_cell(raw_header)
        if header:
            indexes[header] = column_index
    missing = [header for header in REQUIRED_PURCHASE_COLUMNS if header not in indexes]
    if missing:
        raise RuntimeError(f"采购汇总表 {PURCHASE_SUMMARY_SHEET} sheet 缺少必需列: {', '.join(missing)}")
    quantity_header = next(
        (header for header in PURCHASE_QUANTITY_COLUMNS if header in indexes),
        None,
    )
    if quantity_header is None:
        raise RuntimeError(
            f"采购汇总表 {PURCHASE_SUMMARY_SHEET} sheet 缺少必需列: "
            f"{PURCHASE_QUANTITY_COLUMNS[0]} 或 {PURCHASE_QUANTITY_COLUMNS[1]}"
        )
    indexes["数量"] = indexes[quantity_header]
    return indexes


def _is_purchase_summary_total_row(row: tuple[Any, ...]) -> bool:
    first_cell = row[0] if row else None
    return _clean_cell(first_cell) == "合计"


def load_purchase_summary_lines(
    purchase_summary_xlsx: str | Path,
) -> OrderedDict[str, list[PurchaseContractLine]]:
    source_path = Path(purchase_summary_xlsx).expanduser()
    if not source_path.is_file():
        raise RuntimeError(f"找不到采购汇总表: {source_path}")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取采购汇总表") from exc

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    try:
        if PURCHASE_SUMMARY_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"采购汇总表缺少 sheet: {PURCHASE_SUMMARY_SHEET}")
        worksheet = workbook[PURCHASE_SUMMARY_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            header_values = ()
        indexes = _header_indexes(tuple(header_values or ()))

        data_rows = [
            (row_number, tuple(row or ()))
            for row_number, row in enumerate(rows, start=2)
            if any(_clean_cell(value) for value in row)
        ]
        grouped: OrderedDict[str, list[PurchaseContractLine]] = OrderedDict()
        for position, (row_number, row) in enumerate(data_rows):
            if _is_purchase_summary_total_row(row):
                if position != len(data_rows) - 1:
                    raise RuntimeError(f"采购汇总表 第{row_number}行 合计行必须是最后一行")
                continue
            values = {
                header: row[indexes[header] - 1] if indexes[header] - 1 < len(row) else None
                for header in REQUIRED_PURCHASE_COLUMNS
            }
            values["数量"] = (
                row[indexes["数量"] - 1]
                if indexes["数量"] - 1 < len(row)
                else None
            )
            manufacturer = _clean_cell(values["厂家"])
            if not manufacturer:
                raise RuntimeError(f"采购汇总表 第{row_number}行 厂家不能为空")
            quantity = _decimal_from_value(values["数量"], field_name="数量", row_number=row_number)
            if quantity <= 0:
                continue
            tax_unit_price = _decimal_from_value(values["原价"], field_name="原价", row_number=row_number)
            tax_amount = _decimal_from_value(values["总价"], field_name="总价", row_number=row_number)
            if purchase_summary._is_zhengfei_manufacturer(manufacturer):
                average_price_value = _row_value_by_header(row, indexes, "均价")
                if not _clean_cell(average_price_value):
                    raise RuntimeError(f"采购汇总表 第{row_number}行 正飞均价不能为空")
                tax_unit_price = _decimal_from_value(
                    average_price_value,
                    field_name="正飞均价",
                    row_number=row_number,
                )
                tax_amount = tax_unit_price * quantity
            line = PurchaseContractLine(
                manufacturer=manufacturer,
                product_name=_clean_cell(values["合同产品名称"]),
                model=_clean_cell(values["型号"]),
                unit=_clean_cell(values["单位"]),
                quantity=quantity,
                tax_unit_price=tax_unit_price,
                tax_amount=tax_amount,
                tax_rate=_clean_cell(values["税率"]),
            )
            grouped.setdefault(manufacturer, []).append(line)
        if not grouped:
            raise RuntimeError("采购汇总表没有可填写合同的厂家明细")
        return grouped
    finally:
        workbook.close()


def _safe_file_stem(value: str) -> str:
    cleaned = INVALID_FILE_NAME_CHARS.sub("_", _clean_cell(value)).strip(". ")
    return (cleaned or "未命名厂家")[:80]


def _resolve_company_sheet(sheet_names: list[str], manufacturer: str) -> tuple[str | None, str | None]:
    exact_matches = [sheet_name for sheet_name in sheet_names if sheet_name == manufacturer]
    if len(exact_matches) == 1:
        return exact_matches[0], None
    if len(exact_matches) > 1:
        return None, f"合同模板中厂家 `{manufacturer}` 精确匹配到多个 sheet，已跳过"

    contains_matches = [sheet_name for sheet_name in sheet_names if manufacturer and manufacturer in sheet_name]
    if len(contains_matches) == 1:
        return contains_matches[0], None
    if not contains_matches:
        return None, f"合同模板中未找到厂家 `{manufacturer}` 对应的 sheet，已跳过"
    return None, f"合同模板中厂家 `{manufacturer}` 匹配到多个 sheet: {', '.join(contains_matches)}，已跳过"


def _merged_range_exists(worksheet: Any, bounds: tuple[int, int, int, int]) -> bool:
    min_col, min_row, max_col, max_row = bounds
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_col == min_col
            and merged_range.min_row == min_row
            and merged_range.max_col == max_col
            and merged_range.max_row == max_row
        ):
            return True
    return False


def _remove_merged_range_registration(worksheet: Any, bounds: tuple[int, int, int, int]) -> None:
    min_col, min_row, max_col, max_row = bounds
    for merged_range in list(worksheet.merged_cells.ranges):
        if (
            merged_range.min_col == min_col
            and merged_range.min_row == min_row
            and merged_range.max_col == max_col
            and merged_range.max_row == max_row
        ):
            worksheet.merged_cells.remove(merged_range)
            return


def _merge_cells_if_missing(worksheet: Any, bounds: tuple[int, int, int, int]) -> None:
    if _merged_range_exists(worksheet, bounds):
        return
    worksheet.merge_cells(
        start_column=bounds[0],
        start_row=bounds[1],
        end_column=bounds[2],
        end_row=bounds[3],
    )


def _merged_range_shifts_for_insert(
    worksheet: Any,
    *,
    row: int,
    amount: int,
) -> list[tuple[tuple[int, int, int, int], tuple[int, int, int, int]]]:
    shifted_bounds: list[tuple[tuple[int, int, int, int], tuple[int, int, int, int]]] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        old_bounds = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        if merged_range.min_row >= row:
            new_bounds = (
                merged_range.min_col,
                merged_range.min_row + amount,
                merged_range.max_col,
                merged_range.max_row + amount,
            )
        elif merged_range.min_row < row <= merged_range.max_row:
            new_bounds = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row + amount,
            )
        else:
            continue
        shifted_bounds.append((old_bounds, new_bounds))
    return shifted_bounds


def _row_dimension_shifts_for_insert(worksheet: Any, *, row: int) -> list[tuple[int, Any]]:
    return [
        (row_index, copy(dimension))
        for row_index, dimension in worksheet.row_dimensions.items()
        if row_index >= row
    ]


def _shift_row_dimensions_after_insert(
    worksheet: Any,
    *,
    row: int,
    amount: int,
    shifted_dimensions: list[tuple[int, Any]],
) -> None:
    for row_index in list(worksheet.row_dimensions.keys()):
        if row_index >= row:
            del worksheet.row_dimensions[row_index]
    for old_row_index, dimension in shifted_dimensions:
        new_row_index = old_row_index + amount
        dimension.index = new_row_index
        worksheet.row_dimensions[new_row_index] = dimension


def _insert_rows_preserving_merged_ranges(worksheet: Any, *, row: int, amount: int) -> None:
    if amount <= 0:
        return
    shifted_bounds = _merged_range_shifts_for_insert(worksheet, row=row, amount=amount)
    shifted_dimensions = _row_dimension_shifts_for_insert(worksheet, row=row)
    worksheet.insert_rows(row, amount=amount)
    _shift_row_dimensions_after_insert(
        worksheet,
        row=row,
        amount=amount,
        shifted_dimensions=shifted_dimensions,
    )
    for old_bounds, _ in shifted_bounds:
        _remove_merged_range_registration(worksheet, old_bounds)
    for _, new_bounds in shifted_bounds:
        _merge_cells_if_missing(worksheet, new_bounds)


def _copy_cell_style(source_cell: Any, target_cell: Any) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)


def _safe_unmerge_single_row_range(worksheet: Any, merged_range: Any) -> None:
    cells = list(merged_range.cells)
    worksheet.merged_cells.remove(merged_range)
    for row, column in cells[1:]:
        worksheet._cells.pop((row, column), None)


def _unmerge_single_row_ranges(worksheet: Any, *, row: int) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row == row and merged_range.max_row == row:
            try:
                _safe_unmerge_single_row_range(worksheet, merged_range)
            except Exception as exc:
                raise RuntimeError(
                    f"{worksheet.title} sheet 模板明细区异常合并单元格，"
                    f"清理第{row}行 {merged_range} 失败: {_exception_text(exc)}"
                ) from exc


def _copy_row_format(worksheet: Any, *, source_row: int, target_row: int) -> None:
    if source_row != target_row:
        _unmerge_single_row_ranges(worksheet, row=target_row)
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column_index in range(1, worksheet.max_column + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        if target_cell.__class__.__name__ == "MergedCell":
            continue
        _copy_cell_style(source_cell, target_cell)

    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row != source_row or merged_range.max_row != source_row:
            continue
        bounds = (
            merged_range.min_col,
            target_row,
            merged_range.max_col,
            target_row,
        )
        if not _merged_range_exists(worksheet, bounds):
            worksheet.merge_cells(
                start_column=bounds[0],
                start_row=bounds[1],
                end_column=bounds[2],
                end_row=bounds[3],
            )


def _single_row_span_for_cell(worksheet: Any, *, row: int, column: int) -> tuple[int, int]:
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row == row
            and merged_range.max_row == row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return merged_range.min_col, merged_range.max_col
    return column, column


def _apply_column_spans(worksheet: Any, *, layout: DetailTableLayout, row: int) -> None:
    for start_column, end_column in layout.header_column_spans.values():
        if end_column <= start_column:
            continue
        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.min_row != row or merged_range.max_row != row:
                continue
            if merged_range.max_col < start_column or merged_range.min_col > end_column:
                continue
            _safe_unmerge_single_row_range(worksheet, merged_range)
        worksheet.merge_cells(
            start_row=row,
            start_column=start_column,
            end_row=row,
            end_column=end_column,
        )


def _writable_cell(worksheet: Any, *, row: int, column: int) -> Any:
    cell = worksheet.cell(row=row, column=column)
    if cell.__class__.__name__ != "MergedCell":
        return cell
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _detail_header_key(value: Any) -> str | None:
    normalized = _normalize_header(value)
    if not normalized:
        return None
    for key, candidates in DETAIL_HEADER_MATCHERS.items():
        if any(candidate in normalized for candidate in candidates):
            return key
    return None


def _find_detail_table_layout(worksheet: Any) -> DetailTableLayout:
    for row in worksheet.iter_rows():
        columns: dict[str, int] = {}
        header_column_spans: dict[str, tuple[int, int]] = {}
        for cell in row:
            key = _detail_header_key(cell.value)
            if key and key not in columns:
                columns[key] = cell.column
                header_column_spans[key] = _single_row_span_for_cell(
                    worksheet,
                    row=cell.row,
                    column=cell.column,
                )
        if all(key in columns for key in REQUIRED_DETAIL_HEADER_KEYS):
            header_row = row[0].row
            detail_start_row = header_row + 1
            summary_row = _find_summary_row(worksheet, start_row=detail_start_row + 1)
            if summary_row <= detail_start_row:
                raise RuntimeError(f"{worksheet.title} sheet 明细表没有可复制的数据行")
            return DetailTableLayout(
                header_row=header_row,
                detail_start_row=detail_start_row,
                summary_row=summary_row,
                columns=columns,
                header_column_spans=header_column_spans,
            )
    raise RuntimeError(f"{worksheet.title} sheet 缺少合同明细表头")


def _find_summary_row(worksheet: Any, *, start_row: int) -> int:
    for row_index in range(start_row, worksheet.max_row + 1):
        for column_index in range(1, worksheet.max_column + 1):
            if "合计" in _clean_cell(worksheet.cell(row=row_index, column=column_index).value):
                return row_index
    raise RuntimeError(f"{worksheet.title} sheet 缺少合计行")


def _set_cell_value(worksheet: Any, *, row: int, column: int, value: Any) -> None:
    _writable_cell(worksheet, row=row, column=column).value = value


def _make_single_row_cell_writable(worksheet: Any, *, row: int, column: int) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if (
            merged_range.min_row != row
            or merged_range.max_row != row
            or not (merged_range.min_col < column <= merged_range.max_col)
        ):
            continue

        start_column = merged_range.min_col
        end_column = merged_range.max_col
        source_cell = worksheet.cell(row=row, column=start_column)
        _safe_unmerge_single_row_range(worksheet, merged_range)
        if start_column < column - 1:
            worksheet.merge_cells(
                start_row=row,
                start_column=start_column,
                end_row=row,
                end_column=column - 1,
            )
        if column + 1 < end_column:
            worksheet.merge_cells(
                start_row=row,
                start_column=column + 1,
                end_row=row,
                end_column=end_column,
            )
        _copy_cell_style(source_cell, worksheet.cell(row=row, column=column))
        return


def _clear_row_values(worksheet: Any, *, row: int) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        cell = _writable_cell(worksheet, row=row, column=column_index)
        if cell.row == row:
            cell.value = None


def _format_date_text(value: date) -> str:
    return f"{value.year}年{value.month}月{value.day}日"


def _apply_info_block_alignment(cell: Any) -> None:
    try:
        from openpyxl.styles import Alignment
    except Exception:
        return
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )


def _extract_contract_number(value: str) -> str:
    match = re.search(
        r"合同编号\s*[:：]\s*(.*?)(?=\s*(?:Date|日期)\s*[:：]|[\r\n]|$)",
        value,
        flags=re.IGNORECASE,
    )
    return _clean_cell(match.group(1)) if match else ""


def _fill_contract_info_block(worksheet: Any, *, contract_date: date) -> bool:
    date_text = _format_date_text(contract_date)
    for row in worksheet.iter_rows():
        for cell in row:
            raw_text = cell.value
            if not isinstance(raw_text, str):
                continue
            if "交货日期" in raw_text:
                continue
            if "合同编号" not in raw_text and "Date" not in raw_text and "日期" not in raw_text:
                continue
            contract_number = _extract_contract_number(raw_text)
            if not contract_number:
                continue
            cell.value = f"合同编号：{contract_number}\nDate：{date_text}"
            _apply_info_block_alignment(cell)
            return True
    return False


def _fill_delivery_payment_tax_block(
    worksheet: Any,
    *,
    delivery_date: date,
    tax_rate_text: str,
) -> bool:
    delivery_date_text = _format_date_text(delivery_date)
    for row in worksheet.iter_rows():
        for cell in row:
            raw_text = cell.value
            if not isinstance(raw_text, str) or "交货日期" not in raw_text:
                continue
            cell.value = (
                f"交货日期：{delivery_date_text}\n"
                "付款期限：发货验收付款\n"
                f"币种：人民币 税率：{tax_rate_text}"
            )
            _apply_info_block_alignment(cell)
            return True
    return False


def _format_tax_rate(value: str) -> str:
    text = _clean_cell(value)
    if not text:
        return ""
    if text.endswith("%"):
        return text
    try:
        rate = Decimal(text)
    except (InvalidOperation, ValueError):
        return text
    if not rate.is_finite():
        return text
    if abs(rate) <= 1:
        rate *= Decimal("100")
    normalized = rate.normalize()
    display = format(normalized, "f").rstrip("0").rstrip(".")
    return f"{display}%"


def _fill_date_and_tax(
    worksheet: Any,
    *,
    contract_date: date,
    delivery_date: date,
    tax_rate: str,
    warnings: list[str],
    manufacturer: str,
) -> None:
    tax_rate_text = _format_tax_rate(tax_rate)

    if not _fill_contract_info_block(worksheet, contract_date=contract_date):
        warnings.append(f"厂家 `{manufacturer}` 合同模板未找到合同日期位置")
    if not tax_rate_text:
        warnings.append(f"厂家 `{manufacturer}` 采购汇总表税率为空")
        return
    if not _fill_delivery_payment_tax_block(
        worksheet,
        delivery_date=delivery_date,
        tax_rate_text=tax_rate_text,
    ):
        warnings.append(f"厂家 `{manufacturer}` 合同模板未找到税率位置")


def _fill_detail_rows(worksheet: Any, lines: list[PurchaseContractLine]) -> int:
    layout = _find_detail_table_layout(worksheet)
    row_count = len(lines)
    available_rows = layout.summary_row - layout.detail_start_row
    if row_count > available_rows:
        missing_rows = row_count - available_rows
        _insert_rows_preserving_merged_ranges(worksheet, row=layout.summary_row, amount=missing_rows)
        for target_row in range(layout.summary_row, layout.summary_row + missing_rows):
            _copy_row_format(worksheet, source_row=layout.detail_start_row, target_row=target_row)
        layout.summary_row += missing_rows
        available_rows = row_count

    for offset, line in enumerate(lines):
        row_index = layout.detail_start_row + offset
        if offset > 0:
            _copy_row_format(worksheet, source_row=layout.detail_start_row, target_row=row_index)
        _apply_column_spans(worksheet, layout=layout, row=row_index)
        values = {
            "sequence": offset + 1,
            "product_name": line.product_name,
            "model": line.model,
            "unit": line.unit,
            "quantity": _decimal_to_cell_value(line.quantity),
            "tax_unit_price": _decimal_to_cell_value(line.tax_unit_price),
            "tax_amount": _decimal_to_cell_value(line.tax_amount),
            "remark": None,
        }
        for key, value in values.items():
            column = layout.columns.get(key)
            if column is not None:
                _set_cell_value(worksheet, row=row_index, column=column, value=value)

    for row_index in range(layout.detail_start_row + row_count, layout.detail_start_row + available_rows):
        _copy_row_format(worksheet, source_row=layout.detail_start_row, target_row=row_index)
        _apply_column_spans(worksheet, layout=layout, row=row_index)
        _clear_row_values(worksheet, row=row_index)

    total_quantity = sum((line.quantity for line in lines), Decimal("0"))
    total_amount = sum((line.tax_amount for line in lines), Decimal("0"))
    _apply_column_spans(worksheet, layout=layout, row=layout.summary_row)
    _make_single_row_cell_writable(
        worksheet,
        row=layout.summary_row,
        column=layout.columns["quantity"],
    )
    _set_cell_value(
        worksheet,
        row=layout.summary_row,
        column=layout.columns["quantity"],
        value=_decimal_to_cell_value(total_quantity),
    )
    _set_cell_value(
        worksheet,
        row=layout.summary_row,
        column=layout.columns["tax_amount"],
        value=_decimal_to_cell_value(total_amount),
    )
    return row_count


def _copy_addendum_sheet(workbook: Any) -> Any:
    if ADDENDUM_TEMPLATE_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"合同汇总模板缺少 sheet: {ADDENDUM_TEMPLATE_SHEET}")
    if ADDENDUM_OUTPUT_SHEET in workbook.sheetnames:
        workbook.remove(workbook[ADDENDUM_OUTPUT_SHEET])
    worksheet = workbook.copy_worksheet(workbook[ADDENDUM_TEMPLATE_SHEET])
    worksheet.title = ADDENDUM_OUTPUT_SHEET
    return worksheet


def _save_single_company_contract(
    *,
    template_xlsx: Path,
    output_dir: Path,
    manufacturer: str,
    sheet_name: str,
    lines: list[PurchaseContractLine],
    contract_date: date,
    warnings: list[str],
) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法填写采购合同") from exc

    workbook = load_workbook(template_xlsx)
    try:
        worksheet = workbook[sheet_name]
        addendum_worksheet = _copy_addendum_sheet(workbook)
        for other_sheet in list(workbook.worksheets):
            if other_sheet.title not in {sheet_name, ADDENDUM_OUTPUT_SHEET}:
                workbook.remove(other_sheet)

        _fill_date_and_tax(
            worksheet,
            contract_date=contract_date,
            delivery_date=contract_date + timedelta(days=3),
            tax_rate=lines[0].tax_rate if lines else "",
            warnings=warnings,
            manufacturer=manufacturer,
        )
        _fill_detail_rows(worksheet, lines)
        _fill_detail_rows(addendum_worksheet, lines)

        output_path = output_dir / f"{_safe_file_stem(manufacturer)}_purchase_contract.xlsx"
        workbook.save(output_path)
        return str(output_path)
    finally:
        workbook.close()


def fill_purchase_contracts(
    *,
    purchase_summary_xlsx: str | Path,
    contract_template_xlsx: str | Path,
    output_dir: str | Path | None = None,
    today: date | None = None,
) -> dict[str, Any]:
    purchase_summary_path = Path(purchase_summary_xlsx).expanduser()
    template_path = Path(contract_template_xlsx).expanduser()
    if not template_path.is_file():
        raise RuntimeError(f"找不到合同汇总模板: {template_path}")

    grouped_lines = load_purchase_summary_lines(purchase_summary_path)
    directory = Path(OUTPUT_DIR if output_dir is None else output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    contract_date = today or date.today()

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法读取合同汇总模板") from exc

    template_workbook = load_workbook(template_path, read_only=True)
    try:
        sheet_names = list(template_workbook.sheetnames)
    finally:
        template_workbook.close()
    if ADDENDUM_TEMPLATE_SHEET not in sheet_names:
        raise RuntimeError(f"合同汇总模板缺少 sheet: {ADDENDUM_TEMPLATE_SHEET}")

    warnings: list[str] = []
    output_files: list[dict[str, str]] = []
    skipped_manufacturers: list[str] = []
    for manufacturer, lines in grouped_lines.items():
        sheet_name, warning = _resolve_company_sheet(sheet_names, manufacturer)
        if warning:
            warnings.append(warning)
            skipped_manufacturers.append(manufacturer)
            continue
        assert sheet_name is not None
        try:
            output_xlsx = _save_single_company_contract(
                template_xlsx=template_path,
                output_dir=directory,
                manufacturer=manufacturer,
                sheet_name=sheet_name,
                lines=lines,
                contract_date=contract_date,
                warnings=warnings,
            )
        except Exception as exc:
            warnings.append(f"厂家 `{manufacturer}` 合同填写失败: {_exception_text(exc)}")
            skipped_manufacturers.append(manufacturer)
            continue
        output_files.append(
            {
                "manufacturer": manufacturer,
                "sheet_name": sheet_name,
                "output_xlsx": output_xlsx,
            }
        )

    return {
        "success": True,
        "purchase_summary_xlsx": str(purchase_summary_path),
        "contract_template_xlsx": str(template_path),
        "output_dir": str(directory),
        "output_files": output_files,
        "generated_count": len(output_files),
        "skipped_manufacturer_count": len(skipped_manufacturers),
        "skipped_manufacturers": skipped_manufacturers,
        "warnings": warnings,
        "source": SOURCE,
    }


def run(arguments: dict[str, Any]) -> dict[str, Any]:
    """lxeskill entrypoint — the catalog input_schema is the argument contract."""
    purchase_summary_xlsx = str(arguments.get("purchase_summary_xlsx") or "")
    contract_template_xlsx = str(arguments.get("contract_template_xlsx") or "")
    try:
        return fill_purchase_contracts(
            purchase_summary_xlsx=purchase_summary_xlsx,
            contract_template_xlsx=contract_template_xlsx,
        )
    except Exception as exc:  # noqa: BLE001 — failure context belongs in the payload
        return {
            "success": False,
            "purchase_summary_xlsx": purchase_summary_xlsx,
            "contract_template_xlsx": contract_template_xlsx,
            "exception": _exception_text(exc),
            "source": SOURCE,
        }
