from __future__ import annotations

from pathlib import Path


def _form_value(call: dict, name: str) -> str:
    for key, value in call.get("data", []):
        if key == name:
            return value
    raise AssertionError(f"missing form field: {name}")


def _form_values(call: dict, name: str) -> list[str]:
    return [value for key, value in call.get("data", []) if key == name]


def _sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _write_mabang_msku(
    path: Path,
    mskus: list[str],
    *,
    site: str = "美国站",
    include_site: bool = True,
) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = "mskulist"
        headers = ["店铺名称", "MSKU"]
        if include_site:
            headers.insert(1, "站点")
        worksheet.append(headers)
        for msku in mskus:
            row = {"店铺名称": "Amazon-Test-US", "站点": site, "MSKU": msku}
            worksheet.append([row.get(header, "") for header in headers])
        workbook.save(path)
    finally:
        workbook.close()
    return path


def _xlsx_bytes(rows: list[dict], *, columns: list[str]) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
