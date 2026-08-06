from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from services.agent_cli.mabang import fill_purchase_contracts as contract_workbook
from services.agent_cli.mabang import generate_fba_restock_workbook as restock_workbook
from services.agent_cli.mabang import generate_restock_workbook as purchase_summary
from services.agent_cli.mabang import regenerate_purchase_files as cli


BATCH_ID = "00000000-0000-0000-0000-000000000001"
REVISION_ID = "00000000-0000-0000-0000-000000000002"
CONTRACT_ID = "00000000-0000-0000-0000-000000000003"
BATCH_NO = "PB20260723-0001"
SP_NO = "SP260710001"


def _write_contract_sheet(worksheet, *, marker: str) -> None:
    worksheet["A1"] = marker
    worksheet["E2"] = "合同编号：OLD\nDate：2000年1月1日"
    worksheet["E3"] = (
        "交货日期：2000年1月4日\n付款期限：发货验收付款\n币种：人民币 税率：0%"
    )
    headers = [
        "序号",
        "产品名称",
        "规格型号",
        "单位",
        "数量",
        "含税单价",
        "含税金额（元）",
        "备注",
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=4, column=column, value=header)
    worksheet.append([None] * len(headers))
    worksheet.append(["合计", None, None, None, None, None, 0, None])


def _write_template(path: Path, suppliers: list[str], *, marker: str = "最新模板") -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for supplier in suppliers:
        _write_contract_sheet(workbook.create_sheet(supplier), marker=marker)
    addendum = workbook.create_sheet(contract_workbook.ADDENDUM_TEMPLATE_SHEET)
    _write_contract_sheet(addendum, marker="附加件")
    addendum["A2"] = "采购合同编号：OLD"
    workbook.save(path)
    workbook.close()


def _planned_line(*, quantity: str = "10") -> dict:
    return {
        "stock_sku": "SKU-A",
        "product_name": "冻结产品A",
        "model": "M-1",
        "planned_shipment_quantity": quantity,
        "supplier_name": "厂家A",
    }


def _allocation(
    *,
    quantity: str,
    source_kind: str,
    sp_no: str = SP_NO,
) -> dict:
    carryover = source_kind == "carryover"
    return {
        "sp_no": sp_no,
        "stock_sku": "SKU-A",
        "source_kind": source_kind,
        "carryover_entry_id": "carryover-1" if carryover else None,
        "quantity": quantity,
        "source_contract_no": "OLD20260101001" if carryover else "",
        "historical_tax_unit_price": "3.75" if carryover else None,
    }


def _purchase_line(
    *,
    purchase_quantity: str = "6",
    carryover_quantity: str = "4",
    allocations: list[dict] | None = None,
) -> dict:
    if allocations is None:
        allocations = [
            _allocation(quantity=purchase_quantity, source_kind="current_purchase"),
            _allocation(quantity=carryover_quantity, source_kind="carryover"),
        ]
    return {
        "supplier_name": "厂家A",
        "tax_rate": "13%",
        "line_ref": "L0001",
        "contract_product_name": "合同产品A",
        "model": "M-1",
        "unit": "个",
        "source_tax_unit_price": "4.00",
        "planned_shipment_quantity": "10",
        "carryover_applied_quantity": carryover_quantity,
        "purchase_quantity": purchase_quantity,
        "inventory_sources": [],
        "applications": [],
        "allocation_details": allocations,
        "contract_no": "HT20260723001" if purchase_quantity != "0" else "",
        "contract_id": CONTRACT_ID if purchase_quantity != "0" else None,
        "tax_unit_price": "4.25" if purchase_quantity != "0" else None,
    }


def _contract() -> dict:
    return {
        "id": CONTRACT_ID,
        "contract_id": CONTRACT_ID,
        "revision_id": REVISION_ID,
        "supplier_name": "厂家A",
        "contract_no": "HT20260723001",
        "contract_date": "2026-07-23",
        "daily_sequence": 1,
        "supplier_contract_sequence": 1,
        "source_kind": "generated",
        "purchase_amount": "25.50",
        "status": "current",
        "tax_rate": "13%",
        "lines": [
            {
                "contract_product_name": "合同产品A",
                "model": "M-1",
                "unit": "个",
                "purchase_quantity": "6",
                "tax_unit_price": "4.25",
            }
        ],
    }


def _snapshot(
    *,
    purchase_line: dict | None = None,
    contracts: list[dict] | None = None,
    sps: list[dict] | None = None,
) -> dict:
    return {
        "snapshot_schema": "lxe.erp.purchase-artifact-snapshot.v1",
        "batch_id": BATCH_ID,
        "batch_no": BATCH_NO,
        "revision_id": REVISION_ID,
        "version_no": 1,
        "status": "current",
        "business_date": "2026-07-23",
        "sps": sps
        or [
            {
                "sp_no": SP_NO,
                "country": "英国",
                "planned_lines": [_planned_line()],
                "unmatched_lines": [],
            }
        ],
        "purchase_lines": [purchase_line or _purchase_line()],
        "contracts": [_contract()] if contracts is None else contracts,
    }


def _fake_snapshot(monkeypatch, snapshot: dict) -> list[str]:
    calls: list[str] = []

    def fake_request(method: str, path: str, **_kwargs):
        calls.append(path)
        assert method == "GET"
        return 200, deepcopy(snapshot)

    monkeypatch.setattr(cli.erp_http, "request_json", fake_request)
    return calls


def _output_dirs(monkeypatch, tmp_path: Path) -> tuple[Path, Path, Path]:
    summary_dir = tmp_path / "purchase-summary"
    restock_dir = tmp_path / "restock"
    contract_dir = tmp_path / "contracts"
    monkeypatch.setattr(purchase_summary, "OUTPUT_DIR", summary_dir)
    monkeypatch.setattr(restock_workbook, "OUTPUT_DIR", restock_dir)
    monkeypatch.setattr(contract_workbook, "OUTPUT_DIR", contract_dir)
    return summary_dir, restock_dir, contract_dir


def _rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [cell.value for cell in worksheet[1]]
        return [
            dict(zip(headers, values, strict=True))
            for values in worksheet.iter_rows(min_row=2, values_only=True)
            if values and values[0] != "合计"
        ]
    finally:
        workbook.close()


def _cell(path: Path, sheet_name: str, coordinate: str):
    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet_name][coordinate].value
    finally:
        workbook.close()


def test_regenerates_summary_restock_and_contract_from_frozen_snapshot(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    summary_dir, _restock_dir, _contract_dir = _output_dirs(monkeypatch, tmp_path)
    snapshot = _snapshot()
    snapshot["sps"][0]["unmatched_lines"] = [
        {
            "stock_sku": "SKU-UNMATCHED",
            "product_name": "未匹配产品",
            "planned_shipment_quantity": "2",
            "issue_code": "export_tax_master_stock_sku_not_found",
            "issue_message": "出口退税总表未找到库存sku",
        }
    ]
    calls = _fake_snapshot(monkeypatch, snapshot)

    result = cli.run(
        {
            "batch_no": " pb20260723-0001 ",
            "contract_template_xlsx": str(template),
        }
    )

    assert result["success"] is True
    assert result["result_schema"] == cli.RESULT_SCHEMA
    assert result["gross_margin"] == "0.3"
    assert result["artifact_summary"] == {
        "delivery_count": 1,
        "restock_count": 1,
        "contract_count": 1,
        "deliverable_file_count": 3,
    }
    assert calls == [
        f"/api/v1/erp/purchase-batches/by-number/{BATCH_NO}/artifact-snapshot"
    ]
    summary_path = Path(result["purchase_summary_xlsx"])
    assert summary_path == summary_dir / f"{SP_NO}_purchase_summary.xlsx"
    summary_rows = _rows(summary_path, purchase_summary.SUMMARY_SHEET_NAME)
    assert [row["本次采购量"] for row in summary_rows] == [6, 0]
    assert [row["留存库存抵扣量"] for row in summary_rows] == [0, 4]
    assert summary_rows[0]["本次采购合同编号"] == "HT20260723001"
    assert summary_rows[0]["产品名称"] == "冻结产品A"
    assert summary_rows[1]["历史库存合同编号"] == "OLD20260101001 × 4"
    unmatched_rows = _rows(summary_path, purchase_summary.UNMATCHED_SHEET_NAME)
    assert unmatched_rows[0]["库存sku"] == "SKU-UNMATCHED"
    assert unmatched_rows[0]["数量"] == 2

    restock_path = Path(result["restock_xlsx_paths"][0])
    restock_rows = _rows(restock_path, restock_workbook.RESTOCK_SHEET_NAME)
    assert [row["毛利率"] for row in restock_rows] == [0.3, 0.3]
    assert [row["本次采购量"] for row in restock_rows] == [6, 0]
    assert [row["留存库存抵扣量"] for row in restock_rows] == [0, 4]
    assert restock_rows[0]["日期"] == "2026-07-23"

    contract_path = Path(result["contract_xlsx_paths"][0])
    assert _cell(contract_path, "厂家A", "A1") == "最新模板"
    assert _cell(contract_path, "厂家A", "E2") == (
        "合同编号：HT20260723001\nDate：2026年7月23日"
    )
    assert _cell(contract_path, "厂家A", "E5") == 6
    assert _cell(contract_path, "厂家A", "F5") == 4.25


def test_regenerates_multiple_supplier_contracts(monkeypatch, tmp_path: Path) -> None:
    snapshot = _snapshot()
    snapshot["sps"][0]["planned_lines"].append(
        {
            "stock_sku": "SKU-B",
            "product_name": "冻结产品B",
            "model": "M-2",
            "planned_shipment_quantity": "2",
            "supplier_name": "厂家B",
        }
    )
    second_line = deepcopy(_purchase_line())
    second_line.update(
        {
            "supplier_name": "厂家B",
            "line_ref": "L0002",
            "contract_product_name": "合同产品B",
            "model": "M-2",
            "planned_shipment_quantity": "2",
            "carryover_applied_quantity": "0",
            "purchase_quantity": "2",
            "contract_no": "HT20260723002",
            "contract_id": "00000000-0000-0000-0000-000000000004",
            "tax_unit_price": "5.00",
            "source_tax_unit_price": "5.00",
            "allocation_details": [
                {
                    "sp_no": SP_NO,
                    "stock_sku": "SKU-B",
                    "source_kind": "current_purchase",
                    "carryover_entry_id": None,
                    "quantity": "2",
                    "source_contract_no": "",
                    "historical_tax_unit_price": None,
                }
            ],
        }
    )
    snapshot["purchase_lines"].append(second_line)
    second_contract = deepcopy(_contract())
    second_contract.update(
        {
            "id": "00000000-0000-0000-0000-000000000004",
            "contract_id": "00000000-0000-0000-0000-000000000004",
            "supplier_name": "厂家B",
            "contract_no": "HT20260723002",
        }
    )
    snapshot["contracts"].append(second_contract)
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A", "厂家B"])
    _output_dirs(monkeypatch, tmp_path)
    _fake_snapshot(monkeypatch, snapshot)

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is True
    assert result["artifact_summary"]["contract_count"] == 2
    assert [contract["supplier_name"] for contract in result["contracts"]] == [
        "厂家A",
        "厂家B",
    ]


def test_regenerates_every_sp_with_explicit_new_margin(monkeypatch, tmp_path: Path) -> None:
    second_sp = "SP260710002"
    sps = [
        {
            "sp_no": SP_NO,
            "country": "英国",
            "planned_lines": [_planned_line(quantity="4")],
            "unmatched_lines": [],
        },
        {
            "sp_no": second_sp,
            "country": "德国",
            "planned_lines": [_planned_line(quantity="6")],
            "unmatched_lines": [],
        },
    ]
    line = _purchase_line(
        purchase_quantity="10",
        carryover_quantity="0",
        allocations=[
            _allocation(quantity="4", source_kind="current_purchase"),
            _allocation(
                quantity="6",
                source_kind="current_purchase",
                sp_no=second_sp,
            ),
        ],
    )
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    _output_dirs(monkeypatch, tmp_path)
    _fake_snapshot(monkeypatch, _snapshot(purchase_line=line, sps=sps))

    result = cli.run(
        {
            "batch_no": BATCH_NO,
            "gross_margin": "0.35",
            "contract_template_xlsx": str(template),
        }
    )

    assert result["success"] is True
    assert result["gross_margin"] == "0.35"
    assert len(result["restock_xlsx_paths"]) == 2
    for path in result["restock_xlsx_paths"]:
        assert _rows(Path(path), restock_workbook.RESTOCK_SHEET_NAME)[0]["毛利率"] == 0.35


def test_current_batch_without_contracts_still_regenerates_summary_and_restock(
    monkeypatch,
    tmp_path: Path,
) -> None:
    carryover_only = _purchase_line(
        purchase_quantity="0",
        carryover_quantity="10",
        allocations=[_allocation(quantity="10", source_kind="carryover")],
    )
    _output_dirs(monkeypatch, tmp_path)
    _fake_snapshot(
        monkeypatch,
        _snapshot(purchase_line=carryover_only, contracts=[]),
    )

    result = cli.run(
        {
            "batch_no": BATCH_NO,
            "contract_template_xlsx": str(tmp_path / "not-needed.xlsx"),
        }
    )

    assert result["success"] is True
    assert result["artifact_summary"]["contract_count"] == 0
    assert result["contract_xlsx_paths"] == []
    assert Path(result["purchase_summary_xlsx"]).is_file()
    assert len(result["restock_xlsx_paths"]) == 1


def test_rejects_snapshot_contract_from_another_revision_before_writing(
    monkeypatch,
    tmp_path: Path,
) -> None:
    snapshot = _snapshot()
    snapshot["contracts"][0]["revision_id"] = "another-revision"
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    summary_dir, _restock_dir, _contract_dir = _output_dirs(monkeypatch, tmp_path)
    _fake_snapshot(monkeypatch, snapshot)

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "purchase_batch_artifact_snapshot_invalid"
    assert not summary_dir.exists()


def test_atomic_summary_overwrite_keeps_old_file_when_windows_rejects_replace(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    summary_dir, _restock_dir, _contract_dir = _output_dirs(monkeypatch, tmp_path)
    summary_dir.mkdir()
    summary_path = summary_dir / f"{SP_NO}_purchase_summary.xlsx"
    original = b"original summary"
    summary_path.write_bytes(original)
    _fake_snapshot(monkeypatch, _snapshot())

    def locked_replace(_source, _target):
        raise PermissionError("采购汇总被 Excel 占用")

    monkeypatch.setattr(purchase_summary.os, "replace", locked_replace)

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "purchase_files_regeneration_failed"
    assert "采购汇总被 Excel 占用" in result["error"]["message"]
    assert summary_path.read_bytes() == original
    assert not list(summary_dir.glob(".*.part"))


def test_locked_restock_reports_already_regenerated_summary(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    summary_dir, restock_dir, _contract_dir = _output_dirs(monkeypatch, tmp_path)
    restock_dir.mkdir()
    restock_path = restock_dir / "7.23-SP260710001-新棱镜备货-英国.xlsx"
    original = b"original restock"
    restock_path.write_bytes(original)
    _fake_snapshot(monkeypatch, _snapshot())
    real_replace = purchase_summary.os.replace

    def locked_restock_replace(source, target):
        if Path(target).parent == restock_dir:
            raise PermissionError("备货单被 Excel 占用")
        return real_replace(source, target)

    monkeypatch.setattr(purchase_summary.os, "replace", locked_restock_replace)

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["purchase_summary_xlsx"] == str(
        summary_dir / f"{SP_NO}_purchase_summary.xlsx"
    )
    assert result["restock_xlsx_paths"] == []
    assert restock_path.read_bytes() == original
    assert not list(restock_dir.glob(".*.part"))
