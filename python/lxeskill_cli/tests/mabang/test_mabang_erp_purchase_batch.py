from __future__ import annotations

import json
from copy import deepcopy
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pytest
from openpyxl import Workbook, load_workbook

from services.agent_cli.mabang import erp_purchase_batch as erp
from services.agent_cli.mabang import fill_purchase_contracts as contract_cli
from services.agent_cli.mabang import generate_purchase_batch_workbooks as cli


def _write_delivery_csv(
    path: Path,
    *,
    msku_quantity: str = "5",
    sku_quantity: str = "10",
) -> None:
    path.write_text(
        "\n".join(
            [
                '"发货单号","MSKU","MSKU发货量","SKU发货量","国家","备注"',
                f'"SP260710001","MSKU-X","{msku_quantity}","SKU-A × {sku_quantity}","德国",""',
            ]
        ),
        encoding="utf-8-sig",
    )


def _write_master(path: Path) -> None:
    workbook = Workbook()
    sku = workbook.active
    sku.title = "SKU表"
    sku.append(["库存sku", "产品名称", "型号", "原价", "厂家", "备用厂家"])
    sku.append(["SKU-A", "产品A", "A-1", 3.5, "深圳正飞科技", ""])
    contracts = workbook.create_sheet("供应商合同信息")
    contracts.append(["供货方", "单位", "合同产品名称", "合同编号前缀", "税率"])
    contracts.append(["深圳正飞科技", "个", "合同产品A", "ZF", "13%"])
    workbook.save(path)


def _fixture_inputs(tmp_path: Path, *, msku_quantity: str = "5", sku_quantity: str = "10") -> tuple[Path, Path]:
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    _write_delivery_csv(
        csv_dir / "SP260710001_1.csv",
        msku_quantity=msku_quantity,
        sku_quantity=sku_quantity,
    )
    master = tmp_path / "master.xlsx"
    _write_master(master)
    return csv_dir, master


def _fixture_mixed_inputs(tmp_path: Path) -> tuple[Path, Path]:
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    (csv_dir / "SP260710001_1.csv").write_text(
        "\n".join(
            [
                '"发货单号","MSKU","MSKU发货量","SKU发货量","品名","国家","备注"',
                '"SP260710001","MSKU-X","5","SKU-A × 10\nSKU-B × 5","未匹配组合品名","德国",""',
            ]
        ),
        encoding="utf-8-sig",
    )
    master = tmp_path / "master.xlsx"
    _write_master(master)
    return csv_dir, master


def _erp_result() -> dict[str, Any]:
    return {
        "status": "created",
        "batch_id": "00000000-0000-0000-0000-000000000001",
        "batch_no": "PB20260723-0001",
        "revision_id": "00000000-0000-0000-0000-000000000006",
        "version_no": 1,
        "sp_nos": ["SP260710001"],
        "unmatched_stock_sku_count": 0,
        "unmatched_component_count": 0,
        "contracts": [
            {
                "contract_id": "00000000-0000-0000-0000-000000000002",
                "supplier_name": "深圳正飞科技",
                "contract_no": "ZF202607230001",
                "daily_sequence": 1,
                "supplier_contract_sequence": 223,
                "supplier_contract_count": 223,
            }
        ],
        "purchase_lines": [
            {
                "supplier_name": "深圳正飞科技",
                "tax_rate": "13%",
                "line_ref": "L0001",
                "contract_product_name": "合同产品A",
                "model": "A-1",
                "unit": "个",
                "source_tax_unit_price": 3.5,
                "planned_shipment_quantity": 10,
                "carryover_applied_quantity": 4,
                "purchase_quantity": 6,
                "contract_no": "ZF202607230001",
                "contract_id": "00000000-0000-0000-0000-000000000002",
                "tax_unit_price": 3.5,
                "inventory_sources": [
                    {
                        "carryover_entry_id": "00000000-0000-0000-0000-000000000003",
                        "source_contract_no": "ZF20260601001",
                        "historical_tax_unit_price": 3,
                        "source_reference": "SP-OLD",
                        "acquired_on": "2026-06-01",
                        "available_quantity": 4,
                        "suggested_applied_quantity": 4,
                        "version_no": 1,
                    }
                ],
                "applications": [
                    {
                        "carryover_entry_id": "00000000-0000-0000-0000-000000000003",
                        "applied_quantity": 4,
                        "source_contract_no": "ZF20260601001",
                        "historical_tax_unit_price": 3,
                        "source_reference": "SP-OLD",
                        "acquired_on": "2026-06-01",
                        "version_no": 1,
                    }
                ],
                "allocation_details": [
                    {
                        "sp_no": "SP260710001",
                        "stock_sku": "SKU-A",
                        "source_kind": "carryover",
                        "carryover_entry_id": "00000000-0000-0000-0000-000000000003",
                        "quantity": 4,
                        "source_contract_no": "ZF20260601001",
                        "historical_tax_unit_price": 3,
                    },
                    {
                        "sp_no": "SP260710001",
                        "stock_sku": "SKU-A",
                        "source_kind": "current_purchase",
                        "carryover_entry_id": None,
                        "quantity": 6,
                        "source_contract_no": "",
                        "historical_tax_unit_price": None,
                    },
                ],
            }
        ],
    }


def _request_payload() -> dict[str, Any]:
    return {
        "request_id": "purchase-1",
        "source_sha256": "a" * 64,
        "suppliers": [{"name": "深圳正飞科技", "contract_prefix": "ZF"}],
        "sps": [
            {
                "sp_no": "SP260710001",
                "planned_lines": [
                    {
                        "stock_sku": "SKU-A",
                        "product_name": "产品A",
                        "model": "A-1",
                        "supplier_name": "深圳正飞科技",
                        "planned_shipment_quantity": "10",
                    }
                ],
            }
        ],
        "contracts": [
            {
                "supplier_name": "深圳正飞科技",
                "tax_rate": "13%",
                "lines": [
                    {
                        "line_ref": "L0001",
                        "contract_product_name": "合同产品A",
                        "model": "A-1",
                        "unit": "个",
                        "source_tax_unit_price": "3.5",
                        "planned_shipment_quantity": "10",
                        "allocations": [
                            {
                                "sp_no": "SP260710001",
                                "stock_sku": "SKU-A",
                                "planned_quantity": "10",
                            }
                        ],
                    }
                ],
            }
        ],
    }


def _quote_response(*, status: str = "confirmation_required") -> dict[str, Any]:
    error = {
        "code": (
            "purchase_inventory_quote_stale"
            if status == "quote_stale"
            else "purchase_inventory_confirmation_required"
        ),
        "message": (
            "the previous inventory quote is stale; review the latest inventory "
            "deduction proposal. No inventory changes have been committed."
            if status == "quote_stale"
            else (
                "ERP prepared an inventory deduction proposal. No inventory changes "
                "have been committed; confirmation is required."
            )
        ),
    }
    confirmation = {
        "kind": "inventory_quote",
        "quote_id": "00000000-0000-0000-0000-000000000004",
        "inventory_changes_committed": False,
        "planned_shipment_quantity": 10,
        "proposed_inventory_deduction_quantity": 4,
        "proposed_purchase_quantity": 6,
        "all_line_count": 1,
        "affected_line_count": 1,
        "omitted_unaffected_line_count": 0,
        "omitted_unused_inventory_source_count": 0,
        "affected_lines": [
            {
                "supplier_name": "深圳正飞科技",
                "model": "A-1",
                "contract_product_name": "合同产品A",
                "planned_shipment_quantity": 10,
                "proposed_inventory_deduction_quantity": 4,
                "proposed_purchase_quantity": 6,
                "inventory_sources": [
                    {
                        "carryover_entry_id": "00000000-0000-0000-0000-000000000003",
                        "source_kind": "opening_inventory",
                        "source_contract_no": "ZF20260601001",
                        "source_sp_no": "SP260601001",
                        "source_reference": "SP260601001-row-1",
                        "historical_tax_unit_price": 3,
                        "original_quantity": 4,
                        "current_remaining_quantity": 4,
                        "replacement_released_quantity": 0,
                        "available_after_release": 4,
                        "proposed_applied_quantity": 4,
                    }
                ],
            }
        ],
    }
    return {
        "response_schema": "lxe.erp.purchase-confirmation.v2",
        "status": status,
        "request_id": "purchase-1",
        "error": error,
        "confirmation": confirmation,
    }


def test_build_intent_uploads_exact_unit_components_without_declared_quantity(tmp_path: Path) -> None:
    csv_dir, master = _fixture_inputs(tmp_path)

    payload, context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )

    assert context["delivery_nos"] == ["SP260710001"]
    assert payload["sps"][0]["mskus"] == [
        {
            "msku": "MSKU-X",
            "components": [
                {
                    "stock_sku": "SKU-A",
                    "tracking_mode": "tracked",
                    "quantity_per_msku": "2",
                }
            ],
        }
    ]
    assert "declared_ship_quantity" not in payload["sps"][0]["mskus"][0]
    assert payload["sps"][0]["unmatched_lines"] == []
    assert payload["sps"][0]["planned_lines"][0]["planned_shipment_quantity"] == "10"
    line = payload["contracts"][0]["lines"][0]
    assert line["source_tax_unit_price"] == "3.5"
    assert "purchase_quantity" not in line
    assert "carryover_applied_quantity" not in line
    assert payload["request_id"].startswith("purchase-")


def test_build_intent_preserves_unmatched_components_without_planning_them(
    tmp_path: Path,
) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)

    payload, context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )

    assert [line["stock_sku"] for line in payload["sps"][0]["planned_lines"]] == [
        "SKU-A"
    ]
    assert payload["contracts"][0]["lines"][0]["planned_shipment_quantity"] == "10"
    assert payload["sps"][0]["mskus"] == [
        {
            "msku": "MSKU-X",
            "components": [
                {
                    "stock_sku": "SKU-A",
                    "tracking_mode": "tracked",
                    "quantity_per_msku": "2",
                },
                {
                    "stock_sku": "SKU-B",
                    "tracking_mode": "unmatched",
                    "quantity_per_msku": "1",
                },
            ],
        }
    ]
    assert payload["sps"][0]["unmatched_lines"] == [
        {
            "stock_sku": "SKU-B",
            "product_name": "未匹配组合品名",
            "planned_shipment_quantity": "5",
            "issue_code": "export_tax_master_stock_sku_not_found",
        }
    ]
    assert context["unmatched_summary"] == {
        "stock_sku_count": 1,
        "sp_sku_count": 1,
        "component_count": 1,
        "planned_shipment_quantity": "5",
        "items": [
            {
                "sp_no": "SP260710001",
                "stock_sku": "SKU-B",
                "product_name": "未匹配组合品名",
                "planned_shipment_quantity": "5",
                "affected_mskus": [
                    {"msku": "MSKU-X", "quantity_per_msku": "1"}
                ],
            }
        ],
    }
    assert context["confirm_unmatched_sku_token"].startswith("unmatched-")


def test_build_intent_deduplicates_delivery_names_for_unmatched_sp_sku(
    tmp_path: Path,
) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    (csv_dir / "SP260710001_1.csv").write_text(
        "\n".join(
            [
                '"发货单号","MSKU","MSKU发货量","SKU发货量","品名","国家","备注"',
                '"SP260710001","MSKU-X","3","SKU-A × 6\nSKU-B × 3","组合品名A","德国",""',
                '"SP260710001","MSKU-Y","2","SKU-A × 4\nSKU-B × 2","组合品名B","德国",""',
                '"SP260710001","MSKU-Z","1","SKU-A × 2\nSKU-B × 1","组合品名A","德国",""',
            ]
        ),
        encoding="utf-8-sig",
    )

    payload, context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )

    line = payload["sps"][0]["unmatched_lines"][0]
    assert line["product_name"] == "组合品名A\n组合品名B"
    assert line["planned_shipment_quantity"] == "6"
    assert context["unmatched_summary"]["items"][0]["product_name"] == (
        "组合品名A\n组合品名B"
    )


def test_build_intent_allows_blank_unmatched_product_name(tmp_path: Path) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    (csv_dir / "SP260710001_1.csv").write_text(
        "\n".join(
            [
                '"发货单号","MSKU","MSKU发货量","SKU发货量","国家","备注"',
                '"SP260710001","MSKU-X","5","SKU-A × 10\nSKU-B × 5","德国",""',
            ]
        ),
        encoding="utf-8-sig",
    )

    payload, context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )

    assert payload["sps"][0]["unmatched_lines"][0]["product_name"] == ""
    assert context["unmatched_summary"]["items"][0]["product_name"] == ""


def test_build_intent_keeps_unmatched_lines_separate_by_sp(tmp_path: Path) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    (csv_dir / "SP260710002_1.csv").write_text(
        "\n".join(
            [
                '"发货单号","MSKU","MSKU发货量","SKU发货量","品名","国家","备注"',
                '"SP260710002","MSKU-Y","2","SKU-A × 4\nSKU-B × 2","第二个SP品名","法国",""',
            ]
        ),
        encoding="utf-8-sig",
    )

    payload, context = erp.build_purchase_intent(
        ["SP260710001", "SP260710002"],
        master_xlsx=master,
        csv_dir=csv_dir,
    )

    assert [sp["unmatched_lines"] for sp in payload["sps"]] == [
        [
            {
                "stock_sku": "SKU-B",
                "product_name": "未匹配组合品名",
                "planned_shipment_quantity": "5",
                "issue_code": "export_tax_master_stock_sku_not_found",
            }
        ],
        [
            {
                "stock_sku": "SKU-B",
                "product_name": "第二个SP品名",
                "planned_shipment_quantity": "2",
                "issue_code": "export_tax_master_stock_sku_not_found",
            }
        ],
    ]
    assert [item["sp_no"] for item in context["unmatched_summary"]["items"]] == [
        "SP260710001",
        "SP260710002",
    ]


def test_build_intent_rejects_batch_with_only_unmatched_skus(tmp_path: Path) -> None:
    csv_dir, master = _fixture_inputs(tmp_path)
    csv_path = csv_dir / "SP260710001_1.csv"
    csv_path.write_text(
        csv_path.read_text(encoding="utf-8-sig").replace("SKU-A", "SKU-B"),
        encoding="utf-8-sig",
    )

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.build_purchase_intent(
            ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
        )

    assert captured.value.code == "purchase_intent_no_tracked_stock_sku"
    assert captured.value.detail["unmatched_summary"]["stock_sku_count"] == 1


def test_formal_workbooks_keep_existing_unmatched_sheets(tmp_path: Path) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    request_payload, _context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    response = _erp_result()
    response["unmatched_stock_sku_count"] = 1
    response["unmatched_component_count"] = 1
    erp.validate_purchase_response(
        status_code=201,
        response=response,
        request_payload=request_payload,
    )
    generated = cli.generate_purchase_batch_workbooks(
        ["SP260710001"],
        master_xlsx=master,
        gross_margin="0.3",
        csv_dir=csv_dir,
        purchase_output_dir=tmp_path / "purchase",
        restock_output_dir=tmp_path / "restock",
    )

    formal = erp.apply_formal_erp_result(
        generated,
        response,
        request_payload=request_payload,
    )

    for path in [
        formal["purchase_summary_xlsx"],
        *formal["restock_xlsx_paths"],
    ]:
        workbook = load_workbook(path, data_only=True)
        try:
            assert "未匹配" in workbook.sheetnames
            sheet = workbook["未匹配"]
            assert sheet["A2"].value == "SKU-B"
            headers = [cell.value for cell in sheet[1]]
            product_name_column = headers.index("品名") + 1
            quantity_column = headers.index("数量") + 1
            assert sheet.cell(2, product_name_column).value == "未匹配组合品名"
            assert sheet.cell(2, quantity_column).value == 5
        finally:
            workbook.close()


def test_build_intent_rejects_non_integer_unit_component(tmp_path: Path) -> None:
    csv_dir, master = _fixture_inputs(tmp_path, msku_quantity="4", sku_quantity="10")

    with pytest.raises(erp.PurchaseBatchClientError, match="无法推导整数 quantity_per_msku"):
        erp.build_purchase_intent(["SP260710001"], master_xlsx=master, csv_dir=csv_dir)


@pytest.mark.parametrize("tax_rate", ["", "not-a-tax-rate", "NaN%"])
def test_build_intent_rejects_missing_or_invalid_tax_rate_before_http(
    tmp_path: Path,
    tax_rate: str,
) -> None:
    csv_dir, master = _fixture_inputs(tmp_path)
    workbook = load_workbook(master)
    workbook["供应商合同信息"]["E2"] = tax_rate
    workbook.save(master)
    workbook.close()

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.build_purchase_intent(
            ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
        )

    assert captured.value.code == "purchase_intent_invalid"
    assert "税率" in str(captured.value)


def test_confirmation_response_does_not_generate_files(monkeypatch) -> None:
    payload = _request_payload()
    response = _quote_response()
    response["server_optional"] = {"trace": "preserved"}
    response["confirmation"]["optional_note"] = "preserved"
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (
            payload,
            {"delivery_nos": ["SP260710001"], "csv_paths": [], "master_xlsx": "master.xlsx"},
        ),
    )
    monkeypatch.setattr(erp, "import_purchase_intent", lambda _payload: (409, response))
    monkeypatch.setattr(contract_cli, "validate_contract_template", lambda *args, **kwargs: {})
    monkeypatch.setattr(
        cli,
        "generate_purchase_batch_workbooks",
        lambda *args, **kwargs: pytest.fail("confirmation must not generate files"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
        }
    )

    assert result["success"] is False
    assert result["status"] == "confirmation_required"
    assert result["error"]["code"] == "purchase_inventory_confirmation_required"
    assert "detail" not in result["error"]
    assert result["erp"] == response
    assert result["erp"]["server_optional"] == {"trace": "preserved"}
    assert result["erp"]["confirmation"]["optional_note"] == "preserved"


def test_unmatched_sku_requires_local_confirmation_before_template_or_http(
    monkeypatch,
    tmp_path: Path,
) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    payload, context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (payload, context),
    )
    monkeypatch.setattr(
        contract_cli,
        "validate_contract_template",
        lambda *args, **kwargs: pytest.fail("template validation must wait for confirmation"),
    )
    monkeypatch.setattr(
        erp,
        "import_purchase_intent",
        lambda _payload: pytest.fail("ERP must not be called before confirmation"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": str(master),
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
        }
    )

    assert result["status"] == "confirmation_required"
    assert result["error"]["code"] == "purchase_unmatched_sku_confirmation_required"
    assert result["confirmation"]["token"] == context["confirm_unmatched_sku_token"]
    assert result["confirmation"]["items"][0]["stock_sku"] == "SKU-B"


def test_confirmed_unmatched_sku_continues_to_erp_and_preserves_token(
    monkeypatch,
    tmp_path: Path,
) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    payload, context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    response = _quote_response()
    response["request_id"] = payload["request_id"]
    calls: list[dict[str, Any]] = []
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (payload, context),
    )
    monkeypatch.setattr(contract_cli, "validate_contract_template", lambda *args, **kwargs: {})
    monkeypatch.setattr(
        erp,
        "import_purchase_intent",
        lambda request: (calls.append(request) or (409, response)),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": str(master),
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
            "confirm_unmatched_sku_token": context["confirm_unmatched_sku_token"],
        }
    )

    assert calls == [payload]
    assert result["error"]["code"] == "purchase_inventory_confirmation_required"
    assert result["confirm_unmatched_sku_token"] == context[
        "confirm_unmatched_sku_token"
    ]
    assert result["unmatched_summary"]["planned_shipment_quantity"] == "5"


def test_unmatched_confirmation_token_becomes_stale_when_inputs_change(
    monkeypatch,
    tmp_path: Path,
) -> None:
    csv_dir, master = _fixture_mixed_inputs(tmp_path)
    _payload, original_context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    csv_path = csv_dir / "SP260710001_1.csv"
    csv_path.write_text(
        csv_path.read_text(encoding="utf-8-sig").replace("SKU-B × 5", "SKU-B × 10"),
        encoding="utf-8-sig",
    )
    changed_payload, changed_context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    assert changed_context["confirm_unmatched_sku_token"] != original_context[
        "confirm_unmatched_sku_token"
    ]
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (changed_payload, changed_context),
    )
    monkeypatch.setattr(
        erp,
        "import_purchase_intent",
        lambda _payload: pytest.fail("stale confirmation must not call ERP"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": str(master),
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
            "confirm_unmatched_sku_token": original_context[
                "confirm_unmatched_sku_token"
            ],
        }
    )

    assert result["status"] == "confirmation_required"
    assert result["error"]["code"] == "purchase_unmatched_sku_confirmation_stale"
    assert result["confirmation"]["token"] == changed_context[
        "confirm_unmatched_sku_token"
    ]


def test_stale_quote_passes_through_latest_confirmation(monkeypatch) -> None:
    payload = _request_payload()
    response = _quote_response(status="quote_stale")
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (
            payload,
            {"delivery_nos": ["SP260710001"], "csv_paths": [], "master_xlsx": "master.xlsx"},
        ),
    )
    monkeypatch.setattr(erp, "import_purchase_intent", lambda _payload: (409, response))
    monkeypatch.setattr(contract_cli, "validate_contract_template", lambda *args, **kwargs: {})

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
        }
    )

    assert result["success"] is False
    assert result["status"] == "quote_stale"
    assert result["error"]["code"] == "purchase_inventory_quote_stale"
    assert "detail" not in result["error"]
    assert result["erp"] == response
    assert result["erp"]["confirmation"]["quote_id"] == response["confirmation"]["quote_id"]
    assert result["erp"]["confirmation"]["affected_lines"][0]["model"] == "A-1"


def test_replace_confirmation_passes_through_server_response() -> None:
    response = {
        "response_schema": "lxe.erp.purchase-confirmation.v2",
        "status": "confirmation_required",
        "request_id": "purchase-1",
        "error": {
            "code": "purchase_batch_replace_confirmation_required",
            "message": "one or more SP numbers already belong to a current batch",
        },
        "confirmation": {
            "kind": "batch_replacement",
            "conflicts": [
                {
                    "sp_no": "SP260710001",
                    "batch_id": "00000000-0000-0000-0000-000000000010",
                    "batch_no": "PB20260723-0002",
                    "version_no": 3,
                }
            ],
        },
    }

    erp.validate_purchase_response(
        status_code=409,
        response=response,
        request_payload=_request_payload(),
    )
    result = erp.confirmation_result(
        response=response,
        status_code=409,
        request_payload=_request_payload(),
    )

    assert result["erp"] == response
    assert "detail" not in result["error"]


@pytest.mark.parametrize(
    ("mutate", "expected_code"),
    [
        (
            lambda response: response.__setitem__(
                "response_schema", "lxe.erp.purchase-confirmation.v3"
            ),
            "erp_purchase_confirmation_schema_unsupported",
        ),
        (
            lambda response: response.__setitem__("request_id", "purchase-other"),
            "erp_purchase_result_incomplete",
        ),
        (
            lambda response: response["confirmation"].__setitem__(
                "proposed_purchase_quantity", 7
            ),
            "erp_purchase_result_incomplete",
        ),
        (
            lambda response: response["confirmation"]["affected_lines"][0][
                "inventory_sources"
            ][0].__setitem__("proposed_applied_quantity", 3),
            "erp_purchase_result_incomplete",
        ),
        (
            lambda response: response["confirmation"].__setitem__(
                "omitted_unaffected_line_count", 1
            ),
            "erp_purchase_result_incomplete",
        ),
    ],
    ids=[
        "unknown-schema",
        "request-id-mismatch",
        "total-invariant",
        "inventory-source-total",
        "omitted-line-count",
    ],
)
def test_confirmation_response_fails_closed_on_invalid_contract(
    mutate, expected_code: str
) -> None:
    response = _quote_response()
    mutate(response)

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.validate_purchase_response(
            status_code=409,
            response=response,
            request_payload=_request_payload(),
        )

    assert captured.value.code == expected_code


def test_draft_never_calls_erp_and_marks_outputs(monkeypatch, tmp_path: Path) -> None:
    purchase_path = tmp_path / "purchase.xlsx"
    restock_path = tmp_path / "restock.xlsx"
    for path, sheet_name in ((purchase_path, "采购汇总"), (restock_path, "备货单")):
        workbook = Workbook()
        workbook.active.title = sheet_name
        columns = (
            erp.purchase_summary.MANUFACTURER_COLUMNS
            if path == purchase_path
            else ("数量",)
        )
        workbook.active.append(list(columns))
        row = [""] * len(columns)
        row[columns.index("数量")] = 10
        workbook.active.append(row)
        workbook.save(path)
    generated = {
        "success": True,
        "delivery_nos": ["SP260710001"],
        "purchase_summary_xlsx": str(purchase_path),
        "restock_xlsx_paths": [str(restock_path)],
        "restock_outputs": [{"delivery_no": "SP260710001", "output_xlsx": str(restock_path)}],
    }
    monkeypatch.setattr(cli, "generate_purchase_batch_workbooks", lambda *args, **kwargs: dict(generated))
    monkeypatch.setattr(
        erp,
        "import_purchase_intent",
        lambda _payload: pytest.fail("draft must not call ERP"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "gross_margin": "0.3",
            "draft": True,
        }
    )

    assert result["success"] is True
    assert result["mode"] == "draft"
    assert result["erp_synced"] is False
    assert Path(result["purchase_summary_xlsx"]).name == "purchase-DRAFT.xlsx"
    workbook = load_workbook(result["purchase_summary_xlsx"], read_only=True)
    try:
        assert workbook.sheetnames[0] == "草稿-未同步ERP"
        sheet = workbook["采购汇总"]
        headers = [cell.value for cell in sheet[1]]
        assert "合同编号前缀" not in headers
        assert "本次采购合同编号" in headers
        assert "历史库存合同编号" in headers
        assert sheet.cell(2, headers.index("本次采购合同编号") + 1).value is None
        assert sheet.cell(2, headers.index("历史库存合同编号") + 1).value is None
        assert sheet.cell(2, headers.index("计划发货量") + 1).value == 10
        assert sheet.cell(2, headers.index("本次采购量") + 1).value == 10
        assert sheet.cell(2, headers.index("留存库存抵扣量") + 1).value == 0
    finally:
        workbook.close()


def test_formal_workbooks_split_and_move_yellow_inventory_rows_to_end(tmp_path: Path) -> None:
    csv_dir, master = _fixture_inputs(tmp_path)
    generated = cli.generate_purchase_batch_workbooks(
        ["SP260710001"],
        master_xlsx=master,
        gross_margin="0.3",
        csv_dir=csv_dir,
        purchase_output_dir=tmp_path / "purchase",
        restock_output_dir=tmp_path / "restock",
    )

    request_payload, _context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    erp.validate_purchase_response(
        status_code=201,
        response=_erp_result(),
        request_payload=request_payload,
    )
    result = erp.apply_formal_erp_result(
        generated,
        _erp_result(),
        request_payload=request_payload,
    )

    assert result["success"] is True
    assert result["mode"] == "formal"
    summary = load_workbook(result["purchase_summary_xlsx"], data_only=True)
    try:
        sheet = summary["采购汇总"]
        headers = [cell.value for cell in sheet[1]]
        assert headers[headers.index("计划发货量") : headers.index("计划发货量") + 3] == [
            "计划发货量",
            "本次采购量",
            "留存库存抵扣量",
        ]
        planned_col = headers.index("计划发货量") + 1
        purchase_col = headers.index("本次采购量") + 1
        carryover_col = headers.index("留存库存抵扣量") + 1
        assert (
            sheet.cell(2, planned_col).value,
            sheet.cell(2, purchase_col).value,
            sheet.cell(2, carryover_col).value,
        ) == (6, 6, 0)
        assert (
            sheet.cell(3, planned_col).value,
            sheet.cell(3, purchase_col).value,
            sheet.cell(3, carryover_col).value,
        ) == (4, 0, 4)
        assert (
            sheet.cell(4, planned_col).value,
            sheet.cell(4, purchase_col).value,
            sheet.cell(4, carryover_col).value,
        ) == (10, 6, 4)
        assert "合同编号前缀" not in headers
        assert sheet.cell(2, headers.index("本次采购合同编号") + 1).value == "ZF202607230001"
        assert sheet.cell(2, headers.index("历史库存合同编号") + 1).value is None
        assert sheet.cell(3, headers.index("本次采购合同编号") + 1).value is None
        assert sheet.cell(3, headers.index("历史库存合同编号") + 1).value == "ZF20260601001"
        assert sheet.cell(2, headers.index("库存sku") + 1).value == "SKU-A × 6"
        assert sheet.cell(3, headers.index("库存sku") + 1).value == "SKU-A × 4"
        assert sheet.cell(2, headers.index("原价") + 1).value == 3.5
        assert sheet.cell(3, headers.index("原价") + 1).value == 3
        assert sheet.cell(2, headers.index("总价") + 1).value == 21
        assert sheet.cell(3, headers.index("总价") + 1).value == 12
        assert sheet.cell(4, headers.index("总价") + 1).value == 33
        assert all(cell.fill.fgColor.rgb == "FFFFFF00" for cell in sheet[3])
        assert any(cell.fill.fgColor.rgb != "FFFFFF00" for cell in sheet[2])
    finally:
        summary.close()

    restock = load_workbook(result["restock_xlsx_paths"][0], data_only=True)
    try:
        sheet = restock["备货单"]
        headers = [cell.value for cell in sheet[1]]
        order_col = headers.index("采购订单号") + 1
        planned_col = headers.index("计划发货量") + 1
        purchase_col = headers.index("本次采购量") + 1
        carryover_col = headers.index("留存库存抵扣量") + 1
        assert sheet.cell(2, order_col).value == "ZF202607230001"
        assert (sheet.cell(2, planned_col).value, sheet.cell(2, purchase_col).value, sheet.cell(2, carryover_col).value) == (6, 6, 0)
        assert sheet.cell(3, order_col).value == "ZF20260601001"
        assert (sheet.cell(3, planned_col).value, sheet.cell(3, purchase_col).value, sheet.cell(3, carryover_col).value) == (4, 0, 4)
        assert all(cell.fill.fgColor.rgb == "FFFFFF00" for cell in sheet[3])
    finally:
        restock.close()


def test_purchase_source_rows_group_inventory_by_contract_and_price() -> None:
    columns = erp._replace_quantity_header(list(erp.purchase_summary.MANUFACTURER_COLUMNS))
    original = {column: "" for column in erp.purchase_summary.MANUFACTURER_COLUMNS}
    original.update({"型号": "A-1", "厂家": "供应商A", "原价": 5})
    line = {
        "planned_shipment_quantity": 7,
        "purchase_quantity": 2,
        "carryover_applied_quantity": 5,
        "contract_no": "NEW-001",
        "allocation_details": [
            {
                "sp_no": "SP-1",
                "stock_sku": "SKU-A",
                "source_kind": "current_purchase",
                "quantity": 2,
            },
            {
                "sp_no": "SP-1",
                "stock_sku": "SKU-A",
                "source_kind": "carryover",
                "source_contract_no": "OLD-001",
                "historical_tax_unit_price": 3,
                "quantity": 1,
            },
            {
                "sp_no": "SP-2",
                "stock_sku": "SKU-B",
                "source_kind": "carryover",
                "source_contract_no": "OLD-002",
                "historical_tax_unit_price": "2.5",
                "quantity": 2,
            },
            {
                "sp_no": "SP-1",
                "stock_sku": "SKU-A",
                "source_kind": "carryover",
                "source_contract_no": "OLD-001",
                "historical_tax_unit_price": 3,
                "quantity": 2,
            },
        ],
    }

    current, inventory = erp._purchase_source_rows(
        original,
        line,
        columns=columns,
        product_names={("SP-1", "SKU-A"): "产品A", ("SP-2", "SKU-B"): "产品B"},
    )

    assert len(current) == 1
    assert len(inventory) == 2
    current_values = dict(zip(columns, current[0], strict=True))
    first_inventory = dict(zip(columns, inventory[0], strict=True))
    second_inventory = dict(zip(columns, inventory[1], strict=True))
    assert (
        current_values["计划发货量"],
        current_values["本次采购量"],
        current_values["留存库存抵扣量"],
    ) == (2, 2, 0)
    assert current_values["本次采购合同编号"] == "NEW-001"
    assert current_values["历史库存合同编号"] == ""
    assert current_values["库存sku"] == "SKU-A × 2"
    assert (
        first_inventory["计划发货量"],
        first_inventory["本次采购量"],
        first_inventory["留存库存抵扣量"],
    ) == (3, 0, 3)
    assert first_inventory["历史库存合同编号"] == "OLD-001"
    assert first_inventory["原价"] == 3
    assert first_inventory["总价"] == 9
    assert second_inventory["历史库存合同编号"] == "OLD-002"
    assert second_inventory["原价"] == 2.5
    assert second_inventory["总价"] == 5


@pytest.mark.parametrize(
    "mutate",
    [
        lambda response: response["purchase_lines"][0].pop("purchase_quantity"),
        lambda response: response["purchase_lines"][0].__setitem__("purchase_quantity", 7),
        lambda response: response["purchase_lines"][0].__setitem__(
            "allocation_details",
            response["purchase_lines"][0]["allocation_details"][:-1],
        ),
        lambda response: response["contracts"][0].pop("daily_sequence"),
        lambda response: response["contracts"][0].__setitem__(
            "supplier_contract_count", 222
        ),
    ],
    ids=[
        "missing-required-field",
        "line-invariant",
        "sp-sku-conservation",
        "missing-daily-sequence",
        "supplier-count-behind-sequence",
    ],
)
def test_success_response_must_be_complete_and_conserve_allocations(mutate) -> None:
    response = deepcopy(_erp_result())
    mutate(response)

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.validate_purchase_response(
            status_code=201,
            response=response,
            request_payload=_request_payload(),
        )

    assert captured.value.code == "erp_purchase_result_incomplete"


def test_quote_response_must_preserve_inventory_source_lineage() -> None:
    response = _quote_response()
    response["confirmation"]["affected_lines"][0]["inventory_sources"][0].pop(
        "source_contract_no"
    )

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.validate_purchase_response(
            status_code=409,
            response=response,
            request_payload=_request_payload(),
        )

    assert captured.value.code == "erp_purchase_result_incomplete"
    assert "source_contract_no" in str(captured.value)


def test_quote_response_preserves_complete_v2_inventory_batch_semantics() -> None:
    response = _quote_response()

    erp.validate_purchase_response(
        status_code=409,
        response=response,
        request_payload=_request_payload(),
    )

    confirmation = response["confirmation"]
    assert confirmation["inventory_changes_committed"] is False
    line = confirmation["affected_lines"][0]
    source = line["inventory_sources"][0]
    assert line["proposed_inventory_deduction_quantity"] == 4
    assert line["proposed_purchase_quantity"] == 6
    assert source == {
        "carryover_entry_id": "00000000-0000-0000-0000-000000000003",
        "source_kind": "opening_inventory",
        "source_contract_no": "ZF20260601001",
        "source_sp_no": "SP260601001",
        "source_reference": "SP260601001-row-1",
        "historical_tax_unit_price": 3,
        "original_quantity": 4,
        "current_remaining_quantity": 4,
        "replacement_released_quantity": 0,
        "available_after_release": 4,
        "proposed_applied_quantity": 4,
    }
    assert "suggested_applied_quantity" not in source
    assert "carryover_applied_quantity" not in line
    assert "purchase_quantity" not in line


def test_quote_response_accepts_manual_inventory_import_source() -> None:
    response = _quote_response()
    source = response["confirmation"]["affected_lines"][0]["inventory_sources"][0]
    source["source_kind"] = "manual_import"

    erp.validate_purchase_response(
        status_code=409,
        response=response,
        request_payload=_request_payload(),
    )

    assert source["source_kind"] == "manual_import"


def test_quote_response_keeps_same_contract_in_distinct_sp_inventory_batches() -> None:
    response = _quote_response()
    line = response["confirmation"]["affected_lines"][0]
    first = line["inventory_sources"][0]
    first["proposed_applied_quantity"] = 2
    second = deepcopy(first)
    second.update(
        {
            "carryover_entry_id": "00000000-0000-0000-0000-000000000005",
            "source_sp_no": "SP260601002",
            "source_reference": "SP260601002-row-2",
            "original_quantity": 2,
            "current_remaining_quantity": 2,
            "available_after_release": 2,
            "proposed_applied_quantity": 2,
        }
    )
    line["inventory_sources"].append(second)

    erp.validate_purchase_response(
        status_code=409,
        response=response,
        request_payload=_request_payload(),
    )

    assert [source["source_sp_no"] for source in line["inventory_sources"]] == [
        "SP260601001",
        "SP260601002",
    ]


def test_quote_response_rejects_duplicate_inventory_batch_identity() -> None:
    response = _quote_response()
    line = response["confirmation"]["affected_lines"][0]
    first = line["inventory_sources"][0]
    first["proposed_applied_quantity"] = 2
    duplicate = deepcopy(first)
    duplicate["source_sp_no"] = "SP260601002"
    duplicate["source_reference"] = "SP260601002-row-2"
    line["inventory_sources"].append(duplicate)

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.validate_purchase_response(
            status_code=409,
            response=response,
            request_payload=_request_payload(),
        )

    assert captured.value.code == "erp_purchase_result_incomplete"
    assert "库存批次身份重复" in str(captured.value)


@pytest.mark.parametrize(
    "mutate",
    [
        lambda source: source.pop("current_remaining_quantity"),
        lambda source: source.__setitem__("replacement_released_quantity", 1),
        lambda source: source.__setitem__("proposed_applied_quantity", 5),
        lambda source: source.__setitem__("source_kind", "unknown"),
    ],
    ids=[
        "missing-current-balance",
        "release-equation",
        "proposed-exceeds-available",
        "unknown-source-kind",
    ],
)
def test_quote_response_rejects_incomplete_or_invalid_v2_inventory_source(mutate) -> None:
    response = _quote_response()
    source = response["confirmation"]["affected_lines"][0]["inventory_sources"][0]
    mutate(source)

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.validate_purchase_response(
            status_code=409,
            response=response,
            request_payload=_request_payload(),
        )

    assert captured.value.code == "erp_purchase_result_incomplete"


def test_quote_response_rejects_v1_confirmation_schema() -> None:
    response = _quote_response()
    response["response_schema"] = "lxe.erp.purchase-confirmation.v1"

    with pytest.raises(erp.PurchaseBatchClientError) as captured:
        erp.validate_purchase_response(
            status_code=409,
            response=response,
            request_payload=_request_payload(),
        )

    assert captured.value.code == "erp_purchase_confirmation_schema_unsupported"


def test_cli_marks_success_validation_failure_as_committed(monkeypatch) -> None:
    response = deepcopy(_erp_result())
    response["purchase_lines"][0].pop("purchase_quantity")
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (
            _request_payload(),
            {
                "delivery_nos": ["SP260710001"],
                "csv_paths": [],
                "master_xlsx": "master.xlsx",
            },
        ),
    )
    monkeypatch.setattr(erp, "import_purchase_intent", lambda _payload: (201, response))
    monkeypatch.setattr(contract_cli, "validate_contract_template", lambda *args, **kwargs: {})
    monkeypatch.setattr(
        cli,
        "generate_purchase_batch_workbooks",
        lambda *args, **kwargs: pytest.fail("invalid ERP response must not generate files"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
        }
    )

    assert result["success"] is False
    assert result["status"] == "batch_committed_artifact_generation_failed"
    assert result["error"]["code"] == "batch_committed_artifact_generation_failed"
    assert result["artifact_error"]["code"] == "erp_purchase_result_incomplete"
    assert result["batch_id"] == response["batch_id"]


def test_formal_success_renders_contracts_locally_without_erp_download(
    monkeypatch,
    tmp_path: Path,
) -> None:
    response = deepcopy(_erp_result())
    purchase_path = tmp_path / "purchase.xlsx"
    restock_path = tmp_path / "restock.xlsx"
    contract_path = tmp_path / "ZF202607230001-深圳正飞科技.xlsx"
    for path in (purchase_path, restock_path, contract_path):
        path.write_text("artifact", encoding="utf-8")
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (
            _request_payload(),
            {
                "delivery_nos": ["SP260710001"],
                "csv_paths": [],
                "master_xlsx": "master.xlsx",
            },
        ),
    )
    monkeypatch.setattr(contract_cli, "validate_contract_template", lambda *args, **kwargs: {})
    monkeypatch.setattr(erp, "import_purchase_intent", lambda _payload: (201, response))
    monkeypatch.setattr(
        cli,
        "generate_purchase_batch_workbooks",
        lambda *args, **kwargs: {
            "success": True,
            "purchase_summary_xlsx": str(purchase_path),
            "restock_xlsx_paths": [str(restock_path)],
            "restock_outputs": [
                {"delivery_no": "SP260710001", "output_xlsx": str(restock_path)}
            ],
        },
    )
    monkeypatch.setattr(
        erp,
        "apply_formal_erp_result",
        lambda generated, erp_result, **kwargs: {
            **generated,
            **erp_result,
            "mode": "formal",
        },
    )
    seen: list[dict] = []

    def fake_fill_formal_purchase_contracts(**kwargs):
        seen.append(kwargs)
        return {
            "success": True,
            "contract_template_xlsx": kwargs["contract_template_xlsx"],
            "output_files": [
                {
                    "manufacturer": "深圳正飞科技",
                    "sheet_name": "深圳正飞科技",
                    "contract_no": "ZF202607230001",
                    "output_xlsx": str(contract_path),
                }
            ],
        }

    monkeypatch.setattr(
        contract_cli,
        "fill_formal_purchase_contracts",
        fake_fill_formal_purchase_contracts,
    )
    monkeypatch.setattr(
        erp,
        "download_contract_workbooks",
        lambda _result: pytest.fail("formal mode must not download contracts from ERP"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
        }
    )

    assert result["success"] is True
    assert result["status"] == "completed"
    assert result["result_schema"] == "lxe.fba.purchase-summary-result.v1"
    assert result["quantity_summary"] == {
        "planned_shipment_quantity": "10",
        "carryover_applied_quantity": "4",
        "purchase_quantity": "6",
    }
    assert result["artifact_summary"] == {
        "delivery_count": 1,
        "restock_count": 1,
        "contract_count": 1,
        "deliverable_file_count": 3,
    }
    assert result["contract_xlsx_paths"] == [str(contract_path)]
    assert result["contracts"] == [
        {
            "contract_id": "00000000-0000-0000-0000-000000000002",
            "supplier_name": "深圳正飞科技",
            "contract_no": "ZF202607230001",
            "daily_sequence": 1,
            "supplier_contract_sequence": 223,
            "supplier_contract_count": 223,
            "output_xlsx": str(contract_path),
        }
    ]
    assert "erp" not in result
    assert "purchase_lines" not in result
    assert "restock_outputs" not in result
    assert "contract_outputs" not in result
    assert seen[0]["contracts"] == response["contracts"]


def test_formal_success_result_is_compact_and_keeps_every_deliverable_path() -> None:
    response = deepcopy(_erp_result())
    base_line = response["purchase_lines"][0]
    response["purchase_lines"] = []
    for index in range(16):
        line = deepcopy(base_line)
        line["line_ref"] = f"L{index + 1:04d}"
        line["model"] = f"MODEL-{index + 1:02d}-完整型号"
        line["allocation_details"] = [
            {
                **deepcopy(base_line["allocation_details"][detail_index % 2]),
                "stock_sku": f"SKU-{index + 1:02d}-{detail_index + 1:02d}",
            }
            for detail_index in range(24)
        ]
        response["purchase_lines"].append(line)

    response["contracts"] = [
        {
            "contract_id": f"contract-{index + 1:02d}",
            "supplier_name": f"供应商-{index + 1:02d}",
            "contract_no": f"HT20260731{index + 1:04d}",
            "daily_sequence": index + 1,
            "supplier_contract_sequence": index + 1,
            "supplier_contract_count": index + 1,
        }
        for index in range(14)
    ]
    restock_paths = [f"/artifacts/SP-{index + 1:02d}-备货单.xlsx" for index in range(10)]
    contract_outputs = [
        {
            "manufacturer": contract["supplier_name"],
            "sheet_name": contract["supplier_name"],
            "contract_no": contract["contract_no"],
            "output_xlsx": f"/artifacts/{contract['contract_no']}.xlsx",
        }
        for contract in response["contracts"]
    ]
    formal = {
        "success": True,
        "delivery_nos": [f"SP-{index + 1:02d}" for index in range(10)],
        "csv_paths": [f"/inputs/SP-{index + 1:02d}.csv" for index in range(10)],
        "master_xlsx": "/inputs/master.xlsx",
        "gross_margin": "0.3",
        "pricing_basis": "fixture",
        "purchase_summary_xlsx": "/artifacts/purchase-summary.xlsx",
        "restock_xlsx_paths": restock_paths,
        "restock_outputs": [
            {"delivery_no": f"SP-{index + 1:02d}", "output_xlsx": path}
            for index, path in enumerate(restock_paths)
        ],
        "sku_count": 415,
        "manufacturer_count": 15,
        "warnings": [],
        "source": "fba_purchase_batch_workbooks",
        "erp": deepcopy(response),
        "contracts": deepcopy(response["contracts"]),
        "purchase_lines": deepcopy(response["purchase_lines"]),
        "contract_outputs": deepcopy(contract_outputs),
        "contract_xlsx_paths": [item["output_xlsx"] for item in contract_outputs],
    }
    contract_result = {"output_files": contract_outputs, "warnings": []}

    compact = cli._formal_success_result(
        formal,
        response,
        contract_result,
        {
            "unmatched_summary": {
                "stock_sku_count": 1,
                "sp_sku_count": 1,
                "component_count": 2,
                "planned_shipment_quantity": "5",
                "items": [],
            }
        },
    )

    old_size = len(json.dumps(formal, ensure_ascii=False, separators=(",", ":")))
    compact_size = len(json.dumps(compact, ensure_ascii=False, separators=(",", ":")))
    assert compact_size <= old_size * 0.35
    assert compact_size < 12_000
    assert compact["artifact_summary"] == {
        "delivery_count": 10,
        "restock_count": 10,
        "contract_count": 14,
        "deliverable_file_count": 25,
    }
    assert compact["manufacturer_count"] == 15
    assert compact["unmatched_summary"]["planned_shipment_quantity"] == "5"
    assert any("未匹配库存 SKU" in warning for warning in compact["warnings"])
    assert compact["contract_xlsx_paths"] == [
        item["output_xlsx"] for item in contract_outputs
    ]
    assert compact["restock_xlsx_paths"] == restock_paths
    assert "erp" not in compact
    assert "purchase_lines" not in compact


def test_cli_preserves_local_artifacts_when_formal_contract_generation_fails(
    monkeypatch,
    tmp_path: Path,
) -> None:
    response = deepcopy(_erp_result())
    purchase_path = tmp_path / "purchase.xlsx"
    restock_path = tmp_path / "restock.xlsx"
    completed_contract_path = tmp_path / "ZF202607230001-厂家A.xlsx"
    for path in (purchase_path, restock_path, completed_contract_path):
        path.write_text("artifact", encoding="utf-8")
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (
            _request_payload(),
            {
                "delivery_nos": ["SP260710001"],
                "csv_paths": [],
                "master_xlsx": "master.xlsx",
            },
        ),
    )
    monkeypatch.setattr(erp, "import_purchase_intent", lambda _payload: (201, response))
    monkeypatch.setattr(contract_cli, "validate_contract_template", lambda *args, **kwargs: {})
    monkeypatch.setattr(
        cli,
        "generate_purchase_batch_workbooks",
        lambda *args, **kwargs: {
            "success": True,
            "purchase_summary_xlsx": str(purchase_path),
            "restock_xlsx_paths": [str(restock_path)],
            "restock_outputs": [
                {"delivery_no": "SP260710001", "output_xlsx": str(restock_path)}
            ],
        },
    )
    monkeypatch.setattr(
        erp,
        "apply_formal_erp_result",
        lambda generated, erp_result, **kwargs: {
            **generated,
            **erp_result,
            "mode": "formal",
        },
    )
    monkeypatch.setattr(
        contract_cli,
        "fill_formal_purchase_contracts",
        lambda **kwargs: (_ for _ in ()).throw(
            contract_cli.FormalContractGenerationError(
                "disk full",
                output_files=[
                    {
                        "manufacturer": "厂家A",
                        "sheet_name": "厂家A",
                        "contract_no": "ZF202607230001",
                        "output_xlsx": str(completed_contract_path),
                    }
                ],
            )
        ),
    )
    monkeypatch.setattr(
        erp,
        "download_contract_workbooks",
        lambda _result: pytest.fail("formal mode must not download contracts from ERP"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "contract_template_xlsx": "contracts.xlsx",
            "gross_margin": "0.3",
        }
    )

    assert result["status"] == "batch_committed_artifact_generation_failed"
    assert result["batch_id"] == response["batch_id"]
    assert result["purchase_summary_xlsx"] == str(purchase_path)
    assert result["restock_xlsx_paths"] == [str(restock_path)]
    assert result["contract_xlsx_paths"] == [str(completed_contract_path)]
    assert result["artifact_error"] == {
        "code": "FormalContractGenerationError",
        "message": "disk full",
    }


def test_invalid_contract_template_stops_before_erp_commit(monkeypatch, tmp_path: Path) -> None:
    invalid_template = tmp_path / "invalid-contract-template.xlsx"
    invalid_template.write_text("not an xlsx", encoding="utf-8")
    monkeypatch.setattr(
        erp,
        "build_purchase_intent",
        lambda *args, **kwargs: (
            _request_payload(),
            {
                "delivery_nos": ["SP260710001"],
                "csv_paths": [],
                "master_xlsx": "master.xlsx",
            },
        ),
    )
    monkeypatch.setattr(
        erp,
        "import_purchase_intent",
        lambda _payload: pytest.fail("invalid local template must stop before ERP commit"),
    )

    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "contract_template_xlsx": str(invalid_template),
            "gross_margin": "0.3",
        }
    )

    assert result["success"] is False
    assert result["error"]["code"] == "purchase_batch_generation_failed"
    assert "不是有效 xlsx" in result["error"]["message"]


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    workbook.active["A1"] = "ERP正式合同"
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_contract_download_is_safe_deliverable_and_retry_restores_missing_file(
    monkeypatch,
    tmp_path: Path,
) -> None:
    from services.agent_cli.mabang import erp_http

    calls: list[str] = []
    encoded_name = quote("../../正飞合同?.xlsx", safe="")

    def fake_request_bytes(method: str, path: str, **_kwargs):
        calls.append(path)
        assert method == "GET"
        return (
            200,
            _xlsx_bytes(),
            {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
        )

    monkeypatch.setattr(erp_http, "request_bytes", fake_request_bytes)
    source = {"contracts": deepcopy(_erp_result()["contracts"])}

    first = erp.download_contract_workbooks(dict(source), output_dir=tmp_path)
    contract_path = Path(first["contract_xlsx_paths"][0])
    assert contract_path.parent == tmp_path.resolve()
    assert contract_path.name == "正飞合同_.xlsx"
    assert first["contract_outputs"][0]["output_xlsx"] == str(contract_path)
    workbook = load_workbook(contract_path, read_only=True)
    try:
        assert workbook.active["A1"].value == "ERP正式合同"
    finally:
        workbook.close()

    contract_path.unlink()
    second = erp.download_contract_workbooks(dict(source), output_dir=tmp_path)
    assert Path(second["contract_xlsx_paths"][0]).is_file()
    assert calls == [
        "/api/v1/erp/contracts/00000000-0000-0000-0000-000000000002/download",
        "/api/v1/erp/contracts/00000000-0000-0000-0000-000000000002/download",
    ]


def test_formal_restock_rebuilds_current_and_inventory_skus_and_product_names(
    tmp_path: Path,
) -> None:
    csv_dir = tmp_path / "csv"
    csv_dir.mkdir()
    (csv_dir / "SP260710001_1.csv").write_text(
        "\n".join(
            [
                '"发货单号","MSKU","MSKU发货量","SKU发货量","国家","备注"',
                '"SP260710001","MSKU-X","5","SKU-A × 10\nSKU-B × 5","德国",""',
            ]
        ),
        encoding="utf-8-sig",
    )
    master = tmp_path / "master.xlsx"
    _write_master(master)
    workbook = load_workbook(master)
    workbook["SKU表"].append(["SKU-B", "产品B", "A-1", 3.5, "深圳正飞科技", ""])
    workbook.save(master)
    workbook.close()

    request_payload, _context = erp.build_purchase_intent(
        ["SP260710001"], master_xlsx=master, csv_dir=csv_dir
    )
    response = deepcopy(_erp_result())
    line = response["purchase_lines"][0]
    line["planned_shipment_quantity"] = 15
    line["carryover_applied_quantity"] = 4
    line["purchase_quantity"] = 11
    line["allocation_details"] = [
        line["allocation_details"][0],
        {**line["allocation_details"][1], "quantity": 6},
        {
            **line["allocation_details"][1],
            "stock_sku": "SKU-B",
            "quantity": 5,
        },
    ]
    erp.validate_purchase_response(
        status_code=201,
        response=response,
        request_payload=request_payload,
    )
    generated = cli.generate_purchase_batch_workbooks(
        ["SP260710001"],
        master_xlsx=master,
        gross_margin="0.3",
        csv_dir=csv_dir,
        purchase_output_dir=tmp_path / "purchase",
        restock_output_dir=tmp_path / "restock",
    )
    formal = erp.apply_formal_erp_result(
        generated,
        response,
        request_payload=request_payload,
    )

    workbook = load_workbook(formal["restock_xlsx_paths"][0], data_only=True)
    try:
        sheet = workbook["备货单"]
        headers = [cell.value for cell in sheet[1]]
        sku_column = headers.index("库存sku") + 1
        product_column = headers.index("产品名称") + 1
        assert sheet.cell(2, sku_column).value == "SKU-A × 6\nSKU-B × 5"
        assert sheet.cell(2, product_column).value == "产品A\n产品B"
        assert sheet.cell(3, sku_column).value == "SKU-A × 4"
        assert sheet.cell(3, product_column).value == "产品A"
        assert all(cell.fill.fgColor.rgb == "FFFFFF00" for cell in sheet[3])
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "confirmation_argument",
    [
        {"confirm_inventory_quote_id": "quote-1"},
        {"confirm_unmatched_sku_token": "unmatched-token"},
    ],
)
def test_draft_rejects_confirmation_flags_without_http(
    confirmation_argument: dict[str, str],
) -> None:
    result = cli.run(
        {
            "delivery_no": ["SP260710001"],
            "master_xlsx": "master.xlsx",
            "gross_margin": "0.3",
            "draft": True,
            **confirmation_argument,
        }
    )
    assert result["success"] is False
    assert result["error"]["code"] == "draft_arguments_conflict"


def test_contract_fill_uses_actual_purchase_quantity_and_skips_zero(tmp_path: Path) -> None:
    path = tmp_path / "formal-summary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购汇总"
    sheet.append(
        [
            "厂家",
            "合同产品名称",
            "型号",
            "单位",
            "计划发货量",
            "本次采购量",
            "留存库存抵扣量",
            "原价",
            "总价",
            "税率",
        ]
    )
    sheet.append(["厂家A", "产品A", "A-1", "个", 10, 6, 4, 3.5, 21, "13%"])
    sheet.append(["厂家A", "产品B", "A-2", "个", 5, 0, 5, 4, 0, "13%"])
    workbook.save(path)

    grouped = contract_cli.load_purchase_summary_lines(path)

    assert len(grouped["厂家A"]) == 1
    assert grouped["厂家A"][0].quantity == 6
