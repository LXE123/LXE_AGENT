from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from services.agent_cli.mabang import fill_purchase_contracts as contract_workbook
from services.agent_cli.mabang import regenerate_purchase_contracts as cli


BATCH_ID = "00000000-0000-0000-0000-000000000001"
CONTRACT_ID = "00000000-0000-0000-0000-000000000002"
BATCH_NO = "PB20260723-0001"


def _write_detail_sheet(worksheet, *, marker: str) -> None:
    worksheet["A1"] = marker
    worksheet["E2"] = "合同编号：OLD\nDate：2000年1月1日"
    worksheet["E3"] = (
        "交货日期：2000年1月4日\n付款期限：发货验收付款\n币种：人民币 税率：0%"
    )
    headers = [
        "序号",
        "产品名称",
        "规格型号",
        "单位",
        "数量",
        "含税单价",
        "含税金额（元）",
        "备注",
    ]
    for column, header in enumerate(headers, start=1):
        worksheet.cell(row=4, column=column, value=header)
    worksheet.append([None] * len(headers))
    worksheet.append(["合计", None, None, None, None, None, 0, None])


def _write_template(path: Path, suppliers: list[str], *, marker: str = "最新模板") -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for supplier in suppliers:
        _write_detail_sheet(workbook.create_sheet(supplier), marker=marker)
    addendum = workbook.create_sheet(contract_workbook.ADDENDUM_TEMPLATE_SHEET)
    _write_detail_sheet(addendum, marker="附加件")
    addendum["A2"] = "采购合同编号：OLD"
    workbook.save(path)
    workbook.close()


def _contract_summary(
    *,
    contract_id: str = CONTRACT_ID,
    supplier_name: str = "厂家A",
    contract_no: str = "HT20260723001",
) -> dict:
    return {
        "contract_id": contract_id,
        "supplier_name": supplier_name,
        "contract_no": contract_no,
        "status": "current",
    }


def _batch(*, status: str = "current", contracts: list[dict] | None = None) -> dict:
    return {
        "batch_id": BATCH_ID,
        "batch_no": BATCH_NO,
        "revisions": [
            {
                "batch_id": BATCH_ID,
                "version_no": 1,
                "status": status,
                "contracts": [_contract_summary()] if contracts is None else contracts,
            }
        ],
    }


def _contract_detail(
    *,
    contract_id: str = CONTRACT_ID,
    supplier_name: str = "厂家A",
    contract_no: str = "HT20260723001",
) -> dict:
    return {
        "contract_id": contract_id,
        "supplier_name": supplier_name,
        "contract_no": contract_no,
        "contract_date": "2026-07-23",
        "tax_rate": "13%",
        "source_kind": "generated",
        "status": "current",
        "batch_id": BATCH_ID,
        "batch_no": BATCH_NO,
        "version_no": 1,
        "lines": [
            {
                "contract_product_name": "合同产品A",
                "model": "M-1",
                "unit": "个",
                "purchase_quantity": "6",
                "tax_unit_price": "4.25",
            }
        ],
    }


def _fake_erp(monkeypatch, batch: dict, details: dict[str, dict]) -> list[str]:
    calls: list[str] = []

    def fake_request(method: str, path: str, **_kwargs):
        calls.append(path)
        assert method == "GET"
        if "/purchase-batches/by-number/" in path:
            return 200, deepcopy(batch)
        contract_id = path.rsplit("/", 1)[-1]
        return 200, deepcopy(details[contract_id])

    monkeypatch.setattr(cli.erp_http, "request_json", fake_request)
    return calls


def _cell(path: Path, sheet: str, coordinate: str):
    workbook = load_workbook(path, data_only=True)
    try:
        return workbook[sheet][coordinate].value
    finally:
        workbook.close()


def test_regenerates_current_contract_with_saved_erp_values_and_overwrites(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    output_dir = tmp_path / "contracts"
    output_dir.mkdir()
    output_path = output_dir / "HT20260723001-厂家A.xlsx"
    output_path.write_text("old contract", encoding="utf-8")
    monkeypatch.setattr(contract_workbook, "OUTPUT_DIR", output_dir)
    calls = _fake_erp(monkeypatch, _batch(), {CONTRACT_ID: _contract_detail()})

    result = cli.run(
        {
            "batch_no": " pb20260723-0001 ",
            "contract_template_xlsx": str(template),
        }
    )

    assert result["success"] is True
    assert result["result_schema"] == cli.RESULT_SCHEMA
    assert result["batch_no"] == BATCH_NO
    assert result["version_no"] == 1
    assert result["generated_count"] == 1
    assert result["contract_xlsx_paths"] == [str(output_path)]
    assert calls == [
        f"/api/v1/erp/purchase-batches/by-number/{BATCH_NO}",
        f"/api/v1/erp/contracts/{CONTRACT_ID}",
    ]
    assert _cell(output_path, "厂家A", "A1") == "最新模板"
    assert _cell(output_path, "厂家A", "E2") == "合同编号：HT20260723001\nDate：2026年7月23日"
    assert "交货日期：2026年7月26日" in _cell(output_path, "厂家A", "E3")
    assert _cell(output_path, "厂家A", "B5") == "合同产品A"
    assert _cell(output_path, "厂家A", "E5") == 6
    assert _cell(output_path, "厂家A", "F5") == 4.25
    assert _cell(output_path, "厂家A", "G5") == 25.5
    assert not list(output_dir.glob(".*.part"))


def test_regenerates_every_contract_in_the_current_revision(monkeypatch, tmp_path: Path) -> None:
    second_id = "00000000-0000-0000-0000-000000000003"
    summaries = [
        _contract_summary(),
        _contract_summary(
            contract_id=second_id,
            supplier_name="厂家B",
            contract_no="HT20260723002",
        ),
    ]
    details = {
        CONTRACT_ID: _contract_detail(),
        second_id: _contract_detail(
            contract_id=second_id,
            supplier_name="厂家B",
            contract_no="HT20260723002",
        ),
    }
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A", "厂家B"])
    monkeypatch.setattr(contract_workbook, "OUTPUT_DIR", tmp_path / "contracts")
    _fake_erp(monkeypatch, _batch(contracts=summaries), details)

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is True
    assert result["generated_count"] == 2
    assert [item["supplier_name"] for item in result["contracts"]] == ["厂家A", "厂家B"]


def test_rejects_batch_without_current_revision_before_fetching_contracts(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    calls = _fake_erp(monkeypatch, _batch(status="cancelled"), {})

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "purchase_batch_not_current"
    assert calls == [f"/api/v1/erp/purchase-batches/by-number/{BATCH_NO}"]


def test_reports_current_batch_with_no_contracts(monkeypatch, tmp_path: Path) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    _fake_erp(monkeypatch, _batch(contracts=[]), {})

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "purchase_batch_has_no_current_contracts"


def test_rejects_contract_detail_from_another_batch(monkeypatch, tmp_path: Path) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    detail = _contract_detail()
    detail["batch_no"] = "PB20260723-OTHER"
    _fake_erp(monkeypatch, _batch(), {CONTRACT_ID: detail})

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "erp_contract_detail_mismatch"
    assert result["error"]["detail"]["field"] == "batch_no"


def test_reports_missing_contract_line_value(monkeypatch, tmp_path: Path) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    detail = _contract_detail()
    detail["lines"][0].pop("tax_unit_price")
    _fake_erp(monkeypatch, _batch(), {CONTRACT_ID: detail})

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "erp_contract_detail_invalid"
    assert result["error"]["detail"]["field"].endswith("tax_unit_price")


def test_atomic_overwrite_keeps_old_file_when_windows_rejects_replace(
    monkeypatch,
    tmp_path: Path,
) -> None:
    template = tmp_path / "contract-template.xlsx"
    _write_template(template, ["厂家A"])
    output_dir = tmp_path / "contracts"
    output_dir.mkdir()
    output_path = output_dir / "HT20260723001-厂家A.xlsx"
    original = b"original contract"
    output_path.write_bytes(original)
    monkeypatch.setattr(contract_workbook, "OUTPUT_DIR", output_dir)
    _fake_erp(monkeypatch, _batch(), {CONTRACT_ID: _contract_detail()})

    def locked_replace(_source, _target):
        raise PermissionError("合同文件正在被 Excel 占用")

    monkeypatch.setattr(contract_workbook.os, "replace", locked_replace)

    result = cli.run(
        {"batch_no": BATCH_NO, "contract_template_xlsx": str(template)}
    )

    assert result["success"] is False
    assert result["error"]["code"] == "purchase_contract_regeneration_failed"
    assert "合同文件正在被 Excel 占用" in result["error"]["message"]
    assert output_path.read_bytes() == original
    assert not list(output_dir.glob(".*.part"))
